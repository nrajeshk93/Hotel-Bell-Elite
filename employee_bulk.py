"""Excel template + import for bulk employee create (Add Employee → Bulk)."""

from __future__ import annotations

import io
import re
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

BLANK_ROWS = 200
INSTRUCTIONS_TITLE = "Bulk employee upload"
EMPLOYEES_SHEET = "Employees"
DROPDOWNS_SHEET = "Dropdowns"

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
TITLE_FONT = Font(bold=True, name="Calibri", size=14)
BODY_FONT = Font(name="Calibri", size=11)
REQUIRED_FILL = PatternFill("solid", fgColor="DBEAFE")
REQUIRED_FONT = Font(color="1E3A5F", bold=True, name="Calibri", size=11)

# Headers match Add Employee form. EMP ID is system-assigned (not in template).
# Changing this tuple invalidates old templates.
UNIFIED_HEADERS = (
    "Name *",
    "Mobile *",
    "Guardian Mobile",
    "Sex",
    "Department",
    "Total Off",
    "Status",
    "Company",
    "Gross Salary *",
    "Basic Pay",
    "EPF",
    "ESIC",
    "No EPF",
    "No ESIC",
    "Aadhar",
    "PAN",
    "EPF Number",
    "ESIC Number",
    "Address",
    "Bank Name",
    "Account Holder Name",
    "Account Number",
    "IFSC Code",
)

SEX_LABELS = ("Male", "Female")
STATUS_LABELS = ("active", "inactive")
YES_NO_LABELS = ("Yes", "No")


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("₹", "rs")
    return re.sub(r"[^a-z0-9]+", "", text)


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    return " ".join(text.split()).strip()


def employee_bulk_headers() -> tuple[str, ...]:
    return UNIFIED_HEADERS


def _style_header(ws, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _add_list_validation(ws, formula: str, cells: str, title: str, error: str) -> None:
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle=title,
        error=error,
    )
    dv.add(cells)
    ws.add_data_validation(dv)


def build_employee_bulk_template(
    *,
    departments: list[str] | tuple[str, ...],
    default_company: str = "Hotel Bell Elite",
) -> io.BytesIO:
    """Live workbook matching Add Employee → Single fields + department dropdowns."""
    headers = employee_bulk_headers()
    depts = [str(d).strip() for d in (departments or []) if str(d).strip()]
    company = (default_company or "Hotel Bell Elite").strip() or "Hotel Bell Elite"

    wb = Workbook()
    info = wb.active
    info.title = "Instructions"
    info["A1"] = INSTRUCTIONS_TITLE
    info["A1"].font = TITLE_FONT
    lines = [
        "",
        "This file matches Add Employee → Single.",
        "Download a fresh copy after department or form fields change in the app.",
        "",
        "1. Open the Employees sheet.",
        "2. Enter one employee per row.",
        "3. Required: Name, Mobile (10 digits), Gross Salary.",
        "4. Employee ID (HBE…) is assigned automatically on upload — do not enter it.",
        "5. Department / Sex / Status use dropdowns.",
        "6. No EPF / No ESIC: Yes or No (blank = No).",
        "7. Save and upload from Add Employee → Bulk.",
        "",
        "Do not rename sheets or the header row.",
    ]
    for idx, text in enumerate(lines, start=2):
        cell = info.cell(idx, 1, text)
        cell.font = BODY_FONT
    info.column_dimensions["A"].width = 92

    lists = wb.create_sheet(DROPDOWNS_SHEET)
    lists["A1"] = "Department"
    lists["B1"] = "Sex"
    lists["C1"] = "Status"
    lists["D1"] = "YesNo"
    _style_header(lists, 4)
    for i, name in enumerate(depts, start=2):
        lists.cell(i, 1, name)
    for i, label in enumerate(SEX_LABELS, start=2):
        lists.cell(i, 2, label)
    for i, label in enumerate(STATUS_LABELS, start=2):
        lists.cell(i, 3, label)
    for i, label in enumerate(YES_NO_LABELS, start=2):
        lists.cell(i, 4, label)
    for col, width in (("A", 18), ("B", 12), ("C", 12), ("D", 10)):
        lists.column_dimensions[col].width = width
    lists.freeze_panes = "A2"
    lists.sheet_state = "hidden"

    ws = wb.create_sheet(EMPLOYEES_SHEET, 1)
    ws.append(list(headers))
    _style_header(ws, len(headers))
    required_cols = {1, 2, 9}  # Name, Mobile, Gross Salary
    for col in required_cols:
        cell = ws.cell(1, col)
        cell.fill = REQUIRED_FILL
        cell.font = REQUIRED_FONT

    last_row = 1 + BLANK_ROWS
    for _ in range(BLANK_ROWS):
        row = [""] * len(headers)
        row[6] = "active"  # Status
        row[7] = company  # Company
        row[12] = "No"  # No EPF
        row[13] = "No"  # No ESIC
        ws.append(row)

    ws.freeze_panes = "A2"
    widths = {
        "A": 24,
        "B": 14,
        "C": 16,
        "D": 10,
        "E": 14,
        "F": 10,
        "G": 10,
        "H": 18,
        "I": 14,
        "J": 12,
        "K": 10,
        "L": 10,
        "M": 10,
        "N": 10,
        "O": 14,
        "P": 12,
        "Q": 14,
        "R": 14,
        "S": 28,
        "T": 18,
        "U": 20,
        "V": 16,
        "W": 14,
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    for row in range(2, last_row + 1):
        for col in (9, 10, 11, 12):
            ws.cell(row, col).number_format = "0.00"

    dept_end = 1 + max(len(depts), 1)
    _add_list_validation(
        ws,
        f"Dropdowns!$A$2:$A${dept_end}",
        f"E2:E{last_row}",
        "Department",
        "Pick a department from the list.",
    )
    _add_list_validation(
        ws,
        f"Dropdowns!$B$2:$B${1 + len(SEX_LABELS)}",
        f"D2:D{last_row}",
        "Sex",
        "Pick Male or Female.",
    )
    _add_list_validation(
        ws,
        f"Dropdowns!$C$2:$C${1 + len(STATUS_LABELS)}",
        f"G2:G{last_row}",
        "Status",
        "Pick active or inactive.",
    )
    _add_list_validation(
        ws,
        f"Dropdowns!$D$2:$D${1 + len(YES_NO_LABELS)}",
        f"M2:N{last_row}",
        "Yes/No",
        "Pick Yes or No.",
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _sheet_named(wb, name: str):
    for ws in wb.worksheets:
        if (ws.title or "").strip().casefold() == name.casefold():
            return ws
    return None


def _assert_matches_template(wb) -> None:
    employees = _sheet_named(wb, EMPLOYEES_SHEET)
    dropdowns = _sheet_named(wb, DROPDOWNS_SHEET)
    instructions = _sheet_named(wb, "Instructions")
    if employees is None or dropdowns is None or instructions is None:
        raise ValueError(
            "This file is not the employee template. Download the template and upload that file."
        )
    title = _clean_cell(instructions["A1"].value)
    if title.casefold() != INSTRUCTIONS_TITLE.casefold():
        raise ValueError(
            "This file is not the employee template. Download the template and upload that file."
        )
    expected = [_norm_header(h) for h in UNIFIED_HEADERS]
    actual = []
    for col in range(1, len(UNIFIED_HEADERS) + 1):
        actual.append(_norm_header(employees.cell(1, col).value))
    if actual != expected:
        raise ValueError(
            "This Excel file does not match the current employee template. "
            "Download a fresh template and try again."
        )


def _parse_yes_no(value: str) -> bool:
    v = (value or "").strip().casefold()
    return v in {"yes", "y", "1", "true"}


def _parse_money(value: str, *, default: float = 0.0) -> float:
    text = (value or "").strip().replace(",", "")
    if not text:
        return default
    return float(text)


def _assign_emp_code(
    conn,
    *,
    next_emp_code_fn: Callable,
    emp_code_taken_fn: Callable,
    seen_codes: set[str],
) -> str:
    emp_code = next_emp_code_fn(conn)
    while emp_code_taken_fn(conn, emp_code) or emp_code.upper() in seen_codes:
        match = re.match(r"^HBE(\d+)$", emp_code, re.I)
        if match:
            emp_code = f"HBE{int(match.group(1)) + 1}"
        else:
            emp_code = next_emp_code_fn(conn)
            break
    seen_codes.add(emp_code.upper())
    return emp_code


def import_employee_bulk(
    conn,
    file_stream,
    *,
    departments: list[str] | tuple[str, ...],
    next_emp_code_fn: Callable,
    emp_code_taken_fn: Callable,
    default_company: str = "Hotel Bell Elite",
    epf_max: float = 1800.0,
) -> dict[str, Any]:
    """Import employees from the Download template workbook."""
    raw = file_stream.read() if hasattr(file_stream, "read") else file_stream
    if not raw:
        raise ValueError("Upload an Excel file.")
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as exc:
        raise ValueError("Could not read Excel file. Upload a valid .xlsx.") from exc

    _assert_matches_template(wb)
    ws = _sheet_named(wb, EMPLOYEES_SHEET)
    assert ws is not None

    allowed_depts = {d.strip().casefold(): d.strip() for d in departments if str(d).strip()}
    company_default = (default_company or "Hotel Bell Elite").strip() or "Hotel Bell Elite"

    created = 0
    skipped = 0
    errors: list[dict[str, Any]] = []
    seen_mobiles: set[str] = set()
    seen_codes: set[str] = set()

    for row_idx in range(2, ws.max_row + 1):
        values = [_clean_cell(ws.cell(row_idx, c).value) for c in range(1, len(UNIFIED_HEADERS) + 1)]
        if not any(values):
            continue

        (
            name,
            mobile,
            guardian_mobile,
            sex,
            department,
            total_off_raw,
            status,
            company,
            gross_raw,
            basic_raw,
            epf_raw,
            esic_raw,
            no_epf_raw,
            no_esic_raw,
            aadhar,
            pan,
            epf_number,
            esic_number,
            address,
            bank_name,
            account_holder_name,
            account_number,
            ifsc_code,
        ) = values

        row_errors: list[str] = []
        if not name:
            row_errors.append("Name is required.")
        mobile_digits = re.sub(r"\D+", "", mobile)
        if not re.fullmatch(r"\d{10}", mobile_digits):
            row_errors.append("Mobile must be exactly 10 digits.")
        else:
            mobile = mobile_digits
        if guardian_mobile:
            g_digits = re.sub(r"\D+", "", guardian_mobile)
            if not re.fullmatch(r"\d{10}", g_digits):
                row_errors.append("Guardian mobile must be exactly 10 digits.")
            else:
                guardian_mobile = g_digits
        if aadhar and not re.fullmatch(r"\d{12}", re.sub(r"\D+", "", aadhar)):
            row_errors.append("Aadhar must be exactly 12 digits.")
        elif aadhar:
            aadhar = re.sub(r"\D+", "", aadhar)
        pan = pan.upper()
        if pan and not re.fullmatch(r"[A-Z]{5}\d{4}[A-Z]", pan):
            row_errors.append("PAN must be in format ABCDE1234F.")
        ifsc_code = ifsc_code.upper()
        if ifsc_code and not re.fullmatch(r"[A-Z]{4}0[A-Z0-9]{6}", ifsc_code):
            row_errors.append("IFSC must be in format ABCD0123456.")

        try:
            gross_salary = _parse_money(gross_raw, default=0.0)
            if gross_salary < 0:
                row_errors.append("Gross salary must be a positive number.")
        except ValueError:
            row_errors.append("Gross salary must be a valid number.")
            gross_salary = 0.0

        try:
            basic_salary = _parse_money(basic_raw, default=0.0)
        except ValueError:
            basic_salary = 0.0

        epf_exempt = 1 if _parse_yes_no(no_epf_raw) else 0
        esic_exempt = 1 if _parse_yes_no(no_esic_raw) else 0
        try:
            epf_amount = 0.0 if epf_exempt else _parse_money(epf_raw, default=0.0)
        except ValueError:
            epf_amount = 0.0
        if epf_amount > epf_max:
            epf_amount = float(epf_max)
        try:
            esic_amount = 0.0 if esic_exempt else _parse_money(esic_raw, default=0.0)
        except ValueError:
            esic_amount = 0.0

        try:
            total_off = int(float(total_off_raw)) if total_off_raw else 0
            if total_off < 0 or total_off > 31:
                row_errors.append("Total Off must be between 0 and 31.")
                total_off = 0
        except ValueError:
            row_errors.append("Total Off must be a whole number.")
            total_off = 0

        status_l = (status or "active").strip().lower()
        if status_l not in ("active", "inactive"):
            status_l = "active"
        sex_n = sex.strip().capitalize() if sex else ""
        if sex_n not in ("Male", "Female", ""):
            sex_n = ""

        location = ""
        if department:
            key = department.casefold()
            if key not in allowed_depts:
                row_errors.append(f"Unknown department “{department}”.")
            else:
                location = allowed_depts[key]

        company_val = company or company_default

        if mobile and mobile in seen_mobiles:
            row_errors.append("Duplicate mobile in this file.")
        elif mobile:
            dup_mobile = conn.execute(
                "SELECT id FROM employees WHERE mobile=?",
                (mobile,),
            ).fetchone()
            if dup_mobile:
                row_errors.append("Mobile already exists for another employee.")
            else:
                seen_mobiles.add(mobile)

        if row_errors:
            skipped += 1
            errors.append({"row": row_idx, "name": name, "errors": row_errors})
            continue

        emp_code = _assign_emp_code(
            conn,
            next_emp_code_fn=next_emp_code_fn,
            emp_code_taken_fn=emp_code_taken_fn,
            seen_codes=seen_codes,
        )

        conn.execute(
            """INSERT INTO employees
               (emp_code, name, company, location, mobile, guardian_mobile, sex, address,
                aadhar, pan, epf_number, esic_number, gross_salary, basic_salary,
                epf_amount, esic_amount, credit_repayment, epf_exempt, esic_exempt,
                weekday_shift, sunday_shift, bank_name, account_holder_name,
                account_number, ifsc_code, total_off, status)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                emp_code,
                name,
                company_val,
                location,
                mobile,
                guardian_mobile,
                sex_n,
                address,
                aadhar,
                pan,
                epf_number,
                esic_number,
                gross_salary,
                basic_salary,
                epf_amount,
                esic_amount,
                0.0,
                epf_exempt,
                esic_exempt,
                "",
                "",
                bank_name,
                account_holder_name,
                account_number,
                ifsc_code,
                total_off,
                status_l,
            ),
        )
        created += 1

    return {
        "created_count": created,
        "skipped_count": skipped,
        "error_count": len(errors),
        "errors": errors[:50],
    }
