"""Hotel GST report — room supply invoices with CGST/UGST splits."""

from __future__ import annotations

import calendar
import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO

from flask import render_template, request, send_file, url_for

from db import (
    _HOTEL_INVOICE_STAY_SOURCE_SQL,
    _hotel_normalize_invoice_row_status,
    _hotel_split_inclusive_tax,
    _normalize_agency_gst,
    _normalize_agency_name,
    backfill_hotel_room_invoices_from_layout,
    ensure_hotel_rooms_schema,
    get_db,
    get_hotel_guest_profile,
    get_hotel_tax_rates,
    list_agencies,
)
from reports import report_export_month_filename

GST_HOTEL_TITLE = "HOTEL BELL ELITE - GST REPORT"
GST_HOTEL_EXPORT_TITLE = "GST Hotel"
GST_HOTEL_HEADERS = (
    "AGENT NAME",
    "INVOICE DATE",
    "GUEST NAME",
    "GSTIN",
    "INVOICE NUMBER",
    "ROOM RENT",
    "TAXABLE VALUE",
    "CGST",
    "UGST",
    "TOTAL GST",
    "TOTAL",
    "ROUND OFF",
    "FINAL AMOUNT",
)
_CANCELLED_STATUSES = frozenset({"cancelled", "canceled", "void"})
_STORED_TAXABLE_KEYS = ("taxableValue", "taxable_value", "taxable")
_STORED_CGST_KEYS = ("cgstAmount", "cgst_amount", "cgst")
_STORED_UGST_KEYS = ("ugstAmount", "ugst_amount", "ugst", "sgstAmount", "sgst_amount")
_STORED_GST_KEYS = ("totalGst", "total_gst", "gstAmount", "gst_amount", "gst")
_STORED_RENT_KEYS = ("roomRent", "room_rent")
_STORED_TOTAL_KEYS = ("billTotal", "bill_total", "invoiceTotal", "invoice_total")
_STORED_ROUND_KEYS = ("roundOff", "round_off")
_STORED_FINAL_KEYS = ("finalAmount", "final_amount")
_STORED_GSTIN_KEYS = (
    "agencyGst",
    "agency_gst",
    "gstin",
    "gstIn",
    "gstIN",
    "gstNo",
    "gst_no",
    "gstNumber",
    "gst_number",
    "guestGst",
    "guest_gst",
)


def _round_money(value, dec=2):
    try:
        quantum = Decimal("1").scaleb(-dec)
        return float(Decimal(str(value or 0)).quantize(quantum, rounding=ROUND_HALF_UP))
    except (InvalidOperation, TypeError, ValueError):
        return 0.0


def _as_money(value):
    if value is None or value == "":
        return None
    try:
        return _round_money(value, 2)
    except (TypeError, ValueError):
        return None


def _first_money(*sources_and_keys):
    """Return the first stored numeric amount from dict sources."""
    sources = sources_and_keys[:-1]
    keys = sources_and_keys[-1]
    for source in sources:
        if not isinstance(source, dict):
            continue
        for key in keys:
            if key not in source:
                continue
            amount = _as_money(source.get(key))
            if amount is not None:
                return amount
    return None


def _parse_payload(raw):
    if isinstance(raw, dict):
        return raw
    try:
        payload = json.loads(raw or "{}")
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def _stay_from_payload(payload):
    stay = payload.get("stay") if isinstance(payload, dict) else None
    return stay if isinstance(stay, dict) else {}


def _parse_iso_date(value):
    text = str(value or "").strip()
    if not text:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = text.replace("T", " ", 1)
    for fmt, length in (("%Y-%m-%d %H:%M:%S", 19), ("%Y-%m-%d %H:%M", 16), ("%Y-%m-%d", 10)):
        chunk = text[:length]
        if len(chunk) < length:
            continue
        try:
            return datetime.strptime(chunk, fmt).date()
        except ValueError:
            continue
    return None


def _format_ddmmyyyy(value):
    parsed = _parse_iso_date(value)
    if parsed is None:
        return ""
    return parsed.strftime("%d-%m-%Y")


def _period_from_args(args, *, today=None):
    ref = today or date.today()
    try:
        year = int(args.get("year") or ref.year)
    except (TypeError, ValueError):
        year = ref.year
    try:
        month = int(args.get("month") or ref.month)
    except (TypeError, ValueError):
        month = ref.month
    if year < 2000 or year > 2100:
        year = ref.year
    if month < 1 or month > 12:
        month = ref.month
    start = date(year, month, 1)
    end = date(year, month, calendar.monthrange(year, month)[1])
    return year, month, start, end


def _is_cancelled(row):
    status = _hotel_normalize_invoice_row_status(row.get("status"))
    raw = str(row.get("status") or "").strip().lower()
    if status in _CANCELLED_STATUSES or raw in _CANCELLED_STATUSES:
        return True
    return False


def _agency_gst_index(conn):
    index = {}
    for agency in list_agencies(conn):
        name = _normalize_agency_name(agency.get("name")).lower()
        gst = _normalize_agency_gst(agency.get("gst"))
        if name and gst and name not in index:
            index[name] = gst
    return index


def _first_text(*sources_and_keys):
    sources = sources_and_keys[:-1]
    keys = sources_and_keys[-1]
    for source in sources:
        if not isinstance(source, dict):
            continue
        for key in keys:
            text = str(source.get(key) or "").strip()
            if text:
                return text
    return ""


def _resolve_gstin(stay, payload, guest_profile, agency_name, agency_gst_by_name):
    stored = _first_text(stay, payload, guest_profile or {}, _STORED_GSTIN_KEYS)
    stored = _normalize_agency_gst(stored)
    if stored:
        return stored
    key = _normalize_agency_name(agency_name).lower()
    if key:
        return agency_gst_by_name.get(key, "")
    return ""


def _tax_breakdown(stay, payload, inclusive, tax_rates):
    """Prefer stored invoice tax fields; otherwise split the inclusive room bill."""
    taxable = _first_money(stay, payload, _STORED_TAXABLE_KEYS)
    cgst = _first_money(stay, payload, _STORED_CGST_KEYS)
    ugst = _first_money(stay, payload, _STORED_UGST_KEYS)
    total_gst = _first_money(stay, payload, _STORED_GST_KEYS)
    room_rent = _first_money(stay, payload, _STORED_RENT_KEYS)
    stored_total = _first_money(stay, payload, _STORED_TOTAL_KEYS)
    round_off = _first_money(stay, payload, _STORED_ROUND_KEYS)
    final_amount = _first_money(stay, payload, _STORED_FINAL_KEYS)

    has_stored_split = taxable is not None and cgst is not None and ugst is not None
    if not has_stored_split:
        split_taxable, split_cgst, split_ugst, split_inclusive = _hotel_split_inclusive_tax(
            inclusive, tax_rates
        )
        if taxable is None:
            taxable = split_taxable
        if cgst is None:
            cgst = split_cgst
        if ugst is None:
            ugst = split_ugst
        if stored_total is None:
            stored_total = split_inclusive

    taxable = _round_money(taxable or 0, 2)
    cgst = _round_money(cgst or 0, 2)
    ugst = _round_money(ugst or 0, 2)
    if total_gst is None:
        total_gst = _round_money(cgst + ugst, 2)
    else:
        total_gst = _round_money(total_gst, 2)
    if room_rent is None:
        room_rent = taxable
    else:
        room_rent = _round_money(room_rent, 2)
    if stored_total is None:
        stored_total = _round_money(taxable + cgst + ugst, 2)
    else:
        stored_total = _round_money(stored_total, 2)
    if final_amount is None:
        final_amount = _round_money(stored_total, 0)
    else:
        final_amount = _round_money(final_amount, 2)
        if final_amount == int(final_amount):
            final_amount = float(int(final_amount))
    if round_off is None:
        round_off = _round_money(final_amount - stored_total, 2)
    else:
        round_off = _round_money(round_off, 2)
    return {
        "room_rent": room_rent,
        "taxable": taxable,
        "cgst": cgst,
        "ugst": ugst,
        "total_gst": total_gst,
        "total": stored_total,
        "round_off": round_off,
        "final_amount": final_amount,
    }


def _empty_totals():
    return {
        "room_rent": 0.0,
        "taxable": 0.0,
        "cgst": 0.0,
        "ugst": 0.0,
        "total_gst": 0.0,
        "total": 0.0,
        "round_off": 0.0,
        "final_amount": 0.0,
    }


def load_gst_hotel_rows(conn, *, date_from, date_to):
    """Hotel stay invoices in the month, excluding cancelled/void like sales KPIs."""
    ensure_hotel_rooms_schema(conn)
    backfill_hotel_room_invoices_from_layout(conn)
    start = date_from.isoformat() if hasattr(date_from, "isoformat") else str(date_from)[:10]
    end = date_to.isoformat() if hasattr(date_to, "isoformat") else str(date_to)[:10]
    tax_rates = get_hotel_tax_rates(conn)
    agency_gst_by_name = _agency_gst_index(conn)
    db_rows = conn.execute(
        f"""
        SELECT invoice_number, room_id, room_number, guest_name, booking_number,
               check_in_date, check_out_date, invoice_generated_at, estimated_total,
               status, source, payload_json, cancel_reason, cancelled_at
        FROM hotel_room_invoices
        WHERE {_HOTEL_INVOICE_STAY_SOURCE_SQL}
          AND substr(invoice_generated_at, 1, 10) >= ?
          AND substr(invoice_generated_at, 1, 10) <= ?
        ORDER BY substr(invoice_generated_at, 1, 10) ASC,
                 invoice_number ASC
        """,
        (start, end),
    ).fetchall()

    rows = []
    for raw in db_rows:
        item = dict(raw)
        if _is_cancelled(item):
            continue
        invoice_number = str(item.get("invoice_number") or "").strip()
        if not invoice_number:
            continue
        payload = _parse_payload(item.get("payload_json"))
        stay = _stay_from_payload(payload)
        guest_name = (
            str(item.get("guest_name") or "").strip()
            or str(stay.get("guestName") or stay.get("guest_name") or "").strip()
        )
        agent_name = _normalize_agency_name(
            stay.get("agencyName") or stay.get("agency_name") or item.get("agency_name")
        )
        mobile = stay.get("mobile") or stay.get("guestMobile") or stay.get("guest_mobile")
        guest_profile = get_hotel_guest_profile(conn, mobile) if mobile else None
        gstin = _resolve_gstin(stay, payload, guest_profile, agent_name, agency_gst_by_name)
        invoice_at = item.get("invoice_generated_at") or stay.get("invoiceGeneratedAt") or ""
        inclusive = _as_money(item.get("estimated_total"))
        if inclusive is None:
            inclusive = _as_money(stay.get("estimatedTotal") or stay.get("estimated_total")) or 0.0
        tax = _tax_breakdown(stay, payload, inclusive, tax_rates)
        rows.append(
            {
                "agent_name": agent_name,
                "invoice_date": _format_ddmmyyyy(invoice_at),
                "invoice_date_iso": str(invoice_at or "")[:10],
                "guest_name": guest_name,
                "gstin": gstin,
                "invoice_number": invoice_number,
                **tax,
            }
        )
    return rows


def summarize_gst_hotel_rows(rows):
    totals = _empty_totals()
    for row in rows or []:
        for key in totals:
            totals[key] = _round_money(totals[key] + float(row.get(key) or 0), 2)
    return {
        "invoice_count": len(rows or []),
        **totals,
    }


def build_gst_hotel_report(conn, *, year, month, today=None):
    ref = today or date.today()
    year = int(year)
    month = int(month)
    start = date(year, month, 1)
    end = date(year, month, calendar.monthrange(year, month)[1])
    rows = load_gst_hotel_rows(conn, date_from=start, date_to=end)
    totals = summarize_gst_hotel_rows(rows)
    return {
        "year": year,
        "month": month,
        "month_name": calendar.month_name[month],
        "date_from": start,
        "date_to": end,
        "period_label": f"FROM {start.strftime('%d-%m-%Y')} TO {end.strftime('%d-%m-%Y')}",
        "today": ref,
        "rows": rows,
        "totals": totals,
        "kpis": totals,
    }


def _whole_or_float(value):
    if value is None:
        return 0
    try:
        num = float(value)
    except (TypeError, ValueError):
        return value
    return int(num) if num == int(num) else num


def build_gst_hotel_workbook(payload):
    """Excel export matching other report chrome and the Room Supply layout."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    rows = payload.get("rows") or []
    totals = payload.get("totals") or _empty_totals()
    period_label = payload.get("period_label") or ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Room Supply"

    header_fill = PatternFill(
        fill_type="solid",
        start_color="FF315A78",
        end_color="FF315A78",
    )
    total_fill = PatternFill(
        fill_type="solid",
        start_color="FFD6E4F0",
        end_color="FFD6E4F0",
    )
    title_font = Font(name="Calibri", bold=True, size=16, color="FFFFFFFF")
    period_font = Font(name="Calibri", bold=True, size=12, color="FFFFFFFF")
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFFFF")
    body_font = Font(name="Calibri", size=11, color="FF000000")
    total_font = Font(name="Calibri", bold=True, size=11, color="FF000000")
    thin = Side(style="thin", color="FF000000")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    col_count = len(GST_HOTEL_HEADERS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws.cell(row=1, column=1, value=GST_HOTEL_TITLE)
    ws.cell(row=2, column=1, value=period_label)
    for col, title in enumerate(GST_HOTEL_HEADERS, start=1):
        ws.cell(row=3, column=col, value=title)

    money_cols = set(range(6, col_count + 1))
    data_start = 4
    for offset, row in enumerate(rows):
        ridx = data_start + offset
        values = (
            row.get("agent_name") or "",
            row.get("invoice_date") or "",
            row.get("guest_name") or "",
            row.get("gstin") or "",
            row.get("invoice_number") or "",
            _whole_or_float(row.get("room_rent")),
            _whole_or_float(row.get("taxable")),
            _whole_or_float(row.get("cgst")),
            _whole_or_float(row.get("ugst")),
            _whole_or_float(row.get("total_gst")),
            _whole_or_float(row.get("total")),
            _whole_or_float(row.get("round_off")),
            _whole_or_float(row.get("final_amount")),
        )
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=ridx, column=col, value=value)
            if col in money_cols:
                cell.number_format = "#,##0.00"

    total_row = data_start + len(rows)
    ws.cell(row=total_row, column=1, value="TOTAL")
    total_values = (
        None,
        None,
        None,
        None,
        _whole_or_float(totals.get("room_rent")),
        _whole_or_float(totals.get("taxable")),
        _whole_or_float(totals.get("cgst")),
        _whole_or_float(totals.get("ugst")),
        _whole_or_float(totals.get("total_gst")),
        _whole_or_float(totals.get("total")),
        _whole_or_float(totals.get("round_off")),
        _whole_or_float(totals.get("final_amount")),
    )
    for col, value in enumerate(total_values, start=2):
        cell = ws.cell(row=total_row, column=col, value=value)
        if col in money_cols:
            cell.number_format = "#,##0.00"

    for col in range(1, col_count + 1):
        for ridx in (1, 2):
            cell = ws.cell(row=ridx, column=col)
            cell.fill = header_fill
            cell.font = title_font if ridx == 1 else period_font
            cell.alignment = center
            cell.border = grid
        header_cell = ws.cell(row=3, column=col)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = center
        header_cell.border = grid

    last_data = total_row - 1
    for ridx in range(data_start, last_data + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=ridx, column=col)
            cell.font = body_font
            cell.border = grid
            cell.alignment = right if col in money_cols else (center if col in (2, 4, 5) else left)

    for col in range(1, col_count + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = grid
        cell.alignment = right if col in money_cols else left

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 28
    col_widths = (36, 14, 28, 20, 20, 14, 16, 12, 12, 14, 14, 12, 16)
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(col_count)}{max(3, last_data)}"
    return wb


def _hub_kwargs():
    from_hub = (request.args.get("from_hub") or "").strip().lower()
    if from_hub != "reports":
        from_hub = ""
    kwargs = {}
    if from_hub:
        kwargs["from_hub"] = "reports"
    return from_hub, kwargs


def _load_page_payload():
    year, month, _start, _end = _period_from_args(request.args)
    conn = get_db()
    try:
        payload = build_gst_hotel_report(conn, year=year, month=month)
        conn.commit()
        return payload
    finally:
        conn.close()


def register_gst_hotel(app):
    """Register GST Hotel page and Excel export. Login is enforced by app.before_request."""

    @app.route("/reports/gst/hotel", endpoint="gst_hotel_report")
    def gst_hotel_report():
        payload = _load_page_payload()
        from_hub, hub_kwargs = _hub_kwargs()
        export_kwargs = dict(hub_kwargs)
        export_kwargs["month"] = payload["month"]
        export_kwargs["year"] = payload["year"]
        return render_template(
            "gst_hotel_report.html",
            de_nav_section="report",
            de_nav_report_view="home",
            page_title="GST Hotel",
            rows=payload["rows"],
            kpis=payload["kpis"],
            totals=payload["totals"],
            sel_year=payload["year"],
            sel_month=payload["month"],
            month_name=payload["month_name"],
            today_year=payload["today"].year,
            period_label=payload["period_label"],
            date_from=payload["date_from"].isoformat(),
            date_to=payload["date_to"].isoformat(),
            filter_form_action=url_for("gst_hotel_report", **hub_kwargs),
            gst_hotel_export_url=url_for("gst_hotel_report_export", **export_kwargs),
            gst_hotel_export_filename=report_export_month_filename(
                GST_HOTEL_EXPORT_TITLE, payload["year"], payload["month"]
            ),
            preserve_from_hub=bool(from_hub),
            from_hub=from_hub,
            back_href=url_for("reports") if from_hub else None,
            back_label="Back to Reports" if from_hub else None,
        )

    @app.route("/reports/gst/hotel.xlsx", endpoint="gst_hotel_report_export")
    def gst_hotel_report_export():
        payload = _load_page_payload()
        wb = build_gst_hotel_workbook(payload)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = report_export_month_filename(
            GST_HOTEL_EXPORT_TITLE, payload["year"], payload["month"]
        )
        response = send_file(
            buf,
            as_attachment=True,
            download_name=fname,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        return response
