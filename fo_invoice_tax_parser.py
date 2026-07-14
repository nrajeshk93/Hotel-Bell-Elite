"""Parse FO Invoice Tax Excel reports into hotel sales ledger line items."""

from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta
from typing import BinaryIO

import openpyxl


PAYMENT_MODE_MAP = {
    "cash": "cash",
    "cash payment": "cash",
    "credit card": "card",
    "card": "card",
    "upi": "upi",
}

# Report - FO Invoices Tax column mapping (0-based row indices).
COL_TAXABLE_AMOUNT = 10  # K
COL_TOTAL_TAX = 14       # O
COL_PAY_MODES = 20       # U


def parse_amount(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    return round(float(text), 2)


def parse_report_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial dates use 1899-12-30 as the practical epoch.
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def normalize_payment_mode(value) -> str:
    mode = str(value or "").strip().lower()
    return PAYMENT_MODE_MAP.get(mode, "room_credit")


def _is_header_or_total(value) -> bool:
    text = str(value or "").strip().lower()
    return not text or text == "date" or text == "total" or text.startswith("report - ")


def parse_fo_invoice_tax_report(file_obj: BinaryIO) -> dict:
    """Return hotel ledger lines grouped by the dates in column A."""
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    try:
        ws = wb["Report - FO Invoices Tax"] if "Report - FO Invoices Tax" in wb.sheetnames else wb.active
        lines = []
        counts_by_date = defaultdict(int)

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            if not row or len(row) <= COL_PAY_MODES:
                continue
            if _is_header_or_total(row[0]):
                continue

            sales_date = parse_report_date(row[0])
            invoice_number = str(row[1] or "").strip()
            taxable_amount = parse_amount(row[COL_TAXABLE_AMOUNT])
            total_tax = parse_amount(row[COL_TOTAL_TAX])
            amount = round(taxable_amount + total_tax, 2)
            if not sales_date or not invoice_number or amount <= 0:
                continue

            sales_date_key = sales_date.isoformat()
            counts_by_date[sales_date_key] += 1
            lines.append({
                "sales_date": sales_date_key,
                "invoice_number": invoice_number,
                "room": invoice_number,
                "reserve_number": str(row[2] or "").strip(),
                "guest_name": str(row[3] or "").strip(),
                "company_name": str(row[4] or "").strip(),
                "room_type": "",
                "travel_agent": "",
                "pax": "",
                "room_plan": "",
                "tariff": taxable_amount,
                "discount": 0,
                "extra_amount": total_tax,
                "amount": amount,
                "payment_mode": normalize_payment_mode(row[COL_PAY_MODES]),
                "report_payment_mode": str(row[COL_PAY_MODES] or "").strip(),
                "sort_order": counts_by_date[sales_date_key],
                "source_row": row_idx,
            })

        imported_dates = sorted(counts_by_date)
        return {
            "lines": lines,
            "lines_by_date": {
                sales_date: [line for line in lines if line["sales_date"] == sales_date]
                for sales_date in imported_dates
            },
            "meta": {
                "sheet": ws.title,
                "line_count": len(lines),
                "imported_dates": imported_dates,
                "counts_by_date": dict(counts_by_date),
                "total_amount": round(sum(line["amount"] for line in lines), 2),
            },
        }
    finally:
        wb.close()
