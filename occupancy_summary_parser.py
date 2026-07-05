"""Parse occupancy summary Excel reports into hotel sales ledger line items."""

from __future__ import annotations

import re
from typing import BinaryIO

import openpyxl

SKIP_ROOM_LABELS = frozenset({
    "room count",
    "total",
    "grand total",
    "revenue",
    "room sold",
    "average room rate",
    "rev par",
    "occupancy %",
    "revenue summary",
    "today",
})
FOOTER_STOP_LABELS = frozenset({"room count", "revenue summary"})


def parse_amount(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    return round(float(text), 2)


def _is_data_row(room_value) -> bool:
    if room_value is None:
        return False
    text = str(room_value).strip()
    if not text:
        return False
    lowered = text.lower()
    if lowered in SKIP_ROOM_LABELS:
        return False
    if lowered.startswith("room count"):
        return False
    if not re.match(r"^[A-Za-z0-9\-]+$", text):
        return False
    return bool(re.search(r"\d", text))


def parse_occupancy_summary_report(file_obj: BinaryIO) -> dict:
    """Return ledger lines from Report - Occupancy Summary sheet."""
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    try:
        ws = wb["Report - Occupancy Summary"] if "Report - Occupancy Summary" in wb.sheetnames else wb.active
        lines = []
        sort_order = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 9:
                continue
            room = row[0]
            room_text = str(room).strip()
            room_lower = room_text.lower()
            if room_lower in FOOTER_STOP_LABELS or room_lower.startswith("room count"):
                break
            if not _is_data_row(room):
                continue
            tariff = parse_amount(row[8])
            discount = parse_amount(row[9]) if len(row) > 9 else 0.0
            extra = parse_amount(row[10]) if len(row) > 10 else 0.0
            amount = round(tariff - discount + extra, 2)
            sort_order += 1
            lines.append({
                "room": str(room).strip(),
                "room_type": str(row[1] or "").strip(),
                "reserve_number": str(row[2] or "").strip(),
                "guest_name": str(row[3] or "").strip(),
                "company_name": str(row[4] or "").strip(),
                "travel_agent": str(row[5] or "").strip(),
                "pax": str(row[6] or "").strip(),
                "room_plan": str(row[7] or "").strip(),
                "tariff": tariff,
                "discount": discount,
                "extra_amount": extra,
                "amount": amount,
                "payment_mode": "",
                "sort_order": sort_order,
                "source_row": row_idx,
            })
        return {
            "lines": lines,
            "meta": {
                "sheet": ws.title,
                "line_count": len(lines),
                "total_amount": round(sum(line["amount"] for line in lines), 2),
            },
        }
    finally:
        wb.close()
