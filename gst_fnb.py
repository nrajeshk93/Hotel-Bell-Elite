"""GST report — Bar and Restaurant POS invoices (month window)."""

from __future__ import annotations

import calendar
from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO

from flask import render_template, request, send_file, url_for

from db import (
    POS_DEFAULT_CGST_PCT,
    POS_DEFAULT_UGST_PCT,
    POS_DEFAULT_VAT_PCT,
    POS_OUTLET_BAR,
    _normalize_pos_payment_method,
    _pos_is_official_order_no,
    ensure_pos_schema,
    get_db,
    get_pos_tax_rates,
    is_pos_liquor_category,
    normalize_pos_outlet,
)
from reports import report_export_month_filename

PAGE_TITLE = "GST — Bar and Restaurant"
EXCEL_TITLE = "HOTEL BELL ELITE - GST REPORT - BAR AND RESTAURANT"
EXPORT_REPORT_TITLE = "GST Restaurant and Bar"
EXCEL_SHEET_NAME = "Bar and Restaurant"

_COLUMNS = (
    "INVOICE NUMBER",
    "INVOICE DATE",
    "LIQUOR",
    "VAT",
    "RESTAURANT",
    "GST",
    "TOTAL",
    "SWIGGY",
    "ZOMATO",
)


def _money(value, default=0.0):
    try:
        quantum = Decimal("0.01")
        return float(Decimal(str(value if value is not None else default)).quantize(
            quantum, rounding=ROUND_HALF_UP
        ))
    except (InvalidOperation, TypeError, ValueError):
        return 0.0


def _parse_period(args):
    """month + year query params; default current calendar month."""
    today = date.today()
    try:
        year = int(args.get("year") or today.year)
    except (TypeError, ValueError):
        year = today.year
    try:
        month = int(args.get("month") or today.month)
    except (TypeError, ValueError):
        month = today.month
    if month < 1 or month > 12:
        month = today.month
    if year < 2000 or year > today.year + 8:
        year = today.year
    start = date(year, month, 1)
    end = date(year, month, calendar.monthrange(year, month)[1])
    return year, month, start, end


def _period_from_to_label(start, end):
    return f"FROM {start.strftime('%d-%m-%Y')} TO {end.strftime('%d-%m-%Y')}"


def _from_hub(args):
    value = (args.get("from_hub") or "").strip().lower()
    return "reports" if value == "reports" else ""


def _line_is_vat_liquor(line, invoice_outlet):
    """Bar liquor lines attract VAT; same rule as POS invoice recompute."""
    kind = str(line.get("item_kind") or "").strip().lower()
    if kind in ("liquour", "alcohol", "bar"):
        kind = "liquor"
    menu_type = str(line.get("menu_type") or "").strip().lower()
    if menu_type in ("liquour", "alcohol"):
        menu_type = "liquor"
    cat = line.get("category_name") or line.get("variant") or ""
    name = line.get("name") or ""
    is_liquor = (
        kind == "liquor"
        or menu_type == "liquor"
        or is_pos_liquor_category(cat)
        or is_pos_liquor_category(name)
    )
    raw_outlet = line.get("menu_outlet")
    if raw_outlet:
        return normalize_pos_outlet(raw_outlet) == POS_OUTLET_BAR and is_liquor
    return normalize_pos_outlet(invoice_outlet) == POS_OUTLET_BAR and is_liquor


def _chunked(values, size=400):
    seq = list(values)
    for i in range(0, len(seq), size):
        yield seq[i : i + size]


def _load_line_buckets(conn, invoice_ids):
    """invoice_id -> {liquor_incl, food_incl} from POS lines + menu category."""
    buckets = {int(i): {"liquor": 0.0, "food": 0.0} for i in invoice_ids}
    if not invoice_ids:
        return buckets
    for chunk in _chunked(invoice_ids):
        placeholders = ",".join("?" for _ in chunk)
        rows = conn.execute(
            f"""
            SELECT
                l.invoice_id,
                l.name,
                l.variant,
                l.rate,
                l.qty,
                l.line_total,
                i.outlet AS invoice_outlet,
                m.item_kind,
                m.menu_type,
                m.outlet AS menu_outlet,
                c.name AS category_name
            FROM pos_invoice_lines l
            JOIN pos_invoices i ON i.id = l.invoice_id
            LEFT JOIN pos_menu_items m ON m.id = l.menu_item_id
            LEFT JOIN pos_menu_categories c ON c.id = m.category_id
            WHERE l.invoice_id IN ({placeholders})
            """,
            chunk,
        ).fetchall()
        for row in rows:
            inv_id = int(row["invoice_id"])
            line_total = _money(row["line_total"])
            if abs(line_total) < 0.0005:
                line_total = _money(_money(row["rate"]) * _money(row["qty"]))
            line = {
                "name": row["name"] or "",
                "variant": row["variant"] or "",
                "item_kind": row["item_kind"] if "item_kind" in row.keys() else "",
                "menu_type": row["menu_type"] if "menu_type" in row.keys() else "",
                "menu_outlet": row["menu_outlet"] if "menu_outlet" in row.keys() else "",
                "category_name": row["category_name"] if "category_name" in row.keys() else "",
            }
            outlet = row["invoice_outlet"] if "invoice_outlet" in row.keys() else ""
            key = "liquor" if _line_is_vat_liquor(line, outlet) else "food"
            buckets[inv_id][key] = _money(buckets[inv_id][key] + line_total)
    return buckets


def _load_aggregator_flags(conn, invoice_ids):
    """invoice_id -> {swiggy, zomato} from stored payment methods."""
    flags = {int(i): {"swiggy": False, "zomato": False} for i in invoice_ids}
    if not invoice_ids:
        return flags
    for chunk in _chunked(invoice_ids):
        placeholders = ",".join("?" for _ in chunk)
        rows = conn.execute(
            f"""
            SELECT invoice_id, payment_method, amount
            FROM pos_invoice_payments
            WHERE invoice_id IN ({placeholders})
            """,
            chunk,
        ).fetchall()
        for row in rows:
            if abs(_money(row["amount"])) < 0.0005:
                continue
            inv_id = int(row["invoice_id"])
            key = _normalize_pos_payment_method(row["payment_method"])
            if key in ("swiggy", "zomato") and inv_id in flags:
                flags[inv_id][key] = True
    return flags


def _gst_frac_for_invoice(inv, rates):
    cgst = inv.get("tax_cgst_pct")
    ugst = inv.get("tax_ugst_pct")
    if cgst is not None or ugst is not None:
        cgst_pct = float(cgst) if cgst is not None else rates.get("cgst_pct", POS_DEFAULT_CGST_PCT)
        ugst_pct = float(ugst) if ugst is not None else rates.get("ugst_pct", POS_DEFAULT_UGST_PCT)
        return max(0.0, (cgst_pct + ugst_pct) / 100.0)
    return float(rates.get("cgst") or 0) + float(rates.get("ugst") or 0)


def _vat_frac_for_invoice(rates):
    return float(rates.get("vat") or 0) or (POS_DEFAULT_VAT_PCT / 100.0)


def _split_exclusive(inv, bucket, rates):
    """Liquor/restaurant exclusive values; VAT/GST from stored invoice tax."""
    vat = _money(inv.get("vat_amount"))
    gst = _money(inv.get("gst_amount"))
    subtotal = _money(inv.get("subtotal"))
    discount = _money(inv.get("discount_amount"))
    net_exclusive = _money(max(0.0, subtotal - discount))
    vat_frac = _vat_frac_for_invoice(rates)
    gst_frac = _gst_frac_for_invoice(inv, rates)

    liquor = _money(vat / vat_frac) if vat > 0.0005 and vat_frac > 0 else 0.0
    restaurant = _money(gst / gst_frac) if gst > 0.0005 and gst_frac > 0 else 0.0

    remainder = _money(net_exclusive - liquor - restaurant)
    liquor_incl = _money((bucket or {}).get("liquor"))
    food_incl = _money((bucket or {}).get("food"))
    line_sum = liquor_incl + food_incl
    if remainder > 0.004:
        if line_sum > 0.004:
            liquor_part = _money(remainder * (liquor_incl / line_sum))
            liquor = _money(liquor + liquor_part)
            restaurant = _money(restaurant + remainder - liquor_part)
        elif liquor_incl >= food_incl and liquor_incl > 0:
            liquor = _money(liquor + remainder)
        else:
            restaurant = _money(restaurant + remainder)
    elif remainder < -0.004:
        # Rounding vs stored tax: keep tax columns, trim the larger exclusive base.
        extra = _money(-remainder)
        if restaurant >= extra:
            restaurant = _money(restaurant - extra)
        elif liquor >= extra:
            liquor = _money(liquor - extra)

    return liquor, restaurant, vat, gst


def load_gst_fnb_rows(conn, *, date_from, date_to):
    """Official Restaurant + Bar POS invoices in [date_from, date_to], not cancelled."""
    ensure_pos_schema(conn)
    start = date_from.isoformat() if hasattr(date_from, "isoformat") else str(date_from)
    end = date_to.isoformat() if hasattr(date_to, "isoformat") else str(date_to)
    invoice_rows = conn.execute(
        """
        SELECT
            id, order_no, order_date, outlet, status,
            gst_amount, vat_amount, subtotal, discount_amount, grand_total,
            tax_cgst_pct, tax_ugst_pct
        FROM pos_invoices
        WHERE is_active = 1
          AND lower(COALESCE(status, 'open')) != 'cancelled'
          AND TRIM(COALESCE(cancelled_at, '')) = ''
          AND order_date >= ?
          AND order_date <= ?
        ORDER BY order_date ASC, order_no ASC, id ASC
        """,
        (start, end),
    ).fetchall()

    invoices = []
    for row in invoice_rows:
        order_no = str(row["order_no"] or "").strip()
        outlet = normalize_pos_outlet(row["outlet"] if "outlet" in row.keys() else None)
        if not _pos_is_official_order_no(order_no, outlet):
            continue
        cgst_raw = row["tax_cgst_pct"] if "tax_cgst_pct" in row.keys() else None
        ugst_raw = row["tax_ugst_pct"] if "tax_ugst_pct" in row.keys() else None
        invoices.append(
            {
                "id": int(row["id"]),
                "order_no": order_no,
                "order_date": str(row["order_date"] or "")[:10],
                "outlet": outlet,
                "gst_amount": row["gst_amount"],
                "vat_amount": row["vat_amount"],
                "subtotal": row["subtotal"],
                "discount_amount": row["discount_amount"],
                "grand_total": row["grand_total"],
                "tax_cgst_pct": cgst_raw,
                "tax_ugst_pct": ugst_raw,
            }
        )

    ids = [inv["id"] for inv in invoices]
    buckets = _load_line_buckets(conn, ids)
    aggregators = _load_aggregator_flags(conn, ids)
    rates_by_outlet = {
        "restaurant": get_pos_tax_rates(conn, "restaurant"),
        POS_OUTLET_BAR: get_pos_tax_rates(conn, POS_OUTLET_BAR),
    }

    rows = []
    for inv in invoices:
        rates = rates_by_outlet.get(inv["outlet"]) or rates_by_outlet["restaurant"]
        liquor, restaurant, vat, gst = _split_exclusive(
            inv, buckets.get(inv["id"]), rates
        )
        total = _money(inv.get("grand_total"))
        flags = aggregators.get(inv["id"]) or {}
        swiggy = total if flags.get("swiggy") else 0.0
        zomato = total if flags.get("zomato") else 0.0
        order_date = inv["order_date"]
        try:
            date_obj = date.fromisoformat(order_date)
            date_display = date_obj.strftime("%d-%m-%Y")
        except ValueError:
            date_obj = None
            date_display = order_date
        rows.append(
            {
                "id": inv["id"],
                "invoice_number": inv["order_no"],
                "invoice_date": order_date,
                "invoice_date_display": date_display,
                "invoice_date_obj": date_obj,
                "liquor": liquor,
                "vat": vat,
                "restaurant": restaurant,
                "gst": gst,
                "total": total,
                "swiggy": swiggy,
                "zomato": zomato,
                "outlet": inv["outlet"],
            }
        )
    return rows


def summarize_gst_fnb(rows):
    kpis = {
        "invoice_count": len(rows or []),
        "liquor": 0.0,
        "vat": 0.0,
        "restaurant": 0.0,
        "gst": 0.0,
        "total": 0.0,
        "swiggy": 0.0,
        "zomato": 0.0,
    }
    for row in rows or []:
        for key in ("liquor", "vat", "restaurant", "gst", "total", "swiggy", "zomato"):
            kpis[key] = _money(kpis[key] + _money(row.get(key)))
    return kpis


def build_gst_fnb_report(conn, *, year, month):
    start = date(int(year), int(month), 1)
    end = date(int(year), int(month), calendar.monthrange(int(year), int(month))[1])
    rows = load_gst_fnb_rows(conn, date_from=start, date_to=end)
    return {
        "rows": rows,
        "kpis": summarize_gst_fnb(rows),
        "year": int(year),
        "month": int(month),
        "date_from": start,
        "date_to": end,
        "period_label": _period_from_to_label(start, end),
        "month_name": calendar.month_name[int(month)],
    }


def _excel_number(value):
    num = _money(value)
    if abs(num - int(num)) < 0.0005:
        return int(num)
    return num


def build_gst_fnb_workbook(payload):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    rows = payload.get("rows") or []
    kpis = payload.get("kpis") or {}
    period_label = payload.get("period_label") or ""

    wb = Workbook()
    ws = wb.active
    ws.title = EXCEL_SHEET_NAME

    header_fill = PatternFill(
        fill_type="solid", start_color="FF315A78", end_color="FF315A78"
    )
    total_fill = PatternFill(
        fill_type="solid", start_color="FFE8EEF2", end_color="FFE8EEF2"
    )
    title_font = Font(name="Calibri", bold=True, size=14, color="FF000000")
    period_font = Font(name="Calibri", bold=True, size=11, color="FF000000")
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFFFF")
    body_font = Font(name="Calibri", size=11, color="FF000000")
    total_font = Font(name="Calibri", bold=True, size=11, color="FF000000")
    thin = Side(style="thin", color="FF000000")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws["A1"] = EXCEL_TITLE
    ws["A1"].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_COLUMNS))
    ws["A2"] = period_label
    ws["A2"].font = period_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(_COLUMNS))

    header_row = 4
    for col, title in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = grid
        cell.alignment = center

    money_cols = {3, 4, 5, 6, 7, 8, 9}
    for idx, entry in enumerate(rows, start=header_row + 1):
        values = (
            entry.get("invoice_number") or "",
            entry.get("invoice_date_display") or entry.get("invoice_date") or "",
            _excel_number(entry.get("liquor")),
            _excel_number(entry.get("vat")),
            _excel_number(entry.get("restaurant")),
            _excel_number(entry.get("gst")),
            _excel_number(entry.get("total")),
            _excel_number(entry.get("swiggy")),
            _excel_number(entry.get("zomato")),
        )
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col, value=value)
            cell.font = body_font
            cell.border = grid
            cell.alignment = right if col in money_cols else left
            if col in money_cols:
                cell.number_format = "#,##0.00"

    total_row = header_row + 1 + len(rows)
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=2, value="")
    totals = (
        None,
        None,
        kpis.get("liquor"),
        kpis.get("vat"),
        kpis.get("restaurant"),
        kpis.get("gst"),
        kpis.get("total"),
        kpis.get("swiggy"),
        kpis.get("zomato"),
    )
    for col, value in enumerate(totals, start=1):
        cell = ws.cell(row=total_row, column=col)
        if value is not None:
            cell.value = _excel_number(value)
            cell.number_format = "#,##0.00"
            cell.alignment = right
        else:
            cell.alignment = left
        cell.font = total_font
        cell.fill = total_fill
        cell.border = grid

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[header_row].height = 20
    widths = (22, 16, 14, 12, 16, 12, 14, 14, 14)
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(_COLUMNS))}{max(header_row, total_row - 1)}"
    return wb


def gst_fnb_excel_bytes(payload):
    wb = build_gst_fnb_workbook(payload)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _report_context(payload, *, from_hub=""):
    year = payload["year"]
    month = payload["month"]
    filter_kwargs = {}
    if from_hub:
        filter_kwargs["from_hub"] = from_hub
    export_kwargs = {"year": year, "month": month}
    return {
        "page_title": PAGE_TITLE,
        "rows": payload["rows"],
        "kpis": payload["kpis"],
        "sel_year": year,
        "sel_month": month,
        "month_name": payload["month_name"],
        "period_label": payload["period_label"],
        "today_year": date.today().year,
        "filter_form_action": url_for("gst_fnb_report", **filter_kwargs),
        "export_url": url_for("gst_fnb_report_export", **export_kwargs),
        "export_filename": report_export_month_filename(
            EXPORT_REPORT_TITLE, year, month
        ),
        "from_hub": from_hub,
        "back_href": url_for("reports") if from_hub == "reports" else None,
        "back_label": "Back to Reports" if from_hub == "reports" else None,
        "de_nav_section": "report",
        "de_nav_report_view": "home",
        "de_nav_host": "report",
    }


def gst_fnb_report():
    """GST Bar and Restaurant report — invoice-wise liquor/VAT and food/GST."""
    year, month, _start, _end = _parse_period(request.args)
    from_hub = _from_hub(request.args)
    conn = get_db()
    try:
        payload = build_gst_fnb_report(conn, year=year, month=month)
    finally:
        conn.close()
    return render_template("gst_fnb_report.html", **_report_context(payload, from_hub=from_hub))


def gst_fnb_report_export():
    """Excel export for GST Bar and Restaurant report."""
    year, month, _start, _end = _parse_period(request.args)
    conn = get_db()
    try:
        payload = build_gst_fnb_report(conn, year=year, month=month)
    finally:
        conn.close()
    buf = gst_fnb_excel_bytes(payload)
    fname = report_export_month_filename(EXPORT_REPORT_TITLE, year, month)
    response = send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response


def register_gst_fnb(app):
    """Wire GST F&B report + Excel export. Called from app.py by the wiring agent."""
    app.add_url_rule(
        "/reports/gst/restaurant-bar",
        endpoint="gst_fnb_report",
        view_func=gst_fnb_report,
        methods=["GET"],
    )
    app.add_url_rule(
        "/reports/gst/restaurant-bar.xlsx",
        endpoint="gst_fnb_report_export",
        view_func=gst_fnb_report_export,
        methods=["GET"],
    )
