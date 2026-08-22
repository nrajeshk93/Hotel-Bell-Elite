"""Tests for Back Office Receipt ledger."""

import os
import tempfile
import unittest
from datetime import date
from io import BytesIO
from unittest import mock

import db as db_mod
from back_office_receipt import (
    amount_in_indian_words,
    create_back_office_receipt,
    insert_back_office_receipt_allocations,
    list_back_office_receipt_filter_agencies,
    list_back_office_receipt_ledger_entries,
    list_back_office_receipts,
    update_back_office_receipt,
)
from workspace_access import (
    get_endpoint_accounts_submodule,
    get_endpoint_dashboard_module,
)


class AmountInWordsTests(unittest.TestCase):
    def test_basic_rupees(self):
        self.assertEqual(amount_in_indian_words(1500), "Rupees One Thousand Five Hundred Only")

    def test_zero(self):
        self.assertEqual(amount_in_indian_words(0), "Rupees Zero Only")

    def test_paise(self):
        self.assertEqual(
            amount_in_indian_words(10.25),
            "Rupees Ten and Twenty Five Paise Only",
        )


class BackOfficeReceiptDbTests(unittest.TestCase):
    def setUp(self):
        import sqlite3

        self.conn = sqlite3.connect(":memory:")
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys=ON")
        db_mod.ensure_agencies_schema(self.conn)
        db_mod.ensure_back_office_receipt_schema(self.conn)
        cur = self.conn.execute(
            "INSERT INTO agencies (name, gst) VALUES (?, ?)",
            ("MakeMyTrip", "29AAAAA0000A1Z5"),
        )
        self.agency_id = cur.lastrowid

    def tearDown(self):
        self.conn.close()

    def test_allocate_receipt_no_fiscal_format(self):
        short_fy, seq, receipt_no = db_mod.allocate_back_office_receipt_no(
            self.conn, date(2026, 8, 21)
        )
        self.assertEqual(short_fy, "26-27")
        self.assertEqual(seq, 1)
        self.assertEqual(receipt_no, "HBE/BOR/26-27/1")
        _, seq2, no2 = db_mod.allocate_back_office_receipt_no(self.conn, date(2026, 8, 21))
        self.assertEqual(seq2, 2)
        self.assertEqual(no2, "HBE/BOR/26-27/2")

    def test_create_with_agency(self):
        row = create_back_office_receipt(
            self.conn,
            receipt_date=date(2026, 8, 21),
            payer_name="",
            agency_id=self.agency_id,
            amount=5550,
            payment_mode="cash",
            towards="Room advance",
        )
        self.conn.commit()
        self.assertEqual(row["payer_name"], "MakeMyTrip")
        self.assertEqual(row["receipt_no"], "HBE/BOR/26-27/1")
        self.assertTrue(row["amount_words"].startswith("Rupees"))
        self.assertEqual(row["payment_mode"], "cash")

    def test_ledger_shows_received_and_applied(self):
        receipt = create_back_office_receipt(
            self.conn,
            receipt_date=date(2026, 8, 20),
            payer_name="MakeMyTrip",
            agency_id=self.agency_id,
            amount=10000,
            payment_mode="cash",
            towards="Advance",
        )
        db_mod.ensure_hotel_invoice_credits_schema(self.conn)
        payment_id = self.conn.execute(
            """INSERT INTO hotel_invoice_credit_payments
               (company, agency_name, payment_date, payment_method, transaction_id, total_amount, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            ("Hotel Bell Elite", "MakeMyTrip", "2026-08-21", "bor", "", 4000, ""),
        ).lastrowid
        insert_back_office_receipt_allocations(
            self.conn,
            payment_id,
            [{"receipt_id": receipt["id"], "amount": 4000}],
        )
        self.conn.commit()
        entries = list_back_office_receipt_ledger_entries(
            self.conn, date(2026, 8, 1), date(2026, 8, 31)
        )
        self.assertEqual(len(entries), 2)
        # Newest-first display order
        applied = next(e for e in entries if e["entry_type"] == "applied")
        received = next(e for e in entries if e["entry_type"] == "receipt")
        self.assertAlmostEqual(applied["signed_amount"], -4000, places=2)
        self.assertAlmostEqual(received["signed_amount"], 10000, places=2)
        self.assertIn("Hotel credit", applied["detail"])
        # Chronological balance ends at 6000 on newest applied row when display is reversed
        chronological = list(reversed(entries))
        self.assertAlmostEqual(chronological[0]["running_balance"], 10000, places=2)
        self.assertAlmostEqual(chronological[1]["running_balance"], 6000, places=2)

    def test_create_free_text_payer(self):
        row = create_back_office_receipt(
            self.conn,
            receipt_date=date(2026, 8, 21),
            payer_name="Walk-in Guest",
            agency_id=None,
            amount=100,
            payment_mode="bank_transfer",
            instrument_no="UTR123",
            instrument_date=date(2026, 8, 20),
            towards="Banquet",
        )
        self.assertEqual(row["payer_name"], "Walk-in Guest")
        self.assertIsNone(row["agency_id"])
        self.assertEqual(row["instrument_no"], "UTR123")

    def test_update_preserves_receipt_no(self):
        created = create_back_office_receipt(
            self.conn,
            receipt_date=date(2026, 8, 21),
            payer_name="Walk-in",
            agency_id=None,
            amount=100,
            payment_mode="cash",
            towards="Advance",
        )
        self.conn.commit()
        updated = update_back_office_receipt(
            self.conn,
            created["id"],
            receipt_date=date(2026, 8, 19),
            payer_name="Walk-in Updated",
            agency_id=None,
            amount=250,
            payment_mode="cash",
            towards="Room",
        )
        self.conn.commit()
        self.assertEqual(updated["receipt_no"], created["receipt_no"])
        self.assertEqual(updated["payer_name"], "Walk-in Updated")
        self.assertEqual(updated["amount"], 250)
        self.assertEqual(updated["receipt_date"], "2026-08-19")

    def test_reject_invalid_amount_and_missing_payer(self):
        with self.assertRaises(ValueError):
            create_back_office_receipt(
                self.conn,
                receipt_date=date(2026, 8, 21),
                payer_name="",
                agency_id=None,
                amount=10,
                payment_mode="cash",
            )
        with self.assertRaises(ValueError):
            create_back_office_receipt(
                self.conn,
                receipt_date=date(2026, 8, 21),
                payer_name="Someone",
                agency_id=None,
                amount=0,
                payment_mode="cash",
            )

    def test_list_filters_by_date(self):
        create_back_office_receipt(
            self.conn,
            receipt_date=date(2026, 8, 10),
            payer_name="A",
            agency_id=None,
            amount=10,
            payment_mode="cash",
        )
        create_back_office_receipt(
            self.conn,
            receipt_date=date(2026, 8, 20),
            payer_name="B",
            agency_id=None,
            amount=20,
            payment_mode="cash",
        )
        self.conn.commit()
        rows = list_back_office_receipts(self.conn, date(2026, 8, 15), date(2026, 8, 31))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["payer_name"], "B")

    def test_list_filters_by_agency(self):
        create_back_office_receipt(
            self.conn,
            receipt_date=date(2026, 8, 10),
            payer_name="Goibibo",
            agency_id=self.agency_id,
            amount=100,
            payment_mode="cash",
        )
        create_back_office_receipt(
            self.conn,
            receipt_date=date(2026, 8, 11),
            payer_name="Walk-in",
            agency_id=None,
            amount=50,
            payment_mode="cash",
        )
        self.conn.commit()
        rows = list_back_office_receipts(
            self.conn, date(2026, 8, 1), date(2026, 8, 31), agency_id=self.agency_id
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["payer_name"], "Goibibo")
        entries = list_back_office_receipt_ledger_entries(
            self.conn,
            date(2026, 8, 1),
            date(2026, 8, 31),
            agency_id=self.agency_id,
        )
        receipt_entries = [e for e in entries if e["entry_type"] == "receipt"]
        self.assertEqual(len(receipt_entries), 1)
        self.assertEqual(receipt_entries[0]["party_name"], "Goibibo")


class BackOfficeReceiptRouteTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db_path = self.tmp.name
        self._orig_path = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = self.db_path
        db_mod.init_db()

        import app as app_mod

        self.app_mod = app_mod
        self.app = app_mod.app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

        conn = db_mod.get_db()
        try:
            admin = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            self.admin_id = admin["id"]
            cur = conn.execute(
                "INSERT INTO agencies (name, gst) VALUES (?, ?)",
                ("Goibibo", "29BBBBB0000B1Z5"),
            )
            self.agency_id = cur.lastrowid
            conn.commit()
        finally:
            conn.close()

        self.user = {
            "id": self.admin_id,
            "username": "admin",
            "full_name": "Administrator",
            "is_admin": True,
            "dashboard_access": set(),
            "accounts_access": set(),
        }
        self._user_patch = mock.patch.object(
            self.app_mod, "get_current_user", return_value=self.user
        )
        self._user_patch.start()

    def tearDown(self):
        self._user_patch.stop()
        db_mod.DATABASE_PATH = self._orig_path
        try:
            os.unlink(self.db_path)
        except OSError:
            pass

    def test_page_renders_marker_and_nav(self):
        resp = self.client.get("/accounts/back-office-receipt")
        self.assertEqual(resp.status_code, 200)
        html = resp.get_data(as_text=True)
        self.assertIn('id="back-office-receipt-page"', html)
        self.assertIn('id="de-nav-back-office-receipt"', html)
        self.assertIn("is-active", html)
        self.assertIn("Back Office Receipt", html)
        self.assertIn('id="bor-agency-filter-listbox"', html)
        self.assertIn("All agencies", html)
        # Agency master entries appear in Add modal, not the filter until used on a receipt.
        self.assertIn('id="bor-agencies-json"', html)

    def test_page_filters_by_agency(self):
        self.client.post(
            "/accounts/back-office-receipt/add",
            json={
                "receipt_date": "2026-08-10",
                "payer_name": "Goibibo",
                "agency_id": self.agency_id,
                "amount": 100,
                "payment_mode": "cash",
            },
        )
        self.client.post(
            "/accounts/back-office-receipt/add",
            json={
                "receipt_date": "2026-08-11",
                "payer_name": "Other Party",
                "amount": 50,
                "payment_mode": "cash",
            },
        )
        filtered = self.client.get(
            f"/accounts/back-office-receipt?date_from=2026-08-01&date_to=2026-08-31&agency={self.agency_id}"
        )
        self.assertEqual(filtered.status_code, 200)
        html = filtered.get_data(as_text=True)
        self.assertIn("Goibibo", html)
        self.assertNotIn("Other Party", html)
        self.assertIn(f'name="agency" value="{self.agency_id}"', html)
        # Filter dropdown only lists agencies present on the ledger.
        options_start = html.find('id="bor-agency-filter-options"')
        options_end = html.find("</div>", options_start)
        options_html = html[options_start:options_end]
        self.assertIn("Goibibo", options_html)
        self.assertIn(f'data-value="{self.agency_id}"', options_html)

    def test_filter_agencies_only_ledger_parties(self):
        conn = db_mod.get_db()
        try:
            cur = conn.execute(
                "INSERT INTO agencies (name, gst) VALUES (?, ?)",
                ("Unused Agency", "29UNUSED0000U1Z5"),
            )
            unused_id = cur.lastrowid
            conn.commit()
            self.assertTrue(unused_id)
        finally:
            conn.close()

        self.client.post(
            "/accounts/back-office-receipt/add",
            json={
                "receipt_date": "2026-08-10",
                "payer_name": "Goibibo",
                "agency_id": self.agency_id,
                "amount": 100,
                "payment_mode": "cash",
            },
        )
        page = self.client.get(
            "/accounts/back-office-receipt?date_from=2026-08-01&date_to=2026-08-31"
        )
        html = page.get_data(as_text=True)
        options_start = html.find('id="bor-agency-filter-options"')
        options_end = html.find('id="bor-agency-filter-empty"', options_start)
        options_html = html[options_start:options_end]
        self.assertIn("Goibibo", options_html)
        self.assertNotIn("Unused Agency", options_html)
        self.assertIn("Unused Agency", html)  # still available in Add receipt modal

        conn = db_mod.get_db()
        try:
            names = [
                a["name"]
                for a in list_back_office_receipt_filter_agencies(
                    conn, date(2026, 8, 1), date(2026, 8, 31)
                )
            ]
        finally:
            conn.close()
        self.assertEqual(names, ["Goibibo"])

    def test_edit_receipt(self):
        created = self.client.post(
            "/accounts/back-office-receipt/add",
            json={
                "receipt_date": "2026-08-21",
                "payer_name": "Walk-in",
                "amount": 100,
                "payment_mode": "cash",
                "towards": "Advance",
            },
        )
        receipt_id = created.get_json()["receipt"]["id"]
        receipt_no = created.get_json()["receipt"]["receipt_no"]
        edited = self.client.post(
            "/accounts/back-office-receipt/edit",
            json={
                "id": receipt_id,
                "receipt_date": "2026-08-18",
                "payer_name": "Walk-in Updated",
                "amount": 200,
                "payment_mode": "cash",
                "towards": "Room",
            },
        )
        self.assertEqual(edited.status_code, 200)
        payload = edited.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["receipt"]["receipt_no"], receipt_no)
        self.assertEqual(payload["receipt"]["payer_name"], "Walk-in Updated")
        self.assertEqual(payload["receipt"]["amount"], 200)

    def test_add_and_list_and_export(self):
        add = self.client.post(
            "/accounts/back-office-receipt/add",
            json={
                "receipt_date": "2026-08-21",
                "payer_name": "Goibibo",
                "agency_id": self.agency_id,
                "amount": 555,
                "payment_mode": "cheque",
                "instrument_no": "CHQ-9",
                "instrument_date": "2026-08-20",
                "towards": "Room stay",
            },
        )
        self.assertEqual(add.status_code, 200)
        payload = add.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["receipt"]["receipt_no"], "HBE/BOR/26-27/1")
        self.assertIn("Rupees", payload["receipt"]["amount_words"])

        page = self.client.get(
            "/accounts/back-office-receipt?date_from=2026-04-01&date_to=2026-08-21"
        )
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn("HBE/BOR/26-27/1", html)
        self.assertIn("Goibibo", html)

        export = self.client.get(
            "/accounts/back-office-receipt/report?date_from=2026-04-01&date_to=2026-08-21"
        )
        self.assertEqual(export.status_code, 200)
        self.assertIn(
            "spreadsheetml",
            export.headers.get("Content-Type", ""),
        )
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(export.data))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        self.assertEqual(headers[0], "Date")
        self.assertEqual(headers[1], "Type")
        self.assertEqual(headers[2], "Receipt No.")
        self.assertEqual(headers[7], "Amount")
        self.assertEqual(headers[8], "Balance")
        self.assertEqual(ws[2][1].value, "Received")
        self.assertEqual(ws[2][2].value, "HBE/BOR/26-27/1")

    def test_add_rejects_bad_payload(self):
        resp = self.client.post(
            "/accounts/back-office-receipt/add",
            json={"amount": 0, "payer_name": "", "payment_mode": "cash"},
        )
        self.assertEqual(resp.status_code, 400)
        self.assertFalse(resp.get_json()["ok"])


class BackOfficeReceiptAccessTests(unittest.TestCase):
    def test_endpoint_mapping(self):
        self.assertEqual(get_endpoint_dashboard_module("back_office_receipt"), "accounts")
        self.assertEqual(
            get_endpoint_dashboard_module("back_office_receipt_add"), "accounts"
        )
        self.assertEqual(
            get_endpoint_dashboard_module("export_back_office_receipt_report"),
            "accounts",
        )
        self.assertEqual(
            get_endpoint_accounts_submodule("back_office_receipt"),
            "back_office_receipt",
        )
        self.assertEqual(
            get_endpoint_accounts_submodule("back_office_receipt_delete"),
            "back_office_receipt",
        )
        self.assertEqual(
            get_endpoint_accounts_submodule("back_office_receipt_edit"),
            "back_office_receipt",
        )


if __name__ == "__main__":
    unittest.main()
