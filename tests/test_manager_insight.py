"""Manager Insight — occupancy math, fixed MTD/YTD, hub card, Excel."""

import json
import os
import re
import tempfile
import unittest
from datetime import date
from io import BytesIO
from unittest import mock

import db as db_mod
from manager_insight import (
    build_manager_insight,
    manager_insight_windows,
    metrics_for_window,
)
from reports import REPORT_DEFINITIONS
from workspace_access import get_endpoint_dashboard_module


FROZEN_TODAY = date(2026, 8, 13)


class ManagerInsightTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db_path = self.tmp.name
        self._orig_path = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = self.db_path
        db_mod.init_db()

        import app as app_mod

        self.app_mod = app_mod
        self.app = app_mod.app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

        conn = db_mod.get_db()
        try:
            admin = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            self.admin_id = admin["id"]
        finally:
            conn.close()

        self.user = {
            "id": self.admin_id,
            "username": "admin",
            "full_name": "Administrator",
            "is_admin": True,
            "is_active": True,
            "dashboard_access": set(),
            "stores_access": set(),
        }
        self._get_user_patch = mock.patch.object(
            app_mod, "get_current_user", return_value=self.user
        )
        self._get_user_patch.start()
        class FrozenDate(date):
            @classmethod
            def today(cls):
                return FROZEN_TODAY

        self._today_app = mock.patch.object(app_mod, "date", FrozenDate)
        self._today_app.start()

    def tearDown(self):
        self._today_app.stop()
        self._get_user_patch.stop()
        db_mod.DATABASE_PATH = self._orig_path
        try:
            os.unlink(self.db_path)
        except OSError:
            pass

    def _insert_stay(
        self,
        *,
        invoice_number="HBE/RM/1/2026-27",
        check_in="2026-08-12",
        check_out="2026-08-14",
        adults=2,
        children=0,
        extra_bed_qty=0,
        rate_plan="EP",
        total=3000,
        payments=None,
        room_number="101",
    ):
        conn = db_mod.get_db()
        try:
            db_mod.ensure_hotel_room_invoices_schema(conn)
            stay = {
                "guestName": "Test Guest",
                "invoiceNumber": invoice_number,
                "checkInDate": check_in,
                "checkOutDate": check_out,
                "adults": adults,
                "children": children,
                "extraBedQty": extra_bed_qty,
                "ratePlan": rate_plan,
                "estimatedTotal": float(total),
                "payments": payments
                if payments is not None
                else [{"method": "cash", "amount": float(total)}],
            }
            payload = {
                "id": f"room-{room_number}",
                "number": room_number,
                "stay": stay,
            }
            conn.execute(
                """
                INSERT INTO hotel_room_invoices (
                    invoice_number, room_id, room_number, room_type_label,
                    guest_name, booking_number, check_in_date, check_out_date,
                    invoice_generated_at, estimated_total, advance_paid,
                    balance_amount, status, payload_json
                ) VALUES (?, ?, ?, 'Deluxe', 'Test Guest', '', ?, ?,
                          ?, ?, 0, 0, 'settled', ?)
                """,
                (
                    invoice_number,
                    f"room-{room_number}",
                    room_number,
                    check_in,
                    check_out,
                    f"{check_out} 10:00:00",
                    float(total),
                    json.dumps(payload),
                ),
            )
            conn.commit()
        finally:
            conn.close()

    def _row_cells(self, html, key):
        match = re.search(
            rf'data-row-key="{re.escape(key)}"[^>]*>.*?</tr>',
            html,
            re.S,
        )
        self.assertIsNotNone(match, f"missing row {key}")
        cells = re.findall(r"<td[^>]*>(.*?)</td>", match.group(0), flags=re.S)
        texts = [re.sub(r"<[^>]+>", "", cell).strip() for cell in cells]
        return texts

    def test_hub_card_present(self):
        ids = [r["id"] for r in REPORT_DEFINITIONS if r.get("category") == "sales"]
        self.assertIn("manager_insight", ids)
        self.assertEqual(get_endpoint_dashboard_module("sales_report_manager_insight"), "reports")
        self.assertEqual(
            get_endpoint_dashboard_module("sales_report_manager_insight_export"),
            "reports",
        )
        hub = self.client.get("/reports")
        self.assertEqual(hub.status_code, 200)
        html = hub.get_data(as_text=True)
        self.assertIn('data-report-id="manager_insight"', html)
        self.assertIn("Manager Insight", html)
        self.assertIn("/reports/sales/manager-insight", html)

    def test_occupancy_math_one_room_two_nights(self):
        stay = {
            "check_in": date(2026, 8, 12),
            "check_out": date(2026, 8, 14),
            "pax": 2,
            "extra_person": 0,
            "plan": "EP",
            "room_units": 1,
            "revenue": 3000.0,
            "payments": {
                "bank": 0.0,
                "cash": 3000.0,
                "credit": 0.0,
                "card": 0.0,
                "upi": 0.0,
            },
        }
        metrics = metrics_for_window(
            [stay], 1, date(2026, 8, 12), date(2026, 8, 13)
        )
        self.assertEqual(metrics["total_rooms"], 2)
        self.assertEqual(metrics["rooms_sold"], 2)
        self.assertEqual(metrics["pax"], 4)
        self.assertEqual(metrics["double"], 2)
        self.assertEqual(metrics["pct_occupancy"], 100.0)
        self.assertEqual(metrics["ep_room"], 2)
        self.assertEqual(metrics["ep_pax"], 4)
        self.assertEqual(metrics["revenue"], 3000.0)
        self.assertEqual(metrics["arr"], 1500.0)
        self.assertEqual(metrics["arp"], 750.0)
        self.assertEqual(metrics["cash"], 3000.0)

    def test_stay_spanning_month_end_splits_nights(self):
        stay = {
            "check_in": date(2026, 7, 31),
            "check_out": date(2026, 8, 2),
            "pax": 1,
            "extra_person": 0,
            "plan": "CP",
            "room_units": 1,
            "revenue": 2000.0,
            "payments": {
                "bank": 0.0,
                "cash": 0.0,
                "credit": 2000.0,
                "card": 0.0,
                "upi": 0.0,
            },
        }
        duration = metrics_for_window(
            [stay], 1, date(2026, 7, 31), date(2026, 7, 31)
        )
        mtd = metrics_for_window([stay], 1, date(2026, 8, 1), date(2026, 8, 13))
        self.assertEqual(duration["rooms_sold"], 1)
        self.assertEqual(mtd["rooms_sold"], 1)
        self.assertEqual(duration["revenue"], 1000.0)
        self.assertEqual(mtd["revenue"], 1000.0)
        self.assertEqual(duration["cp_room"], 1)
        self.assertEqual(mtd["cp_room"], 1)

    def test_mtd_follows_selected_month(self):
        windows = manager_insight_windows(
            FROZEN_TODAY, date(2026, 4, 1), date(2026, 4, 1)
        )
        self.assertEqual(windows["duration"], (date(2026, 4, 1), date(2026, 4, 1)))
        self.assertEqual(windows["mtd"], (date(2026, 4, 1), date(2026, 4, 30)))
        self.assertEqual(windows["mtd_month_label"], "April Month")
        self.assertEqual(windows["ytd"], (date(2026, 4, 1), FROZEN_TODAY))

        current = manager_insight_windows(
            FROZEN_TODAY, date(2026, 8, 1), date(2026, 8, 13)
        )
        self.assertEqual(current["mtd"], (date(2026, 8, 1), FROZEN_TODAY))
        self.assertEqual(current["mtd_month_label"], "August Month")

    def test_page_headers_and_labels(self):
        page = self.client.get("/reports/sales/manager-insight?from_hub=reports")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn(">HEAD<", html)
        self.assertIn(">Duration<", html)
        self.assertIn(">August Month<", html)
        self.assertIn(">YTD<", html)
        self.assertIn("Today", html)
        self.assertIn("Yesterday", html)
        self.assertIn("7 Days", html)
        self.assertIn("30 Days", html)
        self.assertIn("QTD", html)
        self.assertIn('data-mi-period="mtd"', html)
        self.assertIn("is-active", html)
        for label in (
            "Total Rooms",
            "Rooms Sold",
            "Pax",
            "Extraperson",
            "Single Occupied",
            "Double Occupied",
            "Triple Occupied",
            "Quadruple Occupied",
            "% of Occupancy",
            "EP Room",
            "Total Room Revenue",
            "ARR",
            "ARP",
            "Bank Transfer",
            "Cash",
            "Credit",
            "Card",
            "UPI",
        ):
            self.assertIn(label, html)
        self.assertNotIn(">Tip<", html)
        self.assertIn('id="manager-insight-report-page"', html)

    def test_period_today_sets_duration(self):
        page = self.client.get("/reports/sales/manager-insight?period=today")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn('value="today"', html)
        self.assertIn('value="2026-08-13"', html)
        self.assertRegex(
            html,
            r'data-mi-period="today"[^>]*is-active|class="md-period-pill is-active"[^>]*data-mi-period="today"',
        )

    def test_changing_duration_uses_selected_month_column(self):
        self._insert_stay()
        april = self.client.get(
            "/reports/sales/manager-insight?date_from=2026-04-01&date_to=2026-04-01"
        )
        august = self.client.get(
            "/reports/sales/manager-insight?date_from=2026-08-12&date_to=2026-08-13"
        )
        self.assertEqual(april.status_code, 200)
        self.assertEqual(august.status_code, 200)
        april_html = april.get_data(as_text=True)
        august_html = august.get_data(as_text=True)
        self.assertIn(">April Month<", april_html)
        self.assertIn(">August Month<", august_html)
        april_sold = self._row_cells(april_html, "rooms_sold")
        august_sold = self._row_cells(august_html, "rooms_sold")
        self.assertEqual(april_sold[1], "0")
        self.assertEqual(april_sold[2], "0")
        self.assertEqual(august_sold[1], "2")
        self.assertEqual(august_sold[2], "2")
        self.assertEqual(april_sold[3], august_sold[3])

    def test_page_occupancy_matches_layout_inventory(self):
        self._insert_stay()
        page = self.client.get(
            "/reports/sales/manager-insight?date_from=2026-08-12&date_to=2026-08-13"
        )
        html = page.get_data(as_text=True)
        conn = db_mod.get_db()
        try:
            layout = db_mod.get_hotel_rooms_layout(conn)
        finally:
            conn.close()
        sellable = sum(
            1
            for room in layout.get("rooms") or []
            if db_mod._normalize_hotel_room_status(room.get("status")) != "out_of_order"
        )
        total_rooms = self._row_cells(html, "total_rooms")
        self.assertEqual(total_rooms[1], str(sellable * 2))
        sold = self._row_cells(html, "rooms_sold")
        self.assertEqual(sold[1], "2")
        pax = self._row_cells(html, "pax")
        self.assertEqual(pax[1], "4")

    def test_excel_columns_and_no_tip(self):
        self._insert_stay()
        export = self.client.get(
            "/reports/sales/manager-insight/export?date_from=2026-08-12&date_to=2026-08-13"
        )
        self.assertEqual(export.status_code, 200)
        from openpyxl import load_workbook

        ws = load_workbook(BytesIO(export.data)).active
        self.assertEqual(ws.title, "Manager Insight")
        self.assertIn("Manager Insight", ws["A1"].value or "")
        self.assertEqual(
            [ws.cell(2, col).value for col in range(1, 5)],
            ["HEAD", "Duration", "August Month", "YTD"],
        )
        labels = [ws.cell(row, 1).value for row in range(3, (ws.max_row or 2) + 1)]
        self.assertIn("Rooms Sold", labels)
        self.assertIn("Total Room Revenue", labels)
        self.assertNotIn("Tip", labels)
        sold_row = labels.index("Rooms Sold") + 3
        self.assertEqual(int(ws.cell(sold_row, 2).value), 2)

    def test_build_manager_insight_uses_indian_fy(self):
        self._insert_stay(
            check_in="2026-07-31",
            check_out="2026-08-02",
            total=2000,
        )
        conn = db_mod.get_db()
        try:
            payload = build_manager_insight(
                conn,
                date_from=date(2026, 7, 31),
                date_to=date(2026, 7, 31),
                today=FROZEN_TODAY,
            )
        finally:
            conn.close()
        by_key = {row["key"]: row for row in payload["rows"]}
        self.assertEqual(by_key["rooms_sold"]["duration"], 1)
        self.assertEqual(by_key["rooms_sold"]["mtd"], 1)
        self.assertEqual(payload["windows"]["ytd"][0], date(2026, 4, 1))


if __name__ == "__main__":
    unittest.main()
