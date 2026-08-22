"""POS Invoice Ledger — save, list, KPI, upsert, delete."""

import os
import tempfile
import unittest
from datetime import date, timedelta
from unittest import mock

import db as db_mod


class PosInvoiceLedgerTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db_path = self.tmp.name
        self._orig_path = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = self.db_path
        db_mod.init_db()

        import app as app_mod

        self.app_mod = app_mod
        self.app = app_mod.app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

        conn = db_mod.get_db()
        try:
            admin = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            self.admin_id = admin["id"]
        finally:
            conn.close()

        self.user = {
            "id": self.admin_id,
            "username": "admin",
            "full_name": "Administrator",
            "is_admin": True,
            "is_active": True,
            "dashboard_access": set(),
            "stores_access": set(),
        }
        self._get_user_patch = mock.patch.object(app_mod, "get_current_user", return_value=self.user)
        self._get_user_patch.start()

    def tearDown(self):
        self._get_user_patch.stop()
        db_mod.DATABASE_PATH = self._orig_path
        try:
            os.unlink(self.db_path)
        except OSError:
            pass

    def _payload(self, order_no="ORD-2607-0001", total=500, **overrides):
        data = {
            "orderNo": order_no,
            "savedAt": "2026-07-22 18:00:00",
            "orderType": "dine_in",
            "table": "T1",
            "captain": "",
            "customerName": "Guest One",
            "customerMobile": "9876543210",
            "notes": "",
            "discountType": "pct",
            "discountValue": 0,
            "serviceType": "pct",
            "serviceValue": 0,
            "tipAmount": 0,
            "couponCode": "",
            "lines": [
                {
                    "uid": "1",
                    "menuId": None,
                    "name": "Filter Coffee",
                    "variant": "Hot",
                    "rate": 100,
                    "qty": 2,
                },
                {
                    "uid": "2",
                    "menuId": None,
                    "name": "Masala Dosa",
                    "variant": "",
                    "rate": 150,
                    "qty": 2,
                },
            ],
            "totals": {
                "subtotal": 500,
                "discount": 0,
                "discountType": "pct",
                "discountValue": 0,
                "gst": 25,
                "service": 0,
                "serviceType": "pct",
                "serviceValue": 0,
                "tip": 0,
                "roundOff": 0,
                "total": total,
            },
        }
        data.update(overrides)
        return data

    def test_save_list_kpi_upsert_delete(self):
        save = self.client.post("/point-of-sale/api/invoices", json=self._payload())
        self.assertEqual(save.status_code, 200, save.get_data(as_text=True))
        body = save.get_json()
        self.assertTrue(body["ok"])
        invoice_id = body["invoice"]["id"]
        self.assertEqual(body["invoice"]["order_no"], "ORD-2607-0001")
        self.assertEqual(len(body["invoice"]["lines"]), 2)
        self.assertEqual(body["invoice"]["grand_total"], 500)

        page = self.client.get("/point-of-sale/invoice-ledger")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn("ORD-2607-0001", html)
        self.assertIn("Invoice Ledger", html)
        self.assertIn("22 July 26", html)
        self.assertIn("18:00", html)
        self.assertIn("Guest One", html)
        self.assertIn("Payment Mode", html)
        self.assertNotIn('data-sort="table"', html)

        save2 = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._payload(total=550, customerName="Guest Updated"),
        )
        self.assertEqual(save2.status_code, 200)
        body2 = save2.get_json()
        self.assertTrue(body2["ok"])
        self.assertEqual(body2["invoice"]["id"], invoice_id)
        self.assertEqual(body2["invoice"]["customer_name"], "Guest Updated")
        self.assertEqual(body2["invoice"]["grand_total"], 550)

        detail = self.client.get(f"/point-of-sale/api/invoices/{invoice_id}")
        self.assertEqual(detail.status_code, 200)
        detail_body = detail.get_json()
        self.assertTrue(detail_body["ok"])
        self.assertEqual(detail_body["invoice"]["customer_name"], "Guest Updated")

        # Second invoice for KPI count
        save_b = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._payload(order_no="ORD-2607-0002", total=200, customerName="Guest Two"),
        )
        self.assertEqual(save_b.status_code, 200)

        conn = db_mod.get_db()
        try:
            rows = db_mod.list_pos_invoices(conn)
            kpis = db_mod.pos_invoice_kpis(conn, rows, today="2026-07-22")
        finally:
            conn.close()
        self.assertEqual(kpis["invoice_count"], 2)
        self.assertEqual(kpis["total_sales"], 750)
        self.assertEqual(kpis["average_bill"], 375)
        self.assertEqual(kpis["today_sales"], 750)

        deleted = self.client.post(
            f"/point-of-sale/api/invoices/{invoice_id}/delete",
            json={"reason": "Test cancel"},
        )
        self.assertEqual(deleted.status_code, 200)
        deleted_body = deleted.get_json()
        self.assertTrue(deleted_body["ok"])
        self.assertEqual(deleted_body.get("mode"), "cancelled")

        cancelled = self.client.get(f"/point-of-sale/api/invoices/{invoice_id}")
        self.assertEqual(cancelled.status_code, 200)
        self.assertEqual(cancelled.get_json()["invoice"]["status"], "cancelled")
        self.assertEqual(cancelled.get_json()["invoice"]["payment_mode_label"], "Cancelled")

        conn = db_mod.get_db()
        try:
            rows = db_mod.list_pos_invoices(conn)
            kpis_after = db_mod.pos_invoice_kpis(conn, rows, today="2026-07-22")
            cancelled_row = next(r for r in rows if r["id"] == invoice_id)
        finally:
            conn.close()
        self.assertEqual(cancelled_row["payment_mode_label"], "Cancelled")
        # Cancelled invoice remains listed but is excluded from sales KPIs.
        self.assertEqual(kpis_after["invoice_count"], 1)
        self.assertEqual(kpis_after["total_sales"], 200)

        page2 = self.client.get("/point-of-sale/invoice-ledger")
        html2 = page2.get_data(as_text=True)
        self.assertIn("ORD-2607-0001", html2)
        self.assertIn("ORD-2607-0002", html2)
        self.assertIn("Cancelled", html2)

    def test_ledger_hides_delete_for_generated_and_settled(self):
        from datetime import date

        today = date.today().isoformat()
        draft = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._payload(order_no="SPC/DRAFT1/26-27", total=100),
        )
        self.assertEqual(draft.status_code, 200, draft.get_data(as_text=True))
        draft_id = draft.get_json()["invoice"]["id"]

        generated = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._payload(
                order_no="SPC/AAAAAA/26-27",
                total=200,
                customerBill=True,
                orderDate=today,
                savedAt=today + " 12:00:00",
                orderType="takeaway",
                table="",
            ),
        )
        self.assertEqual(generated.status_code, 200, generated.get_data(as_text=True))
        gen_inv = generated.get_json()["invoice"]
        gen_id = gen_inv["id"]
        self.assertTrue(gen_inv.get("customer_bill_sent"))
        gen_no = gen_inv["order_no"]

        settled = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._payload(order_no="ORD-SETTLE-01", total=300, table="T2"),
        )
        self.assertEqual(settled.status_code, 200, settled.get_data(as_text=True))
        settled_id = settled.get_json()["invoice"]["id"]
        close = self.client.post(f"/point-of-sale/api/invoices/{settled_id}/close")
        self.assertEqual(close.status_code, 200, close.get_data(as_text=True))

        page = self.client.get("/point-of-sale/invoice-ledger")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)

        self.assertIn(f'data-invoice-id="{draft_id}"', html)
        self.assertRegex(
            html,
            rf'pos-il-delete-btn[^>]*data-invoice-id="{draft_id}"|data-invoice-id="{draft_id}"[^>]*pos-il-delete-btn',
        )
        self.assertRegex(
            html,
            rf'pos-il-edit-btn[^>]*data-invoice-id="{draft_id}"|data-invoice-id="{draft_id}"[^>]*pos-il-edit-btn',
        )
        self.assertNotIn(
            f'pos-il-cancel-btn" data-tip="Cancel" aria-label="Cancel invoice SPC/DRAFT1/26-27"',
            html,
        )

        self.assertIn(f'data-invoice-id="{gen_id}"', html)
        self.assertIn(f'Cancel invoice {gen_no}', html)
        self.assertIn("pos-il-cancel-btn", html)
        self.assertRegex(
            html,
            rf'pos-il-edit-btn[^>]*data-invoice-id="{gen_id}"|data-invoice-id="{gen_id}"[^>]*pos-il-edit-btn',
        )
        self.assertNotRegex(
            html,
            rf'pos-il-delete-btn[^>]*data-invoice-id="{gen_id}"|data-invoice-id="{gen_id}"[^>]*pos-il-delete-btn',
        )

        self.assertIn(f'data-invoice-id="{settled_id}"', html)
        settled_block = html.split(f'data-invoice-id="{settled_id}"', 1)[1].split(
            "</tr>", 1
        )[0]
        self.assertNotIn("pos-il-delete-btn", settled_block)
        self.assertNotIn("pos-il-cancel-btn", settled_block)
        self.assertNotIn("pos-il-edit-btn", settled_block)

        # Settled cancel blocked for everyone (admin included).
        blocked = self.client.post(
            f"/point-of-sale/api/invoices/{settled_id}/delete",
            json={"reason": "Should fail"},
        )
        self.assertEqual(blocked.status_code, 400)
        self.assertIn("Settled", blocked.get_json()["error"])

    def test_export_report(self):
        save = self.client.post("/point-of-sale/api/invoices", json=self._payload())
        self.assertEqual(save.status_code, 200)
        export = self.client.get("/point-of-sale/invoice-ledger/report")
        self.assertEqual(export.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            export.content_type or "",
        )
        self.assertIn(b"PK", export.data[:4])
        fy_start, today = db_mod.indian_fiscal_year_bounds()
        from reports import report_export_filename

        expected_name = report_export_filename(
            "Invoice Ledger - Restaurant",
            date_from=fy_start,
            date_to=today,
            date_filter_active=True,
        )
        self.assertIn(expected_name, export.headers.get("Content-Disposition") or "")

        from io import BytesIO
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(export.data))
        ws = wb.active
        self.assertEqual(
            ws["A1"].value, "Hotel Bell Elite — Invoice Ledger — Restaurant"
        )
        self.assertTrue(str(ws["A2"].value or "").startswith("From "))
        headers = [ws.cell(3, col).value for col in range(1, ws.max_column + 1)]
        self.assertEqual(headers[0], "Order No")
        self.assertEqual(headers[1], "Date")
        self.assertEqual(headers[2], "Customer")
        self.assertNotIn("Saved At", headers)
        self.assertEqual(headers[5], "Payment Mode")
        self.assertEqual(headers[6], "Cash")
        self.assertEqual(headers[7], "UPI")
        self.assertEqual(headers[8], "Card")
        self.assertEqual(headers[9], "Room Transfer")
        self.assertEqual(headers[-1], "Total")
        self.assertEqual(ws["A1"].fill.fgColor.rgb, "FF315A78")
        self.assertEqual(ws.cell(3, 1).fill.fgColor.rgb, "FF315A78")
        self.assertEqual(ws["A1"].font.color.rgb, "FFFFFFFF")
        self.assertEqual(ws.cell(3, 1).font.color.rgb, "FFFFFFFF")
        self.assertEqual(ws["A1"].font.size, 14)
        self.assertEqual(ws.cell(3, 1).font.size, 11)
        self.assertTrue(ws["A1"].font.bold)
        self.assertEqual(ws.cell(4, 1).font.size, 11)
        self.assertEqual(ws.cell(4, 1).font.color.rgb, "FF000000")
        self.assertEqual(ws.cell(4, 1).alignment.horizontal, "left")
        self.assertEqual(ws.cell(4, ws.max_column).alignment.horizontal, "right")
        self.assertIsNotNone(ws.cell(4, 1).border.left.style)
        self.assertEqual(ws.cell(4, 1).value, "ORD-2607-0001")
        self.assertEqual(ws.cell(4, 2).value, "2026-07-22")
        self.assertEqual(ws.cell(4, 3).value, "Guest One")

    def test_save_validation(self):
        empty = self.client.post(
            "/point-of-sale/api/invoices",
            json={"orderNo": "ORD-X", "customerName": "A", "lines": [], "totals": {}},
        )
        self.assertEqual(empty.status_code, 400)

        no_name = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._payload(customerName=""),
        )
        self.assertEqual(no_name.status_code, 400)

    def test_today_invoices_lists_todays_bills_newest_first(self):
        from datetime import datetime, timedelta

        empty = self.client.get("/point-of-sale/api/today-invoices")
        self.assertEqual(empty.status_code, 200)
        body = empty.get_json()
        self.assertTrue(body["ok"])
        self.assertEqual(body["invoice_count"], 0)
        self.assertEqual(body["invoices"], [])
        today = body["date"]
        self.assertEqual(today, datetime.now().strftime("%Y-%m-%d"))

        # Older day must not appear in the hub.
        yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        self.client.post(
            "/point-of-sale/api/invoices",
            json=self._payload(
                order_no="ORD-OLD-0001",
                savedAt=f"{yesterday} 10:00:00",
                table="T9",
            ),
        )

        older = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._payload(
                order_no="ORD-TODAY-0001",
                savedAt=f"{today} 09:00:00",
                orderType="takeaway",
                table="",
            ),
        )
        self.assertEqual(older.status_code, 200, older.get_data(as_text=True))
        newer = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._payload(
                order_no="ORD-TODAY-0002",
                savedAt=f"{today} 18:30:00",
                orderType="dine_in",
                table="T2",
            ),
        )
        self.assertEqual(newer.status_code, 200, newer.get_data(as_text=True))

        res = self.client.get("/point-of-sale/api/today-invoices")
        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["date"], today)
        self.assertEqual(payload["invoice_count"], 2)
        orders = [inv["order_no"] for inv in payload["invoices"]]
        self.assertEqual(orders, ["ORD-TODAY-0002", "ORD-TODAY-0001"])
        first = payload["invoices"][0]
        self.assertEqual(first["table_label"], "T2")
        self.assertEqual(first["order_type"], "dine_in")
        self.assertEqual(first["status"], "open")
        self.assertIn("grand_total", first)
        self.assertIn("saved_at", first)

    def test_ledger_shows_settlement_payment_mode(self):
        saved = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._payload(order_no="ORD-PAY-0001", total=500, table="T1"),
        )
        self.assertEqual(saved.status_code, 200, saved.get_data(as_text=True))
        invoice = saved.get_json()["invoice"]
        invoice_id = invoice["id"]
        total = float(invoice["grand_total"])

        settle = self.client.post(
            f"/point-of-sale/api/invoices/{invoice_id}/settle",
            json={
                "payment_splits": [
                    {"payment_method": "cash", "amount": 200},
                    {"payment_method": "upi", "amount": total - 200},
                ],
            },
        )
        self.assertEqual(settle.status_code, 200, settle.get_data(as_text=True))
        settled = settle.get_json()["invoice"]
        self.assertEqual(settled.get("payment_mode_label"), "Cash + UPI")

        page = self.client.get("/point-of-sale/invoice-ledger")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn("Cash + UPI", html)
        self.assertIn("Payment Mode", html)
        self.assertIn(">Cash<", html)
        self.assertIn(">UPI<", html)
        self.assertIn(">Room Transfer<", html)
        self.assertIn('data-sort="pay_cash"', html)
        self.assertNotIn('aria-label="Settlement by payment mode"', html)
        self.assertNotIn("pos-il-settlement-summary", html)

        conn = db_mod.get_db()
        try:
            rows = db_mod.list_pos_invoices(conn)
        finally:
            conn.close()
        match = next(r for r in rows if r["id"] == invoice_id)
        self.assertEqual(match["payment_modes"], ["cash", "upi"])
        self.assertEqual(match["payment_mode_label"], "Cash + UPI")
        self.assertEqual(match["payment_amounts"]["cash"], 200.0)
        self.assertAlmostEqual(
            match["payment_amounts"]["upi"], total - 200, places=2
        )
        self.assertEqual(match["payment_amounts"]["card"], 0.0)
        self.assertEqual(match["payment_amounts"]["room_transfer"], 0.0)

    def test_ledger_split_cash_and_room_transfer_amounts(self):
        today = date.today()
        checkin = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Asha",
                    "lastName": "Nair",
                    "mobile": "9000000001",
                    "checkInDate": today.isoformat(),
                    "checkOutDate": (today + timedelta(days=1)).isoformat(),
                    "roomRate": 2000,
                    "nights": 1,
                    "advancePaid": 0,
                },
            },
        )
        self.assertEqual(checkin.status_code, 200, checkin.get_data(as_text=True))

        saved = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._payload(order_no="ORD-SPLIT-RT-1", total=500, table="T1"),
        )
        self.assertEqual(saved.status_code, 200, saved.get_data(as_text=True))
        invoice = saved.get_json()["invoice"]
        invoice_id = invoice["id"]
        total = float(invoice["grand_total"])
        cash_amt = 150.0
        transfer_amt = round(total - cash_amt, 2)

        settle = self.client.post(
            f"/point-of-sale/api/invoices/{invoice_id}/settle",
            json={
                "payment_splits": [
                    {"payment_method": "cash", "amount": cash_amt},
                    {"payment_method": "room_transfer", "amount": transfer_amt},
                ],
                "hotel_room_id": "room-101",
            },
        )
        self.assertEqual(settle.status_code, 200, settle.get_data(as_text=True))
        settled = settle.get_json()["invoice"]
        self.assertIn("Cash", settled.get("payment_mode_label") or "")
        self.assertIn("Room Transfer", settled.get("payment_mode_label") or "")

        conn = db_mod.get_db()
        try:
            rows = db_mod.list_pos_invoices(conn)
            kpis = db_mod.pos_invoice_kpis(conn, rows)
        finally:
            conn.close()
        match = next(r for r in rows if r["id"] == invoice_id)
        self.assertEqual(match["payment_amounts"]["cash"], cash_amt)
        self.assertAlmostEqual(
            match["payment_amounts"]["room_transfer"], transfer_amt, places=2
        )
        self.assertGreaterEqual(kpis["payment_totals"]["cash"], cash_amt)
        self.assertGreaterEqual(
            kpis["payment_totals"]["room_transfer"], transfer_amt - 0.01
        )

        export = self.client.get("/point-of-sale/invoice-ledger/report")
        self.assertEqual(export.status_code, 200)
        from io import BytesIO
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(export.data))
        ws = wb.active
        headers = [ws.cell(3, col).value for col in range(1, ws.max_column + 1)]
        cash_col = headers.index("Cash") + 1
        room_col = headers.index("Room Transfer") + 1
        order_row = None
        for row in range(4, ws.max_row + 1):
            if ws.cell(row, 1).value == "ORD-SPLIT-RT-1":
                order_row = row
                break
        self.assertIsNotNone(order_row)
        self.assertEqual(ws.cell(order_row, cash_col).value, cash_amt)
        self.assertAlmostEqual(
            float(ws.cell(order_row, room_col).value), transfer_amt, places=2
        )

    def test_ledger_room_transfer_payment_mode_shows_fbe_reference(self):
        today = date.today()
        checkin = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Asha",
                    "lastName": "Nair",
                    "mobile": "9000000001",
                    "checkInDate": today.isoformat(),
                    "checkOutDate": (today + timedelta(days=1)).isoformat(),
                    "roomRate": 2000,
                    "nights": 1,
                    "advancePaid": 0,
                },
            },
        )
        self.assertEqual(checkin.status_code, 200, checkin.get_data(as_text=True))

        saved = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._payload(order_no="ORD-FBE-0709", total=500, table="T1"),
        )
        self.assertEqual(saved.status_code, 200, saved.get_data(as_text=True))
        invoice = saved.get_json()["invoice"]
        invoice_id = invoice["id"]
        total = float(invoice["grand_total"])

        settle = self.client.post(
            f"/point-of-sale/api/invoices/{invoice_id}/settle",
            json={
                "payment_splits": [
                    {"payment_method": "room_transfer", "amount": total},
                ],
                "hotel_room_id": "room-101",
            },
        )
        self.assertEqual(settle.status_code, 200, settle.get_data(as_text=True))
        pending_label = "Room Transfer · 101 (Invoice yet to generate)"
        self.assertEqual(
            settle.get_json()["invoice"].get("payment_mode_label"),
            pending_label,
        )

        pending_page = self.client.get("/point-of-sale/invoice-ledger")
        self.assertEqual(pending_page.status_code, 200)
        pending_html = pending_page.get_data(as_text=True)
        self.assertIn(pending_label, pending_html)

        conn = db_mod.get_db()
        try:
            rows = db_mod.list_pos_invoices(conn)
        finally:
            conn.close()
        match = next(r for r in rows if r["id"] == invoice_id)
        self.assertEqual(match["payment_modes"], ["room_transfer"])
        self.assertEqual(match["payment_mode_label"], pending_label)

        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "payment_splits": []},
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))
        fb_no = gen.get_json()["room"]["stay"]["fbTransferInvoiceNumber"]
        self.assertTrue(str(fb_no).startswith("FBE/"))

        generated_page = self.client.get("/point-of-sale/invoice-ledger")
        self.assertEqual(generated_page.status_code, 200)
        generated_html = generated_page.get_data(as_text=True)
        expected = f"Room Transfer · 101 ({fb_no})"
        self.assertIn(expected, generated_html)

        conn = db_mod.get_db()
        try:
            rows = db_mod.list_pos_invoices(conn)
        finally:
            conn.close()
        match = next(r for r in rows if r["id"] == invoice_id)
        self.assertEqual(match["payment_mode_label"], expected)
        conn = db_mod.get_db()
        try:
            single = db_mod.get_pos_invoice(conn, invoice_id)
        finally:
            conn.close()
        self.assertEqual(single["payment_mode_label"], expected)

    def test_ledger_shows_unsettled_when_no_payment(self):
        saved = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._payload(order_no="ORD-PAY-OPEN", total=500, table="T1"),
        )
        self.assertEqual(saved.status_code, 200, saved.get_data(as_text=True))
        invoice_id = saved.get_json()["invoice"]["id"]

        page = self.client.get("/point-of-sale/invoice-ledger")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn("Unsettled", html)
        self.assertIn('id="pos-inv-settle-modal"', html)
        self.assertIn("pos_settle.js", html)
        self.assertIn("data-se-listbox-combobox", html)
        self.assertIn('id="pos-inv-settle-hotel-room-trigger"', html)
        self.assertNotIn("ep-listbox-search", html)
        self.assertIn("Settle invoice ORD-PAY-OPEN", html)
        self.assertIn('class="pos-il-row is-unsettled"', html)

        conn = db_mod.get_db()
        try:
            rows = db_mod.list_pos_invoices(conn)
        finally:
            conn.close()
        match = next(r for r in rows if r["id"] == invoice_id)
        self.assertEqual(match["payment_modes"], [])
        self.assertEqual(match["payment_mode_label"], "Unsettled")

    def test_ledger_filters_by_settlement_status(self):
        open_res = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._payload(order_no="ORD-FILT-OPEN", total=500, table="T1"),
        )
        self.assertEqual(open_res.status_code, 200, open_res.get_data(as_text=True))
        open_id = open_res.get_json()["invoice"]["id"]

        settled_res = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._payload(order_no="ORD-FILT-PAID", total=500, table="T2"),
        )
        self.assertEqual(settled_res.status_code, 200, settled_res.get_data(as_text=True))
        settled_inv = settled_res.get_json()["invoice"]
        settled_id = settled_inv["id"]
        total = float(settled_inv["grand_total"])
        settle = self.client.post(
            f"/point-of-sale/api/invoices/{settled_id}/settle",
            json={"payment_splits": [{"payment_method": "cash", "amount": total}]},
        )
        self.assertEqual(settle.status_code, 200, settle.get_data(as_text=True))

        unsettled_page = self.client.get("/point-of-sale/invoice-ledger?settlement=unsettled")
        self.assertEqual(unsettled_page.status_code, 200)
        unsettled_html = unsettled_page.get_data(as_text=True)
        self.assertIn("ORD-FILT-OPEN", unsettled_html)
        self.assertNotIn("ORD-FILT-PAID", unsettled_html)
        self.assertIn('id="pos-il-settlement"', unsettled_html)
        self.assertIn("Un Settled", unsettled_html)

        settled_page = self.client.get("/point-of-sale/invoice-ledger?settlement=settled")
        self.assertEqual(settled_page.status_code, 200)
        settled_html = settled_page.get_data(as_text=True)
        self.assertIn("ORD-FILT-PAID", settled_html)
        self.assertNotIn("ORD-FILT-OPEN", settled_html)

        conn = db_mod.get_db()
        try:
            unsettled_rows = db_mod.list_pos_invoices(conn, settlement="unsettled")
            settled_rows = db_mod.list_pos_invoices(conn, settlement="settled")
        finally:
            conn.close()
        self.assertTrue(any(r["id"] == open_id for r in unsettled_rows))
        self.assertFalse(any(r["id"] == settled_id for r in unsettled_rows))
        self.assertTrue(any(r["id"] == settled_id for r in settled_rows))
        self.assertFalse(any(r["id"] == open_id for r in settled_rows))

    def test_ledger_settle_selected_invoices(self):
        first = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._payload(order_no="ORD-SEL-0001", total=200, table="T1"),
        )
        second = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._payload(order_no="ORD-SEL-0002", total=300, table="T2"),
        )
        self.assertEqual(first.status_code, 200, first.get_data(as_text=True))
        self.assertEqual(second.status_code, 200, second.get_data(as_text=True))
        inv_a = first.get_json()["invoice"]
        inv_b = second.get_json()["invoice"]
        combined = round(float(inv_a["grand_total"]) + float(inv_b["grand_total"]), 2)

        page = self.client.get("/point-of-sale/invoice-ledger")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn('id="pos-il-select-all"', html)
        self.assertIn('id="pos-il-settle-selected"', html)
        self.assertIn("pos-il-row-check", html)
        self.assertIn("Settle selected", html)
        self.assertIn("/point-of-sale/api/invoices/settle-selected", html)
        self.assertIn("ORD-SEL-0001", html)
        self.assertIn("ORD-SEL-0002", html)

        settle = self.client.post(
            "/point-of-sale/api/invoices/settle-selected",
            json={
                "invoice_ids": [inv_a["id"], inv_b["id"]],
                "payment_splits": [{"payment_method": "cash", "amount": combined}],
            },
        )
        self.assertEqual(settle.status_code, 200, settle.get_data(as_text=True))
        payload = settle.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["paid_count"], 2)
        self.assertEqual(payload["settled_count"], 2)
        ids = {row["id"] for row in payload["invoices"]}
        self.assertEqual(ids, {inv_a["id"], inv_b["id"]})
        for row in payload["invoices"]:
            self.assertEqual(row["status"], "closed")
            self.assertTrue(row.get("payments") or row.get("payment_modes"))

        again = self.client.post(
            "/point-of-sale/api/invoices/settle-selected",
            json={
                "invoice_ids": [inv_a["id"]],
                "payment_splits": [{"payment_method": "cash", "amount": 1}],
            },
        )
        self.assertEqual(again.status_code, 400)

        dup = self.client.post(
            "/point-of-sale/api/invoices/settle-selected",
            json={
                "invoice_ids": [inv_b["id"], inv_b["id"]],
                "payment_splits": [{"payment_method": "cash", "amount": 1}],
            },
        )
        self.assertEqual(dup.status_code, 400)

    def test_bar_ledger_settle_selected_invoices(self):
        first = self.client.post(
            "/bar-point-of-sale/api/invoices",
            json=self._payload(order_no="BAR-SEL-0001", total=180, table="B1"),
        )
        second = self.client.post(
            "/bar-point-of-sale/api/invoices",
            json=self._payload(order_no="BAR-SEL-0002", total=220, table="B2"),
        )
        self.assertEqual(first.status_code, 200, first.get_data(as_text=True))
        self.assertEqual(second.status_code, 200, second.get_data(as_text=True))
        inv_a = first.get_json()["invoice"]
        inv_b = second.get_json()["invoice"]
        combined = round(float(inv_a["grand_total"]) + float(inv_b["grand_total"]), 2)

        page = self.client.get("/bar-point-of-sale/invoice-ledger")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn('id="pos-il-select-all"', html)
        self.assertIn("/bar-point-of-sale/api/invoices/settle-selected", html)
        self.assertIn("BAR-SEL-0001", html)

        restaurant = self.client.post(
            "/point-of-sale/api/invoices/settle-selected",
            json={
                "invoice_ids": [inv_a["id"], inv_b["id"]],
                "payment_splits": [{"payment_method": "cash", "amount": combined}],
            },
        )
        self.assertEqual(restaurant.status_code, 404)

        settle = self.client.post(
            "/bar-point-of-sale/api/invoices/settle-selected",
            json={
                "invoice_ids": [inv_a["id"], inv_b["id"]],
                "payment_splits": [
                    {"payment_method": "cash", "amount": 150},
                    {"payment_method": "upi", "amount": round(combined - 150, 2)},
                ],
            },
        )
        self.assertEqual(settle.status_code, 200, settle.get_data(as_text=True))
        payload = settle.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["paid_count"], 2)
        self.assertEqual(payload["settled_count"], 2)
        for row in payload["invoices"]:
            self.assertEqual(row["status"], "closed")


if __name__ == "__main__":
    unittest.main()
