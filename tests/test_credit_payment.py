"""Tests for credit payment settlement helpers and validation."""

import sqlite3
import unittest
from datetime import date

import app as app_module
import db as db_mod


def _memory_conn():
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    conn.executescript(
        """
        CREATE TABLE suppliers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            gst TEXT NOT NULL DEFAULT '',
            address TEXT NOT NULL DEFAULT '',
            phone TEXT NOT NULL DEFAULT '',
            bank_name TEXT NOT NULL DEFAULT '',
            bank_account_number TEXT NOT NULL DEFAULT '',
            ifsc_code TEXT NOT NULL DEFAULT ''
        );
        CREATE TABLE sales_update_expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company TEXT NOT NULL,
            location TEXT NOT NULL,
            sales_date TEXT NOT NULL,
            description TEXT NOT NULL DEFAULT '',
            amount REAL NOT NULL DEFAULT 0,
            payment_type TEXT NOT NULL DEFAULT 'cash',
            transaction_id TEXT NOT NULL DEFAULT '',
            invoice_number TEXT NOT NULL DEFAULT '',
            expense_code TEXT NOT NULL DEFAULT '',
            supplier_id INTEGER,
            category TEXT NOT NULL DEFAULT '',
            entry_kind TEXT NOT NULL DEFAULT 'expense',
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            updated_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE credit_payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company TEXT NOT NULL,
            supplier_id INTEGER NOT NULL,
            payment_date TEXT NOT NULL,
            payment_method TEXT NOT NULL DEFAULT 'cash',
            transaction_id TEXT NOT NULL DEFAULT '',
            total_amount REAL NOT NULL DEFAULT 0,
            notes TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            updated_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE credit_payment_allocations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            credit_payment_id INTEGER NOT NULL,
            expense_id INTEGER NOT NULL,
            amount REAL NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE purchase_verifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company TEXT NOT NULL,
            supplier_id INTEGER NOT NULL,
            verification_date TEXT NOT NULL,
            verification_method TEXT NOT NULL DEFAULT 'cash',
            verification_account TEXT NOT NULL DEFAULT '',
            transaction_id TEXT NOT NULL DEFAULT '',
            total_amount REAL NOT NULL DEFAULT 0,
            notes TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            updated_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE purchase_verification_allocations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            purchase_verification_id INTEGER NOT NULL,
            expense_id INTEGER NOT NULL,
            amount REAL NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        );
        """
    )
    return conn


def _seed_supplier(conn, name="Acme Foods", gst="29AAAAA0000A1Z5"):
    cur = conn.execute(
        "INSERT INTO suppliers (name, gst) VALUES (?, ?)",
        (name, gst),
    )
    return cur.lastrowid


def _seed_expense(conn, supplier_id, amount, payment_type="credit", code="HBE-EX-1", sales_date="2026-07-01", entry_kind="expense"):
    cur = conn.execute(
        """INSERT INTO sales_update_expenses
           (company, location, sales_date, description, amount, payment_type, supplier_id, category, expense_code, entry_kind)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        ("HBE", "Hotel", sales_date, "Test expense", amount, payment_type, supplier_id, "grocery", code, entry_kind),
    )
    return cur.lastrowid


def _seed_verification(conn, supplier_id, expense_id, amount, verification_date="2026-07-10"):
    cur = conn.execute(
        """INSERT INTO purchase_verifications
           (company, supplier_id, verification_date, verification_method, verification_account, total_amount)
           VALUES ('HBE', ?, ?, 'cash', 'administrator', ?)""",
        (supplier_id, verification_date, amount),
    )
    verification_id = cur.lastrowid
    conn.execute(
        """INSERT INTO purchase_verification_allocations
           (purchase_verification_id, expense_id, amount)
           VALUES (?, ?, ?)""",
        (verification_id, expense_id, amount),
    )
    return verification_id


class CreditPaymentBalanceTests(unittest.TestCase):
    def test_balance_none_partial_full(self):
        self.assertEqual(app_module._credit_expense_balance(10000, 0), 10000)
        self.assertEqual(app_module._credit_expense_balance(10000, 2500), 7500)
        self.assertEqual(app_module._credit_expense_balance(10000, 10000), 0)
        self.assertEqual(app_module._credit_expense_balance(10000, 12000), 0)

    def test_settlement_status_labels(self):
        self.assertEqual(app_module._credit_settlement_status("credit", 100, 0), "outstanding")
        self.assertEqual(app_module._credit_settlement_status("credit", 100, 40), "partial")
        self.assertEqual(app_module._credit_settlement_status("credit", 100, 100), "cleared")
        self.assertEqual(app_module._credit_settlement_status("cash", 100, 0), "cleared")
        self.assertEqual(app_module._credit_settlement_status("bank_transfer", 100, 0), "cleared")

    def test_optional_filter_date_range_inactive_when_missing(self):
        date_from, date_to, active = app_module._resolve_optional_filter_date_range({}, "date_from", "date_to")
        self.assertIsNone(date_from)
        self.assertIsNone(date_to)
        self.assertFalse(active)

    def test_optional_filter_date_range_default_fy_when_requested(self):
        date_from, date_to, active = app_module._resolve_optional_filter_date_range(
            {}, "date_from", "date_to", default_fy=True
        )
        fy_start, today = db_mod.indian_fiscal_year_bounds()
        self.assertTrue(active)
        self.assertEqual(date_from, fy_start)
        self.assertEqual(date_to, today)

    def test_payment_date_filter_stays_inactive_when_missing(self):
        date_from, date_to, active = app_module._resolve_optional_filter_date_range(
            {}, "payment_date_from", "payment_date_to"
        )
        self.assertIsNone(date_from)
        self.assertIsNone(date_to)
        self.assertFalse(active)

    def test_optional_filter_date_range_active_when_provided(self):
        date_from, date_to, active = app_module._resolve_optional_filter_date_range(
            {"date_from": "2026-07-01", "date_to": "2026-07-14"},
            "date_from",
            "date_to",
        )
        self.assertEqual(date_from, date(2026, 7, 1))
        self.assertEqual(date_to, date(2026, 7, 14))
        self.assertTrue(active)


class CreditPaymentValidationTests(unittest.TestCase):
    def setUp(self):
        self.conn = _memory_conn()
        self.supplier_a = _seed_supplier(self.conn, "Supplier A", "29AAAAA0000A1Z5")
        self.supplier_b = _seed_supplier(self.conn, "Supplier B", "29BBBBB0000B1Z5")
        self.expense_a1 = _seed_expense(self.conn, self.supplier_a, 10000, code="HBE-EX-1")
        self.expense_a2 = _seed_expense(self.conn, self.supplier_a, 5000, code="HBE-EX-2", sales_date="2026-07-02")
        self.expense_b1 = _seed_expense(self.conn, self.supplier_b, 3000, code="HBE-EX-3")
        self.expense_cash = _seed_expense(
            self.conn, self.supplier_a, 2000, payment_type="cash", code="HBE-EX-4"
        )
        _seed_verification(self.conn, self.supplier_a, self.expense_a1, 10000)
        _seed_verification(self.conn, self.supplier_a, self.expense_a2, 5000)
        _seed_verification(self.conn, self.supplier_b, self.expense_b1, 3000)
        self.conn.commit()

    def tearDown(self):
        self.conn.close()

    def _create_payment(self, allocations, **overrides):
        payload = {
            "supplier_id": self.supplier_a,
            "payment_date": "2026-07-13",
            "payment_method": "cash",
            "transaction_id": "",
            "notes": "",
            "allocations": allocations,
        }
        payload.update(overrides)
        return app_module._validate_credit_payment_payload(self.conn, payload)

    def test_valid_multi_expense_same_supplier(self):
        payload, errors = self._create_payment([
            {"expense_id": self.expense_a1, "amount": 6000},
            {"expense_id": self.expense_a2, "amount": 3000},
        ])
        self.assertEqual(errors, [])
        self.assertEqual(payload["total_amount"], 9000)
        self.assertEqual(len(payload["allocations"]), 2)

    def test_reject_mixed_suppliers(self):
        payload, errors = self._create_payment([
            {"expense_id": self.expense_a1, "amount": 1000},
            {"expense_id": self.expense_b1, "amount": 1000},
        ])
        self.assertIsNone(payload)
        self.assertTrue(any("same supplier" in err.lower() for err in errors))

    def test_reject_over_allocation(self):
        payload, errors = self._create_payment([
            {"expense_id": self.expense_a1, "amount": 15000},
        ])
        self.assertIsNone(payload)
        self.assertTrue(any("exceeds outstanding" in err.lower() for err in errors))

    def test_reject_non_credit_expense(self):
        payload, errors = self._create_payment([
            {"expense_id": self.expense_cash, "amount": 500},
        ])
        self.assertIsNone(payload)
        self.assertTrue(any("credit" in err.lower() for err in errors))

    def test_reject_card_without_transaction_id(self):
        payload, errors = self._create_payment(
            [{"expense_id": self.expense_a1, "amount": 1000}],
            payment_method="card",
            transaction_id="",
        )
        self.assertIsNone(payload)
        self.assertTrue(any("transaction id" in err.lower() for err in errors))

    def test_partial_then_remaining_balance(self):
        payload, errors = self._create_payment([
            {"expense_id": self.expense_a1, "amount": 4000},
        ])
        self.assertEqual(errors, [])
        cur = self.conn.execute(
            """INSERT INTO credit_payments
               (company, supplier_id, payment_date, payment_method, transaction_id, total_amount, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (
                payload["company"],
                payload["supplier_id"],
                payload["payment_date"],
                payload["payment_method"],
                payload["transaction_id"],
                payload["total_amount"],
                payload["notes"],
            ),
        )
        payment_id = cur.lastrowid
        for allocation in payload["allocations"]:
            self.conn.execute(
                """INSERT INTO credit_payment_allocations (credit_payment_id, expense_id, amount)
                   VALUES (?, ?, ?)""",
                (payment_id, allocation["expense_id"], allocation["amount"]),
            )
        self.conn.commit()

        paid = app_module._credit_expense_paid_total(self.conn, self.expense_a1)
        self.assertEqual(paid, 4000)
        balance = app_module._credit_expense_balance(10000, paid)
        self.assertEqual(balance, 6000)

        payload2, errors2 = self._create_payment([
            {"expense_id": self.expense_a1, "amount": 6000},
        ])
        self.assertEqual(errors2, [])
        self.assertEqual(payload2["total_amount"], 6000)

        payload3, errors3 = self._create_payment([
            {"expense_id": self.expense_a1, "amount": 6000.01},
        ])
        self.assertIsNone(payload3)
        self.assertTrue(errors3)

    def test_outstanding_filters_cleared_expenses(self):
        cur = self.conn.execute(
            """INSERT INTO credit_payments
               (company, supplier_id, payment_date, payment_method, total_amount)
               VALUES ('HBE', ?, '2026-07-13', 'cash', 10000)""",
            (self.supplier_a,),
        )
        payment_id = cur.lastrowid
        self.conn.execute(
            """INSERT INTO credit_payment_allocations (credit_payment_id, expense_id, amount)
               VALUES (?, ?, 10000)""",
            (payment_id, self.expense_a1),
        )
        self.conn.commit()

        entries = app_module._outstanding_credit_expenses(
            self.conn, date(2026, 7, 1), date(2026, 7, 31), supplier_id=self.supplier_a
        )
        ids = {entry["id"] for entry in entries}
        self.assertNotIn(self.expense_a1, ids)
        self.assertIn(self.expense_a2, ids)

    def test_outstanding_excludes_unverified_credit_expenses(self):
        unverified = _seed_expense(
            self.conn, self.supplier_a, 2500, code="HBE-EX-UNVERIFIED", sales_date="2026-07-03"
        )
        self.conn.commit()

        entries = app_module._outstanding_credit_expenses(
            self.conn, date(2026, 7, 1), date(2026, 7, 31), supplier_id=self.supplier_a
        )
        ids = {entry["id"] for entry in entries}
        self.assertNotIn(unverified, ids)
        self.assertIn(self.expense_a1, ids)
        self.assertIn(self.expense_a2, ids)

    def test_outstanding_filters_by_entry_kind(self):
        purchase = _seed_expense(
            self.conn,
            self.supplier_a,
            1800,
            code="HBE-PU-KIND",
            sales_date="2026-07-04",
            entry_kind="purchase",
        )
        _seed_verification(self.conn, self.supplier_a, purchase, 1800, verification_date="2026-07-04")
        self.conn.commit()

        purchases = app_module._outstanding_credit_expenses(
            self.conn,
            date(2026, 7, 1),
            date(2026, 7, 31),
            supplier_id=self.supplier_a,
            entry_kind="purchase",
        )
        expenses = app_module._outstanding_credit_expenses(
            self.conn,
            date(2026, 7, 1),
            date(2026, 7, 31),
            supplier_id=self.supplier_a,
            entry_kind="expense",
        )
        self.assertEqual({entry["id"] for entry in purchases}, {purchase})
        self.assertIn(self.expense_a1, {entry["id"] for entry in expenses})
        self.assertNotIn(purchase, {entry["id"] for entry in expenses})

    def test_reject_unverified_credit_expense(self):
        unverified = _seed_expense(
            self.conn, self.supplier_a, 2500, code="HBE-EX-UNVERIFIED", sales_date="2026-07-03"
        )
        self.conn.commit()
        payload, errors = self._create_payment([
            {"expense_id": unverified, "amount": 2500},
        ])
        self.assertIsNone(payload)
        self.assertTrue(any("verified" in err.lower() for err in errors))

    def test_delete_restores_balance(self):
        cur = self.conn.execute(
            """INSERT INTO credit_payments
               (company, supplier_id, payment_date, payment_method, total_amount)
               VALUES ('HBE', ?, '2026-07-13', 'cash', 4000)""",
            (self.supplier_a,),
        )
        payment_id = cur.lastrowid
        self.conn.execute(
            """INSERT INTO credit_payment_allocations (credit_payment_id, expense_id, amount)
               VALUES (?, ?, 4000)""",
            (payment_id, self.expense_a1),
        )
        self.conn.commit()
        self.assertEqual(app_module._credit_expense_paid_total(self.conn, self.expense_a1), 4000)

        self.conn.execute(
            "DELETE FROM credit_payment_allocations WHERE credit_payment_id = ?",
            (payment_id,),
        )
        self.conn.execute("DELETE FROM credit_payments WHERE id = ?", (payment_id,))
        self.conn.commit()
        self.assertEqual(app_module._credit_expense_paid_total(self.conn, self.expense_a1), 0)

    def test_purchase_ledger_reflects_cleared_credit_payment_mode(self):
        cur = self.conn.execute(
            """INSERT INTO credit_payments
               (company, supplier_id, payment_date, payment_method, total_amount)
               VALUES ('HBE', ?, '2026-07-13', 'cash', 10000)""",
            (self.supplier_a,),
        )
        payment_id = cur.lastrowid
        self.conn.execute(
            """INSERT INTO credit_payment_allocations (credit_payment_id, expense_id, amount)
               VALUES (?, ?, 10000)""",
            (payment_id, self.expense_a1),
        )
        self.conn.commit()
        app_module._sync_expense_payment_after_clearance(self.conn, self.expense_a1)
        self.conn.commit()

        entries = app_module._purchase_ledger_entries(
            self.conn, date(2026, 7, 1), date(2026, 7, 31)
        )
        cleared = next(entry for entry in entries if entry["id"] == self.expense_a1)
        self.assertEqual(cleared["display_payment_type"], "cash")
        self.assertEqual(cleared["settlement_status"], "cleared")

        cash_entry = next(entry for entry in entries if entry["id"] == self.expense_cash)
        self.assertEqual(cash_entry["display_payment_type"], "cash")
        self.assertEqual(cash_entry["settlement_status"], "cleared")

    def test_update_outstanding_purchase_from_ledger(self):
        recent = date.today().isoformat()
        self.conn.execute(
            "UPDATE sales_update_expenses SET sales_date = ? WHERE id = ?",
            (recent, self.expense_a1),
        )
        self.conn.commit()
        result, error = app_module._update_purchase_ledger_expense(
            self.conn,
            {"is_admin": True},
            {
                "expense_id": self.expense_a1,
                "date": recent,
                "description": "Updated credit purchase",
                "amount": 12000,
                "payment_type": "credit",
                "category": "grocery",
                "supplier_id": self.supplier_a,
                "invoice_number": "INV-EDIT-1",
            },
        )
        self.assertIsNone(error)
        self.assertEqual(result["expense_id"], self.expense_a1)
        self.assertEqual(result["sales_date"], recent)
        row = self.conn.execute(
            "SELECT description, amount, sales_date, invoice_number FROM sales_update_expenses WHERE id = ?",
            (self.expense_a1,),
        ).fetchone()
        self.assertEqual(row["description"], "Updated credit purchase")
        self.assertEqual(float(row["amount"]), 12000.0)
        self.assertEqual(row["sales_date"], recent)
        self.assertEqual(row["invoice_number"], "INV-EDIT-1")

    def test_reject_edit_for_cleared_purchase(self):
        recent = date.today().isoformat()
        self.conn.execute(
            "UPDATE sales_update_expenses SET sales_date = ? WHERE id = ?",
            (recent, self.expense_a1),
        )
        cur = self.conn.execute(
            """INSERT INTO credit_payments
               (company, supplier_id, payment_date, payment_method, total_amount)
               VALUES ('HBE', ?, ?, 'cash', 10000)""",
            (self.supplier_a, recent),
        )
        payment_id = cur.lastrowid
        self.conn.execute(
            """INSERT INTO credit_payment_allocations (credit_payment_id, expense_id, amount)
               VALUES (?, ?, 10000)""",
            (payment_id, self.expense_a1),
        )
        self.conn.commit()
        result, error = app_module._update_purchase_ledger_expense(
            self.conn,
            {"is_admin": True},
            {
                "expense_id": self.expense_a1,
                "date": recent,
                "description": "Should fail",
                "amount": 50,
                "payment_type": "credit",
                "category": "grocery",
                "supplier_id": self.supplier_a,
            },
        )
        self.assertIsNone(result)
        self.assertEqual(error, app_module.PURCHASE_LEDGER_CREDIT_SETTLED_EDIT_MESSAGE)

    def test_delete_outstanding_purchase_from_ledger(self):
        recent = date.today().isoformat()
        self.conn.execute(
            "UPDATE sales_update_expenses SET sales_date = ? WHERE id = ?",
            (recent, self.expense_a2),
        )
        self.conn.commit()
        result, error = app_module._delete_purchase_ledger_expense(
            self.conn,
            {"is_admin": True},
            {"expense_id": self.expense_a2},
        )
        self.assertIsNone(error)
        self.assertEqual(result["expense_id"], self.expense_a2)
        gone = self.conn.execute(
            "SELECT id FROM sales_update_expenses WHERE id = ?",
            (self.expense_a2,),
        ).fetchone()
        self.assertIsNone(gone)

    def test_reject_delete_for_cleared_purchase(self):
        result, error = app_module._delete_purchase_ledger_expense(
            self.conn,
            {"is_admin": True},
            {"expense_id": self.expense_cash},
        )
        self.assertIsNone(result)
        self.assertIn("outstanding", (error or "").lower())

    def test_reject_duplicate_supplier_invoice(self):
        self.conn.execute(
            "UPDATE sales_update_expenses SET invoice_number = ? WHERE id = ?",
            ("INV-1001", self.expense_a1),
        )
        self.conn.commit()

        result, error = app_module._create_sales_expense(
            self.conn,
            {"is_admin": True},
            {
                "company": "HBE",
                "location": "Hotel",
                "date": "2026-07-13",
                "description": "Duplicate invoice",
                "amount": 500,
                "payment_type": "cash",
                "category": "grocery",
                "supplier_id": self.supplier_a,
                "invoice_number": "inv-1001",
            },
            skip_cash_check=True,
        )
        self.assertIsNone(result)
        self.assertIn("already exists", error.lower())

        result2, error2 = app_module._create_sales_expense(
            self.conn,
            {"is_admin": True},
            {
                "company": "HBE",
                "location": "Hotel",
                "date": "2026-07-13",
                "description": "Unique invoice",
                "amount": 500,
                "payment_type": "cash",
                "category": "grocery",
                "supplier_id": self.supplier_a,
                "invoice_number": "INV-2002",
            },
            skip_cash_check=True,
        )
        self.assertIsNotNone(result2)
        self.assertIsNone(error2)


class PurchaseVerificationIsolationTests(unittest.TestCase):
    def setUp(self):
        self.conn = _memory_conn()
        self.user = {"username": "administrator", "full_name": "Administrator"}
        self.supplier_a = _seed_supplier(self.conn, "ABC Supplies", "29ABCDE1234F1Z5")
        self.expense_credit = _seed_expense(self.conn, self.supplier_a, 100, code="HBE-EX-1")
        self.expense_cash = _seed_expense(
            self.conn, self.supplier_a, 50, payment_type="cash", code="HBE-EX-2"
        )
        self.conn.commit()

    def tearDown(self):
        self.conn.close()

    def _record_credit_payment(self, expense_id, amount):
        cur = self.conn.execute(
            """INSERT INTO credit_payments
               (company, supplier_id, payment_date, payment_method, total_amount)
               VALUES ('HBE', ?, '2026-07-13', 'card', 100)""",
            (self.supplier_a,),
        )
        payment_id = cur.lastrowid
        self.conn.execute(
            """INSERT INTO credit_payment_allocations (credit_payment_id, expense_id, amount)
               VALUES (?, ?, ?)""",
            (payment_id, expense_id, amount),
        )
        self.conn.commit()
        return payment_id

    def test_credit_payment_does_not_appear_in_verification_history(self):
        self._record_credit_payment(self.expense_credit, 100)

        history = app_module._purchase_verification_entries(
            self.conn, verification_date_from=date(2026, 7, 1), verification_date_to=date(2026, 7, 31)
        )
        self.assertEqual(history, [])

    def test_credit_payment_does_not_clear_pending_verification(self):
        self._record_credit_payment(self.expense_credit, 100)

        pending = app_module._pending_purchase_verifications(
            self.conn, date(2026, 7, 1), date(2026, 7, 31), supplier_id=self.supplier_a
        )
        ids = {entry["id"] for entry in pending}
        self.assertIn(self.expense_credit, ids)
        self.assertIn(self.expense_cash, ids)

    def test_verification_accepts_any_hotel_purchase_type(self):
        payload, errors = app_module._validate_purchase_verification_payload(
            self.conn,
            {
                "supplier_id": self.supplier_a,
                "payment_date": "2026-07-13",
                "payment_method": "cash",
                "allocations": [
                    {"expense_id": self.expense_credit, "amount": 100},
                    {"expense_id": self.expense_cash, "amount": 50},
                ],
            },
            user=self.user,
        )
        self.assertEqual(errors, [])
        self.assertEqual(payload["total_amount"], 150)
        self.assertEqual(payload["verification_account"], "administrator")

    def test_verification_requires_logged_in_user(self):
        payload, errors = app_module._validate_purchase_verification_payload(
            self.conn,
            {
                "supplier_id": self.supplier_a,
                "payment_date": "2026-07-13",
                "allocations": [{"expense_id": self.expense_credit, "amount": 100}],
            },
        )
        self.assertIsNone(payload)
        self.assertTrue(any("logged in" in err.lower() for err in errors))

    def test_verification_removes_expense_from_pending_list(self):
        payload, errors = app_module._validate_purchase_verification_payload(
            self.conn,
            {
                "supplier_id": self.supplier_a,
                "payment_date": "2026-07-13",
                "payment_method": "cash",
                "allocations": [{"expense_id": self.expense_credit, "amount": 100}],
            },
            user=self.user,
        )
        self.assertEqual(errors, [])
        cur = self.conn.execute(
            """INSERT INTO purchase_verifications
               (company, supplier_id, verification_date, verification_method, verification_account, total_amount)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (
                payload["company"],
                payload["supplier_id"],
                payload["verification_date"],
                payload["verification_method"],
                payload["verification_account"],
                payload["total_amount"],
            ),
        )
        verification_id = cur.lastrowid
        self.conn.execute(
            """INSERT INTO purchase_verification_allocations
               (purchase_verification_id, expense_id, amount)
               VALUES (?, ?, ?)""",
            (verification_id, self.expense_credit, 100),
        )
        self.conn.commit()

        pending = app_module._pending_purchase_verifications(
            self.conn, date(2026, 7, 1), date(2026, 7, 31), supplier_id=self.supplier_a
        )
        ids = {entry["id"] for entry in pending}
        self.assertNotIn(self.expense_credit, ids)
        self.assertIn(self.expense_cash, ids)

        history = app_module._purchase_verification_entries(
            self.conn, verification_date_from=date(2026, 7, 1), verification_date_to=date(2026, 7, 31)
        )
        self.assertEqual(len(history), 1)
        self.assertEqual(history[0]["total_amount"], 100)
        self.assertEqual(history[0]["verification_account"], "administrator")
        self.assertEqual(history[0]["expense_codes"], "HBE-EX-1")


class CreditPaymentAccessTests(unittest.TestCase):
    def test_endpoints_map_to_accounts(self):
        from workspace_access import get_endpoint_dashboard_module

        for endpoint in (
            "credit_payment",
            "purchase_verification",
            "create_credit_payment",
            "delete_credit_payment",
            "credit_payment_detail",
            "create_purchase_verification",
            "delete_purchase_verification",
            "purchase_verification_detail",
            "export_credit_payment_report",
            "export_purchase_verification_report",
        ):
            self.assertEqual(get_endpoint_dashboard_module(endpoint), "accounts")


class PurchaseVerificationApprovalGateTests(unittest.TestCase):
    """Verify/Approve/Revert require Approval module; Accounts alone is view-only."""

    def setUp(self):
        import os
        import tempfile
        from unittest import mock

        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db_path = self.tmp.name
        self._orig_path = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = self.db_path
        db_mod.init_db()

        self.app = app_module.app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

        conn = db_mod.get_db()
        try:
            admin = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            self.admin_id = admin["id"]
            cur = conn.execute(
                """INSERT INTO suppliers
                   (name, gst, bank_account_number, ifsc_code)
                   VALUES (?, ?, ?, ?)""",
                ("Gate Vendor", "29BBBBB0000B1Z5", "987654321098", "HDFC0001234"),
            )
            self.supplier_id = cur.lastrowid
            cur = conn.execute(
                """INSERT INTO sales_update_expenses
                   (company, location, sales_date, description, amount, payment_type,
                    expense_code, supplier_id, category)
                   VALUES ('HBE', 'Hotel', '2026-07-10', 'Gate expense', 500, 'credit',
                           'HBE-EX-GATE', ?, 'vegetables')""",
                (self.supplier_id,),
            )
            self.expense_id = cur.lastrowid
            cur = conn.execute(
                """INSERT INTO sales_update_expenses
                   (company, location, sales_date, description, amount, payment_type,
                    expense_code, supplier_id, category)
                   VALUES ('HBE', 'Hotel', '2026-07-12', 'Open expense', 300, 'credit',
                           'HBE-EX-OPEN', ?, 'vegetables')""",
                (self.supplier_id,),
            )
            self.open_expense_id = cur.lastrowid
            cur = conn.execute(
                """INSERT INTO purchase_verifications
                   (company, supplier_id, verification_date, verification_method,
                    verification_account, total_amount)
                   VALUES ('HBE', ?, '2026-07-11', 'cash', 'administrator', 500)""",
                (self.supplier_id,),
            )
            self.verification_id = cur.lastrowid
            conn.execute(
                """INSERT INTO purchase_verification_allocations
                   (purchase_verification_id, expense_id, amount)
                   VALUES (?, ?, ?)""",
                (self.verification_id, self.expense_id, 500),
            )
            conn.commit()
        finally:
            conn.close()

        self._mock = mock
        self._os = os
        self.viewer = {
            "id": self.admin_id,
            "username": "accounts_viewer",
            "full_name": "Accounts Viewer",
            "is_admin": False,
            "is_active": True,
            "dashboard_access": {"accounts"},
            "stores_access": set(),
        }
        self.approver = {
            "id": self.admin_id,
            "username": "approver",
            "full_name": "Approver",
            "is_admin": False,
            "is_active": True,
            "dashboard_access": {"accounts", "approval"},
            "stores_access": set(),
        }

    def tearDown(self):
        db_mod.DATABASE_PATH = self._orig_path
        try:
            self._os.unlink(self.db_path)
        except OSError:
            pass

    def test_page_hides_verify_without_approval(self):
        with self._mock.patch.object(app_module, "get_current_user", return_value=self.viewer):
            page = self.client.get("/accounts/purchase-verification")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn('data-can-mutate="0"', html)
        self.assertNotIn('id="cp-header-verify-btn"', html)
        self.assertNotIn('class="cp-row-approve-btn"', html)

        with self._mock.patch.object(app_module, "get_current_user", return_value=self.viewer):
            history = self.client.get("/accounts/purchase-verification?view=history")
        self.assertEqual(history.status_code, 200)
        history_html = history.get_data(as_text=True)
        self.assertIn('data-can-mutate="0"', history_html)
        self.assertIn("cp-view-payment", history_html)
        self.assertNotIn('class="act-btn del cp-delete-payment"', history_html)

    def test_page_shows_verify_with_approval(self):
        with self._mock.patch.object(app_module, "get_current_user", return_value=self.approver):
            page = self.client.get("/accounts/purchase-verification")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn('data-can-mutate="1"', html)
        self.assertIn('id="cp-open-select-btn"', html)

    def test_create_requires_approval(self):
        payload = {
            "supplier_id": self.supplier_id,
            "verification_date": "2026-07-12",
            "verification_method": "cash",
            "allocations": [{"expense_id": self.expense_id, "amount": 500}],
        }
        with self._mock.patch.object(app_module, "get_current_user", return_value=self.viewer):
            denied = self.client.post(
                "/accounts/purchase-verification/create",
                json=payload,
            )
        self.assertEqual(denied.status_code, 403)
        body = denied.get_json()
        self.assertFalse(body.get("ok"))
        self.assertIn("Approval", body.get("error") or "")

    def test_delete_requires_approval(self):
        with self._mock.patch.object(app_module, "get_current_user", return_value=self.viewer):
            denied = self.client.post(
                "/accounts/purchase-verification/delete",
                json={"payment_id": self.verification_id},
            )
        self.assertEqual(denied.status_code, 403)
        body = denied.get_json()
        self.assertFalse(body.get("ok"))
        self.assertIn("Approval", body.get("error") or "")

        with self._mock.patch.object(app_module, "get_current_user", return_value=self.approver):
            allowed = self.client.post(
                "/accounts/purchase-verification/delete",
                json={"payment_id": self.verification_id},
            )
        self.assertEqual(allowed.status_code, 200, allowed.get_data(as_text=True))
        self.assertTrue(allowed.get_json().get("ok"))


class CreditPaymentApprovalGateTests(unittest.TestCase):
    """Clear Payment / Pay / Revert require Approval module; Accounts alone is view-only."""

    def setUp(self):
        import os
        import tempfile
        from unittest import mock

        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db_path = self.tmp.name
        self._orig_path = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = self.db_path
        db_mod.init_db()

        self.app = app_module.app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

        conn = db_mod.get_db()
        try:
            admin = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            self.admin_id = admin["id"]
            cur = conn.execute(
                """INSERT INTO suppliers
                   (name, gst, bank_account_number, ifsc_code)
                   VALUES (?, ?, ?, ?)""",
                ("Clear Vendor", "29CCCCC0000C1Z5", "111122223333", "SBIN0001234"),
            )
            self.supplier_id = cur.lastrowid
            cur = conn.execute(
                """INSERT INTO sales_update_expenses
                   (company, location, sales_date, description, amount, payment_type,
                    expense_code, supplier_id, category)
                   VALUES ('HBE', 'Hotel', '2026-07-10', 'Cleared expense', 400, 'credit',
                           'HBE-EX-CLR', ?, 'vegetables')""",
                (self.supplier_id,),
            )
            self.paid_expense_id = cur.lastrowid
            cur = conn.execute(
                """INSERT INTO sales_update_expenses
                   (company, location, sales_date, description, amount, payment_type,
                    expense_code, supplier_id, category)
                   VALUES ('HBE', 'Hotel', '2026-07-12', 'Open credit', 250, 'credit',
                           'HBE-EX-OPENCP', ?, 'vegetables')""",
                (self.supplier_id,),
            )
            self.open_expense_id = cur.lastrowid
            # Both expenses fully verified so they appear on Credit Payment.
            for expense_id, amount in (
                (self.paid_expense_id, 400),
                (self.open_expense_id, 250),
            ):
                cur = conn.execute(
                    """INSERT INTO purchase_verifications
                       (company, supplier_id, verification_date, verification_method,
                        verification_account, total_amount)
                       VALUES ('HBE', ?, '2026-07-13', 'cash', 'administrator', ?)""",
                    (self.supplier_id, amount),
                )
                verification_id = cur.lastrowid
                conn.execute(
                    """INSERT INTO purchase_verification_allocations
                       (purchase_verification_id, expense_id, amount)
                       VALUES (?, ?, ?)""",
                    (verification_id, expense_id, amount),
                )
            cur = conn.execute(
                """INSERT INTO credit_payments
                   (company, supplier_id, payment_date, payment_method, total_amount)
                   VALUES ('HBE', ?, '2026-07-14', 'cash', 400)""",
                (self.supplier_id,),
            )
            self.payment_id = cur.lastrowid
            conn.execute(
                """INSERT INTO credit_payment_allocations
                   (credit_payment_id, expense_id, amount)
                   VALUES (?, ?, ?)""",
                (self.payment_id, self.paid_expense_id, 400),
            )
            conn.commit()
        finally:
            conn.close()

        self._mock = mock
        self._os = os
        self.viewer = {
            "id": self.admin_id,
            "username": "accounts_viewer",
            "full_name": "Accounts Viewer",
            "is_admin": False,
            "is_active": True,
            "dashboard_access": {"accounts"},
            "stores_access": set(),
        }
        self.approver = {
            "id": self.admin_id,
            "username": "approver",
            "full_name": "Approver",
            "is_admin": False,
            "is_active": True,
            "dashboard_access": {"accounts", "approval"},
            "stores_access": set(),
        }

    def tearDown(self):
        db_mod.DATABASE_PATH = self._orig_path
        try:
            self._os.unlink(self.db_path)
        except OSError:
            pass

    def test_page_hides_clear_payment_without_approval(self):
        with self._mock.patch.object(app_module, "get_current_user", return_value=self.viewer):
            page = self.client.get("/accounts/credit-payment")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn('data-can-mutate="0"', html)
        self.assertNotIn('id="cp-header-verify-btn"', html)
        self.assertNotIn('class="cp-row-approve-btn"', html)

        with self._mock.patch.object(app_module, "get_current_user", return_value=self.viewer):
            history = self.client.get("/accounts/credit-payment?view=history")
        self.assertEqual(history.status_code, 200)
        history_html = history.get_data(as_text=True)
        self.assertIn('data-can-mutate="0"', history_html)
        self.assertIn("cp-view-payment", history_html)
        self.assertNotIn('class="act-btn del cp-delete-payment"', history_html)

    def test_page_shows_clear_payment_with_approval(self):
        with self._mock.patch.object(app_module, "get_current_user", return_value=self.approver):
            page = self.client.get("/accounts/credit-payment")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn('data-can-mutate="1"', html)
        self.assertIn('id="cp-open-select-btn"', html)
        self.assertIn("Clear Payment", html)

    def test_create_requires_approval(self):
        payload = {
            "supplier_id": self.supplier_id,
            "payment_date": "2026-07-15",
            "payment_method": "cash",
            "allocations": [{"expense_id": self.open_expense_id, "amount": 250}],
        }
        with self._mock.patch.object(app_module, "get_current_user", return_value=self.viewer):
            denied = self.client.post(
                "/accounts/credit-payment/create",
                json=payload,
            )
        self.assertEqual(denied.status_code, 403)
        body = denied.get_json()
        self.assertFalse(body.get("ok"))
        self.assertIn("Approval", body.get("error") or "")

    def test_delete_requires_approval(self):
        with self._mock.patch.object(app_module, "get_current_user", return_value=self.viewer):
            denied = self.client.post(
                "/accounts/credit-payment/delete",
                json={"payment_id": self.payment_id},
            )
        self.assertEqual(denied.status_code, 403)
        body = denied.get_json()
        self.assertFalse(body.get("ok"))
        self.assertIn("Approval", body.get("error") or "")

        with self._mock.patch.object(app_module, "get_current_user", return_value=self.approver):
            allowed = self.client.post(
                "/accounts/credit-payment/delete",
                json={"payment_id": self.payment_id},
            )
        self.assertEqual(allowed.status_code, 200, allowed.get_data(as_text=True))
        self.assertTrue(allowed.get_json().get("ok"))


class CreditPaymentReportExportTests(unittest.TestCase):
    """Excel export must work even when vendor_payment_template.xlsx is absent."""

    def setUp(self):
        import os
        import tempfile
        from unittest import mock

        import db as db_mod

        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db_path = self.tmp.name
        self._orig_path = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = self.db_path
        db_mod.init_db()

        self.app = app_module.app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

        conn = db_mod.get_db()
        try:
            admin = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            self.admin_id = admin["id"]
            cur = conn.execute(
                """INSERT INTO suppliers
                   (name, gst, bank_account_number, ifsc_code)
                   VALUES (?, ?, ?, ?)""",
                ("ICICI Vendor", "29AAAAA0000A1Z5", "123456789012", "ICIC0001234"),
            )
            supplier_id = cur.lastrowid
            cur = conn.execute(
                """INSERT INTO sales_update_expenses
                   (company, location, sales_date, description, amount, payment_type,
                    expense_code, supplier_id, category)
                   VALUES ('HBE', 'Hotel', '2026-07-10', 'Veg supplies', 2500, 'credit',
                           'HBE-EX-RPT', ?, 'vegetables')""",
                (supplier_id,),
            )
            expense_id = cur.lastrowid
            cur = conn.execute(
                """INSERT INTO purchase_verifications
                   (company, supplier_id, verification_date, verification_method,
                    verification_account, total_amount)
                   VALUES ('HBE', ?, '2026-07-11', 'cash', 'administrator', 2500)""",
                (supplier_id,),
            )
            verification_id = cur.lastrowid
            conn.execute(
                """INSERT INTO purchase_verification_allocations
                   (purchase_verification_id, expense_id, amount)
                   VALUES (?, ?, ?)""",
                (verification_id, expense_id, 2500),
            )
            conn.commit()
        finally:
            conn.close()

        self.user = {
            "id": self.admin_id,
            "username": "admin",
            "full_name": "Administrator",
            "is_admin": True,
            "is_active": True,
            "dashboard_access": set(),
            "stores_access": set(),
        }
        self._get_user_patch = mock.patch.object(
            app_module, "get_current_user", return_value=self.user
        )
        self._get_user_patch.start()
        self._os = os
        self._db_mod = db_mod

    def tearDown(self):
        self._get_user_patch.stop()
        self._db_mod.DATABASE_PATH = self._orig_path
        try:
            self._os.unlink(self.db_path)
        except OSError:
            pass

    def test_export_generates_xlsx_without_template(self):
        self.assertFalse(self._os.path.isfile(app_module._VENDOR_PAYMENT_TEMPLATE))
        resp = self.client.get("/accounts/credit-payment/report?supplier=all")
        self.assertEqual(resp.status_code, 200, f"status={resp.status_code}")
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resp.content_type or "",
        )
        self.assertEqual(resp.data[:2], b"PK")
        cd = resp.headers.get("Content-Disposition") or ""
        fy_start, today = db_mod.indian_fiscal_year_bounds()
        from reports import report_export_filename

        expected_name = report_export_filename(
            "Credit Payment",
            date_from=fy_start,
            date_to=today,
            date_filter_active=True,
        )
        self.assertIn(expected_name, cd)

        from openpyxl import load_workbook
        import io

        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb.active
        self.assertEqual(ws.cell(1, 1).value, "PYMT_PROD_TYPE_CODE")
        self.assertEqual(ws.cell(2, 1).value, "PAB_VENDOR")
        self.assertEqual(ws.cell(2, 4).value, "ICICI Vendor")
        self.assertEqual(float(ws.cell(2, 7).value), 2500.0)


if __name__ == "__main__":
    unittest.main()
