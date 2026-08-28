import os
import tempfile
import unittest
from datetime import datetime, timedelta
from unittest import mock

import db as db_mod


class StoresFlowTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db_path = self.tmp.name
        self._orig_path = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = self.db_path
        db_mod.init_db()

        import app as app_mod
        import stores as stores_mod

        self.app_mod = app_mod
        self.stores_mod = stores_mod
        self.app = app_mod.app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

        conn = db_mod.get_db()
        try:
            admin = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            self.admin_id = admin["id"]
        finally:
            conn.close()

        self.user = {
            "id": self.admin_id,
            "username": "admin",
            "full_name": "Administrator",
            "is_admin": True,
            "is_active": True,
            "dashboard_access": set(),
            "stores_access": set(),
        }
        self._get_user_patch = mock.patch.object(app_mod, "get_current_user", return_value=self.user)
        self._stores_user_patch = mock.patch.object(stores_mod, "_get_user", return_value=self.user)
        self._get_user_patch.start()
        self._stores_user_patch.start()

        # CRITICAL: never hit live Meta/WhatsApp while exercising indent submit flows.
        # .env may load real credentials; each pending indent would otherwise send paid messages.
        self._wa_buttons_patch = mock.patch(
            "whatsapp_indent.wa.send_interactive_buttons",
            return_value=(True, "", {"messages": [{"id": "wamid.BTN"}]}),
        )
        self._wa_buttons_patch.start()

    def tearDown(self):
        self._wa_buttons_patch.stop()
        self._get_user_patch.stop()
        self._stores_user_patch.stop()
        db_mod.DATABASE_PATH = self._orig_path
        try:
            os.unlink(self.db_path)
        except OSError:
            pass

    def test_format_stores_dt(self):
        fmt = self.stores_mod._format_stores_dt
        self.assertEqual(fmt("2026-07-19 10:05:57"), "19-July 10.05 AM")
        self.assertEqual(fmt("2026-07-19 13:18:06"), "19-July 1.18 PM")
        self.assertEqual(fmt("2026-07-19"), "19-July")
        self.assertEqual(fmt(""), "")
        self.assertEqual(fmt(None), "")
        self.assertEqual(
            self.stores_mod._format_stores_date_line("2026-07-19 10:05:57"),
            "19 July",
        )
        self.assertEqual(
            self.stores_mod._format_stores_time_line("2026-07-19 10:05:57"),
            "10:05 AM",
        )

    def test_indent_no_format_per_outlet_fiscal_year(self):
        from datetime import date

        fy = db_mod.indian_fiscal_year_label(date(2026, 7, 29))
        self.assertEqual(fy, "2026-27")
        next_no = self.stores_mod._next_indent_no
        conn = db_mod.get_db()
        try:
            bar1 = next_no(conn, "bar", when=date(2026, 7, 29))
            self.assertEqual(bar1, "IND/BAR/26-27/1")
            conn.execute(
                """
                INSERT INTO store_indents
                    (outlet, indent_no, status, notes, created_at)
                VALUES ('bar', ?, 'draft', '', datetime('now','localtime'))
                """,
                (bar1,),
            )
            bar2 = next_no(conn, "bar", when=date(2026, 7, 29))
            self.assertEqual(bar2, "IND/BAR/26-27/2")
            rest1 = next_no(conn, "restaurant", when=date(2026, 7, 29))
            self.assertEqual(rest1, "IND/RES/26-27/1")
            # Prior FY does not advance current-year series
            conn.execute(
                """
                INSERT INTO store_indents
                    (outlet, indent_no, status, notes, created_at)
                VALUES ('bar', 'IND/BAR/25-26/9', 'draft', '', datetime('now','localtime'))
                """,
            )
            bar_still_2 = next_no(conn, "bar", when=date(2026, 7, 29))
            self.assertEqual(bar_still_2, "IND/BAR/26-27/2")
            # Legacy IND/Bar/n/YYYY-YY numbers still advance the same FY series.
            conn.execute(
                """
                INSERT INTO store_indents
                    (outlet, indent_no, status, notes, created_at)
                VALUES ('restaurant', 'IND/Restaurant/3/2026-27', 'draft', '', datetime('now','localtime'))
                """,
            )
            rest_next = next_no(conn, "restaurant", when=date(2026, 7, 29))
            self.assertEqual(rest_next, "IND/RES/26-27/4")
        finally:
            conn.close()

        create_bar = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Indent no format bar",
                "item_name": ["Onion"],
                "quantity": ["1"],
                "unit": ["kg"],
                "approximate_price": ["10"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create_bar.status_code, 302)
        create_rest = self.client.post(
            "/stores/indent?outlet=restaurant",
            data={
                "outlet": "restaurant",
                "action": "submit",
                "notes": "Indent no format restaurant",
                "item_name": ["Onion"],
                "quantity": ["1"],
                "unit": ["kg"],
                "approximate_price": ["10"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create_rest.status_code, 302)
        conn = db_mod.get_db()
        try:
            bar = conn.execute(
                "SELECT indent_no FROM store_indents WHERE notes = 'Indent no format bar'"
            ).fetchone()
            rest = conn.execute(
                "SELECT indent_no FROM store_indents WHERE notes = 'Indent no format restaurant'"
            ).fetchone()
            self.assertIsNotNone(bar)
            self.assertIsNotNone(rest)
            self.assertRegex(bar["indent_no"], r"^IND/BAR/\d{2}-\d{2}/\d+$")
            self.assertRegex(rest["indent_no"], r"^IND/RES/\d{2}-\d{2}/\d+$")
        finally:
            conn.close()

    def test_product_master_seeded(self):
        conn = db_mod.get_db()
        try:
            cats = {
                row["name"]: row["id"]
                for row in conn.execute("SELECT id, name FROM store_product_categories").fetchall()
            }
            self.assertIn("Non-Veg", cats)
            self.assertIn("Dairy Products", cats)
            self.assertIn("Fruits", cats)
            self.assertIn("Vegetable", cats)
            fruit_cat = conn.execute(
                """
                SELECT c.name
                FROM store_products p
                JOIN store_product_categories c ON c.id = p.category_id
                WHERE lower(p.name) = lower('Anar') AND p.is_active = 1
                LIMIT 1
                """
            ).fetchone()
            self.assertIsNotNone(fruit_cat)
            self.assertEqual(fruit_cat["name"], "Fruits")
            count = conn.execute("SELECT COUNT(*) AS c FROM store_products").fetchone()["c"]
            self.assertGreaterEqual(count, 60)
        finally:
            conn.close()

        page = self.client.get("/stores/product-master")
        self.assertEqual(page.status_code, 200)
        self.assertIn(b"Products", page.data)
        self.assertNotIn(b'id="st-products-pagination"', page.data)
        self.assertNotIn(b"st-product-pack-row", page.data)
        self.assertIn(b"Non-Veg", page.data)
        self.assertIn(b"Outlet", page.data)
        self.assertIn(b'id="st-outlet-listbox"', page.data)
        self.assertIn(b"All", page.data)
        self.assertIn(b"Approximate Price", page.data)
        self.assertIn(b"Restaurant", page.data)
        self.assertIn(b"Edit", page.data)
        self.assertIn(b"Delete", page.data)

        bar_page = self.client.get("/stores/product-master?outlet=bar")
        self.assertEqual(bar_page.status_code, 200)
        self.assertIn(b'id="st-outlet-listbox"', bar_page.data)
        self.assertIn(b'data-value="bar"', bar_page.data)
        self.assertRegex(bar_page.data, rb'id="st-outlet-value"[^>]*>\s*Bar')

        conn = db_mod.get_db()
        try:
            product = conn.execute(
                """
                SELECT id, name, category_id, default_unit
                FROM store_products WHERE is_active = 1 ORDER BY id LIMIT 1
                """
            ).fetchone()
            self.assertIsNotNone(product)
            pid = product["id"]
            category_id = product["category_id"]
            unit = product["default_unit"] or "kg"
        finally:
            conn.close()

        edit_page = self.client.get(f"/stores/product-master?edit={pid}")
        self.assertEqual(edit_page.status_code, 200)
        self.assertIn(b"Edit product", edit_page.data)
        self.assertIn(b"Pack variants", edit_page.data)
        self.assertNotIn(b'id="st-product-approx-price"', edit_page.data)

        update = self.client.post(
            "/stores/product-master",
            data={
                "action": "save_product",
                "product_id": str(pid),
                "category_id": str(category_id),
                "name": "Updated Product Name",
                "outlet": "bar",
                "variant_qty": ["1"],
                "variant_unit": [unit],
                "variant_approximate_price": ["250"],
            },
            follow_redirects=True,
        )
        self.assertEqual(update.status_code, 200)
        self.assertIn(b"Product updated", update.data)
        self.assertIn(b"Bar", update.data)
        self.assertIn(b"\xe2\x82\xb9250", update.data)  # ₹250

        conn = db_mod.get_db()
        try:
            saved = conn.execute(
                "SELECT outlet, approximate_price FROM store_products WHERE id = ?", (pid,)
            ).fetchone()
            self.assertEqual(saved["outlet"], "bar")
            self.assertEqual(float(saved["approximate_price"]), 250.0)
        finally:
            conn.close()

        delete = self.client.get(f"/stores/product-master/{pid}/delete", follow_redirects=True)
        self.assertEqual(delete.status_code, 200)
        conn = db_mod.get_db()
        try:
            gone = conn.execute(
                "SELECT is_active FROM store_products WHERE id = ?", (pid,)
            ).fetchone()
            self.assertEqual(int(gone["is_active"]), 0)
        finally:
            conn.close()

    def test_product_delete_json_stays_ok(self):
        conn = db_mod.get_db()
        try:
            category = conn.execute(
                "SELECT id FROM store_product_categories WHERE is_active = 1 ORDER BY id LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(category)
            category_id = category["id"]
            conn.execute(
                """
                INSERT INTO store_products
                  (category_id, name, default_unit, outlet, approximate_price, is_active, created_at, updated_at)
                VALUES (?, 'Ajax Delete Me', 'kg', 'restaurant', 10, 1, datetime('now'), datetime('now'))
                """,
                (category_id,),
            )
            conn.commit()
            pid = conn.execute(
                "SELECT id FROM store_products WHERE lower(name)=lower('Ajax Delete Me') AND is_active=1"
            ).fetchone()["id"]
        finally:
            conn.close()

        resp = self.client.post(
            f"/stores/product-master/{pid}/delete",
            headers={
                "Accept": "application/json",
                "X-Requested-With": "XMLHttpRequest",
            },
        )
        self.assertEqual(resp.status_code, 200)
        payload = resp.get_json()
        self.assertTrue(payload.get("ok"))
        self.assertEqual(int(payload.get("product_id")), pid)
        conn = db_mod.get_db()
        try:
            gone = conn.execute(
                "SELECT is_active FROM store_products WHERE id = ?", (pid,)
            ).fetchone()
            self.assertEqual(int(gone["is_active"]), 0)
        finally:
            conn.close()

    def test_product_preferred_suppliers_save_and_reload(self):
        conn = db_mod.get_db()
        try:
            category = conn.execute(
                "SELECT id FROM store_product_categories WHERE is_active = 1 ORDER BY id LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(category)
            category_id = category["id"]
            conn.execute(
                "INSERT INTO suppliers (name, gst, address, phone) VALUES (?, '', '', '')",
                ("Preferred Alpha",),
            )
            conn.execute(
                "INSERT INTO suppliers (name, gst, address, phone) VALUES (?, '', '', '')",
                ("Preferred Beta",),
            )
            conn.execute(
                "INSERT INTO suppliers (name, gst, address, phone) VALUES (?, '', '', '')",
                ("Preferred Gamma",),
            )
            conn.commit()
            s1, s2, s3 = [
                row["id"]
                for row in conn.execute(
                    """
                    SELECT id FROM suppliers
                    WHERE name IN ('Preferred Alpha', 'Preferred Beta', 'Preferred Gamma')
                    ORDER BY name
                    """
                ).fetchall()
            ]
        finally:
            conn.close()

        create = self.client.post(
            "/stores/product-master",
            data={
                "action": "save_product",
                "category_id": str(category_id),
                "name": "Preferred Supplier Product",
                "outlet": "restaurant",
                "preferred_supplier_1_id": str(s1),
                "preferred_supplier_2_id": str(s2),
                "preferred_supplier_3_id": str(s3),
                "variant_qty": ["1"],
                "variant_unit": ["kg"],
                "variant_approximate_price": ["100"],
            },
            follow_redirects=True,
        )
        self.assertEqual(create.status_code, 200)
        self.assertIn(b"Product added to master", create.data)

        conn = db_mod.get_db()
        try:
            product = conn.execute(
                """
                SELECT id, preferred_supplier_1_id, preferred_supplier_2_id, preferred_supplier_3_id
                FROM store_products
                WHERE lower(name) = lower('Preferred Supplier Product') AND is_active = 1
                """
            ).fetchone()
            self.assertIsNotNone(product)
            pid = product["id"]
            self.assertEqual(int(product["preferred_supplier_1_id"]), s1)
            self.assertEqual(int(product["preferred_supplier_2_id"]), s2)
            self.assertEqual(int(product["preferred_supplier_3_id"]), s3)
        finally:
            conn.close()

        edit_page = self.client.get(f"/stores/product-master?edit={pid}")
        self.assertEqual(edit_page.status_code, 200)
        self.assertIn(b"Supplier 1", edit_page.data)
        self.assertIn(b"Supplier 2", edit_page.data)
        self.assertIn(b"Supplier 3", edit_page.data)
        self.assertIn(b'id="st-product-supplier-1"', edit_page.data)
        self.assertIn(str(s1).encode(), edit_page.data)

    def test_inward_auto_updates_product_preferred_suppliers(self):
        """Last inward supplier becomes preferred #1; cheaper PO history fills #2/#3."""
        from datetime import date

        conn = db_mod.get_db()
        try:
            category = conn.execute(
                "SELECT id FROM store_product_categories WHERE is_active = 1 ORDER BY id LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(category)
            category_id = int(category["id"])
            for name in ("Cheap Co", "Mid Co", "Last Inward Co"):
                conn.execute(
                    "INSERT INTO suppliers (name, gst, address, phone) VALUES (?, '', '', '')",
                    (name,),
                )
            conn.commit()
            ids = {
                row["name"]: int(row["id"])
                for row in conn.execute(
                    """
                    SELECT id, name FROM suppliers
                    WHERE name IN ('Cheap Co', 'Mid Co', 'Last Inward Co')
                    """
                ).fetchall()
            }
            cheap_id = ids["Cheap Co"]
            mid_id = ids["Mid Co"]
            last_id = ids["Last Inward Co"]
            cur = conn.execute(
                """
                INSERT INTO store_products
                    (category_id, name, default_unit, outlet, is_active, sort_order, created_at,
                     preferred_supplier_1_id)
                VALUES (?, 'Auto Pref Tomato', 'kg', 'restaurant', 1, 0, datetime('now','localtime'), ?)
                """,
                (category_id, mid_id),
            )
            product_id = int(cur.lastrowid)
            conn.execute(
                """
                INSERT INTO store_indents
                    (outlet, indent_no, status, notes, created_by, created_at)
                VALUES ('restaurant', 'IND/PREF/TEST', 'approved', '', ?, datetime('now','localtime'))
                """,
                (self.admin_id,),
            )
            indent_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
            for item_supplier, rate in ((cheap_id, 40), (mid_id, 55)):
                conn.execute(
                    """
                    INSERT INTO store_indent_lines
                        (indent_id, item_name, quantity, unit, notes, approximate_price)
                    VALUES (?, 'Auto Pref Tomato', 1, 'kg', '', ?)
                    """,
                    (indent_id, rate),
                )
                line_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
                conn.execute(
                    """
                    INSERT INTO store_po_lines
                        (indent_id, line_id, supplier_id, rate, quantity, updated_at)
                    VALUES (?, ?, ?, ?, 1, datetime('now','localtime'))
                    """,
                    (indent_id, line_id, item_supplier, rate),
                )
            conn.commit()
        finally:
            conn.close()

        confirm = self.client.post(
            "/stores/purchase-requests/confirm-direct-with-expense",
            json={
                "outlet": "restaurant",
                "notes": "",
                "lines": [
                    {
                        "item_name": "Auto Pref Tomato",
                        "qty": 2,
                        "unit": "kg",
                        "unit_price": 60,
                        "tax_percent": 0,
                    }
                ],
                "date": date.today().isoformat(),
                "description": "Pref supplier inward",
                "amount": 120,
                "payment_type": "credit",
                "category": "grocery",
                "invoice_number": "INV-PREF-AUTO-1",
                "supplier_id": last_id,
            },
        )
        self.assertEqual(confirm.status_code, 200, confirm.data)
        self.assertTrue(confirm.get_json().get("ok"), confirm.get_json())

        conn = db_mod.get_db()
        try:
            product = conn.execute(
                """
                SELECT preferred_supplier_1_id, preferred_supplier_2_id, preferred_supplier_3_id
                FROM store_products WHERE id = ?
                """,
                (product_id,),
            ).fetchone()
            self.assertIsNotNone(product)
            self.assertEqual(int(product["preferred_supplier_1_id"]), last_id)
            self.assertEqual(int(product["preferred_supplier_2_id"]), cheap_id)
            self.assertEqual(int(product["preferred_supplier_3_id"]), mid_id)
        finally:
            conn.close()

    def test_product_edit_heals_preferred_supplier_from_last_inward(self):
        """Opening edit backfills Supplier 1 from the newest stock inward."""
        conn = db_mod.get_db()
        try:
            category = conn.execute(
                "SELECT id FROM store_product_categories WHERE is_active = 1 ORDER BY id LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(category)
            category_id = int(category["id"])
            conn.execute(
                "INSERT INTO suppliers (name, gst, address, phone) VALUES (?, '', '', '')",
                ("Heal Pref Co",),
            )
            supplier_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
            cur = conn.execute(
                """
                INSERT INTO store_products
                    (category_id, name, default_unit, outlet, is_active, sort_order, created_at)
                VALUES (?, 'Heal Pref Item', 'kg', 'restaurant', 1, 0, datetime('now','localtime'))
                """,
                (category_id,),
            )
            product_id = int(cur.lastrowid)
            conn.execute(
                """
                INSERT INTO store_indents
                    (outlet, indent_no, status, notes, created_by, created_at)
                VALUES ('restaurant', 'IND/HEAL/PREF', 'approved', '', ?, datetime('now','localtime'))
                """,
                (self.admin_id,),
            )
            indent_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
            conn.execute(
                """
                INSERT INTO store_stock_movements
                    (outlet, item_name, unit, qty_delta, movement_type, ref_type, ref_id,
                     notes, created_by, unit_cost, created_at)
                VALUES (
                    'restaurant', 'Heal Pref Item', 'kg', 1, 'receive', 'stock_inward', ?,
                    'heal test', ?, 77, datetime('now','localtime')
                )
                """,
                (indent_id, self.admin_id),
            )
            conn.execute(
                """
                INSERT INTO sales_update_expenses
                    (company, location, sales_date, category, description, amount, payment_type,
                     supplier_id, invoice_number, created_at)
                VALUES (
                    'HBE', 'Hotel', date('now'), 'grocery',
                    'Stock inward IND/HEAL/PREF', 77, 'credit', ?, 'INV-HEAL-PREF',
                    datetime('now','localtime')
                )
                """,
                (supplier_id,),
            )
            conn.commit()
            before = conn.execute(
                "SELECT preferred_supplier_1_id FROM store_products WHERE id = ?",
                (product_id,),
            ).fetchone()
            self.assertIsNone(before["preferred_supplier_1_id"])
        finally:
            conn.close()

        resp = self.client.get(f"/stores/product-master?edit={product_id}&outlet=restaurant")
        self.assertEqual(resp.status_code, 200, resp.data)
        html = resp.get_data(as_text=True)
        self.assertIn(f'id="st-product-supplier-1"', html)
        self.assertIn(f'value="{supplier_id}"', html)
        self.assertIn("Heal Pref Co", html)

        conn = db_mod.get_db()
        try:
            after = conn.execute(
                "SELECT preferred_supplier_1_id FROM store_products WHERE id = ?",
                (product_id,),
            ).fetchone()
            self.assertEqual(int(after["preferred_supplier_1_id"]), supplier_id)
        finally:
            conn.close()

    def test_product_master_heals_approx_price_from_last_inward(self):
        """Blank Approx Price is filled from the newest stock inward unit cost."""
        conn = db_mod.get_db()
        try:
            category = conn.execute(
                "SELECT id FROM store_product_categories WHERE is_active = 1 ORDER BY id LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(category)
            category_id = int(category["id"])
            cur = conn.execute(
                """
                INSERT INTO store_products
                    (category_id, name, default_unit, outlet, is_active, sort_order, created_at)
                VALUES (?, 'Heal Price Item', 'kg', 'restaurant', 1, 0, datetime('now','localtime'))
                """,
                (category_id,),
            )
            product_id = int(cur.lastrowid)
            conn.execute(
                """
                INSERT INTO store_stock_movements
                    (outlet, item_name, unit, qty_delta, movement_type, ref_type, ref_id,
                     notes, created_by, unit_cost, created_at)
                VALUES (
                    'restaurant', 'Heal Price Item', 'kg', 2, 'receive', 'stock_inward_direct', 1,
                    'heal price', ?, 88.5, datetime('now','localtime')
                )
                """,
                (self.admin_id,),
            )
            conn.commit()
            before = conn.execute(
                "SELECT approximate_price FROM store_products WHERE id = ?",
                (product_id,),
            ).fetchone()
            self.assertIsNone(before["approximate_price"])
        finally:
            conn.close()

        page = self.client.get("/stores/product-master?outlet=restaurant")
        self.assertEqual(page.status_code, 200)
        self.assertIn(b"Heal Price Item", page.data)

        conn = db_mod.get_db()
        try:
            after = conn.execute(
                "SELECT approximate_price FROM store_products WHERE id = ?",
                (product_id,),
            ).fetchone()
            self.assertAlmostEqual(float(after["approximate_price"]), 88.5)
        finally:
            conn.close()

    def test_product_edit_shows_inward_price_on_pack_row(self):
        """Edit modal Pack variants ₹ field shows the latest inward / product rate."""
        conn = db_mod.get_db()
        try:
            category = conn.execute(
                "SELECT id FROM store_product_categories WHERE is_active = 1 ORDER BY id LIMIT 1"
            ).fetchone()
            category_id = int(category["id"])
            cur = conn.execute(
                """
                INSERT INTO store_products
                    (category_id, name, default_unit, outlet, approximate_price,
                     is_active, sort_order, created_at)
                VALUES (?, 'Pack Price Show', 'kg', 'restaurant', 105, 1, 0,
                        datetime('now','localtime'))
                """,
                (category_id,),
            )
            product_id = int(cur.lastrowid)
            conn.commit()
        finally:
            conn.close()

        resp = self.client.get(f"/stores/product-master?edit={product_id}&outlet=restaurant")
        self.assertEqual(resp.status_code, 200, resp.data)
        html = resp.get_data(as_text=True)
        self.assertIn('name="variant_approximate_price"', html)
        self.assertIn('value="105"', html)
        self.assertIn('name="variant_qty"', html)
        self.assertRegex(html, r'name="variant_qty"[^>]*value="1"|value="1"[^>]*name="variant_qty"')

    def test_product_pack_variants_save_and_reload(self):
        conn = db_mod.get_db()
        try:
            category = conn.execute(
                "SELECT id FROM store_product_categories WHERE is_active = 1 ORDER BY id LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(category)
            category_id = category["id"]
        finally:
            conn.close()

        create = self.client.post(
            "/stores/product-master",
            data={
                "action": "save_product",
                "category_id": str(category_id),
                "name": "Test Masala Pack",
                "outlet": "restaurant",
                "variant_qty": ["200", "500"],
                "variant_unit": ["gram", "gram"],
                "variant_approximate_price": ["40", "90"],
            },
            follow_redirects=True,
        )
        self.assertEqual(create.status_code, 200)
        self.assertIn(b"Product added to master", create.data)
        self.assertIn(b"2 packs", create.data)

        conn = db_mod.get_db()
        try:
            product = conn.execute(
                """
                SELECT id, default_unit, approximate_price FROM store_products
                WHERE lower(name) = lower('Test Masala Pack') AND is_active = 1
                """
            ).fetchone()
            self.assertIsNotNone(product)
            pid = product["id"]
            self.assertEqual(product["default_unit"], "kg")
            # 500 gram @ ₹90 → ₹180 / kg (largest pack anchors product price)
            self.assertAlmostEqual(float(product["approximate_price"]), 180.0)
            variants = conn.execute(
                """
                SELECT label, qty_in_base, approximate_price
                FROM store_product_variants
                WHERE product_id = ? AND is_active = 1
                ORDER BY sort_order, id
                """,
                (pid,),
            ).fetchall()
            self.assertEqual(len(variants), 2)
            self.assertEqual(variants[0]["label"], "200 gram")
            self.assertAlmostEqual(float(variants[0]["qty_in_base"]), 0.2)
            self.assertAlmostEqual(float(variants[0]["approximate_price"]), 40.0)
            self.assertEqual(variants[1]["label"], "500 gram")
            self.assertAlmostEqual(float(variants[1]["qty_in_base"]), 0.5)
        finally:
            conn.close()

        edit_page = self.client.get(f"/stores/product-master?edit={pid}")
        self.assertEqual(edit_page.status_code, 200)
        self.assertIn(b'name="variant_qty"', edit_page.data)
        self.assertIn(b'value="200"', edit_page.data)
        self.assertIn(b'value="500"', edit_page.data)
        self.assertIn(b'value="gram"', edit_page.data)
        self.assertIn(b"Pack variants", edit_page.data)
        self.assertNotIn(b'id="st-product-unit-listbox"', edit_page.data)
        self.assertNotIn(b'id="st-product-approx-price"', edit_page.data)

        update = self.client.post(
            "/stores/product-master",
            data={
                "action": "save_product",
                "product_id": str(pid),
                "category_id": str(category_id),
                "name": "Test Masala Pack",
                "outlet": "restaurant",
                "variant_qty": ["500", "1"],
                "variant_unit": ["gram", "kg"],
                "variant_approximate_price": ["90", "170"],
            },
            follow_redirects=True,
        )
        self.assertEqual(update.status_code, 200)
        self.assertIn(b"Product updated", update.data)

        conn = db_mod.get_db()
        try:
            active = conn.execute(
                """
                SELECT label, qty_in_base FROM store_product_variants
                WHERE product_id = ? AND is_active = 1
                ORDER BY sort_order, id
                """,
                (pid,),
            ).fetchall()
            self.assertEqual([row["label"] for row in active], ["500 gram", "1 kg"])
            self.assertAlmostEqual(float(active[0]["qty_in_base"]), 0.5)
            self.assertAlmostEqual(float(active[1]["qty_in_base"]), 1.0)
            inactive = conn.execute(
                """
                SELECT label FROM store_product_variants
                WHERE product_id = ? AND is_active = 0 AND lower(label) = lower('200 gram')
                """,
                (pid,),
            ).fetchone()
            self.assertIsNotNone(inactive)
        finally:
            conn.close()

        # Alphanumeric pack size is allowed and counts as 1 of the pack unit.
        alpha = self.client.post(
            "/stores/product-master",
            data={
                "action": "save_product",
                "category_id": str(category_id),
                "name": "Test Paneer Pack Alpha",
                "outlet": "restaurant",
                "variant_qty": ["Half"],
                "variant_unit": ["kg"],
                "variant_approximate_price": ["120"],
            },
            follow_redirects=True,
        )
        self.assertEqual(alpha.status_code, 200)
        self.assertIn(b"Product added to master", alpha.data)
        conn = db_mod.get_db()
        try:
            alpha_row = conn.execute(
                """
                SELECT v.label, v.qty_in_base
                FROM store_product_variants v
                JOIN store_products p ON p.id = v.product_id
                WHERE lower(p.name) = lower('Test Paneer Pack Alpha')
                  AND v.is_active = 1 AND p.is_active = 1
                """
            ).fetchone()
            self.assertIsNotNone(alpha_row)
            self.assertEqual(alpha_row["label"], "Half kg")
            self.assertAlmostEqual(float(alpha_row["qty_in_base"]), 1.0)
        finally:
            conn.close()

        # Product with no variants still saves cleanly.
        plain = self.client.post(
            "/stores/product-master",
            data={
                "action": "save_product",
                "category_id": str(category_id),
                "name": "Plain No Pack Product",
                "outlet": "bar",
                "variant_qty": [""],
                "variant_unit": ["gram"],
                "variant_approximate_price": [""],
            },
            follow_redirects=True,
        )
        self.assertEqual(plain.status_code, 200)
        self.assertIn(b"Product added to master", plain.data)
        conn = db_mod.get_db()
        try:
            plain_row = conn.execute(
                """
                SELECT id FROM store_products
                WHERE lower(name) = lower('Plain No Pack Product') AND is_active = 1
                """
            ).fetchone()
            count = conn.execute(
                """
                SELECT COUNT(*) AS c FROM store_product_variants
                WHERE product_id = ? AND is_active = 1
                """,
                (plain_row["id"],),
            ).fetchone()["c"]
            self.assertEqual(count, 0)
        finally:
            conn.close()

    def test_indent_pack_converts_to_base_qty_on_stock_receive(self):
        from datetime import date

        conn = db_mod.get_db()
        try:
            category = conn.execute(
                "SELECT id FROM store_product_categories WHERE is_active = 1 ORDER BY id LIMIT 1"
            ).fetchone()
            category_id = category["id"]
            cur = conn.execute(
                """
                INSERT INTO store_products
                    (category_id, name, default_unit, outlet, approximate_price, is_active, sort_order)
                VALUES (?, 'Pack Convert Masala', 'kg', 'bar', 100, 1, 9990)
                """,
                (category_id,),
            )
            product_id = cur.lastrowid
            conn.execute(
                """
                INSERT INTO store_product_variants
                    (product_id, label, qty_in_base, approximate_price, sort_order, is_active)
                VALUES (?, '500 gram', 0.5, 90, 10, 1)
                """,
                (product_id,),
            )
            conn.execute(
                "INSERT INTO suppliers (name) VALUES (?)",
                ("Pack Inward Supplier",),
            )
            supplier_id = conn.execute(
                "SELECT id FROM suppliers WHERE name = 'Pack Inward Supplier'"
            ).fetchone()["id"]
            conn.execute(
                "INSERT INTO cash_ledger_loads (company, load_date, description, amount) VALUES (?,?,?,?)",
                ("HBE", date.today().isoformat(), "Pack test float", 5000),
            )
            conn.commit()
        finally:
            conn.close()

        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Pack inward",
                "item_name": ["Pack Convert Masala"],
                "quantity": ["2"],
                "unit": ["kg"],
                "approximate_price": ["90"],
                "pack_label": ["500 gram"],
                "pack_qty_in_base": ["0.5"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)

        conn = db_mod.get_db()
        try:
            indent = conn.execute(
                "SELECT id FROM store_indents WHERE notes = 'Pack inward' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(indent)
            indent_id = indent["id"]
            line = conn.execute(
                """
                SELECT id, quantity, unit, pack_label, pack_qty_in_base
                FROM store_indent_lines WHERE indent_id = ?
                """,
                (indent_id,),
            ).fetchone()
            self.assertEqual(float(line["quantity"]), 2.0)
            self.assertEqual(line["unit"], "kg")
            self.assertEqual(line["pack_label"], "500 gram")
            self.assertAlmostEqual(float(line["pack_qty_in_base"]), 0.5)
            line_id = int(line["id"])
        finally:
            conn.close()

        decide = self.client.post(
            f"/stores/indent/{indent_id}/decide",
            data={"outlet": "bar", "decision": "approved", "decision_note": ""},
            follow_redirects=False,
        )
        self.assertEqual(decide.status_code, 302)

        confirm = self.client.post(
            "/stores/purchase-requests/confirm-with-expense",
            json={
                "indent_id": indent_id,
                "notes": "",
                "lines": [
                    {
                        "line_id": line_id,
                        "received_qty": 2,
                        "unit_price": 95,
                        "tax_percent": 0,
                    }
                ],
                "date": date.today().isoformat(),
                "description": "Pack stock inward",
                "amount": 190,
                "payment_type": "cash",
                "category": "grocery",
                "supplier_id": supplier_id,
            },
        )
        self.assertEqual(confirm.status_code, 200, confirm.data)
        payload = confirm.get_json()
        self.assertTrue(payload.get("ok"), payload)

        conn = db_mod.get_db()
        try:
            stock = conn.execute(
                """
                SELECT qty_on_hand, unit FROM store_stock_items
                WHERE outlet = 'bar' AND lower(item_name) = lower('Pack Convert Masala')
                """
            ).fetchone()
            self.assertIsNotNone(stock)
            self.assertEqual(stock["unit"], "kg")
            # 2 packs × 0.5 kg = 1 kg stock
            self.assertAlmostEqual(float(stock["qty_on_hand"]), 1.0)
            movement = conn.execute(
                """
                SELECT qty_delta, unit_cost FROM store_stock_movements
                WHERE outlet = 'bar' AND lower(item_name) = lower('Pack Convert Masala')
                ORDER BY id DESC LIMIT 1
                """
            ).fetchone()
            self.assertAlmostEqual(float(movement["qty_delta"]), 1.0)
            # ₹95 / pack ÷ 0.5 kg = ₹190 per kg
            self.assertAlmostEqual(float(movement["unit_cost"]), 190.0)
            variant = conn.execute(
                """
                SELECT approximate_price FROM store_product_variants
                WHERE product_id = ? AND lower(label) = lower('500 gram') AND is_active = 1
                """,
                (product_id,),
            ).fetchone()
            self.assertAlmostEqual(float(variant["approximate_price"]), 95.0)
            product = conn.execute(
                "SELECT approximate_price FROM store_products WHERE id = ?",
                (product_id,),
            ).fetchone()
            self.assertAlmostEqual(float(product["approximate_price"]), 190.0)
        finally:
            conn.close()

        # Ordering in base unit (no pack) still receives 1:1 into stock.
        create2 = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "No pack inward",
                "item_name": ["Pack Convert Masala"],
                "quantity": ["3"],
                "unit": ["kg"],
                "approximate_price": ["100"],
                "pack_label": [""],
                "pack_qty_in_base": [""],
            },
            follow_redirects=False,
        )
        self.assertEqual(create2.status_code, 302)
        conn = db_mod.get_db()
        try:
            indent2 = conn.execute(
                "SELECT id FROM store_indents WHERE notes = 'No pack inward' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            indent2_id = indent2["id"]
            line2 = conn.execute(
                """
                SELECT id FROM store_indent_lines WHERE indent_id = ?
                """,
                (indent2_id,),
            ).fetchone()
            line2_id = int(line2["id"])
            before = conn.execute(
                """
                SELECT COALESCE(qty_on_hand, 0) AS qty FROM store_stock_items
                WHERE outlet = 'bar' AND lower(item_name) = lower('Pack Convert Masala') AND lower(unit) = 'kg'
                """
            ).fetchone()
            before_qty = float(before["qty"]) if before else 0.0
        finally:
            conn.close()
        self.client.post(
            f"/stores/indent/{indent2_id}/decide",
            data={"outlet": "bar", "decision": "approved", "decision_note": ""},
            follow_redirects=False,
        )
        confirm2 = self.client.post(
            "/stores/purchase-requests/confirm-with-expense",
            json={
                "indent_id": indent2_id,
                "notes": "",
                "lines": [
                    {
                        "line_id": line2_id,
                        "received_qty": 3,
                        "unit_price": 100,
                        "tax_percent": 0,
                    }
                ],
                "date": date.today().isoformat(),
                "description": "Base unit inward",
                "amount": 300,
                "payment_type": "cash",
                "category": "grocery",
                "supplier_id": supplier_id,
            },
        )
        self.assertEqual(confirm2.status_code, 200, confirm2.data)
        self.assertTrue(confirm2.get_json().get("ok"), confirm2.get_json())
        conn = db_mod.get_db()
        try:
            after = conn.execute(
                """
                SELECT qty_on_hand FROM store_stock_items
                WHERE outlet = 'bar' AND lower(item_name) = lower('Pack Convert Masala') AND lower(unit) = 'kg'
                """
            ).fetchone()
            self.assertAlmostEqual(float(after["qty_on_hand"]), before_qty + 3.0)
        finally:
            conn.close()

    def test_indent_submit_requires_quantity_for_each_item(self):
        resp = self.client.post(
            "/stores/indent?outlet=bar&focus=form",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Missing qty",
                "item_name": ["Onion", "Potato"],
                "quantity": ["10", ""],
                "unit": ["kg", "kg"],
                "approximate_price": ["30", "20"],
            },
            follow_redirects=True,
        )
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"Enter a quantity greater than 0 for each item.", resp.data)
        conn = db_mod.get_db()
        try:
            row = conn.execute(
                "SELECT id FROM store_indents WHERE notes = 'Missing qty'"
            ).fetchone()
            self.assertIsNone(row)
        finally:
            conn.close()

    def test_indent_to_stock_happy_path(self):
        # Create + submit indent
        resp = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Evening bar needs",
                "item_name": ["Onion", "Potato"],
                "quantity": ["10", "24"],
                "unit": ["kg", "kg"],
                "approximate_price": ["30", "20"],
            },
            follow_redirects=False,
        )
        self.assertEqual(resp.status_code, 302)

        conn = db_mod.get_db()
        try:
            indent = conn.execute(
                "SELECT * FROM store_indents WHERE outlet = 'bar' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(indent)
            self.assertEqual(indent["status"], "pending")
            indent_id = indent["id"]
        finally:
            conn.close()

        # Approve
        resp = self.client.post(
            f"/stores/indent/{indent_id}/decide",
            data={"outlet": "bar", "decision": "approved", "decision_note": ""},
            follow_redirects=False,
        )
        self.assertEqual(resp.status_code, 302)

        # Create PR from indent
        resp = self.client.post(
            "/stores/purchase-requests?outlet=bar",
            data={
                "outlet": "bar",
                "action": "create_from_indent",
                "indent_id": str(indent_id),
            },
            follow_redirects=False,
        )
        self.assertEqual(resp.status_code, 302)

        conn = db_mod.get_db()
        try:
            pr = conn.execute(
                "SELECT * FROM store_purchase_requests WHERE indent_id = ?",
                (indent_id,),
            ).fetchone()
            self.assertIsNotNone(pr)
            pr_id = pr["id"]
        finally:
            conn.close()

        # Receive into stock
        resp = self.client.post(
            f"/stores/purchase-requests/{pr_id}/receive",
            data={"outlet": "bar"},
            follow_redirects=False,
        )
        self.assertEqual(resp.status_code, 302)

        conn = db_mod.get_db()
        try:
            stock = {
                (row["item_name"], row["unit"]): row["qty_on_hand"]
                for row in conn.execute(
                    "SELECT item_name, unit, qty_on_hand FROM store_stock_items WHERE outlet = 'bar'"
                ).fetchall()
            }
            self.assertEqual(stock[("Onion", "kg")], 10)
            self.assertEqual(stock[("Potato", "kg")], 24)
        finally:
            conn.close()

        # Pages render
        for path in (
            "/stores/indent?outlet=bar",
            "/stores/approvals?outlet=bar",
            "/stores/purchase-requests?outlet=bar",
            "/stores/stock?outlet=bar",
            "/stores?outlet=restaurant",
            "/stores?outlet=kitchen",  # legacy alias → restaurant
        ):
            page = self.client.get(path, follow_redirects=True)
            self.assertEqual(page.status_code, 200, path)

        inward_page = self.client.get("/stores/purchase-requests?outlet=bar", follow_redirects=True)
        self.assertIn(b"Stock Inward", inward_page.data)

    def test_stock_inward_confirms_into_stock_with_cash_expense(self):
        from datetime import date

        conn = db_mod.get_db()
        try:
            conn.execute(
                "INSERT INTO suppliers (name) VALUES (?)",
                ("Inward Cash Supplier",),
            )
            supplier_id = conn.execute(
                "SELECT id FROM suppliers WHERE name = 'Inward Cash Supplier'"
            ).fetchone()["id"]
            conn.execute(
                "INSERT INTO cash_ledger_loads (company, load_date, description, amount) VALUES (?,?,?,?)",
                ("HBE", date.today().isoformat(), "Test cash float", 5000),
            )
            conn.commit()
        finally:
            conn.close()

        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Inward me",
                "item_name": ["Onion", "Potato"],
                "quantity": ["10", "24"],
                "unit": ["kg", "kg"],
                "approximate_price": ["30", "20"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)
        conn = db_mod.get_db()
        try:
            indent = conn.execute(
                "SELECT id, indent_no FROM store_indents WHERE notes = 'Inward me' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(indent)
            indent_id = indent["id"]
            lines = conn.execute(
                "SELECT id, item_name, quantity FROM store_indent_lines WHERE indent_id = ? ORDER BY id",
                (indent_id,),
            ).fetchall()
            self.assertEqual(len(lines), 2)
            line_ids = [int(row["id"]) for row in lines]
        finally:
            conn.close()

        decide = self.client.post(
            f"/stores/indent/{indent_id}/decide",
            data={"outlet": "bar", "decision": "approved", "decision_note": ""},
            follow_redirects=False,
        )
        self.assertEqual(decide.status_code, 302)
        po_id = self._generate_po_for_supplier(
            indent_id,
            supplier_id,
            "bar",
            line_ids,
            rates={line_ids[0]: "30", line_ids[1]: "20"},
        )

        page = self.client.get(
            f"/stores/purchase-requests?outlet=bar&indent={indent_id}",
            follow_redirects=True,
        )
        self.assertEqual(page.status_code, 200)
        self.assertIn(b"Confirm purchase", page.data)
        self.assertIn(b"st-inward-expense-modal", page.data)

        form_blocked = self.client.post(
            "/stores/purchase-requests?outlet=bar",
            data={
                "outlet": "bar",
                "action": "confirm_stock_inward",
                "indent_id": str(indent_id),
                "notes": "",
                "selected_line": [str(line_ids[0]), str(line_ids[1])],
                f"received_qty_{line_ids[0]}": "10",
                f"received_qty_{line_ids[1]}": "20",
            },
            follow_redirects=True,
        )
        self.assertEqual(form_blocked.status_code, 200)
        self.assertIn(b"purchase dialog", form_blocked.data)

        empty = self.client.post(
            "/stores/purchase-requests/confirm-with-expense",
            json={
                "indent_id": indent_id,
                "lines": [],
                "date": date.today().isoformat(),
                "description": "Stock inward",
                "amount": 700,
                "payment_type": "cash",
                "category": "grocery",
                "supplier_id": supplier_id,
            },
        )
        self.assertEqual(empty.status_code, 400)
        self.assertIn(b"Select at least one item", empty.data)

        confirm = self.client.post(
            "/stores/purchase-requests/confirm-with-expense",
            json={
                "indent_id": indent_id,
                "po_id": po_id,
                "notes": "Received evening delivery",
                "lines": [
                    {"line_id": line_ids[0], "received_qty": 10},
                    {"line_id": line_ids[1], "received_qty": 24},
                ],
                "date": date.today().isoformat(),
                "description": "Stock inward cash",
                "amount": 780,
                "payment_type": "cash",
                "category": "grocery",
                "supplier_id": supplier_id,
            },
        )
        self.assertEqual(confirm.status_code, 200)
        payload = confirm.get_json()
        self.assertTrue(payload.get("ok"))
        self.assertIn("/stores/stock", payload.get("redirect", ""))

        conn = db_mod.get_db()
        try:
            status = conn.execute(
                "SELECT status FROM store_indents WHERE id = ?", (indent_id,)
            ).fetchone()["status"]
            self.assertEqual(status, "stocked")
            stock = {
                (row["item_name"], row["unit"]): float(row["qty_on_hand"])
                for row in conn.execute(
                    """
                    SELECT item_name, unit, qty_on_hand FROM store_stock_items
                    WHERE outlet = 'bar' AND place = 'warehouse'
                    """
                ).fetchall()
            }
            self.assertEqual(stock[("Onion", "kg")], 10.0)
            self.assertEqual(stock[("Potato", "kg")], 24.0)
            counter_count = conn.execute(
                """
                SELECT COUNT(*) AS c FROM store_stock_items
                WHERE outlet = 'bar' AND place = 'counter'
                  AND lower(item_name) IN ('onion', 'potato')
                """
            ).fetchone()["c"]
            self.assertEqual(int(counter_count), 0)
            movement = conn.execute(
                """
                SELECT COUNT(*) AS c FROM store_stock_movements
                WHERE ref_type = 'stock_inward' AND ref_id = ?
                """,
                (indent_id,),
            ).fetchone()["c"]
            self.assertEqual(int(movement), 2)
            expense = conn.execute(
                """
                SELECT id, amount, payment_type, location, description
                FROM sales_update_expenses
                WHERE id = ?
                """,
                (payload["expense_id"],),
            ).fetchone()
            self.assertIsNotNone(expense)
            self.assertEqual(float(expense["amount"]), 780.0)
            self.assertEqual(expense["payment_type"], "cash")
            self.assertEqual(expense["location"], "Hotel")
            self.assertIn("Stock inward cash", expense["description"])
        finally:
            conn.close()

        gone = self.client.get(
            f"/stores/purchase-requests?outlet=bar&indent={indent_id}",
            follow_redirects=True,
        )
        self.assertEqual(gone.status_code, 200)
        self.assertIn(b"No purchase orders yet", gone.data)

    def test_stock_inward_partial_keeps_remaining_on_page(self):
        from datetime import date

        conn = db_mod.get_db()
        try:
            conn.execute(
                "INSERT INTO suppliers (name) VALUES (?)",
                ("Inward Partial Supplier",),
            )
            supplier_id = conn.execute(
                "SELECT id FROM suppliers WHERE name = 'Inward Partial Supplier'"
            ).fetchone()["id"]
            conn.execute(
                "INSERT INTO cash_ledger_loads (company, load_date, description, amount) VALUES (?,?,?,?)",
                ("HBE", date.today().isoformat(), "Partial inward float", 5000),
            )
            conn.commit()
        finally:
            conn.close()

        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Partial inward indent",
                "item_name": ["Onion", "Potato"],
                "quantity": ["10", "24"],
                "unit": ["kg", "kg"],
                "approximate_price": ["30", "20"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)
        conn = db_mod.get_db()
        try:
            indent = conn.execute(
                "SELECT id FROM store_indents WHERE notes = 'Partial inward indent' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            indent_id = indent["id"]
            lines = conn.execute(
                "SELECT id, item_name FROM store_indent_lines WHERE indent_id = ? ORDER BY id",
                (indent_id,),
            ).fetchall()
            line_ids = [int(row["id"]) for row in lines]
        finally:
            conn.close()

        decide = self.client.post(
            f"/stores/indent/{indent_id}/decide",
            data={"outlet": "bar", "decision": "approved", "decision_note": ""},
            follow_redirects=False,
        )
        self.assertEqual(decide.status_code, 302)
        po_id = self._generate_po_for_supplier(
            indent_id,
            supplier_id,
            "bar",
            line_ids,
            rates={line_ids[0]: "30", line_ids[1]: "20"},
        )

        # Receive only first line fully; leave Potato pending.
        partial = self.client.post(
            "/stores/purchase-requests/confirm-with-expense",
            json={
                "indent_id": indent_id,
                "po_id": po_id,
                "notes": "First delivery",
                "lines": [{"line_id": line_ids[0], "received_qty": 10}],
                "date": date.today().isoformat(),
                "description": "Partial stock inward",
                "amount": 300,
                "payment_type": "cash",
                "category": "grocery",
                "supplier_id": supplier_id,
            },
        )
        self.assertEqual(partial.status_code, 200)
        payload = partial.get_json()
        self.assertTrue(payload.get("ok"))
        self.assertTrue(payload.get("partial"))
        self.assertIn("/stores/purchase-requests", payload.get("redirect", ""))

        conn = db_mod.get_db()
        try:
            status = conn.execute(
                "SELECT status FROM store_indents WHERE id = ?", (indent_id,)
            ).fetchone()["status"]
            self.assertEqual(status, "approved")
            rows = {
                int(row["id"]): float(row["quantity_received"] or 0)
                for row in conn.execute(
                    "SELECT id, quantity_received FROM store_indent_lines WHERE indent_id = ?",
                    (indent_id,),
                ).fetchall()
            }
            self.assertEqual(rows[line_ids[0]], 10.0)
            self.assertEqual(rows[line_ids[1]], 0.0)
        finally:
            conn.close()

        page = self.client.get(
            f"/stores/purchase-requests?outlet=bar&indent={indent_id}",
            follow_redirects=True,
        )
        self.assertEqual(page.status_code, 200)
        self.assertIn(b'st-inward-item-name">Potato', page.data)
        self.assertNotIn(b'st-inward-item-name">Onion', page.data)

        # Over-remaining should fail.
        over = self.client.post(
            "/stores/purchase-requests/confirm-with-expense",
            json={
                "indent_id": indent_id,
                "po_id": po_id,
                "lines": [{"line_id": line_ids[1], "received_qty": 25}],
                "date": date.today().isoformat(),
                "description": "Too much",
                "amount": 500,
                "payment_type": "cash",
                "category": "grocery",
                "supplier_id": supplier_id,
            },
        )
        self.assertEqual(over.status_code, 400)

        # Finish remaining.
        finish = self.client.post(
            "/stores/purchase-requests/confirm-with-expense",
            json={
                "indent_id": indent_id,
                "po_id": po_id,
                "lines": [{"line_id": line_ids[1], "received_qty": 24}],
                "date": date.today().isoformat(),
                "description": "Remainder stock inward",
                "amount": 480,
                "payment_type": "cash",
                "category": "grocery",
                "supplier_id": supplier_id,
            },
        )
        self.assertEqual(finish.status_code, 200)
        finish_payload = finish.get_json()
        self.assertTrue(finish_payload.get("ok"))
        self.assertFalse(finish_payload.get("partial"))
        self.assertIn("/stores/stock", finish_payload.get("redirect", ""))

        conn = db_mod.get_db()
        try:
            status = conn.execute(
                "SELECT status FROM store_indents WHERE id = ?", (indent_id,)
            ).fetchone()["status"]
            self.assertEqual(status, "stocked")
        finally:
            conn.close()

        gone = self.client.get(
            f"/stores/purchase-requests?outlet=bar&indent={indent_id}",
            follow_redirects=True,
        )
        self.assertEqual(gone.status_code, 200)
        self.assertIn(b"No purchase orders yet", gone.data)

    def test_stock_inward_confirms_into_stock_with_credit_expense(self):
        from datetime import date

        conn = db_mod.get_db()
        try:
            conn.execute(
                "INSERT INTO suppliers (name) VALUES (?)",
                ("Inward Credit Supplier",),
            )
            supplier_id = conn.execute(
                "SELECT id FROM suppliers WHERE name = 'Inward Credit Supplier'"
            ).fetchone()["id"]
            conn.commit()
        finally:
            conn.close()

        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Inward credit",
                "item_name": ["Tomato"],
                "quantity": ["5"],
                "unit": ["kg"],
                "approximate_price": ["40"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)
        conn = db_mod.get_db()
        try:
            indent = conn.execute(
                "SELECT id FROM store_indents WHERE notes = 'Inward credit' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            indent_id = indent["id"]
            line_id = conn.execute(
                "SELECT id FROM store_indent_lines WHERE indent_id = ?",
                (indent_id,),
            ).fetchone()["id"]
        finally:
            conn.close()

        decide = self.client.post(
            f"/stores/indent/{indent_id}/decide",
            data={"outlet": "bar", "decision": "approved", "decision_note": ""},
            follow_redirects=False,
        )
        self.assertEqual(decide.status_code, 302)

        confirm = self.client.post(
            "/stores/purchase-requests/confirm-with-expense",
            json={
                "indent_id": indent_id,
                "notes": "credit delivery",
                "lines": [{"line_id": int(line_id), "received_qty": 5}],
                "date": date.today().isoformat(),
                "description": "Stock inward credit",
                "amount": 200,
                "payment_type": "credit",
                "category": "grocery",
                "supplier_id": supplier_id,
            },
        )
        self.assertEqual(confirm.status_code, 200)
        payload = confirm.get_json()
        self.assertTrue(payload.get("ok"))
        expense_id = payload["expense_id"]

        conn = db_mod.get_db()
        try:
            status = conn.execute(
                "SELECT status FROM store_indents WHERE id = ?", (indent_id,)
            ).fetchone()["status"]
            self.assertEqual(status, "stocked")
            expense = conn.execute(
                "SELECT amount, payment_type, location FROM sales_update_expenses WHERE id = ?",
                (expense_id,),
            ).fetchone()
            self.assertEqual(float(expense["amount"]), 200.0)
            self.assertEqual(expense["payment_type"], "credit")
            self.assertEqual(expense["location"], "Hotel")
            verified = conn.execute(
                """
                SELECT COALESCE(SUM(amount), 0) AS total
                FROM purchase_verification_allocations
                WHERE expense_id = ?
                """,
                (expense_id,),
            ).fetchone()["total"]
            self.assertEqual(float(verified), 200.0)
            pending = self.app_mod._pending_purchase_verifications(conn)
            self.assertFalse(any(int(row["id"]) == int(expense_id) for row in pending))
            outstanding = self.app_mod._outstanding_credit_expenses(conn)
            match = [row for row in outstanding if int(row["id"]) == int(expense_id)]
            self.assertEqual(len(match), 1)
            self.assertEqual(float(match[0]["balance"]), 200.0)
            stock_qty = conn.execute(
                """
                SELECT qty_on_hand FROM store_stock_items
                WHERE outlet = 'bar' AND item_name = 'Tomato' AND unit = 'kg'
                """
            ).fetchone()["qty_on_hand"]
            self.assertEqual(float(stock_qty), 5.0)
        finally:
            conn.close()

    def test_stock_inward_credit_over_approved_goes_to_purchase_verification(self):
        from datetime import date

        conn = db_mod.get_db()
        try:
            conn.execute(
                "INSERT INTO suppliers (name) VALUES (?)",
                ("Inward Over Approved Supplier",),
            )
            supplier_id = conn.execute(
                "SELECT id FROM suppliers WHERE name = 'Inward Over Approved Supplier'"
            ).fetchone()["id"]
            conn.commit()
        finally:
            conn.close()

        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Inward over approved",
                "item_name": ["Potato"],
                "quantity": ["5"],
                "unit": ["kg"],
                "approximate_price": ["40"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)
        conn = db_mod.get_db()
        try:
            indent = conn.execute(
                "SELECT id FROM store_indents WHERE notes = 'Inward over approved' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            indent_id = indent["id"]
            line_id = conn.execute(
                "SELECT id FROM store_indent_lines WHERE indent_id = ?",
                (indent_id,),
            ).fetchone()["id"]
        finally:
            conn.close()

        decide = self.client.post(
            f"/stores/indent/{indent_id}/decide",
            data={"outlet": "bar", "decision": "approved", "decision_note": ""},
            follow_redirects=False,
        )
        self.assertEqual(decide.status_code, 302)

        # Approved total = 5 × 40 = 200; entered rate 40.2 → 201 must not auto-verify.
        confirm = self.client.post(
            "/stores/purchase-requests/confirm-with-expense",
            json={
                "indent_id": indent_id,
                "notes": "over price delivery",
                "lines": [
                    {
                        "line_id": int(line_id),
                        "received_qty": 5,
                        "unit_price": 40.2,
                        "tax_percent": 0,
                    }
                ],
                "date": date.today().isoformat(),
                "description": "Stock inward over approved",
                "amount": 201,
                "payment_type": "credit",
                "supplier_id": supplier_id,
            },
        )
        self.assertEqual(confirm.status_code, 200)
        payload = confirm.get_json()
        self.assertTrue(payload.get("ok"))
        expense_id = payload["expense_id"]

        conn = db_mod.get_db()
        try:
            status = conn.execute(
                "SELECT status FROM store_indents WHERE id = ?", (indent_id,)
            ).fetchone()["status"]
            self.assertEqual(status, "stocked")
            expense = conn.execute(
                "SELECT amount, payment_type FROM sales_update_expenses WHERE id = ?",
                (expense_id,),
            ).fetchone()
            self.assertEqual(float(expense["amount"]), 201.0)
            self.assertEqual(expense["payment_type"], "credit")
            verified = conn.execute(
                """
                SELECT COALESCE(SUM(amount), 0) AS total
                FROM purchase_verification_allocations
                WHERE expense_id = ?
                """,
                (expense_id,),
            ).fetchone()["total"]
            self.assertEqual(float(verified), 0.0)
            pending = self.app_mod._pending_purchase_verifications(conn)
            self.assertTrue(any(int(row["id"]) == int(expense_id) for row in pending))
            outstanding = self.app_mod._outstanding_credit_expenses(conn)
            self.assertFalse(any(int(row["id"]) == int(expense_id) for row in outstanding))
        finally:
            conn.close()

    def test_stock_inward_rejects_non_approved(self):
        from datetime import date

        conn = db_mod.get_db()
        try:
            conn.execute(
                "INSERT INTO suppliers (name) VALUES (?)",
                ("Inward Reject Supplier",),
            )
            supplier_id = conn.execute(
                "SELECT id FROM suppliers WHERE name = 'Inward Reject Supplier'"
            ).fetchone()["id"]
            conn.commit()
        finally:
            conn.close()

        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Still pending",
                "item_name": ["Onion"],
                "quantity": ["1"],
                "unit": ["kg"],
                "approximate_price": ["10"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)
        conn = db_mod.get_db()
        try:
            indent = conn.execute(
                "SELECT id FROM store_indents WHERE notes = 'Still pending' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            indent_id = indent["id"]
            line_id = conn.execute(
                "SELECT id FROM store_indent_lines WHERE indent_id = ?",
                (indent_id,),
            ).fetchone()["id"]
        finally:
            conn.close()

        bad = self.client.post(
            "/stores/purchase-requests/confirm-with-expense",
            json={
                "indent_id": indent_id,
                "lines": [{"line_id": int(line_id), "received_qty": 1}],
                "date": date.today().isoformat(),
                "description": "Should fail",
                "amount": 10,
                "payment_type": "credit",
                "category": "grocery",
                "supplier_id": supplier_id,
            },
        )
        self.assertEqual(bad.status_code, 400)
        self.assertIn(b"Select an approved indent", bad.data)

    def test_stock_inward_mode_tabs_and_direct_confirm(self):
        from datetime import date

        approved = self.client.get("/stores/purchase-requests?outlet=bar&view=approved")
        self.assertEqual(approved.status_code, 200)
        self.assertIn(b"Indent Approved", approved.data)
        self.assertIn(b"Without Indent Approval", approved.data)
        self.assertIn(b'id="st-inward-indent-listbox"', approved.data)
        self.assertIn(b"Purchase Order", approved.data)
        self.assertIn(b"Select purchase order", approved.data)
        self.assertIn(b'cp-view-tab is-active', approved.data)
        # Stock Inward outlet filter includes All / Bar / Restaurant.
        self.assertIn(b'data-value="both"', approved.data)
        self.assertIn(b">All</button>", approved.data)
        self.assertIn(b'data-value="bar"', approved.data)
        self.assertIn(b'data-value="restaurant"', approved.data)

        # ?outlet=both stays on All (no redirect to Bar).
        both_page = self.client.get(
            "/stores/purchase-requests?outlet=both&view=direct",
            follow_redirects=False,
        )
        self.assertEqual(both_page.status_code, 200)
        self.assertIn(b'data-value="both"', both_page.data)
        self.assertIn(b">All</button>", both_page.data)

        direct = self.client.get("/stores/purchase-requests?outlet=bar&view=direct")
        self.assertEqual(direct.status_code, 200)
        self.assertNotIn(b'id="st-inward-indent-listbox"', direct.data)
        self.assertIn(b"Invoice Items", direct.data)
        self.assertIn(b'data-st-inward-view="direct"', direct.data)
        self.assertIn(b"st-inward-direct-line", direct.data)

        missing_outlet = self.client.post(
            "/stores/purchase-requests/confirm-direct-with-expense",
            json={
                "outlet": "both",
                "lines": [{"item_name": "Onion", "qty": 1, "unit": "kg", "unit_price": 20}],
                "date": date.today().isoformat(),
                "description": "Direct inward",
                "amount": 20,
                "payment_type": "credit",
                "category": "grocery",
                "supplier_id": 1,
            },
        )
        self.assertEqual(missing_outlet.status_code, 400)
        self.assertIn(b"Choose Bar or Restaurant", missing_outlet.data)

        empty_lines = self.client.post(
            "/stores/purchase-requests/confirm-direct-with-expense",
            json={
                "outlet": "restaurant",
                "lines": [],
                "date": date.today().isoformat(),
                "description": "Direct inward",
                "amount": 20,
                "payment_type": "credit",
                "category": "grocery",
                "supplier_id": 1,
            },
        )
        self.assertEqual(empty_lines.status_code, 400)

        conn = db_mod.get_db()
        try:
            conn.execute(
                "INSERT INTO suppliers (name) VALUES (?)",
                ("Direct Inward Supplier",),
            )
            supplier_id = conn.execute(
                "SELECT id FROM suppliers WHERE name = 'Direct Inward Supplier'"
            ).fetchone()["id"]
            before = conn.execute(
                """
                SELECT COALESCE(qty_on_hand, 0) AS qty FROM store_stock_items
                WHERE outlet = 'restaurant' AND item_name = 'Onion' AND unit = 'kg'
                """
            ).fetchone()
            before_qty = float(before["qty"]) if before else 0.0
            conn.commit()
        finally:
            conn.close()

        # Seeded Product Master maps Onion to restaurant (not bar).
        confirm = self.client.post(
            "/stores/purchase-requests/confirm-direct-with-expense",
            json={
                "outlet": "restaurant",
                "notes": "No indent",
                "lines": [
                    {
                        "item_name": "Onion",
                        "qty": 3,
                        "unit": "kg",
                        "unit_price": 25,
                        "tax_percent": 0,
                    }
                ],
                "date": date.today().isoformat(),
                "description": "Stock inward without indent approval",
                "amount": 75,
                "payment_type": "credit",
                "category": "grocery",
                "invoice_number": "INV-DIRECT-1",
                "supplier_id": supplier_id,
            },
        )
        self.assertEqual(confirm.status_code, 200, confirm.data)
        payload = confirm.get_json()
        self.assertTrue(payload.get("ok"))
        self.assertIn("/stores/stock", payload.get("redirect", ""))
        expense_id = payload.get("expense_id")
        self.assertTrue(expense_id)
        expenses = payload.get("expenses") or []
        self.assertEqual(len(expenses), 1)
        self.assertEqual(expenses[0].get("expense_id"), expense_id)

        conn = db_mod.get_db()
        try:
            stock = conn.execute(
                """
                SELECT qty_on_hand FROM store_stock_items
                WHERE outlet = 'restaurant' AND item_name = 'Onion' AND unit = 'kg'
                """
            ).fetchone()
            self.assertIsNotNone(stock)
            self.assertAlmostEqual(float(stock["qty_on_hand"]), before_qty + 3.0)
            movement = conn.execute(
                """
                SELECT ref_type, ref_id FROM store_stock_movements
                WHERE ref_type = 'stock_inward_direct' AND ref_id = ?
                ORDER BY id DESC LIMIT 1
                """,
                (int(expense_id),),
            ).fetchone()
            self.assertIsNotNone(movement)
            verified = conn.execute(
                """
                SELECT COALESCE(SUM(amount), 0) AS total
                FROM purchase_verification_allocations
                WHERE expense_id = ?
                """,
                (int(expense_id),),
            ).fetchone()["total"]
            self.assertEqual(float(verified), 0.0)
            pending = self.app_mod._pending_purchase_verifications(conn)
            self.assertTrue(any(int(row["id"]) == int(expense_id) for row in pending))
            expense = conn.execute(
                "SELECT description, invoice_number FROM sales_update_expenses WHERE id = ?",
                (int(expense_id),),
            ).fetchone()
            self.assertIsNotNone(expense)
            self.assertIn("without indent", (expense["description"] or "").lower())
            self.assertEqual(expense["invoice_number"] or "", "INV-DIRECT-1")
        finally:
            conn.close()

    def test_direct_stock_inward_multi_category_expenses(self):
        from datetime import date

        conn = db_mod.get_db()
        try:
            conn.execute(
                "INSERT INTO suppliers (name) VALUES (?)",
                ("Multi Cat Supplier",),
            )
            supplier_id = conn.execute(
                "SELECT id FROM suppliers WHERE name = 'Multi Cat Supplier'"
            ).fetchone()["id"]
            before_rows = {
                row["item_name"]: float(row["qty"] or 0)
                for row in conn.execute(
                    """
                    SELECT item_name, COALESCE(qty_on_hand, 0) AS qty
                    FROM store_stock_items
                    WHERE outlet = 'restaurant'
                      AND item_name IN ('Chicken Whole', 'Butter')
                      AND unit = 'kg'
                    """
                ).fetchall()
            }
            conn.commit()
        finally:
            conn.close()

        confirm = self.client.post(
            "/stores/purchase-requests/confirm-direct-with-expense",
            json={
                "outlet": "restaurant",
                "notes": "Split categories",
                "lines": [
                    {
                        "item_name": "Chicken Whole",
                        "qty": 2,
                        "unit": "kg",
                        "unit_price": 1000,
                        "tax_percent": 0,
                    },
                    {
                        "item_name": "Butter",
                        "qty": 1,
                        "unit": "kg",
                        "unit_price": 500,
                        "tax_percent": 0,
                    },
                ],
                "date": date.today().isoformat(),
                "description": "Stock inward without indent approval",
                "amount": 2500,
                "payment_type": "credit",
                "invoice_number": "INV-MULTI-1",
                "supplier_id": supplier_id,
            },
        )
        self.assertEqual(confirm.status_code, 200, confirm.data)
        payload = confirm.get_json()
        self.assertTrue(payload.get("ok"))
        expenses = payload.get("expenses") or []
        self.assertEqual(len(expenses), 2)
        self.assertEqual(payload.get("expense_id"), expenses[0]["expense_id"])

        conn = db_mod.get_db()
        try:
            rows = conn.execute(
                """
                SELECT id, category, amount, invoice_number, supplier_id
                FROM sales_update_expenses
                WHERE supplier_id = ? AND invoice_number = ?
                ORDER BY id
                """,
                (supplier_id, "INV-MULTI-1"),
            ).fetchall()
            self.assertEqual(len(rows), 2)
            cats = {str(row["category"]) for row in rows}
            self.assertIn("non_veg", cats)
            self.assertIn("dairy_products", cats)
            amounts = {str(row["category"]): float(row["amount"]) for row in rows}
            self.assertAlmostEqual(amounts["non_veg"], 2000.0)
            self.assertAlmostEqual(amounts["dairy_products"], 500.0)

            for item_name, qty_delta in (("Chicken Whole", 2.0), ("Butter", 1.0)):
                stock = conn.execute(
                    """
                    SELECT qty_on_hand FROM store_stock_items
                    WHERE outlet = 'restaurant' AND item_name = ? AND unit = 'kg'
                    """,
                    (item_name,),
                ).fetchone()
                self.assertIsNotNone(stock)
                self.assertAlmostEqual(
                    float(stock["qty_on_hand"]),
                    before_rows.get(item_name, 0.0) + qty_delta,
                )

            for row in rows:
                item_name = "Chicken Whole" if row["category"] == "non_veg" else "Butter"
                movement = conn.execute(
                    """
                    SELECT item_name, ref_type, ref_id
                    FROM store_stock_movements
                    WHERE ref_type = 'stock_inward_direct' AND ref_id = ?
                    ORDER BY id DESC LIMIT 1
                    """,
                    (int(row["id"]),),
                ).fetchone()
                self.assertIsNotNone(movement)
                self.assertEqual(movement["item_name"], item_name)
        finally:
            conn.close()

    def test_direct_stock_inward_missing_product_category_rejected(self):
        from datetime import date

        conn = db_mod.get_db()
        try:
            conn.execute(
                "INSERT INTO suppliers (name) VALUES (?)",
                ("No Cat Supplier",),
            )
            supplier_id = conn.execute(
                "SELECT id FROM suppliers WHERE name = 'No Cat Supplier'"
            ).fetchone()["id"]
            # Insert a product without a Product Master category name (orphan category_id).
            conn.execute(
                """
                INSERT INTO store_product_categories (name, sort_order, is_active)
                VALUES ('Temp Blank Cat', 999, 1)
                """
            )
            cat_id = conn.execute(
                "SELECT id FROM store_product_categories WHERE name = 'Temp Blank Cat'"
            ).fetchone()["id"]
            conn.execute(
                """
                INSERT INTO store_products
                    (category_id, name, default_unit, outlet, is_active, sort_order)
                VALUES (?, 'Uncategorized Widget', 'kg', 'restaurant', 1, 10)
                """,
                (cat_id,),
            )
            # Wipe category name so mapping fails (simulate missing PM category).
            conn.execute(
                "UPDATE store_product_categories SET name = '' WHERE id = ?",
                (cat_id,),
            )
            conn.commit()
        finally:
            conn.close()

        confirm = self.client.post(
            "/stores/purchase-requests/confirm-direct-with-expense",
            json={
                "outlet": "restaurant",
                "lines": [
                    {
                        "item_name": "Uncategorized Widget",
                        "qty": 1,
                        "unit": "kg",
                        "unit_price": 10,
                        "tax_percent": 0,
                    }
                ],
                "date": date.today().isoformat(),
                "description": "Stock inward without indent approval",
                "amount": 10,
                "payment_type": "credit",
                "invoice_number": "INV-NOCAT-1",
                "supplier_id": supplier_id,
            },
        )
        self.assertEqual(confirm.status_code, 400, confirm.data)
        body = confirm.get_json() or {}
        self.assertFalse(body.get("ok", True))
        self.assertIn("category", (body.get("error") or "").lower())

    def test_indent_approved_view_includes_approver(self):
        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Approve me",
                "item_name": ["Onion"],
                "quantity": ["1"],
                "unit": ["kg"],
                "approximate_price": ["10"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)
        conn = db_mod.get_db()
        try:
            indent = conn.execute(
                "SELECT id FROM store_indents WHERE notes = 'Approve me' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(indent)
            indent_id = indent["id"]
        finally:
            conn.close()

        decide = self.client.post(
            f"/stores/indent/{indent_id}/decide",
            data={"outlet": "bar", "decision": "approved", "decision_note": ""},
            follow_redirects=False,
        )
        self.assertEqual(decide.status_code, 302)

        page = self.client.get("/stores/indent?outlet=bar&view=approved")
        self.assertEqual(page.status_code, 200)
        self.assertIn(b'"decided_by_name": "Administrator"', page.data)
        self.assertIn(b'"decided_by_username": "admin"', page.data)
        self.assertIn(b'"decided_at":', page.data)

    def test_indent_defaults_to_all_outlet(self):
        page = self.client.get("/stores/indent")
        self.assertEqual(page.status_code, 200)
        self.assertIn(b'id="st-outlet-value">All</span>', page.data)
        self.assertIn(b'data-value="both"', page.data)
        self.assertIn(b"Pending Approval", page.data)
        self.assertNotIn(b'href="/stores/indent?outlet=both&amp;view=draft"', page.data)
        self.assertNotIn(b'href="/stores/indent?view=draft"', page.data)
        self.assertIn(b"Approved", page.data)
        self.assertIn(b"Rejected", page.data)
        self.assertIn(b"cp-view-tabs", page.data)
        self.assertIn(b'id="de-nav-stores-purchase-order"', page.data)
        self.assertIn(b">Purchase Order</a>", page.data)
        self.assertIn(b'id="de-nav-stores-indent"', page.data)
        # Default Indent list keeps Indent nav active (not Purchase Order).
        self.assertRegex(
            page.data,
            rb'class="de-nav-subitem is-active"[^>]*id="de-nav-stores-indent"|id="de-nav-stores-indent"[^>]*class="de-nav-subitem is-active"',
        )
        # Approved tab stays on Indent (does not jump to Purchase Order).
        approved_tab = self.client.get("/stores/indent?view=approved")
        self.assertEqual(approved_tab.status_code, 200)
        self.assertIn(b'href="/stores/indent?outlet=both&amp;view=approved"', approved_tab.data)
        self.assertRegex(
            approved_tab.data,
            rb'class="de-nav-subitem is-active"[^>]*id="de-nav-stores-indent"|id="de-nav-stores-indent"[^>]*class="de-nav-subitem is-active"',
        )
        po_page = self.client.get("/stores/orders")
        self.assertEqual(po_page.status_code, 200)
        self.assertIn(b"Purchase Order", po_page.data)
        self.assertRegex(
            po_page.data,
            rb'class="de-nav-subitem is-active"[^>]*id="de-nav-stores-purchase-order"|id="de-nav-stores-purchase-order"[^>]*class="de-nav-subitem is-active"',
        )

    def test_indent_rejected_tab_allows_edit_and_resubmit(self):
        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Reject then fix",
                "item_name": ["Onion"],
                "quantity": ["1"],
                "unit": ["kg"],
                "approximate_price": ["10"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)
        conn = db_mod.get_db()
        try:
            indent = conn.execute(
                "SELECT id FROM store_indents WHERE notes = 'Reject then fix' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(indent)
            indent_id = indent["id"]
        finally:
            conn.close()

        self.client.post(
            f"/stores/indent/{indent_id}/decide",
            data={"outlet": "bar", "decision": "rejected", "decision_note": "Wrong qty"},
            follow_redirects=False,
        )

        rejected_page = self.client.get("/stores/indent?outlet=bar&view=rejected")
        self.assertEqual(rejected_page.status_code, 200)
        self.assertIn(b"Reject then fix", rejected_page.data)
        self.assertIn(b"Wrong qty", rejected_page.data)
        self.assertIn(b'data-st-edit-indent="%d"' % indent_id, rejected_page.data)

        save = self.client.post(
            f"/stores/indent?outlet=bar&edit={indent_id}",
            data={
                "outlet": "bar",
                "indent_id": str(indent_id),
                "action": "save",
                "notes": "Reject then fix",
                "item_name": ["Onion"],
                "quantity": ["4"],
                "unit": ["kg"],
                "approximate_price": ["10"],
            },
            follow_redirects=False,
        )
        self.assertEqual(save.status_code, 302)
        self.assertIn("view=pending", save.headers.get("Location", ""))

        conn = db_mod.get_db()
        try:
            row = conn.execute(
                "SELECT status, decision_note, decided_at FROM store_indents WHERE id = ?",
                (indent_id,),
            ).fetchone()
            self.assertEqual(row["status"], "pending")
            self.assertEqual(row["decision_note"] or "", "")
            self.assertFalse(row["decided_at"])
            qty = conn.execute(
                "SELECT quantity FROM store_indent_lines WHERE indent_id = ?",
                (indent_id,),
            ).fetchone()
            self.assertEqual(float(qty["quantity"]), 4.0)
        finally:
            conn.close()

        pending_page = self.client.get("/stores/indent?outlet=bar&view=pending")
        self.assertIn(b"Reject then fix", pending_page.data)
        rejected_again = self.client.get("/stores/indent?outlet=bar&view=rejected")
        self.assertNotIn(b"Reject then fix", rejected_again.data)

    def test_stores_outlet_filter_includes_all(self):
        for path in (
            "/stores/approvals",
            "/stores/stock",
        ):
            page = self.client.get(path)
            self.assertEqual(page.status_code, 200, path)
            self.assertIn(b'data-value="both"', page.data, path)
            self.assertIn(b">All</button>", page.data, path)

        # Stock Inward: All / Bar / Restaurant; bare URL defaults to All.
        inward = self.client.get("/stores/purchase-requests", follow_redirects=False)
        self.assertEqual(inward.status_code, 200)
        self.assertIn(b'data-value="both"', inward.data)
        self.assertIn(b">All</button>", inward.data)
        self.assertRegex(
            inward.data,
            rb'id="st-outlet-value"[^>]*>\s*All\s*<',
        )
        self.assertRegex(
            inward.data,
            rb'id="st-outlet"[^>]*value="both"',
        )
        inward_bar = self.client.get("/stores/purchase-requests?outlet=bar")
        self.assertEqual(inward_bar.status_code, 200)
        self.assertIn(b'data-value="both"', inward_bar.data)
        self.assertIn(b">All</button>", inward_bar.data)
        self.assertIn(b'data-value="bar"', inward_bar.data)
        self.assertIn(b'data-value="restaurant"', inward_bar.data)

    def test_approvals_table_is_sortable(self):
        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Sort me",
                "item_name": ["Onion"],
                "quantity": ["2"],
                "unit": ["kg"],
                "approximate_price": ["10"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)
        page = self.client.get("/stores/approvals?outlet=bar")
        self.assertEqual(page.status_code, 200)
        self.assertIn(b'id="st-approvals-pending-table"', page.data)
        self.assertIn(b"pl-sortable", page.data)
        self.assertIn(b'data-sort="indent"', page.data)
        self.assertIn(b'data-sort="submitted"', page.data)
        self.assertIn(b'data-sort-row', page.data)
        self.assertIn(b"pl-list-panel--scroll", page.data)
        self.assertIn(b"hbe-scroll-panel", page.data)
        self.assertIn(b"Click to sort", page.data)
        self.assertIn(b"Approx. price", page.data)
        self.assertIn(b"\xe2\x82\xb920", page.data)  # ₹20 (2 × 10)
        self.assertIn(b"Indents awaiting your approval", page.data)
        self.assertIn(b"st-appr-btn--approve", page.data)
        self.assertRegex(page.data.decode("utf-8"), r"\d{1,2} [A-Za-z]+")
        self.assertRegex(page.data.decode("utf-8"), r"\d{1,2}:\d{2} (AM|PM)")
        self.assertNotRegex(page.data.decode("utf-8", errors="ignore"), r">\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}<")
        self.assertNotIn(b"st-appr-icon--cal", page.data)

    def test_indent_approval_button_glows_when_pending(self):
        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Glow me",
                "item_name": ["Onion"],
                "quantity": ["2"],
                "unit": ["kg"],
                "approximate_price": ["10"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)

        indent_page = self.client.get("/stores/indent?outlet=bar")
        self.assertEqual(indent_page.status_code, 200)
        self.assertIn(b'id="st-approvals-open"', indent_page.data)
        self.assertIn(b"st-approvals-open--attention", indent_page.data)
        self.assertIn(b'data-pending-count="1"', indent_page.data)

        conn = db_mod.get_db()
        try:
            indent_id = conn.execute(
                "SELECT id FROM store_indents WHERE notes = 'Glow me' ORDER BY id DESC LIMIT 1"
            ).fetchone()["id"]
        finally:
            conn.close()

        decide = self.client.post(
            f"/stores/indent/{indent_id}/decide",
            data={"outlet": "bar", "decision": "approved", "decision_note": ""},
            follow_redirects=False,
        )
        self.assertEqual(decide.status_code, 302)

        cleared = self.client.get("/stores/indent?outlet=bar")
        self.assertEqual(cleared.status_code, 200)
        self.assertIn(b'id="st-approvals-open"', cleared.data)
        self.assertNotIn(b"st-approvals-open--attention", cleared.data)
        self.assertIn(b'data-pending-count="0"', cleared.data)

    def test_home_shows_approvals_notification_only_for_approvers(self):
        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Notify approvers",
                "item_name": ["Onion"],
                "quantity": ["2"],
                "unit": ["kg"],
                "approximate_price": ["10"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)

        home = self.client.get("/home")
        self.assertEqual(home.status_code, 200)
        self.assertIn(b"Indents awaiting approval", home.data)
        self.assertIn(b'href="/stores/approvals"', home.data)

        # Stores Approvals submodule alone is not enough — need Module Tree → Approval.
        stores_clerk = {
            "id": self.admin_id,
            "username": "storeclerk",
            "full_name": "Store Clerk",
            "is_admin": False,
            "is_active": True,
            "dashboard_access": {"stores"},
            "stores_access": {"indent", "stock", "approvals"},
            "sales_analytics_access": set(),
            "user_access": set(),
            "payroll_access": set(),
            "accounts_access": set(),
        }
        with mock.patch.object(self.app_mod, "get_current_user", return_value=stores_clerk):
            denied = self.client.get("/home")
        self.assertEqual(denied.status_code, 200)
        self.assertNotIn(b"Indents awaiting approval", denied.data)

        approver = dict(stores_clerk)
        approver["dashboard_access"] = {"stores", "approval"}
        with mock.patch.object(self.app_mod, "get_current_user", return_value=approver):
            allowed = self.client.get("/home")
        self.assertEqual(allowed.status_code, 200)
        self.assertIn(b"Indents awaiting approval", allowed.data)
        self.assertIn(b'href="/stores/approvals"', allowed.data)

    def test_approvals_reject_popup_and_reopen(self):
        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Needs decision",
                "item_name": ["Onion"],
                "quantity": ["3"],
                "unit": ["kg"],
                "approximate_price": ["12"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)
        page = self.client.get("/stores/approvals?outlet=bar")
        self.assertEqual(page.status_code, 200)
        self.assertIn(b'data-st-reject-open', page.data)
        self.assertIn(b'id="st-reject-modal"', page.data)
        self.assertIn(b'name="decision_note"', page.data)
        self.assertIn(b'name="decision" value="approved"', page.data)

        conn = db_mod.get_db()
        try:
            indent = conn.execute(
                "SELECT id FROM store_indents WHERE notes = 'Needs decision' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(indent)
            indent_id = indent["id"]
        finally:
            conn.close()

        blocked = self.client.post(
            f"/stores/indent/{indent_id}/decide",
            data={"outlet": "bar", "decision": "rejected", "decision_note": ""},
            follow_redirects=False,
        )
        self.assertEqual(blocked.status_code, 302)
        conn = db_mod.get_db()
        try:
            row = conn.execute(
                "SELECT status FROM store_indents WHERE id = ?",
                (indent_id,),
            ).fetchone()
            self.assertEqual(row["status"], "pending")
        finally:
            conn.close()

        rejected = self.client.post(
            f"/stores/indent/{indent_id}/decide",
            data={"outlet": "bar", "decision": "rejected", "decision_note": "Out of season"},
            follow_redirects=False,
        )
        self.assertEqual(rejected.status_code, 302)
        conn = db_mod.get_db()
        try:
            row = conn.execute(
                "SELECT status, decision_note FROM store_indents WHERE id = ?",
                (indent_id,),
            ).fetchone()
            self.assertEqual(row["status"], "rejected")
            self.assertEqual(row["decision_note"], "Out of season")
        finally:
            conn.close()

        recent = self.client.get("/stores/approvals?outlet=bar")
        self.assertIn(b"Return to waiting", recent.data)

        reopened = self.client.post(
            f"/stores/indent/{indent_id}/reopen",
            data={"outlet": "bar"},
            follow_redirects=False,
        )
        self.assertEqual(reopened.status_code, 302)
        conn = db_mod.get_db()
        try:
            row = conn.execute(
                "SELECT status, decision_note, decided_at FROM store_indents WHERE id = ?",
                (indent_id,),
            ).fetchone()
            self.assertEqual(row["status"], "pending")
            self.assertEqual(row["decision_note"] or "", "")
            self.assertFalse(row["decided_at"])
        finally:
            conn.close()

        waiting = self.client.get("/stores/approvals?outlet=bar")
        self.assertIn(b"Needs decision", waiting.data)

    def test_indent_list_view_filters_by_status(self):
        draft = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "save",
                "notes": "Draft only",
                "item_name": ["Onion"],
                "quantity": ["1"],
                "unit": ["kg"],
                "approximate_price": ["10"],
            },
            follow_redirects=False,
        )
        self.assertEqual(draft.status_code, 302)
        self.assertIn("view=draft", draft.headers.get("Location", ""))
        pending = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Waiting",
                "item_name": ["Onion"],
                "quantity": ["2"],
                "unit": ["kg"],
                "approximate_price": ["20"],
            },
            follow_redirects=False,
        )
        self.assertEqual(pending.status_code, 302)

        conn = db_mod.get_db()
        try:
            waiting = conn.execute(
                "SELECT id FROM store_indents WHERE notes = 'Waiting' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(waiting)
            conn.execute(
                "UPDATE store_indents SET status = 'approved' WHERE id = ?",
                (waiting["id"],),
            )
            conn.commit()
        finally:
            conn.close()

        draft_page = self.client.get("/stores/indent?outlet=bar&view=draft")
        self.assertEqual(draft_page.status_code, 200)
        self.assertIn(b"Draft only", draft_page.data)
        self.assertNotIn(b"Waiting", draft_page.data)
        self.assertIn(b'href="/stores/indent?outlet=bar&amp;view=draft"', draft_page.data)

        pending_page = self.client.get("/stores/indent?outlet=bar&view=pending")
        self.assertEqual(pending_page.status_code, 200)
        self.assertNotIn(b"Draft only", pending_page.data)
        self.assertNotIn(b"Waiting", pending_page.data)

        approved_page = self.client.get("/stores/indent?outlet=bar&view=approved")
        self.assertEqual(approved_page.status_code, 200)
        self.assertIn(b"Waiting", approved_page.data)
        self.assertNotIn(b"Draft only", approved_page.data)
        self.assertIn(b"Download PO", approved_page.data)
        self.assertIn(b"/purchase-order", approved_page.data)

    def test_indent_draft_tab_hidden_until_draft_exists(self):
        empty = self.client.get("/stores/indent?outlet=bar&view=pending")
        self.assertEqual(empty.status_code, 200)
        self.assertNotIn(b">Draft</a>", empty.data)
        self.assertNotIn(b">Draft</span>", empty.data)
        self.assertNotIn(b'href="/stores/indent?outlet=bar&amp;view=draft"', empty.data)
        self.assertNotIn(b"cp-view-tab is-disabled", empty.data)

        redirect = self.client.get("/stores/indent?outlet=bar&view=draft", follow_redirects=False)
        self.assertEqual(redirect.status_code, 302)
        self.assertIn("view=pending", redirect.headers.get("Location", ""))

        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "save",
                "notes": "Enable draft tab",
                "item_name": ["Onion"],
                "quantity": ["1"],
                "unit": ["kg"],
                "approximate_price": ["10"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)
        self.assertIn("view=draft", create.headers.get("Location", ""))

        with_draft = self.client.get("/stores/indent?outlet=bar&view=draft")
        self.assertEqual(with_draft.status_code, 200)
        self.assertIn(b"Enable draft tab", with_draft.data)
        self.assertIn(b'href="/stores/indent?outlet=bar&amp;view=draft"', with_draft.data)
        self.assertIn(b">Draft</a>", with_draft.data)
        self.assertNotIn(b"cp-view-tab is-disabled", with_draft.data)

    def test_indent_purchase_order_download(self):
        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "PO export",
                "item_name": ["Onion", "Potato"],
                "quantity": ["10", "5"],
                "unit": ["kg", "kg"],
                "approximate_price": ["30", "20"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)

        conn = db_mod.get_db()
        try:
            indent = conn.execute(
                "SELECT id, indent_no FROM store_indents WHERE notes = 'PO export' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(indent)
            indent_id = indent["id"]
            indent_no = indent["indent_no"]
            conn.execute(
                "UPDATE store_indents SET status = 'approved' WHERE id = ?",
                (indent_id,),
            )
            conn.commit()
        finally:
            conn.close()

        approved_page = self.client.get("/stores/indent?outlet=bar&view=approved")
        self.assertEqual(approved_page.status_code, 200)
        self.assertIn(indent_no.encode(), approved_page.data)
        self.assertIn(f"/stores/indent/{indent_id}/purchase-order".encode(), approved_page.data)

        po = self.client.get(f"/stores/indent/{indent_id}/purchase-order")
        self.assertEqual(po.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            po.content_type,
        )
        self.assertTrue(po.data[:2] == b"PK")

        pending_blocked = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "save",
                "notes": "Draft blocked",
                "item_name": ["Onion"],
                "quantity": ["1"],
                "unit": ["kg"],
                "approximate_price": ["10"],
            },
            follow_redirects=False,
        )
        self.assertEqual(pending_blocked.status_code, 302)
        conn = db_mod.get_db()
        try:
            draft = conn.execute(
                "SELECT id FROM store_indents WHERE notes = 'Draft blocked' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(draft)
            draft_id = draft["id"]
        finally:
            conn.close()
        blocked = self.client.get(
            f"/stores/indent/{draft_id}/purchase-order",
            follow_redirects=False,
        )
        self.assertEqual(blocked.status_code, 302)

    def test_indent_product_list_filters_by_outlet(self):
        conn = db_mod.get_db()
        try:
            db_mod.ensure_stores_schema(conn)
            cat = conn.execute(
                "SELECT id FROM store_product_categories WHERE is_active = 1 ORDER BY id LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(cat)
            cat_id = cat["id"]
            conn.execute(
                """
                INSERT INTO store_products
                    (category_id, name, default_unit, outlet, is_active, sort_order, updated_at)
                VALUES (?, 'Bar Only Mixer', 'bottle', 'bar', 1, 9990, datetime('now','localtime'))
                """,
                (cat_id,),
            )
            conn.execute(
                """
                INSERT INTO store_products
                    (category_id, name, default_unit, outlet, is_active, sort_order, updated_at)
                VALUES (?, 'Restaurant Only Herb', 'bunch', 'restaurant', 1, 9991, datetime('now','localtime'))
                """,
                (cat_id,),
            )
            conn.execute(
                """
                INSERT INTO store_products
                    (category_id, name, default_unit, outlet, is_active, sort_order, updated_at)
                VALUES (?, 'Shared Oil', 'liter', 'both', 1, 9992, datetime('now','localtime'))
                """,
                (cat_id,),
            )
            conn.commit()
        finally:
            conn.close()

        bar_form = self.client.get("/stores/indent?outlet=bar&focus=form")
        self.assertEqual(bar_form.status_code, 200)
        self.assertIn(b"<h1>Create Indent</h1>", bar_form.data)
        self.assertIn(b'class="su-page-back"', bar_form.data)
        self.assertIn(b'href="/stores/indent?outlet=bar"', bar_form.data)
        self.assertIn(b'aria-label="Back to Indent"', bar_form.data)
        self.assertIn(b"Bar Only Mixer", bar_form.data)
        self.assertIn(b"Shared Oil", bar_form.data)
        self.assertNotIn(b"Restaurant Only Herb", bar_form.data)
        # Create form: Bar/Restaurant only — All is list/view filter.
        form_outlet = bar_form.data.split(b'id="st-outlet-listbox"', 1)[1].split(b"</article>", 1)[0]
        self.assertNotIn(b'data-value="both"', form_outlet)
        self.assertIn(b'data-value="bar"', form_outlet)
        self.assertIn(b'data-value="restaurant"', form_outlet)

        rest_form = self.client.get("/stores/indent?outlet=restaurant&focus=form")
        self.assertEqual(rest_form.status_code, 200)
        self.assertIn(b"Restaurant Only Herb", rest_form.data)
        self.assertIn(b"Shared Oil", rest_form.data)
        self.assertNotIn(b"Bar Only Mixer", rest_form.data)
        rest_outlet = rest_form.data.split(b'id="st-outlet-listbox"', 1)[1].split(b"</article>", 1)[0]
        self.assertNotIn(b'data-value="both"', rest_outlet)

        # New Indent with All / no outlet defaults to Restaurant.
        unset_form = self.client.get("/stores/indent?focus=form", follow_redirects=False)
        self.assertEqual(unset_form.status_code, 302)
        self.assertIn("outlet=restaurant", unset_form.headers.get("Location", ""))
        self.assertIn("focus=form", unset_form.headers.get("Location", ""))

        unset_landed = self.client.get("/stores/indent?focus=form", follow_redirects=True)
        self.assertEqual(unset_landed.status_code, 200)
        self.assertIn(b"<h1>Create Indent</h1>", unset_landed.data)
        self.assertIn(b'id="st-outlet-value">Restaurant</span>', unset_landed.data)
        self.assertNotIn(b"Select outlet", unset_landed.data)
        self.assertNotIn(b'id="st-outlet-value" class="se-filter-chip-value is-placeholder"', unset_landed.data)
        self.assertNotRegex(unset_landed.data, rb'id="st-outlet-value"[^>]*is-placeholder')
        self.assertIn(b"Restaurant Only Herb", unset_landed.data)
        self.assertIn(b"Shared Oil", unset_landed.data)
        self.assertNotIn(b"Bar Only Mixer", unset_landed.data)
        unset_outlet = unset_landed.data.split(b'id="st-outlet-listbox"', 1)[1].split(b"</article>", 1)[0]
        self.assertNotIn(b'data-value="both"', unset_outlet)

        both_form = self.client.get("/stores/indent?outlet=both&focus=form", follow_redirects=False)
        self.assertEqual(both_form.status_code, 302)
        self.assertIn("outlet=restaurant", both_form.headers.get("Location", ""))

        both_landed = self.client.get("/stores/indent?outlet=both&focus=form", follow_redirects=True)
        self.assertEqual(both_landed.status_code, 200)
        self.assertIn(b'id="st-outlet-value">Restaurant</span>', both_landed.data)
        self.assertNotIn(b"Select outlet", both_landed.data)
        self.assertIn(b"Restaurant Only Herb", both_landed.data)
        self.assertNotIn(b"Bar Only Mixer", both_landed.data)

    def test_indent_list_view_edit_delete_actions(self):
        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "save",
                "notes": "Editable draft",
                "item_name": ["Onion"],
                "quantity": ["2"],
                "unit": ["kg"],
                "approximate_price": ["40"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)

        conn = db_mod.get_db()
        try:
            indent = conn.execute(
                "SELECT id, indent_no FROM store_indents WHERE outlet = 'bar' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(indent)
            indent_id = indent["id"]
            line = conn.execute(
                "SELECT approximate_price FROM store_indent_lines WHERE indent_id = ?",
                (indent_id,),
            ).fetchone()
            self.assertEqual(float(line["approximate_price"]), 40.0)
        finally:
            conn.close()

        listing = self.client.get("/stores/indent?outlet=bar&view=draft")
        self.assertEqual(listing.status_code, 200)
        self.assertIn(b"Editable draft", listing.data)
        self.assertIn(b'data-st-view-indent=', listing.data)
        self.assertIn(b'id="st-indent-view-modal"', listing.data)
        self.assertIn(b'data-st-edit-indent=', listing.data)
        self.assertIn(b'id="st-indent-edit-modal"', listing.data)
        # Soft-nav only executes scripts inside .de-main-wrapper; stores.js must
        # appear before the shell close scripts or View/Edit never bind.
        stores_js_idx = listing.data.find(b"/static/stores.js")
        shell_nav_idx = listing.data.find(b"/static/de_workspace_nav.js")
        self.assertGreaterEqual(stores_js_idx, 0)
        self.assertGreaterEqual(shell_nav_idx, 0)
        self.assertLess(stores_js_idx, shell_nav_idx)
        self.assertIn(b'Approximate Price', listing.data)
        self.assertIn(b'pl-sortable', listing.data)
        self.assertIn(b'data-sort="indent"', listing.data)
        self.assertIn(b'data-tip="Edit"', listing.data)
        self.assertIn(b'data-tip="Delete"', listing.data)

        edit_page = self.client.get(f"/stores/indent?outlet=bar&edit={indent_id}&focus=form")
        self.assertEqual(edit_page.status_code, 200)
        self.assertIn(b'id="st-indent-edit-modal"', edit_page.data)
        self.assertIn(f'data-st-open-edit="{indent_id}"'.encode(), edit_page.data)
        self.assertIn(b"Edit indent", edit_page.data)
        self.assertIn(b"Onion", edit_page.data)

        update = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "indent_id": str(indent_id),
                "action": "save",
                "notes": "Updated note",
                "item_name": ["Potato"],
                "quantity": ["5"],
                "unit": ["kg"],
                "approximate_price": ["55"],
            },
            follow_redirects=True,
        )
        self.assertEqual(update.status_code, 200)
        self.assertIn(b"Indent sent for approval", update.data)
        self.assertIn(b"Waiting approval", update.data)

        delete = self.client.get(
            f"/stores/indent/{indent_id}/delete?outlet=bar",
            follow_redirects=True,
        )
        self.assertEqual(delete.status_code, 200)
        self.assertIn(b"Deleted", delete.data)
        conn = db_mod.get_db()
        try:
            gone = conn.execute(
                "SELECT id FROM store_indents WHERE id = ?", (indent_id,)
            ).fetchone()
            self.assertIsNone(gone)
        finally:
            conn.close()

    def test_edit_save_sends_draft_for_approval(self):
        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "save",
                "notes": "Editable draft",
                "item_name": ["Onion"],
                "quantity": ["2"],
                "unit": ["kg"],
                "approximate_price": ["40"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)

        conn = db_mod.get_db()
        try:
            indent = conn.execute(
                "SELECT id, status FROM store_indents WHERE outlet = 'bar' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(indent)
            self.assertEqual(indent["status"], "draft")
            indent_id = indent["id"]
        finally:
            conn.close()

        # Edit modal Save is the final save → Waiting approval.
        update = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "indent_id": str(indent_id),
                "action": "save",
                "notes": "Ready for approval",
                "item_name": ["Onion"],
                "quantity": ["3"],
                "unit": ["kg"],
                "approximate_price": ["45"],
            },
            follow_redirects=True,
        )
        self.assertEqual(update.status_code, 200)
        self.assertIn(b"Indent sent for approval", update.data)

        conn = db_mod.get_db()
        try:
            row = conn.execute(
                "SELECT status, notes, submitted_at FROM store_indents WHERE id = ?",
                (indent_id,),
            ).fetchone()
            self.assertEqual(row["status"], "pending")
            self.assertEqual(row["notes"], "Ready for approval")
            self.assertTrue(row["submitted_at"])
            line = conn.execute(
                "SELECT quantity, approximate_price FROM store_indent_lines WHERE indent_id = ?",
                (indent_id,),
            ).fetchone()
            self.assertEqual(float(line["quantity"]), 3.0)
            self.assertEqual(float(line["approximate_price"]), 45.0)
        finally:
            conn.close()

        # Soft-nav can drop the hidden field — ?edit= on the action URL must still update.
        update_via_query = self.client.post(
            f"/stores/indent?outlet=bar&edit={indent_id}",
            data={
                "outlet": "bar",
                "action": "save",
                "notes": "Via query edit id",
                "item_name": ["Onion"],
                "quantity": ["4"],
                "unit": ["kg"],
                "approximate_price": ["50"],
            },
            follow_redirects=True,
        )
        self.assertEqual(update_via_query.status_code, 200)
        self.assertIn(b"Indent updated", update_via_query.data)
        conn = db_mod.get_db()
        try:
            row = conn.execute(
                "SELECT status, notes FROM store_indents WHERE id = ?",
                (indent_id,),
            ).fetchone()
            self.assertEqual(row["status"], "pending")
            self.assertEqual(row["notes"], "Via query edit id")
            count = conn.execute(
                "SELECT COUNT(*) AS c FROM store_indents WHERE outlet = 'bar'"
            ).fetchone()["c"]
            self.assertEqual(count, 1)
        finally:
            conn.close()

    def test_admin_delete_approved_indent_removes_non_inwarded_only(self):
        from datetime import date

        conn = db_mod.get_db()
        try:
            conn.execute(
                "INSERT INTO suppliers (name) VALUES (?)",
                ("Approved Delete Supplier",),
            )
            supplier_id = conn.execute(
                "SELECT id FROM suppliers WHERE name = 'Approved Delete Supplier'"
            ).fetchone()["id"]
            conn.execute(
                "INSERT INTO cash_ledger_loads (company, load_date, description, amount) VALUES (?,?,?,?)",
                ("HBE", date.today().isoformat(), "Approved delete float", 5000),
            )
            conn.commit()
        finally:
            conn.close()

        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Admin delete approved partial",
                "item_name": ["Onion", "Potato", "Tomato"],
                "quantity": ["10", "24", "8"],
                "unit": ["kg", "kg", "kg"],
                "approximate_price": ["30", "20", "15"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)
        conn = db_mod.get_db()
        try:
            indent = conn.execute(
                "SELECT id, indent_no FROM store_indents WHERE notes = 'Admin delete approved partial' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            indent_id = indent["id"]
            lines = {
                row["item_name"]: int(row["id"])
                for row in conn.execute(
                    "SELECT id, item_name FROM store_indent_lines WHERE indent_id = ?",
                    (indent_id,),
                ).fetchall()
            }
        finally:
            conn.close()

        decide = self.client.post(
            f"/stores/indent/{indent_id}/decide",
            data={"outlet": "bar", "decision": "approved", "decision_note": ""},
            follow_redirects=False,
        )
        self.assertEqual(decide.status_code, 302)

        # Fully receive Onion; partially receive Potato (10 of 24); leave Tomato pending.
        partial = self.client.post(
            "/stores/purchase-requests/confirm-with-expense",
            json={
                "indent_id": indent_id,
                "notes": "First delivery before cancel",
                "lines": [
                    {"line_id": lines["Onion"], "received_qty": 10},
                    {"line_id": lines["Potato"], "received_qty": 10},
                ],
                "date": date.today().isoformat(),
                "description": "Partial before approved delete",
                "amount": 500,
                "payment_type": "cash",
                "category": "grocery",
                "supplier_id": supplier_id,
            },
        )
        self.assertEqual(partial.status_code, 200)
        self.assertTrue(partial.get_json().get("ok"))

        conn = db_mod.get_db()
        try:
            stock_before = {
                (row["item_name"], row["unit"]): float(row["qty_on_hand"])
                for row in conn.execute(
                    "SELECT item_name, unit, qty_on_hand FROM store_stock_items WHERE outlet = 'bar'"
                ).fetchall()
            }
            movements_before = conn.execute(
                """
                SELECT COUNT(*) AS c FROM store_stock_movements
                WHERE ref_type = 'stock_inward' AND ref_id = ?
                """,
                (indent_id,),
            ).fetchone()["c"]
        finally:
            conn.close()
        self.assertEqual(stock_before[("Onion", "kg")], 10.0)
        self.assertEqual(stock_before[("Potato", "kg")], 10.0)
        self.assertNotIn(("Tomato", "kg"), stock_before)

        approved_list = self.client.get("/stores/indent?outlet=bar&view=approved")
        self.assertEqual(approved_list.status_code, 200)
        self.assertIn(b'data-tip="Delete remaining"', approved_list.data)
        self.assertIn(b"Already inwarded stock will not be changed", approved_list.data)

        delete = self.client.get(
            f"/stores/indent/{indent_id}/delete?outlet=bar",
            follow_redirects=True,
        )
        self.assertEqual(delete.status_code, 200)
        self.assertIn(b"Inwarded stock unchanged", delete.data)

        conn = db_mod.get_db()
        try:
            indent_row = conn.execute(
                "SELECT status FROM store_indents WHERE id = ?", (indent_id,)
            ).fetchone()
            self.assertIsNotNone(indent_row)
            # Remains approved so PO download stays available; open qty is gone.
            self.assertEqual(indent_row["status"], "approved")
            line_rows = {
                row["item_name"]: (
                    float(row["quantity"]),
                    float(row["quantity_received"] or 0),
                )
                for row in conn.execute(
                    "SELECT item_name, quantity, quantity_received FROM store_indent_lines WHERE indent_id = ?",
                    (indent_id,),
                ).fetchall()
            }
            self.assertEqual(set(line_rows), {"Onion", "Potato"})
            self.assertEqual(line_rows["Onion"], (10.0, 10.0))
            self.assertEqual(line_rows["Potato"], (10.0, 10.0))
            self.assertNotIn("Tomato", line_rows)

            stock_after = {
                (row["item_name"], row["unit"]): float(row["qty_on_hand"])
                for row in conn.execute(
                    "SELECT item_name, unit, qty_on_hand FROM store_stock_items WHERE outlet = 'bar'"
                ).fetchall()
            }
            movements_after = conn.execute(
                """
                SELECT COUNT(*) AS c FROM store_stock_movements
                WHERE ref_type = 'stock_inward' AND ref_id = ?
                """,
                (indent_id,),
            ).fetchone()["c"]
        finally:
            conn.close()

        self.assertEqual(stock_after[("Onion", "kg")], 10.0)
        self.assertEqual(stock_after[("Potato", "kg")], 10.0)
        self.assertNotIn(("Tomato", "kg"), stock_after)
        self.assertEqual(movements_after, movements_before)

        # Nothing left to cancel — clear error, indent and stock unchanged.
        blocked = self.client.get(
            f"/stores/indent/{indent_id}/delete?outlet=bar",
            follow_redirects=True,
        )
        self.assertEqual(blocked.status_code, 200)
        self.assertIn(b"fully inwarded", blocked.data)
        self.assertIn(b"nothing left to delete", blocked.data)

    def test_non_admin_cannot_delete_approved_indent(self):
        create = self.client.post(
            "/stores/indent?outlet=bar",
            data={
                "outlet": "bar",
                "action": "submit",
                "notes": "Non-admin approved delete",
                "item_name": ["Onion"],
                "quantity": ["5"],
                "unit": ["kg"],
                "approximate_price": ["25"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)
        conn = db_mod.get_db()
        try:
            indent_id = conn.execute(
                "SELECT id FROM store_indents WHERE notes = 'Non-admin approved delete' ORDER BY id DESC LIMIT 1"
            ).fetchone()["id"]
        finally:
            conn.close()

        decide = self.client.post(
            f"/stores/indent/{indent_id}/decide",
            data={"outlet": "bar", "decision": "approved", "decision_note": ""},
            follow_redirects=False,
        )
        self.assertEqual(decide.status_code, 302)

        clerk = {
            "id": self.admin_id,
            "username": "storeclerk",
            "full_name": "Store Clerk",
            "is_admin": False,
            "is_active": True,
            "dashboard_access": {"stores"},
            "stores_access": {"indent", "stock", "purchase_requests", "approvals"},
            "sales_analytics_access": set(),
            "user_access": set(),
            "payroll_access": set(),
            "accounts_access": set(),
        }
        with mock.patch.object(self.app_mod, "get_current_user", return_value=clerk), mock.patch.object(
            self.stores_mod, "_get_user", return_value=clerk
        ):
            listing = self.client.get("/stores/indent?outlet=bar&view=approved")
            self.assertEqual(listing.status_code, 200)
            self.assertIn(b'data-tip="Download PO"', listing.data)
            self.assertNotIn(b'data-tip="Delete remaining"', listing.data)

            denied = self.client.get(
                f"/stores/indent/{indent_id}/delete?outlet=bar",
                follow_redirects=True,
            )
            self.assertEqual(denied.status_code, 200)
            self.assertIn(b"Only administrators can delete approved indents", denied.data)

        conn = db_mod.get_db()
        try:
            row = conn.execute(
                "SELECT status FROM store_indents WHERE id = ?", (indent_id,)
            ).fetchone()
            self.assertEqual(row["status"], "approved")
            line = conn.execute(
                "SELECT quantity, COALESCE(quantity_received, 0) AS quantity_received FROM store_indent_lines WHERE indent_id = ?",
                (indent_id,),
            ).fetchone()
            self.assertEqual(float(line["quantity"]), 5.0)
            self.assertEqual(float(line["quantity_received"]), 0.0)
        finally:
            conn.close()


    def _seed_stock_item(self, *, outlet="restaurant", item_name="Tomato", unit="kg", qty=35.0):
        conn = db_mod.get_db()
        try:
            db_mod.ensure_stores_schema(conn)
            conn.execute(
                """
                INSERT INTO store_stock_items (outlet, item_name, unit, qty_on_hand, updated_at)
                VALUES (?, ?, ?, ?, datetime('now','localtime'))
                """,
                (outlet, item_name, unit, qty),
            )
            conn.commit()
            row = conn.execute(
                """
                SELECT id FROM store_stock_items
                WHERE outlet = ? AND item_name = ? AND unit = ?
                ORDER BY id DESC LIMIT 1
                """,
                (outlet, item_name, unit),
            ).fetchone()
            return int(row["id"])
        finally:
            conn.close()

    def test_stock_audit_seeds_and_zero_variance_verify(self):
        self._seed_stock_item(qty=10.0)
        page = self.client.get("/stores/stock-audit?outlet=restaurant")
        self.assertEqual(page.status_code, 200)
        self.assertIn(b"Stock Audit", page.data)
        self.assertIn(b"Tomato", page.data)
        self.assertIn(b'id="st-audit-page"', page.data)

        conn = db_mod.get_db()
        try:
            line = conn.execute(
                """
                SELECT l.id, l.system_qty, l.status
                FROM store_stock_audit_lines l
                JOIN store_stock_audits a ON a.id = l.audit_id
                WHERE a.outlet = 'restaurant' AND a.status = 'open'
                  AND l.item_name = 'Tomato'
                ORDER BY l.id DESC LIMIT 1
                """
            ).fetchone()
            self.assertIsNotNone(line)
            self.assertEqual(line["status"], "pending")
            self.assertAlmostEqual(float(line["system_qty"]), 10.0)
            line_id = int(line["id"])
        finally:
            conn.close()

        verify = self.client.post(
            "/stores/stock-audit/verify",
            json={"line_id": line_id, "actual_qty": 10, "reason": "", "remarks": ""},
        )
        self.assertEqual(verify.status_code, 200)
        payload = verify.get_json()
        self.assertTrue(payload.get("ok"))

        conn = db_mod.get_db()
        try:
            line = conn.execute(
                "SELECT status, variance_qty FROM store_stock_audit_lines WHERE id = ?",
                (line_id,),
            ).fetchone()
            self.assertEqual(line["status"], "verified")
            self.assertAlmostEqual(float(line["variance_qty"] or 0), 0.0)
            stock = conn.execute(
                "SELECT qty_on_hand FROM store_stock_items WHERE item_name = 'Tomato' AND outlet = 'restaurant'"
            ).fetchone()
            self.assertAlmostEqual(float(stock["qty_on_hand"]), 10.0)
            moves = conn.execute(
                """
                SELECT COUNT(*) AS c FROM store_stock_movements
                WHERE ref_type = 'stock_audit' AND item_name = 'Tomato'
                """
            ).fetchone()
            self.assertEqual(int(moves["c"]), 0)
        finally:
            conn.close()

    def test_stock_audit_syncs_newly_inwarded_stock_items(self):
        """Open audit queue picks up stock items created after the audit started."""
        self._seed_stock_item(item_name="Tomato", qty=10.0)
        page = self.client.get("/stores/stock-audit?outlet=restaurant")
        self.assertEqual(page.status_code, 200)
        self.assertIn(b"Tomato", page.data)
        self.assertNotIn(b"Anar", page.data)

        conn = db_mod.get_db()
        try:
            conn.execute(
                """
                INSERT INTO store_stock_items (outlet, item_name, unit, qty_on_hand, updated_at)
                VALUES ('restaurant', 'Anar', 'kg', 10, datetime('now','localtime'))
                """
            )
            conn.commit()
        finally:
            conn.close()

        page2 = self.client.get("/stores/stock-audit?outlet=restaurant")
        self.assertEqual(page2.status_code, 200)
        self.assertIn(b"Anar", page2.data)

        conn = db_mod.get_db()
        try:
            line = conn.execute(
                """
                SELECT l.item_name, l.system_qty, l.status
                FROM store_stock_audit_lines l
                JOIN store_stock_audits a ON a.id = l.audit_id
                WHERE a.outlet = 'restaurant' AND a.status = 'open'
                  AND lower(l.item_name) = lower('Anar')
                ORDER BY l.id DESC LIMIT 1
                """
            ).fetchone()
            self.assertIsNotNone(line)
            self.assertEqual(line["status"], "pending")
            self.assertAlmostEqual(float(line["system_qty"]), 10.0)
        finally:
            conn.close()

    def test_stock_audit_variance_requires_reason_and_adjusts(self):
        self._seed_stock_item(item_name="Onion", qty=20.0)
        page = self.client.get("/stores/stock-audit?outlet=restaurant")
        self.assertEqual(page.status_code, 200)

        conn = db_mod.get_db()
        try:
            line = conn.execute(
                """
                SELECT l.id FROM store_stock_audit_lines l
                JOIN store_stock_audits a ON a.id = l.audit_id
                WHERE a.outlet = 'restaurant' AND a.status = 'open'
                  AND l.item_name = 'Onion'
                ORDER BY l.id DESC LIMIT 1
                """
            ).fetchone()
            line_id = int(line["id"])
        finally:
            conn.close()

        blocked = self.client.post(
            "/stores/stock-audit/verify",
            json={"line_id": line_id, "actual_qty": 18, "reason": "", "remarks": ""},
        )
        self.assertEqual(blocked.status_code, 400)
        self.assertIn("reason", (blocked.get_json() or {}).get("error", "").lower())

        ok = self.client.post(
            "/stores/stock-audit/verify",
            json={
                "line_id": line_id,
                "actual_qty": 18,
                "reason": "kitchen_wastage",
                "remarks": "Trim loss",
                "go_next": "1",
            },
        )
        self.assertEqual(ok.status_code, 200)
        self.assertTrue((ok.get_json() or {}).get("ok"))

        conn = db_mod.get_db()
        try:
            stock = conn.execute(
                "SELECT qty_on_hand FROM store_stock_items WHERE item_name = 'Onion' AND outlet = 'restaurant'"
            ).fetchone()
            self.assertAlmostEqual(float(stock["qty_on_hand"]), 18.0)
            move = conn.execute(
                """
                SELECT qty_delta, movement_type, notes
                FROM store_stock_movements
                WHERE ref_type = 'stock_audit' AND item_name = 'Onion'
                ORDER BY id DESC LIMIT 1
                """
            ).fetchone()
            self.assertIsNotNone(move)
            self.assertEqual(move["movement_type"], "adjustment")
            self.assertAlmostEqual(float(move["qty_delta"]), -2.0)
            self.assertIn("Kitchen Wastage", move["notes"] or "")
        finally:
            conn.close()

    def test_stock_audit_skip_stays_pending_and_verified_expires(self):
        self._seed_stock_item(item_name="Pepper", qty=12.0)
        page = self.client.get("/stores/stock-audit?outlet=restaurant")
        self.assertEqual(page.status_code, 200)

        conn = db_mod.get_db()
        try:
            line = conn.execute(
                """
                SELECT l.id FROM store_stock_audit_lines l
                JOIN store_stock_audits a ON a.id = l.audit_id
                WHERE a.outlet = 'restaurant' AND a.status = 'open'
                  AND l.item_name = 'Pepper'
                ORDER BY l.id DESC LIMIT 1
                """
            ).fetchone()
            line_id = int(line["id"])
        finally:
            conn.close()

        skipped = self.client.post(
            "/stores/stock-audit/skip",
            json={"line_id": line_id},
        )
        self.assertEqual(skipped.status_code, 200)
        self.assertTrue((skipped.get_json() or {}).get("ok"))

        conn = db_mod.get_db()
        try:
            line = conn.execute(
                "SELECT status FROM store_stock_audit_lines WHERE id = ?",
                (line_id,),
            ).fetchone()
            self.assertEqual(line["status"], "pending")

            verify = self.client.post(
                "/stores/stock-audit/verify",
                json={"line_id": line_id, "actual_qty": 12, "reason": "", "remarks": ""},
            )
            self.assertEqual(verify.status_code, 200)
            self.assertTrue((verify.get_json() or {}).get("ok"))

            stale = (datetime.now() - timedelta(days=8)).strftime("%Y-%m-%d %H:%M:%S")
            conn.execute(
                """
                UPDATE store_stock_audit_lines
                SET verified_at = ?, status = 'verified'
                WHERE id = ?
                """,
                (stale, line_id),
            )
            conn.commit()
        finally:
            conn.close()

        page2 = self.client.get("/stores/stock-audit?outlet=restaurant")
        self.assertEqual(page2.status_code, 200)
        self.assertNotIn(b"Skipped", page2.data)

        conn = db_mod.get_db()
        try:
            line = conn.execute(
                "SELECT status, actual_qty, verified_at FROM store_stock_audit_lines WHERE id = ?",
                (line_id,),
            ).fetchone()
            self.assertEqual(line["status"], "pending")
            self.assertIsNone(line["actual_qty"])
            self.assertIsNone(line["verified_at"])
        finally:
            conn.close()

    def test_stock_audit_report_shows_adjustments_and_export(self):
        self._seed_stock_item(item_name="Carrot", qty=10.0)
        self._seed_stock_item(item_name="Beans", qty=8.0)
        page = self.client.get("/stores/stock-audit?outlet=restaurant")
        self.assertEqual(page.status_code, 200)

        conn = db_mod.get_db()
        try:
            carrot = conn.execute(
                """
                SELECT l.id FROM store_stock_audit_lines l
                JOIN store_stock_audits a ON a.id = l.audit_id
                WHERE a.outlet = 'restaurant' AND a.status = 'open'
                  AND l.item_name = 'Carrot'
                ORDER BY l.id DESC LIMIT 1
                """
            ).fetchone()
            beans = conn.execute(
                """
                SELECT l.id FROM store_stock_audit_lines l
                JOIN store_stock_audits a ON a.id = l.audit_id
                WHERE a.outlet = 'restaurant' AND a.status = 'open'
                  AND l.item_name = 'Beans'
                ORDER BY l.id DESC LIMIT 1
                """
            ).fetchone()
            self.assertIsNotNone(carrot)
            self.assertIsNotNone(beans)
            carrot_id = int(carrot["id"])
            beans_id = int(beans["id"])
        finally:
            conn.close()

        zero = self.client.post(
            "/stores/stock-audit/verify",
            json={"line_id": carrot_id, "actual_qty": 10, "reason": "", "remarks": ""},
        )
        self.assertEqual(zero.status_code, 200)
        self.assertTrue((zero.get_json() or {}).get("ok"))

        adjusted = self.client.post(
            "/stores/stock-audit/verify",
            json={
                "line_id": beans_id,
                "actual_qty": 6,
                "reason": "kitchen_wastage",
                "remarks": "Prep loss",
            },
        )
        self.assertEqual(adjusted.status_code, 200)
        self.assertTrue((adjusted.get_json() or {}).get("ok"))

        report = self.client.get("/stores/stock-audit/report?outlet=restaurant")
        self.assertEqual(report.status_code, 200)
        self.assertIn(b"Stock Audit Report", report.data)
        self.assertIn(b"Beans", report.data)
        self.assertIn(b"Kitchen Wastage", report.data)
        self.assertIn(b"Prep loss", report.data)
        self.assertNotIn(b"Carrot", report.data)

        export = self.client.get("/stores/stock-audit/report/export?outlet=restaurant")
        self.assertEqual(export.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            export.content_type,
        )
        self.assertTrue(export.data[:2] == b"PK")
        self.assertIn("no-store", export.headers.get("Cache-Control", ""))

        from io import BytesIO

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(export.data))
        self.assertEqual(wb.sheetnames, ["Summary", "Line Items"])
        summary = wb["Summary"]
        details = wb["Line Items"]
        self.assertTrue(
            (summary["A1"].value or "").startswith("Hotel Bell Elite — Stock Audit")
        )
        self.assertEqual(summary["A1"].fill.fgColor.rgb, "FF315A78")
        self.assertEqual(summary["A2"].value, "Adjustments")
        self.assertEqual(int(summary["B2"].value), 1)
        self.assertEqual(summary["A3"].value, "Gains")
        self.assertEqual(summary["A4"].value, "Losses")
        self.assertEqual(int(summary["B4"].value), 1)
        self.assertEqual(summary["A5"].value, "Net Variance Value")
        self.assertEqual(summary["A7"].value, "Reason")
        self.assertEqual(summary["A7"].fill.fgColor.rgb, "FF315A78")
        self.assertIn("A1:B1", {str(r) for r in summary.merged_cells.ranges})

        col_a = [
            details.cell(row, 1).value
            for row in range(1, (details.max_row or 1) + 1)
        ]
        banner = "Hotel Bell Elite — Stock Audit - Kitchen Wastage"
        self.assertIn(banner, col_a)
        banner_row = col_a.index(banner) + 1
        self.assertEqual(details.cell(banner_row, 1).fill.fgColor.rgb, "FF315A78")
        headers = [
            details.cell(banner_row + 1, col).value for col in range(1, 15)
        ]
        self.assertEqual(
            headers,
            [
                "Verified at",
                "Outlet",
                "Place",
                "Audit",
                "Product",
                "Category",
                "Unit",
                "System qty",
                "Actual qty",
                "Variance qty",
                "Variance value",
                "Reason",
                "Remarks",
                "Verified by",
            ],
        )
        self.assertEqual(details.cell(banner_row + 1, 1).fill.fgColor.rgb, "FF315A78")
        self.assertEqual(details.cell(banner_row + 2, 5).value, "Beans")
        self.assertEqual(details.cell(banner_row + 2, 13).value, "Prep loss")

        viewer = {
            "id": self.admin_id,
            "username": "stockonly2",
            "full_name": "Stock Only 2",
            "is_admin": False,
            "is_active": True,
            "dashboard_access": {"stores"},
            "stores_access": {"stock"},
            "sales_analytics_access": set(),
            "user_access": set(),
            "payroll_access": set(),
            "accounts_access": set(),
        }
        with mock.patch.object(self.app_mod, "get_current_user", return_value=viewer), mock.patch.object(
            self.stores_mod, "_get_user", return_value=viewer
        ):
            denied = self.client.get("/stores/stock-audit/report?outlet=restaurant")
        self.assertIn(denied.status_code, (302, 403))

    def test_stock_report_export(self):
        self._seed_stock_item(item_name="Cabbage", qty=13.0)
        self._seed_stock_item(item_name="Onion", qty=0.0)
        page = self.client.get("/stores/stock?outlet=restaurant")
        self.assertEqual(page.status_code, 200)
        self.assertIn(b'id="de-nav-stores-stock"', page.data)
        self.assertIn(b">Store</a>", page.data)
        self.assertIn(b"<h1>Store</h1>", page.data)
        self.assertIn(b"id=\"st-stock-place-host\"", page.data)
        self.assertIn(b">Warehouse</button>", page.data)
        self.assertIn(b">Counter</button>", page.data)
        self.assertIn(b">Stock Audit</a>", page.data)
        self.assertIn(b"/stores/stock/export", page.data)
        self.assertIn(b"Download Stock Report Excel", page.data)

        export = self.client.get("/stores/stock/export?outlet=restaurant")
        self.assertEqual(export.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            export.content_type,
        )
        self.assertTrue(export.data[:2] == b"PK")
        self.assertIn(b"Hotel Bell Elite Store Warehouse.xlsx", export.headers.get("Content-Disposition", "").encode())

        from openpyxl import load_workbook
        import io

        wb = load_workbook(io.BytesIO(export.data))
        ws = wb.active
        self.assertEqual(ws.title, "Stock Report")
        headers = [ws.cell(1, c).value for c in range(1, 10)]
        self.assertEqual(
            headers,
            [
                "Product",
                "Category",
                "On hand",
                "Unit",
                "Status",
                "Unit price",
                "Value",
                "Outlet",
                "Place",
            ],
        )
        self.assertTrue(ws.cell(1, 1).font.bold)
        self.assertEqual(ws.cell(1, 1).border.left.style, "thin")
        self.assertEqual(ws.freeze_panes, "A2")

        out_only = self.client.get("/stores/stock/export?outlet=restaurant&status=out")
        self.assertEqual(out_only.status_code, 200)
        self.assertTrue(out_only.data[:2] == b"PK")

    def test_stock_audit_access_gate(self):
        self._seed_stock_item(item_name="Chicken", qty=5.0)
        viewer = {
            "id": self.admin_id,
            "username": "stockonly",
            "full_name": "Stock Only",
            "is_admin": False,
            "is_active": True,
            "dashboard_access": {"stores"},
            "stores_access": {"stock"},
            "sales_analytics_access": set(),
            "user_access": set(),
            "payroll_access": set(),
            "accounts_access": set(),
        }
        with mock.patch.object(self.app_mod, "get_current_user", return_value=viewer), mock.patch.object(
            self.stores_mod, "_get_user", return_value=viewer
        ):
            denied = self.client.get("/stores/stock-audit?outlet=restaurant")
        self.assertIn(denied.status_code, (302, 403))
        if denied.status_code == 302:
            # Permission helper usually flashes and redirects away from the page.
            self.assertNotIn(b'id="st-audit-page"', denied.data)

    def _create_approved_indent_for_po(self, *, notes="PO send indent"):
        """Create product with preferred supplier, raise + approve an indent, return ids."""
        conn = db_mod.get_db()
        try:
            category = conn.execute(
                "SELECT id FROM store_product_categories WHERE is_active = 1 ORDER BY id LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(category)
            category_id = int(category["id"])
            conn.execute(
                "INSERT INTO suppliers (name, gst, address, phone) VALUES (?, '', '', ?)",
                ("PO Alpha Traders", "9876543210"),
            )
            supplier_id = int(
                conn.execute(
                    "SELECT id FROM suppliers WHERE name = 'PO Alpha Traders'"
                ).fetchone()["id"]
            )
            conn.execute(
                """
                INSERT INTO store_products
                    (category_id, name, default_unit, outlet, approximate_price, is_active,
                     preferred_supplier_1_id)
                VALUES (?, 'PO Rice Bag', 'kg', 'restaurant', 50, 1, ?)
                """,
                (category_id, supplier_id),
            )
            conn.execute(
                """
                INSERT INTO store_products
                    (category_id, name, default_unit, outlet, approximate_price, is_active)
                VALUES (?, 'Mystery Unmapped Item', 'pcs', 'restaurant', 20, 1)
                """,
                (category_id,),
            )
            conn.commit()
        finally:
            conn.close()

        create = self.client.post(
            "/stores/indent?outlet=restaurant",
            data={
                "outlet": "restaurant",
                "action": "submit",
                "notes": notes,
                "item_name": ["PO Rice Bag", "Mystery Unmapped Item"],
                "quantity": ["10", "2"],
                "unit": ["kg", "pcs"],
                "approximate_price": ["50", "20"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)

        conn = db_mod.get_db()
        try:
            indent = conn.execute(
                "SELECT id, indent_no, status FROM store_indents WHERE notes = ? ORDER BY id DESC LIMIT 1",
                (notes,),
            ).fetchone()
            self.assertIsNotNone(indent)
            indent_id = int(indent["id"])
            lines = {
                row["item_name"]: int(row["id"])
                for row in conn.execute(
                    "SELECT id, item_name FROM store_indent_lines WHERE indent_id = ?",
                    (indent_id,),
                ).fetchall()
            }
        finally:
            conn.close()

        decide = self.client.post(
            f"/stores/indent/{indent_id}/decide",
            data={"outlet": "restaurant", "decision": "approved", "decision_note": ""},
            follow_redirects=False,
        )
        self.assertEqual(decide.status_code, 302)
        return {
            "indent_id": indent_id,
            "supplier_id": supplier_id,
            "lines": lines,
        }

    def _generate_po_for_supplier(self, indent_id, supplier_id, outlet, line_ids, rates=None):
        """Assign lines to a supplier and generate a purchase order. Returns po id."""
        data = {
            "outlet": outlet,
            "selectable_supplier": [str(supplier_id)],
            "selected_supplier": [str(supplier_id)],
        }
        for lid in line_ids:
            data[f"line_supplier_{lid}"] = str(supplier_id)
            data[f"line_rate_{lid}"] = str((rates or {}).get(lid, "30"))
        res = self.client.post(
            f"/stores/orders/{indent_id}/lines/next",
            data=data,
            follow_redirects=False,
        )
        self.assertEqual(res.status_code, 302)
        conn = db_mod.get_db()
        try:
            po = conn.execute(
                "SELECT id FROM store_purchase_orders WHERE indent_id = ? AND supplier_id = ?",
                (indent_id, supplier_id),
            ).fetchone()
            self.assertIsNotNone(po)
            return int(po["id"])
        finally:
            conn.close()

    def test_po_groups_by_preferred_supplier(self):
        ids = self._create_approved_indent_for_po(notes="PO group test")
        page = self.client.get(f"/stores/orders/{ids['indent_id']}")
        self.assertEqual(page.status_code, 200)
        self.assertIn(b"Suppliers &amp; Items (Grouped)", page.data)
        self.assertIn(b"PO Alpha Traders", page.data)
        self.assertIn(b"Unassigned", page.data)
        self.assertIn(b"<th>Total</th>", page.data)
        self.assertIn(b"st-po-tabs", page.data)
        self.assertIn(b"Generate PO", page.data)
        self.assertIn(b"Send to Supplier", page.data)
        self.assertIn(b'data-po-tab="generate"', page.data)
        # Step 1 reviews items; Generate purchase order issues numbers and leaves them off this page.
        self.assertIn(b"st-po-next-btn", page.data)
        self.assertIn(b"Generate purchase order", page.data)
        self.assertNotIn(b"Send via WhatsApp", page.data)
        # Mixed indent (assigned + unassigned) can proceed — generate stays enabled.
        self.assertNotIn(b"st-po-next-btn is-disabled", page.data)
        # Assigned suppliers are selectable; unassigned groups are not.
        self.assertIn(b"st-po-group-checkbox", page.data)
        self.assertIn(b'name="selected_supplier"', page.data)
        self.assertIn(f'value="{ids["supplier_id"]}"'.encode(), page.data)
        self.assertIn(b"Select all", page.data)

        # Compose preview retired — ?step=compose redirects to the items step.
        compose = self.client.get(
            f"/stores/orders/{ids['indent_id']}?step=compose",
            follow_redirects=False,
        )
        self.assertEqual(compose.status_code, 302)
        self.assertNotIn("step=compose", (compose.headers.get("Location") or ""))

        items = self.client.get(f"/stores/orders/{ids['indent_id']}")
        self.assertEqual(items.status_code, 200)
        self.assertIn(b"Generate purchase order", items.data)
        self.assertNotIn(b"st-po-message-preview", items.data)
        self.assertNotIn(b"Send Purchase Order to Supplier", items.data)
        self.assertNotIn(b"Send via WhatsApp", items.data)

        conn = db_mod.get_db()
        try:
            payload = self.stores_mod._load_po_supplier_groups(conn, ids["indent_id"])
        finally:
            conn.close()
        self.assertTrue(payload)
        names = [g["supplier_name"] for g in payload["groups"]]
        self.assertIn("Unassigned", names)
        self.assertIn("PO Alpha Traders", names)
        assigned = next(g for g in payload["groups"] if g["supplier_name"] == "PO Alpha Traders")
        self.assertEqual(assigned["item_count"], 1)
        self.assertAlmostEqual(assigned["estimated_value"], 500.0)
        self.assertTrue(assigned["can_send"])
        unassigned = next(g for g in payload["groups"] if g.get("is_unassigned"))
        self.assertEqual(unassigned["item_count"], 1)
        self.assertFalse(unassigned["can_send"])

    def test_po_next_respects_selected_suppliers(self):
        ids = self._create_approved_indent_for_po(notes="PO select suppliers")
        rice_line = ids["lines"]["PO Rice Bag"]
        mystery_line = ids["lines"]["Mystery Unmapped Item"]

        conn = db_mod.get_db()
        try:
            conn.execute(
                "INSERT INTO suppliers (name, gst, address, phone) VALUES (?, '', '', ?)",
                ("PO Beta Foods", "9123456780"),
            )
            beta_id = int(
                conn.execute(
                    "SELECT id FROM suppliers WHERE name = 'PO Beta Foods'"
                ).fetchone()["id"]
            )
            conn.commit()
        finally:
            conn.close()

        # Assign both lines so two supplier groups exist.
        save = self.client.post(
            f"/stores/orders/{ids['indent_id']}/lines",
            json={
                "lines": [
                    {"line_id": rice_line, "supplier_id": ids["supplier_id"], "rate": 50},
                    {"line_id": mystery_line, "supplier_id": beta_id, "rate": 25},
                ]
            },
        )
        self.assertEqual(save.status_code, 200)

        # Generate with only Alpha selected — issues PO, does not auto-send WhatsApp.
        nxt = self.client.post(
            f"/stores/orders/{ids['indent_id']}/lines/next",
            data={
                "outlet": "restaurant",
                f"line_supplier_{rice_line}": str(ids["supplier_id"]),
                f"line_rate_{rice_line}": "50",
                f"line_supplier_{mystery_line}": str(beta_id),
                f"line_rate_{mystery_line}": "25",
                "selectable_supplier": [str(ids["supplier_id"]), str(beta_id)],
                "selected_supplier": [str(ids["supplier_id"])],
            },
            follow_redirects=False,
        )
        self.assertEqual(nxt.status_code, 302)
        loc = nxt.headers.get("Location") or ""
        self.assertNotIn("step=compose", loc)
        # After generate, land on Send to Supplier (not back on Generate).
        self.assertIn("/stores/orders", loc)
        self.assertIn("tab=send", loc)
        self.assertNotIn(f"/stores/orders/{ids['indent_id']}", loc.split("?")[0])

        items = self.client.get(f"/stores/orders/{ids['indent_id']}")
        self.assertEqual(items.status_code, 200)
        self.assertIn(b"Generate purchase order", items.data)
        self.assertNotIn(b"Send Purchase Order to Supplier", items.data)
        self.assertNotIn(b"st-po-message-preview", items.data)
        # Generated Alpha group leaves the Send list; Beta remains pending.
        self.assertNotIn(b'<strong class="st-po-group-name">PO Alpha Traders</strong>', items.data)
        self.assertIn(b'<strong class="st-po-group-name">PO Beta Foods</strong>', items.data)
        self.assertIn(b'>1 Supplier</span>', items.data)

        send_tab = self.client.get("/stores/orders?outlet=restaurant&tab=send")
        self.assertEqual(send_tab.status_code, 200)
        self.assertIn(b"Send to Supplier", send_tab.data)

        conn = db_mod.get_db()
        try:
            po_row = conn.execute(
                "SELECT id, po_no FROM store_purchase_orders WHERE indent_id = ? AND supplier_id = ?",
                (ids["indent_id"], ids["supplier_id"]),
            ).fetchone()
            frozen = conn.execute(
                "SELECT COUNT(*) AS n FROM store_purchase_order_lines WHERE purchase_order_id = ?",
                (int(po_row["id"]),),
            ).fetchone()
            send_row = conn.execute(
                """
                SELECT status, supplier_id, po_no FROM store_po_sends
                WHERE indent_id = ? AND supplier_id = ?
                ORDER BY id DESC LIMIT 1
                """,
                (ids["indent_id"], ids["supplier_id"]),
            ).fetchone()
        finally:
            conn.close()
        self.assertIsNotNone(po_row)
        self.assertRegex(str(po_row["po_no"]), r"^PO/RES/\d{2}-\d{2}/\d+$")
        self.assertGreaterEqual(int(frozen["n"]), 1)
        # WhatsApp is deferred to the send-confirm popup / Purchase Orders action.
        self.assertIsNone(send_row)

        # Clearing every supplier is rejected.
        refused = self.client.post(
            f"/stores/orders/{ids['indent_id']}/lines/next",
            data={
                "outlet": "restaurant",
                f"line_supplier_{mystery_line}": str(beta_id),
                f"line_rate_{mystery_line}": "25",
                "selectable_supplier": [str(beta_id)],
            },
            follow_redirects=False,
        )
        self.assertEqual(refused.status_code, 302)
        self.assertNotIn("step=compose", refused.headers.get("Location") or "")
        # Beta is still pending until generated.
        still = self.client.get(f"/stores/orders/{ids['indent_id']}")
        self.assertIn(b"PO Beta Foods", still.data)
        # Indent stays in the Generate PO picker while any supplier is pending.
        self.assertIn(f'id="st-po-indent" value="{ids["indent_id"]}"'.encode(), still.data)
        self.assertIn(b'st-po-indent-list', still.data)

        with mock.patch.dict(os.environ, {"WHATSAPP_DRY_RUN": "1"}, clear=False):
            done = self.client.post(
                f"/stores/orders/{ids['indent_id']}/lines/next",
                data={
                    "outlet": "restaurant",
                    f"line_supplier_{mystery_line}": str(beta_id),
                    f"line_rate_{mystery_line}": "25",
                    "selectable_supplier": [str(beta_id)],
                    "selected_supplier": [str(beta_id)],
                },
                follow_redirects=False,
            )
        self.assertEqual(done.status_code, 302)
        # Opening the finished indent bounces away — it is no longer in Generate PO.
        bounced = self.client.get(
            f"/stores/orders/{ids['indent_id']}",
            follow_redirects=False,
        )
        self.assertEqual(bounced.status_code, 302)
        bounce_path = (bounced.headers.get("Location") or "").split("?")[0]
        self.assertNotEqual(bounce_path.rstrip("/"), f"/stores/orders/{ids['indent_id']}")

        listing = self.client.get("/stores/orders?outlet=restaurant&tab=send")
        self.assertEqual(listing.status_code, 200)
        # Finished indent is absent from the Generate PO indent options.
        self.assertNotIn(
            f'data-value="{ids["indent_id"]}"'.encode(),
            listing.data,
        )
        # Generate PO must not deep-link a finished indent (server 302s back → looks broken).
        self.assertNotRegex(
            listing.data,
            rf'<a class="cp-view-tab[^"]*"[^>]*href="/stores/orders/{ids["indent_id"]}(\?|")'.encode(),
        )

    def test_po_line_overrides_persist(self):
        ids = self._create_approved_indent_for_po(notes="PO override test")
        rice_line = ids["lines"]["PO Rice Bag"]
        mystery_line = ids["lines"]["Mystery Unmapped Item"]

        conn = db_mod.get_db()
        try:
            conn.execute(
                "INSERT INTO suppliers (name, gst, address, phone) VALUES (?, '', '', ?)",
                ("PO Beta Foods", "9123456780"),
            )
            beta_id = int(
                conn.execute(
                    "SELECT id FROM suppliers WHERE name = 'PO Beta Foods'"
                ).fetchone()["id"]
            )
            conn.commit()
        finally:
            conn.close()

        save = self.client.post(
            f"/stores/orders/{ids['indent_id']}/lines",
            json={
                "lines": [
                    {"line_id": rice_line, "supplier_id": beta_id, "rate": 55},
                    {"line_id": mystery_line, "supplier_id": beta_id, "rate": 25},
                ]
            },
        )
        self.assertEqual(save.status_code, 200)
        payload = save.get_json()
        self.assertTrue(payload.get("ok"))

        conn = db_mod.get_db()
        try:
            rows = {
                int(row["line_id"]): dict(row)
                for row in conn.execute(
                    "SELECT line_id, supplier_id, rate FROM store_po_lines WHERE indent_id = ?",
                    (ids["indent_id"],),
                ).fetchall()
            }
            grouped = self.stores_mod._load_po_supplier_groups(conn, ids["indent_id"])
        finally:
            conn.close()

        self.assertEqual(int(rows[rice_line]["supplier_id"]), beta_id)
        self.assertAlmostEqual(float(rows[rice_line]["rate"]), 55.0)
        self.assertEqual(int(rows[mystery_line]["supplier_id"]), beta_id)
        beta_group = next(g for g in grouped["groups"] if g["supplier_id"] == beta_id)
        self.assertEqual(beta_group["item_count"], 2)
        self.assertAlmostEqual(beta_group["estimated_value"], 10 * 55 + 2 * 25)

    def test_po_partial_quantity_override(self):
        ids = self._create_approved_indent_for_po(notes="PO partial qty")
        rice_line = ids["lines"]["PO Rice Bag"]
        mystery_line = ids["lines"]["Mystery Unmapped Item"]

        page = self.client.get(f"/stores/orders/{ids['indent_id']}")
        self.assertEqual(page.status_code, 200)
        self.assertIn(b'st-po-line-qty', page.data)
        self.assertIn(f'name="line_qty_{rice_line}"'.encode(), page.data)

        save = self.client.post(
            f"/stores/orders/{ids['indent_id']}/lines",
            json={
                "lines": [
                    {
                        "line_id": rice_line,
                        "supplier_id": ids["supplier_id"],
                        "rate": 50,
                        "quantity": 4,
                    },
                    {
                        "line_id": mystery_line,
                        "supplier_id": ids["supplier_id"],
                        "rate": 25,
                        "quantity": 1,
                    },
                ]
            },
        )
        self.assertEqual(save.status_code, 200)
        self.assertTrue(save.get_json().get("ok"))

        conn = db_mod.get_db()
        try:
            rows = {
                int(row["line_id"]): dict(row)
                for row in conn.execute(
                    "SELECT line_id, supplier_id, rate, quantity FROM store_po_lines WHERE indent_id = ?",
                    (ids["indent_id"],),
                ).fetchall()
            }
            grouped = self.stores_mod._load_po_supplier_groups(conn, ids["indent_id"])
            # Over-indent qty is capped to the indent line quantity (10).
            over = self.client.post(
                f"/stores/orders/{ids['indent_id']}/lines",
                json={
                    "lines": [
                        {
                            "line_id": rice_line,
                            "supplier_id": ids["supplier_id"],
                            "rate": 50,
                            "quantity": 99,
                        }
                    ]
                },
            )
            self.assertEqual(over.status_code, 200)
            capped = conn.execute(
                "SELECT quantity FROM store_po_lines WHERE line_id = ?",
                (rice_line,),
            ).fetchone()
        finally:
            conn.close()

        self.assertAlmostEqual(float(rows[rice_line]["quantity"]), 4.0)
        self.assertAlmostEqual(float(rows[mystery_line]["quantity"]), 1.0)
        alpha = next(g for g in grouped["groups"] if g["supplier_id"] == ids["supplier_id"])
        rice = next(l for l in alpha["lines"] if l["line_id"] == rice_line)
        self.assertAlmostEqual(float(rice["quantity"]), 4.0)
        self.assertAlmostEqual(float(rice["indent_quantity"]), 10.0)
        self.assertTrue(rice["quantity_is_partial"])
        self.assertAlmostEqual(float(capped["quantity"]), 10.0)

    def test_po_partial_generate_keeps_remaining_qty(self):
        """After generating a partial PO, remaining indent qty stays on Generate."""
        ids = self._create_approved_indent_for_po(notes="PO remaining after partial")
        rice_line = ids["lines"]["PO Rice Bag"]
        mystery_line = ids["lines"]["Mystery Unmapped Item"]

        self.client.post(
            f"/stores/orders/{ids['indent_id']}/lines",
            json={
                "lines": [
                    {
                        "line_id": rice_line,
                        "supplier_id": ids["supplier_id"],
                        "rate": 50,
                        "quantity": 4,
                    },
                    {
                        "line_id": mystery_line,
                        "supplier_id": ids["supplier_id"],
                        "rate": 25,
                        "quantity": 1,
                    },
                ]
            },
        )
        with mock.patch.dict(os.environ, {"WHATSAPP_DRY_RUN": "1"}, clear=False):
            gen = self.client.post(
                f"/stores/orders/{ids['indent_id']}/lines/next",
                data={
                    "outlet": "restaurant",
                    f"line_supplier_{rice_line}": str(ids["supplier_id"]),
                    f"line_rate_{rice_line}": "50",
                    f"line_qty_{rice_line}": "4",
                    f"line_supplier_{mystery_line}": str(ids["supplier_id"]),
                    f"line_rate_{mystery_line}": "25",
                    f"line_qty_{mystery_line}": "1",
                    "selectable_supplier": [str(ids["supplier_id"])],
                    "selected_supplier": [str(ids["supplier_id"])],
                },
                follow_redirects=False,
            )
        self.assertEqual(gen.status_code, 302)

        conn = db_mod.get_db()
        try:
            rice = conn.execute(
                "SELECT quantity, quantity_ordered FROM store_indent_lines WHERE id = ?",
                (rice_line,),
            ).fetchone()
            mystery = conn.execute(
                "SELECT quantity, quantity_ordered FROM store_indent_lines WHERE id = ?",
                (mystery_line,),
            ).fetchone()
            grouped = self.stores_mod._load_po_supplier_groups(conn, ids["indent_id"])
            pending = self.stores_mod._pending_po_groups(grouped.get("groups") or [])
        finally:
            conn.close()

        self.assertAlmostEqual(float(rice["quantity_ordered"]), 4.0)
        self.assertAlmostEqual(float(mystery["quantity_ordered"]), 1.0)
        self.assertTrue(pending)
        alpha = next(g for g in pending if g["supplier_id"] == ids["supplier_id"])
        rice_row = next(l for l in alpha["lines"] if l["line_id"] == rice_line)
        self.assertAlmostEqual(float(rice_row["remaining_quantity"]), 6.0)
        self.assertAlmostEqual(float(rice_row["quantity"]), 6.0)
        self.assertAlmostEqual(float(rice_row["indent_quantity"]), 10.0)

        page = self.client.get(f"/stores/orders/{ids['indent_id']}")
        self.assertEqual(page.status_code, 200)
        self.assertIn(b"PO Rice Bag", page.data)
        self.assertIn(b'value="6"', page.data)
        # Hint shows available remaining, not the original approved indent qty.
        self.assertIn(b"of 6", page.data)
        self.assertNotIn(b"of 10", page.data)

    def test_po_pack_filled_from_product_master_when_indent_missing(self):
        """Generate PO Pack column uses Product Master variant when indent pack is blank."""
        ids = self._create_approved_indent_for_po(notes="PO pack from product master")
        indent_id = int(ids["indent_id"])
        rice_line = int(ids["lines"]["PO Rice Bag"])
        product_id = None

        conn = db_mod.get_db()
        try:
            product = conn.execute(
                "SELECT id FROM store_products WHERE name = 'PO Rice Bag' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            self.assertIsNotNone(product)
            product_id = int(product["id"])
            conn.execute(
                """
                INSERT INTO store_product_variants
                    (product_id, label, qty_in_base, approximate_price, sort_order, is_active)
                VALUES (?, '1 kg', 1, 50, 10, 1)
                """,
                (product_id,),
            )
            conn.execute(
                """
                UPDATE store_indent_lines
                SET pack_label = '', pack_qty_in_base = NULL
                WHERE id = ?
                """,
                (rice_line,),
            )
            conn.commit()
            grouped = self.stores_mod._load_po_supplier_groups(conn, indent_id)
            healed = conn.execute(
                "SELECT pack_label, pack_qty_in_base FROM store_indent_lines WHERE id = ?",
                (rice_line,),
            ).fetchone()
        finally:
            conn.close()

        alpha = next(
            g for g in (grouped.get("groups") or []) if g.get("supplier_id") == ids["supplier_id"]
        )
        rice_row = next(l for l in alpha["lines"] if l["line_id"] == rice_line)
        self.assertEqual(rice_row["pack_label"], "1 kg")
        self.assertAlmostEqual(float(rice_row["pack_qty_in_base"]), 1.0)
        self.assertEqual(healed["pack_label"], "1 kg")
        self.assertAlmostEqual(float(healed["pack_qty_in_base"]), 1.0)

        page = self.client.get(f"/stores/orders/{indent_id}")
        self.assertEqual(page.status_code, 200)
        self.assertIn(b"1 kg", page.data)

    def test_po_number_series_is_stable_per_supplier(self):
        ids = self._create_approved_indent_for_po(notes="PO number test")
        rice_line = ids["lines"]["PO Rice Bag"]
        with mock.patch.dict(os.environ, {"WHATSAPP_DRY_RUN": "1"}, clear=False):
            first = self.client.post(
                f"/stores/orders/{ids['indent_id']}/lines/next",
                data={
                    "outlet": "restaurant",
                    f"line_supplier_{rice_line}": str(ids["supplier_id"]),
                    f"line_rate_{rice_line}": "50",
                    "selectable_supplier": [str(ids["supplier_id"])],
                    "selected_supplier": [str(ids["supplier_id"])],
                },
                follow_redirects=False,
            )
        self.assertEqual(first.status_code, 302)

        conn = db_mod.get_db()
        try:
            row = conn.execute(
                "SELECT po_no FROM store_purchase_orders WHERE indent_id = ? AND supplier_id = ?",
                (ids["indent_id"], ids["supplier_id"]),
            ).fetchone()
            self.assertIsNotNone(row)
            po_no = str(row["po_no"])
            short_fy = self.stores_mod._short_fiscal_year_label()
        finally:
            conn.close()
        self.assertRegex(po_no, r"^PO/RES/\d{2}-\d{2}/\d+$")
        self.assertIn(f"/{short_fy}/", po_no)

        # Generating again for the same supplier reuses the issued number.
        with mock.patch.dict(os.environ, {"WHATSAPP_DRY_RUN": "1"}, clear=False):
            again = self.client.post(
                f"/stores/orders/{ids['indent_id']}/lines/next",
                data={
                    "outlet": "restaurant",
                    f"line_supplier_{rice_line}": str(ids["supplier_id"]),
                    f"line_rate_{rice_line}": "50",
                    "selectable_supplier": [str(ids["supplier_id"])],
                    "selected_supplier": [str(ids["supplier_id"])],
                },
                follow_redirects=False,
            )
        self.assertEqual(again.status_code, 302)

        # A second supplier group on the same indent takes the next number in the series.
        conn = db_mod.get_db()
        try:
            conn.execute(
                "INSERT INTO suppliers (name, gst, address, phone) VALUES (?, '', '', ?)",
                ("PO Gamma Supplies", "9000000001"),
            )
            gamma_id = int(
                conn.execute(
                    "SELECT id FROM suppliers WHERE name = 'PO Gamma Supplies'"
                ).fetchone()["id"]
            )
            conn.commit()
        finally:
            conn.close()
        assign = self.client.post(
            f"/stores/orders/{ids['indent_id']}/lines",
            json={"lines": [{"line_id": ids["lines"]["Mystery Unmapped Item"], "supplier_id": gamma_id}]},
        )
        self.assertEqual(assign.status_code, 200)
        mystery_line = ids["lines"]["Mystery Unmapped Item"]
        with mock.patch.dict(os.environ, {"WHATSAPP_DRY_RUN": "1"}, clear=False):
            gamma_gen = self.client.post(
                f"/stores/orders/{ids['indent_id']}/lines/next",
                data={
                    "outlet": "restaurant",
                    f"line_supplier_{mystery_line}": str(gamma_id),
                    f"line_rate_{mystery_line}": "25",
                    "selectable_supplier": [str(gamma_id)],
                    "selected_supplier": [str(gamma_id)],
                },
                follow_redirects=False,
            )
        self.assertEqual(gamma_gen.status_code, 302)

        conn = db_mod.get_db()
        try:
            numbers = {
                int(r["supplier_id"]): str(r["po_no"])
                for r in conn.execute(
                    "SELECT supplier_id, po_no FROM store_purchase_orders WHERE indent_id = ?",
                    (ids["indent_id"],),
                ).fetchall()
            }
            # Alpha still has the original number after the no-op second generate attempt.
            alpha_count = conn.execute(
                "SELECT COUNT(*) AS c FROM store_purchase_orders WHERE indent_id = ? AND supplier_id = ?",
                (ids["indent_id"], ids["supplier_id"]),
            ).fetchone()["c"]
        finally:
            conn.close()
        self.assertEqual(int(alpha_count), 1)
        self.assertEqual(len(numbers), 2)
        self.assertEqual(numbers[ids["supplier_id"]], po_no)
        self.assertNotEqual(numbers[gamma_id], po_no)
        seq_of = lambda value: int(value.rstrip("/").split("/")[-1])
        self.assertEqual(seq_of(numbers[gamma_id]), seq_of(po_no) + 1)

        # Purchase Orders tab lists every generated PO number.
        listing = self.client.get("/stores/orders?outlet=restaurant")
        self.assertEqual(listing.status_code, 200)
        self.assertIn(b"Stock Inward - Pending", listing.data)
        self.assertIn(po_no.encode(), listing.data)
        self.assertIn(numbers[gamma_id].encode(), listing.data)
        self.assertIn(b"PO Alpha Traders", listing.data)
        self.assertIn(b"PO Gamma Supplies", listing.data)
        self.assertIn(b"View PDF", listing.data)
        self.assertIn(b'data-st-po-pdf', listing.data)
        self.assertIn(b'id="st-po-pdf-modal"', listing.data)

    def test_po_pdf_endpoint(self):
        ids = self._create_approved_indent_for_po(notes="PO pdf test")
        ok = self.client.get(
            f"/stores/orders/{ids['indent_id']}/pdf/{ids['supplier_id']}"
        )
        self.assertEqual(ok.status_code, 200)
        self.assertEqual(ok.mimetype, "application/pdf")
        self.assertTrue(ok.data.startswith(b"%PDF"))
        # The document is filed under its PO number, not the indent number.
        self.assertIn("PO_RES_", ok.headers.get("Content-Disposition", ""))
        conn = db_mod.get_db()
        try:
            po_no = str(
                conn.execute(
                    "SELECT po_no FROM store_purchase_orders WHERE indent_id = ? AND supplier_id = ?",
                    (ids["indent_id"], ids["supplier_id"]),
                ).fetchone()["po_no"]
            )
        finally:
            conn.close()
        self.assertIn(po_no.replace("/", "_"), ok.headers.get("Content-Disposition", ""))

        # Pending indent must be refused.
        create = self.client.post(
            "/stores/indent?outlet=restaurant",
            data={
                "outlet": "restaurant",
                "action": "submit",
                "notes": "PO pending pdf",
                "item_name": ["PO Rice Bag"],
                "quantity": ["1"],
                "unit": ["kg"],
                "approximate_price": ["50"],
            },
            follow_redirects=False,
        )
        self.assertEqual(create.status_code, 302)
        conn = db_mod.get_db()
        try:
            pending_id = int(
                conn.execute(
                    "SELECT id FROM store_indents WHERE notes = 'PO pending pdf' ORDER BY id DESC LIMIT 1"
                ).fetchone()["id"]
            )
        finally:
            conn.close()
        refused = self.client.get(
            f"/stores/orders/{pending_id}/pdf/{ids['supplier_id']}",
            follow_redirects=False,
        )
        self.assertIn(refused.status_code, (302, 400))

    def test_po_send_dry_run_records_history(self):
        ids = self._create_approved_indent_for_po(notes="PO dry run send")
        rice_line = ids["lines"]["PO Rice Bag"]
        gen = self.client.post(
            f"/stores/orders/{ids['indent_id']}/lines/next",
            data={
                "outlet": "restaurant",
                f"line_supplier_{rice_line}": str(ids["supplier_id"]),
                f"line_rate_{rice_line}": "50",
                "selectable_supplier": [str(ids["supplier_id"])],
                "selected_supplier": [str(ids["supplier_id"])],
            },
            follow_redirects=False,
        )
        self.assertEqual(gen.status_code, 302)
        conn = db_mod.get_db()
        try:
            po_row = conn.execute(
                "SELECT po_no FROM store_purchase_orders WHERE indent_id = ? AND supplier_id = ?",
                (ids["indent_id"], ids["supplier_id"]),
            ).fetchone()
            send_before = conn.execute(
                "SELECT COUNT(*) AS n FROM store_po_sends WHERE indent_id = ?",
                (ids["indent_id"],),
            ).fetchone()
        finally:
            conn.close()
        self.assertIsNotNone(po_row)
        self.assertEqual(int(send_before["n"]), 0)

        with mock.patch.dict(os.environ, {"WHATSAPP_DRY_RUN": "1"}, clear=False):
            send = self.client.post(
                f"/stores/orders/{ids['indent_id']}/send",
                json={
                    "supplier_id": ids["supplier_id"],
                    "po_no": str(po_row["po_no"]),
                    "include_pdf": True,
                    "message": "Hello PO Alpha Traders, please supply rice.",
                },
            )
        self.assertEqual(send.status_code, 200)
        payload = send.get_json()
        self.assertTrue(payload.get("ok"))
        self.assertTrue(payload.get("dry_run"))
        self.assertTrue(payload.get("conversation_id"))
        self.assertEqual(str(payload.get("po_no") or ""), str(po_row["po_no"]))

        conn = db_mod.get_db()
        try:
            row = conn.execute(
                """
                SELECT status, supplier_id, phone, pdf_name, conversation_id, include_pdf, po_no
                FROM store_po_sends
                WHERE indent_id = ?
                ORDER BY id DESC LIMIT 1
                """,
                (ids["indent_id"],),
            ).fetchone()
            self.assertIsNotNone(row)
            self.assertEqual(row["status"], "sent")
            self.assertEqual(int(row["supplier_id"]), ids["supplier_id"])
            self.assertTrue(row["pdf_name"])
            self.assertEqual(int(row["include_pdf"]), 1)
            self.assertEqual(str(row["po_no"]), payload.get("po_no"))
            self.assertIn(str(row["po_no"]).replace("/", "_"), str(row["pdf_name"]))
            conv = conn.execute(
                "SELECT id, phone_e164 FROM wa_conversations WHERE id = ?",
                (int(row["conversation_id"]),),
            ).fetchone()
            self.assertIsNotNone(conv)
            self.assertTrue(str(conv["phone_e164"]).endswith("9876543210") or "9876543210" in str(conv["phone_e164"]))
        finally:
            conn.close()

        history = self.client.get("/stores/orders/history", follow_redirects=False)
        self.assertEqual(history.status_code, 302)
        self.assertIn("/stores/orders", history.headers.get("Location", ""))

        # Send still records history rows even though the History tab is gone.
        conn = db_mod.get_db()
        try:
            hist = conn.execute(
                """
                SELECT s.po_no, sp.name AS supplier_name
                FROM store_po_sends s
                LEFT JOIN suppliers sp ON sp.id = s.supplier_id
                WHERE s.po_no = ?
                ORDER BY s.id DESC LIMIT 1
                """,
                (payload.get("po_no"),),
            ).fetchone()
            self.assertIsNotNone(hist)
            self.assertIn("PO Alpha Traders", hist["supplier_name"] or "")
        finally:
            conn.close()

    def test_po_send_reconstructs_missing_frozen_lines(self):
        """Older POs without store_purchase_order_lines rows can still be sent."""
        ids = self._create_approved_indent_for_po(notes="PO send reconstruct")
        rice_line = ids["lines"]["PO Rice Bag"]
        gen = self.client.post(
            f"/stores/orders/{ids['indent_id']}/lines/next",
            data={
                "outlet": "restaurant",
                f"line_supplier_{rice_line}": str(ids["supplier_id"]),
                f"line_rate_{rice_line}": "50",
                "selectable_supplier": [str(ids["supplier_id"])],
                "selected_supplier": [str(ids["supplier_id"])],
            },
            follow_redirects=False,
        )
        self.assertEqual(gen.status_code, 302)
        conn = db_mod.get_db()
        try:
            po_row = conn.execute(
                "SELECT id, po_no FROM store_purchase_orders WHERE indent_id = ? AND supplier_id = ?",
                (ids["indent_id"], ids["supplier_id"]),
            ).fetchone()
            self.assertIsNotNone(po_row)
            conn.execute(
                "DELETE FROM store_purchase_order_lines WHERE purchase_order_id = ?",
                (int(po_row["id"]),),
            )
            conn.commit()
            left = conn.execute(
                "SELECT COUNT(*) AS n FROM store_purchase_order_lines WHERE purchase_order_id = ?",
                (int(po_row["id"]),),
            ).fetchone()
            self.assertEqual(int(left["n"]), 0)
        finally:
            conn.close()

        with mock.patch.dict(os.environ, {"WHATSAPP_DRY_RUN": "1"}, clear=False):
            send = self.client.post(
                f"/stores/orders/{ids['indent_id']}/send",
                json={
                    "supplier_id": ids["supplier_id"],
                    "po_no": str(po_row["po_no"]),
                    "include_pdf": True,
                },
            )
        self.assertEqual(send.status_code, 200)
        payload = send.get_json()
        self.assertTrue(payload.get("ok"))
        self.assertEqual(str(payload.get("supplier_name") or ""), "PO Alpha Traders")

        conn = db_mod.get_db()
        try:
            backfilled = conn.execute(
                "SELECT COUNT(*) AS n FROM store_purchase_order_lines WHERE purchase_order_id = ?",
                (int(po_row["id"]),),
            ).fetchone()
            self.assertGreaterEqual(int(backfilled["n"]), 1)
            sent = conn.execute(
                "SELECT status FROM store_po_sends WHERE indent_id = ? ORDER BY id DESC LIMIT 1",
                (ids["indent_id"],),
            ).fetchone()
            self.assertEqual(sent["status"], "sent")
        finally:
            conn.close()

    def test_po_generate_json_returns_issued_without_sending(self):
        ids = self._create_approved_indent_for_po(notes="PO generate JSON")
        rice_line = ids["lines"]["PO Rice Bag"]
        mystery_line = ids["lines"]["Mystery Unmapped Item"]
        self.client.post(
            f"/stores/orders/{ids['indent_id']}/lines",
            json={
                "lines": [
                    {"line_id": rice_line, "supplier_id": ids["supplier_id"], "rate": 50},
                    {"line_id": mystery_line, "supplier_id": ids["supplier_id"], "rate": 25},
                ]
            },
        )
        gen = self.client.post(
            f"/stores/orders/{ids['indent_id']}/lines/next",
            data={
                "outlet": "restaurant",
                f"line_supplier_{rice_line}": str(ids["supplier_id"]),
                f"line_rate_{rice_line}": "50",
                f"line_supplier_{mystery_line}": str(ids["supplier_id"]),
                f"line_rate_{mystery_line}": "25",
                "selectable_supplier": [str(ids["supplier_id"])],
                "selected_supplier": [str(ids["supplier_id"])],
            },
            headers={
                "Accept": "application/json",
                "X-Requested-With": "XMLHttpRequest",
            },
        )
        self.assertEqual(gen.status_code, 200)
        payload = gen.get_json()
        self.assertTrue(payload.get("ok"))
        issued = payload.get("issued") or []
        self.assertEqual(len(issued), 1)
        self.assertEqual(int(issued[0]["supplier_id"]), ids["supplier_id"])
        self.assertTrue(issued[0].get("can_send"))
        self.assertRegex(str(issued[0].get("po_no") or ""), r"^PO/RES/\d{2}-\d{2}/\d+$")
        self.assertIn("tab=send", str(payload.get("redirect") or ""))

        conn = db_mod.get_db()
        try:
            sends = conn.execute(
                "SELECT COUNT(*) AS n FROM store_po_sends WHERE indent_id = ?",
                (ids["indent_id"],),
            ).fetchone()
            frozen = conn.execute(
                """
                SELECT COUNT(*) AS n
                FROM store_purchase_order_lines pol
                JOIN store_purchase_orders po ON po.id = pol.purchase_order_id
                WHERE po.indent_id = ?
                """,
                (ids["indent_id"],),
            ).fetchone()
        finally:
            conn.close()
        self.assertEqual(int(sends["n"]), 0)
        self.assertEqual(int(frozen["n"]), 2)

    def test_po_send_tab_shows_unsent_pending_inward_only(self):
        ids = self._create_approved_indent_for_po(notes="PO send queue filter")
        rice_line = ids["lines"]["PO Rice Bag"]
        gen = self.client.post(
            f"/stores/orders/{ids['indent_id']}/lines/next",
            data={
                "outlet": "restaurant",
                f"line_supplier_{rice_line}": str(ids["supplier_id"]),
                f"line_rate_{rice_line}": "50",
                "selectable_supplier": [str(ids["supplier_id"])],
                "selected_supplier": [str(ids["supplier_id"])],
            },
            follow_redirects=False,
        )
        self.assertEqual(gen.status_code, 302)

        conn = db_mod.get_db()
        try:
            po_row = conn.execute(
                "SELECT po_no FROM store_purchase_orders WHERE indent_id = ? AND supplier_id = ?",
                (ids["indent_id"], ids["supplier_id"]),
            ).fetchone()
        finally:
            conn.close()
        self.assertIsNotNone(po_row)
        po_no = str(po_row["po_no"])

        send_tab = self.client.get("/stores/orders?tab=send&outlet=restaurant")
        self.assertEqual(send_tab.status_code, 200)
        self.assertIn(b"Send to Supplier", send_tab.data)
        self.assertIn(po_no.encode(), send_tab.data)
        # Unsent POs stay eligible for Stock Inward from the Send tab.
        self.assertIn(
            f"/stores/purchase-requests?outlet=restaurant&amp;view=approved&amp;po_id=".encode(),
            send_tab.data,
        )
        self.assertIn(b"Open stock inward for", send_tab.data)

        orders_tab = self.client.get("/stores/orders?outlet=restaurant")
        self.assertEqual(orders_tab.status_code, 200)
        self.assertIn(po_no.encode(), orders_tab.data)
        self.assertIn(b"Stock Inward - Pending", orders_tab.data)
        self.assertIn(b"Stock Inward - Completed", orders_tab.data)
        self.assertIn(b'id="st-po-pending-table"', orders_tab.data)
        self.assertIn(
            f"/stores/purchase-requests?outlet=restaurant&amp;view=approved&amp;po_id=".encode(),
            orders_tab.data,
        )

        with mock.patch.dict(os.environ, {"WHATSAPP_DRY_RUN": "1"}, clear=False):
            send = self.client.post(
                f"/stores/orders/{ids['indent_id']}/send",
                json={
                    "supplier_id": ids["supplier_id"],
                    "po_no": po_no,
                    "include_pdf": True,
                },
            )
        self.assertEqual(send.status_code, 200)
        self.assertTrue(send.get_json().get("ok"))

        send_tab_after = self.client.get("/stores/orders?tab=send&outlet=restaurant")
        self.assertEqual(send_tab_after.status_code, 200)
        self.assertNotIn(po_no.encode(), send_tab_after.data)

        orders_after = self.client.get("/stores/orders?outlet=restaurant")
        self.assertIn(po_no.encode(), orders_after.data)

        # Orphan PO (no frozen lines + no current supplier assignment) stays off Send queue.
        conn = db_mod.get_db()
        try:
            cur = conn.execute(
                "INSERT INTO suppliers (name, gst, address, phone) VALUES (?, '', '', ?)",
                ("PO Orphan Supplier", "9000000000"),
            )
            orphan_supplier_id = int(cur.lastrowid)
            conn.execute(
                """
                INSERT INTO store_purchase_orders (indent_id, supplier_id, po_no, created_at)
                VALUES (?, ?, ?, datetime('now'))
                """,
                (ids["indent_id"], orphan_supplier_id, "PO/ORPHAN/TEST/1"),
            )
            conn.commit()
            queue = self.stores_mod._load_generated_purchase_orders(
                conn, "restaurant", send_queue=True
            )
            self.assertFalse(
                any(str(row.get("po_no")) == "PO/ORPHAN/TEST/1" for row in queue)
            )
        finally:
            conn.close()

        # Unsent PO whose indent is fully inwarded must leave the Send queue.
        # Create a second indent reusing products already seeded by the helper.
        conn = db_mod.get_db()
        try:
            supplier_id = int(ids["supplier_id"])
            user_id = int(
                conn.execute("SELECT id FROM users ORDER BY id LIMIT 1").fetchone()["id"]
            )
            cur = conn.execute(
                """
                INSERT INTO store_indents
                    (outlet, indent_no, status, notes, created_by, created_at, submitted_at, decided_at, decided_by)
                VALUES (
                    'restaurant', 'IND/RES/26-27/TEST-INWARD', 'approved',
                    'PO send queue inwarded', ?, datetime('now','localtime'),
                    datetime('now','localtime'), datetime('now','localtime'), ?
                )
                """,
                (user_id, user_id),
            )
            indent2_id = int(cur.lastrowid)
            conn.execute(
                """
                INSERT INTO store_indent_lines
                    (indent_id, item_name, quantity, unit, approximate_price, quantity_received)
                VALUES (?, 'PO Rice Bag', 10, 'kg', 50, 0)
                """,
                (indent2_id,),
            )
            line2 = conn.execute(
                "SELECT id FROM store_indent_lines WHERE indent_id = ? ORDER BY id DESC LIMIT 1",
                (indent2_id,),
            ).fetchone()
            line2_id = int(line2["id"])
            conn.execute(
                """
                INSERT INTO store_po_lines (indent_id, line_id, supplier_id, rate, quantity, updated_at)
                VALUES (?, ?, ?, 50, 10, datetime('now','localtime'))
                """,
                (indent2_id, line2_id, supplier_id),
            )
            conn.commit()
        finally:
            conn.close()

        gen2 = self.client.post(
            f"/stores/orders/{indent2_id}/lines/next",
            data={
                "outlet": "restaurant",
                f"line_supplier_{line2_id}": str(supplier_id),
                f"line_rate_{line2_id}": "50",
                f"line_qty_{line2_id}": "10",
                "selectable_supplier": [str(supplier_id)],
                "selected_supplier": [str(supplier_id)],
            },
            follow_redirects=False,
        )
        self.assertEqual(gen2.status_code, 302)

        conn = db_mod.get_db()
        try:
            po2 = conn.execute(
                "SELECT po_no FROM store_purchase_orders WHERE indent_id = ?",
                (indent2_id,),
            ).fetchone()
            self.assertIsNotNone(po2)
            po2_no = str(po2["po_no"])
            before = self.stores_mod._load_generated_purchase_orders(
                conn, "restaurant", send_queue=True
            )
            self.assertTrue(any(str(row.get("po_no")) == po2_no for row in before))
            conn.execute(
                "UPDATE store_indent_lines SET quantity_received = quantity WHERE indent_id = ?",
                (indent2_id,),
            )
            conn.commit()
            after = self.stores_mod._load_generated_purchase_orders(
                conn, "restaurant", send_queue=True
            )
        finally:
            conn.close()
        self.assertFalse(any(str(row.get("po_no")) == po2_no for row in after))

    def test_stock_inward_removes_po_from_send_queue(self):
        """Any stock inward on a PO treats it as sent — it leaves Send to Supplier."""
        ids = self._create_approved_indent_for_po(notes="PO inward clears send queue")
        indent_id = int(ids["indent_id"])
        supplier_id = int(ids["supplier_id"])
        rice_line = int(ids["lines"]["PO Rice Bag"])

        gen = self.client.post(
            f"/stores/orders/{indent_id}/lines/next",
            data={
                "outlet": "restaurant",
                f"line_supplier_{rice_line}": str(supplier_id),
                f"line_rate_{rice_line}": "50",
                f"line_qty_{rice_line}": "10",
                "selectable_supplier": [str(supplier_id)],
                "selected_supplier": [str(supplier_id)],
            },
            follow_redirects=False,
        )
        self.assertEqual(gen.status_code, 302)

        conn = db_mod.get_db()
        try:
            po = conn.execute(
                "SELECT id, po_no FROM store_purchase_orders WHERE indent_id = ? AND supplier_id = ?",
                (indent_id, supplier_id),
            ).fetchone()
            self.assertIsNotNone(po)
            po_id = int(po["id"])
            po_no = str(po["po_no"])
            before = self.stores_mod._load_generated_purchase_orders(
                conn, "restaurant", send_queue=True
            )
            self.assertTrue(any(str(row.get("po_no")) == po_no for row in before))

            # Partial stock inward on the PO (indent still has remaining).
            conn.execute(
                """
                UPDATE store_purchase_order_lines
                SET quantity_received = 1
                WHERE purchase_order_id = ?
                """,
                (po_id,),
            )
            conn.execute(
                """
                UPDATE store_indent_lines
                SET quantity_received = 1
                WHERE id = ?
                """,
                (rice_line,),
            )
            conn.commit()

            after = self.stores_mod._load_generated_purchase_orders(
                conn, "restaurant", send_queue=True
            )
            self.assertFalse(any(str(row.get("po_no")) == po_no for row in after))

            # Still eligible for further stock inward while qty remains.
            pending = self.stores_mod._load_generated_purchase_orders(
                conn, "restaurant", pending_inward=True
            )
            self.assertTrue(any(str(row.get("po_no")) == po_no for row in pending))
        finally:
            conn.close()

    def test_stock_inward_po_picker_hides_fully_received_pos(self):
        """Purchase Order dropdown only lists POs that still have pending inward qty."""
        ids = self._create_approved_indent_for_po(notes="PO inward picker filter")
        indent_id = int(ids["indent_id"])
        supplier_id = int(ids["supplier_id"])
        rice_line = int(ids["lines"]["PO Rice Bag"])

        gen = self.client.post(
            f"/stores/orders/{indent_id}/lines/next",
            data={
                "outlet": "restaurant",
                f"line_supplier_{rice_line}": str(supplier_id),
                f"line_rate_{rice_line}": "50",
                f"line_qty_{rice_line}": "10",
                "selectable_supplier": [str(supplier_id)],
                "selected_supplier": [str(supplier_id)],
            },
            follow_redirects=False,
        )
        self.assertEqual(gen.status_code, 302)

        conn = db_mod.get_db()
        try:
            po = conn.execute(
                "SELECT id, po_no FROM store_purchase_orders WHERE indent_id = ? AND supplier_id = ?",
                (indent_id, supplier_id),
            ).fetchone()
            self.assertIsNotNone(po)
            po_no = str(po["po_no"])
            before = self.stores_mod._load_generated_purchase_orders(
                conn, "restaurant", pending_inward=True
            )
            self.assertTrue(any(str(row.get("po_no")) == po_no for row in before))

            page_before = self.client.get(
                "/stores/purchase-requests?outlet=restaurant&view=approved"
            )
            self.assertEqual(page_before.status_code, 200)
            self.assertIn(po_no.encode(), page_before.data)
            # Unsent generated POs must appear in the Stock Inward PO dropdown.
            row = next(r for r in before if str(r.get("po_no")) == po_no)
            self.assertNotEqual(str(row.get("status") or "").lower(), "sent")

            # Supplier chip must not hide other suppliers' generated POs from the list.
            other = self.client.get(
                "/stores/purchase-requests?outlet=restaurant&view=approved&supplier_id=999999"
            )
            self.assertEqual(other.status_code, 200)
            self.assertIn(po_no.encode(), other.data)

            conn.execute(
                "UPDATE store_indent_lines SET quantity_received = quantity WHERE indent_id = ?",
                (indent_id,),
            )
            conn.execute(
                """
                UPDATE store_purchase_order_lines
                SET quantity_received = quantity
                WHERE purchase_order_id = ?
                """,
                (int(po["id"]),),
            )
            conn.commit()

            after = self.stores_mod._load_generated_purchase_orders(
                conn, "restaurant", pending_inward=True
            )
        finally:
            conn.close()

        self.assertFalse(any(str(row.get("po_no")) == po_no for row in after))
        page_after = self.client.get(
            "/stores/purchase-requests?outlet=restaurant&view=approved"
        )
        self.assertEqual(page_after.status_code, 200)
        self.assertNotIn(po_no.encode(), page_after.data)

    def test_stock_inward_hides_po_when_po_qty_received_but_indent_open(self):
        """A fully received PO must leave Stock Inward even if indent qty remains."""
        ids = self._create_approved_indent_for_po(notes="PO partial inward remaining indent")
        indent_id = int(ids["indent_id"])
        supplier_id = int(ids["supplier_id"])
        rice_line = int(ids["lines"]["PO Rice Bag"])

        # Indent line qty is 10; generate a PO for only 3.
        gen = self.client.post(
            f"/stores/orders/{indent_id}/lines/next",
            data={
                "outlet": "restaurant",
                f"line_supplier_{rice_line}": str(supplier_id),
                f"line_rate_{rice_line}": "50",
                f"line_qty_{rice_line}": "3",
                "selectable_supplier": [str(supplier_id)],
                "selected_supplier": [str(supplier_id)],
            },
            follow_redirects=False,
        )
        self.assertEqual(gen.status_code, 302)

        conn = db_mod.get_db()
        try:
            po = conn.execute(
                "SELECT id, po_no FROM store_purchase_orders WHERE indent_id = ? AND supplier_id = ?",
                (indent_id, supplier_id),
            ).fetchone()
            self.assertIsNotNone(po)
            po_id = int(po["id"])
            po_no = str(po["po_no"])

            # Simulate inward of the full PO qty while indent still has remaining.
            conn.execute(
                "UPDATE store_indent_lines SET quantity_received = 3 WHERE id = ?",
                (rice_line,),
            )
            conn.execute(
                """
                UPDATE store_purchase_order_lines
                SET quantity_received = quantity
                WHERE purchase_order_id = ?
                """,
                (po_id,),
            )
            conn.commit()

            pending = self.stores_mod._load_generated_purchase_orders(
                conn, "restaurant", pending_inward=True
            )
            self.assertFalse(any(str(row.get("po_no")) == po_no for row in pending))

            po_row = {
                "po_id": po_id,
                "id": po_id,
                "indent_id": indent_id,
                "supplier_id": supplier_id,
                "po_no": po_no,
                "supplier_name": "Test",
            }
            _indent, lines = self.stores_mod._build_inward_lines_for_po(
                conn, po_row, outlet="restaurant"
            )
            self.assertEqual(lines, [])
        finally:
            conn.close()

        page = self.client.get(
            f"/stores/purchase-requests?outlet=restaurant&view=approved&po_id={po_id}"
        )
        self.assertEqual(page.status_code, 200)
        self.assertNotIn(po_no.encode(), page.data)
        self.assertNotIn(b"Purchase Order Items", page.data)

    def test_po_orders_tab_splits_pending_and_completed_inward(self):
        """Purchase Orders tab shows pending vs completed stock-inward blocks."""
        ids = self._create_approved_indent_for_po(notes="PO orders inward blocks")
        indent_id = int(ids["indent_id"])
        supplier_id = int(ids["supplier_id"])
        rice_line = int(ids["lines"]["PO Rice Bag"])

        gen = self.client.post(
            f"/stores/orders/{indent_id}/lines/next",
            data={
                "outlet": "restaurant",
                f"line_supplier_{rice_line}": str(supplier_id),
                f"line_rate_{rice_line}": "50",
                f"line_qty_{rice_line}": "10",
                "selectable_supplier": [str(supplier_id)],
                "selected_supplier": [str(supplier_id)],
            },
            follow_redirects=False,
        )
        self.assertEqual(gen.status_code, 302)

        conn = db_mod.get_db()
        try:
            po = conn.execute(
                "SELECT id, po_no FROM store_purchase_orders WHERE indent_id = ? AND supplier_id = ?",
                (indent_id, supplier_id),
            ).fetchone()
            self.assertIsNotNone(po)
            po_id = int(po["id"])
            po_no = str(po["po_no"])

            pending = self.stores_mod._load_generated_purchase_orders(
                conn, "restaurant", pending_inward=True
            )
            completed = self.stores_mod._load_generated_purchase_orders(
                conn, "restaurant", completed_inward=True
            )
            self.assertTrue(any(str(row.get("po_no")) == po_no for row in pending))
            self.assertFalse(any(str(row.get("po_no")) == po_no for row in completed))
        finally:
            conn.close()

        before = self.client.get("/stores/orders?outlet=restaurant")
        self.assertEqual(before.status_code, 200)
        self.assertIn(b"Stock Inward - Pending", before.data)
        self.assertIn(b"Stock Inward - Completed", before.data)
        self.assertIn(b'data-st-po-inward-block="pending"', before.data)
        self.assertIn(b'data-st-po-inward-block="completed"', before.data)
        self.assertIn(b'id="st-po-pending-table"', before.data)
        pending_html = before.data.split(b'data-st-po-inward-block="pending"', 1)[1].split(
            b'data-st-po-inward-block="completed"', 1
        )[0]
        completed_html = before.data.split(b'data-st-po-inward-block="completed"', 1)[1]
        self.assertIn(po_no.encode(), pending_html)
        self.assertNotIn(po_no.encode(), completed_html)
        self.assertIn(f"po_id={po_id}".encode(), pending_html)
        self.assertIn(b'data-tip="Stock Inward"', pending_html)
        self.assertIn(b"No completed stock inward POs", completed_html)

        conn = db_mod.get_db()
        try:
            conn.execute(
                "UPDATE store_indent_lines SET quantity_received = quantity WHERE indent_id = ?",
                (indent_id,),
            )
            conn.execute(
                """
                UPDATE store_purchase_order_lines
                SET quantity_received = quantity
                WHERE purchase_order_id = ?
                """,
                (po_id,),
            )
            conn.commit()
            pending_after = self.stores_mod._load_generated_purchase_orders(
                conn, "restaurant", pending_inward=True
            )
            completed_after = self.stores_mod._load_generated_purchase_orders(
                conn, "restaurant", completed_inward=True
            )
        finally:
            conn.close()

        self.assertFalse(any(str(row.get("po_no")) == po_no for row in pending_after))
        self.assertTrue(any(str(row.get("po_no")) == po_no for row in completed_after))

        after = self.client.get("/stores/orders?outlet=restaurant")
        self.assertEqual(after.status_code, 200)
        pending_after_html = after.data.split(b'data-st-po-inward-block="pending"', 1)[1].split(
            b'data-st-po-inward-block="completed"', 1
        )[0]
        completed_after_html = after.data.split(b'data-st-po-inward-block="completed"', 1)[1]
        self.assertNotIn(po_no.encode(), pending_after_html)
        self.assertIn(b'id="st-po-completed-table"', completed_after_html)
        self.assertIn(po_no.encode(), completed_after_html)
        self.assertIn(b"No POs pending stock inward", pending_after_html)


if __name__ == "__main__":
    unittest.main()
