"""Tests for Employee Add → Bulk Excel template + import."""

import io
import tempfile
import unittest
from unittest import mock

from openpyxl import Workbook, load_workbook

import db as db_mod
import employee_bulk as bulk
from employee_payroll import (
    _DEFAULT_COMPANY,
    _EPF_MAX,
    _PAYROLL_DEPARTMENTS,
    _emp_code_taken,
    _next_emp_code,
)


def _fill_employee_row(ws, row_idx, **fields):
    headers = list(bulk.UNIFIED_HEADERS)
    index = {h: i + 1 for i, h in enumerate(headers)}
    for key, value in fields.items():
        # allow keys without * suffix
        col = None
        for h, c in index.items():
            if h.replace(" *", "") == key or h == key:
                col = c
                break
        if col is None:
            raise KeyError(key)
        ws.cell(row_idx, col, value)


class EmployeeBulkUnitTests(unittest.TestCase):
    def test_template_fingerprint_and_departments(self):
        buf = bulk.build_employee_bulk_template(
            departments=list(_PAYROLL_DEPARTMENTS) + ["NEWDEPT"],
            default_company=_DEFAULT_COMPANY,
        )
        wb = load_workbook(buf)
        self.assertEqual(wb["Instructions"]["A1"].value, bulk.INSTRUCTIONS_TITLE)
        self.assertIn("Employees", wb.sheetnames)
        self.assertIn("Dropdowns", wb.sheetnames)
        headers = [wb["Employees"].cell(1, c).value for c in range(1, len(bulk.UNIFIED_HEADERS) + 1)]
        self.assertEqual(tuple(headers), bulk.UNIFIED_HEADERS)
        self.assertNotIn("EMP ID", headers)
        instructions = "\n".join(
            str(wb["Instructions"].cell(r, 1).value or "")
            for r in range(1, wb["Instructions"].max_row + 1)
        )
        self.assertIn("automatically", instructions.casefold())
        depts = []
        dd = wb["Dropdowns"]
        for r in range(2, dd.max_row + 1):
            val = dd.cell(r, 1).value
            if val:
                depts.append(val)
        self.assertIn("FO", depts)
        self.assertIn("NEWDEPT", depts)

    def test_assert_rejects_wrong_headers(self):
        wb = Workbook()
        wb.active.title = "Instructions"
        wb.active["A1"] = bulk.INSTRUCTIONS_TITLE
        wb.create_sheet("Dropdowns")
        ws = wb.create_sheet("Employees")
        ws.append(["Wrong", "Headers"])
        with self.assertRaises(ValueError):
            bulk._assert_matches_template(wb)

    def test_import_creates_and_skips(self):
        tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        tmp.close()
        orig = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = tmp.name
        try:
            db_mod.init_db()
            buf = bulk.build_employee_bulk_template(
                departments=list(_PAYROLL_DEPARTMENTS),
                default_company=_DEFAULT_COMPANY,
            )
            wb = load_workbook(buf)
            ws = wb["Employees"]
            _fill_employee_row(
                ws,
                2,
                **{
                    "Name *": "Anita Sharma",
                    "Mobile *": "9876543210",
                    "Department": "FO",
                    "Gross Salary *": 25000,
                    "Status": "active",
                },
            )
            _fill_employee_row(
                ws,
                3,
                **{
                    "Name *": "Bad Mobile",
                    "Mobile *": "123",
                    "Gross Salary *": 20000,
                },
            )
            _fill_employee_row(
                ws,
                4,
                **{
                    "Name *": "",
                    "Mobile *": "9876543211",
                    "Gross Salary *": 20000,
                },
            )
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)

            conn = db_mod.get_db()
            try:
                result = bulk.import_employee_bulk(
                    conn,
                    out,
                    departments=list(_PAYROLL_DEPARTMENTS),
                    next_emp_code_fn=_next_emp_code,
                    emp_code_taken_fn=_emp_code_taken,
                    default_company=_DEFAULT_COMPANY,
                    epf_max=_EPF_MAX,
                )
                conn.commit()
                row = conn.execute(
                    "SELECT name, mobile, location, gross_salary, emp_code FROM employees WHERE mobile=?",
                    ("9876543210",),
                ).fetchone()
            finally:
                conn.close()

            self.assertEqual(result["created_count"], 1)
            self.assertGreaterEqual(result["skipped_count"], 2)
            self.assertIsNotNone(row)
            self.assertEqual(row["name"], "Anita Sharma")
            self.assertEqual(row["location"], "FO")
            self.assertEqual(float(row["gross_salary"]), 25000.0)
            self.assertTrue(str(row["emp_code"] or "").upper().startswith("HBE"))
        finally:
            db_mod.DATABASE_PATH = orig


class EmployeeBulkRouteTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db_path = self.tmp.name
        self._orig_path = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = self.db_path
        db_mod.init_db()

        import app as app_mod

        self.app_mod = app_mod
        self.app = app_mod.app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

        conn = db_mod.get_db()
        try:
            admin = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            self.admin_id = admin["id"]
        finally:
            conn.close()

        self.user = {
            "id": self.admin_id,
            "username": "admin",
            "full_name": "Administrator",
            "is_admin": True,
            "is_active": True,
            "dashboard_access": set(),
            "stores_access": set(),
            "payroll_access": {"employee", "attendance", "payroll"},
        }
        self._get_user_patch = mock.patch.object(
            app_mod, "get_current_user", return_value=self.user
        )
        self._get_user_patch.start()

    def tearDown(self):
        self._get_user_patch.stop()
        db_mod.DATABASE_PATH = self._orig_path
        try:
            import os

            os.unlink(self.db_path)
        except OSError:
            pass

    def test_download_template_route(self):
        resp = self.client.get("/download_employee_template")
        self.assertEqual(resp.status_code, 200)
        self.assertIn(
            "spreadsheetml",
            (resp.headers.get("Content-Type") or "").lower(),
        )
        wb = load_workbook(io.BytesIO(resp.data))
        self.assertEqual(wb["Instructions"]["A1"].value, bulk.INSTRUCTIONS_TITLE)

    def test_upload_json_import(self):
        buf = bulk.build_employee_bulk_template(
            departments=list(_PAYROLL_DEPARTMENTS),
            default_company=_DEFAULT_COMPANY,
        )
        wb = load_workbook(buf)
        _fill_employee_row(
            wb["Employees"],
            2,
            **{
                "Name *": "Ravi Kumar",
                "Mobile *": "9123456780",
                "Department": "HK",
                "Gross Salary *": 30000,
            },
        )
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        resp = self.client.post(
            "/upload_employees",
            data={"file": (out, "employees.xlsx")},
            content_type="multipart/form-data",
            headers={
                "Accept": "application/json",
                "X-Requested-With": "XMLHttpRequest",
            },
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.get_json()
        self.assertTrue(data["ok"])
        self.assertEqual(data["created_count"], 1)

    def test_add_employee_page_has_bulk_mode(self):
        resp = self.client.get("/add_employee?from=employee_master")
        self.assertEqual(resp.status_code, 200)
        html = resp.get_data(as_text=True)
        self.assertIn("ep-emp-mode-host", html)
        self.assertIn("data-ep-emp-mode=\"bulk\"", html)
        self.assertIn("ep-emp-bulk-template", html)


if __name__ == "__main__":
    unittest.main()
