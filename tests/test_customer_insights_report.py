"""Customer Insights report — per-mobile aggregation, page, export."""

import json
import os
import tempfile
import unittest
from unittest import mock

import db as db_mod


class CustomerInsightsReportTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db_path = self.tmp.name
        self._orig_path = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = self.db_path
        db_mod.init_db()

        import app as app_mod

        self.app_mod = app_mod
        self.app = app_mod.app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

        conn = db_mod.get_db()
        try:
            admin = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            self.admin_id = admin["id"]
            db_mod.ensure_pos_schema(conn)
            cat = db_mod.save_pos_menu_category(
                conn, name="Mains", outlet=db_mod.POS_OUTLET_RESTAURANT
            )
            bar_cat = db_mod.save_pos_menu_category(
                conn, name="Spirits", outlet=db_mod.POS_OUTLET_BAR
            )
            self.butter = db_mod.save_pos_menu_item(
                conn,
                category_id=cat["id"],
                name="Chicken Butter Masala",
                rate=320,
                outlet=db_mod.POS_OUTLET_RESTAURANT,
            )
            self.whisky = db_mod.save_pos_menu_item(
                conn,
                category_id=bar_cat["id"],
                name="Whisky Peg",
                rate=250,
                outlet=db_mod.POS_OUTLET_BAR,
                item_kind="liquor",
                menu_type="liquor",
            )
            conn.commit()
        finally:
            conn.close()

        self.user = {
            "id": self.admin_id,
            "username": "admin",
            "full_name": "Administrator",
            "is_admin": True,
            "is_active": True,
            "dashboard_access": set(),
            "stores_access": set(),
        }
        self._get_user_patch = mock.patch.object(
            app_mod, "get_current_user", return_value=self.user
        )
        self._get_user_patch.start()

    def tearDown(self):
        self._get_user_patch.stop()
        db_mod.DATABASE_PATH = self._orig_path
        try:
            os.unlink(self.db_path)
        except OSError:
            pass

    def _pos_invoice(
        self,
        *,
        outlet,
        order_no,
        menu_id,
        name,
        rate,
        qty,
        customer_name,
        customer_mobile,
    ):
        line_total = round(rate * qty, 2)
        payload = {
            "outlet": outlet,
            "orderNo": order_no,
            "savedAt": "2026-08-01 18:00:00",
            "orderType": "dine_in",
            "table": "T1" if outlet == "restaurant" else "B1",
            "captain": "",
            "customerName": customer_name,
            "customerMobile": customer_mobile,
            "notes": "",
            "discountType": "pct",
            "discountValue": 0,
            "serviceType": "pct",
            "serviceValue": 0,
            "tipAmount": 0,
            "couponCode": "",
            "lines": [
                {
                    "uid": "1",
                    "menuId": menu_id,
                    "name": name,
                    "variant": "",
                    "rate": rate,
                    "qty": qty,
                }
            ],
            "totals": {
                "subtotal": line_total,
                "discount": 0,
                "discountType": "pct",
                "discountValue": 0,
                "gst": 0,
                "service": 0,
                "serviceType": "pct",
                "serviceValue": 0,
                "tip": 0,
                "roundOff": 0,
                "total": line_total,
            },
        }
        conn = db_mod.get_db()
        try:
            inv = db_mod.save_pos_invoice(conn, payload)
            conn.commit()
            return inv
        finally:
            conn.close()

    def _hotel_invoice(self, *, invoice_number, guest_name, mobile, total, status="settled"):
        conn = db_mod.get_db()
        try:
            db_mod.ensure_hotel_room_invoices_schema(conn)
            payload = {
                "id": "r101",
                "number": "101",
                "stay": {
                    "guestName": guest_name,
                    "mobile": mobile,
                    "invoiceNumber": invoice_number,
                },
            }
            conn.execute(
                """
                INSERT INTO hotel_room_invoices (
                    invoice_number, room_id, room_number, room_type_label,
                    guest_name, booking_number, check_in_date, check_out_date,
                    invoice_generated_at, estimated_total, advance_paid,
                    balance_amount, status, payload_json
                ) VALUES (?, 'r101', '101', 'Deluxe', ?, '', '2026-08-01', '2026-08-02',
                          '2026-08-02 10:00:00', ?, 0, 0, ?, ?)
                """,
                (
                    invoice_number,
                    guest_name,
                    float(total),
                    status,
                    json.dumps(payload),
                ),
            )
            conn.commit()
        finally:
            conn.close()

    def test_list_customer_insights_merges_pos_and_hotel(self):
        self._pos_invoice(
            outlet="restaurant",
            order_no="ORD-CI-01",
            menu_id=self.butter["id"],
            name="Chicken Butter Masala",
            rate=320,
            qty=2,
            customer_name="Rajesh",
            customer_mobile="9876543210",
        )
        self._pos_invoice(
            outlet="bar",
            order_no="ORD-CI-02",
            menu_id=self.whisky["id"],
            name="Whisky Peg",
            rate=250,
            qty=1,
            customer_name="Rajesh",
            customer_mobile="9876543210",
        )
        self._hotel_invoice(
            invoice_number="HBE/RM/1/2026-27",
            guest_name="Rajesh Kumar",
            mobile="9876543210",
            total=4500,
        )

        conn = db_mod.get_db()
        try:
            rows = db_mod.list_customer_insights(
                conn,
                date_from="2026-08-01",
                date_to="2026-08-08",
                channel="all",
            )
            kpis = db_mod.customer_insights_kpis(rows)
        finally:
            conn.close()

        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["mobile"], "9876543210")
        self.assertEqual(row["order_count"], 3)
        self.assertEqual(row["top_item"], "Chicken Butter Masala")
        self.assertEqual(row["restaurant_value"], 640.0)
        self.assertEqual(row["bar_value"], 250.0)
        self.assertEqual(row["hotel_value"], 4500.0)
        self.assertEqual(row["total_value"], 5390.0)
        self.assertEqual(kpis["customer_count"], 1)
        self.assertEqual(kpis["total_value_sum"], 5390.0)

    def test_named_customers_without_mobile_are_included(self):
        self._pos_invoice(
            outlet="restaurant",
            order_no="ORD-CI-NM-01",
            menu_id=self.butter["id"],
            name="Chicken Butter Masala",
            rate=320,
            qty=1,
            customer_name="Mr. Baborsekh",
            customer_mobile="",
        )
        self._pos_invoice(
            outlet="bar",
            order_no="ORD-CI-NM-02",
            menu_id=self.whisky["id"],
            name="Whisky Peg",
            rate=250,
            qty=2,
            customer_name="Mr. Baborsekh",
            customer_mobile="",
        )
        self._pos_invoice(
            outlet="restaurant",
            order_no="ORD-CI-NM-03",
            menu_id=self.butter["id"],
            name="Chicken Butter Masala",
            rate=320,
            qty=1,
            customer_name="Guest",
            customer_mobile="",
        )
        conn = db_mod.get_db()
        try:
            rows = db_mod.list_customer_insights(
                conn,
                date_from="2026-08-01",
                date_to="2026-08-08",
                channel="all",
            )
        finally:
            conn.close()
        by_name = {row["customer_name"]: row for row in rows}
        self.assertIn("Mr. Baborsekh", by_name)
        named = by_name["Mr. Baborsekh"]
        self.assertEqual(named["mobile"], "")
        self.assertEqual(named["order_count"], 2)
        self.assertEqual(named["restaurant_value"], 320.0)
        self.assertEqual(named["bar_value"], 500.0)
        self.assertEqual(named["top_item"], "Whisky Peg")
        self.assertIn("Guest", by_name)
        self.assertEqual(by_name["Guest"]["order_count"], 1)

    def test_channel_filter_hotel_only(self):
        self._pos_invoice(
            outlet="restaurant",
            order_no="ORD-CI-03",
            menu_id=self.butter["id"],
            name="Chicken Butter Masala",
            rate=320,
            qty=1,
            customer_name="Asha",
            customer_mobile="9123456780",
        )
        self._hotel_invoice(
            invoice_number="HBE/RM/2/2026-27",
            guest_name="Asha",
            mobile="9123456780",
            total=2000,
        )
        conn = db_mod.get_db()
        try:
            rows = db_mod.list_customer_insights(
                conn,
                date_from="2026-08-01",
                date_to="2026-08-08",
                channel="hotel",
            )
        finally:
            conn.close()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["hotel_value"], 2000.0)
        self.assertEqual(rows[0]["restaurant_value"], 0.0)
        self.assertEqual(rows[0]["top_item"], "")

    def test_page_and_export(self):
        from datetime import date, datetime

        today = date.today()
        # Invoice must fall in the default FY window used when hub opens with no dates.
        conn = db_mod.get_db()
        try:
            payload = {
                "outlet": "restaurant",
                "orderNo": "ORD-CI-04",
                "savedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "orderDate": today.isoformat(),
                "orderType": "dine_in",
                "table": "T1",
                "captain": "",
                "customerName": "Meera",
                "customerMobile": "9000011111",
                "notes": "",
                "discountType": "pct",
                "discountValue": 0,
                "serviceType": "pct",
                "serviceValue": 0,
                "tipAmount": 0,
                "couponCode": "",
                "lines": [
                    {
                        "uid": "1",
                        "menuId": self.butter["id"],
                        "name": "Chicken Butter Masala",
                        "variant": "",
                        "rate": 320,
                        "qty": 1,
                    }
                ],
                "totals": {
                    "subtotal": 320,
                    "discount": 0,
                    "discountType": "pct",
                    "discountValue": 0,
                    "gst": 0,
                    "service": 0,
                    "serviceType": "pct",
                    "serviceValue": 0,
                    "tip": 0,
                    "roundOff": 0,
                    "total": 320,
                },
            }
            db_mod.save_pos_invoice(conn, payload)
            payload["orderNo"] = "ORD-CI-05"
            payload["customerName"] = "Mr. Baborsekh"
            payload["customerMobile"] = ""
            db_mod.save_pos_invoice(conn, payload)
            conn.commit()
        finally:
            conn.close()

        page = self.client.get("/reports/sales/customer-insights")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn("Customer Insights", html)
        self.assertIn("Unique customers", html)
        self.assertIn("Meera", html)
        self.assertIn("9000011111", html)
        self.assertIn("Mr. Baborsekh", html)
        fy_start_year = today.year if today.month >= 4 else today.year - 1
        self.assertIn(f'value="{date(fy_start_year, 4, 1).isoformat()}"', html)
        self.assertIn(f'value="{today.isoformat()}"', html)

        export = self.client.get("/reports/sales/customer-insights/export")
        self.assertEqual(export.status_code, 200)
        cd = export.headers.get("Content-Disposition") or ""
        fy_start, fy_today = db_mod.indian_fiscal_year_bounds()
        from reports import report_export_filename

        expected_name = report_export_filename(
            "Customer Insights",
            date_from=fy_start,
            date_to=fy_today,
            date_filter_active=True,
        )
        self.assertIn(expected_name, cd)

        from io import BytesIO
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(export.data))
        self.assertEqual(
            wb.sheetnames, ["Summary", "Hotel", "Restaurant", "Bar"]
        )
        ws = wb["Summary"]
        self.assertEqual(ws["A1"].value, "Hotel Bell Elite — Customer Insights")
        self.assertEqual(ws["A1"].fill.fgColor.rgb, "FF315A78")
        self.assertEqual(
            [ws.cell(2, col).value for col in range(1, 9)],
            [
                "Customer",
                "Mobile",
                "Orders",
                "Top Item",
                "Restaurant",
                "Bar",
                "Hotel",
                "Total",
            ],
        )
        summary_names = {
            ws.cell(row, 1).value
            for row in range(3, (ws.max_row or 2) + 1)
            if ws.cell(row, 1).value
        }
        self.assertIn("Meera", summary_names)
        self.assertIn("Mr. Baborsekh", summary_names)

        self.assertEqual(
            wb["Hotel"]["A1"].value,
            "Hotel Bell Elite — Customer Insights - Hotel",
        )
        self.assertEqual(
            wb["Restaurant"]["A1"].value,
            "Hotel Bell Elite — Customer Insights - Restaurant",
        )
        self.assertEqual(
            wb["Bar"]["A1"].value,
            "Hotel Bell Elite — Customer Insights - Bar",
        )
        for sheet_name in ("Hotel", "Restaurant", "Bar"):
            sheet = wb[sheet_name]
            self.assertEqual(sheet["A1"].fill.fgColor.rgb, "FF315A78")
            self.assertEqual(sheet["A2"].value, "Customer")

        restaurant = wb["Restaurant"]
        restaurant_names = {
            restaurant.cell(row, 1).value
            for row in range(3, (restaurant.max_row or 2) + 1)
            if restaurant.cell(row, 1).value
        }
        self.assertIn("Meera", restaurant_names)
        self.assertIn("Mr. Baborsekh", restaurant_names)

        bar = wb["Bar"]
        bar_names = {
            bar.cell(row, 1).value
            for row in range(3, (bar.max_row or 2) + 1)
            if bar.cell(row, 1).value
        }
        # ORD-CI-05 has empty mobile + name Mr. Baborsekh but restaurant outlet only
        # in setUp data for page test — no bar invoice in this test's FY seed.
        self.assertEqual(bar_names, set())
        self.assertEqual(
            {
                wb["Hotel"].cell(row, 1).value
                for row in range(3, (wb["Hotel"].max_row or 2) + 1)
                if wb["Hotel"].cell(row, 1).value
            },
            set(),
        )

    def test_export_outlet_sheets_filter_by_channel(self):
        """Hotel / Restaurant / Bar sheets contain only that outlet's customers."""
        self._pos_invoice(
            outlet="restaurant",
            order_no="ORD-CI-OUT-R",
            menu_id=self.butter["id"],
            name="Chicken Butter Masala",
            rate=320,
            qty=1,
            customer_name="Rest Only",
            customer_mobile="9111000001",
        )
        self._pos_invoice(
            outlet="bar",
            order_no="ORD-CI-OUT-B",
            menu_id=self.whisky["id"],
            name="Whisky Peg",
            rate=250,
            qty=1,
            customer_name="Bar Only",
            customer_mobile="9111000002",
        )
        self._hotel_invoice(
            invoice_number="HBE/RM/OUT/2026-27",
            guest_name="Hotel Only",
            mobile="9111000003",
            total=1800,
        )
        self._pos_invoice(
            outlet="restaurant",
            order_no="ORD-CI-OUT-BOTH-R",
            menu_id=self.butter["id"],
            name="Chicken Butter Masala",
            rate=320,
            qty=1,
            customer_name="Multi Channel",
            customer_mobile="9111000004",
        )
        self._pos_invoice(
            outlet="bar",
            order_no="ORD-CI-OUT-BOTH-B",
            menu_id=self.whisky["id"],
            name="Whisky Peg",
            rate=250,
            qty=2,
            customer_name="Multi Channel",
            customer_mobile="9111000004",
        )

        from io import BytesIO
        from openpyxl import load_workbook

        export = self.client.get(
            "/reports/sales/customer-insights/export"
            "?date_from=2026-08-01&date_to=2026-08-08"
        )
        self.assertEqual(export.status_code, 200)
        wb = load_workbook(BytesIO(export.data))
        self.assertEqual(
            wb.sheetnames, ["Summary", "Hotel", "Restaurant", "Bar"]
        )

        def _names(sheet):
            return {
                sheet.cell(row, 1).value
                for row in range(3, (sheet.max_row or 2) + 1)
                if sheet.cell(row, 1).value
            }

        summary_names = _names(wb["Summary"])
        self.assertEqual(
            summary_names,
            {"Rest Only", "Bar Only", "Hotel Only", "Multi Channel"},
        )
        self.assertEqual(_names(wb["Hotel"]), {"Hotel Only"})
        self.assertEqual(
            _names(wb["Restaurant"]), {"Rest Only", "Multi Channel"}
        )
        self.assertEqual(
            _names(wb["Bar"]), {"Bar Only", "Multi Channel"}
        )

        hotel = wb["Hotel"]
        self.assertEqual(hotel["A1"].value, "Hotel Bell Elite — Customer Insights - Hotel")
        hotel_row = next(
            row
            for row in range(3, (hotel.max_row or 2) + 1)
            if hotel.cell(row, 1).value == "Hotel Only"
        )
        self.assertEqual(float(hotel.cell(hotel_row, 7).value), 1800.0)
        self.assertEqual(float(hotel.cell(hotel_row, 5).value or 0), 0.0)
        self.assertEqual(float(hotel.cell(hotel_row, 6).value or 0), 0.0)

        restaurant = wb["Restaurant"]
        rest_row = next(
            row
            for row in range(3, (restaurant.max_row or 2) + 1)
            if restaurant.cell(row, 1).value == "Rest Only"
        )
        self.assertEqual(float(restaurant.cell(rest_row, 5).value), 320.0)
        self.assertEqual(restaurant.cell(rest_row, 4).value, "Chicken Butter Masala")


if __name__ == "__main__":
    unittest.main()
