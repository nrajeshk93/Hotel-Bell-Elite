"""Hotel Rooms floor board seed and API."""

import json
import os
import tempfile
import unittest
from datetime import datetime, timedelta
from unittest import mock

import db as db_mod


class HotelRoomsTests(unittest.TestCase):
    @staticmethod
    def _stay_window(nights=2, end_offset_days=1):
        """Return check-in / check-out ISO dates that avoid overstay billing."""
        nights = max(1, int(nights or 1))
        check_out = datetime.now().date() + timedelta(days=max(1, int(end_offset_days or 1)))
        check_in = check_out - timedelta(days=nights)
        return check_in.isoformat(), check_out.isoformat()

    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db_path = self.tmp.name
        self._orig_path = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = self.db_path
        db_mod.init_db()

        import app as app_mod

        self.app_mod = app_mod
        self.app = app_mod.app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

        conn = db_mod.get_db()
        try:
            admin = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            self.admin_id = admin["id"]
        finally:
            conn.close()

        self.user = {
            "id": self.admin_id,
            "username": "admin",
            "full_name": "Administrator",
            "is_admin": True,
            "is_active": True,
            "dashboard_access": set(),
            "stores_access": set(),
        }
        self._get_user_patch = mock.patch.object(app_mod, "get_current_user", return_value=self.user)
        self._get_user_patch.start()

    def tearDown(self):
        self._get_user_patch.stop()
        db_mod.DATABASE_PATH = self._orig_path
        try:
            os.unlink(self.db_path)
        except OSError:
            pass

    def test_default_seed_inventory(self):
        layout = db_mod.default_hotel_rooms_layout()
        self.assertEqual(len(layout["floors"]), 3)
        self.assertEqual(len(layout["rooms"]), 20)
        by_type = {}
        for room in layout["rooms"]:
            by_type[room["roomType"]] = by_type.get(room["roomType"], 0) + 1
            self.assertEqual(room["status"], "vacant")
            self.assertEqual(room["floorId"], f"floor-{room['number'][0]}")
        self.assertEqual(
            by_type,
            {
                "premium_without_balcony": 6,
                "premium_deluxe_balcony": 12,
                "premium_suite_tub": 2,
            },
        )

    def test_get_layout_seeds_when_empty(self):
        conn = db_mod.get_db()
        try:
            layout = db_mod.get_hotel_rooms_layout(conn)
            counts = db_mod.hotel_rooms_status_counts(layout)
            self.assertEqual(counts["total"], 20)
            self.assertEqual(counts["vacant"], 20)
            updated = db_mod.update_hotel_room_status(conn, "room-207", "dirty")
            room = next(r for r in updated["rooms"] if r["id"] == "room-207")
            self.assertEqual(room["status"], "dirty")
            conn.commit()
        finally:
            conn.close()

    def test_rooms_api_get_and_status_put(self):
        get_resp = self.client.get("/hotel/api/rooms")
        self.assertEqual(get_resp.status_code, 200)
        payload = get_resp.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(len(payload["rooms"]), 20)
        self.assertEqual(payload["counts"]["total"], 20)

        put_resp = self.client.put(
            "/hotel/api/rooms",
            json={"roomId": "room-101", "status": "occupied"},
        )
        self.assertEqual(put_resp.status_code, 200)
        put_payload = put_resp.get_json()
        self.assertTrue(put_payload["ok"])
        room = next(r for r in put_payload["rooms"] if r["id"] == "room-101")
        self.assertEqual(room["status"], "occupied")
        self.assertEqual(put_payload["counts"]["occupied"], 1)

    def test_rooms_api_reserve_binds_stay_to_as_of_date(self):
        put_resp = self.client.put(
            "/hotel/api/rooms",
            json={
                "roomId": "room-104",
                "status": "reserved",
                "checkInDate": "2026-07-31",
                "checkOutDate": "2026-08-01",
            },
        )
        self.assertEqual(put_resp.status_code, 200)
        payload = put_resp.get_json()
        self.assertTrue(payload["ok"])
        room = next(r for r in payload["rooms"] if r["id"] == "room-104")
        self.assertEqual(room["status"], "reserved")
        self.assertEqual(room["stay"]["checkInDate"], "2026-07-31")
        self.assertEqual(room["stay"]["checkOutDate"], "2026-08-01")
        # Reservation without an explicit rate seeds the room-type tariff.
        self.assertGreater(float(room["stay"]["roomRate"]), 0)
        self.assertEqual(payload["counts"]["reserved"], 1)

    def test_rooms_page_renders(self):
        resp = self.client.get("/hotel/rooms")
        self.assertEqual(resp.status_code, 200)
        html = resp.get_data(as_text=True)
        self.assertIn("hotel-rooms-page", html)
        self.assertIn("Rooms", html)
        self.assertIn("hotel-rooms-status-listbox", html)
        self.assertIn("hotel-rooms-quick-reservations", html)
        self.assertIn("Reservations", html)
        self.assertIn("hr-board-reserve-modal", html)
        self.assertIn("hr-board-reserve-form", html)
        self.assertIn("hr-board-reserve-rooms-select", html)
        self.assertIn("Save Reservation", html)
        self.assertIn('data-customers-api="/hotel/api/customers"', html)
        self.assertIn("hotel-rooms-date-filter", html)
        self.assertIn('data-kpi="expected_checkout"', html)
        self.assertIn("Expected Check Out", html)
        self.assertNotIn('data-kpi="reserved"', html)

    def test_board_multi_room_reserve_via_per_room_api(self):
        """Board Reserve loops single-room reserve; same dates work on free rooms."""
        stay = {
            "guestName": "Board Guest",
            "mobile": "9876543210",
            "email": "board@example.com",
            "additionalRequests": "Late arrival after 10 PM",
        }
        dates = {
            "action": "reserve",
            "checkInDate": "2026-09-10",
            "checkOutDate": "2026-09-12",
            "stay": stay,
        }
        room_a = self.client.put("/hotel/api/rooms/room-201", json=dates)
        self.assertEqual(room_a.status_code, 200)
        self.assertTrue(room_a.get_json()["ok"])
        self.assertEqual(room_a.get_json()["room"]["status"], "reserved")
        self.assertEqual(room_a.get_json()["room"]["stay"]["checkInDate"], "2026-09-10")
        self.assertGreater(float(room_a.get_json()["room"]["stay"]["roomRate"]), 0)
        self.assertEqual(
            room_a.get_json()["room"]["stay"]["additionalRequests"],
            "Late arrival after 10 PM",
        )

        room_b = self.client.put("/hotel/api/rooms/room-202", json=dates)
        self.assertEqual(room_b.status_code, 200)
        self.assertTrue(room_b.get_json()["ok"])
        self.assertEqual(room_b.get_json()["room"]["status"], "reserved")
        self.assertEqual(
            room_b.get_json()["room"]["stay"]["guestName"], "Board Guest"
        )

        occupied = self.client.put(
            "/hotel/api/rooms/room-203",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "In",
                    "lastName": "House",
                    "mobile": "9000099999",
                    "checkInDate": "2026-07-28",
                    "nights": 2,
                    "roomRate": 4000,
                },
            },
        )
        self.assertEqual(occupied.status_code, 200)
        self.assertEqual(occupied.get_json()["room"]["status"], "occupied")

        # Future non-overlapping dates queue as upcomingStay while still occupied.
        future = self.client.put(
            "/hotel/api/rooms/room-203",
            json={
                "action": "reserve",
                "checkInDate": "2026-09-10",
                "checkOutDate": "2026-09-12",
                "stay": stay,
            },
        )
        self.assertEqual(future.status_code, 200)
        future_room = future.get_json()["room"]
        self.assertTrue(future.get_json()["ok"])
        self.assertEqual(future_room["status"], "occupied")
        self.assertEqual(future_room["stay"]["guestName"], "In House")
        self.assertEqual(future_room["upcomingStay"]["guestName"], "Board Guest")
        self.assertEqual(future_room["upcomingStay"]["checkInDate"], "2026-09-10")

        overlap_occupied = self.client.put(
            "/hotel/api/rooms/room-203",
            json={
                "action": "reserve",
                "checkInDate": "2026-07-29",
                "checkOutDate": "2026-07-31",
                "stay": {
                    "guestName": "Overlap Occ",
                    "mobile": "9111222333",
                },
            },
        )
        self.assertEqual(overlap_occupied.status_code, 400)
        self.assertFalse(overlap_occupied.get_json()["ok"])

        overlap_replace = self.client.put(
            "/hotel/api/rooms/room-201",
            json={
                "action": "reserve",
                "replace": True,
                "checkInDate": "2026-09-11",
                "checkOutDate": "2026-09-13",
                "stay": {
                    "guestName": "Overlap Guest",
                    "mobile": "9123456780",
                },
            },
        )
        self.assertEqual(overlap_replace.status_code, 400)
        self.assertFalse(overlap_replace.get_json()["ok"])
        self.assertIn("already reserved", overlap_replace.get_json()["error"].lower())

    def test_occupied_future_reserve_and_checkout_promotes_upcoming(self):
        """Occupied rooms accept non-overlapping future reserve; checkout promotes it."""
        checkin = self.client.put(
            "/hotel/api/rooms/room-204",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Today",
                    "lastName": "Guest",
                    "mobile": "9000088888",
                    "checkInDate": "2026-07-28",
                    "nights": 2,
                    "roomRate": 3500,
                },
            },
        )
        self.assertEqual(checkin.status_code, 200)
        self.assertEqual(checkin.get_json()["room"]["status"], "occupied")
        self.assertEqual(
            checkin.get_json()["room"]["stay"]["checkOutDate"], "2026-07-30"
        )
        self._generate_stay_invoice("room-204")

        future = self.client.put(
            "/hotel/api/rooms/room-204",
            json={
                "action": "reserve",
                "checkInDate": "2026-08-20",
                "checkOutDate": "2026-08-22",
                "stay": {
                    "guestName": "Next Guest",
                    "mobile": "9888777666",
                    "email": "next@example.com",
                },
            },
        )
        self.assertEqual(future.status_code, 200)
        room = future.get_json()["room"]
        self.assertEqual(room["status"], "occupied")
        self.assertEqual(room["stay"]["guestName"], "Today Guest")
        self.assertEqual(room["upcomingStay"]["guestName"], "Next Guest")
        self.assertEqual(room["upcomingStay"]["checkInDate"], "2026-08-20")
        self.assertEqual(room["upcomingStay"]["checkOutDate"], "2026-08-22")

        # Overlapping the in-house window still fails.
        overlap = self.client.put(
            "/hotel/api/rooms/room-204",
            json={
                "action": "reserve",
                "checkInDate": "2026-07-29",
                "checkOutDate": "2026-07-31",
                "stay": {"guestName": "Clash", "mobile": "9777666555"},
            },
        )
        self.assertEqual(overlap.status_code, 400)
        self.assertIn("occupied", overlap.get_json()["error"].lower())

        checkout = self.client.put(
            "/hotel/api/rooms/room-204",
            json={"action": "checkout"},
        )
        self.assertEqual(checkout.status_code, 200)
        after = checkout.get_json()["room"]
        self.assertEqual(after["status"], "reserved")
        self.assertEqual(after["stay"]["guestName"], "Next Guest")
        self.assertEqual(after["stay"]["checkInDate"], "2026-08-20")
        self.assertNotIn("upcomingStay", after)

    def test_room_detail_page_renders(self):
        resp = self.client.get("/hotel/rooms/room-101")
        self.assertEqual(resp.status_code, 200)
        html = resp.get_data(as_text=True)
        self.assertIn("hotel-room-detail-page", html)
        self.assertIn("Room 101", html)
        self.assertIn("Start Check-In", html)
        self.assertIn("hrd-add-guest-id", html)
        self.assertIn("Add Guest ID", html)
        self.assertIn('data-action="add-guest-id"', html)
        self.assertIn("hrd-ci-id-proof", html)
        self.assertIn("hrd-id-doc-name", html)
        self.assertIn("hrd-id-preview-modal", html)
        self.assertIn("hrd-checkout-invoice-modal", html)
        self.assertIn("Generate Invoice to check out", html)
        self.assertNotIn("New Check-in", html)
        self.assertIn("hrd-reserve", html)
        self.assertIn("hrd-reserve-new", html)
        self.assertIn("hrd-reserve-modal", html)
        self.assertIn("hrd-form-listbox--countries", html)
        box_at = html.find('id="hrd-ci-nationality-listbox"')
        self.assertGreater(box_at, 0)
        nat_at = html.find('id="hrd-ci-nationality-list"')
        self.assertGreater(nat_at, box_at)
        trigger_chunk = html[box_at:nat_at]
        self.assertIn('role="combobox"', trigger_chunk)
        self.assertIn('id="hrd-ci-nationality-trigger"', trigger_chunk)
        self.assertNotIn("ep-listbox-search", trigger_chunk)
        nat_chunk = html[nat_at:nat_at + 120000]
        self.assertNotIn("ep-listbox-search", nat_chunk.split('<div class="ep-listbox-options"', 1)[0])
        self.assertIn("Afghanistan", nat_chunk)
        self.assertIn("Indonesia", nat_chunk)
        self.assertGreater(nat_chunk.count("se-filter-listbox-option"), 180)

    def test_reserved_future_room_detail_renders(self):
        reserved = self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "reserve",
                "checkInDate": "2026-08-14",
                "checkOutDate": "2026-08-15",
                "stay": {
                    "guestName": "Pratik",
                    "mobile": "8587836993",
                    "email": "pratikmmt@gmail.com",
                },
            },
        )
        self.assertEqual(reserved.status_code, 200, reserved.get_data(as_text=True))
        self.assertEqual(reserved.get_json()["room"]["status"], "reserved")

        detail = self.client.get("/hotel/rooms/room-102")
        self.assertEqual(detail.status_code, 200, detail.get_data(as_text=True))
        html = detail.get_data(as_text=True)
        self.assertIn("hotel-room-detail-page", html)
        self.assertIn("Room 102", html)
        self.assertIn("hrd-start-checkin-reserved", html)
        self.assertIn("Edit Reservation", html)

        with_date = self.client.get("/hotel/rooms/room-102?date=2026-08-14")
        self.assertEqual(with_date.status_code, 200)
        self.assertIn("hotel-room-detail-page", with_date.get_data(as_text=True))

    def test_room_detail_reserve_and_occupied_reject(self):
        reserved = self.client.put(
            "/hotel/api/rooms/room-105",
            json={
                "action": "reserve",
                "checkInDate": "2026-08-06",
                "checkOutDate": "2026-08-08",
                "stay": {
                    "guestName": "Priya Shah",
                    "mobile": "9887766554",
                    "email": "priya@example.com",
                    "agencyName": "Travel Co",
                    "agencyGst": "27AAAAA0000A1Z5",
                    "agencyAddress": "MG Road",
                    "agencyBilling": True,
                },
            },
        )
        self.assertEqual(reserved.status_code, 200)
        payload = reserved.get_json()
        self.assertTrue(payload["ok"])
        stay = payload["room"]["stay"]
        self.assertEqual(payload["room"]["status"], "reserved")
        self.assertEqual(stay["checkInDate"], "2026-08-06")
        self.assertEqual(stay["checkOutDate"], "2026-08-08")
        self.assertEqual(stay["nights"], 2)
        self.assertEqual(stay["guestName"], "Priya Shah")
        self.assertEqual(stay["firstName"], "Priya")
        self.assertEqual(stay["lastName"], "Shah")
        self.assertEqual(stay["mobile"], "9887766554")
        self.assertEqual(stay["email"], "priya@example.com")
        self.assertEqual(stay["agencyName"], "Travel Co")
        self.assertEqual(stay["agencyGst"], "27AAAAA0000A1Z5")
        self.assertTrue(stay["agencyBilling"])
        self.assertTrue(stay["agencyRoomBilling"])
        self.assertTrue(stay["agencyFbBilling"])
        self.assertEqual(stay["invoiceTo"], "Travel Co")

        missing = self.client.put(
            "/hotel/api/rooms/room-104",
            json={
                "action": "reserve",
                "checkInDate": "2026-08-10",
                "checkOutDate": "2026-08-12",
                "stay": {"guestName": "No Phone"},
            },
        )
        self.assertEqual(missing.status_code, 400)
        self.assertFalse(missing.get_json()["ok"])
        self.assertIn("Mobile", missing.get_json()["error"])

        missing_name = self.client.put(
            "/hotel/api/rooms/room-104",
            json={
                "action": "reserve",
                "checkInDate": "2026-08-10",
                "checkOutDate": "2026-08-12",
                "stay": {"mobile": "9000000001"},
            },
        )
        self.assertEqual(missing_name.status_code, 400)
        self.assertFalse(missing_name.get_json()["ok"])
        self.assertIn("Guest name", missing_name.get_json()["error"])

        checkin = self.client.put(
            "/hotel/api/rooms/room-106",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Sam",
                    "lastName": "Lee",
                    "mobile": "9000012345",
                    "checkInDate": "2026-07-30",
                    "nights": 1,
                    "roomRate": 4500,
                },
            },
        )
        self.assertEqual(checkin.status_code, 200)
        self.assertEqual(checkin.get_json()["room"]["status"], "occupied")

        # Overlapping the in-house stay is still rejected.
        blocked = self.client.put(
            "/hotel/api/rooms/room-106",
            json={
                "action": "reserve",
                "checkInDate": "2026-07-30",
                "checkOutDate": "2026-08-01",
                "stay": {"guestName": "Blocked Guest", "mobile": "9111111111"},
            },
        )
        self.assertEqual(blocked.status_code, 400)
        self.assertFalse(blocked.get_json()["ok"])
        self.assertIn("occupied", blocked.get_json()["error"].lower())

        # Non-overlapping future dates queue as upcomingStay.
        future = self.client.put(
            "/hotel/api/rooms/room-106",
            json={
                "action": "reserve",
                "checkInDate": "2026-08-10",
                "checkOutDate": "2026-08-12",
                "stay": {"guestName": "Future Guest", "mobile": "9111111111"},
            },
        )
        self.assertEqual(future.status_code, 200)
        self.assertEqual(future.get_json()["room"]["status"], "occupied")
        self.assertEqual(
            future.get_json()["room"]["upcomingStay"]["guestName"], "Future Guest"
        )

    def test_room_detail_reserve_replace_clears_prior_guest(self):
        first = self.client.put(
            "/hotel/api/rooms/room-105",
            json={
                "action": "reserve",
                "checkInDate": "2026-08-06",
                "checkOutDate": "2026-08-08",
                "stay": {
                    "guestName": "Priya Shah",
                    "mobile": "9887766554",
                    "email": "priya@example.com",
                    "agencyName": "Travel Co",
                    "agencyGst": "27AAAAA0000A1Z5",
                    "agencyAddress": "MG Road",
                    "agencyBilling": True,
                },
            },
        )
        self.assertEqual(first.status_code, 200)
        self.assertEqual(first.get_json()["room"]["stay"]["guestName"], "Priya Shah")

        overlap = self.client.put(
            "/hotel/api/rooms/room-105",
            json={
                "action": "reserve",
                "replace": True,
                "checkInDate": "2026-08-07",
                "checkOutDate": "2026-08-09",
                "stay": {
                    "guestName": "Overlap Guest",
                    "mobile": "9000099999",
                },
            },
        )
        self.assertEqual(overlap.status_code, 400)
        self.assertFalse(overlap.get_json()["ok"])
        self.assertIn("already reserved", overlap.get_json()["error"].lower())

        replaced = self.client.put(
            "/hotel/api/rooms/room-105",
            json={
                "action": "reserve",
                "replace": True,
                "checkInDate": "2026-08-10",
                "checkOutDate": "2026-08-12",
                "stay": {
                    "guestName": "Asha Rao",
                    "mobile": "9000012345",
                    "email": "asha@example.com",
                },
            },
        )
        self.assertEqual(replaced.status_code, 200)
        payload = replaced.get_json()
        self.assertTrue(payload["ok"])
        stay = payload["room"]["stay"]
        self.assertEqual(payload["room"]["status"], "reserved")
        self.assertEqual(stay["guestName"], "Asha Rao")
        self.assertEqual(stay["mobile"], "9000012345")
        self.assertEqual(stay["email"], "asha@example.com")
        self.assertEqual(stay["checkInDate"], "2026-08-10")
        self.assertEqual(stay["checkOutDate"], "2026-08-12")
        self.assertEqual(stay.get("agencyName") or "", "")
        self.assertEqual(stay.get("agencyGst") or "", "")
        self.assertFalse(stay.get("agencyBilling"))

    def test_agency_room_and_fb_billing_flags(self):
        check_in, check_out = self._stay_window(nights=1)
        room_only = self.client.put(
            "/hotel/api/rooms/room-103",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Lukas",
                    "lastName": "Wong",
                    "mobile": "9000000227",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 2000,
                    "agencyName": "ATPI India Pvt. Ltd",
                    "agencyRoomBilling": True,
                    "agencyFbBilling": False,
                },
            },
        )
        self.assertEqual(room_only.status_code, 200, room_only.get_data(as_text=True))
        stay = room_only.get_json()["room"]["stay"]
        self.assertTrue(stay["agencyRoomBilling"])
        self.assertFalse(stay["agencyFbBilling"])
        self.assertTrue(stay["agencyBilling"])
        self.assertEqual(stay["invoiceTo"], "ATPI India Pvt. Ltd")
        self.assertTrue(db_mod._hotel_stay_bills_room_to_agency(stay))
        self.assertFalse(db_mod._hotel_stay_bills_fb_to_agency(stay))

        fb_only = self.client.put(
            "/hotel/api/rooms/room-104",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Guest",
                    "lastName": "Two",
                    "mobile": "9000000228",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 2000,
                    "agencyName": "ATPI India Pvt. Ltd",
                    "agencyRoomBilling": False,
                    "agencyFbBilling": True,
                },
            },
        )
        self.assertEqual(fb_only.status_code, 200, fb_only.get_data(as_text=True))
        stay = fb_only.get_json()["room"]["stay"]
        self.assertFalse(stay["agencyRoomBilling"])
        self.assertTrue(stay["agencyFbBilling"])
        self.assertTrue(stay["agencyBilling"])
        self.assertEqual(stay.get("invoiceTo") or "", "")
        self.assertFalse(db_mod._hotel_stay_bills_room_to_agency(stay))
        self.assertTrue(db_mod._hotel_stay_bills_fb_to_agency(stay))

        missing = self.client.put(
            "/hotel/api/rooms/room-105",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "No",
                    "lastName": "Agency",
                    "mobile": "9000000229",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 2000,
                    "agencyRoomBilling": True,
                },
            },
        )
        self.assertEqual(missing.status_code, 400)
        self.assertIn("agency", (missing.get_json() or {}).get("error", "").lower())

        missing_fb = self.client.put(
            "/hotel/api/rooms/room-105",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "No",
                    "lastName": "AgencyFb",
                    "mobile": "9000000230",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 2000,
                    "agencyFbBilling": True,
                },
            },
        )
        self.assertEqual(missing_fb.status_code, 400)

        flags = db_mod._hotel_stay_agency_bill_flags({"agencyBilling": True})
        self.assertEqual(flags, (True, True))
        flags = db_mod._hotel_stay_agency_bill_flags(
            {"agencyRoomBilling": True, "agencyFbBilling": False}
        )
        self.assertEqual(flags, (True, False))

    def test_room_detail_api_get_and_mark_clean(self):
        put_dirty = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"status": "dirty"},
        )
        self.assertEqual(put_dirty.status_code, 200)
        self.assertEqual(put_dirty.get_json()["room"]["status"], "dirty")

        put_clean = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"status": "vacant"},
        )
        self.assertEqual(put_clean.status_code, 200)
        cleaned = put_clean.get_json()["room"]
        self.assertEqual(cleaned["status"], "vacant")
        self.assertEqual(cleaned.get("statusLabel"), "Vacant")

    def test_occupied_status_cannot_become_vacant_while_checked_in(self):
        """Checked-in rooms stay Occupied until checkout — no Vacant shortcut."""
        check_in, check_out = self._stay_window(nights=3)
        self.client.put(
            "/hotel/api/rooms/room-103",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "In",
                    "lastName": "House",
                    "mobile": "9000000103",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 3,
                    "roomRate": 3000,
                },
            },
        )
        blocked_vacant = self.client.put(
            "/hotel/api/rooms/room-103",
            json={"status": "vacant"},
        )
        self.assertEqual(blocked_vacant.status_code, 400)
        self.assertIn("checked in", blocked_vacant.get_json().get("error", "").lower())

        still = self.client.get("/hotel/api/rooms/room-103").get_json()["room"]
        self.assertEqual(still["status"], "occupied")
        self.assertEqual(still.get("statusLabel"), "Occupied")
        self.assertTrue(still.get("stay", {}).get("guestName") or still.get("stay", {}).get("firstName"))

        # Dirty while occupied is allowed — checks guest out and marks Dirty.
        self._generate_stay_invoice("room-103")
        forced_dirty = self.client.put(
            "/hotel/api/rooms/room-103",
            json={"status": "dirty"},
        )
        self.assertEqual(forced_dirty.status_code, 200, forced_dirty.get_data(as_text=True))
        dirty_room = forced_dirty.get_json()["room"]
        self.assertEqual(dirty_room["status"], "dirty")
        self.assertFalse(dirty_room.get("stay"))

        cleaned = self.client.put(
            "/hotel/api/rooms/room-103",
            json={"status": "vacant"},
        )
        self.assertEqual(cleaned.status_code, 200)
        self.assertEqual(cleaned.get_json()["room"]["status"], "vacant")

    def test_checkout_sets_room_dirty(self):
        check_in, check_out = self._stay_window(nights=1)
        self.client.put(
            "/hotel/api/rooms/room-104",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Out",
                    "lastName": "Guest",
                    "mobile": "9000000104",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 2000,
                },
            },
        )
        self._generate_stay_invoice("room-104")
        closed = self.client.put(
            "/hotel/api/rooms/room-104",
            json={"action": "checkout"},
        )
        self.assertEqual(closed.status_code, 200)
        room = closed.get_json()["room"]
        self.assertEqual(room["status"], "dirty")
        self.assertEqual(room.get("statusLabel"), "Dirty")
        self.assertFalse(room.get("stay"))

    def test_checkout_allows_legacy_invoice_without_night_snapshots(self):
        """Invoiced stays missing hotelInvoiced* must still check out."""
        self._checkin_with_charges("room-104", advance=0)
        self._generate_stay_invoice("room-104")
        # Force legacy shape: clear snapshots + leave an untagged folio line.
        conn = db_mod.get_db()
        try:
            layout = db_mod.get_hotel_rooms_layout(conn)
            rooms = layout.get("rooms") or []
            room = next(r for r in rooms if r.get("id") == "room-104")
            stay = dict(room.get("stay") or {})
            stay["hotelInvoicedBillableNights"] = 0
            stay["hotelInvoicedEstimatedTotal"] = 0
            stay["hotelInvoicedExtraBedAmount"] = 0
            stay["hotelInvoicedEarlyCheckinAmount"] = 0
            stay["hotelInvoicedLateCheckoutAmount"] = 0
            folio = list(stay.get("folioCharges") or [])
            folio.append(
                {
                    "id": "legacy-untagged",
                    "kind": "other",
                    "label": "Room 102 — stay charges",
                    "amount": 2000,
                    "source": "room_merge",
                    "sourceRoomId": "room-102",
                }
            )
            stay["folioCharges"] = folio
            room["stay"] = db_mod._normalize_hotel_room_stay(stay)
            db_mod.save_hotel_rooms_layout(conn, layout.get("floors") or [], rooms)
            conn.commit()
        finally:
            conn.close()

        closed = self.client.put(
            "/hotel/api/rooms/room-104",
            json={"action": "checkout"},
        )
        self.assertEqual(closed.status_code, 200, closed.get_data(as_text=True))
        self.assertEqual(closed.get_json()["room"]["status"], "dirty")

    def test_checkout_rejected_without_invoice(self):
        self._checkin_with_charges("room-104", advance=0)
        before = self.client.get("/hotel/api/rooms/room-104").get_json()["room"]
        self.assertFalse(before["stay"].get("invoiceGenerated"))
        self.assertFalse(before["stay"].get("invoiceNumber"))

        closed = self.client.put(
            "/hotel/api/rooms/room-104",
            json={"action": "checkout"},
        )
        self.assertEqual(closed.status_code, 400, closed.get_data(as_text=True))
        self.assertIn(
            "generate invoice to check out",
            closed.get_json().get("error", "").lower(),
        )
        still = self.client.get("/hotel/api/rooms/room-104").get_json()["room"]
        self.assertEqual(still["status"], "occupied")
        self.assertTrue(still.get("stay"))
        self.assertFalse(still["stay"].get("invoiceGenerated"))

    def test_checkout_member_does_not_require_own_invoice(self):
        self._checkin_with_charges("room-101", advance=0)
        self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Asha",
                    "lastName": "Nair",
                    "mobile": "9000000011",
                    "checkInDate": datetime.now().date().isoformat(),
                    "nights": 1,
                    "roomRate": 2000,
                    "advancePaid": 0,
                },
            },
        )
        self.client.put(
            "/hotel/api/rooms/room-102",
            json={"action": "merge_rooms", "fromRoomId": "room-102", "toRoomId": "room-101"},
        )
        member = self.client.get("/hotel/api/rooms/room-102").get_json()["room"]
        self.assertTrue(member.get("isMergeMember"))
        self.assertFalse(member["stay"].get("invoiceGenerated"))

        closed = self.client.put(
            "/hotel/api/rooms/room-102",
            json={"action": "checkout"},
        )
        self.assertEqual(closed.status_code, 200, closed.get_data(as_text=True))
        self.assertEqual(closed.get_json()["room"]["status"], "dirty")
        primary = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        self.assertEqual(primary["status"], "occupied")
        self.assertFalse(primary["stay"].get("invoiceGenerated"))

    def test_expected_checkout_kpi_counts_departures_for_day(self):
        today = datetime.now().date()
        check_out_today = today.isoformat()
        check_in = (today - timedelta(days=2)).isoformat()
        future_out = (today + timedelta(days=2)).isoformat()

        self.client.put(
            "/hotel/api/rooms/room-105",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Due",
                    "lastName": "Today",
                    "mobile": "9000000105",
                    "checkInDate": check_in,
                    "checkOutDate": check_out_today,
                    "nights": 2,
                    "roomRate": 3000,
                },
            },
        )
        self.client.put(
            "/hotel/api/rooms/room-106",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Staying",
                    "lastName": "On",
                    "mobile": "9000000106",
                    "checkInDate": check_in,
                    "checkOutDate": future_out,
                    "nights": 4,
                    "roomRate": 3000,
                },
            },
        )

        layout = self.client.get("/hotel/api/rooms").get_json()
        self.assertEqual(layout["counts"]["expected_checkout"], 1)
        self.assertGreaterEqual(layout["counts"]["occupied"], 2)

        # Board date filter for a future day should count that day's departures.
        conn = db_mod.get_db()
        try:
            counts = db_mod.hotel_rooms_status_counts(
                {"rooms": layout["rooms"]}, as_of=future_out
            )
            self.assertEqual(counts["expected_checkout"], 1)
        finally:
            conn.close()

    def test_expected_checkout_counts_each_merged_occupied_room(self):
        """Expected checkout is physical rooms due that day, including merge members."""
        today = datetime.now().date()
        check_out_today = today.isoformat()
        check_in = (today - timedelta(days=1)).isoformat()
        stay = {
            "firstName": "Merge",
            "lastName": "Out",
            "mobile": "9000000201",
            "checkInDate": check_in,
            "checkOutDate": check_out_today,
            "nights": 1,
            "roomRate": 3000,
        }
        self.assertEqual(
            self.client.put(
                "/hotel/api/rooms/room-201",
                json={"action": "checkin", "stay": stay},
            ).status_code,
            200,
        )
        self.assertEqual(
            self.client.put(
                "/hotel/api/rooms/room-202",
                json={"action": "checkin", "stay": dict(stay, mobile="9000000202")},
            ).status_code,
            200,
        )
        merged = self.client.put(
            "/hotel/api/rooms/room-202",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-202",
                "toRoomId": "room-201",
            },
        )
        self.assertEqual(merged.status_code, 200, merged.get_data(as_text=True))
        layout = self.client.get("/hotel/api/rooms").get_json()
        self.assertEqual(layout["counts"]["expected_checkout"], 2)
        rooms = {r["id"]: r for r in layout.get("rooms") or []}
        self.assertEqual(rooms["room-201"]["status"], "occupied")
        self.assertEqual(rooms["room-202"]["status"], "occupied")
        self.assertTrue(rooms["room-201"].get("mergeGroupId"))
        self.assertEqual(
            rooms["room-201"].get("mergeGroupId"),
            rooms["room-202"].get("mergeGroupId"),
        )

    def test_heals_orphan_vacant_with_inhouse_stay(self):
        """Vacant inventory with an active stay is restored to Occupied on load."""
        import json

        conn = db_mod.get_db()
        try:
            layout = db_mod.get_hotel_rooms_layout(conn)
            rooms = layout.get("rooms") or []
            room = next(r for r in rooms if r.get("id") == "room-104")
            room["status"] = "vacant"
            room["stay"] = {
                "firstName": "Still",
                "lastName": "Here",
                "guestName": "Still Here",
                "mobile": "9000000104",
                "checkInDate": "2026-07-28",
                "checkOutDate": "2026-08-05",
                "nights": 8,
                "roomRate": 4000,
                "checkedInAt": "2026-07-28 14:00:00",
            }
            room.pop("mergeGroupId", None)
            room.pop("mergePrimary", None)
            blob = json.dumps(
                {"floors": layout.get("floors") or [], "rooms": rooms},
                separators=(",", ":"),
            )
            conn.execute(
                """
                INSERT INTO hotel_rooms_layout (id, payload, updated_at)
                VALUES (1, ?, datetime('now','localtime'))
                ON CONFLICT(id) DO UPDATE SET
                    payload = excluded.payload,
                    updated_at = excluded.updated_at
                """,
                (blob,),
            )
            conn.commit()
        finally:
            conn.close()

        healed = self.client.get("/hotel/api/rooms/room-104").get_json()["room"]
        self.assertEqual(healed["status"], "occupied")
        self.assertEqual(healed.get("statusLabel"), "Occupied")

    def test_mark_clean_blocked_when_merge_member_occupied(self):
        """Vacant billing primary stays vacant; only the occupied room is blocked."""
        self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Raj",
                    "lastName": "Kumar",
                    "mobile": "9000000101",
                    "checkInDate": "2026-07-29",
                    "nights": 2,
                    "roomRate": 2500,
                },
            },
        )
        merged = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-101",
                "toRoomId": "room-106",
            },
        )
        self.assertEqual(merged.status_code, 200, merged.get_data(as_text=True))
        primary = merged.get_json()["primaryRoom"]
        self.assertEqual(primary["id"], "room-106")
        self.assertEqual(primary["status"], "vacant")
        member = merged.get_json()["memberRoom"]
        self.assertEqual(member["id"], "room-101")
        self.assertEqual(member["status"], "occupied")

        allowed = self.client.put(
            "/hotel/api/rooms/room-106",
            json={"status": "vacant"},
        )
        self.assertEqual(allowed.status_code, 200, allowed.get_data(as_text=True))
        self.assertEqual(allowed.get_json()["room"]["status"], "vacant")

        blocked = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"status": "vacant"},
        )
        self.assertEqual(blocked.status_code, 400, blocked.get_data(as_text=True))
        err = blocked.get_json().get("error", "").lower()
        self.assertIn("checked in", err)

        still = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        self.assertEqual(still["status"], "occupied")

    def test_heals_vacant_merge_primary_when_member_occupied(self):
        """Vacant billing primary stays vacant when a member is occupied."""
        import json

        conn = db_mod.get_db()
        try:
            layout = db_mod.get_hotel_rooms_layout(conn)
            rooms = layout.get("rooms") or []
            primary = next(r for r in rooms if r.get("id") == "room-106")
            member = next(r for r in rooms if r.get("id") == "room-101")
            gid = "hmg_test_heal_occ"
            primary["mergeGroupId"] = gid
            primary["mergePrimary"] = True
            primary["status"] = "vacant"
            primary["stay"] = {
                "mergeRole": "primary",
                "firstName": "Host",
                "guestName": "Host",
            }
            member["mergeGroupId"] = gid
            member["mergePrimary"] = False
            member["status"] = "occupied"
            member["stay"] = {
                "firstName": "Guest",
                "lastName": "One",
                "guestName": "Guest One",
                "mobile": "9000000102",
                "checkInDate": "2026-07-29",
                "nights": 1,
                "roomRate": 2000,
                "checkedInAt": "2026-07-29 14:00:00",
                "mergeRole": "member",
                "billingRoomId": "room-106",
            }
            # Bypass save_hotel_rooms_layout heal so we can seed the stale row.
            blob = json.dumps(
                {"floors": layout.get("floors") or [], "rooms": rooms},
                separators=(",", ":"),
            )
            conn.execute(
                """
                INSERT INTO hotel_rooms_layout (id, payload, updated_at)
                VALUES (1, ?, datetime('now','localtime'))
                ON CONFLICT(id) DO UPDATE SET
                    payload = excluded.payload,
                    updated_at = excluded.updated_at
                """,
                (blob,),
            )
            conn.commit()
        finally:
            conn.close()

        healed = self.client.get("/hotel/api/rooms/room-106").get_json()["room"]
        self.assertEqual(healed["status"], "vacant")
        member_healed = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        self.assertEqual(member_healed["status"], "occupied")

    def test_room_detail_404(self):
        resp = self.client.get("/hotel/rooms/room-999")
        self.assertEqual(resp.status_code, 404)

    def test_room_checkin_and_checkout(self):
        check_in, check_out = self._stay_window(nights=2)
        checkin = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Raj",
                    "lastName": "Kumar",
                    "mobile": "9876543210",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 2,
                    "roomRate": 4500,
                    "idType": "Aadhaar",
                    "idNumber": "1234-5678",
                },
            },
        )
        self.assertEqual(checkin.status_code, 200)
        payload = checkin.get_json()
        self.assertTrue(payload["ok"])
        room = payload["room"]
        self.assertEqual(room["status"], "occupied")
        self.assertEqual(room["stay"]["firstName"], "Raj")
        self.assertEqual(room["stay"]["guestName"], "Raj Kumar")
        self.assertTrue(room["stay"]["bookingNumber"].startswith("BK"))
        self.assertEqual(room["stay"]["totalRate"], 9000.0)

        detail = self.client.get("/hotel/rooms/room-101")
        self.assertEqual(detail.status_code, 200)
        html = detail.get_data(as_text=True)
        self.assertIn("New Check-In", html)
        self.assertIn("hrd-checkin-modal", html)
        self.assertIn('id="hrd-checkin-form"', html)
        self.assertIn('autocomplete="off"', html)
        self.assertIn('id="hrd-ci-mobile"', html)
        self.assertIn(
            'id="hrd-ci-mobile" required autocomplete="off"',
            html.replace("  ", " "),
        )
        self.assertIn('name="firstName" required autocomplete="off"', html)
        self.assertIn('name="lastName" required autocomplete="off"', html)
        self.assertNotIn('name="firstName" required autocomplete="given-name"', html)
        self.assertIn('id="hrd-ci-adults"', html)
        self.assertIn('id="hrd-ci-adults" name="adults" value=""', html)
        self.assertIn("Adults *", html)
        self.assertIn('data-se-listbox-change="hrdCiMobileCountryChanged"', html)
        self.assertIn("Indonesia", html)

        self._generate_stay_invoice("room-101")
        checkout = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "checkout"},
        )
        self.assertEqual(checkout.status_code, 200)
        out = checkout.get_json()["room"]
        self.assertEqual(out["status"], "dirty")
        self.assertNotIn("stay", out)

    def test_checkin_rejected_when_room_is_dirty(self):
        check_in, check_out = self._stay_window(nights=1)
        self.assertEqual(
            self.client.put("/hotel/api/rooms/room-101", json={"status": "dirty"}).status_code,
            200,
        )
        blocked = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Raj",
                    "lastName": "Kumar",
                    "mobile": "9876543210",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 4500,
                },
            },
        )
        self.assertEqual(blocked.status_code, 400, blocked.get_data(as_text=True))
        payload = blocked.get_json()
        self.assertFalse(payload["ok"])
        self.assertIn("dirty", (payload.get("error") or "").lower())
        room = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        self.assertEqual(room["status"], "dirty")
        self.assertFalse(room.get("stay"))

        cleaned = self.client.put("/hotel/api/rooms/room-101", json={"status": "vacant"})
        self.assertEqual(cleaned.status_code, 200)
        allowed = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Raj",
                    "lastName": "Kumar",
                    "mobile": "9876543210",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 4500,
                },
            },
        )
        self.assertEqual(allowed.status_code, 200, allowed.get_data(as_text=True))
        self.assertEqual(allowed.get_json()["room"]["status"], "occupied")

    def test_room_checkin_requires_fields(self):
        resp = self.client.put(
            "/hotel/api/rooms/room-102",
            json={"action": "checkin", "stay": {"firstName": "Only"}},
        )
        self.assertEqual(resp.status_code, 400)
        self.assertFalse(resp.get_json()["ok"])

        missing_rate = self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Asha",
                    "lastName": "Nair",
                    "mobile": "9000000001",
                    "checkInDate": "2026-07-29",
                    "roomRate": 0,
                },
            },
        )
        self.assertEqual(missing_rate.status_code, 400)
        self.assertEqual(missing_rate.get_json()["error"], "Room rate is required.")

        missing_plan = self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Asha",
                    "lastName": "Nair",
                    "mobile": "9000000001",
                    "checkInDate": "2026-07-29",
                    "roomRate": 2500,
                    "ratePlan": "",
                },
            },
        )
        self.assertEqual(missing_plan.status_code, 400)
        self.assertEqual(missing_plan.get_json()["error"], "Meal plan is required.")

    def test_room_transfer_vacant_only(self):
        checkin = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Asha",
                    "lastName": "Nair",
                    "mobile": "9000000001",
                    "checkInDate": "2026-07-29",
                    "roomRate": 2500,
                },
            },
        )
        self.assertEqual(checkin.status_code, 200)

        # Occupy another room so it is not offered as vacant.
        self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Other",
                    "lastName": "Guest",
                    "mobile": "9000000002",
                    "checkInDate": "2026-07-29",
                    "roomRate": 2500,
                },
            },
        )

        bad = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "transfer", "toRoomId": "room-102"},
        )
        self.assertEqual(bad.status_code, 400)
        self.assertIn("vacant", bad.get_json()["error"].lower())

        ok = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "transfer", "toRoomId": "room-103", "note": "Upgrade"},
        )
        self.assertEqual(ok.status_code, 200)
        payload = ok.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["fromRoom"]["status"], "dirty")
        self.assertNotIn("stay", payload["fromRoom"])
        self.assertEqual(payload["toRoom"]["status"], "occupied")
        self.assertEqual(payload["toRoom"]["stay"]["firstName"], "Asha")
        self.assertEqual(payload["toRoom"]["stay"]["transferCount"], 1)
        self.assertEqual(payload["toRoom"]["stay"]["transferHistory"][0]["fromRoomNumber"], "101")
        self.assertEqual(payload["toRoom"]["stay"]["transferHistory"][0]["toRoomNumber"], "103")

        vacant_only = self.client.put(
            "/hotel/api/rooms/room-103",
            json={"action": "transfer", "toRoomId": "room-101"},
        )
        self.assertEqual(vacant_only.status_code, 400)

    def test_room_transfer_clears_stale_vacant_stay_shell(self):
        """Vacant destinations with a cancelled reservation shell must still accept transfer."""
        import json

        checkin = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Asha",
                    "lastName": "Nair",
                    "mobile": "9000000011",
                    "checkInDate": "2026-07-29",
                    "roomRate": 2500,
                },
            },
        )
        self.assertEqual(checkin.status_code, 200)

        conn = db_mod.get_db()
        try:
            layout = db_mod.get_hotel_rooms_layout(conn)
            rooms = layout.get("rooms") or []
            for room in rooms:
                if room.get("id") != "room-105":
                    continue
                room["status"] = "vacant"
                room["stay"] = {
                    "checkInDate": "2026-08-01",
                    "checkOutDate": "2026-08-02",
                    "firstName": "",
                    "lastName": "",
                    "guestName": "",
                }
                break
            # Bypass save heal so the cancelled shell is present in storage.
            blob = json.dumps(
                {"floors": layout.get("floors") or [], "rooms": rooms},
                separators=(",", ":"),
            )
            conn.execute(
                """
                INSERT INTO hotel_rooms_layout (id, payload, updated_at)
                VALUES (1, ?, datetime('now','localtime'))
                ON CONFLICT(id) DO UPDATE SET
                    payload = excluded.payload,
                    updated_at = excluded.updated_at
                """,
                (blob,),
            )
            conn.commit()
        finally:
            conn.close()

        moved = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "transfer", "toRoomId": "room-105"},
        )
        self.assertEqual(moved.status_code, 200, moved.get_data(as_text=True))
        payload = moved.get_json()
        self.assertEqual(payload["toRoom"]["id"], "room-105")
        self.assertEqual(payload["toRoom"]["status"], "occupied")
        self.assertEqual(payload["toRoom"]["stay"]["firstName"], "Asha")
        self.assertEqual(payload["fromRoom"]["status"], "dirty")
        self.assertNotIn("stay", payload["fromRoom"])

    def test_guest_lookup_by_mobile(self):
        miss = self.client.get("/hotel/api/guests/lookup?mobile=9111111111")
        self.assertEqual(miss.status_code, 200)
        self.assertFalse(miss.get_json()["found"])

        self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Priya",
                    "lastName": "Shah",
                    "mobile": "9876501234",
                    "email": "priya@example.com",
                    "checkInDate": "2026-07-29",
                    "city": "Mumbai",
                    "idType": "Aadhaar",
                    "idNumber": "9999",
                    "roomRate": 2500,
                },
            },
        )
        # Vacate so lookup still finds historical stay on the room layout.
        hit = self.client.get("/hotel/api/guests/lookup?mobile=9876501234")
        self.assertEqual(hit.status_code, 200)
        body = hit.get_json()
        self.assertTrue(body["found"])
        guest = body["guest"]
        self.assertEqual(guest["firstName"], "Priya")
        self.assertEqual(guest["lastName"], "Shah")
        self.assertEqual(guest["email"], "priya@example.com")
        self.assertEqual(guest["city"], "Mumbai")
        self.assertEqual(guest["returningGuest"], "Yes")

        # After checkout, stay is cleared — full profile still comes from guest profiles.
        self._generate_stay_invoice("room-101")
        self.client.put("/hotel/api/rooms/room-101", json={"action": "checkout"})
        after = self.client.get("/hotel/api/guests/lookup?mobile=9876501234")
        self.assertEqual(after.status_code, 200)
        after_body = after.get_json()
        self.assertTrue(after_body["found"])
        self.assertEqual(after_body["guest"]["firstName"], "Priya")
        self.assertEqual(after_body["guest"]["lastName"], "Shah")
        self.assertEqual(after_body["guest"]["email"], "priya@example.com")
        self.assertEqual(after_body["guest"]["city"], "Mumbai")
        self.assertEqual(after_body["guest"]["returningGuest"], "Yes")
        self.assertTrue(after_body.get("nameMatch", True))

    def test_hotel_guest_names_match(self):
        guest = {"firstName": "Priya", "lastName": "Shah"}
        self.assertTrue(db_mod.hotel_guest_names_match(guest, "", ""))
        self.assertTrue(db_mod.hotel_guest_names_match(guest, "priya", "SHAH"))
        self.assertTrue(db_mod.hotel_guest_names_match(guest, "Mr. Priya", "Shah"))
        self.assertFalse(db_mod.hotel_guest_names_match(guest, "Amit", "Khan"))
        self.assertFalse(db_mod.hotel_guest_names_match(guest, "Priya", "Khan"))

    def test_guest_lookup_returns_id_after_checkout(self):
        doc_path = "/hotel/api/id-documents/aa111111-2222-3333-4444-555555555555.pdf"
        extra_path = "/hotel/api/id-documents/bb111111-2222-3333-4444-555555555555.pdf"
        healed_doc = "/hotel/api/id-documents/view/aa111111-2222-3333-4444-555555555555.pdf/raw"
        healed_extra = "/hotel/api/id-documents/view/bb111111-2222-3333-4444-555555555555.pdf/raw"
        check_in, check_out = self._stay_window(nights=1)
        checkin = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Priya",
                    "lastName": "Shah",
                    "mobile": "9876501234",
                    "address": "12 Marine Drive",
                    "city": "Mumbai",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "roomRate": 2500,
                    "idType": "Aadhaar",
                    "idDocumentName": "Priya Shah Aadhaar.pdf",
                    "idDocumentPath": doc_path,
                    "idDocumentMime": "application/pdf",
                    "idDocumentStoredName": "aa111111-2222-3333-4444-555555555555.pdf",
                    "additionalGuests": [
                        {
                            "name": "Amit Shah",
                            "idType": "Passport",
                            "idDocumentName": "Amit Shah Passport.pdf",
                            "idDocumentPath": extra_path,
                            "idDocumentMime": "application/pdf",
                        }
                    ],
                },
            },
        )
        self.assertEqual(checkin.status_code, 200, checkin.get_data(as_text=True))
        self._generate_stay_invoice("room-101")
        self.client.put("/hotel/api/rooms/room-101", json={"action": "checkout"})

        hit = self.client.get("/hotel/api/guests/lookup?mobile=9876501234")
        self.assertEqual(hit.status_code, 200)
        body = hit.get_json()
        self.assertTrue(body["found"])
        self.assertTrue(body["nameMatch"])
        guest = body["guest"]
        self.assertEqual(guest["firstName"], "Priya")
        self.assertEqual(guest["idType"], "Aadhaar")
        self.assertEqual(guest["idDocumentPath"], healed_doc)
        self.assertEqual(guest["address"], "12 Marine Drive")
        extras = guest.get("additionalGuests") or []
        self.assertEqual(len(extras), 1)
        self.assertEqual(extras[0]["name"], "Amit Shah")
        self.assertEqual(extras[0]["idDocumentPath"], healed_extra)

        mismatch = self.client.get(
            "/hotel/api/guests/lookup?mobile=9876501234&firstName=Amit&lastName=Khan"
        )
        mismatch_body = mismatch.get_json()
        self.assertTrue(mismatch_body["found"])
        self.assertFalse(mismatch_body["nameMatch"])
        self.assertEqual(mismatch_body["guest"]["firstName"], "Priya")
        self.assertEqual(mismatch_body["guest"]["idDocumentPath"], healed_doc)

        titled = self.client.get(
            "/hotel/api/guests/lookup?mobile=9876501234&firstName=Mr%20Priya&lastName=Shah"
        )
        self.assertTrue(titled.get_json()["nameMatch"])

    def test_guest_profile_merge_keeps_id_without_reupload(self):
        doc_path = "/hotel/api/id-documents/cc111111-2222-3333-4444-555555555555.pdf"
        healed_doc = "/hotel/api/id-documents/view/cc111111-2222-3333-4444-555555555555.pdf/raw"
        check_in, check_out = self._stay_window(nights=1)
        first = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Arun",
                    "lastName": "Shetty",
                    "mobile": "9000011122",
                    "address": "9 MG Road",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "roomRate": 2500,
                    "idType": "Aadhaar",
                    "idDocumentName": "Arun Shetty Aadhaar.pdf",
                    "idDocumentPath": doc_path,
                    "idDocumentMime": "application/pdf",
                    "idDocumentStoredName": "cc111111-2222-3333-4444-555555555555.pdf",
                },
            },
        )
        self.assertEqual(first.status_code, 200, first.get_data(as_text=True))
        self._generate_stay_invoice("room-101")
        self.client.put("/hotel/api/rooms/room-101", json={"action": "checkout"})
        self.client.put("/hotel/api/rooms/room-101", json={"status": "vacant"})

        second = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Arun",
                    "lastName": "Shetty",
                    "mobile": "9000011122",
                    "address": "9 MG Road",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "roomRate": 2500,
                    "idType": "Aadhaar",
                    "idDocumentPath": "",
                    "idDocumentName": "",
                    "idDocumentMime": "",
                    "idDocumentStoredName": "",
                },
            },
        )
        self.assertEqual(second.status_code, 200, second.get_data(as_text=True))
        self.assertEqual(second.get_json()["room"]["stay"].get("idDocumentPath") or "", "")
        self._generate_stay_invoice("room-101")
        self.client.put("/hotel/api/rooms/room-101", json={"action": "checkout"})

        after = self.client.get("/hotel/api/guests/lookup?mobile=9000011122")
        self.assertEqual(after.status_code, 200)
        guest = after.get_json()["guest"]
        self.assertEqual(guest["idDocumentPath"], healed_doc)
        self.assertEqual(guest["idType"], "Aadhaar")
        self.assertEqual(guest["idDocumentName"], "Arun Shetty Aadhaar.pdf")

    def test_checkin_estimated_charges_and_folio_moves_on_transfer(self):
        check_in, check_out = self._stay_window(nights=2)
        checkin = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Ravi",
                    "lastName": "Menon",
                    "mobile": "9000000099",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 2,
                    "roomRate": 2500,
                    "totalRate": 5000,
                    "extraBedAmount": 500,
                    "earlyCheckinAmount": 250,
                    "advancePaid": 1000,
                    "folioCharges": [
                        {
                            "id": "fc-test-1",
                            "kind": "restaurant_room_transfer",
                            "label": "Restaurant Room Transfer · ORD-1",
                            "amount": 420,
                            "source": "pos",
                            "invoiceId": "9",
                            "outlet": "restaurant",
                        }
                    ],
                },
            },
        )
        self.assertEqual(checkin.status_code, 200, checkin.get_data(as_text=True))
        stay = checkin.get_json()["room"]["stay"]
        self.assertEqual(stay["roomRate"], 2500)
        self.assertEqual(stay["extraBedAmount"], 500)
        self.assertEqual(stay["earlyCheckinAmount"], 250)
        self.assertEqual(len(stay["folioCharges"]), 1)
        self.assertEqual(stay["folioCharges"][0]["kind"], "restaurant_room_transfer")
        self.assertEqual(stay["folioCharges"][0]["orderNo"], "ORD-1")
        self.assertEqual(
            stay["folioCharges"][0]["label"],
            "Restaurant Room Transfer · ORD-1",
        )
        # room 5000 + extras 750 = 5750 hotel; POS transfer tracked separately
        self.assertEqual(stay["estimatedTotal"], 5750.0)
        self.assertEqual(stay["fbTransferTotal"], 420.0)
        self.assertEqual(stay["advancePaid"], 1000)
        self.assertEqual(stay["balanceAmount"], 4750.0)
        self.assertEqual(stay["combinedBalanceDue"], 5170.0)

        detail = self.client.get("/hotel/rooms/room-101")
        self.assertEqual(detail.status_code, 200)
        html = detail.get_data(as_text=True)
        self.assertIn("hrd-charges-card", html)
        self.assertIn("Estimated Charges", html)
        self.assertIn("data-charges-cgst", html)
        self.assertIn("data-charges-ugst", html)
        self.assertIn("CGST", html)
        self.assertIn("UGST", html)

        moved = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "transfer", "toRoomId": "room-105", "note": "Upgrade"},
        )
        self.assertEqual(moved.status_code, 200, moved.get_data(as_text=True))
        dest = moved.get_json()["toRoom"]["stay"]
        self.assertEqual(dest["firstName"], "Ravi")
        self.assertEqual(len(dest["folioCharges"]), 1)
        self.assertEqual(dest["folioCharges"][0]["amount"], 420)
        self.assertEqual(dest["estimatedTotal"], 5750.0)
        self.assertEqual(dest["fbTransferTotal"], 420.0)
        self.assertEqual(dest["balanceAmount"], 4750.0)
        self.assertEqual(dest["combinedBalanceDue"], 5170.0)
        self.assertNotIn("stay", moved.get_json()["fromRoom"])

    def test_pos_occupied_rooms_api_lists_in_house_only(self):
        empty = self.client.get("/point-of-sale/api/hotel-rooms/occupied")
        self.assertEqual(empty.status_code, 200)
        self.assertEqual(empty.get_json()["rooms"], [])

        self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "In",
                    "lastName": "House",
                    "mobile": "9000000088",
                    "checkInDate": "2026-07-29",
                    "roomRate": 2500,
                },
            },
        )
        self.client.put("/hotel/api/rooms/room-102", json={"status": "dirty"})

        listed = self.client.get("/point-of-sale/api/hotel-rooms/occupied")
        self.assertEqual(listed.status_code, 200)
        rooms = listed.get_json()["rooms"]
        self.assertEqual(len(rooms), 1)
        self.assertEqual(rooms[0]["id"], "room-101")
        self.assertIn("In", rooms[0]["guestName"])

    def _checkin_with_charges(self, room_id="room-101", advance=1000):
        check_in, check_out = self._stay_window(nights=2)
        res = self.client.put(
            f"/hotel/api/rooms/{room_id}",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Ravi",
                    "lastName": "Menon",
                    "mobile": "9000000099",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 2,
                    "roomRate": 2500,
                    "totalRate": 5000,
                    "extraBedAmount": 500,
                    "earlyCheckinAmount": 250,
                    "advancePaid": advance,
                },
            },
        )
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        return res.get_json()["room"]

    def _generate_stay_invoice(self, room_id="room-101"):
        res = self.client.put(
            f"/hotel/api/rooms/{room_id}",
            json={"action": "generate_invoice", "payment_splits": []},
        )
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        return res.get_json()["room"]

    def test_overstay_adds_room_charge_per_extra_day(self):
        today = datetime.now().date()
        check_in = (today - timedelta(days=2)).isoformat()
        check_out = (today - timedelta(days=1)).isoformat()  # expected yesterday
        res = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Over",
                    "lastName": "Stay",
                    "mobile": "9000000888",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 3500,
                    "advancePaid": 0,
                },
            },
        )
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        stay = res.get_json()["room"]["stay"]
        self.assertEqual(stay["nights"], 1)
        self.assertEqual(stay["overstayNights"], 1)
        self.assertEqual(stay["billableNights"], 2)
        # 3500 * 2 nights (tax-inclusive)
        self.assertEqual(stay["estimatedTotal"], 7000.0)
        self.assertEqual(stay["balanceAmount"], 7000.0)

        # Two days past expected checkout → two overstay nights.
        older_out = (today - timedelta(days=2)).isoformat()
        older_in = (today - timedelta(days=3)).isoformat()
        res2 = self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Long",
                    "lastName": "Stay",
                    "mobile": "9000000889",
                    "checkInDate": older_in,
                    "checkOutDate": older_out,
                    "nights": 1,
                    "roomRate": 3500,
                    "advancePaid": 0,
                },
            },
        )
        self.assertEqual(res2.status_code, 200, res2.get_data(as_text=True))
        stay2 = res2.get_json()["room"]["stay"]
        self.assertEqual(stay2["overstayNights"], 2)
        self.assertEqual(stay2["billableNights"], 3)
        self.assertEqual(stay2["estimatedTotal"], 10500.0)

    def test_generate_invoice_mints_number_and_allows_partial_payment(self):
        room = self._checkin_with_charges()
        stay = room["stay"]
        self.assertEqual(stay["estimatedTotal"], 5750.0)
        self.assertEqual(stay["advancePaid"], 1000)
        self.assertEqual(stay["balanceAmount"], 4750.0)
        self.assertFalse(stay.get("invoiceGenerated"))

        detail = self.client.get("/hotel/rooms/room-101")
        html = detail.get_data(as_text=True)
        self.assertIn("hrd-generate-hotel-invoice", html)
        self.assertIn("hrd-generate-fb-invoice", html)
        self.assertIn("hrd-invoice-modal", html)
        self.assertIn("hrd-invoice-add-split", html)
        self.assertIn("Split Payment", html)

        invoice_page = self.client.get("/hotel/rooms/room-101/invoice")
        self.assertEqual(invoice_page.status_code, 200, invoice_page.get_data(as_text=True))
        inv_html = invoice_page.get_data(as_text=True)
        self.assertIn("hotel-room-invoice-page", inv_html)
        self.assertIn("pos-invoice-page", inv_html)
        self.assertIn("pos_invoice.css", inv_html)
        self.assertIn("hri-generate", inv_html)
        self.assertIn("Generate Room Invoice", inv_html)
        self.assertIn('data-invoice-kind="hotel"', inv_html)
        self.assertIn("Settle Bill", inv_html)
        self.assertIn("Bill Summary", inv_html)
        self.assertIn("pos-inv-settle-modal", inv_html)
        self.assertIn("hri-sum-cgst", inv_html)
        self.assertIn("hri-sum-ugst", inv_html)
        self.assertIn("UGST", inv_html)
        self.assertIn("(2.5% incl.)", inv_html)
        self.assertIn("hri-tool-discount", inv_html)
        self.assertIn("Add Discount", inv_html)
        self.assertIn("hri-discount-modal", inv_html)
        self.assertIn("hri-add-custom", inv_html)
        self.assertIn("Custom Charges", inv_html)
        self.assertIn("hri-custom-modal", inv_html)

        generated = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "generate_invoice",
                "payment": {
                    "amount": 1500,
                    "method": "upi",
                    "reference": "UPI123",
                    "note": "Partial",
                },
            },
        )
        self.assertEqual(generated.status_code, 200, generated.get_data(as_text=True))
        body = generated.get_json()
        stay = body["room"]["stay"]
        self.assertTrue(body.get("minted"))
        self.assertTrue(stay["invoiceGenerated"])
        inv_no = str(stay["invoiceNumber"])
        fy = db_mod.indian_fiscal_year_label()
        self.assertEqual(inv_no, f"HBE/1/{fy}")
        self.assertEqual(stay["checkInAdvancePaid"], 1000)
        self.assertEqual(len(stay["payments"]), 1)
        self.assertEqual(stay["payments"][0]["amount"], 1500)
        self.assertEqual(stay["payments"][0]["method"], "upi")
        self.assertEqual(stay["advancePaid"], 2500)
        self.assertEqual(stay["balanceAmount"], 3250.0)
        self.assertEqual(body["room"]["status"], "occupied")

        # Second generate without pending charges is rejected (invoice is locked).
        again = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "amount": 0},
        )
        self.assertEqual(again.status_code, 400, again.get_data(as_text=True))
        self.assertIn("No pending charges", again.get_data(as_text=True))
        unchanged = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]["stay"]
        self.assertEqual(unchanged["invoiceNumber"], stay["invoiceNumber"])

    def test_generate_invoice_continues_hbe_short_fy_series(self):
        fy = db_mod.indian_fiscal_year_label()
        conn = db_mod.get_db()
        try:
            db_mod.ensure_hotel_rooms_schema(conn)
            conn.execute(
                """
                INSERT INTO hotel_room_invoices (
                    invoice_number, guest_name, estimated_total, balance_amount, status
                ) VALUES (?, 'Seed', 1000, 1000, 'open')
                """,
                (f"HBE/219/{fy}",),
            )
            conn.commit()
        finally:
            conn.close()

        self._checkin_with_charges("room-101", advance=0)
        room = self._generate_stay_invoice("room-101")
        self.assertEqual(room["stay"]["invoiceNumber"], f"HBE/220/{fy}")

        self._checkin_with_charges("room-102", advance=0)
        room_b = self._generate_stay_invoice("room-102")
        self.assertEqual(room_b["stay"]["invoiceNumber"], f"HBE/221/{fy}")

    def test_set_discount_updates_balance_and_locks_after_generate(self):
        room = self._checkin_with_charges()
        self.assertEqual(room["stay"]["estimatedTotal"], 5750.0)
        self.assertEqual(room["stay"]["balanceAmount"], 4750.0)

        discounted = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "set_discount",
                "discountType": "pct",
                "discountValue": 10,
            },
        )
        self.assertEqual(discounted.status_code, 200, discounted.get_data(as_text=True))
        stay = discounted.get_json()["room"]["stay"]
        self.assertEqual(stay["discountType"], "pct")
        self.assertEqual(stay["discountValue"], 10)
        self.assertEqual(stay["discountAmount"], 575.0)
        self.assertEqual(stay["estimatedTotal"], 5175.0)
        self.assertEqual(stay["balanceAmount"], 4175.0)

        needs_reason = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "set_discount",
                "discountType": "pct",
                "discountValue": 20,
            },
        )
        self.assertEqual(needs_reason.status_code, 400)
        self.assertIn("reason", needs_reason.get_json().get("error", "").lower())

        with_reason = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "set_discount",
                "discountType": "pct",
                "discountValue": 20,
                "discountReason": "Loyalty guest",
            },
        )
        self.assertEqual(with_reason.status_code, 200, with_reason.get_data(as_text=True))
        stay = with_reason.get_json()["room"]["stay"]
        self.assertEqual(stay["discountAmount"], 1150.0)
        self.assertEqual(stay["discountReason"], "Loyalty guest")

        generated = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "amount": 0},
        )
        self.assertEqual(generated.status_code, 200, generated.get_data(as_text=True))
        locked = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "set_discount", "discountType": "pct", "discountValue": 5},
        )
        self.assertEqual(locked.status_code, 400)
        self.assertIn("generated", locked.get_json().get("error", "").lower())

    def test_add_custom_charge_updates_folio_and_locks_after_generate(self):
        room = self._checkin_with_charges()
        before = room["stay"]["estimatedTotal"]
        added = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "add_custom_charge",
                "label": "Mini bar",
                "amount": 350,
            },
        )
        self.assertEqual(added.status_code, 200, added.get_data(as_text=True))
        stay = added.get_json()["room"]["stay"]
        self.assertEqual(stay["estimatedTotal"], round(before + 350, 2))
        labels = [f.get("label") for f in stay.get("folioCharges") or []]
        self.assertIn("Mini bar", labels)

        generated = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "amount": 0},
        )
        self.assertEqual(generated.status_code, 200, generated.get_data(as_text=True))
        locked = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "add_custom_charge", "label": "Late fee", "amount": 100},
        )
        self.assertEqual(locked.status_code, 400)
        self.assertIn("generated", locked.get_json().get("error", "").lower())

    def test_update_and_delete_folio_charges(self):
        room = self._checkin_with_charges()
        self.assertEqual(room["stay"]["extraBedAmount"], 500)

        updated = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "update_charge",
                "chargeKey": "extra_bed",
                "amount": 750,
            },
        )
        self.assertEqual(updated.status_code, 200, updated.get_data(as_text=True))
        self.assertEqual(updated.get_json()["room"]["stay"]["extraBedAmount"], 750)

        room_rate = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "update_charge",
                "chargeKey": "room",
                "rate": 3000,
                "label": "Deluxe Suite Tariff",
            },
        )
        self.assertEqual(room_rate.status_code, 200, room_rate.get_data(as_text=True))
        stay = room_rate.get_json()["room"]["stay"]
        self.assertEqual(stay["roomRate"], 3000)
        self.assertEqual(stay["totalRate"], 6000)
        self.assertEqual(
            (stay.get("chargeLabels") or {}).get("room"),
            "Deluxe Suite Tariff",
        )

        deleted = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "delete_charge", "chargeKey": "extra_bed"},
        )
        self.assertEqual(deleted.status_code, 200, deleted.get_data(as_text=True))
        self.assertEqual(deleted.get_json()["room"]["stay"]["extraBedAmount"], 0)

        refuse_room = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "delete_charge", "chargeKey": "room"},
        )
        self.assertEqual(refuse_room.status_code, 400)

        custom = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "add_custom_charge",
                "label": "Laundry",
                "amount": 200,
            },
        )
        self.assertEqual(custom.status_code, 200, custom.get_data(as_text=True))
        folio = custom.get_json()["room"]["stay"]["folioCharges"]
        laundry = next(f for f in folio if f.get("label") == "Laundry")
        folio_key = "folio:" + laundry["id"]
        renamed = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "update_charge",
                "chargeKey": folio_key,
                "label": "Laundry service",
                "amount": 250,
            },
        )
        self.assertEqual(renamed.status_code, 200, renamed.get_data(as_text=True))
        folio2 = renamed.get_json()["room"]["stay"]["folioCharges"]
        item = next(f for f in folio2 if f.get("id") == laundry["id"])
        self.assertEqual(item["label"], "Laundry service")
        self.assertEqual(item["amount"], 250)

        removed = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "delete_charge", "chargeKey": folio_key},
        )
        self.assertEqual(removed.status_code, 200, removed.get_data(as_text=True))
        labels = [f.get("label") for f in removed.get_json()["room"]["stay"]["folioCharges"]]
        self.assertNotIn("Laundry service", labels)

    def test_checkin_other_special_folio_charges(self):
        check_in, check_out = self._stay_window(nights=2)
        night2 = (datetime.fromisoformat(check_in).date() + timedelta(days=1)).isoformat()
        res = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Custom",
                    "lastName": "Charges",
                    "mobile": "9000000123",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 2,
                    "roomRate": 2500,
                    "specialRequests": ["Other"],
                    "folioCharges": [
                        {
                            "id": "fc111",
                            "kind": "other",
                            "source": "checkin_special",
                            "label": "Laundry",
                            "serviceDate": check_in,
                            "qty": 2,
                            "rate": 200,
                            "amount": 400,
                        },
                        {
                            "id": "fc222",
                            "kind": "other",
                            "source": "checkin_special",
                            "label": "Spa",
                            "serviceDate": night2,
                            "qty": 1,
                            "rate": 1500,
                            "amount": 1500,
                        },
                    ],
                },
            },
        )
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        stay = res.get_json()["room"]["stay"]
        folio = stay["folioCharges"]
        self.assertEqual(len(folio), 2)
        laundry = next(f for f in folio if f.get("label") == "Laundry")
        spa = next(f for f in folio if f.get("label") == "Spa")
        self.assertEqual(laundry["serviceDate"], check_in)
        self.assertEqual(laundry["qty"], 2)
        self.assertEqual(laundry["rate"], 200)
        self.assertEqual(laundry["amount"], 400)
        self.assertEqual(spa["serviceDate"], night2)
        self.assertEqual(spa["amount"], 1500)
        self.assertIn("Other", stay.get("specialRequests") or [])
        self.assertEqual(stay["estimatedTotal"], 6900.0)

    def test_update_room_rate_propagates_to_nightly_rates(self):
        """Editing Rate ₹/night must rewrite booked nightlyRates, not only roomRate."""
        check_in, check_out = self._stay_window(nights=2)
        night2 = (datetime.fromisoformat(check_in).date() + timedelta(days=1)).isoformat()
        res = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Nightly",
                    "lastName": "Edit",
                    "mobile": "9000000111",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 2,
                    "roomRate": 3500,
                    "ratePlan": "EP",
                    "nightlyRates": [
                        {"date": check_in, "roomRate": 3500, "ratePlan": "EP"},
                        {"date": night2, "roomRate": 3500, "ratePlan": "EP"},
                    ],
                    "advancePaid": 0,
                },
            },
        )
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        stay = res.get_json()["room"]["stay"]
        self.assertEqual(stay["totalRate"], 7000.0)

        updated = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "update_charge",
                "chargeKey": "room",
                "rate": 4000,
                "label": "Deluxe with Balcony",
            },
        )
        self.assertEqual(updated.status_code, 200, updated.get_data(as_text=True))
        stay2 = updated.get_json()["room"]["stay"]
        self.assertEqual(stay2["roomRate"], 4000.0)
        self.assertEqual(stay2["totalRate"], 8000.0)
        self.assertEqual(
            [row["roomRate"] for row in (stay2.get("nightlyRates") or [])],
            [4000.0, 4000.0],
        )
        self.assertEqual(
            (stay2.get("chargeLabels") or {}).get("room"),
            "Deluxe with Balcony",
        )

        night_edit = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "update_charge",
                "chargeKey": "night:1",
                "rate": 4200,
            },
        )
        self.assertEqual(night_edit.status_code, 200, night_edit.get_data(as_text=True))
        stay3 = night_edit.get_json()["room"]["stay"]
        self.assertEqual(stay3["nightlyRates"][0]["roomRate"], 4000.0)
        self.assertEqual(stay3["nightlyRates"][1]["roomRate"], 4200.0)
        self.assertEqual(stay3["totalRate"], 8200.0)

    def test_record_payment_rejects_before_generate_and_overpay(self):
        self._checkin_with_charges()
        early = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "record_payment", "amount": 100, "method": "cash"},
        )
        self.assertEqual(early.status_code, 400)
        self.assertIn("Generate", early.get_json().get("error", ""))

        self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "amount": 0},
        )
        over = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "record_payment",
                "amount": 99999,
                "method": "cash",
            },
        )
        self.assertEqual(over.status_code, 400)
        self.assertIn("exceeds", over.get_json().get("error", "").lower())

        bank = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "record_payment",
                "amount": 100,
                "method": "bank_transfer",
            },
        )
        self.assertEqual(bank.status_code, 400)
        self.assertIn("reference", bank.get_json().get("error", "").lower())

    def test_full_payment_then_checkout(self):
        self._checkin_with_charges(advance=0)
        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "generate_invoice",
                "payment": {"amount": 2000, "method": "cash"},
            },
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))
        stay = gen.get_json()["room"]["stay"]
        self.assertEqual(stay["advancePaid"], 2000)
        balance = stay["balanceAmount"]
        self.assertGreater(balance, 0)

        paid = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "record_payment",
                "amount": balance,
                "method": "card",
            },
        )
        self.assertEqual(paid.status_code, 200, paid.get_data(as_text=True))
        stay = paid.get_json()["room"]["stay"]
        self.assertEqual(stay["balanceAmount"], 0)
        self.assertEqual(len(stay["payments"]), 2)
        self.assertEqual(stay["advancePaid"], stay["estimatedTotal"])

        closed = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "checkout"},
        )
        self.assertEqual(closed.status_code, 200, closed.get_data(as_text=True))
        room = closed.get_json()["room"]
        self.assertEqual(room["status"], "dirty")
        self.assertNotIn("stay", room)

    def test_generate_invoice_split_payment(self):
        self._checkin_with_charges(advance=0)
        split = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "generate_invoice",
                "payment_splits": [
                    {"payment_method": "cash", "amount": 3000},
                    {
                        "payment_method": "upi",
                        "amount": 2000,
                        "transaction_id": "",
                    },
                ],
                "note": "Split at desk",
            },
        )
        self.assertEqual(split.status_code, 200, split.get_data(as_text=True))
        stay = split.get_json()["room"]["stay"]
        self.assertTrue(stay["invoiceGenerated"])
        self.assertEqual(stay["advancePaid"], 5000)
        self.assertEqual(stay["balanceAmount"], 750.0)
        self.assertEqual(len(stay["payments"]), 2)
        methods = sorted(p["method"] for p in stay["payments"])
        self.assertEqual(methods, ["cash", "upi"])

        over = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "record_payment",
                "payment_splits": [
                    {"payment_method": "cash", "amount": 600},
                    {"payment_method": "card", "amount": 600},
                ],
            },
        )
        self.assertEqual(over.status_code, 400)

        done = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "record_payment",
                "payment_splits": [
                    {"payment_method": "cash", "amount": 400},
                    {
                        "payment_method": "bank_transfer",
                        "amount": 350,
                        "transaction_id": "NEFT991",
                    },
                ],
            },
        )
        self.assertEqual(done.status_code, 200, done.get_data(as_text=True))
        stay = done.get_json()["room"]["stay"]
        self.assertEqual(stay["balanceAmount"], 0)
        self.assertEqual(len(stay["payments"]), 4)

    def test_credit_payment_requires_agency(self):
        self._checkin_with_charges(advance=0)
        denied = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "generate_invoice",
                "payment_splits": [
                    {"payment_method": "credit", "amount": 1000},
                ],
            },
        )
        self.assertEqual(denied.status_code, 400)
        self.assertIn("agency", denied.get_json().get("error", "").lower())

        with_agency = self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Asha",
                    "lastName": "Iyer",
                    "mobile": "9000000077",
                    "checkInDate": "2026-07-29",
                    "nights": 1,
                    "roomRate": 2000,
                    "advancePaid": 0,
                    "agencyName": "Travel Desk Co",
                    "agencyBilling": True,
                },
            },
        )
        self.assertEqual(with_agency.status_code, 200, with_agency.get_data(as_text=True))
        credited = self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "generate_invoice",
                "payment_splits": [
                    {"payment_method": "credit", "amount": 500},
                    {"payment_method": "cash", "amount": 500},
                ],
            },
        )
        self.assertEqual(credited.status_code, 200, credited.get_data(as_text=True))
        stay = credited.get_json()["room"]["stay"]
        methods = sorted(p["method"] for p in stay["payments"])
        self.assertEqual(methods, ["cash", "credit"])
        self.assertEqual(stay["advancePaid"], 1000)

    def test_invoice_ledger_lists_generated_invoice(self):
        room = self._checkin_with_charges(advance=0)
        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "generate_invoice",
                "payment": {"amount": 500, "method": "cash"},
            },
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))
        inv_no = gen.get_json()["room"]["stay"]["invoiceNumber"]
        self.assertTrue(inv_no)

        page = self.client.get("/hotel/invoice-ledger")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn("Invoice Ledger", html)
        self.assertIn(inv_no, html)
        self.assertIn("is-open", html)
        self.assertIn("Un Settled", html)
        self.assertIn("Created by", html)
        self.assertIn('data-sort="created_by"', html)
        self.assertIn("Administrator", html)
        self.assertIn('data-hil-kpi="settled"', html)
        self.assertIn('data-sort="guest"', html)
        self.assertIn('data-sort="agency"', html)
        self.assertIn(">Agency</th>", html)
        self.assertIn("hil-agency-listbox", html)
        self.assertIn('name="agency"', html)
        self.assertIn("All agencies", html)
        self.assertIn("pos-inv-settle-modal", html)
        self.assertIn("hotel_settle_modal.js", html)
        self.assertIn("data-hil-settle", html)
        self.assertIn('data-room-id="room-101"', html)
        self.assertIn("Settle invoice", html)
        self.assertIn("hil-view-btn", html)
        self.assertIn("hil-print-btn", html)
        self.assertIn("hil-edit-btn", html)
        self.assertIn("hil-cancel-btn", html)
        self.assertIn("hil-select-all", html)
        self.assertIn("hil-settle-selected", html)
        self.assertIn("hil-status-settle", html)
        self.assertIn("hil-row-check", html)
        self.assertIn("settle-selected", html)
        self.assertIn("de-nav-hotel-invoice-ledger", html)
        self.assertIn("de-nav-hotel-credit", html)
        self.assertIn("hil-invoice-listbox", html)
        self.assertIn("hil-invoice-tabs", html)
        self.assertIn("Room Transfer", html)
        self.assertIn("hil-open-room-transfer", html)
        self.assertIn("/hotel/room-transfer-invoices", html)
        self.assertNotIn("hil-rt-overlay", html)
        self.assertNotIn('data-value="room_transfer"', html)
        self.assertIn(">Invoice</span>", html)

        api = self.client.get(f"/hotel/invoice-ledger/api/{inv_no}")
        self.assertEqual(api.status_code, 200)
        payload = api.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["invoice"]["invoice_number"], inv_no)
        self.assertEqual(payload["room"]["stay"]["invoiceNumber"], inv_no)
        self.assertGreater(float(payload["invoice"]["balance_amount"]), 0)
        self.assertEqual(payload["invoice"]["status"], "open")

        settle = self.client.post(
            f"/hotel/invoice-ledger/api/{inv_no}/settle",
            json={
                "payment_splits": [
                    {"method": "cash", "amount": payload["invoice"]["balance_amount"]}
                ]
            },
        )
        self.assertEqual(settle.status_code, 200, settle.get_data(as_text=True))
        settled = settle.get_json()
        self.assertTrue(settled["ok"])
        self.assertEqual(settled["invoice"]["status"], "settled")
        self.assertLessEqual(float(settled["invoice"]["balance_amount"]), 0.009)
        ledger = self.client.get("/hotel/invoice-ledger")
        self.assertIn("Payment Mode", ledger.get_data(as_text=True))
        conn = db_mod.get_db()
        try:
            rows = db_mod.list_hotel_room_invoices(conn, q=inv_no)
        finally:
            conn.close()
        match = next(row for row in rows if row["invoice_number"] == inv_no)
        self.assertEqual(match["status"], "settled")
        self.assertEqual(match["payment_mode_label"], "Cash")
        self.assertEqual(match["payment_modes"], ["cash"])
        self.assertAlmostEqual(float(match["payment_amounts"]["cash"]), float(settled["invoice"]["estimated_total"]), places=2)

    def test_invoice_ledger_payment_mode_shows_split_tenders(self):
        self._checkin_with_charges(advance=0)
        stay = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]["stay"]
        total = round(float(stay["balanceAmount"]), 2)
        cash_part = round(total / 2, 2)
        upi_part = round(total - cash_part, 2)
        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "generate_invoice",
                "payment_splits": [
                    {"payment_method": "cash", "amount": cash_part},
                    {"payment_method": "upi", "amount": upi_part},
                ],
            },
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))
        inv_no = gen.get_json()["room"]["stay"]["invoiceNumber"]
        conn = db_mod.get_db()
        try:
            rows = db_mod.list_hotel_room_invoices(conn, q=inv_no)
            kpis = db_mod.hotel_room_invoice_kpis(rows)
        finally:
            conn.close()
        match = next(row for row in rows if row["invoice_number"] == inv_no)
        self.assertEqual(match["status"], "settled")
        self.assertEqual(match["payment_mode_label"], "Cash + UPI")
        self.assertEqual(match["payment_modes"], ["cash", "upi"])
        self.assertAlmostEqual(float(match["payment_amounts"]["cash"]), cash_part, places=2)
        self.assertAlmostEqual(float(match["payment_amounts"]["upi"]), upi_part, places=2)
        self.assertAlmostEqual(float(kpis["payment_totals"]["cash"]), cash_part, places=2)
        self.assertAlmostEqual(float(kpis["payment_totals"]["upi"]), upi_part, places=2)

        page = self.client.get("/hotel/invoice-ledger")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn('data-sort="pay_cash"', html)
        self.assertNotIn('aria-label="Settlement by payment mode"', html)
        self.assertNotIn("hil-settlement-summary", html)
        self.assertIn(">Cash<", html)
        self.assertIn(">UPI<", html)

        export = self.client.get("/hotel/invoice-ledger/export")
        self.assertEqual(export.status_code, 200)
        from io import BytesIO
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(export.data))
        ws = wb.active
        headers = [ws.cell(3, col).value for col in range(1, ws.max_column + 1)]
        self.assertEqual(headers[11], "Payment Mode")
        self.assertEqual(headers[12], "Cash")
        self.assertEqual(headers[13], "UPI")
        self.assertEqual(headers[14], "Card")
        self.assertEqual(headers[-1], "Created by")
        order_row = None
        for row in range(4, ws.max_row + 1):
            if ws.cell(row, 1).value == inv_no:
                order_row = row
                break
        self.assertIsNotNone(order_row)
        self.assertAlmostEqual(float(ws.cell(order_row, 13).value), cash_part, places=2)
        self.assertAlmostEqual(float(ws.cell(order_row, 14).value), upi_part, places=2)
        self.assertEqual(ws.cell(order_row, ws.max_column).value, "Administrator")

    def test_fb_transfer_folio_keeps_pos_vat_not_hotel_gst(self):
        """Room-transfer folio tax must follow the POS invoice (e.g. VAT 10%)."""
        self._checkin_with_charges(advance=0)
        conn = db_mod.get_db()
        try:
            db_mod.ensure_pos_schema(conn)
            cur = conn.execute(
                """
                INSERT INTO pos_invoices
                   (order_no, order_date, order_type, table_label, customer_name, customer_mobile,
                    captain, status, outlet, subtotal, discount_amount, gst_amount, vat_amount,
                    service_amount, tip, round_off, grand_total, saved_at, is_active)
                VALUES (?, date('now','localtime'), 'dine_in', 'T1', 'Guest', '',
                        '', 'closed', ?, 1000, 0, 0, 100, 0, 0, 0, 1100,
                        datetime('now','localtime'), 1)
                """,
                ("INV/26-27/497", db_mod.POS_OUTLET_BAR),
            )
            pos_id = cur.lastrowid
            result = db_mod.append_hotel_room_folio_charge(
                conn,
                "room-101",
                amount=1100,
                kind="bar_room_transfer",
                label="Bar Room Transfer · INV/26-27/497",
                source="pos",
                invoice_id=str(pos_id),
                order_no="INV/26-27/497",
                outlet="bar",
            )
            charge = result["charge"]
            self.assertAlmostEqual(float(charge.get("vat") or 0), 100.0, places=2)
            self.assertAlmostEqual(float(charge.get("gst") or 0), 0.0, places=2)
            self.assertAlmostEqual(float(charge.get("subtotal") or 0), 1000.0, places=2)
            self.assertAlmostEqual(float(charge.get("vatPct") or 0), 10.0, places=2)
            conn.commit()

            room = db_mod.get_hotel_room(conn, "room-101")
            fb = [
                line
                for line in (room["stay"].get("folioCharges") or [])
                if line.get("orderNo") == "INV/26-27/497"
            ]
            self.assertEqual(len(fb), 1)
            self.assertAlmostEqual(float(fb[0].get("vat") or 0), 100.0, places=2)
            self.assertAlmostEqual(float(fb[0].get("gst") or 0), 0.0, places=2)
            self.assertAlmostEqual(float(fb[0].get("subtotal") or 0), 1000.0, places=2)

            # Existing folio without tax snapshot is backfilled on room load.
            bare = {
                "id": "fc-bare",
                "kind": "bar_room_transfer",
                "label": "Bar Room Transfer · INV/26-27/497",
                "amount": 1100,
                "invoiceId": str(pos_id),
                "orderNo": "INV/26-27/497",
                "outlet": "bar",
                "settled": False,
            }
            enriched_stay = db_mod._hotel_enrich_folio_transfer_tax(
                conn, {"folioCharges": [bare]}
            )
            line = enriched_stay["folioCharges"][0]
            self.assertAlmostEqual(float(line.get("vat") or 0), 100.0, places=2)
            self.assertAlmostEqual(float(line.get("subtotal") or 0), 1000.0, places=2)
        finally:
            conn.close()

    def test_fbe_ledger_settlement_records_real_tender_amounts(self):
        self._checkin_with_charges(advance=0)
        conn = db_mod.get_db()
        try:
            result = db_mod.append_hotel_room_folio_charge(
                conn,
                "room-101",
                amount=400,
                kind="restaurant_room_transfer",
                label="Restaurant Room Transfer · SPC/26-27/88",
                source="pos",
                invoice_id="88",
                order_no="SPC/26-27/88",
                outlet="restaurant",
            )
            conn.commit()
        finally:
            conn.close()
        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "invoice_kind": "fb", "payment_splits": []},
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))
        fb_no = gen.get_json()["room"]["stay"]["fbTransferInvoiceNumber"]
        self.assertTrue(str(fb_no).startswith("FBE/"))

        settle = self.client.post(
            f"/hotel/invoice-ledger/api/{fb_no}/settle",
            json={
                "payment_splits": [
                    {"method": "cash", "amount": 150},
                    {"method": "upi", "amount": 250},
                ]
            },
        )
        self.assertEqual(settle.status_code, 200, settle.get_data(as_text=True))
        settled = settle.get_json()["invoice"]
        self.assertEqual(settled["status"], "settled")
        self.assertEqual(settled["payment_mode_label"], "Cash + UPI")
        self.assertEqual(settled["payment_modes"], ["cash", "upi"])
        self.assertAlmostEqual(float(settled["payment_amounts"]["cash"]), 150, places=2)
        self.assertAlmostEqual(float(settled["payment_amounts"]["upi"]), 250, places=2)

        conn = db_mod.get_db()
        try:
            room = db_mod.get_hotel_room(conn, "room-101")
            stay = room["stay"]
            methods = [
                p.get("method")
                for p in (stay.get("fbTransferPayments") or [])
                if isinstance(p, dict)
            ]
        finally:
            conn.close()
        self.assertIn("cash", methods)
        self.assertIn("upi", methods)
        self.assertNotIn("checkout", methods)

    def test_pos_room_transfer_lists_on_invoice_ledger(self):
        """Per-POS room transfers list on Room Transfer, not Invoice Ledger."""
        self._checkin_with_charges(advance=0)
        conn = db_mod.get_db()
        try:
            result = db_mod.append_hotel_room_folio_charge(
                conn,
                "room-101",
                amount=150.48,
                kind="restaurant_room_transfer",
                label="Restaurant Room Transfer · SPC/26-27/12",
                source="pos",
                invoice_id="99",
                order_no="SPC/26-27/12",
                outlet="restaurant",
            )
            db_mod.upsert_pos_room_transfer_invoice(
                conn, result["room"], result["charge"]
            )
            bar = db_mod.append_hotel_room_folio_charge(
                conn,
                "room-101",
                amount=40.0,
                kind="bar_room_transfer",
                label="Bar Room Transfer · BAR/26-27/1",
                source="pos",
                invoice_id="100",
                order_no="BAR/26-27/1",
                outlet="bar",
            )
            db_mod.upsert_pos_room_transfer_invoice(
                conn, bar["room"], bar["charge"]
            )
            rt_spc = db_mod._pos_room_transfer_ledger_invoice_number(
                conn, "SPC/26-27/12"
            )
            rt_bar = db_mod._pos_room_transfer_ledger_invoice_number(
                conn, "BAR/26-27/1"
            )
            conn.commit()
        finally:
            conn.close()

        self.assertTrue(str(rt_spc or "").startswith("RT/"))
        self.assertTrue(str(rt_bar or "").startswith("RT/"))

        page = self.client.get("/hotel/invoice-ledger")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertNotIn(rt_spc, html)
        self.assertIn("hil-open-room-transfer", html)
        self.assertIn("/hotel/room-transfer-invoices", html)
        self.assertNotIn("hil-rt-overlay", html)
        self.assertNotIn('id="hil-rt-frame"', html)
        self.assertNotIn('data-value="room_transfer"', html)

        room_transfer = self.client.get("/hotel/room-transfer-invoices")
        self.assertEqual(room_transfer.status_code, 200)
        rt_html = room_transfer.get_data(as_text=True)
        self.assertIn(rt_spc, rt_html)
        self.assertIn(rt_bar, rt_html)
        self.assertIn("Invoice yet to generate", rt_html)
        self.assertIn('id="hrt-kpi-open">2<', rt_html)
        self.assertIn("Yet to generate", rt_html)
        self.assertIn("Invoice Generated", rt_html)
        self.assertIn('id="hrt-kpi-outstanding">0<', rt_html)
        self.assertIn('id="hrt-outlet-tabs"', rt_html)
        self.assertIn('data-value="bar"', rt_html)
        self.assertIn('data-value="restaurant"', rt_html)
        self.assertIn(">Status<", rt_html)
        self.assertNotIn(">Payment Mode<", rt_html)
        self.assertIn('id="hrt-filter-form"', rt_html)
        self.assertIn('data-invoice-source="pos_room_transfer"', rt_html)
        self.assertIn('data-ledger-prefix="hrt"', rt_html)
        self.assertIn('data-room-transfer-ledger="1"', rt_html)
        self.assertNotIn('data-hil-kpi="settled"', rt_html)
        self.assertNotIn('id="hrt-kpi-settled"', rt_html)
        self.assertNotIn("hil-status-settle", rt_html)
        self.assertNotIn("data-hil-settle", rt_html)
        self.assertNotIn("hrt-settle-selected", rt_html)
        self.assertNotIn("hil-row-check", rt_html)
        self.assertNotIn("hrt-select-all", rt_html)
        self.assertIn("de-nav-hotel-invoice-ledger", rt_html)
        self.assertIn("Back to Invoice Ledger", rt_html)
        self.assertIn('class="su-page-back"', rt_html)
        self.assertNotIn("hil-invoice-listbox", rt_html)
        self.assertNotIn("hil-open-room-transfer", rt_html)

        restaurant_only = self.client.get(
            "/hotel/room-transfer-invoices?outlet=restaurant"
        )
        restaurant_html = restaurant_only.get_data(as_text=True)
        self.assertIn(rt_spc, restaurant_html)
        self.assertNotIn(rt_bar, restaurant_html)
        self.assertIn('id="hrt-kpi-open">1<', restaurant_html)
        bar_only = self.client.get("/hotel/room-transfer-invoices?outlet=bar")
        bar_html = bar_only.get_data(as_text=True)
        self.assertIn(rt_bar, bar_html)
        self.assertNotIn(rt_spc, bar_html)

        hotel_only = self.client.get("/hotel/invoice-ledger?invoice=hotel")
        self.assertNotIn(rt_spc, hotel_only.get_data(as_text=True))

        stay = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]["stay"]
        balance_before = float(stay["balanceAmount"])
        settle = self.client.post(
            "/hotel/invoice-ledger/api/SPC/26-27/12/settle",
            json={
                "payment_splits": [{"method": "cash", "amount": 150.48}],
            },
        )
        self.assertEqual(settle.status_code, 200, settle.get_data(as_text=True))
        settled = settle.get_json()
        self.assertTrue(settled["ok"])
        self.assertEqual(settled["invoice"]["status"], "settled")
        self.assertEqual(settled["invoice"]["source"], "pos_room_transfer")
        stay_after = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]["stay"]
        self.assertAlmostEqual(
            float(stay_after["balanceAmount"]),
            balance_before,
            places=2,
        )
        folio = stay_after["folioCharges"]
        self.assertTrue(any(f.get("settled") for f in folio))

        # Payment-settled transfer still counts as Un Settled until FBE is generated.
        rt_after = self.client.get("/hotel/room-transfer-invoices").get_data(as_text=True)
        self.assertIn("Invoice yet to generate", rt_after)
        self.assertIn('id="hrt-kpi-open">2<', rt_after)
        filtered = self.client.get("/hotel/room-transfer-invoices?status=open")
        self.assertEqual(filtered.status_code, 200)
        filtered_html = filtered.get_data(as_text=True)
        self.assertIn(rt_spc, filtered_html)
        self.assertIn("Invoice yet to generate", filtered_html)

    def test_invoice_ledger_filters_by_agency_name(self):
        """Agency filter shows only invoices booked through that agency."""
        check_in, check_out = self._stay_window(nights=1)
        agency = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Agency",
                    "lastName": "Guest",
                    "mobile": "9000000091",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 1000,
                    "totalRate": 1000,
                    "advancePaid": 0,
                    "agencyName": "ATPI India Pvt. Ltd",
                },
            },
        )
        self.assertEqual(agency.status_code, 200, agency.get_data(as_text=True))
        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "payment_splits": []},
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))
        inv_no = gen.get_json()["room"]["stay"]["invoiceNumber"]

        page = self.client.get(
            "/hotel/invoice-ledger",
            query_string={"agency": "ATPI India Pvt. Ltd"},
        )
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn(inv_no, html)
        self.assertIn("ATPI India Pvt. Ltd", html)
        self.assertIn('name="agency"', html)
        self.assertIn('value="ATPI India Pvt. Ltd"', html)

        other = self.client.get(
            "/hotel/invoice-ledger",
            query_string={"agency": "Some Other Agency"},
        )
        self.assertEqual(other.status_code, 200)
        other_html = other.get_data(as_text=True)
        self.assertNotIn(inv_no, other_html)

    def test_invoice_ledger_settle_with_back_office_receipt(self):
        """Agency stay can settle hotel invoice using a pending Back Office Receipt."""
        from datetime import date

        from back_office_receipt import (
            create_back_office_receipt,
            list_pending_back_office_receipts_for_agency,
        )

        check_in, check_out = self._stay_window(nights=1)
        res = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Bor",
                    "lastName": "Guest",
                    "mobile": "9000000088",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 1000,
                    "totalRate": 1000,
                    "advancePaid": 0,
                    "agencyName": "Travel Co",
                    "agencyGst": "27AAAAA0000A1Z5",
                },
            },
        )
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))

        conn = db_mod.get_db()
        try:
            receipt = create_back_office_receipt(
                conn,
                receipt_date=date.today(),
                payer_name="Travel Co",
                agency_id=None,
                amount=5000,
                payment_mode="cash",
                towards="Advance",
            )
            conn.commit()
            receipt_id = receipt["id"]
            pending = list_pending_back_office_receipts_for_agency(
                conn, agency_name="Travel Co"
            )
            self.assertTrue(any(int(p["id"]) == int(receipt_id) for p in pending))
        finally:
            conn.close()

        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "payment_splits": []},
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))
        stay = gen.get_json()["room"]["stay"]
        inv_no = stay["invoiceNumber"]
        balance = float(stay["balanceAmount"])
        self.assertGreater(balance, 0)

        settle = self.client.post(
            f"/hotel/invoice-ledger/api/{inv_no}/settle",
            json={
                "payment_splits": [
                    {
                        "method": "bor",
                        "amount": balance,
                        "receipt_id": receipt_id,
                    }
                ],
            },
        )
        self.assertEqual(settle.status_code, 200, settle.get_data(as_text=True))
        settled = settle.get_json()
        self.assertTrue(settled["ok"])
        self.assertEqual(settled["invoice"]["status"], "settled")

        conn = db_mod.get_db()
        try:
            pending_after = list_pending_back_office_receipts_for_agency(
                conn, agency_name="Travel Co"
            )
            left = next(
                (p for p in pending_after if int(p["id"]) == int(receipt_id)), None
            )
            if left:
                self.assertAlmostEqual(
                    float(left["pending_amount"]), 5000 - balance, places=2
                )
            else:
                self.assertGreaterEqual(balance, 5000 - 0.01)
            alloc = conn.execute(
                """
                SELECT amount, hotel_invoice_number
                FROM back_office_receipt_allocations
                WHERE receipt_id = ? AND hotel_invoice_number = ?
                """,
                (receipt_id, inv_no),
            ).fetchone()
            self.assertIsNotNone(alloc)
            self.assertAlmostEqual(float(alloc["amount"]), balance, places=2)
        finally:
            conn.close()

        page = self.client.get("/hotel/invoice-ledger")
        html = page.get_data(as_text=True)
        self.assertIn("Back Office Receipt", html)
        self.assertIn("hil-settle-bor-field", html)

    def test_pos_room_transfer_ledger_allow_credit_when_stay_has_agency(self):
        """Room-transfer bills on the ledger expose Credit when the live stay has an agency."""
        check_in, check_out = self._stay_window(nights=2)
        res = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Agency",
                    "lastName": "Guest",
                    "mobile": "9000000099",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 2,
                    "roomRate": 2500,
                    "totalRate": 5000,
                    "advancePaid": 0,
                    "agencyName": "Travel Co",
                    "agencyGst": "27AAAAA0000A1Z5",
                },
            },
        )
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        conn = db_mod.get_db()
        try:
            result = db_mod.append_hotel_room_folio_charge(
                conn,
                "room-101",
                amount=158.0,
                kind="restaurant_room_transfer",
                label="Restaurant Room Transfer · SPC/26-27/99",
                source="pos",
                invoice_id="199",
                order_no="SPC/26-27/99",
                outlet="restaurant",
            )
            db_mod.upsert_pos_room_transfer_invoice(
                conn, result["room"], result["charge"]
            )
            conn.commit()
        finally:
            conn.close()

        api = self.client.get("/hotel/invoice-ledger/api/SPC/26-27/99")
        self.assertEqual(api.status_code, 200, api.get_data(as_text=True))
        payload = api.get_json()
        self.assertTrue(payload["allow_credit"])
        self.assertEqual(
            payload["room"]["stay"].get("agencyName")
            or payload["room"]["stay"].get("agency_name"),
            "Travel Co",
        )

        settle = self.client.post(
            "/hotel/invoice-ledger/api/SPC/26-27/99/settle",
            json={
                "payment_splits": [{"method": "credit", "amount": 158.0}],
            },
        )
        self.assertEqual(settle.status_code, 200, settle.get_data(as_text=True))
        settled = settle.get_json()
        self.assertTrue(settled["ok"])
        self.assertEqual(settled["invoice"]["status"], "settled")

    def test_stay_payment_settles_linked_pos_room_transfer(self):
        self._checkin_with_charges(advance=0)
        conn = db_mod.get_db()
        try:
            result = db_mod.append_hotel_room_folio_charge(
                conn,
                "room-101",
                amount=80,
                kind="bar_room_transfer",
                label="Bar Room Transfer · INV/26-27/4",
                source="pos",
                invoice_id="100",
                order_no="INV/26-27/4",
                outlet="bar",
            )
            db_mod.upsert_pos_room_transfer_invoice(
                conn, result["room"], result["charge"]
            )
            conn.commit()
        finally:
            conn.close()
        stay = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]["stay"]
        combined_due = float(stay["combinedBalanceDue"])
        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "generate_invoice",
                "payment_splits": [
                    {"method": "upi", "amount": combined_due},
                ],
            },
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))
        payload = gen.get_json()
        fb_no = payload["room"]["stay"]["fbTransferInvoiceNumber"]
        self.assertTrue(str(fb_no).startswith("FBE/"))
        api = self.client.get(f"/hotel/invoice-ledger/api/{fb_no}")
        self.assertEqual(api.status_code, 200)
        item = api.get_json()["invoice"]
        self.assertEqual(item["status"], "settled")
        self.assertLessEqual(float(item["balance_amount"]), 0.009)
        self.assertEqual(item["source"], "fb_combined_transfer")
        spc = self.client.get("/hotel/invoice-ledger/api/INV/26-27/4")
        self.assertEqual(spc.status_code, 200, spc.get_data(as_text=True))
        spc_item = spc.get_json()["invoice"]
        self.assertEqual(spc_item["status"], "cancelled")
        self.assertEqual(spc_item["payment_mode_label"], f"Invoice Generated ({fb_no})")
        rt_html = self.client.get("/hotel/room-transfer-invoices").get_data(as_text=True)
        self.assertIn(f"Invoice Generated ({fb_no})", rt_html)
        self.assertIn("cp-status-pill--cleared", rt_html)
        self.assertIn("Total Transfer Bill", rt_html)
        self.assertIn('id="hrt-kpi-total" data-amount="80.0"', rt_html)
        self.assertIn('id="hrt-kpi-open">0<', rt_html)
        self.assertIn('id="hrt-kpi-outstanding">1<', rt_html)
        open_only = self.client.get("/hotel/room-transfer-invoices?status=open")
        self.assertNotIn("INV/26-27/4", open_only.get_data(as_text=True))
        generated_only = self.client.get(
            "/hotel/room-transfer-invoices?status=cancelled"
        )
        generated_html = generated_only.get_data(as_text=True)
        self.assertIn("INV/26-27/4", generated_html)
        self.assertIn(f"Invoice Generated ({fb_no})", generated_html)

    def test_invoice_ledger_settle_after_checkout(self):
        self._checkin_with_charges(advance=0)
        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "generate_invoice",
                "payment_splits": [],
            },
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))
        stay = gen.get_json()["room"]["stay"]
        inv_no = stay["invoiceNumber"]
        balance = stay["balanceAmount"]
        closed = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "checkout"},
        )
        self.assertEqual(closed.status_code, 200, closed.get_data(as_text=True))
        self.assertNotIn("stay", closed.get_json()["room"])

        settle = self.client.post(
            f"/hotel/invoice-ledger/api/{inv_no}/settle",
            json={
                "payment_splits": [{"method": "upi", "amount": balance}],
            },
        )
        self.assertEqual(settle.status_code, 200, settle.get_data(as_text=True))
        settled = settle.get_json()
        self.assertTrue(settled["ok"])
        self.assertEqual(settled["invoice"]["status"], "settled")
        self.assertLessEqual(float(settled["invoice"]["balance_amount"]), 0.009)

        api = self.client.get(f"/hotel/invoice-ledger/api/{inv_no}")
        self.assertEqual(api.status_code, 200)
        item = api.get_json()["invoice"]
        self.assertEqual(item["status"], "settled")

    def test_invoice_ledger_settle_selected_invoices(self):
        self._checkin_with_charges("room-101", advance=0)
        self._checkin_with_charges("room-102", advance=0)
        gen_a = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "payment_splits": []},
        )
        gen_b = self.client.put(
            "/hotel/api/rooms/room-102",
            json={"action": "generate_invoice", "payment_splits": []},
        )
        self.assertEqual(gen_a.status_code, 200, gen_a.get_data(as_text=True))
        self.assertEqual(gen_b.status_code, 200, gen_b.get_data(as_text=True))
        stay_a = gen_a.get_json()["room"]["stay"]
        stay_b = gen_b.get_json()["room"]["stay"]
        inv_a = stay_a["invoiceNumber"]
        inv_b = stay_b["invoiceNumber"]
        combined = round(
            float(stay_a["balanceAmount"]) + float(stay_b["balanceAmount"]), 2
        )
        self.assertNotEqual(inv_a, inv_b)
        self.assertGreater(combined, 0)

        page = self.client.get("/hotel/invoice-ledger")
        html = page.get_data(as_text=True)
        self.assertIn('id="hil-select-all"', html)
        self.assertIn('id="hil-settle-selected"', html)
        self.assertIn(inv_a, html)
        self.assertIn(inv_b, html)

        settle = self.client.post(
            "/hotel/invoice-ledger/api/settle-selected",
            json={
                "invoice_numbers": [inv_a, inv_b],
                "payment_splits": [{"method": "cash", "amount": combined}],
            },
        )
        self.assertEqual(settle.status_code, 200, settle.get_data(as_text=True))
        payload = settle.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["paid_count"], 2)
        self.assertEqual(payload["settled_count"], 2)
        numbers = {row["invoice_number"] for row in payload["invoices"]}
        self.assertEqual(numbers, {inv_a, inv_b})
        for row in payload["invoices"]:
            self.assertEqual(row["status"], "settled")
            self.assertLessEqual(float(row["balance_amount"]), 0.009)

        for inv_no in (inv_a, inv_b):
            api = self.client.get(f"/hotel/invoice-ledger/api/{inv_no}")
            self.assertEqual(api.status_code, 200)
            item = api.get_json()["invoice"]
            self.assertEqual(item["status"], "settled")

        again = self.client.post(
            "/hotel/invoice-ledger/api/settle-selected",
            json={
                "invoice_numbers": [inv_a],
                "payment_splits": [{"method": "cash", "amount": 1}],
            },
        )
        self.assertEqual(again.status_code, 400)

    def test_invoice_ledger_survives_checkout(self):
        self._checkin_with_charges(advance=0)
        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "generate_invoice",
                "payment": {"amount": 2000, "method": "cash"},
            },
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))
        stay = gen.get_json()["room"]["stay"]
        inv_no = stay["invoiceNumber"]
        balance = stay["balanceAmount"]
        paid = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "record_payment",
                "amount": balance,
                "method": "card",
            },
        )
        self.assertEqual(paid.status_code, 200, paid.get_data(as_text=True))
        closed = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "checkout"},
        )
        self.assertEqual(closed.status_code, 200, closed.get_data(as_text=True))
        self.assertNotIn("stay", closed.get_json()["room"])

        page = self.client.get("/hotel/invoice-ledger")
        self.assertEqual(page.status_code, 200)
        self.assertIn(inv_no, page.get_data(as_text=True))

        api = self.client.get(f"/hotel/invoice-ledger/api/{inv_no}")
        self.assertEqual(api.status_code, 200)
        payload = api.get_json()
        self.assertTrue(payload["ok"])
        self.assertIn("stay", payload["room"])
        self.assertEqual(payload["room"]["stay"]["invoiceNumber"], inv_no)
        self.assertEqual(payload["invoice"]["status"], "settled")
        self.assertLessEqual(float(payload["invoice"]["balance_amount"]), 0.009)

    def test_invoice_ledger_access_gate(self):
        from workspace_access import get_endpoint_dashboard_module

        self.assertEqual(
            get_endpoint_dashboard_module("hotel_invoice_ledger"), "hotel_rooms"
        )
        self.assertEqual(
            get_endpoint_dashboard_module("hotel_invoice_ledger_api"), "hotel_rooms"
        )
        self.assertEqual(
            get_endpoint_dashboard_module("hotel_invoice_ledger_export"), "hotel_rooms"
        )
        self.assertEqual(
            get_endpoint_dashboard_module("hotel_room_transfer_invoices"),
            "hotel_rooms",
        )
        self.assertEqual(
            get_endpoint_dashboard_module("hotel_room_transfer_invoices_export"),
            "hotel_rooms",
        )
        self.assertEqual(
            get_endpoint_dashboard_module("hotel_invoice_ledger_settle_api"),
            "hotel_rooms",
        )
        self.assertEqual(
            get_endpoint_dashboard_module("hotel_invoice_ledger_settle_selected_api"),
            "hotel_rooms",
        )
        self.assertEqual(
            get_endpoint_dashboard_module("hotel_invoice_ledger_cancel_api"),
            "hotel_rooms",
        )
        self.assertEqual(
            get_endpoint_dashboard_module("hotel_invoice_ledger_reopen_edit_api"),
            "hotel_rooms",
        )
        self.assertEqual(
            get_endpoint_dashboard_module("hotel_invoice_ledger_edit_page"),
            "hotel_rooms",
        )
        self.assertEqual(
            get_endpoint_dashboard_module("hotel_invoice_ledger_edit_api"),
            "hotel_rooms",
        )

        viewer = {
            "id": self.admin_id,
            "username": "posonly",
            "full_name": "POS Only",
            "is_admin": False,
            "is_active": True,
            "dashboard_access": {"point_of_sale"},
            "stores_access": set(),
            "sales_analytics_access": set(),
            "user_access": set(),
            "payroll_access": set(),
            "accounts_access": set(),
        }
        with mock.patch.object(self.app_mod, "get_current_user", return_value=viewer):
            denied_page = self.client.get("/hotel/invoice-ledger")
            denied_rt = self.client.get("/hotel/room-transfer-invoices")
            denied_rooms = self.client.get("/hotel/rooms")
            denied_api = self.client.get(
                "/hotel/invoice-ledger/api/HBE/RM/1/2025-26",
                headers={"X-Requested-With": "XMLHttpRequest"},
            )
        self.assertIn(denied_page.status_code, (302, 403))
        self.assertIn(denied_rt.status_code, (302, 403))
        self.assertIn(denied_rooms.status_code, (302, 403))
        self.assertEqual(denied_api.status_code, 403)

    def test_invoice_ledger_edit_cancel_require_separate_access(self):
        self._checkin_with_charges(advance=0)
        room = self._generate_stay_invoice()
        inv_no = room["stay"]["invoiceNumber"]

        page = self.client.get("/hotel/invoice-ledger?status=open")
        html = page.get_data(as_text=True)
        self.assertIn("hil-edit-btn", html)
        self.assertIn("hil-cancel-btn", html)
        self.assertIn(f'aria-label="Edit invoice {inv_no}"', html)

        hotel_only = {
            "id": self.admin_id,
            "username": "frontdesk",
            "full_name": "Front Desk",
            "is_admin": False,
            "is_active": True,
            "dashboard_access": {"hotel_rooms"},
            "stores_access": set(),
            "sales_analytics_access": set(),
            "user_access": set(),
            "payroll_access": set(),
            "accounts_access": set(),
        }
        with mock.patch.object(self.app_mod, "get_current_user", return_value=hotel_only):
            locked_page = self.client.get("/hotel/invoice-ledger?status=open")
            denied_edit = self.client.post(
                f"/hotel/invoice-ledger/api/{inv_no}/reopen-edit",
                json={},
                headers={"X-Requested-With": "XMLHttpRequest"},
            )
            denied_cancel = self.client.post(
                f"/hotel/invoice-ledger/api/{inv_no}/cancel",
                json={"reason": "Guest left"},
                headers={"X-Requested-With": "XMLHttpRequest"},
            )
        locked_html = locked_page.get_data(as_text=True)
        self.assertEqual(locked_page.status_code, 200)
        self.assertIn(inv_no, locked_html)
        self.assertNotIn("hil-edit-btn", locked_html)
        self.assertNotIn("hil-cancel-btn", locked_html)
        self.assertEqual(denied_edit.status_code, 403)
        self.assertIn("Edit Access", (denied_edit.get_json() or {}).get("error", ""))
        self.assertEqual(denied_cancel.status_code, 403)
        self.assertIn(
            "Cancellation", (denied_cancel.get_json() or {}).get("error", "")
        )

        cancel_only = dict(hotel_only)
        cancel_only["dashboard_access"] = {"hotel_rooms", "cancellation_access"}
        with mock.patch.object(self.app_mod, "get_current_user", return_value=cancel_only):
            cancel_page = self.client.get("/hotel/invoice-ledger?status=open")
            cancel_denied_edit = self.client.post(
                f"/hotel/invoice-ledger/api/{inv_no}/reopen-edit",
                json={},
                headers={"X-Requested-With": "XMLHttpRequest"},
            )
        cancel_html = cancel_page.get_data(as_text=True)
        self.assertNotIn("hil-edit-btn", cancel_html)
        self.assertIn("hil-cancel-btn", cancel_html)
        self.assertEqual(cancel_denied_edit.status_code, 403)

        edit_only = dict(hotel_only)
        edit_only["dashboard_access"] = {"hotel_rooms", "edit_access"}
        with mock.patch.object(self.app_mod, "get_current_user", return_value=edit_only):
            edit_page = self.client.get("/hotel/invoice-ledger?status=open")
            edit_denied_cancel = self.client.post(
                f"/hotel/invoice-ledger/api/{inv_no}/cancel",
                json={"reason": "Guest left"},
                headers={"X-Requested-With": "XMLHttpRequest"},
            )
        edit_html = edit_page.get_data(as_text=True)
        self.assertIn("hil-edit-btn", edit_html)
        self.assertNotIn("hil-cancel-btn", edit_html)
        self.assertEqual(edit_denied_cancel.status_code, 403)

        missing_reason = self.client.post(
            f"/hotel/invoice-ledger/api/{inv_no}/cancel",
            json={},
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(missing_reason.status_code, 400)
        self.assertIn("reason", (missing_reason.get_json() or {}).get("error", "").lower())

        reopen = self.client.post(
            f"/hotel/invoice-ledger/api/{inv_no}/reopen-edit",
            json={},
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(reopen.status_code, 200, reopen.get_data(as_text=True))
        reopened = reopen.get_json()
        self.assertTrue(reopened["ok"])
        self.assertTrue(reopened["room"]["stay"]["invoiceGenerated"])
        self.assertTrue(reopened["room"]["stay"]["invoiceEditOpen"])
        self.assertEqual(reopened["room"]["stay"]["invoiceNumber"], inv_no)
        self.assertIn(f"/hotel/invoice-ledger/{inv_no}/edit", reopened["edit_url"])
        self.assertNotIn("/hotel/rooms/room-101/invoice", reopened["edit_url"])

        edited = self.client.put(
            f"/hotel/invoice-ledger/api/{inv_no}/edit",
            json={"action": "update_charge", "chargeKey": "extra_bed", "amount": 800},
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(edited.status_code, 200, edited.get_data(as_text=True))
        self.assertEqual(edited.get_json()["room"]["stay"]["extraBedAmount"], 800)
        live_after_edit = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]["stay"]
        self.assertEqual(live_after_edit["extraBedAmount"], 800)

        regen = self.client.put(
            f"/hotel/invoice-ledger/api/{inv_no}/edit",
            json={"action": "generate_invoice", "payment_splits": []},
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(regen.status_code, 200, regen.get_data(as_text=True))
        regen_body = regen.get_json()
        self.assertFalse(regen_body.get("minted"))
        self.assertEqual(regen_body["room"]["stay"]["invoiceNumber"], inv_no)
        self.assertTrue(regen_body["room"]["stay"]["invoiceGenerated"])

        cancel = self.client.post(
            f"/hotel/invoice-ledger/api/{inv_no}/cancel",
            json={"reason": "Guest changed rooms"},
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(cancel.status_code, 200, cancel.get_data(as_text=True))
        cancelled = cancel.get_json()
        self.assertTrue(cancelled["ok"])
        self.assertEqual(cancelled["invoice"]["status"], "cancelled")
        self.assertEqual(cancelled["invoice"]["payment_mode_label"], "Cancelled")

        live = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]["stay"]
        self.assertFalse(live.get("invoiceGenerated"))
        self.assertFalse(live.get("invoiceNumber"))

        settle = self.client.post(
            f"/hotel/invoice-ledger/api/{inv_no}/settle",
            json={"payment_splits": [{"method": "cash", "amount": 100}]},
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(settle.status_code, 400)
        self.assertIn("cancelled", (settle.get_json() or {}).get("error", "").lower())

        again = self._generate_stay_invoice()
        self.assertNotEqual(again["stay"]["invoiceNumber"], inv_no)

        cancelled_page = self.client.get("/hotel/invoice-ledger?status=cancelled")
        cancelled_html = cancelled_page.get_data(as_text=True)
        self.assertIn(inv_no, cancelled_html)
        self.assertIn("Cancelled", cancelled_html)
        self.assertNotIn(f'aria-label="Edit invoice {inv_no}"', cancelled_html)

    def test_invoice_ledger_edit_after_checkout(self):
        self._checkin_with_charges(advance=0)
        room = self._generate_stay_invoice()
        inv_no = room["stay"]["invoiceNumber"]

        checkout = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "checkout"},
        )
        self.assertEqual(checkout.status_code, 200, checkout.get_data(as_text=True))
        checked_out = checkout.get_json()["room"]
        self.assertNotIn("stay", checked_out)

        reopen = self.client.post(
            f"/hotel/invoice-ledger/api/{inv_no}/reopen-edit",
            json={},
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(reopen.status_code, 200, reopen.get_data(as_text=True))
        reopened = reopen.get_json()
        self.assertTrue(reopened["ok"])
        self.assertTrue(reopened["room"]["stay"]["invoiceEditOpen"])
        self.assertEqual(reopened["room"]["stay"]["invoiceNumber"], inv_no)
        self.assertIn(f"/hotel/invoice-ledger/{inv_no}/edit", reopened["edit_url"])
        self.assertNotIn("/hotel/rooms/room-101/invoice", reopened["edit_url"])

        live_after_reopen = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        self.assertNotIn("stay", live_after_reopen)

        edited = self.client.put(
            f"/hotel/invoice-ledger/api/{inv_no}/edit",
            json={"action": "update_charge", "chargeKey": "extra_bed", "amount": 900},
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(edited.status_code, 200, edited.get_data(as_text=True))
        self.assertEqual(edited.get_json()["room"]["stay"]["extraBedAmount"], 900)

        live_still_empty = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        self.assertNotIn("stay", live_still_empty)

        regen = self.client.put(
            f"/hotel/invoice-ledger/api/{inv_no}/edit",
            json={"action": "generate_invoice", "payment_splits": []},
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(regen.status_code, 200, regen.get_data(as_text=True))
        regen_body = regen.get_json()
        self.assertFalse(regen_body.get("minted"))
        self.assertFalse((regen_body.get("room") or {}).get("stay", {}).get("invoiceEditOpen"))

        live_after_regen = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        self.assertNotIn("stay", live_after_regen)

        api = self.client.get(f"/hotel/invoice-ledger/api/{inv_no}")
        self.assertEqual(api.status_code, 200)
        invoice = api.get_json()["invoice"]
        self.assertEqual(invoice["invoice_number"], inv_no)
        self.assertEqual(invoice["status"], "open")
        self.assertGreater(float(invoice["estimated_total"]), 0)

    def test_invoice_ledger_edit_blocked_when_room_reoccupied(self):
        self._checkin_with_charges(advance=0)
        room = self._generate_stay_invoice()
        inv_no = room["stay"]["invoiceNumber"]

        checkout = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "checkout"},
        )
        self.assertEqual(checkout.status_code, 200, checkout.get_data(as_text=True))

        cleaned = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"status": "vacant"},
        )
        self.assertEqual(cleaned.status_code, 200, cleaned.get_data(as_text=True))

        check_in, check_out = self._stay_window(nights=1)
        checkin = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "New",
                    "lastName": "Guest",
                    "mobile": "9000000101",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 3200,
                    "totalRate": 3200,
                    "advancePaid": 0,
                },
            },
        )
        self.assertEqual(checkin.status_code, 200, checkin.get_data(as_text=True))

        reopen = self.client.post(
            f"/hotel/invoice-ledger/api/{inv_no}/reopen-edit",
            json={},
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(reopen.status_code, 400, reopen.get_data(as_text=True))
        self.assertIn("another guest", (reopen.get_json() or {}).get("error", "").lower())

    def test_room_invoice_folio_mutations_require_edit_access(self):
        self._checkin_with_charges(advance=0)

        page = self.client.get("/hotel/rooms/room-101/invoice?kind=hotel")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn('data-can-edit="1"', html)
        self.assertIn("Folio charges", html)

        hotel_only = {
            "id": self.admin_id,
            "username": "frontdesk",
            "full_name": "Front Desk",
            "is_admin": False,
            "is_active": True,
            "dashboard_access": {"hotel_rooms"},
            "stores_access": set(),
            "sales_analytics_access": set(),
            "user_access": set(),
            "payroll_access": set(),
            "accounts_access": set(),
        }
        with mock.patch.object(self.app_mod, "get_current_user", return_value=hotel_only):
            locked_page = self.client.get("/hotel/rooms/room-101/invoice?kind=hotel")
            denied_update = self.client.put(
                "/hotel/api/rooms/room-101",
                json={"action": "update_charge", "chargeKey": "extra_bed", "amount": 900},
            )
            denied_discount = self.client.put(
                "/hotel/api/rooms/room-101",
                json={
                    "action": "set_discount",
                    "discountType": "pct",
                    "discountValue": 5,
                },
            )
        locked_html = locked_page.get_data(as_text=True)
        self.assertEqual(locked_page.status_code, 200)
        self.assertIn('data-can-edit="0"', locked_html)
        self.assertIn("hri-generate", locked_html)
        self.assertIn("Generate Room Invoice", locked_html)
        self.assertEqual(denied_update.status_code, 403)
        self.assertIn("Edit Access", (denied_update.get_json() or {}).get("error", ""))
        self.assertEqual(denied_discount.status_code, 403)

        with_edit = dict(hotel_only)
        with_edit["dashboard_access"] = {"hotel_rooms", "edit_access"}
        with mock.patch.object(self.app_mod, "get_current_user", return_value=with_edit):
            allowed_page = self.client.get("/hotel/rooms/room-101/invoice?kind=hotel")
            allowed_update = self.client.put(
                "/hotel/api/rooms/room-101",
                json={"action": "update_charge", "chargeKey": "extra_bed", "amount": 900},
            )
        self.assertIn('data-can-edit="1"', allowed_page.get_data(as_text=True))
        self.assertEqual(allowed_update.status_code, 200, allowed_update.get_data(as_text=True))
        self.assertEqual(allowed_update.get_json()["room"]["stay"]["extraBedAmount"], 900)

        # Generate is allowed without Edit Access; folio edit/delete are not.
        with mock.patch.object(self.app_mod, "get_current_user", return_value=hotel_only):
            allowed_generate = self.client.put(
                "/hotel/api/rooms/room-101",
                json={"action": "generate_invoice", "payment_splits": []},
            )
            denied_delete_after = self.client.put(
                "/hotel/api/rooms/room-101",
                json={"action": "delete_charge", "chargeKey": "extra_bed"},
            )
        self.assertEqual(allowed_generate.status_code, 200, allowed_generate.get_data(as_text=True))
        self.assertTrue((allowed_generate.get_json() or {}).get("ok"))
        stay = ((allowed_generate.get_json() or {}).get("room") or {}).get("stay") or {}
        self.assertTrue(stay.get("invoiceNumber") or stay.get("invoice_number"))
        self.assertEqual(denied_delete_after.status_code, 403)

    def test_invoice_ledger_edit_page_requires_open_session(self):
        self._checkin_with_charges(advance=0)
        room = self._generate_stay_invoice()
        inv_no = room["stay"]["invoiceNumber"]

        blocked = self.client.get(f"/hotel/invoice-ledger/{inv_no}/edit")
        self.assertEqual(blocked.status_code, 403)

        reopen = self.client.post(
            f"/hotel/invoice-ledger/api/{inv_no}/reopen-edit",
            json={},
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(reopen.status_code, 200, reopen.get_data(as_text=True))

        page = self.client.get(f"/hotel/invoice-ledger/{inv_no}/edit")
        self.assertEqual(page.status_code, 200, page.get_data(as_text=True))
        html = page.get_data(as_text=True)
        self.assertIn("data-ledger-edit=\"1\"", html)
        self.assertIn("Edit Invoice", html)
        self.assertIn("/hotel/invoice-ledger", html)

    def test_merge_rooms_combines_billing_onto_primary(self):
        self._checkin_with_charges("room-101", advance=500)
        self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Asha",
                    "lastName": "Nair",
                    "mobile": "9000000011",
                    "checkInDate": "2026-07-29",
                    "nights": 1,
                    "roomRate": 3000,
                    "totalRate": 3000,
                    "advancePaid": 200,
                },
            },
        )
        merged = self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-102",
                "toRoomId": "room-101",
            },
        )
        self.assertEqual(merged.status_code, 200, merged.get_data(as_text=True))
        body = merged.get_json()
        primary = body["primaryRoom"]
        member = body["memberRoom"]
        self.assertTrue(primary["isMergePrimary"])
        self.assertTrue(member["isMergeMember"])
        self.assertEqual(member["stay"]["billingRoomId"], "room-101")
        self.assertEqual(member["stay"]["mergeRole"], "member")
        # Member room charge + advances moved onto primary folio/advance.
        folio_labels = [f.get("label") for f in primary["stay"]["folioCharges"]]
        self.assertTrue(any("Room 102" in (label or "") for label in folio_labels))
        absorb_lines = [
            f
            for f in primary["stay"]["folioCharges"]
            if f.get("source") == "room_merge"
            and str(f.get("sourceRoomId") or "") == "room-102"
        ]
        self.assertEqual(len(absorb_lines), 1, primary["stay"]["folioCharges"])
        # No duplicate auto rate line alongside the absorb for the same member.
        rate_dupes = [
            f
            for f in primary["stay"]["folioCharges"]
            if f.get("source") == "merged_room_rate"
            and str(f.get("sourceRoomId") or "") == "room-102"
        ]
        self.assertEqual(rate_dupes, [])
        self.assertGreaterEqual(float(primary["stay"]["advancePaid"]), 700)
        self.assertGreater(float(primary["stay"]["estimatedTotal"]), 5750)

        blocked = self.client.put(
            "/hotel/api/rooms/room-102",
            json={"action": "generate_invoice", "amount": 0},
        )
        self.assertEqual(blocked.status_code, 400)
        self.assertIn("primary", blocked.get_json().get("error", "").lower())

        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "amount": 0},
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))
        gen_stay = gen.get_json()["room"]["stay"]
        self.assertTrue(gen_stay["invoiceNumber"])
        self.assertEqual(gen_stay.get("mergeRoomLabel"), "101 + 102")
        self.assertEqual(gen_stay.get("mergeRoomNumbers"), ["101", "102"])

        # Sibling merge members are stamped billed via primary — no separate Generate.
        member_after = self.client.get("/hotel/api/rooms/room-102").get_json()["room"]
        self.assertTrue(member_after.get("isMergeMember"))
        mstay = member_after.get("stay") or {}
        self.assertTrue(mstay.get("billedInvoiceGenerated"), mstay)
        self.assertEqual(mstay.get("billedInvoiceNumber"), gen_stay["invoiceNumber"])
        # Overlay may also mirror live invoice flags from primary for display.
        self.assertTrue(mstay.get("invoiceGenerated") or mstay.get("invoiceNumber"))
        blocked_again = self.client.put(
            "/hotel/api/rooms/room-102",
            json={"action": "generate_invoice", "amount": 0},
        )
        self.assertEqual(blocked_again.status_code, 400)
        self.assertIn("primary", blocked_again.get_json().get("error", "").lower())
        # Single shared invoice remains on the primary (ledger sync target).
        primary_after = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        self.assertEqual(
            primary_after["stay"]["invoiceNumber"], gen_stay["invoiceNumber"]
        )
        self.assertTrue(primary_after["stay"].get("invoiceGenerated"))

    def test_merge_promotes_guest_onto_vacant_primary(self):
        """Occupied member guest becomes the primary bill guest for board/invoice."""
        self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Priya",
                    "lastName": "Menon",
                    "mobile": "9000000099",
                    "checkInDate": "2026-07-29",
                    "nights": 1,
                    "roomRate": 2500,
                    "advancePaid": 0,
                },
            },
        )
        merged = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-101",
                "toRoomId": "room-106",
            },
        )
        self.assertEqual(merged.status_code, 200, merged.get_data(as_text=True))
        primary = merged.get_json()["primaryRoom"]
        self.assertEqual(primary["id"], "room-106")
        stay = primary["stay"]
        self.assertIn("Priya", stay.get("guestName") or stay.get("firstName") or "")
        self.assertEqual(stay.get("mobile"), "9000000099")

        member = merged.get_json()["memberRoom"]
        mstay = member["stay"]
        self.assertIn("Priya", mstay.get("guestName") or mstay.get("firstName") or "")
        self.assertEqual(mstay.get("mobile"), "9000000099")

        # Both detail APIs expose the same guest; member also sees shared bill totals.
        primary_get = self.client.get("/hotel/api/rooms/room-106").get_json()["room"]
        member_get = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        p_name = primary_get["stay"].get("guestName") or primary_get["stay"].get("firstName")
        m_name = member_get["stay"].get("guestName") or member_get["stay"].get("firstName")
        self.assertIn("Priya", p_name or "")
        self.assertEqual(p_name, m_name)
        self.assertEqual(
            primary_get["stay"].get("mobile"), member_get["stay"].get("mobile")
        )
        self.assertGreater(float(primary_get["stay"].get("estimatedTotal") or 0), 0)
        self.assertEqual(
            float(primary_get["stay"].get("estimatedTotal") or 0),
            float(member_get["stay"].get("estimatedTotal") or 0),
        )

    def test_checkin_on_merge_primary_replicates_to_member(self):
        """After vacant-vacant merge, check-in on primary copies guest onto the member."""
        check_in, check_out = self._stay_window(nights=2)
        merged = self.client.put(
            "/hotel/api/rooms/room-105",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-105",
                "toRoomId": "room-106",
            },
        )
        self.assertEqual(merged.status_code, 200, merged.get_data(as_text=True))
        checked = self.client.put(
            "/hotel/api/rooms/room-106",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Karan",
                    "lastName": "Singh",
                    "mobile": "9000000088",
                    "email": "karan@example.com",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 2,
                    "roomRate": 3000,
                    "advancePaid": 500,
                },
            },
        )
        self.assertEqual(checked.status_code, 200, checked.get_data(as_text=True))
        primary = self.client.get("/hotel/api/rooms/room-106").get_json()["room"]
        member = self.client.get("/hotel/api/rooms/room-105").get_json()["room"]
        self.assertIn("Karan", primary["stay"].get("guestName") or "")
        self.assertIn("Karan", member["stay"].get("guestName") or "")
        self.assertEqual(primary["stay"].get("mobile"), "9000000088")
        self.assertEqual(member["stay"].get("mobile"), "9000000088")
        self.assertEqual(primary["stay"].get("email"), member["stay"].get("email"))
        self.assertGreater(float(primary["stay"].get("estimatedTotal") or 0), 0)
        self.assertEqual(
            float(primary["stay"].get("estimatedTotal") or 0),
            float(member["stay"].get("estimatedTotal") or 0),
        )
        # Member tariff is typed at check-in; it is not filled from Hotel Settings.
        folio = primary["stay"].get("folioCharges") or []
        rate_lines = [f for f in folio if f.get("source") == "merged_room_rate"]
        self.assertEqual(rate_lines, [])
        # Primary ₹3000 × 2 nights
        self.assertAlmostEqual(
            float(primary["stay"].get("estimatedTotal") or 0), 6000.0, places=2
        )
        self.assertEqual(primary["status"], "occupied")
        self.assertEqual(member["status"], "vacant")

    def test_merged_occupied_rooms_keep_distinct_guest_names(self):
        """Each occupied merge peer keeps its own guest; folio stays on primary."""
        check_in, check_out = self._stay_window(nights=1)
        self.assertEqual(
            self.client.put(
                "/hotel/api/rooms/room-201",
                json={
                    "action": "checkin",
                    "stay": {
                        "firstName": "Anita",
                        "lastName": "Rao",
                        "mobile": "9000000301",
                        "checkInDate": check_in,
                        "checkOutDate": check_out,
                        "nights": 1,
                        "roomRate": 2500,
                    },
                },
            ).status_code,
            200,
        )
        self.assertEqual(
            self.client.put(
                "/hotel/api/rooms/room-202",
                json={
                    "action": "checkin",
                    "stay": {
                        "firstName": "Vikram",
                        "lastName": "Shah",
                        "mobile": "9000000302",
                        "checkInDate": check_in,
                        "checkOutDate": check_out,
                        "nights": 1,
                        "roomRate": 2600,
                    },
                },
            ).status_code,
            200,
        )
        merged = self.client.put(
            "/hotel/api/rooms/room-202",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-202",
                "toRoomId": "room-201",
            },
        )
        self.assertEqual(merged.status_code, 200, merged.get_data(as_text=True))
        primary = self.client.get("/hotel/api/rooms/room-201").get_json()["room"]
        member = self.client.get("/hotel/api/rooms/room-202").get_json()["room"]
        self.assertIn("Anita", primary["stay"].get("guestName") or "")
        self.assertIn("Vikram", member["stay"].get("guestName") or "")
        self.assertEqual(primary["stay"].get("mobile"), "9000000301")
        self.assertEqual(member["stay"].get("mobile"), "9000000302")
        layout = self.client.get("/hotel/api/rooms").get_json()
        rooms = {r["id"]: r for r in layout.get("rooms") or []}
        p_name = rooms["room-201"]["stay"].get("guestName") or rooms["room-201"][
            "stay"
        ].get("firstName")
        m_name = rooms["room-202"]["stay"].get("guestName") or rooms["room-202"][
            "stay"
        ].get("firstName")
        self.assertIn("Anita", p_name or "")
        self.assertIn("Vikram", m_name or "")
        self.assertGreater(float(primary["stay"].get("estimatedTotal") or 0), 0)
        self.assertEqual(
            float(primary["stay"].get("estimatedTotal") or 0),
            float(member["stay"].get("estimatedTotal") or 0),
        )
        member_html = self.client.get("/hotel/rooms/room-202").get_data(as_text=True)
        self.assertIn("Vikram", member_html)
        name_bit = member_html.split('id="hrd-guest-name"', 1)[-1][:120]
        self.assertIn("Vikram", name_bit)
        self.assertNotIn("Anita", name_bit)

    def test_merge_member_agency_comes_from_primary(self):
        """Merged member guest card shows the billing primary's agency."""
        check_in, check_out = self._stay_window(nights=1)
        self.assertEqual(
            self.client.put(
                "/hotel/api/rooms/room-201",
                json={
                    "action": "checkin",
                    "stay": {
                        "firstName": "Anita",
                        "lastName": "Rao",
                        "mobile": "9000000301",
                        "checkInDate": check_in,
                        "checkOutDate": check_out,
                        "nights": 1,
                        "roomRate": 2500,
                        "agencyName": "ATPI India Pvt. Ltd",
                        "agencyGst": "27AABCU9603R1ZM",
                        "agencyAddress": "Bhandup West, Mumbai",
                        "agencyBilling": True,
                        "invoiceTo": "ATPI India Pvt. Ltd",
                    },
                },
            ).status_code,
            200,
        )
        self.assertEqual(
            self.client.put(
                "/hotel/api/rooms/room-202",
                json={
                    "action": "checkin",
                    "stay": {
                        "firstName": "Vikram",
                        "lastName": "Shah",
                        "mobile": "9000000302",
                        "checkInDate": check_in,
                        "checkOutDate": check_out,
                        "nights": 1,
                        "roomRate": 2600,
                    },
                },
            ).status_code,
            200,
        )
        merged = self.client.put(
            "/hotel/api/rooms/room-202",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-202",
                "toRoomId": "room-201",
            },
        )
        self.assertEqual(merged.status_code, 200, merged.get_data(as_text=True))
        member = self.client.get("/hotel/api/rooms/room-202").get_json()["room"]
        stay = member.get("stay") or {}
        self.assertIn("Vikram", stay.get("guestName") or "")
        self.assertEqual(stay.get("agencyName"), "ATPI India Pvt. Ltd")
        self.assertEqual(stay.get("agencyGst"), "27AABCU9603R1ZM")
        self.assertEqual(stay.get("agencyAddress"), "Bhandup West, Mumbai")
        html = self.client.get("/hotel/rooms/room-202").get_data(as_text=True)
        self.assertIn("ATPI India Pvt. Ltd", html)
        self.assertIn("27AABCU9603R1ZM", html)
        self.assertNotIn('id="hrd-agency-card" hidden', html)

    def test_checkin_on_vacant_merge_member_keeps_own_guest(self):
        """Check-in on a vacant merge member is not overwritten by the primary guest."""
        check_in, check_out = self._stay_window(nights=2)
        merged = self.client.put(
            "/hotel/api/rooms/room-105",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-105",
                "toRoomId": "room-106",
            },
        )
        self.assertEqual(merged.status_code, 200, merged.get_data(as_text=True))
        self.assertEqual(
            self.client.put(
                "/hotel/api/rooms/room-106",
                json={
                    "action": "checkin",
                    "stay": {
                        "firstName": "Karan",
                        "lastName": "Singh",
                        "mobile": "9000000088",
                        "checkInDate": check_in,
                        "checkOutDate": check_out,
                        "nights": 2,
                        "roomRate": 3000,
                    },
                },
            ).status_code,
            200,
        )
        checked = self.client.put(
            "/hotel/api/rooms/room-105",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Meera",
                    "lastName": "Nair",
                    "mobile": "9000000089",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 2,
                    "roomRate": 3500,
                },
            },
        )
        self.assertEqual(checked.status_code, 200, checked.get_data(as_text=True))
        primary = self.client.get("/hotel/api/rooms/room-106").get_json()["room"]
        member = self.client.get("/hotel/api/rooms/room-105").get_json()["room"]
        self.assertIn("Karan", primary["stay"].get("guestName") or "")
        self.assertIn("Meera", member["stay"].get("guestName") or "")
        self.assertEqual(primary["stay"].get("mobile"), "9000000088")
        self.assertEqual(member["stay"].get("mobile"), "9000000089")
        self.assertEqual(member["status"], "occupied")
        self.assertEqual(primary["status"], "occupied")

    def test_checkin_on_reserved_merge_member_occupies_only_that_room(self):
        """Check-in on a reserved merge member occupies that room only."""
        check_in, check_out = self._stay_window(nights=1)
        stay = {
            "guestName": "Manoj Vijayan",
            "firstName": "Manoj",
            "lastName": "Vijayan",
            "mobile": "9876500101",
        }
        for room_id in ("room-101", "room-103"):
            reserved = self.client.put(
                f"/hotel/api/rooms/{room_id}",
                json={
                    "action": "reserve",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "stay": stay,
                },
            )
            self.assertEqual(reserved.status_code, 200, reserved.get_data(as_text=True))
            self.assertEqual(reserved.get_json()["room"]["status"], "reserved")
        merged = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-101",
                "toRoomId": "room-103",
            },
        )
        self.assertEqual(merged.status_code, 200, merged.get_data(as_text=True))
        checked = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Manoj",
                    "lastName": "Vijayan",
                    "mobile": "9876500101",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 4500,
                },
            },
        )
        self.assertEqual(checked.status_code, 200, checked.get_data(as_text=True))
        member = checked.get_json()["room"]
        self.assertEqual(member["id"], "room-101")
        self.assertEqual(member["status"], "occupied", member)
        primary = self.client.get("/hotel/api/rooms/room-103").get_json()["room"]
        self.assertEqual(primary["status"], "reserved", primary)
        self.assertTrue(member.get("stay", {}).get("checkedInAt"))
        self.assertFalse(primary.get("stay", {}).get("checkedInAt"))

        primary_in = self.client.put(
            "/hotel/api/rooms/room-103",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Manoj",
                    "lastName": "Vijayan",
                    "mobile": "9876500101",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 4500,
                },
            },
        )
        self.assertEqual(primary_in.status_code, 200, primary_in.get_data(as_text=True))
        primary = primary_in.get_json()["room"]
        self.assertEqual(primary["status"], "occupied")
        member = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        self.assertEqual(member["status"], "occupied")

        closed = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "checkout"},
        )
        self.assertEqual(closed.status_code, 200, closed.get_data(as_text=True))
        self.assertEqual(closed.get_json()["room"]["status"], "dirty")
        still_primary = self.client.get("/hotel/api/rooms/room-103").get_json()["room"]
        self.assertEqual(still_primary["status"], "occupied")
        self.assertTrue(still_primary.get("isMergePrimary") or not still_primary.get("mergeGroupId"))

    def test_merged_checkin_uses_submitted_room_rates(self):
        """Suite primary + Deluxe member bill the typed stay prices, not settings tariffs."""
        check_in, check_out = self._stay_window(nights=1)
        merged = self.client.put(
            "/hotel/api/rooms/room-306",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-306",
                "toRoomId": "room-307",
            },
        )
        self.assertEqual(merged.status_code, 200, merged.get_data(as_text=True))
        checked = self.client.put(
            "/hotel/api/rooms/room-307",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Asha",
                    "lastName": "Nair",
                    "mobile": "9000000306",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 7500,
                    "mergeRoomRates": [
                        {
                            "roomId": "room-307",
                            "number": "307",
                            "roomType": "premium_suite_tub",
                            "ratePlan": "EP",
                            "roomRate": 7500,
                            "isPrimary": True,
                        },
                        {
                            "roomId": "room-306",
                            "number": "306",
                            "roomType": "premium_deluxe_balcony",
                            "ratePlan": "EP",
                            "roomRate": 4500,
                            "isPrimary": False,
                        },
                    ],
                    "advancePaid": 0,
                },
            },
        )
        self.assertEqual(checked.status_code, 200, checked.get_data(as_text=True))
        primary = self.client.get("/hotel/api/rooms/room-307").get_json()["room"]
        folio = primary["stay"].get("folioCharges") or []
        rate_lines = [f for f in folio if f.get("source") == "merged_room_rate"]
        self.assertEqual(len(rate_lines), 1, folio)
        self.assertIn("306", rate_lines[0].get("label") or "")
        self.assertEqual(float(rate_lines[0].get("amount") or 0), 4500.0)
        rates = primary["stay"].get("mergeRoomRates") or []
        by_num = {str(r.get("number") or ""): r for r in rates}
        self.assertEqual(float(by_num["307"]["roomRate"]), 7500.0)
        self.assertEqual(float(by_num["306"]["roomRate"]), 4500.0)
        # ₹7500 + ₹4500 = ₹12,000 (tax-inclusive)
        self.assertAlmostEqual(
            float(primary["stay"].get("estimatedTotal") or 0), 12000.0, places=2
        )

    def test_merged_checkin_keeps_manual_member_rate(self):
        """Typed mergeRoomRates are billed as entered, even if they match the primary."""
        check_in, check_out = self._stay_window(nights=1)
        merged = self.client.put(
            "/hotel/api/rooms/room-306",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-306",
                "toRoomId": "room-307",
            },
        )
        self.assertEqual(merged.status_code, 200, merged.get_data(as_text=True))
        checked = self.client.put(
            "/hotel/api/rooms/room-307",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Dev",
                    "lastName": "Iyer",
                    "mobile": "9000000307",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 7500,
                    "mergeRoomRates": [
                        {
                            "roomId": "room-307",
                            "number": "307",
                            "roomRate": 7500,
                            "isPrimary": True,
                        },
                        {
                            "roomId": "room-306",
                            "number": "306",
                            "roomRate": 7500,
                            "isPrimary": False,
                        },
                    ],
                    "advancePaid": 0,
                },
            },
        )
        self.assertEqual(checked.status_code, 200, checked.get_data(as_text=True))
        primary = self.client.get("/hotel/api/rooms/room-307").get_json()["room"]
        rate_lines = [
            f
            for f in (primary["stay"].get("folioCharges") or [])
            if f.get("source") == "merged_room_rate"
        ]
        self.assertEqual(len(rate_lines), 1, primary["stay"].get("folioCharges"))
        self.assertEqual(float(rate_lines[0].get("amount") or 0), 7500.0)
        by_num = {
            str(r.get("number") or ""): r
            for r in (primary["stay"].get("mergeRoomRates") or [])
        }
        self.assertEqual(float(by_num["306"]["roomRate"]), 7500.0)
        self.assertAlmostEqual(
            float(primary["stay"].get("estimatedTotal") or 0), 15000.0, places=2
        )

    def test_member_rate_edit_updates_absorb_folio_and_total(self):
        """Editing a merged member tariff must refresh the primary absorb line."""
        check_in, check_out = self._stay_window(nights=1)
        self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Pri",
                    "lastName": "Mary",
                    "mobile": "9000000207",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 12600,
                    "advancePaid": 0,
                },
            },
        )
        self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Mem",
                    "lastName": "Ber",
                    "mobile": "9000000307",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 4200,
                    "advancePaid": 0,
                },
            },
        )
        merged = self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-102",
                "toRoomId": "room-101",
            },
        )
        self.assertEqual(merged.status_code, 200, merged.get_data(as_text=True))
        primary = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        absorb = [
            f
            for f in (primary["stay"].get("folioCharges") or [])
            if f.get("source") == "room_merge"
            and str(f.get("sourceRoomId") or "") == "room-102"
        ]
        self.assertEqual(len(absorb), 1, primary["stay"].get("folioCharges"))
        self.assertAlmostEqual(float(absorb[0].get("amount") or 0), 4200.0, places=2)

        # Member edit form typically sends only this room's rate row.
        edited = self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Mem",
                    "lastName": "Ber",
                    "mobile": "9000000307",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 1,
                    "ratePlan": "MAP",
                    "mergeRoomRates": [
                        {
                            "roomId": "room-102",
                            "number": "102",
                            "roomType": "premium_without_balcony",
                            "ratePlan": "MAP",
                            "roomRate": 1,
                            "isPrimary": False,
                            "nightlyRates": [
                                {"date": check_in, "roomRate": 1, "ratePlan": "MAP"}
                            ],
                        }
                    ],
                    "nightlyRates": [
                        {"date": check_in, "roomRate": 1, "ratePlan": "MAP"}
                    ],
                    "advancePaid": 0,
                },
            },
        )
        self.assertEqual(edited.status_code, 200, edited.get_data(as_text=True))

        primary_after = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        absorb_after = [
            f
            for f in (primary_after["stay"].get("folioCharges") or [])
            if f.get("source") == "room_merge"
            and str(f.get("sourceRoomId") or "") == "room-102"
        ]
        self.assertEqual(len(absorb_after), 1, primary_after["stay"].get("folioCharges"))
        self.assertAlmostEqual(float(absorb_after[0].get("amount") or 0), 1.0, places=2)
        by_num = {
            str(r.get("number") or ""): r
            for r in (primary_after["stay"].get("mergeRoomRates") or [])
        }
        self.assertAlmostEqual(float(by_num["102"]["roomRate"]), 1.0, places=2)
        self.assertAlmostEqual(
            float(primary_after["stay"].get("estimatedTotal") or 0), 12601.0, places=2
        )

    def test_merge_estimated_total_parity_primary_and_member_overlay(self):
        """Primary and member overlay must share folio + estimatedTotal after merge."""
        check_in, check_out = self._stay_window(nights=1)
        self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Pri",
                    "lastName": "Mary",
                    "mobile": "9000000208",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 4200,
                    "advancePaid": 0,
                },
            },
        )
        self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Mem",
                    "lastName": "Ber",
                    "mobile": "9000000308",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 4200,
                    "advancePaid": 0,
                },
            },
        )
        merged = self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-102",
                "toRoomId": "room-101",
            },
        )
        self.assertEqual(merged.status_code, 200, merged.get_data(as_text=True))

        primary = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        member = self.client.get("/hotel/api/rooms/room-102").get_json()["room"]
        pstay = primary["stay"]
        mstay = member["stay"]

        absorb = [
            f
            for f in (pstay.get("folioCharges") or [])
            if f.get("source") == "room_merge"
            and str(f.get("sourceRoomId") or "") == "room-102"
        ]
        self.assertEqual(len(absorb), 1, pstay.get("folioCharges"))
        absorb_amt = float(absorb[0].get("amount") or 0)
        primary_rate = float(pstay.get("roomRate") or 0)
        nights = max(1, int(float(pstay.get("nights") or 1)))
        expected = round(primary_rate * nights + absorb_amt, 2)
        self.assertAlmostEqual(float(pstay.get("estimatedTotal") or 0), expected, places=2)

        self.assertEqual(mstay.get("mergeRole"), "member")
        self.assertEqual(mstay.get("billingRoomId"), "room-101")
        self.assertAlmostEqual(
            float(mstay.get("estimatedTotal") or 0),
            float(pstay.get("estimatedTotal") or 0),
            places=2,
        )
        self.assertEqual(
            [
                (f.get("source"), f.get("sourceRoomId"), float(f.get("amount") or 0))
                for f in (mstay.get("folioCharges") or [])
                if str(f.get("source") or "") in ("room_merge", "merged_room_rate")
            ],
            [
                (f.get("source"), f.get("sourceRoomId"), float(f.get("amount") or 0))
                for f in (pstay.get("folioCharges") or [])
                if str(f.get("source") or "") in ("room_merge", "merged_room_rate")
            ],
        )
        # FE contract: member overlay includes absorb labels for Estimated Charges.
        absorb_labels = [
            str(f.get("label") or "")
            for f in (mstay.get("folioCharges") or [])
            if f.get("source") == "room_merge"
        ]
        self.assertTrue(any("102" in label for label in absorb_labels), absorb_labels)

    def test_merge_allows_any_rooms_without_stay(self):
        """Vacant / status-only rooms can join a billing merge group."""
        merged = self.client.put(
            "/hotel/api/rooms/room-105",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-105",
                "toRoomId": "room-106",
            },
        )
        self.assertEqual(merged.status_code, 200, merged.get_data(as_text=True))
        body = merged.get_json()
        primary = body["primaryRoom"]
        member = body["memberRoom"]
        self.assertEqual(primary["id"], "room-106")
        self.assertEqual(member["id"], "room-105")
        self.assertTrue(primary["isMergePrimary"])
        self.assertTrue(member["isMergeMember"])
        # Vacant rooms stay vacant until a guest is checked in on the bill.
        self.assertEqual(primary["status"], "vacant")
        self.assertEqual(member["status"], "vacant")
        self.assertEqual(member["stay"]["billingRoomId"], "room-106")
        self.assertEqual(member["stay"]["mergeRole"], "member")
        self.assertEqual(primary["stay"]["mergeRole"], "primary")
        self.assertTrue(primary.get("mergeGroupId"))
        self.assertEqual(primary["mergeGroupId"], member["mergeGroupId"])

    def test_merge_board_tile_exposes_guest_for_hover(self):
        """Merged primary board payload carries guest so the hover card can render."""
        self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Neha",
                    "lastName": "Iyer",
                    "mobile": "9000000077",
                    "checkInDate": "2026-08-01",
                    "nights": 2,
                    "roomRate": 2800,
                    "adults": 2,
                    "advancePaid": 0,
                },
            },
        )
        self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-101",
                "toRoomId": "room-106",
            },
        )
        layout = self.client.get("/hotel/api/rooms").get_json()
        rooms = layout.get("rooms") or []
        primary = next((r for r in rooms if r.get("id") == "room-106"), None)
        self.assertIsNotNone(primary)
        self.assertTrue(primary.get("isMergePrimary"))
        self.assertEqual(primary.get("status"), "vacant")
        member = next((r for r in rooms if r.get("id") == "room-101"), None)
        self.assertIsNotNone(member)
        self.assertEqual(member.get("status"), "occupied")
        stay = primary.get("stay") or {}
        self.assertIn("Neha", stay.get("guestName") or stay.get("firstName") or "")
        self.assertEqual(stay.get("mobile"), "9000000077")
        self.assertTrue(stay.get("checkInDate"))

    def test_unmerge_splits_folio_onto_member(self):
        self._checkin_with_charges("room-101", advance=0)
        self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Asha",
                    "lastName": "Nair",
                    "mobile": "9000000011",
                    "checkInDate": "2026-07-29",
                    "nights": 1,
                    "roomRate": 2000,
                    "advancePaid": 0,
                },
            },
        )
        self.client.put(
            "/hotel/api/rooms/room-102",
            json={"action": "merge_rooms", "fromRoomId": "room-102", "toRoomId": "room-101"},
        )
        before = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        folio_count = len(before["stay"]["folioCharges"])
        self.assertGreater(folio_count, 0)
        self.assertTrue(
            any(
                str(f.get("sourceRoomId") or "") == "room-102"
                for f in before["stay"]["folioCharges"]
            )
        )

        unmerged = self.client.put(
            "/hotel/api/rooms/room-102",
            json={"action": "unmerge_rooms", "scope": "one"},
        )
        self.assertEqual(unmerged.status_code, 200, unmerged.get_data(as_text=True))
        member = unmerged.get_json()["room"]
        self.assertFalse(member.get("isMergeMember"))
        self.assertFalse(member.get("mergeGroupId"))
        self.assertEqual(member["stay"].get("billingRoomId") or "", "")
        self.assertTrue(member["stay"].get("independentBilling"))
        self.assertGreater(float(member["stay"].get("roomRate") or 0), 0)
        self.assertGreater(float(member["stay"].get("estimatedTotal") or 0), 0)

        primary = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        self.assertFalse(
            any(
                str(f.get("sourceRoomId") or "") == "room-102"
                for f in primary["stay"].get("folioCharges") or []
            )
        )
        self.assertLess(len(primary["stay"]["folioCharges"]), folio_count)
        self.assertTrue(primary["stay"].get("independentBilling"))
        self.assertFalse(primary.get("mergeGroupId"))
        for stay in (member["stay"], primary["stay"]):
            merge_folio = [
                f
                for f in stay.get("folioCharges") or []
                if str(f.get("source") or "") in ("room_merge", "merged_room_rate")
            ]
            self.assertEqual(merge_folio, [])

    def test_independent_stay_strips_stale_merge_folio(self):
        """Unmerged rooms must not keep other rooms' absorb lines on Estimated Charges."""
        check_in, check_out = self._stay_window(nights=1)
        stay = db_mod._normalize_hotel_room_stay(
            {
                "firstName": "Asha",
                "lastName": "Nair",
                "checkInDate": check_in,
                "checkOutDate": check_out,
                "nights": 1,
                "roomRate": 5000,
                "independentBilling": True,
                "folioCharges": [
                    {
                        "kind": "other",
                        "label": "Room 102 — stay charges",
                        "amount": 5000,
                        "source": "room_merge",
                        "sourceRoomId": "room-102",
                        "sourceRoomNumber": "102",
                    },
                    {
                        "kind": "other",
                        "label": "Room 103 — stay charges",
                        "amount": 5000,
                        "source": "room_merge",
                        "sourceRoomId": "room-103",
                        "sourceRoomNumber": "103",
                    },
                    {
                        "kind": "restaurant_room_transfer",
                        "label": "Dinner",
                        "amount": 800,
                        "source": "pos",
                    },
                ],
            }
        )
        sources = [f.get("source") for f in stay.get("folioCharges") or []]
        self.assertNotIn("room_merge", sources)
        self.assertNotIn("merged_room_rate", sources)
        self.assertEqual(len(stay["folioCharges"]), 1)
        self.assertEqual(stay["folioCharges"][0]["kind"], "restaurant_room_transfer")
        self.assertEqual(float(stay["estimatedTotal"]), 5000.0)
        self.assertEqual(float(stay["fbTransferTotal"]), 800.0)
        self.assertEqual(float(stay["combinedBalanceDue"]), 5800.0)

    def test_fb_invoice_balance_allocation_uses_unsettled_lines_only(self):
        """Unpaid F&B due on an FBE must not include settled or pending-untyped amounts."""
        stay = {
            "fbTransferBalance": 526.0,
            "invoiceHistory": [
                {
                    "kind": "fb",
                    "invoiceNumber": "FBE/26-27/00001",
                    "estimatedTotal": 2216.0,
                }
            ],
            "folioCharges": [
                {
                    "kind": "restaurant_room_transfer",
                    "amount": 2216.0,
                    "invoicedInvoiceNumber": "FBE/26-27/00001",
                    "settled": True,
                    "source": "pos",
                },
                {
                    "kind": "restaurant_room_transfer",
                    "amount": 526.0,
                    "source": "pos",
                    "settled": False,
                },
            ],
        }
        allocated = db_mod._hotel_allocate_fb_invoice_balances(stay)
        self.assertEqual(allocated.get("FBE/26-27/00001"), 0.0)

        stay["folioCharges"][0]["settled"] = False
        stay["folioCharges"] = [stay["folioCharges"][0]]
        stay["fbTransferBalance"] = 526.0
        allocated = db_mod._hotel_allocate_fb_invoice_balances(stay)
        self.assertEqual(allocated.get("FBE/26-27/00001"), 526.0)

    def test_folio_keeps_each_pos_invoice_as_own_line(self):
        """Each restaurant/bar transfer must stay a distinct folio line with its order number."""
        check_in, check_out = self._stay_window(nights=1)
        stay = db_mod._normalize_hotel_room_stay(
            {
                "firstName": "Asha",
                "lastName": "Nair",
                "checkInDate": check_in,
                "checkOutDate": check_out,
                "nights": 1,
                "roomRate": 2000,
                "folioCharges": [
                    {
                        "id": "fc1",
                        "kind": "restaurant_room_transfer",
                        "label": "Restaurant Room Transfer · ORD-1",
                        "amount": 150.48,
                        "source": "pos",
                        "invoiceId": "11",
                    },
                    {
                        "id": "fc2",
                        "kind": "restaurant_room_transfer",
                        "label": "Restaurant Room Transfer · ORD-2",
                        "amount": 80,
                        "source": "pos",
                        "invoiceId": "12",
                    },
                    {
                        "id": "fc3",
                        "kind": "bar_room_transfer",
                        "label": "Bar Room Transfer",
                        "amount": 40,
                        "source": "pos",
                        "invoiceId": "13",
                        "orderNo": "ORD-3",
                    },
                ],
            }
        )
        folio = stay["folioCharges"]
        self.assertEqual(len(folio), 3)
        self.assertEqual(folio[0]["orderNo"], "ORD-1")
        self.assertEqual(folio[1]["orderNo"], "ORD-2")
        self.assertEqual(folio[2]["orderNo"], "ORD-3")
        self.assertEqual(folio[0]["label"], "Restaurant Room Transfer · ORD-1")
        self.assertEqual(folio[1]["label"], "Restaurant Room Transfer · ORD-2")
        self.assertEqual(folio[2]["amount"], 40)
        self.assertAlmostEqual(float(stay["fbTransferTotal"]), 270.48, places=2)
        self.assertAlmostEqual(float(stay["estimatedTotal"]), 2000.0, places=2)

    def test_generate_invoice_mints_hbe_and_fbe_with_linked_pos_orders(self):
        """Generate Invoice mints hotel + combined F&B invoices when transfers exist."""
        self._checkin_with_charges(advance=0)
        conn = db_mod.get_db()
        try:
            db_mod.append_hotel_room_folio_charge(
                conn,
                "room-101",
                amount=150.48,
                kind="restaurant_room_transfer",
                label="Restaurant Room Transfer · SPC/26-27/709",
                source="pos",
                invoice_id="501",
                order_no="SPC/26-27/709",
                outlet="restaurant",
            )
            db_mod.append_hotel_room_folio_charge(
                conn,
                "room-101",
                amount=80.0,
                kind="bar_room_transfer",
                label="Bar Room Transfer · INV/26-27/12",
                source="pos",
                invoice_id="502",
                order_no="INV/26-27/12",
                outlet="bar",
            )
            conn.commit()
        finally:
            conn.close()

        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "payment_splits": []},
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))
        payload = gen.get_json()
        stay = payload["room"]["stay"]
        self.assertTrue(str(stay["invoiceNumber"]).startswith("HBE/"))
        self.assertTrue(str(stay["fbTransferInvoiceNumber"]).startswith("FBE/"))
        linked = payload.get("linkedPosOrders") or []
        order_nos = {row.get("orderNo") for row in linked}
        self.assertIn("SPC/26-27/709", order_nos)
        self.assertIn("INV/26-27/12", order_nos)

        conn = db_mod.get_db()
        try:
            rows = db_mod.list_hotel_room_invoices(
                conn, source="fb_combined_transfer", q=stay["fbTransferInvoiceNumber"]
            )
        finally:
            conn.close()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["invoice_number"], stay["fbTransferInvoiceNumber"])

    def test_hotel_only_stay_does_not_mint_fbe(self):
        self._checkin_with_charges(advance=0)
        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "payment_splits": []},
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))
        stay = gen.get_json()["room"]["stay"]
        self.assertTrue(str(stay["invoiceNumber"]).startswith("HBE/"))
        self.assertFalse(stay.get("fbTransferInvoiceNumber"))
        self.assertIsNone(gen.get_json().get("fbInvoice"))

    def test_generate_invoice_kind_hotel_and_fb_separately(self):
        """Room and F&B invoices can be minted independently via invoice_kind."""
        self._checkin_with_charges(advance=0)
        conn = db_mod.get_db()
        try:
            db_mod.append_hotel_room_folio_charge(
                conn,
                "room-101",
                amount=150.48,
                kind="restaurant_room_transfer",
                label="Restaurant Room Transfer · ORD-SEP",
                source="pos",
                invoice_id="701",
                order_no="ORD-SEP",
                outlet="restaurant",
            )
            conn.commit()
        finally:
            conn.close()

        hotel_only = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "generate_invoice",
                "invoice_kind": "hotel",
                "payment_splits": [],
            },
        )
        self.assertEqual(hotel_only.status_code, 200, hotel_only.get_data(as_text=True))
        hotel_body = hotel_only.get_json()
        stay = hotel_body["room"]["stay"]
        self.assertTrue(hotel_body.get("minted"))
        self.assertFalse(hotel_body.get("fbMinted"))
        self.assertTrue(str(stay["invoiceNumber"]).startswith("HBE/"))
        self.assertFalse(stay.get("fbTransferInvoiceNumber"))

        fb_only = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "generate_invoice",
                "invoice_kind": "fb",
                "payment_splits": [],
            },
        )
        self.assertEqual(fb_only.status_code, 200, fb_only.get_data(as_text=True))
        fb_body = fb_only.get_json()
        stay2 = fb_body["room"]["stay"]
        self.assertFalse(fb_body.get("minted"))
        self.assertTrue(fb_body.get("fbMinted"))
        self.assertEqual(stay2["invoiceNumber"], stay["invoiceNumber"])
        self.assertTrue(str(stay2["fbTransferInvoiceNumber"]).startswith("FBE/"))

        again_hotel = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "generate_invoice",
                "invoice_kind": "hotel",
                "payment_splits": [],
            },
        )
        self.assertEqual(again_hotel.status_code, 400, again_hotel.get_data(as_text=True))
        self.assertIn("No pending room charges", again_hotel.get_data(as_text=True))

    def test_supplemental_fbe_locks_first_invoice(self):
        """New F&B transfers after FBE generation mint a supplemental invoice."""
        self._checkin_with_charges(advance=0)
        conn = db_mod.get_db()
        try:
            db_mod.append_hotel_room_folio_charge(
                conn,
                "room-101",
                amount=150.48,
                kind="restaurant_room_transfer",
                label="Restaurant Room Transfer · ORD-A",
                source="pos",
                invoice_id="601",
                order_no="ORD-A",
                outlet="restaurant",
            )
            conn.commit()
        finally:
            conn.close()

        gen1 = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "payment_splits": []},
        )
        self.assertEqual(gen1.status_code, 200, gen1.get_data(as_text=True))
        stay1 = gen1.get_json()["room"]["stay"]
        fbe1 = stay1["fbTransferInvoiceNumber"]
        conn = db_mod.get_db()
        try:
            row1 = db_mod.get_hotel_room_invoice(conn, fbe1)
        finally:
            conn.close()
        est1 = float(row1["estimated_total"])
        lines1 = ((row1.get("room") or {}).get("stay") or {}).get("folioCharges") or []
        self.assertEqual(len(lines1), 1)

        conn = db_mod.get_db()
        try:
            result_b = db_mod.append_hotel_room_folio_charge(
                conn,
                "room-101",
                amount=80.0,
                kind="bar_room_transfer",
                label="Bar Room Transfer · ORD-B",
                source="pos",
                invoice_id="602",
                order_no="ORD-B",
                outlet="bar",
            )
            db_mod.upsert_pos_room_transfer_invoice(
                conn, result_b["room"], result_b["charge"]
            )
            conn.commit()
        finally:
            conn.close()

        conn = db_mod.get_db()
        try:
            row1_locked = db_mod.get_hotel_room_invoice(conn, fbe1)
            live = db_mod.get_hotel_room(conn, "room-101")
            pending_fb = db_mod._hotel_pending_fb_total(live["stay"])
        finally:
            conn.close()
        self.assertAlmostEqual(float(row1_locked["estimated_total"]), est1, places=2)
        self.assertGreater(pending_fb, 0.009)

        gen2 = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "payment_splits": []},
        )
        self.assertEqual(gen2.status_code, 200, gen2.get_data(as_text=True))
        stay2 = gen2.get_json()["room"]["stay"]
        fb_history = [
            row["invoiceNumber"]
            for row in (stay2.get("invoiceHistory") or [])
            if str(row.get("kind") or "").lower() == "fb"
        ]
        self.assertIn(fbe1, fb_history)
        self.assertEqual(len(fb_history), 2)
        fbe2 = next(n for n in fb_history if n != fbe1)

        conn = db_mod.get_db()
        try:
            row1_after = db_mod.get_hotel_room_invoice(conn, fbe1)
            row2 = db_mod.get_hotel_room_invoice(conn, fbe2)
        finally:
            conn.close()
        payload1_after = ((row1_after.get("room") or {}).get("stay") or {})
        payload2_stay = (row2.get("room") or {}).get("stay") or {}
        self.assertAlmostEqual(float(row1_after["estimated_total"]), est1, places=2)
        self.assertEqual(len(payload1_after.get("folioCharges") or []), 1)
        self.assertEqual(len(payload2_stay.get("folioCharges") or []), 1)

        # Second POS transfer must show as Invoice Generated on Room Transfer ledger.
        spc_b = self.client.get("/hotel/invoice-ledger/api/ORD-B")
        self.assertEqual(spc_b.status_code, 200, spc_b.get_data(as_text=True))
        spc_b_item = spc_b.get_json()["invoice"]
        self.assertEqual(spc_b_item["status"], "cancelled")
        self.assertEqual(
            spc_b_item["payment_mode_label"], f"Invoice Generated ({fbe2})"
        )
        rt_html = self.client.get("/hotel/room-transfer-invoices").get_data(as_text=True)
        self.assertIn(f"Invoice Generated ({fbe2})", rt_html)
        self.assertNotRegex(
            rt_html,
            r"ORD-B[\s\S]{0,400}Invoice yet to generate",
        )

    def test_room_transfer_ledger_heals_status_after_fbe_tag(self):
        """Backfill must not leave tagged F&B transfers as 'yet to generate'."""
        self._checkin_with_charges(advance=0)
        conn = db_mod.get_db()
        try:
            result = db_mod.append_hotel_room_folio_charge(
                conn,
                "room-101",
                amount=399,
                kind="bar_room_transfer",
                label="Bar Room Transfer · INV/26-27/499",
                source="pos",
                invoice_id="499",
                order_no="INV/26-27/499",
                outlet="bar",
            )
            db_mod.upsert_pos_room_transfer_invoice(
                conn, result["room"], result["charge"]
            )
            conn.commit()
        finally:
            conn.close()

        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "generate_invoice",
                "invoice_kind": "fb",
                "payment_splits": [],
            },
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))
        fb_no = gen.get_json()["room"]["stay"]["fbTransferInvoiceNumber"]
        self.assertTrue(str(fb_no).startswith("FBE/"))

        # Simulate a stale open row (pre-heal) then list/backfill.
        conn = db_mod.get_db()
        try:
            conn.execute(
                """
                UPDATE hotel_room_invoices
                SET status = 'open',
                    cancel_reason = '',
                    cancelled_at = '',
                    balance_amount = 399
                WHERE invoice_number = ?
                """,
                ("INV/26-27/499",),
            )
            conn.commit()
            db_mod.backfill_pos_room_transfer_invoices_from_layout(conn)
            conn.commit()
            item = db_mod.get_hotel_room_invoice(conn, "INV/26-27/499")
        finally:
            conn.close()
        self.assertEqual(item["status"], "cancelled")
        self.assertEqual(item["payment_mode_label"], f"Invoice Generated ({fb_no})")

        rt_html = self.client.get("/hotel/room-transfer-invoices").get_data(as_text=True)
        self.assertIn(f"Invoice Generated ({fb_no})", rt_html)

    def test_supplemental_hbe_locks_first_invoice_on_overstay(self):
        """Overstay after HBE generation mints a supplemental hotel invoice."""
        check_in, check_out = self._stay_window(nights=1)
        res = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Lock",
                    "lastName": "Stay",
                    "mobile": "9000000777",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 3500,
                    "advancePaid": 0,
                },
            },
        )
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))

        gen1 = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "payment_splits": []},
        )
        self.assertEqual(gen1.status_code, 200, gen1.get_data(as_text=True))
        stay1 = gen1.get_json()["room"]["stay"]
        hbe1 = stay1["invoiceNumber"]
        conn = db_mod.get_db()
        try:
            row1 = db_mod.get_hotel_room_invoice(conn, hbe1)
        finally:
            conn.close()
        est1 = float(row1["estimated_total"])

        conn = db_mod.get_db()
        try:
            layout = db_mod.get_hotel_rooms_layout(conn)
            rooms = layout.get("rooms") or []
            for room in rooms:
                if room.get("id") != "room-101":
                    continue
                stay = room.get("stay") or {}
                stay["checkOutDate"] = (
                    datetime.now().date() - timedelta(days=1)
                ).isoformat()
                room["stay"] = stay
            db_mod.save_hotel_rooms_layout(conn, layout.get("floors") or [], rooms)
            conn.commit()
        finally:
            conn.close()

        refreshed = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        stay_live = refreshed["stay"]
        self.assertGreater(int(stay_live.get("billableNights") or 0), int(stay1.get("hotelInvoicedBillableNights") or 0))

        conn = db_mod.get_db()
        try:
            pending_hotel, _ = db_mod._hotel_pending_hotel_amount(stay_live)
        finally:
            conn.close()
        self.assertGreater(pending_hotel, 0.009)

        gen2 = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "payment_splits": []},
        )
        self.assertEqual(gen2.status_code, 200, gen2.get_data(as_text=True))
        stay2 = gen2.get_json()["room"]["stay"]
        hbe_history = [
            row["invoiceNumber"]
            for row in (stay2.get("invoiceHistory") or [])
            if str(row.get("kind") or "").lower() == "hotel"
        ]
        self.assertIn(hbe1, hbe_history)
        self.assertEqual(len(hbe_history), 2)
        hbe2 = next(n for n in hbe_history if n != hbe1)
        self.assertNotEqual(hbe2, hbe1)

        conn = db_mod.get_db()
        try:
            row1_after = db_mod.get_hotel_room_invoice(conn, hbe1)
        finally:
            conn.close()
        self.assertAlmostEqual(float(row1_after["estimated_total"]), est1, places=2)

    def test_checkout_blocked_when_pending_charges_remain(self):
        self._checkin_with_charges(advance=0)
        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "payment_splits": []},
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))

        conn = db_mod.get_db()
        try:
            db_mod.append_hotel_room_folio_charge(
                conn,
                "room-101",
                amount=120.0,
                kind="restaurant_room_transfer",
                label="Restaurant Room Transfer · ORD-C",
                source="pos",
                invoice_id="603",
                order_no="ORD-C",
                outlet="restaurant",
            )
            conn.commit()
        finally:
            conn.close()

        checkout = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "checkout"},
        )
        self.assertEqual(checkout.status_code, 400, checkout.get_data(as_text=True))
        self.assertIn("Additional Invoice", checkout.get_data(as_text=True))

    def test_extra_bed_after_invoice_allows_additional_then_checkout(self):
        """Extra Bed added after HBE must mint Additional — then checkout works."""
        self._checkin_with_charges("room-101", advance=0)
        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "payment_splits": []},
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))
        stay1 = gen.get_json()["room"]["stay"]
        hbe1 = stay1["invoiceNumber"]
        self.assertEqual(float(stay1.get("hotelInvoicedExtraBedAmount") or 0), 500.0)

        # Simulate Extra Bed raised after the first invoice (layout mutation —
        # update_charge is locked once HBE exists; front desk still ends up here
        # via stay edits / merge flows).
        conn = db_mod.get_db()
        try:
            layout = db_mod.get_hotel_rooms_layout(conn)
            rooms = layout.get("rooms") or []
            room = next(r for r in rooms if r.get("id") == "room-101")
            stay = dict(room.get("stay") or {})
            stay["extraBedAmount"] = 2400.0
            stay["extraBedQty"] = 1
            stay["extraBedRate"] = 2400.0
            stay["extraBedNights"] = 1
            room["stay"] = db_mod._normalize_hotel_room_stay(stay)
            db_mod.save_hotel_rooms_layout(conn, layout.get("floors") or [], rooms)
            conn.commit()
        finally:
            conn.close()

        stay_pending = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]["stay"]
        self.assertEqual(float(stay_pending.get("extraBedAmount") or 0), 2400.0)
        pending, lines = db_mod._hotel_pending_hotel_amount(stay_pending)
        self.assertGreater(pending, 0.009)
        self.assertTrue(any("Extra" in (row.get("label") or "") for row in lines))

        blocked = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "checkout"},
        )
        self.assertEqual(blocked.status_code, 400, blocked.get_data(as_text=True))
        self.assertIn("Additional Invoice", blocked.get_data(as_text=True))

        gen2 = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "payment_splits": [], "invoice_kind": "hotel"},
        )
        self.assertEqual(gen2.status_code, 200, gen2.get_data(as_text=True))
        stay2 = gen2.get_json()["room"]["stay"]
        self.assertEqual(stay2["invoiceNumber"], hbe1)
        self.assertEqual(float(stay2.get("hotelInvoicedExtraBedAmount") or 0), 2400.0)
        hist = stay2.get("invoiceHistory") or []
        hotel_hist = [e for e in hist if (e.get("kind") or "") == "hotel"]
        self.assertGreaterEqual(len(hotel_hist), 2)

        closed = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "checkout"},
        )
        self.assertEqual(closed.status_code, 200, closed.get_data(as_text=True))
        self.assertEqual(closed.get_json()["room"]["status"], "dirty")

    def test_unmerge_primary_scope_one_shows_former_member(self):
        """Unmerge Room on the primary must not leave the member hidden on the board."""
        self._checkin_with_charges("room-106", advance=0)
        self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Orphan",
                    "lastName": "Member",
                    "mobile": "9000000101",
                    "checkInDate": "2026-07-29",
                    "nights": 1,
                    "roomRate": 2000,
                    "advancePaid": 0,
                },
            },
        )
        self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-101",
                "toRoomId": "room-106",
            },
        )
        layout_before = self.client.get("/hotel/api/rooms").get_json()
        member_before = next(
            r for r in layout_before["rooms"] if r.get("id") == "room-101"
        )
        self.assertTrue(member_before.get("isMergeMember"))

        unmerged = self.client.put(
            "/hotel/api/rooms/room-106",
            json={"action": "unmerge_rooms", "scope": "one"},
        )
        self.assertEqual(unmerged.status_code, 200, unmerged.get_data(as_text=True))

        layout = self.client.get("/hotel/api/rooms").get_json()
        by_id = {r["id"]: r for r in layout["rooms"]}
        primary = by_id["room-106"]
        member = by_id["room-101"]
        self.assertFalse(primary.get("isMergePrimary"))
        self.assertFalse(primary.get("mergeGroupId"))
        self.assertFalse(member.get("isMergeMember"))
        self.assertFalse(member.get("mergeGroupId"))
        self.assertEqual((member.get("stay") or {}).get("mergeRole") or "", "")
        self.assertEqual((member.get("stay") or {}).get("billingRoomId") or "", "")

    def test_heal_orphan_merge_member_clears_hidden_room(self):
        """Layout heal recovers members left behind after a bad primary unmerge."""
        conn = db_mod.get_db()
        try:
            layout = db_mod.get_hotel_rooms_layout(conn)
            rooms = layout["rooms"]
            member = next(r for r in rooms if r["id"] == "room-101")
            member["mergeGroupId"] = "mg-orphan-101"
            member["mergePrimary"] = False
            member["status"] = "vacant"
            member["stay"] = db_mod._normalize_hotel_room_stay(
                {
                    "mergeRole": "member",
                    "billingRoomId": "room-106",
                    "firstName": "",
                    "lastName": "",
                }
            )
            db_mod.save_hotel_rooms_layout(conn, layout["floors"], rooms)
        finally:
            conn.close()

        healed = self.client.get("/hotel/api/rooms").get_json()
        room_101 = next(r for r in healed["rooms"] if r["id"] == "room-101")
        self.assertFalse(room_101.get("isMergeMember"))
        self.assertFalse(room_101.get("mergeGroupId"))
        stay = room_101.get("stay") or {}
        self.assertEqual(stay.get("mergeRole") or "", "")
        self.assertEqual(stay.get("billingRoomId") or "", "")

    def test_transfer_member_and_primary_remap_merge(self):
        self._checkin_with_charges("room-101", advance=0)
        self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Asha",
                    "lastName": "Nair",
                    "mobile": "9000000011",
                    "checkInDate": "2026-07-29",
                    "nights": 1,
                    "roomRate": 2000,
                    "advancePaid": 0,
                },
            },
        )
        self.client.put(
            "/hotel/api/rooms/room-102",
            json={"action": "merge_rooms", "fromRoomId": "room-102", "toRoomId": "room-101"},
        )

        # Transfer member 102 → vacant 103; keep billing on 101.
        moved_member = self.client.put(
            "/hotel/api/rooms/room-102",
            json={"action": "transfer", "toRoomId": "room-103"},
        )
        self.assertEqual(moved_member.status_code, 200, moved_member.get_data(as_text=True))
        to_member = moved_member.get_json()["toRoom"]
        self.assertEqual(to_member["id"], "room-103")
        self.assertTrue(to_member["isMergeMember"])
        self.assertEqual(to_member["stay"]["billingRoomId"], "room-101")
        old_member = moved_member.get_json()["fromRoom"]
        self.assertEqual(old_member["status"], "dirty")
        self.assertFalse(old_member.get("mergeGroupId"))

        # Transfer primary 101 → vacant 104; remap member billingRoomId.
        moved_primary = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "transfer", "toRoomId": "room-104"},
        )
        self.assertEqual(moved_primary.status_code, 200, moved_primary.get_data(as_text=True))
        new_primary = moved_primary.get_json()["toRoom"]
        self.assertTrue(new_primary["isMergePrimary"])
        member = self.client.get("/hotel/api/rooms/room-103").get_json()["room"]
        self.assertEqual(member["stay"]["billingRoomId"], "room-104")

    def test_set_merge_primary_and_checkout_group(self):
        self._checkin_with_charges("room-101", advance=0)
        self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Asha",
                    "lastName": "Nair",
                    "mobile": "9000000011",
                    "checkInDate": "2026-07-29",
                    "nights": 1,
                    "roomRate": 2000,
                    "advancePaid": 0,
                },
            },
        )
        self.client.put(
            "/hotel/api/rooms/room-102",
            json={"action": "merge_rooms", "fromRoomId": "room-102", "toRoomId": "room-101"},
        )
        swapped = self.client.put(
            "/hotel/api/rooms/room-102",
            json={"action": "set_merge_primary"},
        )
        self.assertEqual(swapped.status_code, 200, swapped.get_data(as_text=True))
        new_primary = swapped.get_json()["room"]
        self.assertEqual(new_primary["id"], "room-102")
        self.assertTrue(new_primary["isMergePrimary"])
        self.assertGreater(len(new_primary["stay"]["folioCharges"]), 0)
        old = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        self.assertTrue(old["isMergeMember"])
        self.assertEqual(old["stay"]["billingRoomId"], "room-102")

        self._generate_stay_invoice("room-102")
        closed = self.client.put(
            "/hotel/api/rooms/room-102",
            json={"action": "checkout"},
        )
        self.assertEqual(closed.status_code, 200, closed.get_data(as_text=True))
        self.assertEqual(closed.get_json()["room"]["status"], "dirty")
        other = self.client.get("/hotel/api/rooms/room-101").get_json()["room"]
        self.assertEqual(other["status"], "occupied")
        self.assertIn("stay", other)
        self.assertFalse(other.get("isMergeMember"))

    def test_checkout_group_clears_all_merged_rooms(self):
        self._checkin_with_charges("room-101", advance=0)
        self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Asha",
                    "lastName": "Nair",
                    "mobile": "9000000011",
                    "checkInDate": "2026-07-29",
                    "nights": 1,
                    "roomRate": 2000,
                    "advancePaid": 0,
                },
            },
        )
        self.client.put(
            "/hotel/api/rooms/room-103",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Ravi",
                    "lastName": "Menon",
                    "mobile": "9000000012",
                    "checkInDate": "2026-07-29",
                    "nights": 1,
                    "roomRate": 1800,
                    "advancePaid": 0,
                },
            },
        )
        self.client.put(
            "/hotel/api/rooms/room-102",
            json={"action": "merge_rooms", "fromRoomId": "room-102", "toRoomId": "room-101"},
        )
        self.client.put(
            "/hotel/api/rooms/room-103",
            json={"action": "merge_rooms", "fromRoomId": "room-103", "toRoomId": "room-101"},
        )
        self._generate_stay_invoice("room-101")
        closed = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "checkout_group"},
        )
        self.assertEqual(closed.status_code, 200, closed.get_data(as_text=True))
        body = closed.get_json()
        self.assertTrue(body["ok"])
        ids = set(body.get("checkedOutRoomIds") or [])
        self.assertEqual(ids, {"room-101", "room-102", "room-103"})
        for rid in ("room-101", "room-102", "room-103"):
            room = self.client.get("/hotel/api/rooms/" + rid).get_json()["room"]
            self.assertEqual(room["status"], "dirty", rid)
            self.assertFalse(room.get("isMergePrimary"))
            self.assertFalse(room.get("isMergeMember"))
            self.assertFalse(room.get("mergeGroupId"))
            self.assertFalse(room.get("stay"), rid)

    def test_merged_invoice_snapshots_both_rooms_and_survives_checkout(self):
        """Merged bill prints both room numbers; checkout of primary leaves member occupied."""
        self._checkin_with_charges("room-101", advance=0)
        self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Asha",
                    "lastName": "Nair",
                    "mobile": "9000000011",
                    "checkInDate": "2026-07-29",
                    "nights": 1,
                    "roomRate": 2000,
                    "advancePaid": 0,
                },
            },
        )
        merged = self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-102",
                "toRoomId": "room-101",
            },
        )
        self.assertEqual(merged.status_code, 200, merged.get_data(as_text=True))
        primary = merged.get_json()["primaryRoom"]
        self.assertEqual(primary.get("mergeLabel"), "101 + 102")
        self.assertTrue(primary.get("isMergePrimary"))

        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={
                "action": "generate_invoice",
                "payment": {"amount": 1, "method": "cash"},
            },
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))
        stay = gen.get_json()["room"]["stay"]
        inv_no = stay["invoiceNumber"]
        self.assertEqual(stay.get("mergeRoomNumbers"), ["101", "102"])
        self.assertEqual(stay.get("mergeRoomLabel"), "101 + 102")

        # Settle remaining balance so checkout is allowed in UI flows.
        balance = float(stay.get("balanceAmount") or 0)
        if balance > 0:
            paid = self.client.put(
                "/hotel/api/rooms/room-101",
                json={
                    "action": "record_payment",
                    "amount": balance,
                    "method": "cash",
                },
            )
            self.assertEqual(paid.status_code, 200, paid.get_data(as_text=True))

        closed = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "checkout"},
        )
        self.assertEqual(closed.status_code, 200, closed.get_data(as_text=True))
        primary_after = closed.get_json()["room"]
        self.assertEqual(primary_after["status"], "dirty")
        self.assertFalse(primary_after.get("mergeGroupId"))
        self.assertFalse(primary_after.get("isMergePrimary"))
        self.assertNotIn("stay", primary_after)

        member_after = self.client.get("/hotel/api/rooms/room-102").get_json()["room"]
        self.assertEqual(member_after["status"], "occupied")
        self.assertTrue(member_after.get("isMergePrimary") or not member_after.get("isMergeMember"))
        self.assertIn("stay", member_after)
        # Inherited merge bill must stay locked — no phantom "Generate Additional".
        mstay = member_after["stay"]
        self.assertTrue(mstay.get("invoiceGenerated"))
        self.assertEqual(mstay.get("invoiceNumber"), inv_no)
        self.assertGreaterEqual(
            int(mstay.get("hotelInvoicedBillableNights") or 0),
            int(mstay.get("billableNights") or 0),
        )
        untagged = [
            f
            for f in (mstay.get("folioCharges") or [])
            if isinstance(f, dict)
            and float(f.get("amount") or 0) > 0
            and not str(f.get("invoicedInvoiceNumber") or "").strip()
            and str(f.get("kind") or "")
            not in ("restaurant_room_transfer", "bar_room_transfer")
        ]
        self.assertEqual(untagged, [])

        api = self.client.get(f"/hotel/invoice-ledger/api/{inv_no}")
        self.assertEqual(api.status_code, 200, api.get_data(as_text=True))
        body = api.get_json()
        self.assertTrue(body["ok"])
        ledger_room = body["room"]
        ledger_stay = ledger_room["stay"]
        self.assertCountEqual(ledger_stay.get("mergeRoomNumbers") or [], ["101", "102"])
        self.assertCountEqual(
            [p.strip() for p in str(ledger_stay.get("mergeRoomLabel") or "").split("+") if p.strip()],
            ["101", "102"],
        )
        self.assertCountEqual(
            [p.strip() for p in str(body["invoice"]["room_number"] or "").split("+") if p.strip()],
            ["101", "102"],
        )
        self.assertIn("102", ledger_room.get("mergeRoomLabel") or "")

    def test_ledger_room_label_stays_frozen_after_member_checkout(self):
        """Invoice ROOM column keeps all merge rooms after a member leaves."""
        check_in, check_out = self._stay_window(nights=1)
        for rid, rate, mobile in (
            ("room-101", 4200, "9000000401"),
            ("room-102", 4200, "9000000402"),
            ("room-103", 4200, "9000000403"),
        ):
            res = self.client.put(
                f"/hotel/api/rooms/{rid}",
                json={
                    "action": "checkin",
                    "stay": {
                        "firstName": "Merge",
                        "lastName": rid[-3:],
                        "mobile": mobile,
                        "checkInDate": check_in,
                        "checkOutDate": check_out,
                        "nights": 1,
                        "roomRate": rate,
                        "advancePaid": rate,
                    },
                },
            )
            self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        self.client.put(
            "/hotel/api/rooms/room-102",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-102",
                "toRoomId": "room-101",
            },
        )
        self.client.put(
            "/hotel/api/rooms/room-103",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-103",
                "toRoomId": "room-101",
            },
        )
        gen = self.client.put(
            "/hotel/api/rooms/room-101",
            json={"action": "generate_invoice", "payment_splits": []},
        )
        self.assertEqual(gen.status_code, 200, gen.get_data(as_text=True))
        inv_no = gen.get_json()["room"]["stay"]["invoiceNumber"]
        before = self.client.get(f"/hotel/invoice-ledger/api/{inv_no}").get_json()
        before_rooms = [
            p.strip()
            for p in str(before["invoice"]["room_number"] or "").split("+")
            if p.strip()
        ]
        self.assertCountEqual(before_rooms, ["101", "102", "103"])

        closed = self.client.put(
            "/hotel/api/rooms/room-102",
            json={"action": "checkout"},
        )
        self.assertEqual(closed.status_code, 200, closed.get_data(as_text=True))

        # Payment sync / ledger refresh must not drop checked-out room from ROOM.
        after = self.client.get(f"/hotel/invoice-ledger/api/{inv_no}").get_json()
        after_rooms = [
            p.strip()
            for p in str(after["invoice"]["room_number"] or "").split("+")
            if p.strip()
        ]
        self.assertCountEqual(after_rooms, ["101", "102", "103"])
        page = self.client.get("/hotel/invoice-ledger")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn(inv_no, html)
        self.assertTrue(
            ("101" in html and "102" in html and "103" in html),
            "ledger HTML should still list all invoice rooms",
        )

    def test_normalize_nightly_rates_sum_and_fill(self):
        stay = db_mod._normalize_hotel_room_stay(
            {
                "firstName": "Nightly",
                "checkInDate": "2026-08-15",
                "checkOutDate": "2026-08-17",
                "nights": 2,
                "roomRate": 5000,
                "ratePlan": "EP",
                "nightlyRates": [
                    {"date": "2026-08-15", "roomRate": 5000, "ratePlan": "AP"},
                    {"date": "2026-08-16", "roomRate": 4500, "ratePlan": "CP"},
                ],
            }
        )
        self.assertEqual(stay["totalRate"], 9500.0)
        self.assertEqual(stay["roomRate"], 5000.0)
        self.assertEqual(stay["ratePlan"], "AP")
        self.assertEqual(len(stay["nightlyRates"]), 2)
        self.assertEqual(stay["nightlyRates"][1]["ratePlan"], "CP")

        filled = db_mod._normalize_hotel_room_stay(
            {
                "firstName": "Fill",
                "checkInDate": "2026-08-15",
                "nights": 3,
                "roomRate": 4000,
                "ratePlan": "EP",
                "nightlyRates": [
                    {"date": "2026-08-15", "roomRate": 5000, "ratePlan": "AP"},
                    {"date": "2026-08-17", "roomRate": 4500, "ratePlan": "CP"},
                ],
            }
        )
        self.assertEqual(
            [row["roomRate"] for row in filled["nightlyRates"]],
            [5000.0, 5000.0, 4500.0],
        )
        self.assertEqual(filled["totalRate"], 14500.0)
        self.assertEqual(
            [row["ratePlan"] for row in filled["nightlyRates"]],
            ["AP", "AP", "CP"],
        )

    def test_normalize_legacy_flat_rate_without_nightly(self):
        stay = db_mod._normalize_hotel_room_stay(
            {
                "firstName": "Legacy",
                "checkInDate": "2026-08-15",
                "checkOutDate": "2026-08-17",
                "nights": 2,
                "roomRate": 5000,
                "ratePlan": "EP",
            }
        )
        self.assertEqual(stay.get("nightlyRates"), [])
        self.assertEqual(stay["totalRate"], 10000.0)
        self.assertEqual(
            db_mod._hotel_stay_room_charges_amount(stay),
            10000.0,
        )

    def test_normalize_heals_id_document_path_from_filename(self):
        stay = db_mod._normalize_hotel_room_stay(
            {
                "firstName": "Lukas",
                "idType": "Passport",
                "idDocumentName": "f3dd5b8c-2958-4e62-adbd-b62b9e6f89eb.webp",
                "idDocumentPath": "",
                "idDocumentMime": "image/webp",
            }
        )
        self.assertEqual(
            stay["idDocumentPath"],
            "/hotel/api/id-documents/view/f3dd5b8c-2958-4e62-adbd-b62b9e6f89eb.webp/raw",
        )
        extras = db_mod._normalize_hotel_room_stay(
            {
                "firstName": "Host",
                "additionalGuests": [
                    {
                        "name": "Extra",
                        "idType": "Aadhaar",
                        "idDocumentName": "9bd51354-c325-456a-919c-9d9910c52808.webp",
                    }
                ],
            }
        )
        self.assertEqual(
            extras["additionalGuests"][0]["idDocumentPath"],
            "/hotel/api/id-documents/view/9bd51354-c325-456a-919c-9d9910c52808.webp/raw",
        )

    def test_same_reservation_id_auto_merges_reserved_rooms(self):
        """Two reserved rooms with the same reservation id join one merge group."""
        stay = {
            "guestName": "Rid Guest",
            "mobile": "9000012345",
            "reservationId": "AT-RID-1001",
        }
        payload = {
            "action": "reserve",
            "checkInDate": "2026-11-10",
            "checkOutDate": "2026-11-12",
            "stay": stay,
        }
        first = self.client.put("/hotel/api/rooms/room-201", json=payload)
        second = self.client.put("/hotel/api/rooms/room-202", json=payload)
        self.assertEqual(first.status_code, 200, first.get_data(as_text=True))
        self.assertEqual(second.status_code, 200, second.get_data(as_text=True))

        board = self.client.get("/hotel/api/rooms")
        self.assertEqual(board.status_code, 200)
        rooms = {r["id"]: r for r in board.get_json().get("rooms") or []}
        self.assertTrue(rooms["room-201"].get("mergeGroupId"))
        self.assertEqual(
            rooms["room-201"].get("mergeGroupId"),
            rooms["room-202"].get("mergeGroupId"),
        )
        self.assertTrue(
            rooms["room-201"].get("isMergePrimary")
            or rooms["room-202"].get("isMergePrimary")
        )
        self.assertEqual(rooms["room-201"]["status"], "reserved")
        self.assertEqual(rooms["room-202"]["status"], "reserved")

    def test_unmerge_same_reservation_does_not_remerge_on_get(self):
        stay = {
            "guestName": "Rid Guest",
            "mobile": "9000012345",
            "reservationId": "AT-RID-UNMERGE-1",
        }
        payload = {
            "action": "reserve",
            "checkInDate": "2026-11-10",
            "checkOutDate": "2026-11-12",
            "stay": stay,
        }
        self.assertEqual(
            self.client.put("/hotel/api/rooms/room-201", json=payload).status_code, 200
        )
        self.assertEqual(
            self.client.put("/hotel/api/rooms/room-202", json=payload).status_code, 200
        )
        board = self.client.get("/hotel/api/rooms")
        self.assertEqual(board.status_code, 200)
        rooms = {r["id"]: r for r in board.get_json().get("rooms") or []}
        self.assertTrue(rooms["room-201"].get("mergeGroupId"))
        member_id = (
            "room-202" if rooms["room-202"].get("isMergeMember") else "room-201"
        )
        unmerged = self.client.put(
            f"/hotel/api/rooms/{member_id}",
            json={"action": "unmerge_rooms", "scope": "one"},
        )
        self.assertEqual(unmerged.status_code, 200, unmerged.get_data(as_text=True))

        again = self.client.get("/hotel/api/rooms")
        self.assertEqual(again.status_code, 200)
        rooms = {r["id"]: r for r in again.get_json().get("rooms") or []}
        self.assertFalse(rooms["room-201"].get("mergeGroupId"))
        self.assertFalse(rooms["room-202"].get("mergeGroupId"))
        self.assertFalse(rooms["room-201"].get("isMergeMember"))
        self.assertFalse(rooms["room-202"].get("isMergeMember"))
        self.assertTrue(rooms["room-201"]["stay"].get("independentBilling"))
        self.assertTrue(rooms["room-202"]["stay"].get("independentBilling"))

        rematch = self.client.put(
            "/hotel/api/rooms/room-201",
            json={
                "action": "merge_rooms",
                "fromRoomId": "room-201",
                "toRoomId": "room-202",
            },
        )
        self.assertEqual(rematch.status_code, 200, rematch.get_data(as_text=True))
        rooms = {
            r["id"]: r
            for r in self.client.get("/hotel/api/rooms").get_json().get("rooms") or []
        }
        self.assertTrue(rooms["room-201"].get("mergeGroupId"))
        self.assertEqual(
            rooms["room-201"].get("mergeGroupId"),
            rooms["room-202"].get("mergeGroupId"),
        )
        self.assertFalse(rooms["room-201"]["stay"].get("independentBilling"))
        self.assertFalse(rooms["room-202"]["stay"].get("independentBilling"))

    def test_same_reservation_checkin_occupies_only_that_room(self):
        stay = {
            "guestName": "Stagger Guest",
            "firstName": "Stagger",
            "lastName": "Guest",
            "mobile": "9000012346",
            "reservationId": "AT-RID-1002",
        }
        check_in, check_out = self._stay_window(nights=1)
        payload = {
            "action": "reserve",
            "checkInDate": check_in,
            "checkOutDate": check_out,
            "stay": stay,
        }
        self.assertEqual(
            self.client.put("/hotel/api/rooms/room-201", json=payload).status_code, 200
        )
        self.assertEqual(
            self.client.put("/hotel/api/rooms/room-202", json=payload).status_code, 200
        )
        checked = self.client.put(
            "/hotel/api/rooms/room-201",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Stagger",
                    "lastName": "Guest",
                    "mobile": "9000012346",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 4500,
                    "reservationId": "AT-RID-1002",
                },
            },
        )
        self.assertEqual(checked.status_code, 200, checked.get_data(as_text=True))
        self.assertEqual(checked.get_json()["room"]["status"], "occupied")
        peer = self.client.get("/hotel/api/rooms/room-202").get_json()["room"]
        self.assertEqual(peer["status"], "reserved")

    def test_same_reservation_checkout_clears_this_room_only(self):
        stay = {
            "guestName": "Checkout Group",
            "firstName": "Checkout",
            "lastName": "Group",
            "mobile": "9000012347",
            "reservationId": "AT-RID-1003",
        }
        check_in, check_out = self._stay_window(nights=1)
        payload = {
            "action": "reserve",
            "checkInDate": check_in,
            "checkOutDate": check_out,
            "stay": stay,
        }
        self.assertEqual(
            self.client.put("/hotel/api/rooms/room-201", json=payload).status_code, 200
        )
        self.assertEqual(
            self.client.put("/hotel/api/rooms/room-202", json=payload).status_code, 200
        )
        self.client.put(
            "/hotel/api/rooms/room-201",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Checkout",
                    "lastName": "Group",
                    "mobile": "9000012347",
                    "checkInDate": check_in,
                    "checkOutDate": check_out,
                    "nights": 1,
                    "roomRate": 4500,
                    "reservationId": "AT-RID-1003",
                },
            },
        )
        self._generate_stay_invoice("room-201")
        closed = self.client.put(
            "/hotel/api/rooms/room-201",
            json={"action": "checkout"},
        )
        self.assertEqual(closed.status_code, 200, closed.get_data(as_text=True))
        self.assertEqual(closed.get_json()["room"]["status"], "dirty")
        ids = closed.get_json().get("checkedOutRoomIds") or []
        self.assertEqual(ids, ["room-201"])
        peer = self.client.get("/hotel/api/rooms/room-202").get_json()["room"]
        self.assertEqual(peer["status"], "reserved")
        self.assertTrue(peer.get("stay"))

    def test_same_reservation_checkout_leaves_occupied_peers(self):
        stay = {
            "guestName": "Both In",
            "firstName": "Both",
            "lastName": "In",
            "mobile": "9000012348",
            "reservationId": "AT-RID-1004",
        }
        check_in, check_out = self._stay_window(nights=1)
        payload = {
            "action": "reserve",
            "checkInDate": check_in,
            "checkOutDate": check_out,
            "stay": stay,
        }
        self.client.put("/hotel/api/rooms/room-201", json=payload)
        self.client.put("/hotel/api/rooms/room-202", json=payload)
        for room_id in ("room-201", "room-202"):
            self.client.put(
                f"/hotel/api/rooms/{room_id}",
                json={
                    "action": "checkin",
                    "stay": {
                        "firstName": "Both",
                        "lastName": "In",
                        "mobile": "9000012348",
                        "checkInDate": check_in,
                        "checkOutDate": check_out,
                        "nights": 1,
                        "roomRate": 4500,
                        "reservationId": "AT-RID-1004",
                    },
                },
            )
        self._generate_stay_invoice("room-201")
        closed = self.client.put(
            "/hotel/api/rooms/room-201",
            json={"action": "checkout"},
        )
        self.assertEqual(closed.status_code, 200, closed.get_data(as_text=True))
        self.assertEqual(closed.get_json()["room"]["status"], "dirty")
        peer = self.client.get("/hotel/api/rooms/room-202").get_json()["room"]
        self.assertEqual(peer["status"], "occupied")
        self.assertTrue(peer.get("stay"))

    def test_local_bk_booking_number_does_not_auto_merge(self):
        stay = {
            "guestName": "Local Bk",
            "mobile": "9000012349",
            "bookingNumber": "BK20260816123456",
            "reservationBookingId": "BK20260816123456",
        }
        payload = {
            "action": "reserve",
            "checkInDate": "2026-11-20",
            "checkOutDate": "2026-11-22",
            "stay": stay,
        }
        self.assertEqual(
            self.client.put("/hotel/api/rooms/room-201", json=payload).status_code, 200
        )
        self.assertEqual(
            self.client.put("/hotel/api/rooms/room-202", json=payload).status_code, 200
        )
        board = self.client.get("/hotel/api/rooms")
        rooms = {r["id"]: r for r in board.get_json().get("rooms") or []}
        self.assertFalse(rooms["room-201"].get("mergeGroupId"))
        self.assertFalse(rooms["room-202"].get("mergeGroupId"))


if __name__ == "__main__":
    unittest.main()
