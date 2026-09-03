"""Tests for Cash Ledger aggregations and validation."""

import json
import os
import sqlite3
import tempfile
import unittest
from datetime import date
from unittest import mock

import app as app_module
import db as db_mod


def _memory_conn():
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    conn.executescript(
        """
        CREATE TABLE suppliers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            gst TEXT NOT NULL DEFAULT '',
            address TEXT NOT NULL DEFAULT '',
            phone TEXT NOT NULL DEFAULT '',
            bank_name TEXT NOT NULL DEFAULT '',
            bank_account_number TEXT NOT NULL DEFAULT '',
            ifsc_code TEXT NOT NULL DEFAULT ''
        );
        CREATE TABLE sales_updates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company TEXT NOT NULL,
            location TEXT NOT NULL,
            sales_date TEXT NOT NULL,
            sales_entry_values TEXT NOT NULL DEFAULT '{}'
        );
        CREATE TABLE sales_update_expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company TEXT NOT NULL,
            location TEXT NOT NULL,
            sales_date TEXT NOT NULL,
            description TEXT NOT NULL DEFAULT '',
            amount REAL NOT NULL DEFAULT 0,
            payment_type TEXT NOT NULL DEFAULT 'cash',
            transaction_id TEXT NOT NULL DEFAULT '',
            expense_code TEXT NOT NULL DEFAULT '',
            category TEXT NOT NULL DEFAULT '',
            invoice_number TEXT NOT NULL DEFAULT '',
            supplier_id INTEGER,
            entry_kind TEXT NOT NULL DEFAULT 'expense',
            cancelled_at TEXT,
            cancelled_by INTEGER
        );
        CREATE TABLE cash_ledger_loads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company TEXT NOT NULL,
            load_date TEXT NOT NULL,
            description TEXT NOT NULL DEFAULT '',
            amount REAL NOT NULL DEFAULT 0
        );
        CREATE TABLE cash_ledger_transfers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company TEXT NOT NULL,
            transfer_date TEXT NOT NULL,
            destination TEXT NOT NULL DEFAULT 'bank',
            description TEXT NOT NULL DEFAULT '',
            amount REAL NOT NULL DEFAULT 0
        );
        CREATE TABLE back_office_receipts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            receipt_no TEXT NOT NULL UNIQUE,
            fiscal_year TEXT NOT NULL,
            seq INTEGER NOT NULL,
            receipt_date TEXT NOT NULL,
            payer_name TEXT NOT NULL DEFAULT '',
            agency_id INTEGER,
            amount REAL NOT NULL DEFAULT 0,
            amount_words TEXT NOT NULL DEFAULT '',
            payment_mode TEXT NOT NULL DEFAULT 'cash',
            instrument_no TEXT NOT NULL DEFAULT '',
            instrument_date TEXT,
            towards TEXT NOT NULL DEFAULT ''
        );
        CREATE TABLE room_transfer_payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company TEXT NOT NULL,
            payment_date TEXT NOT NULL,
            payment_method TEXT NOT NULL DEFAULT 'cash',
            transaction_id TEXT NOT NULL DEFAULT '',
            total_amount REAL NOT NULL DEFAULT 0,
            notes TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            updated_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            cancelled_at TEXT,
            cancelled_by INTEGER
        );
        CREATE TABLE room_transfer_payment_allocations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            room_transfer_payment_id INTEGER NOT NULL,
            room_transfer_entry_id INTEGER NOT NULL,
            amount REAL NOT NULL DEFAULT 0,
            invoice_number TEXT NOT NULL DEFAULT '',
            guest_name TEXT NOT NULL DEFAULT '',
            location TEXT NOT NULL DEFAULT '',
            sales_date TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        );
        """
    )
    return conn


class CashLedgerHelperTests(unittest.TestCase):
    def test_available_cash_formula_and_running_balance(self):
        conn = _memory_conn()
        conn.execute(
            "INSERT INTO sales_updates (company, location, sales_date, sales_entry_values) VALUES (?,?,?,?)",
            ("HBE", "Hotel", "2026-07-01", json.dumps({"cash": 9999, "actual_cash": 1000})),
        )
        conn.execute(
            "INSERT INTO sales_updates (company, location, sales_date, sales_entry_values) VALUES (?,?,?,?)",
            ("HBE", "Bar", "2026-07-01", json.dumps({"cash": 9999, "actual_cash": 500})),
        )
        conn.execute(
            "INSERT INTO sales_updates (company, location, sales_date, sales_entry_values) VALUES (?,?,?,?)",
            ("HBE", "Restaurant", "2026-07-02", json.dumps({"cash": 9999, "actual_cash": 250})),
        )
        conn.execute(
            "INSERT INTO cash_ledger_loads (company, load_date, description, amount) VALUES (?,?,?,?)",
            ("HBE", "2026-07-01", "Opening float", 200),
        )
        conn.execute(
            """INSERT INTO sales_update_expenses
               (company, location, sales_date, description, amount, payment_type, expense_code)
               VALUES (?,?,?,?,?,?,?)""",
            ("HBE", "Hotel", "2026-07-01", "Veggies", 100, "cash", "HBE-EX-1"),
        )
        conn.execute(
            """INSERT INTO sales_update_expenses
               (company, location, sales_date, description, amount, payment_type, expense_code)
               VALUES (?,?,?,?,?,?,?)""",
            ("HBE", "Bar", "2026-07-02", "Ice", 50, "cash", "HBE-EX-2"),
        )
        conn.execute(
            """INSERT INTO sales_update_expenses
               (company, location, sales_date, description, amount, payment_type, expense_code)
               VALUES (?,?,?,?,?,?,?)""",
            ("HBE", "Hotel", "2026-07-02", "Bank paid", 80, "bank_transfer", "HBE-EX-3"),
        )
        conn.execute(
            """INSERT INTO cash_ledger_transfers
               (company, transfer_date, destination, description, amount)
               VALUES (?,?,?,?,?)""",
            ("HBE", "2026-07-02", "bank", "Deposit", 300),
        )
        conn.execute(
            """INSERT INTO cash_ledger_transfers
               (company, transfer_date, destination, description, amount)
               VALUES (?,?,?,?,?)""",
            ("HBE", "2026-07-02", "owner", "Owner draw", 150),
        )
        conn.commit()

        entries = app_module._build_cash_ledger_entries(
            conn, "HBE", date(2026, 7, 1), date(2026, 7, 2)
        )
        totals = app_module._cash_ledger_totals(entries)

        self.assertEqual(totals["sales_total"], 1750.0)
        self.assertEqual(totals["load_total"], 200.0)
        self.assertEqual(totals["credit_total"], 0.0)
        self.assertEqual(totals["expense_total"], 150.0)
        self.assertEqual(totals["transfer_total"], 450.0)
        self.assertEqual(totals["available_total"], 1350.0)
        self.assertEqual(entries[-1]["running_balance"], 1350.0)

        outlets = {e["detail"] for e in entries if e["entry_type"] == "sales_cash"}
        self.assertEqual(outlets, {"Hotel", "Bar", "Restaurant"})
        expense_outlets = {
            e["location"] for e in entries if e["entry_type"] == "expense"
        }
        self.assertEqual(expense_outlets, {"Hotel", "Bar"})
        expense_rows = [e for e in entries if e["entry_type"] == "expense"]
        self.assertEqual(
            {(e["expense_code"], e["description"]) for e in expense_rows},
            {("HBE-EX-1", "Veggies"), ("HBE-EX-2", "Ice")},
        )
        conn.close()

    def test_location_filter_scopes_sales_and_expenses(self):
        conn = _memory_conn()
        conn.execute(
            "INSERT INTO sales_updates (company, location, sales_date, sales_entry_values) VALUES (?,?,?,?)",
            ("HBE", "Hotel", "2026-07-01", json.dumps({"actual_cash": 1000})),
        )
        conn.execute(
            "INSERT INTO sales_updates (company, location, sales_date, sales_entry_values) VALUES (?,?,?,?)",
            ("HBE", "Bar", "2026-07-01", json.dumps({"actual_cash": 500})),
        )
        conn.execute(
            "INSERT INTO cash_ledger_loads (company, load_date, description, amount) VALUES (?,?,?,?)",
            ("HBE", "2026-07-01", "Opening float", 200),
        )
        conn.execute(
            """INSERT INTO sales_update_expenses
               (company, location, sales_date, description, amount, payment_type, expense_code)
               VALUES (?,?,?,?,?,?,?)""",
            ("HBE", "Hotel", "2026-07-01", "Veggies", 100, "cash", "HBE-EX-1"),
        )
        conn.execute(
            """INSERT INTO sales_update_expenses
               (company, location, sales_date, description, amount, payment_type, expense_code)
               VALUES (?,?,?,?,?,?,?)""",
            ("HBE", "Bar", "2026-07-01", "Ice", 50, "cash", "HBE-EX-2"),
        )
        conn.execute(
            """INSERT INTO cash_ledger_transfers
               (company, transfer_date, destination, description, amount)
               VALUES (?,?,?,?,?)""",
            ("HBE", "2026-07-01", "bank", "Deposit", 300),
        )
        conn.commit()

        hotel_entries = app_module._build_cash_ledger_entries(
            conn, "HBE", date(2026, 7, 1), date(2026, 7, 1), location="Hotel"
        )
        hotel_totals = app_module._cash_ledger_totals(hotel_entries)
        self.assertEqual(hotel_totals["sales_total"], 1000.0)
        self.assertEqual(hotel_totals["expense_total"], 100.0)
        self.assertEqual(hotel_totals["load_total"], 0.0)
        self.assertEqual(hotel_totals["transfer_total"], 0.0)
        self.assertEqual(
            {e["location"] for e in hotel_entries if e["entry_type"] == "sales_cash"},
            {"Hotel"},
        )

        all_entries = app_module._build_cash_ledger_entries(
            conn, "HBE", date(2026, 7, 1), date(2026, 7, 1), location="All"
        )
        all_totals = app_module._cash_ledger_totals(all_entries)
        self.assertEqual(all_totals["sales_total"], 1500.0)
        self.assertEqual(all_totals["load_total"], 200.0)
        self.assertEqual(all_totals["transfer_total"], 300.0)
        conn.close()

    def test_cash_bor_receipt_appears_in_movements(self):
        conn = _memory_conn()
        conn.execute(
            "INSERT INTO sales_updates (company, location, sales_date, sales_entry_values) VALUES (?,?,?,?)",
            ("HBE", "Hotel", "2026-08-21", json.dumps({"actual_cash": 1000})),
        )
        conn.execute(
            """INSERT INTO back_office_receipts
               (receipt_no, fiscal_year, seq, receipt_date, payer_name, amount,
                amount_words, payment_mode, towards)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (
                "HBE/BOR/26-27/1",
                "26-27",
                1,
                "2026-08-21",
                "ATPI India Pvt. Ltd",
                35000,
                "Rupees Thirty Five Thousand Only",
                "cash",
                "Room Advance",
            ),
        )
        conn.execute(
            """INSERT INTO back_office_receipts
               (receipt_no, fiscal_year, seq, receipt_date, payer_name, amount,
                amount_words, payment_mode, instrument_no, towards)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (
                "HBE/BOR/26-27/2",
                "26-27",
                2,
                "2026-08-21",
                "Bank Payer",
                5000,
                "Rupees Five Thousand Only",
                "cheque",
                "CHQ-1",
                "Advance",
            ),
        )
        conn.commit()

        entries = app_module._build_cash_ledger_entries(
            conn, "HBE", date(2026, 8, 1), date(2026, 8, 31)
        )
        bor_rows = [e for e in entries if e["entry_type"] == "bor_cash"]
        self.assertEqual(len(bor_rows), 1)
        self.assertEqual(bor_rows[0]["expense_code"], "HBE/BOR/26-27/1")
        self.assertEqual(bor_rows[0]["signed_amount"], 35000.0)
        self.assertIn("ATPI India Pvt. Ltd", bor_rows[0]["description"])

        totals = app_module._cash_ledger_totals(entries)
        self.assertEqual(totals["bor_total"], 35000.0)
        self.assertEqual(totals["bor_count"], 1)
        self.assertEqual(totals["available_total"], 36000.0)

        hotel_only = app_module._build_cash_ledger_entries(
            conn, "HBE", date(2026, 8, 1), date(2026, 8, 31), location="Hotel"
        )
        self.assertEqual(
            len([e for e in hotel_only if e["entry_type"] == "bor_cash"]),
            1,
        )
        bar_only = app_module._build_cash_ledger_entries(
            conn, "HBE", date(2026, 8, 1), date(2026, 8, 31), location="Bar"
        )
        self.assertEqual(
            len([e for e in bar_only if e["entry_type"] == "bor_cash"]),
            0,
        )
        conn.close()

    def test_transfer_destination_normalizer(self):
        self.assertEqual(
            app_module._normalize_cash_ledger_transfer_destination("Owner"),
            "owner",
        )
        self.assertEqual(
            app_module._normalize_cash_ledger_transfer_destination("bank"),
            "bank",
        )
        self.assertEqual(
            app_module._normalize_cash_ledger_transfer_destination("petty"),
            "",
        )

    def test_cash_credit_repayment_updates_available(self):
        conn = _memory_conn()
        conn.execute(
            "INSERT INTO sales_updates (company, location, sales_date, sales_entry_values) VALUES (?,?,?,?)",
            ("HBE", "Hotel", "2026-07-01", json.dumps({"actual_cash": 100})),
        )
        cash_payment = conn.execute(
            """INSERT INTO room_transfer_payments
               (company, payment_date, payment_method, total_amount, notes)
               VALUES (?,?,?,?,?)""",
            ("HBE", "2026-07-02", "cash", 400, ""),
        ).lastrowid
        upi_payment = conn.execute(
            """INSERT INTO room_transfer_payments
               (company, payment_date, payment_method, total_amount, notes)
               VALUES (?,?,?,?,?)""",
            ("HBE", "2026-07-02", "upi", 250, ""),
        ).lastrowid
        conn.execute(
            """INSERT INTO room_transfer_payment_allocations
               (room_transfer_payment_id, room_transfer_entry_id, amount,
                invoice_number, guest_name, location, sales_date)
               VALUES (?,?,?,?,?,?,?)""",
            (cash_payment, 1, 250, "INV-1", "Ada", "Hotel", "2026-06-20"),
        )
        conn.execute(
            """INSERT INTO room_transfer_payment_allocations
               (room_transfer_payment_id, room_transfer_entry_id, amount,
                invoice_number, guest_name, location, sales_date)
               VALUES (?,?,?,?,?,?,?)""",
            (cash_payment, 2, 150, "INV-2", "Ada", "Bar", "2026-06-21"),
        )
        conn.execute(
            """INSERT INTO room_transfer_payment_allocations
               (room_transfer_payment_id, room_transfer_entry_id, amount,
                invoice_number, guest_name, location, sales_date)
               VALUES (?,?,?,?,?,?,?)""",
            (upi_payment, 3, 250, "INV-3", "Bob", "Hotel", "2026-06-22"),
        )
        conn.commit()

        entries = app_module._build_cash_ledger_entries(
            conn, "HBE", date(2026, 7, 1), date(2026, 7, 2)
        )
        totals = app_module._cash_ledger_totals(entries)
        self.assertEqual(totals["credit_total"], 400.0)
        self.assertEqual(totals["credit_count"], 1)
        self.assertEqual(totals["available_total"], 500.0)

        credit_rows = [e for e in entries if e["entry_type"] == "credit_cash"]
        self.assertEqual(len(credit_rows), 1)
        self.assertEqual(credit_rows[0]["amount"], 400.0)
        self.assertIn("INV-1", credit_rows[0]["description"])
        self.assertIn("Hotel", credit_rows[0]["detail"])
        self.assertIn("Bar", credit_rows[0]["detail"])

        hotel_entries = app_module._build_cash_ledger_entries(
            conn, "HBE", date(2026, 7, 1), date(2026, 7, 2), location="Hotel"
        )
        hotel_totals = app_module._cash_ledger_totals(hotel_entries)
        self.assertEqual(hotel_totals["credit_total"], 250.0)
        self.assertEqual(hotel_totals["available_total"], 350.0)
        conn.close()

    def test_export_cash_ledger_report_route_registered(self):
        rules = [rule.rule for rule in app_module.app.url_map.iter_rules()]
        self.assertIn("/accounts/cash-ledger/report", rules)

    def test_cash_ledger_purchase_entries_render_as_purchase(self):
        tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        tmp.close()
        db_path = tmp.name
        orig = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = db_path
        try:
            db_mod.init_db()
            conn = db_mod.get_db()
            try:
                conn.execute(
                    """INSERT INTO sales_updates
                       (company, location, sales_date, sales_entry_values, created_at, updated_at)
                       VALUES (?,?,?,?,datetime('now'),datetime('now'))""",
                    ("HBE", "Hotel", "2026-07-01", json.dumps({"actual_cash": 1000})),
                )
                conn.execute(
                    """INSERT INTO sales_update_expenses
                       (company, location, sales_date, description, amount,
                        payment_type, expense_code, entry_kind)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    ("HBE", "Hotel", "2026-07-01", "Some purchase", 120, "cash", "HBE-PU-1", "purchase"),
                )
                conn.commit()
            finally:
                conn.close()

            app = app_module.app
            app.config["TESTING"] = True
            client = app.test_client()
            user = {
                "id": 1,
                "username": "admin",
                "full_name": "Administrator",
                "is_admin": True,
                "is_active": True,
                "dashboard_access": set(),
                "stores_access": set(),
            }
            with mock.patch.object(app_module, "get_current_user", return_value=user):
                page = client.get("/accounts/cash-ledger?date_from=2026-07-01&date_to=2026-07-01&location=Hotel")
            html = page.get_data(as_text=True)
            self.assertIn('data-sort-value="Purchase"', html)
        finally:
            db_mod.DATABASE_PATH = orig
            try:
                os.unlink(db_path)
            except OSError:
                pass

    def test_export_cash_ledger_report_summary_and_line_items(self):
        import os
        import tempfile
        from io import BytesIO
        from unittest import mock

        from openpyxl import load_workbook

        tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        tmp.close()
        db_path = tmp.name
        orig = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = db_path
        try:
            db_mod.init_db()
            conn = db_mod.get_db()
            try:
                conn.execute(
                    """INSERT INTO sales_updates
                       (company, location, sales_date, sales_entry_values, created_at, updated_at)
                       VALUES (?,?,?,?,datetime('now'),datetime('now'))""",
                    ("HBE", "Hotel", "2026-07-01", json.dumps({"actual_cash": 1000})),
                )
                conn.execute(
                    """INSERT INTO sales_update_expenses
                       (company, location, sales_date, description, amount, payment_type, expense_code)
                       VALUES (?,?,?,?,?,?,?)""",
                    ("HBE", "Hotel", "2026-07-01", "Veggies", 100, "cash", "HBE-EX-1"),
                )
                conn.commit()
            finally:
                conn.close()

            app = app_module.app
            app.config["TESTING"] = True
            client = app.test_client()
            user = {
                "id": 1,
                "username": "admin",
                "full_name": "Administrator",
                "is_admin": True,
                "is_active": True,
                "dashboard_access": set(),
                "stores_access": set(),
            }
            with mock.patch.object(app_module, "get_current_user", return_value=user):
                export = client.get(
                    "/accounts/cash-ledger/report?date_from=2026-07-01&date_to=2026-07-01"
                )
            self.assertEqual(export.status_code, 200)
            self.assertIn(
                "spreadsheetml.sheet",
                export.headers.get("Content-Type", ""),
            )
            cd = export.headers.get("Content-Disposition") or ""
            self.assertIn("Hotel Bell Elite Cash Ledger", cd)
            self.assertIn("01 July 26.xlsx", cd)

            wb = load_workbook(BytesIO(export.data))
            self.assertEqual(wb.sheetnames, ["Summary", "Line Items"])
            summary = wb["Summary"]
            details = wb["Line Items"]
            self.assertTrue(
                (summary["A1"].value or "").startswith("Hotel Bell Elite — Cash Ledger")
            )
            self.assertEqual(summary["A2"].value, "Actual Cash")
            self.assertEqual(float(summary["B2"].value), 1000)
            self.assertEqual(summary["A8"].value, "Available Cash")
            self.assertEqual(float(summary["B8"].value), 900)
            self.assertEqual(summary["A10"].value, "Movement Type")
            self.assertEqual(summary["A10"].fill.fgColor.rgb, "FF315A78")
            self.assertIn("A1:B1", {str(r) for r in summary.merged_cells.ranges})

            col_a = [
                details.cell(row, 1).value
                for row in range(1, (details.max_row or 1) + 1)
            ]
            self.assertIn("Hotel Bell Elite — Cash Ledger - Actual Cash", col_a)
            self.assertIn("Hotel Bell Elite — Cash Ledger - Expense", col_a)
            expense_banner = col_a.index(
                "Hotel Bell Elite — Cash Ledger - Expense"
            )
            headers = [
                details.cell(expense_banner + 2, col).value for col in range(1, 8)
            ]
            self.assertEqual(
                headers,
                ["Date", "Type", "Detail", "ID", "Description", "Amount", "Balance"],
            )
            self.assertEqual(
                details.cell(expense_banner + 1, 1).fill.fgColor.rgb, "FF315A78"
            )
            self.assertEqual(details.cell(expense_banner + 3, 4).value, "HBE-EX-1")
        finally:
            db_mod.DATABASE_PATH = orig
            try:
                os.unlink(db_path)
            except OSError:
                pass

    def test_resolve_cash_ledger_date_range_defaults_to_fy(self):
        date_from, date_to, active = app_module._resolve_cash_ledger_date_range({})
        fy_start, today = db_mod.indian_fiscal_year_bounds()
        self.assertTrue(active)
        self.assertEqual(date_from, fy_start)
        self.assertEqual(date_to, today)

    def test_resolve_cash_ledger_date_range_with_params(self):
        date_from, date_to, active = app_module._resolve_cash_ledger_date_range({
            "date_from": "2026-07-01",
            "date_to": "2026-07-10",
        })
        self.assertTrue(active)
        self.assertEqual(date_from, date(2026, 7, 1))
        self.assertEqual(date_to, date(2026, 7, 10))

    def test_wide_explicit_range_includes_older_entries(self):
        conn = _memory_conn()
        conn.execute(
            "INSERT INTO sales_updates (company, location, sales_date, sales_entry_values) VALUES (?,?,?,?)",
            ("HBE", "Hotel", "2025-01-15", json.dumps({"actual_cash": 400})),
        )
        conn.execute(
            "INSERT INTO sales_updates (company, location, sales_date, sales_entry_values) VALUES (?,?,?,?)",
            ("HBE", "Hotel", "2026-07-01", json.dumps({"actual_cash": 100})),
        )
        conn.commit()
        all_entries = app_module._build_cash_ledger_entries(
            conn, "HBE", date(2000, 1, 1), date.today()
        )
        filtered = app_module._build_cash_ledger_entries(
            conn, "HBE", date(2026, 7, 1), date(2026, 7, 31)
        )
        self.assertEqual(len(all_entries), 2)
        self.assertEqual(len(filtered), 1)
        conn.close()

    def test_reject_cash_expense_over_available(self):
        conn = _memory_conn()
        conn.execute(
            "INSERT INTO suppliers (name) VALUES (?)",
            ("ABC Supplies",),
        )
        conn.execute(
            "INSERT INTO sales_updates (company, location, sales_date, sales_entry_values) VALUES (?,?,?,?)",
            ("HBE", "Hotel", "2026-07-01", json.dumps({"cash": 500, "actual_cash": 100})),
        )
        conn.commit()
        available = app_module._cash_ledger_available_as_of(
            conn, "HBE", date(2026, 7, 1)
        )
        self.assertEqual(available, 100.0)
        result, error = app_module._create_sales_expense(
            conn,
            {"is_admin": True},
            {
                "company": "HBE",
                "location": "Hotel",
                "date": "2026-07-01",
                "description": "Too much",
                "amount": 150,
                "payment_type": "cash",
                "category": "grocery",
                "supplier_id": 1,
            },
        )
        self.assertIsNone(result)
        self.assertIn("available cash", error.lower())
        result, error = app_module._create_sales_expense(
            conn,
            {"is_admin": True},
            {
                "company": "HBE",
                "location": "Hotel",
                "date": "2026-07-01",
                "description": "Within limit",
                "amount": 80,
                "payment_type": "cash",
                "category": "grocery",
                "supplier_id": 1,
            },
        )
        self.assertIsNone(error)
        self.assertIsNotNone(result)
        conn.close()

    def test_cash_expense_edit_excludes_existing_amount(self):
        conn = _memory_conn()
        conn.execute("INSERT INTO suppliers (name) VALUES (?)", ("ABC Supplies",))
        conn.execute(
            "INSERT INTO sales_updates (company, location, sales_date, sales_entry_values) VALUES (?,?,?,?)",
            ("HBE", "Hotel", "2026-07-01", json.dumps({"actual_cash": 100})),
        )
        conn.execute(
            """INSERT INTO sales_update_expenses
               (company, location, sales_date, description, amount, payment_type, expense_code, category, supplier_id)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            ("HBE", "Hotel", "2026-07-01", "Veggies", 60, "cash", "HBE-EX-1", "grocery", 1),
        )
        conn.commit()
        available = app_module._cash_ledger_available_as_of(
            conn, "HBE", date(2026, 7, 1), exclude_expense_id=1
        )
        self.assertEqual(available, 100.0)
        error = app_module._validate_cash_expense_against_available(
            conn, "HBE", "2026-07-01", 100, "cash", exclude_expense_id=1
        )
        self.assertIsNone(error)
        error = app_module._validate_cash_expense_against_available(
            conn, "HBE", "2026-07-01", 101, "cash", exclude_expense_id=1
        )
        self.assertIn("available cash", error.lower())
        conn.close()


class CashLedgerRouteTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        app_module.app.config["TESTING"] = True
        cls.client = app_module.app.test_client()

    def test_load_and_transfer_validation(self):
        with self.client.session_transaction() as sess:
            # Routes are protected by before_request; skip if session required.
            sess["user_id"] = 1

        # Without auth these still may redirect; exercise helpers via direct call pattern.
        # Validation is covered by posting through the view functions with flask request context.
        with app_module.app.test_request_context(
            "/accounts/cash-ledger/load",
            method="POST",
            json={"amount": 0, "description": "x"},
        ):
            resp = app_module.cash_ledger_load()
            if isinstance(resp, tuple):
                resp = resp[0]
            data = resp.get_json()
            self.assertFalse(data["ok"])
            self.assertIn("positive amount", data["error"].lower())

        with app_module.app.test_request_context(
            "/accounts/cash-ledger/transfer",
            method="POST",
            json={
                "amount": 10,
                "description": "x",
                "destination": "petty",
            },
        ):
            resp = app_module.cash_ledger_transfer()
            if isinstance(resp, tuple):
                resp = resp[0]
            data = resp.get_json()
            self.assertFalse(data["ok"])
            self.assertIn("Bank or Owner", data["error"])


if __name__ == "__main__":
    unittest.main()
