"""Menu Sales report — item-wise aggregation, filters, page, export."""

import os
import tempfile
import unittest
from unittest import mock

import db as db_mod


class MenuSalesReportTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db_path = self.tmp.name
        self._orig_path = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = self.db_path
        db_mod.init_db()

        import app as app_mod

        self.app_mod = app_mod
        self.app = app_mod.app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

        conn = db_mod.get_db()
        try:
            admin = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            self.admin_id = admin["id"]
            db_mod.ensure_pos_schema(conn)
            self.rest_cat = db_mod.save_pos_menu_category(
                conn, name="Mains", outlet=db_mod.POS_OUTLET_RESTAURANT
            )
            self.bar_cat = db_mod.save_pos_menu_category(
                conn, name="Spirits", outlet=db_mod.POS_OUTLET_BAR
            )
            self.butter = db_mod.save_pos_menu_item(
                conn,
                category_id=self.rest_cat["id"],
                name="Chicken Butter Masala",
                rate=320,
                outlet=db_mod.POS_OUTLET_RESTAURANT,
            )
            self.whisky = db_mod.save_pos_menu_item(
                conn,
                category_id=self.bar_cat["id"],
                name="Whisky Peg",
                rate=250,
                outlet=db_mod.POS_OUTLET_BAR,
                item_kind="liquor",
                menu_type="liquor",
            )
            conn.commit()
        finally:
            conn.close()

        self.user = {
            "id": self.admin_id,
            "username": "admin",
            "full_name": "Administrator",
            "is_admin": True,
            "is_active": True,
            "dashboard_access": set(),
            "stores_access": set(),
        }
        self._get_user_patch = mock.patch.object(
            app_mod, "get_current_user", return_value=self.user
        )
        self._get_user_patch.start()

    def tearDown(self):
        self._get_user_patch.stop()
        db_mod.DATABASE_PATH = self._orig_path
        try:
            os.unlink(self.db_path)
        except OSError:
            pass

    def _invoice(self, *, outlet, order_no, menu_id, name, rate, qty, settle=False):
        line_total = round(rate * qty, 2)
        payload = {
            "outlet": outlet,
            "orderNo": order_no,
            "savedAt": "2026-08-01 18:00:00",
            "orderType": "dine_in",
            "table": "T1" if outlet == "restaurant" else "B1",
            "captain": "",
            "customerName": "Guest",
            "customerMobile": "9000000000",
            "notes": "",
            "discountType": "pct",
            "discountValue": 0,
            "serviceType": "pct",
            "serviceValue": 0,
            "tipAmount": 0,
            "couponCode": "",
            "lines": [
                {
                    "uid": "1",
                    "menuId": menu_id,
                    "name": name,
                    "variant": "",
                    "rate": rate,
                    "qty": qty,
                }
            ],
            "totals": {
                "subtotal": line_total,
                "discount": 0,
                "discountType": "pct",
                "discountValue": 0,
                "gst": 0,
                "service": 0,
                "serviceType": "pct",
                "serviceValue": 0,
                "tip": 0,
                "roundOff": 0,
                "total": line_total,
            },
        }
        path = (
            "/bar-point-of-sale/api/invoices"
            if outlet == "bar"
            else "/point-of-sale/api/invoices"
        )
        res = self.client.post(path, json=payload)
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        body = res.get_json() or {}
        self.assertTrue(body.get("ok"), body)
        invoice_id = (body.get("invoice") or {}).get("id")
        if settle and invoice_id:
            settle_path = (
                f"/bar-point-of-sale/api/invoices/{invoice_id}/settle"
                if outlet == "bar"
                else f"/point-of-sale/api/invoices/{invoice_id}/settle"
            )
            settle_res = self.client.post(
                settle_path,
                json={
                    "payment_date": "2026-08-01",
                    "payment_splits": [
                        {"payment_method": "cash", "amount": line_total}
                    ],
                },
            )
            self.assertEqual(
                settle_res.status_code, 200, settle_res.get_data(as_text=True)
            )
        return body

    def test_aggregation_order_count_qty_and_value(self):
        self._invoice(
            outlet="restaurant",
            order_no="MSR-1",
            menu_id=self.butter["id"],
            name="Chicken Butter Masala",
            rate=320,
            qty=1,
        )
        self._invoice(
            outlet="restaurant",
            order_no="MSR-2",
            menu_id=self.butter["id"],
            name="Chicken Butter Masala",
            rate=320,
            qty=2,
        )
        conn = db_mod.get_db()
        try:
            rows = db_mod.list_pos_menu_sales(
                conn, date_from="2026-08-01", date_to="2026-08-01", outlet="restaurant"
            )
            butter = next(
                (r for r in rows if r["item_name"] == "Chicken Butter Masala"), None
            )
            self.assertIsNotNone(butter)
            self.assertEqual(butter["order_count"], 2)
            self.assertEqual(butter["qty_sold"], 3)
            self.assertEqual(butter["sale_value"], 960.0)
            self.assertEqual(butter["rate"], 320.0)
            self.assertEqual(butter["category_name"], "Mains")
            self.assertEqual(butter["outlet"], "restaurant")

            groups = db_mod.group_pos_menu_sales_by_category(rows)
            self.assertEqual(len(groups), 1)
            self.assertEqual(groups[0]["category_name"], "Mains")
            self.assertEqual(groups[0]["qty_sum"], 3)
            self.assertEqual(groups[0]["sale_sum"], 960.0)
            self.assertEqual(len(groups[0]["item_rows"]), 1)

            kpis = db_mod.pos_menu_sales_kpis(
                rows,
                conn,
                date_from="2026-08-01",
                date_to="2026-08-01",
                outlet="restaurant",
            )
            self.assertEqual(kpis["invoice_count"], 2)
            self.assertEqual(kpis["item_count"], 1)
            self.assertEqual(kpis["qty_sum"], 3)
            self.assertEqual(kpis["sale_value_sum"], 960.0)
        finally:
            conn.close()

    def test_outlet_and_category_filters(self):
        self._invoice(
            outlet="restaurant",
            order_no="MSR-R",
            menu_id=self.butter["id"],
            name="Chicken Butter Masala",
            rate=320,
            qty=1,
        )
        self._invoice(
            outlet="bar",
            order_no="",
            menu_id=self.whisky["id"],
            name="Whisky Peg",
            rate=250,
            qty=2,
        )
        conn = db_mod.get_db()
        try:
            all_rows = db_mod.list_pos_menu_sales(
                conn, date_from="2026-08-01", date_to="2026-08-01", outlet=None
            )
            names = {r["item_name"] for r in all_rows}
            self.assertIn("Chicken Butter Masala", names)
            self.assertIn("Whisky Peg", names)
            mixed_groups = db_mod.group_pos_menu_sales_by_category(
                all_rows, include_outlet_label=True
            )
            labels = {g["category_name"] for g in mixed_groups}
            self.assertIn("Mains (Restaurant)", labels)
            self.assertIn("Spirits (Bar)", labels)

            bar_only = db_mod.list_pos_menu_sales(
                conn, date_from="2026-08-01", date_to="2026-08-01", outlet="bar"
            )
            self.assertEqual(len(bar_only), 1)
            self.assertEqual(bar_only[0]["item_name"], "Whisky Peg")
            self.assertEqual(bar_only[0]["qty_sold"], 2)

            cat_rows = db_mod.list_pos_menu_sales(
                conn,
                date_from="2026-08-01",
                date_to="2026-08-01",
                category_id=self.rest_cat["id"],
            )
            self.assertEqual(len(cat_rows), 1)
            self.assertEqual(cat_rows[0]["item_name"], "Chicken Butter Masala")
        finally:
            conn.close()

    def test_settlement_filter(self):
        self._invoice(
            outlet="restaurant",
            order_no="MSR-OPEN",
            menu_id=self.butter["id"],
            name="Chicken Butter Masala",
            rate=320,
            qty=1,
            settle=False,
        )
        self._invoice(
            outlet="restaurant",
            order_no="MSR-PAID",
            menu_id=self.butter["id"],
            name="Chicken Butter Masala",
            rate=320,
            qty=1,
            settle=True,
        )
        conn = db_mod.get_db()
        try:
            unsettled = db_mod.list_pos_menu_sales(
                conn,
                date_from="2026-08-01",
                date_to="2026-08-01",
                outlet="restaurant",
                settlement="unsettled",
            )
            settled = db_mod.list_pos_menu_sales(
                conn,
                date_from="2026-08-01",
                date_to="2026-08-01",
                outlet="restaurant",
                settlement="settled",
            )
            self.assertEqual(len(unsettled), 1)
            self.assertEqual(unsettled[0]["order_count"], 1)
            self.assertEqual(len(settled), 1)
            self.assertEqual(settled[0]["order_count"], 1)
        finally:
            conn.close()

    def test_page_and_export(self):
        self._invoice(
            outlet="restaurant",
            order_no="MSR-PAGE",
            menu_id=self.butter["id"],
            name="Chicken Butter Masala",
            rate=320,
            qty=1,
        )
        page = self.client.get("/reports/sales/menu")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn("Menu Insights", html)
        self.assertIn('id="menu-sales-report-page"', html)
        self.assertIn("Chicken Butter Masala", html)
        self.assertIn("Item Wise Sales Report", html)
        self.assertIn("ItemName", html)
        self.assertIn("Rate", html)
        self.assertIn("Net Amount", html)
        self.assertIn('class="pl-sortable msr-col-item"', html)
        self.assertIn('data-sort="item"', html)
        self.assertIn('data-sort="qty"', html)
        self.assertIn('data-sort="rate"', html)
        self.assertIn('data-sort="sale"', html)
        self.assertIn("data-sort-row", html)
        self.assertIn("Group Total", html)
        self.assertIn("Total Sales", html)
        self.assertIn("Mains", html)
        fy_start, today = db_mod.indian_fiscal_year_bounds()
        self.assertIn(f'value="{fy_start.isoformat()}"', html)
        self.assertIn(f'value="{today.isoformat()}"', html)
        self.assertIn("Qty", html)
        self.assertNotIn(">Orders<", html)
        self.assertNotIn("msr-col-outlet", html)
        self.assertIn('aria-label="Back to Reports"', html)
        self.assertIn('class="su-page-back"', html)
        self.assertIn('href="/reports"', html)
        self.assertIn("Export Excel", html)
        self.assertIn(
            f'download="Hotel Bell Elite Menu Sales {fy_start.day:02d}',
            html,
        )

        export = self.client.get("/reports/sales/menu/export")
        self.assertEqual(export.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            export.content_type or "",
        )
        self.assertTrue(export.data[:2] == b"PK")
        cd = export.headers.get("Content-Disposition") or ""
        self.assertIn("Hotel Bell Elite Menu Sales", cd)
        self.assertTrue(cd.lower().endswith('.xlsx"') or ".xlsx" in cd.lower())

        from io import BytesIO
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(export.data))
        self.assertEqual(wb.sheetnames, ["By Category", "By Qty Sold"])
        ws = wb["By Category"]
        self.assertEqual(ws["A1"].value, "Hotel Bell Elite — Item Wise Sales Report")
        self.assertIn("A1:D1", {str(r) for r in ws.merged_cells.ranges})
        self.assertIn("A2:D2", {str(r) for r in ws.merged_cells.ranges})
        self.assertTrue(str(ws["A2"].value or "").startswith("From "))
        self.assertEqual(
            [ws.cell(3, c).value for c in range(1, 5)],
            ["ItemName", "Qty", "Rate", "Net Amount"],
        )
        self.assertEqual(ws.cell(4, 1).value, "Category Summary")
        self.assertEqual(ws.cell(5, 1).value, "Mains (Restaurant)")
        self.assertEqual(ws.cell(5, 2).value, 1)
        self.assertEqual(ws.cell(5, 4).value, 320)
        self.assertEqual(ws.cell(6, 1).value, "Category Total")
        self.assertEqual(ws.cell(6, 2).value, 1)
        self.assertEqual(ws.cell(6, 4).value, 320)
        self.assertEqual(ws.cell(8, 1).value, "Item Wise Sales")
        self.assertEqual(ws.cell(9, 1).value, "Mains (Restaurant)")
        self.assertEqual(ws.cell(10, 1).value, "Chicken Butter Masala")
        self.assertEqual(ws.cell(10, 2).value, 1)
        self.assertEqual(ws.cell(10, 3).value, 320)
        self.assertEqual(ws.cell(10, 4).value, 320)
        self.assertEqual(ws.cell(11, 1).value, "Group Total")
        self.assertEqual(ws.cell(11, 2).value, 1)
        self.assertEqual(ws.cell(11, 4).value, 320)
        self.assertEqual(ws.cell(12, 1).value, "Total Sales")
        self.assertEqual(ws.cell(12, 2).value, 1)
        self.assertEqual(ws.cell(12, 4).value, 320)
        self.assertEqual(ws["A1"].alignment.horizontal, "center")
        self.assertEqual(ws.cell(3, 1).alignment.horizontal, "center")
        self.assertEqual(ws.cell(10, 1).alignment.horizontal, "left")
        self.assertEqual(ws.cell(10, 3).alignment.horizontal, "right")
        self.assertEqual(ws["A1"].fill.fgColor.rgb, "FF315A78")
        self.assertEqual(ws.cell(3, 1).fill.fgColor.rgb, "FF315A78")
        self.assertEqual(ws.cell(3, 4).fill.fgColor.rgb, "FF315A78")
        self.assertEqual(ws["A1"].font.color.rgb, "FFFFFFFF")
        self.assertEqual(ws.cell(3, 1).font.color.rgb, "FFFFFFFF")
        # Data rows stay white (no solid fill).
        data_fill = ws.cell(10, 1).fill
        self.assertTrue(
            data_fill.fgColor is None
            or data_fill.patternType in (None, "none")
            or getattr(data_fill.fgColor, "rgb", None) in (None, "00000000", "00FFFFFF")
        )
        self.assertIsNotNone(ws.cell(10, 1).border.left.style)
        self.assertTrue(ws.cell(4, 1).font.bold)
        self.assertTrue(ws.cell(6, 1).font.bold)
        self.assertTrue(ws.cell(8, 1).font.bold)
        self.assertTrue(ws.cell(9, 1).font.bold)
        self.assertTrue(ws.cell(11, 1).font.bold)
        self.assertTrue(ws.cell(12, 1).font.bold)

        ranked = wb["By Qty Sold"]
        self.assertEqual(ranked["A1"].value, "Hotel Bell Elite — Item Wise Sales Report")
        self.assertEqual(
            [ranked.cell(3, c).value for c in range(1, 5)],
            ["ItemName", "Qty", "Rate", "Net Amount"],
        )
        ranked_names = [
            ranked.cell(r, 1).value
            for r in range(4, ranked.max_row + 1)
        ]
        self.assertNotIn("Mains (Restaurant)", ranked_names)
        self.assertNotIn("Group Total", ranked_names)
        self.assertEqual(ranked.cell(4, 1).value, "Chicken Butter Masala")
        self.assertEqual(ranked.cell(4, 2).value, 1)
        self.assertEqual(ranked.cell(4, 3).value, 320)
        self.assertEqual(ranked.cell(4, 4).value, 320)
        self.assertEqual(ranked.cell(5, 1).value, "Total Sales")
        self.assertEqual(ranked.cell(5, 2).value, 1)
        self.assertEqual(ranked.cell(5, 4).value, 320)
        self.assertTrue(ranked.cell(5, 1).font.bold)
        self.assertEqual(ranked.cell(3, 1).fill.fgColor.rgb, "FF315A78")

        dated = self.client.get(
            "/reports/sales/menu/export?date_from=2026-08-01&date_to=2026-08-08"
        )
        self.assertEqual(dated.status_code, 200)
        dated_cd = dated.headers.get("Content-Disposition") or ""
        self.assertIn(
            "Hotel Bell Elite Menu Sales 01 August 26 to 08 August 26.xlsx", dated_cd
        )

    def test_export_ranked_sheet_sorts_most_sold_first(self):
        self._invoice(
            outlet="restaurant",
            order_no="MSR-RANK-R",
            menu_id=self.butter["id"],
            name="Chicken Butter Masala",
            rate=320,
            qty=1,
        )
        self._invoice(
            outlet="bar",
            order_no="MSR-RANK-B",
            menu_id=self.whisky["id"],
            name="Whisky Peg",
            rate=250,
            qty=4,
        )
        export = self.client.get(
            "/reports/sales/menu/export?date_from=2026-08-01&date_to=2026-08-01"
        )
        self.assertEqual(export.status_code, 200)
        from io import BytesIO
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(export.data))
        grouped = wb["By Category"]
        self.assertEqual(grouped.cell(4, 1).value, "Category Summary")
        self.assertEqual(grouped.cell(5, 1).value, "Mains (Restaurant)")
        self.assertEqual(grouped.cell(6, 1).value, "Spirits (Bar)")
        self.assertEqual(grouped.cell(7, 1).value, "Category Total")
        self.assertEqual(grouped.cell(7, 2).value, 5)
        self.assertEqual(grouped.cell(9, 1).value, "Item Wise Sales")
        self.assertIn(
            "Group Total", [grouped.cell(r, 1).value for r in range(9, 20)]
        )

        ranked = wb["By Qty Sold"]
        self.assertEqual(ranked.cell(4, 1).value, "Whisky Peg")
        self.assertEqual(ranked.cell(4, 2).value, 4)
        self.assertEqual(ranked.cell(5, 1).value, "Chicken Butter Masala")
        self.assertEqual(ranked.cell(5, 2).value, 1)
        self.assertEqual(ranked.cell(6, 1).value, "Total Sales")
        self.assertEqual(ranked.cell(6, 2).value, 5)
        self.assertNotIn(
            "Group Total",
            [ranked.cell(r, 1).value for r in range(1, ranked.max_row + 1)],
        )

    def test_imported_snapshot_keeps_menu_item_category(self):
        conn = db_mod.get_db()
        try:
            db_mod.import_settled_pos_invoice_snapshot(
                conn,
                {
                    "order_no": "JUL26/R/TEST",
                    "outlet": "restaurant",
                    "order_date": "2026-07-15",
                    "saved_at": "2026-07-15 19:10:00",
                    "settled_at": "2026-07-15 19:40:00",
                    "customer_name": "Walk-in",
                    "subtotal": 640.0,
                    "gst_amount": 32.0,
                    "grand_total": 672.0,
                    "lines": [
                        {
                            "menu_item_id": self.butter["id"],
                            "name": "Chicken Butter Masala",
                            "rate": 320,
                            "qty": 2,
                            "line_total": 640,
                        }
                    ],
                    "payments": [
                        {
                            "payment_method": "upi",
                            "amount": 672.0,
                            "payment_date": "2026-07-15",
                        }
                    ],
                },
            )
            conn.commit()
            rows = db_mod.list_pos_menu_sales(
                conn, date_from="2026-07-01", date_to="2026-07-31", outlet="restaurant"
            )
            butter = next(
                (r for r in rows if r["item_name"] == "Chicken Butter Masala"), None
            )
            self.assertIsNotNone(butter)
            self.assertEqual(butter["category_name"], "Mains")
            self.assertEqual(butter["qty_sold"], 2)
            self.assertEqual(butter["sale_value"], 640.0)
        finally:
            conn.close()


if __name__ == "__main__":
    unittest.main()
