"""KOT report — statuses, outlets, hub access, export, from_hub."""

import os
import tempfile
import unittest
from datetime import date
from io import BytesIO
from unittest import mock

import db as db_mod
from kot_report import (
    STATUS_CANCELLED,
    STATUS_INVOICE_GENERATED,
    STATUS_OPEN,
    build_kot_report,
    build_kot_workbook,
    classify_kot_status,
    kot_display_no,
    kot_invoice_no,
)
from openpyxl import load_workbook
from workspace_access import get_endpoint_dashboard_module


class KotReportTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db_path = self.tmp.name
        self._orig_path = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = self.db_path
        db_mod.init_db()

        import app as app_mod

        self.app_mod = app_mod
        self.app = app_mod.app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

        conn = db_mod.get_db()
        try:
            admin = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            self.admin_id = admin["id"]
        finally:
            conn.close()

        self.user = {
            "id": self.admin_id,
            "username": "admin",
            "full_name": "Administrator",
            "is_admin": True,
            "is_active": True,
            "dashboard_access": set(),
            "stores_access": set(),
        }
        self._get_user_patch = mock.patch.object(
            app_mod, "get_current_user", return_value=self.user
        )
        self._get_user_patch.start()

        put = self.client.put(
            "/point-of-sale/api/floor",
            json={
                "areas": [{"id": "area_1", "type": "area", "name": "Main"}],
                "tables": [
                    {
                        "id": "t1",
                        "type": "table",
                        "name": "T1",
                        "seats": 4,
                        "shape": "square",
                        "status": "available",
                        "areaId": "area_1",
                    },
                    {
                        "id": "t2",
                        "type": "table",
                        "name": "T2",
                        "seats": 4,
                        "shape": "square",
                        "status": "available",
                        "areaId": "area_1",
                    },
                ],
            },
        )
        self.assertEqual(put.status_code, 200)
        put_bar = self.client.put(
            "/bar-point-of-sale/api/floor",
            json={
                "areas": [{"id": "area_1", "type": "area", "name": "Bar"}],
                "tables": [
                    {
                        "id": "b1",
                        "type": "table",
                        "name": "B1",
                        "seats": 2,
                        "shape": "square",
                        "status": "available",
                        "areaId": "area_1",
                    }
                ],
            },
        )
        self.assertEqual(put_bar.status_code, 200)

    def tearDown(self):
        self._get_user_patch.stop()
        db_mod.DATABASE_PATH = self._orig_path
        try:
            os.unlink(self.db_path)
        except OSError:
            pass

    def _payload(self, order_no, table, *, outlet="restaurant", kot_send=True, **overrides):
        data = {
            "orderNo": order_no,
            "savedAt": "2026-08-24 18:00:00",
            "orderType": "dine_in",
            "table": table,
            "captain": "Ravi",
            "customerName": "Guest",
            "customerMobile": "9876543210",
            "notes": "",
            "kotSend": kot_send,
            "outlet": outlet,
            "discountType": "pct",
            "discountValue": 0,
            "serviceType": "pct",
            "serviceValue": 0,
            "tipAmount": 0,
            "couponCode": "",
            "lines": [
                {
                    "uid": "1",
                    "menuId": None,
                    "name": "Filter Coffee" if outlet == "restaurant" else "Gin Tonic",
                    "variant": "",
                    "rate": 100,
                    "qty": 2,
                    "kotSentQty": 2 if kot_send else 0,
                }
            ],
            "totals": {
                "subtotal": 200,
                "discount": 0,
                "gst": 10,
                "service": 0,
                "tip": 0,
                "roundOff": 0,
                "total": 210,
            },
        }
        data.update(overrides)
        return data

    def _api_base(self, outlet):
        return "/bar-point-of-sale" if outlet == "bar" else "/point-of-sale"

    def test_kot_display_no(self):
        self.assertEqual(kot_display_no("ORD-2608-0001"), "KOT/2608-0001")
        self.assertEqual(kot_display_no("KOT-1"), "KOT/1")
        self.assertEqual(kot_display_no("SPC/26-27/1"), "KOT/SPC/26-27/1")
        self.assertEqual(kot_display_no("SPC/26-27/1", "KOT/SPC/26-27/9"), "KOT/SPC/26-27/9")
        self.assertEqual(kot_display_no("KOT/SPC/26-27/1"), "KOT/SPC/26-27/1")

    def test_classify_status(self):
        self.assertEqual(
            classify_kot_status(
                {"is_active": 1, "status": "open", "customer_bill_sent": 0}
            )[0],
            STATUS_OPEN,
        )
        self.assertEqual(
            classify_kot_status(
                {"is_active": 1, "status": "open", "customer_bill_sent": 1}
            )[0],
            STATUS_INVOICE_GENERATED,
        )
        self.assertEqual(
            classify_kot_status(
                {"is_active": 1, "status": "cancelled", "customer_bill_sent": 1}
            )[0],
            STATUS_CANCELLED,
        )
        self.assertEqual(
            classify_kot_status(
                {"is_active": 0, "status": "open", "customer_bill_sent": 0}
            )[0],
            STATUS_CANCELLED,
        )

    def test_cancelled_rows_never_expose_invoice_no(self):
        """Bill-sent then cancelled must still blank invoice_no in the report."""
        cancelled_with_bill = {
            "status": "cancelled",
            "is_active": 1,
            "customer_bill_sent": 1,
        }
        soft_deleted_billed = {
            "status": "open",
            "is_active": 0,
            "customer_bill_sent": 1,
        }
        for raw in (cancelled_with_bill, soft_deleted_billed):
            key, _ = classify_kot_status(raw)
            self.assertEqual(key, STATUS_CANCELLED)
            self.assertEqual(
                kot_invoice_no(key, "SPC/26-27/71", bill_sent=True),
                "",
            )
        self.assertEqual(
            kot_invoice_no(
                STATUS_INVOICE_GENERATED, "SPC/26-27/71", bill_sent=True
            ),
            "SPC/26-27/71",
        )
        self.assertEqual(
            kot_invoice_no(STATUS_OPEN, "ORD-2608-1", bill_sent=False),
            "",
        )

    def test_export_summary_matches_purchase_expense_fonts(self):
        wb = build_kot_workbook(
            {
                "rows": [],
                "kpis": {
                    "total": 3,
                    "open": 1,
                    "invoice_generated": 1,
                    "cancelled": 1,
                    "restaurant": 2,
                    "bar": 1,
                    "sent_qty": 5,
                },
            },
            title_date=" (01 April 26 to 25 August 26)",
        )
        summary = wb["Summary"]
        self.assertEqual(summary["A1"].fill.fgColor.rgb, "FF315A78")
        self.assertEqual(summary["A1"].font.name, "Calibri")
        self.assertEqual(summary["A1"].font.size, 16)
        self.assertTrue(summary["A1"].font.bold)
        self.assertEqual(summary["A1"].font.color.rgb, "FFFFFFFF")
        self.assertIn("A1:B1", {str(r) for r in summary.merged_cells.ranges})
        self.assertEqual(summary["A2"].value, "Total KOTs")
        self.assertEqual(summary["A2"].font.name, "Calibri")
        self.assertEqual(summary["A2"].font.size, 12)
        self.assertTrue(summary["A2"].font.bold)
        self.assertEqual(summary["A3"].font.name, "Calibri")
        self.assertEqual(summary["A3"].font.size, 12)
        self.assertFalse(summary["A3"].font.bold)
        self.assertEqual(wb["KOTs"].cell(1, 1).fill.fgColor.rgb, "FF315A78")
        self.assertEqual(wb["KOTs"].cell(1, 1).font.color.rgb, "FFFFFFFF")
        self.assertEqual(wb["KOTs"].cell(1, 10).value, "Cancel reason")

    def test_restaurant_and_bar_rows_with_statuses(self):
        # Bill first so Generate Invoice frees T1 for later KOTs.
        bill_save = self.client.post(
            f"{self._api_base('restaurant')}/api/invoices",
            json=self._payload("ORD-2608-KOT-BILL", "T1", outlet="restaurant"),
        )
        self.assertEqual(bill_save.status_code, 200, bill_save.get_data(as_text=True))
        bill_gen = self.client.post(
            f"{self._api_base('restaurant')}/api/invoices",
            json=self._payload(
                "ORD-2608-KOT-BILL",
                "T1",
                outlet="restaurant",
                customerBill=True,
            ),
        )
        self.assertEqual(bill_gen.status_code, 200, bill_gen.get_data(as_text=True))
        self.assertTrue((bill_gen.get_json().get("invoice") or {}).get("customer_bill_sent"))

        soft_save = self.client.post(
            f"{self._api_base('restaurant')}/api/invoices",
            json=self._payload("ORD-2608-KOT-SOFT", "T2", outlet="restaurant"),
        )
        self.assertEqual(soft_save.status_code, 200, soft_save.get_data(as_text=True))
        soft_id = soft_save.get_json()["invoice"]["id"]
        conn = db_mod.get_db()
        try:
            db_mod.soft_delete_pos_invoice(conn, soft_id)
            db_mod._pos_mark_table_available(conn, "T2", db_mod.POS_OUTLET_RESTAURANT)
            conn.commit()
        finally:
            conn.close()

        open_save = self.client.post(
            f"{self._api_base('restaurant')}/api/invoices",
            json=self._payload("ORD-2608-KOT-OPEN", "T2", outlet="restaurant"),
        )
        self.assertEqual(open_save.status_code, 200, open_save.get_data(as_text=True))

        cancel_save = self.client.post(
            f"{self._api_base('bar')}/api/invoices",
            json=self._payload("ORD-2608-KOT-CXL", "B1", outlet="bar"),
        )
        self.assertEqual(cancel_save.status_code, 200, cancel_save.get_data(as_text=True))
        cancel_id = cancel_save.get_json()["invoice"]["id"]
        conn = db_mod.get_db()
        try:
            db_mod.cancel_pos_invoice(conn, cancel_id, reason="Guest left")
            conn.commit()
        finally:
            conn.close()

        conn = db_mod.get_db()
        try:
            payload = build_kot_report(
                conn,
                date_from=date(2026, 8, 1),
                date_to=date(2026, 8, 31),
                outlet="all",
            )
        finally:
            conn.close()

        by_order = {r["order_no"]: r for r in payload["rows"]}
        self.assertIn("ORD-2608-KOT-OPEN", by_order)
        self.assertEqual(by_order["ORD-2608-KOT-OPEN"]["status"], STATUS_OPEN)
        self.assertEqual(by_order["ORD-2608-KOT-OPEN"]["outlet"], "restaurant")
        self.assertRegex(by_order["ORD-2608-KOT-OPEN"]["kot_no"], r"^KOT/SPC/\d{2}-\d{2}/\d+$")

        bill_row = next(
            (r for r in payload["rows"] if r["status"] == STATUS_INVOICE_GENERATED),
            None,
        )
        self.assertIsNotNone(bill_row)
        self.assertTrue(bill_row["invoice_no"] or bill_row["order_no"])

        self.assertIn("ORD-2608-KOT-CXL", by_order)
        self.assertEqual(by_order["ORD-2608-KOT-CXL"]["status"], STATUS_CANCELLED)
        self.assertEqual(by_order["ORD-2608-KOT-CXL"]["outlet"], "bar")
        self.assertEqual(by_order["ORD-2608-KOT-CXL"]["invoice_no"], "")
        self.assertEqual(by_order["ORD-2608-KOT-CXL"]["cancel_reason"], "Guest left")

        self.assertIn("ORD-2608-KOT-SOFT", by_order)
        self.assertEqual(by_order["ORD-2608-KOT-SOFT"]["status"], STATUS_CANCELLED)
        self.assertEqual(by_order["ORD-2608-KOT-SOFT"]["invoice_no"], "")

        self.assertGreaterEqual(payload["kpis"]["total"], 4)
        self.assertGreaterEqual(payload["kpis"]["open"], 1)
        self.assertGreaterEqual(payload["kpis"]["invoice_generated"], 1)
        self.assertGreaterEqual(payload["kpis"]["cancelled"], 2)
        self.assertGreaterEqual(payload["kpis"]["restaurant"], 1)
        self.assertGreaterEqual(payload["kpis"]["bar"], 1)

        conn = db_mod.get_db()
        try:
            bar_only = build_kot_report(
                conn,
                date_from=date(2026, 8, 1),
                date_to=date(2026, 8, 31),
                outlet="bar",
            )
        finally:
            conn.close()
        self.assertTrue(all(r["outlet"] == "bar" for r in bar_only["rows"]))
        self.assertIn("ORD-2608-KOT-CXL", {r["order_no"] for r in bar_only["rows"]})

    def test_kot_endpoints_map_and_access_gate(self):
        self.assertEqual(get_endpoint_dashboard_module("sales_report_kot"), "reports")
        self.assertEqual(
            get_endpoint_dashboard_module("sales_report_kot_export"), "reports"
        )

        viewer = {
            "id": self.admin_id,
            "username": "viewer",
            "full_name": "Viewer",
            "is_admin": False,
            "is_active": True,
            "dashboard_access": {"hotel_rooms"},
            "reports_access": set(),
            "stores_access": set(),
        }
        with mock.patch.object(self.app_mod, "get_current_user", return_value=viewer):
            denied = self.client.get("/reports/sales/kot")
            self.assertIn(denied.status_code, (302, 403))

        viewer["dashboard_access"] = {"reports"}
        viewer["reports_access"] = {"kot"}
        with mock.patch.object(self.app_mod, "get_current_user", return_value=viewer):
            allowed = self.client.get("/reports/sales/kot")
            self.assertEqual(allowed.status_code, 200)
