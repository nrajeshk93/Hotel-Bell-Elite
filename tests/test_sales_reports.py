"""Hotel / Restaurant / Bar Sales report pages — hub, pages, export, access."""

import os
import tempfile
import unittest
from datetime import date, datetime, timedelta
from io import BytesIO
from unittest import mock

import db as db_mod
from reports import (
    REPORT_DEFINITIONS,
    build_reports_dashboard,
    report_export_filename,
    report_export_month_filename,
)
from workspace_access import get_endpoint_dashboard_module


class SalesReportsTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db_path = self.tmp.name
        self._orig_path = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = self.db_path
        db_mod.init_db()

        import app as app_mod

        self.app_mod = app_mod
        self.app = app_mod.app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

        conn = db_mod.get_db()
        try:
            admin = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            self.admin_id = admin["id"]
        finally:
            conn.close()

        self.user = {
            "id": self.admin_id,
            "username": "admin",
            "full_name": "Administrator",
            "is_admin": True,
            "is_active": True,
            "dashboard_access": set(),
            "stores_access": set(),
        }
        self._get_user_patch = mock.patch.object(
            app_mod, "get_current_user", return_value=self.user
        )
        self._get_user_patch.start()

    def tearDown(self):
        self._get_user_patch.stop()
        db_mod.DATABASE_PATH = self._orig_path
        try:
            os.unlink(self.db_path)
        except OSError:
            pass

    def _pos_payload(self, *, outlet="restaurant", order_no="ORD-SR-0001", total=500, **overrides):
        data = {
            "outlet": outlet,
            "orderNo": order_no,
            "savedAt": "2026-07-22 18:00:00",
            "orderType": "dine_in",
            "table": "T1",
            "captain": "",
            "customerName": "Sales Guest",
            "customerMobile": "9876543210",
            "notes": "",
            "discountType": "pct",
            "discountValue": 0,
            "serviceType": "pct",
            "serviceValue": 0,
            "tipAmount": 0,
            "couponCode": "",
            "lines": [
                {
                    "uid": "1",
                    "menuId": None,
                    "name": "Filter Coffee",
                    "variant": "Hot",
                    "rate": 100,
                    "qty": 2,
                },
                {
                    "uid": "2",
                    "menuId": None,
                    "name": "Masala Dosa",
                    "variant": "",
                    "rate": 150,
                    "qty": 2,
                },
            ],
            "totals": {
                "subtotal": 500,
                "discount": 0,
                "discountType": "pct",
                "discountValue": 0,
                "gst": 25,
                "service": 0,
                "serviceType": "pct",
                "serviceValue": 0,
                "tip": 0,
                "roundOff": 0,
                "total": total,
            },
        }
        data.update(overrides)
        return data

    def test_sales_definitions_and_hub_cards(self):
        sales = [r for r in REPORT_DEFINITIONS if r.get("category") == "sales"]
        self.assertEqual(
            [r["id"] for r in sales],
            [
                "hotel_sales",
                "agency_billing",
                "manager_insight",
                "restaurant_sales",
                "menu_sales",
                "unit_insights",
                "customer_insights",
            ],
        )
        self.assertEqual(sales[0]["view_route"], "sales_report_hotel")
        self.assertEqual(sales[1]["view_route"], "sales_report_agency_billing")
        self.assertEqual(sales[2]["view_route"], "sales_report_manager_insight")
        self.assertEqual(sales[3]["view_route"], "sales_report_restaurant")
        self.assertEqual(sales[4]["view_route"], "sales_report_menu")
        self.assertEqual(sales[5]["view_route"], "sales_report_unit_insights")
        self.assertEqual(sales[6]["view_route"], "sales_report_customer_insights")

        restaurant = [r for r in REPORT_DEFINITIONS if r.get("category") == "restaurant"]
        self.assertEqual([r["id"] for r in restaurant], ["menu_margin", "meal_plan", "kot"])
        self.assertEqual(restaurant[1]["view_route"], "sales_report_meal_plan")
        self.assertEqual(restaurant[2]["view_route"], "sales_report_kot")

        with self.app.test_request_context():
            from flask import url_for

            payload = build_reports_dashboard(url_for)
        sales_section = next(
            (s for s in payload["report_sections"] if s["key"] == "sales"), None
        )
        self.assertIsNotNone(sales_section)
        self.assertEqual(sales_section["count"], 7)
        names = [r["name"] for r in sales_section["reports"]]
        self.assertEqual(
            names,
            [
                "Hotel Sales",
                "Agency Ledger",
                "Manager Insight",
                "Sales - Restaurant & Bar",
                "Menu Insights",
                "Unit Insight",
                "Customer Insights",
            ],
        )
        restaurant_section = next(
            (s for s in payload["report_sections"] if s["key"] == "restaurant"), None
        )
        self.assertIsNotNone(restaurant_section)
        self.assertEqual(restaurant_section["count"], 3)
        self.assertEqual(
            [r["name"] for r in restaurant_section["reports"]],
            ["Menu & Margin", "Meal Plan", "KOT"],
        )

        hub = self.client.get("/reports")
        self.assertEqual(hub.status_code, 200)
        html = hub.get_data(as_text=True)
        self.assertIn("Hotel Sales", html)
        self.assertIn("Agency Ledger", html)
        self.assertIn("Manager Insight", html)
        self.assertIn("Meal Plan", html)
        self.assertIn('data-report-id="meal_plan"', html)
        self.assertIn("KOT", html)
        self.assertIn('data-report-id="kot"', html)
        meal_card = html[html.find('data-report-id="meal_plan"') :][:280]
        self.assertIn('data-report-category="restaurant"', meal_card)
        self.assertIn("Sales - Restaurant &amp; Bar", html)
        self.assertNotIn("Bar Sales", html)
        self.assertNotIn('data-report-id="bar_sales"', html)
        self.assertIn("Menu Insights", html)
        self.assertIn("Unit Insight", html)
        self.assertIn("Customer Insights", html)
        self.assertIn('data-report-category="sales"', html)
        self.assertIn("/reports/sales/hotel", html)
        self.assertIn("/reports/sales/agency-billing", html)
        self.assertIn("/reports/sales/manager-insight", html)
        self.assertIn("/reports/sales/meal-plan", html)
        self.assertIn("/reports/sales/kot", html)
        self.assertIn("/reports/sales/restaurant", html)
        self.assertNotIn("/reports/sales/bar\"", html)
        self.assertIn("/reports/sales/menu", html)
        self.assertIn("/reports/sales/units", html)
        self.assertIn("/reports/sales/customer-insights", html)

    def test_report_export_filename_matches_menu_sales_style(self):
        self.assertEqual(
            report_export_filename(
                "Menu Sales",
                date_from=date(2026, 8, 1),
                date_to=date(2026, 8, 8),
                date_filter_active=True,
            ),
            "Hotel Bell Elite Menu Sales 01 August 26 to 08 August 26.xlsx",
        )
        self.assertEqual(
            report_export_filename("Tips"),
            "Hotel Bell Elite Tips.xlsx",
        )
        self.assertEqual(
            report_export_filename(
                "Sales - Restaurant & Bar",
                date_from=date(2026, 4, 1),
                date_to=date(2026, 4, 1),
                date_filter_active=True,
            ),
            "Hotel Bell Elite Sales - Restaurant & Bar 01 April 26.xlsx",
        )
        self.assertEqual(
            report_export_filename(
                "Meal Plan",
                date_from=date(2026, 8, 24),
                date_to=date(2026, 8, 24),
                date_filter_active=True,
            ),
            "Hotel Bell Elite Meal Plan 24 August 26.xlsx",
        )
        self.assertEqual(
            report_export_month_filename("Salary Payment", 2026, 8),
            "Hotel Bell Elite Salary Payment 01 August 26 to 31 August 26.xlsx",
        )

    def test_sales_report_pages_and_export(self):
        # Seed one restaurant + one bar invoice so POS pages have rows.
        rest_save = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._pos_payload(outlet="restaurant", order_no="ORD-SR-REST"),
        )
        self.assertEqual(rest_save.status_code, 200, rest_save.get_data(as_text=True))
        bar_save = self.client.post(
            "/bar-point-of-sale/api/invoices",
            json=self._pos_payload(outlet="bar", order_no="", table="B1"),
        )
        self.assertEqual(bar_save.status_code, 200, bar_save.get_data(as_text=True))
        bar_json = bar_save.get_json() or {}
        self.assertTrue(bar_json.get("ok"))
        bar_order_no = (bar_json.get("invoice") or {}).get("order_no") or ""

        cases = (
            (
                "/reports/sales/hotel",
                "/reports/sales/hotel/export",
                "Hotel Sales",
                "Invoice No",
                "Hotel Sales",
            ),
            (
                "/reports/sales/restaurant",
                "/reports/sales/restaurant/export",
                "Sales - Restaurant &amp; Bar",
                "Order No",
                "Sales - Restaurant & Bar",
            ),
            (
                "/reports/sales/bar",
                "/reports/sales/bar/export",
                "Bar Sales",
                "Order No",
                "Bar Sales",
            ),
        )
        fy_start, today = db_mod.indian_fiscal_year_bounds()
        for page_url, export_url, title, col_marker, export_title in cases:
            page = self.client.get(f"{page_url}?from_hub=reports")
            self.assertEqual(page.status_code, 200, page_url)
            html = page.get_data(as_text=True)
            self.assertIn(title, html)
            self.assertIn('id="sales-report-page"', html)
            self.assertIn("Export", html)
            self.assertIn('aria-label="Back to Reports"', html)
            self.assertTrue(
                'id="sr-table"' in html or "No invoices match these filters." in html,
                f"expected table or empty state on {page_url}",
            )
            if 'id="sr-table"' in html:
                self.assertIn(col_marker, html)

            export = self.client.get(export_url)
            self.assertEqual(export.status_code, 200, export_url)
            self.assertIn(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                export.content_type or "",
            )
            self.assertTrue(export.data[:2] == b"PK")
            cd = export.headers.get("Content-Disposition") or ""
            expected_name = report_export_filename(
                export_title,
                date_from=fy_start,
                date_to=today,
                date_filter_active=True,
            )
            self.assertIn(expected_name, cd)
            self.assertIn(
                f'download="{expected_name.replace("&", "&amp;")}"',
                html,
            )
            self.assertIn(
                "no-store", (export.headers.get("Cache-Control") or "").lower()
            )
            if col_marker == "Order No":
                from io import BytesIO
                from openpyxl import load_workbook

                wb = load_workbook(BytesIO(export.data))
                self.assertEqual(wb.sheetnames, ["Summary", "Line Items"])
                summary = wb["Summary"]
                details = wb["Line Items"]
                self.assertNotIn("Sheet1", wb.sheetnames)
                self.assertIn(" (", summary["A1"].value or "")
                self.assertTrue((summary["A1"].value or "").endswith(")"))
                self.assertEqual(summary["A2"].value, "Restaurant")
                self.assertEqual(summary["A3"].value, "Bar")
                self.assertEqual(summary["A4"].value, "Total Sales")
                self.assertEqual(summary["B6"].value, "Amount")
                self.assertEqual(summary.max_column, 2)
                self.assertIn("A1:B1", {str(r) for r in summary.merged_cells.ranges})
                summary_col_a = [
                    summary.cell(row, 1).value
                    for row in range(1, (summary.max_row or 1) + 1)
                ]
                self.assertIn("Payment Mode", summary_col_a)
                self.assertIn("Un Settled", summary_col_a)
                self.assertEqual(summary_col_a[-1], "Total")
                header_row = next(
                    (
                        row
                        for row in range(1, (details.max_row or 1) + 1)
                        if details.cell(row, 1).value == "Order No"
                    ),
                    None,
                )
                self.assertEqual(header_row, 2)
                headers = [details.cell(header_row, col).value for col in range(1, 20)]
                self.assertEqual(headers[0], "Order No")
                self.assertEqual(headers[1], "Date")
                self.assertEqual(headers[2], "Customer")
                self.assertNotIn("Saved At", headers)
                self.assertNotIn("Tip", headers)
                self.assertEqual(headers[13], "Total")
                self.assertIsNone(headers[14])
                invoice_col_a = [
                    details.cell(row, 1).value
                    for row in range(1, (details.max_row or 1) + 1)
                ]
                guests = [
                    details.cell(row, 3).value
                    for row in range(1, (details.max_row or 1) + 1)
                ]
                self.assertNotIn("Sales Summary", invoice_col_a)
                self.assertNotIn("Total Sales", invoice_col_a)
                self.assertEqual(summary["A1"].fill.fgColor.rgb, "FF315A78")
                self.assertEqual(details.cell(2, 1).fill.fgColor.rgb, "FF315A78")
                self.assertNotEqual(
                    getattr(details.cell(3, 1).fill.fgColor, "rgb", None), "FF315A78"
                )
                rest_banner = "Hotel Bell Elite — Sales - Restaurant"
                bar_banner = "Hotel Bell Elite — Sales - Bar"
                if page_url == "/reports/sales/restaurant":
                    self.assertTrue(
                        (summary["A1"].value or "").startswith(
                            "Hotel Bell Elite — Sales - Restaurant & Bar ("
                        )
                    )
                    self.assertIn(rest_banner, invoice_col_a)
                    self.assertIn(bar_banner, invoice_col_a)
                    self.assertIn("ORD-SR-REST", invoice_col_a)
                    self.assertLess(
                        invoice_col_a.index(rest_banner), invoice_col_a.index(bar_banner)
                    )
                    self.assertLess(
                        invoice_col_a.index(rest_banner),
                        invoice_col_a.index("ORD-SR-REST"),
                    )
                    self.assertLess(
                        invoice_col_a.index("ORD-SR-REST"),
                        invoice_col_a.index(bar_banner),
                    )
                    if bar_order_no:
                        self.assertIn(bar_order_no, invoice_col_a)
                        self.assertLess(
                            invoice_col_a.index(bar_banner),
                            invoice_col_a.index(bar_order_no),
                        )
                    self.assertGreaterEqual(guests.count("Sales Guest"), 2)
                    self.assertIn('aria-label="Restaurant sales"', html)
                    self.assertIn('aria-label="Bar sales"', html)
                    self.assertIn("POS restaurant", html)
                    self.assertIn("POS bar", html)
                else:
                    self.assertTrue(
                        (summary["A1"].value or "").startswith(
                            "Hotel Bell Elite — Bar Sales ("
                        )
                    )
                    self.assertIn("Sales Guest", guests)
                    self.assertIn(bar_banner, invoice_col_a)
                    self.assertNotIn(rest_banner, invoice_col_a)

        rest_page = self.client.get("/reports/sales/restaurant")
        rest_html = rest_page.get_data(as_text=True)
        self.assertIn('id="sr-table"', rest_html)
        self.assertIn("Order No", rest_html)
        self.assertIn("ORD-SR-REST", rest_html)
        self.assertIn("22 July 26", rest_html)
        self.assertIn("18:00", rest_html)
        self.assertNotIn("Saved At", rest_html)
        self.assertNotIn(">Tip<", rest_html)
        self.assertIn("Sales Guest", rest_html)
        self.assertIn('aria-label="Restaurant sales"', rest_html)
        self.assertIn('aria-label="Bar sales"', rest_html)
        self.assertIn(">Bills<", rest_html)
        self.assertIn('id="sr-outlet-listbox"', rest_html)
        self.assertIn('id="sr-status-listbox"', rest_html)
        self.assertIn('data-value="all" data-label="All"', rest_html)
        self.assertIn('data-value="restaurant" data-label="Restaurant"', rest_html)
        self.assertIn('data-value="bar" data-label="Bar"', rest_html)
        self.assertIn('id="sr-date-range-today"', rest_html)
        self.assertIn(">Today<", rest_html)
        self.assertIn(f'value="{fy_start.isoformat()}"', rest_html)
        self.assertIn(f'value="{today.isoformat()}"', rest_html)

        bar_page = self.client.get("/reports/sales/bar")
        bar_html = bar_page.get_data(as_text=True)
        self.assertIn('id="sr-table"', bar_html)
        self.assertIn("Order No", bar_html)
        self.assertNotIn("Saved At", bar_html)
        self.assertIn("Sales Guest", bar_html)
        self.assertNotIn('id="sr-outlet-listbox"', bar_html)

        hotel_page = self.client.get("/reports/sales/hotel")
        hotel_html = hotel_page.get_data(as_text=True)
        self.assertNotIn('id="sr-outlet-listbox"', hotel_html)

    def test_combined_sales_outlet_filter(self):
        rest_save = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._pos_payload(outlet="restaurant", order_no="ORD-OUT-REST"),
        )
        self.assertEqual(rest_save.status_code, 200, rest_save.get_data(as_text=True))
        bar_save = self.client.post(
            "/bar-point-of-sale/api/invoices",
            json=self._pos_payload(outlet="bar", order_no="ORD-OUT-BAR", table="B1"),
        )
        self.assertEqual(bar_save.status_code, 200, bar_save.get_data(as_text=True))

        all_html = self.client.get("/reports/sales/restaurant").get_data(as_text=True)
        self.assertIn("ORD-OUT-REST", all_html)
        self.assertIn("ORD-OUT-BAR", all_html)
        self.assertIn('id="sr-outlet-value">All<', all_html)
        self.assertIn('aria-label="Restaurant sales"', all_html)
        self.assertIn('aria-label="Bar sales"', all_html)
        self.assertIn('class="sr-group-header"', all_html)
        rest_header = all_html.find('<td colspan="8">Restaurant</td>')
        bar_header = all_html.find('<td colspan="8">Bar</td>')
        self.assertNotEqual(rest_header, -1)
        self.assertNotEqual(bar_header, -1)
        self.assertLess(rest_header, bar_header)
        self.assertLess(rest_header, all_html.find("ORD-OUT-REST"))
        self.assertLess(all_html.find("ORD-OUT-REST"), bar_header)
        self.assertLess(bar_header, all_html.find("ORD-OUT-BAR"))

        rest_html = self.client.get(
            "/reports/sales/restaurant?outlet=restaurant"
        ).get_data(as_text=True)
        self.assertIn("ORD-OUT-REST", rest_html)
        self.assertNotIn("ORD-OUT-BAR", rest_html)
        self.assertIn('id="sr-outlet-value">Restaurant<', rest_html)
        self.assertIn("outlet=restaurant", rest_html)
        self.assertNotIn('class="sr-group-header"', rest_html)
        self.assertIn('aria-label="Restaurant sales"', rest_html)
        self.assertNotIn('aria-label="Bar sales"', rest_html)

        bar_html = self.client.get("/reports/sales/restaurant?outlet=bar").get_data(
            as_text=True
        )
        self.assertIn("ORD-OUT-BAR", bar_html)
        self.assertNotIn("ORD-OUT-REST", bar_html)
        self.assertIn('id="sr-outlet-value">Bar<', bar_html)
        self.assertNotIn('class="sr-group-header"', bar_html)
        self.assertIn('aria-label="Bar sales"', bar_html)
        self.assertNotIn('aria-label="Restaurant sales"', bar_html)

        # Bar sales page stays bar-only even if outlet=restaurant is passed.
        scoped = self.client.get("/reports/sales/bar?outlet=restaurant").get_data(
            as_text=True
        )
        self.assertNotIn('id="sr-outlet-listbox"', scoped)
        self.assertNotIn('class="sr-group-header"', scoped)
        self.assertIn("ORD-OUT-BAR", scoped)
        self.assertNotIn("ORD-OUT-REST", scoped)

        from io import BytesIO
        from openpyxl import load_workbook

        def export_col_a(url):
            resp = self.client.get(url)
            self.assertEqual(resp.status_code, 200, url)
            wb = load_workbook(BytesIO(resp.data))
            self.assertIn("Line Items", wb.sheetnames, url)
            ws = wb["Line Items"]
            return [ws.cell(row, 1).value for row in range(1, (ws.max_row or 1) + 1)]

        def invoice_labels(col_a):
            return col_a

        rest_banner = "Hotel Bell Elite — Sales - Restaurant"
        bar_banner = "Hotel Bell Elite — Sales - Bar"

        def assert_all_outlet_grouped(col_a, url):
            labels = invoice_labels(col_a)
            self.assertNotIn("Sales Summary", col_a, url)
            self.assertIn("ORD-OUT-REST", labels, url)
            self.assertIn("ORD-OUT-BAR", labels, url)
            self.assertIn(rest_banner, labels, url)
            self.assertIn(bar_banner, labels, url)
            self.assertLess(labels.index(rest_banner), labels.index(bar_banner), url)
            self.assertLess(
                labels.index(rest_banner), labels.index("ORD-OUT-REST"), url
            )
            self.assertLess(labels.index("ORD-OUT-REST"), labels.index(bar_banner), url)
            self.assertLess(labels.index(bar_banner), labels.index("ORD-OUT-BAR"), url)

        for all_url in (
            "/reports/sales/restaurant/export",
            "/reports/sales/restaurant/export?outlet=all",
        ):
            assert_all_outlet_grouped(export_col_a(all_url), all_url)

        rest_col = invoice_labels(
            export_col_a("/reports/sales/restaurant/export?outlet=restaurant")
        )
        self.assertIn("ORD-OUT-REST", rest_col)
        self.assertNotIn("ORD-OUT-BAR", rest_col)
        self.assertIn(rest_banner, rest_col)
        self.assertNotIn(bar_banner, rest_col)

        bar_col = invoice_labels(
            export_col_a("/reports/sales/restaurant/export?outlet=bar")
        )
        self.assertIn("ORD-OUT-BAR", bar_col)
        self.assertNotIn("ORD-OUT-REST", bar_col)
        self.assertIn(bar_banner, bar_col)
        self.assertNotIn(rest_banner, bar_col)

        bar_page_col = invoice_labels(export_col_a("/reports/sales/bar/export"))
        self.assertIn("ORD-OUT-BAR", bar_page_col)
        self.assertNotIn("ORD-OUT-REST", bar_page_col)
        self.assertIn(bar_banner, bar_page_col)
        self.assertNotIn(rest_banner, bar_page_col)

    def test_sales_report_excludes_tip_from_total(self):
        payload = self._pos_payload(outlet="restaurant", order_no="ORD-SR-TIP", total=575)
        payload["tipAmount"] = 50
        payload["totals"]["tip"] = 50
        payload["totals"]["total"] = 575
        saved = self.client.post("/point-of-sale/api/invoices", json=payload)
        self.assertEqual(saved.status_code, 200, saved.get_data(as_text=True))
        invoice = (saved.get_json() or {}).get("invoice") or {}
        grand_total = float(invoice.get("grand_total") or 0)
        tip = float(invoice.get("tip") or invoice.get("tip_amount") or 0)
        self.assertGreater(tip, 0)
        expected = round(grand_total - tip, 2)

        page = self.client.get("/reports/sales/restaurant")
        html = page.get_data(as_text=True)
        self.assertIn("ORD-SR-TIP", html)
        self.assertIn(f'data-amount="{expected}"', html)
        self.assertIn(f'data-sort-value="{expected}"', html)
        self.assertNotIn(">Tip<", html)

        from io import BytesIO
        from openpyxl import load_workbook

        export = self.client.get("/reports/sales/restaurant/export")
        self.assertEqual(export.status_code, 200)
        wb = load_workbook(BytesIO(export.data))
        self.assertEqual(wb.sheetnames, ["Summary", "Line Items"])
        summary = wb["Summary"]
        details = wb["Line Items"]
        self.assertEqual(float(summary["B4"].value), expected)
        self.assertIn("Payment Mode", [
            summary.cell(row, 1).value for row in range(1, (summary.max_row or 1) + 1)
        ])
        header_row = next(
            (
                row
                for row in range(1, (details.max_row or 1) + 1)
                if details.cell(row, 1).value == "Order No"
            ),
            None,
        )
        self.assertEqual(header_row, 2)
        headers = [details.cell(header_row, col).value for col in range(1, 16)]
        self.assertNotIn("Tip", headers)
        self.assertEqual(headers[13], "Total")
        self.assertIn(" (", summary["A1"].value or "")
        order_row = None
        for row in range(1, (details.max_row or 1) + 1):
            if details.cell(row, 1).value == "ORD-SR-TIP":
                order_row = row
                break
        self.assertIsNotNone(order_row)
        self.assertEqual(float(details.cell(order_row, 14).value), expected)
        self.assertNotEqual(float(details.cell(order_row, 14).value), grand_total)

    def test_sales_report_payment_mode_summary_sheet(self):
        saved = self.client.post(
            "/point-of-sale/api/invoices",
            json=self._pos_payload(outlet="restaurant", order_no="ORD-SR-PAY", total=500),
        )
        self.assertEqual(saved.status_code, 200, saved.get_data(as_text=True))
        invoice = (saved.get_json() or {}).get("invoice") or {}
        invoice_id = invoice["id"]
        total = float(invoice.get("grand_total") or 0)
        cash_amount = 200
        upi_amount = round(total - cash_amount, 2)
        settle = self.client.post(
            f"/point-of-sale/api/invoices/{invoice_id}/settle",
            json={
                "payment_splits": [
                    {"payment_method": "cash", "amount": cash_amount},
                    {"payment_method": "upi", "amount": upi_amount},
                ],
            },
        )
        self.assertEqual(settle.status_code, 200, settle.get_data(as_text=True))

        conn = db_mod.get_db()
        try:
            rows, total_amount = self.app_mod._sales_report_payment_mode_summary(
                conn, [invoice]
            )
        finally:
            conn.close()
        by_label = {item["label"]: item["amount"] for item in rows}
        self.assertEqual(by_label.get("Cash"), cash_amount)
        self.assertEqual(by_label.get("UPI"), upi_amount)
        self.assertNotIn("Un Settled", by_label)
        self.assertEqual(total_amount, round(cash_amount + upi_amount, 2))

        from io import BytesIO
        from openpyxl import load_workbook

        export = self.client.get("/reports/sales/restaurant/export")
        self.assertEqual(export.status_code, 200)
        summary = load_workbook(BytesIO(export.data))["Summary"]
        modes = {
            summary.cell(row, 1).value: summary.cell(row, 2).value
            for row in range(1, (summary.max_row or 1) + 1)
            if summary.cell(row, 1).value
        }
        self.assertEqual(float(modes["Cash"]), cash_amount)
        self.assertEqual(float(modes["UPI"]), upi_amount)
        self.assertNotIn("Un Settled", modes)
        self.assertEqual(float(modes["Total"]), float(modes["Total Sales"]))

    def _checkin_agency_room(self, room_id="room-101", agency="Travel Desk Co"):
        check_out = datetime.now().date() + timedelta(days=1)
        check_in = check_out - timedelta(days=1)
        res = self.client.put(
            f"/hotel/api/rooms/{room_id}",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": "Asha",
                    "lastName": "Iyer",
                    "mobile": "9000000077",
                    "checkInDate": check_in.isoformat(),
                    "checkOutDate": check_out.isoformat(),
                    "nights": 1,
                    "roomRate": 2000,
                    "advancePaid": 0,
                    "agencyName": agency,
                    "agencyBilling": True,
                },
            },
        )
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        return res.get_json()["room"]

    def _credit_settle_room(self, room_id="room-101"):
        room = self.client.get(f"/hotel/api/rooms/{room_id}").get_json()["room"]
        pay = round(float(room["stay"]["balanceAmount"]), 2)
        res = self.client.put(
            f"/hotel/api/rooms/{room_id}",
            json={
                "action": "generate_invoice",
                "payment_splits": [{"payment_method": "credit", "amount": pay}],
            },
        )
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        stay = res.get_json()["room"]["stay"]
        return stay["invoiceNumber"], pay

    def test_agency_billing_includes_open_agency_invoices_without_credit(self):
        self._checkin_agency_room(room_id="room-102", agency="ATPI India Pvt. Ltd")
        res = self.client.put(
            "/hotel/api/rooms/room-102",
            json={"action": "generate_invoice", "payment_splits": []},
        )
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        stay = res.get_json()["room"]["stay"]
        inv_no = stay["invoiceNumber"]
        billed = round(float(stay["estimatedTotal"]), 2)
        conn = db_mod.get_db()
        try:
            credit = conn.execute(
                "SELECT id FROM hotel_invoice_credits WHERE invoice_number = ?",
                (inv_no,),
            ).fetchone()
        finally:
            conn.close()
        self.assertIsNone(credit)

        page = self.client.get("/reports/sales/agency-billing")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn(inv_no, html)
        self.assertIn("ATPI India Pvt. Ltd", html)
        self.assertIn(f'data-amount="{billed}', html)
        self.assertIn("Outstanding", html)

    def test_agency_billing_report_page_and_collections(self):
        self._checkin_agency_room()
        inv_no, billed = self._credit_settle_room()
        page = self.client.get("/reports/sales/agency-billing?from_hub=reports")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn("Agency Ledger", html)
        self.assertIn('id="sales-report-page"', html)
        self.assertIn('data-report-kind="agency"', html)
        self.assertIn('aria-label="Back to Reports"', html)
        self.assertIn("Invoice No", html)
        self.assertIn("Agency", html)
        self.assertIn("Billed", html)
        self.assertIn("Received", html)
        self.assertIn("Travel Desk Co", html)
        self.assertIn(inv_no, html)
        self.assertIn("Outstanding", html)
        self.assertIn("Collected in period", html)
        self.assertIn(">Export</a>", html)
        self.assertNotIn("Export Excel", html)

        conn = db_mod.get_db()
        try:
            credit = conn.execute(
                "SELECT id FROM hotel_invoice_credits WHERE invoice_number = ?",
                (inv_no,),
            ).fetchone()
        finally:
            conn.close()
        self.assertIsNotNone(credit)
        partial = round(min(billed, 500.0), 2)
        created = self.client.post(
            "/hotel/credit/create",
            json={
                "payment_date": date.today().isoformat(),
                "payment_method": "cash",
                "allocations": [{"expense_id": credit["id"], "amount": partial}],
            },
        )
        self.assertEqual(created.status_code, 200, created.get_data(as_text=True))

        after = self.client.get("/reports/sales/agency-billing").get_data(as_text=True)
        self.assertIn(inv_no, after)
        remaining = round(billed - partial, 2)
        self.assertIn(f'data-amount="{partial}', after)
        self.assertIn(f'data-amount="{remaining}', after)

        outside = self.client.get(
            "/reports/sales/agency-billing?date_from=2020-04-01&date_to=2020-04-02"
        )
        self.assertEqual(outside.status_code, 200)
        outside_html = outside.get_data(as_text=True)
        self.assertNotIn(inv_no, outside_html)

        export = self.client.get("/reports/sales/agency-billing/export")
        self.assertEqual(export.status_code, 200)
        self.assertTrue(export.data[:2] == b"PK")
        self.assertIn("no-store", (export.headers.get("Cache-Control") or "").lower())
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(export.data))
        self.assertEqual(wb.sheetnames, ["Summary", "Grouped", "All Items"])
        summary = wb["Summary"]
        grouped = wb["Grouped"]
        all_items = wb["All Items"]
        self.assertEqual(summary.title, "Summary")
        self.assertIn("Agency Ledger", str(summary["A1"].value or ""))
        self.assertEqual(summary["A2"].value, "Billed")
        self.assertEqual(summary["A3"].value, "Received")
        self.assertEqual(summary["A4"].value, "Outstanding")
        self.assertEqual(summary["A5"].value, "Collected in period")
        self.assertEqual(summary["A6"].value, "Total")
        self.assertEqual(summary["A1"].fill.fgColor.rgb, "FF315A78")
        self.assertIn("A1:B1", {str(r) for r in summary.merged_cells.ranges})

        col_a = [
            grouped.cell(row, 1).value
            for row in range(1, (grouped.max_row or 1) + 1)
        ]
        agency_banner = "Hotel Bell Elite — Agency Ledger - Travel Desk Co"
        self.assertIn(agency_banner, col_a)
        self.assertIn(inv_no, col_a)
        self.assertEqual(grouped.cell(2, 1).value, "Invoice No")
        self.assertEqual(
            [grouped.cell(2, col).value for col in range(1, 9)],
            [
                "Invoice No",
                "Date",
                "Guest",
                "Room",
                "Billed",
                "Received",
                "Balance",
                "Status",
            ],
        )
        self.assertEqual(grouped.cell(1, 1).fill.fgColor.rgb, "FF315A78")
        self.assertEqual(grouped.cell(2, 1).fill.fgColor.rgb, "FF315A78")

        self.assertEqual(all_items.cell(1, 1).value, "Hotel Bell Elite — Agency Ledger")
        self.assertEqual(all_items.cell(1, 1).fill.fgColor.rgb, "FF315A78")
        self.assertEqual(
            [all_items.cell(2, col).value for col in range(1, 10)],
            [
                "Agency",
                "Invoice No",
                "Date",
                "Guest",
                "Room",
                "Billed",
                "Received",
                "Balance",
                "Status",
            ],
        )
        invoice_col = [
            all_items.cell(row, 2).value
            for row in range(3, (all_items.max_row or 2) + 1)
        ]
        self.assertIn(inv_no, invoice_col)
        fy_start, today = db_mod.indian_fiscal_year_bounds()
        cd = export.headers.get("Content-Disposition") or ""
        self.assertIn(
            report_export_filename(
                "Agency Ledger",
                date_from=fy_start,
                date_to=today,
                date_filter_active=True,
            ),
            cd,
        )

    def test_sales_report_endpoints_map_to_reports_module(self):
        for endpoint in (
            "sales_report_hotel",
            "sales_report_hotel_export",
            "sales_report_agency_billing",
            "sales_report_agency_billing_export",
            "sales_report_manager_insight",
            "sales_report_manager_insight_export",
            "sales_report_meal_plan",
            "sales_report_meal_plan_export",
            "sales_report_kot",
            "sales_report_kot_export",
            "sales_report_restaurant",
            "sales_report_restaurant_export",
            "sales_report_bar",
            "sales_report_bar_export",
            "sales_report_menu",
            "sales_report_menu_export",
            "sales_report_customer_insights",
            "sales_report_customer_insights_export",
        ):
            self.assertEqual(get_endpoint_dashboard_module(endpoint), "reports")

    def test_sales_report_access_gate(self):
        viewer = {
            "id": self.admin_id,
            "username": "noreports",
            "full_name": "No Reports",
            "is_admin": False,
            "is_active": True,
            "dashboard_access": {"hotel_rooms", "point_of_sale"},
            "stores_access": set(),
            "sales_analytics_access": set(),
            "user_access": set(),
            "payroll_access": set(),
            "accounts_access": set(),
        }
        with mock.patch.object(self.app_mod, "get_current_user", return_value=viewer):
            for path in (
                "/reports",
                "/reports/sales/hotel",
                "/reports/sales/agency-billing",
                "/reports/sales/agency-billing/export",
                "/reports/sales/manager-insight",
                "/reports/sales/manager-insight/export",
                "/reports/sales/meal-plan",
                "/reports/sales/meal-plan/export",
                "/reports/sales/kot",
                "/reports/sales/kot/export",
                "/reports/sales/restaurant",
                "/reports/sales/bar",
                "/reports/sales/menu",
                "/reports/sales/customer-insights",
                "/reports/sales/hotel/export",
                "/reports/sales/menu/export",
            ):
                denied = self.client.get(path)
                self.assertIn(denied.status_code, (302, 403), path)
                if denied.status_code == 302:
                    self.assertNotIn(b'id="sales-report-page"', denied.data)
                    self.assertNotIn(b'id="menu-sales-report-page"', denied.data)
                    self.assertNotIn(b'id="meal-plan-report-page"', denied.data)

    def _checkin_meal_plan_room(
        self,
        room_id="room-101",
        *,
        rate_plan="MAP",
        adults=2,
        children=1,
        nights=2,
        nightly_plans=None,
        first_name="Asha",
        last_name="Nair",
    ):
        today = date.today()
        check_out = today + timedelta(days=max(1, int(nights)))
        nightly = nightly_plans
        if nightly is None:
            nightly = [
                {
                    "date": (today + timedelta(days=i)).isoformat(),
                    "roomRate": 3500,
                    "ratePlan": rate_plan,
                }
                for i in range(max(1, int(nights)))
            ]
        res = self.client.put(
            f"/hotel/api/rooms/{room_id}",
            json={
                "action": "checkin",
                "stay": {
                    "firstName": first_name,
                    "lastName": last_name,
                    "mobile": "9000001101",
                    "checkInDate": today.isoformat(),
                    "checkOutDate": check_out.isoformat(),
                    "nights": max(1, int(nights)),
                    "roomRate": 3500,
                    "adults": adults,
                    "children": children,
                    "ratePlan": rate_plan,
                    "nightlyRates": nightly,
                    "advancePaid": 0,
                },
            },
        )
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        return res.get_json()["room"], today, check_out

    def test_meal_plan_report_covers_by_plan_and_nightly_override(self):
        from meal_plan_report import (
            build_meal_plan_report,
            meal_covers_for_plan,
            plan_for_date,
        )

        self.assertEqual(
            meal_covers_for_plan("MAP", 3),
            {"breakfast": 3, "lunch": 0, "dinner": 3},
        )
        self.assertEqual(
            meal_covers_for_plan("CP", 2),
            {"breakfast": 2, "lunch": 0, "dinner": 0},
        )
        self.assertEqual(
            meal_covers_for_plan("EP", 2),
            {"breakfast": 0, "lunch": 0, "dinner": 0},
        )
        self.assertEqual(
            meal_covers_for_plan("AP", 4),
            {"breakfast": 4, "lunch": 4, "dinner": 4},
        )

        today = date.today()
        night2 = today + timedelta(days=1)
        room, _check_in, _check_out = self._checkin_meal_plan_room(
            room_id="room-103",
            rate_plan="MAP",
            adults=2,
            children=1,
            nights=2,
            nightly_plans=[
                {"date": today.isoformat(), "roomRate": 3500, "ratePlan": "MAP"},
                {"date": night2.isoformat(), "roomRate": 3500, "ratePlan": "CP"},
            ],
        )
        stay = room["stay"]
        self.assertEqual(plan_for_date(stay, today), "MAP")
        self.assertEqual(plan_for_date(stay, night2), "CP")

        page = self.client.get("/reports/sales/meal-plan")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn("Meal Plan", html)
        self.assertIn("id=\"meal-plan-report-page\"", html)
        self.assertIn("MAP — Breakfast + Dinner", html)
        self.assertIn("Asha Nair", html)

        conn = db_mod.get_db()
        try:
            day1 = build_meal_plan_report(conn, report_date=today)
            day2 = build_meal_plan_report(conn, report_date=night2)
        finally:
            conn.close()

        match1 = next(r for r in day1["rows"] if "103" in (r.get("room_number") or ""))
        self.assertEqual(match1["plan"], "MAP")
        self.assertEqual(match1["pax"], 3)
        self.assertEqual(match1["breakfast"], 3)
        self.assertEqual(match1["lunch"], 0)
        self.assertEqual(match1["dinner"], 3)
        self.assertEqual(day1["kpis"]["breakfast"], match1["breakfast"])
        self.assertEqual(day1["kpis"]["dinner"], match1["dinner"])

        match2 = next(r for r in day2["rows"] if "103" in (r.get("room_number") or ""))
        self.assertEqual(match2["plan"], "CP")
        self.assertEqual(match2["breakfast"], 3)
        self.assertEqual(match2["lunch"], 0)
        self.assertEqual(match2["dinner"], 0)

        self._checkin_meal_plan_room(
            room_id="room-104",
            rate_plan="EP",
            adults=1,
            children=0,
            nights=1,
            first_name="Ravi",
            last_name="EP",
        )
        conn = db_mod.get_db()
        try:
            with_ep = build_meal_plan_report(conn, report_date=today)
        finally:
            conn.close()
        ep_row = next(r for r in with_ep["rows"] if "104" in (r.get("room_number") or ""))
        self.assertEqual(ep_row["plan"], "EP")
        self.assertEqual(ep_row["breakfast"], 0)
        self.assertEqual(ep_row["dinner"], 0)

        export = self.client.get(
            f"/reports/sales/meal-plan/export?date={today.isoformat()}"
        )
        self.assertEqual(export.status_code, 200)
        self.assertIn(
            "spreadsheetml.sheet",
            export.headers.get("Content-Type", ""),
        )
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(export.data))
        self.assertEqual(wb.sheetnames, ["Summary", "Line Items"])
        summary = wb["Summary"]
        details = wb["Line Items"]
        self.assertTrue(
            (summary["A1"].value or "").startswith("Hotel Bell Elite — Meal Plan")
        )
        self.assertEqual(summary["A2"].value, "Occupied rooms")
        self.assertEqual(summary["A3"].value, "Pax")
        self.assertEqual(summary["A4"].value, "Breakfast")
        self.assertEqual(summary["A5"].value, "Lunch")
        self.assertEqual(summary["A6"].value, "Dinner")
        self.assertEqual(summary["A8"].value, "Meal Plan")
        self.assertEqual(summary["B8"].value, "Rooms")
        self.assertEqual(summary["A13"].value, "Total")
        self.assertIn("A1:B1", {str(r) for r in summary.merged_cells.ranges})
        self.assertEqual(summary["A1"].fill.fgColor.rgb, "FF315A78")
        self.assertEqual(summary["A8"].fill.fgColor.rgb, "FF315A78")

        col_a = [
            details.cell(row, 1).value
            for row in range(1, (details.max_row or 1) + 1)
        ]
        self.assertIn("Hotel Bell Elite — Meal Plan - MAP", col_a)
        self.assertIn("Hotel Bell Elite — Meal Plan - EP", col_a)
        map_banner = col_a.index("Hotel Bell Elite — Meal Plan - MAP")
        ep_banner = col_a.index("Hotel Bell Elite — Meal Plan - EP")
        self.assertLess(ep_banner, map_banner)
        headers = [
            details.cell(ep_banner + 2, col).value for col in range(1, 10)
        ]
        self.assertEqual(
            headers,
            [
                "Room",
                "Guest",
                "Meal plan",
                "Adults",
                "Children",
                "Pax",
                "Breakfast",
                "Lunch",
                "Dinner",
            ],
        )
        self.assertEqual(details.cell(ep_banner + 1, 1).fill.fgColor.rgb, "FF315A78")
        self.assertEqual(details.cell(ep_banner + 2, 1).fill.fgColor.rgb, "FF315A78")

    def test_meal_plan_from_hub_preserves_back_link(self):
        page = self.client.get("/reports/sales/meal-plan?from_hub=reports")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn("Back to Reports", html)
        self.assertIn('name="from_hub" value="reports"', html)

    def test_meal_plan_dedupes_same_room_across_invoices_and_layout(self):
        from meal_plan_report import build_meal_plan_report

        today = date.today()
        room, _ci, _co = self._checkin_meal_plan_room(
            room_id="room-105",
            rate_plan="AP",
            adults=1,
            children=0,
            nights=1,
            first_name="Antony",
            last_name="Kibson",
        )
        stay = room["stay"]
        check_in = stay["checkInDate"]
        check_out = stay["checkOutDate"]
        guest = "Mr Antony Kibson"
        base_payload = {
            "number": "105",
            "stay": {
                "guestName": guest,
                "firstName": "Antony",
                "lastName": "Kibson",
                "adults": 1,
                "children": 0,
                "ratePlan": "AP",
                "checkInDate": check_in,
                "checkOutDate": check_out,
            },
        }
        ep_payload = {
            "number": "105",
            "stay": {
                "guestName": guest,
                "firstName": "Antony",
                "lastName": "Kibson",
                "adults": 1,
                "children": 0,
                "ratePlan": "EP",
                "checkInDate": check_in,
                "checkOutDate": check_out,
            },
        }
        conn = db_mod.get_db()
        try:
            db_mod.ensure_hotel_room_invoices_schema(conn)
            for inv_no, payload in (
                ("HBE-DUP-1", base_payload),
                ("HBE-DUP-2", base_payload),
                ("HBE-DUP-3", base_payload),
                ("HBE-DUP-EP", ep_payload),
            ):
                conn.execute(
                    """
                    INSERT INTO hotel_room_invoices (
                        invoice_number, room_id, room_number, guest_name,
                        check_in_date, check_out_date, payload_json, status
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, 'open')
                    """,
                    (
                        inv_no,
                        "room-105",
                        "105",
                        guest,
                        check_in,
                        check_out,
                        __import__("json").dumps(payload),
                    ),
                )
            conn.commit()
            payload = build_meal_plan_report(conn, report_date=today)
        finally:
            conn.close()

        room_rows = [
            r for r in payload["rows"] if "105" in (r.get("room_number") or "")
        ]
        self.assertEqual(len(room_rows), 1, room_rows)
        self.assertEqual(room_rows[0]["plan"], "AP")
        self.assertEqual(room_rows[0]["pax"], 1)
        self.assertEqual(payload["kpis"]["breakfast"], 1)
        self.assertEqual(payload["kpis"]["lunch"], 1)
        self.assertEqual(payload["kpis"]["dinner"], 1)


if __name__ == "__main__":
    unittest.main()
