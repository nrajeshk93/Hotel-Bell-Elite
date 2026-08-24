"""Employee Master list markup and ID sort order."""

import os
import re
import tempfile
import unittest
from unittest import mock

import db as db_mod


class EmployeeMasterSortTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db_path = self.tmp.name
        self._orig_path = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = self.db_path
        db_mod.init_db()

        import app as app_mod

        self.app_mod = app_mod
        self.app = app_mod.app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

        conn = db_mod.get_db()
        try:
            admin = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            self.admin_id = admin["id"]
            conn.executemany(
                """INSERT INTO employees (emp_code, name, location, gross_salary, address, mobile, status)
                   VALUES (?, ?, ?, ?, ?, ?, 'active')""",
                [
                    ("HBE10", "Zara", "HK", 30000, "10 Hill Rd", "9000000010"),
                    ("HBE2", "Asha", "FO", 20000, "2 Lake Rd", "9000000002"),
                    ("HBE1", "Mira", "BAR", 10000, "1 Park Rd", "9000000001"),
                ],
            )
            conn.commit()
        finally:
            conn.close()

        self.user = {
            "id": self.admin_id,
            "username": "admin",
            "full_name": "Administrator",
            "is_admin": True,
            "is_active": True,
            "dashboard_access": set(),
            "payroll_access": set(),
        }
        self._get_user_patch = mock.patch.object(
            app_mod, "get_current_user", return_value=self.user
        )
        self._get_user_patch.start()

    def tearDown(self):
        self._get_user_patch.stop()
        db_mod.DATABASE_PATH = self._orig_path
        try:
            os.unlink(self.db_path)
        except OSError:
            pass

    def test_employee_master_table_is_sortable(self):
        page = self.client.get("/employee_master")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn('id="employee-master-page"', html)
        self.assertIn('id="emp-main-table"', html)
        self.assertIn('class="pl-sortable"', html)
        self.assertIn('data-sort="id"', html)
        self.assertIn('data-sort="name"', html)
        self.assertIn('data-sort="department"', html)
        self.assertIn('data-sort="salary"', html)
        self.assertIn('data-sort-type="number"', html)
        self.assertIn("data-sort-row", html)
        self.assertIn("data-sort-id=", html)
        self.assertIn("employee_master.js", html)
        self.assertIn("initEmpMasterTableSort", html)

        codes = re.findall(r'data-sort-id="(HBE\d+)"', html)
        self.assertEqual(codes, ["HBE1", "HBE2", "HBE10"])

        js_path = os.path.join(os.path.dirname(__file__), "..", "static", "employee_master.js")
        with open(js_path, encoding="utf-8") as handle:
            js = handle.read()
        self.assertIn("tbody.appendChild(row)", js)
        self.assertIn("localeCompare", js)
        self.assertIn("numeric: true", js)
        self.assertIn("querySelectorAll('tbody tr[data-sort-row]')", js)

    def test_export_employee_master_uses_ledger_chrome_single_sheet(self):
        from io import BytesIO

        from openpyxl import load_workbook

        export = self.client.get("/export/employee_master")
        self.assertEqual(export.status_code, 200)
        self.assertIn(
            "spreadsheetml.sheet",
            export.headers.get("Content-Type", ""),
        )
        cd = export.headers.get("Content-Disposition") or ""
        self.assertIn("Hotel Bell Elite Employee Master", cd)

        wb = load_workbook(BytesIO(export.data))
        self.assertEqual(wb.sheetnames, ["Employees"])
        self.assertNotIn("Summary", wb.sheetnames)
        ws = wb["Employees"]
        self.assertEqual(ws["A1"].value, "Hotel Bell Elite — Employee Master")
        self.assertEqual(ws["A1"].fill.fgColor.rgb, "FF315A78")
        self.assertEqual(ws["A1"].font.name, "Calibri")
        self.assertEqual(ws["A2"].value, "Emp ID")
        self.assertEqual(ws["A2"].fill.fgColor.rgb, "FF315A78")
        self.assertEqual(ws["B2"].value, "Name")
        self.assertEqual(ws["L2"].value, "Gross Salary")
        # Three seeded employees start on row 3
        names = {
            ws.cell(row=r, column=2).value
            for r in range(3, (ws.max_row or 2) + 1)
        }
        self.assertTrue({"Mira", "Asha", "Zara"}.issubset(names))
        self.assertEqual(ws["A3"].font.name, "Calibri")
        self.assertEqual(ws["L3"].alignment.horizontal, "right")


if __name__ == "__main__":
    unittest.main()
