"""Ledger hub card name and three-sheet Excel export."""

import os
import tempfile
import unittest
from io import BytesIO
from unittest import mock

import db as db_mod
from reports import (
    PURCHASE_EXPENSE_LEDGER_NAME,
    REPORT_CATEGORY_LABELS,
    REPORT_DEFINITIONS,
    SALARY_PAYMENT_NAME,
    format_report_date,
    format_report_datetime,
    format_report_time,
    report_export_filename,
)


class LedgerHubAndExportTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db_path = self.tmp.name
        self._orig_path = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = self.db_path
        db_mod.init_db()

        import app as app_mod

        self.app_mod = app_mod
        self.app = app_mod.app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

        conn = db_mod.get_db()
        try:
            admin = conn.execute(
                "SELECT id FROM users WHERE username = 'admin'"
            ).fetchone()
            self.admin_id = admin["id"]
        finally:
            conn.close()

        self.user = {
            "id": self.admin_id,
            "username": "admin",
            "full_name": "Administrator",
            "is_admin": True,
            "is_active": True,
            "dashboard_access": set(),
            "stores_access": set(),
        }
        self._get_user_patch = mock.patch.object(
            app_mod, "get_current_user", return_value=self.user
        )
        self._get_user_patch.start()

    def tearDown(self):
        self._get_user_patch.stop()
        db_mod.DATABASE_PATH = self._orig_path
        try:
            os.unlink(self.db_path)
        except OSError:
            pass

    def test_hub_card_is_purchase_expense_ledger_and_names_drop_report(self):
        ids = [r["id"] for r in REPORT_DEFINITIONS]
        self.assertNotIn("invoice_ledger", ids)
        self.assertNotIn("purchase_verification", ids)
        self.assertNotIn("payroll_hub", ids)
        self.assertNotIn("supplier", ids)
        self.assertNotIn("customer", ids)
        self.assertNotIn("masters", REPORT_CATEGORY_LABELS)
        ledger = next(r for r in REPORT_DEFINITIONS if r["id"] == "expense_ledger")
        self.assertEqual(ledger["name"], PURCHASE_EXPENSE_LEDGER_NAME)
        bank = next(r for r in REPORT_DEFINITIONS if r["id"] == "bank")
        self.assertEqual(bank["name"], SALARY_PAYMENT_NAME)
        self.assertEqual(bank["view_route"], "bank_report")
        for item in REPORT_DEFINITIONS:
            name = item["name"]
            self.assertFalse(name.endswith(" Report"), name)
            self.assertFalse(name.endswith(" Reports"), name)

        hub = self.client.get("/reports")
        self.assertEqual(hub.status_code, 200)
        html = hub.get_data(as_text=True)
        self.assertIn("Purchase &amp; Expense Ledger", html)
        self.assertIn('data-report-id="expense_ledger"', html)
        self.assertIn('data-report-name="Purchase &amp; Expense Ledger"', html)
        self.assertIn('data-report-id="bank"', html)
        self.assertIn('data-report-name="Salary Payment"', html)
        self.assertIn("Open Salary Payment", html)
        self.assertNotIn('data-report-id="invoice_ledger"', html)
        self.assertNotIn('data-report-id="purchase_verification"', html)
        self.assertNotIn('data-report-id="payroll_hub"', html)
        self.assertNotIn('data-report-id="supplier"', html)
        self.assertNotIn('data-report-id="customer"', html)
        self.assertNotIn('data-rd-category="masters"', html)
        self.assertNotIn('data-rd-section="masters"', html)
        self.assertNotIn("Open Invoice Ledger", html)
        self.assertNotIn("Open Approvals", html)
        self.assertNotIn('aria-label="Open Payroll"', html)
        self.assertNotIn('aria-label="Open Supplier"', html)
        self.assertNotIn('aria-label="Open Customer"', html)

    def test_report_datetime_matches_invoice_ledger(self):
        self.assertEqual(format_report_datetime("2026-07-31 15:49:00"), "31 July 26 15:49")
        self.assertEqual(format_report_datetime("2026-07-31"), "31 July 26")
        self.assertEqual(format_report_date("2026-07-01"), "1 July 26")
        self.assertEqual(format_report_time("2026-07-31 15:49:00"), "15:49")
        self.assertEqual(format_report_time("2026-07-31"), "")
        self.assertEqual(format_report_datetime(None), "—")
        self.assertEqual(format_report_datetime(""), "—")

    def test_page_export_button_uses_hotel_bell_elite_filename(self):
        page = self.client.get("/accounts/purchase-ledger")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn(">Export</a>", html)
        fy_start, today = db_mod.indian_fiscal_year_bounds()
        expected = report_export_filename(
            PURCHASE_EXPENSE_LEDGER_NAME,
            date_from=fy_start,
            date_to=today,
            date_filter_active=True,
        )
        self.assertIn(f'download="{expected.replace("&", "&amp;")}"', html)
        self.assertTrue(expected.startswith("Hotel Bell Elite Purchase & Expense Ledger"))

    def test_export_summary_and_grouped_line_items(self):
        conn = db_mod.get_db()
        try:
            cur = conn.execute(
                "INSERT INTO suppliers (name, gst) VALUES (?, ?)",
                ("Acme Foods", "29AAAAA0000A1Z5"),
            )
            supplier_id = cur.lastrowid
            # Later purchase first in insert order — All Items must still sort by date asc.
            conn.execute(
                """INSERT INTO sales_update_expenses
                   (company, location, sales_date, description, amount, payment_type,
                    supplier_id, category, expense_code, entry_kind, invoice_number)
                   VALUES ('HBE', 'Hotel', '2026-07-02', 'Veg stock', 100, 'cash',
                           ?, 'grocery', 'HBE-PU-9', 'purchase', 'INV-P1')""",
                (supplier_id,),
            )
            conn.execute(
                """INSERT INTO sales_update_expenses
                   (company, location, sales_date, description, amount, payment_type,
                    supplier_id, category, expense_code, entry_kind, invoice_number)
                   VALUES ('HBE', 'Hotel', '2026-07-01', 'Electricity', 40, 'credit',
                           ?, 'utilities', 'HBE-EX-9', 'expense', 'INV-E1')""",
                (supplier_id,),
            )
            conn.execute(
                """INSERT INTO sales_update_expenses
                   (company, location, sales_date, description, amount, payment_type,
                    supplier_id, category, expense_code, entry_kind, invoice_number)
                   VALUES ('HBE', 'Hotel', '2026-07-01', 'Gas refill', 25, 'cash',
                           ?, 'utilities', 'HBE-EX-8', 'expense', 'INV-E2')""",
                (supplier_id,),
            )
            conn.commit()
        finally:
            conn.close()

        export = self.client.get(
            "/accounts/purchase-ledger/report?date_from=2026-07-01&date_to=2026-07-02"
        )
        self.assertEqual(export.status_code, 200)
        cd = export.headers.get("Content-Disposition") or ""
        self.assertIn("Hotel Bell Elite Purchase & Expense Ledger", cd)
        self.assertIn("01 July 26 to 02 July 26.xlsx", cd)
        self.assertIn("no-store", (export.headers.get("Cache-Control") or "").lower())

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(export.data))
        self.assertEqual(wb.sheetnames, ["Summary", "Grouped", "All Items"])
        summary = wb["Summary"]
        details = wb["Grouped"]
        all_items = wb["All Items"]
        self.assertEqual(summary["A2"].value, "Purchase")
        self.assertEqual(summary["A3"].value, "Expense")
        self.assertEqual(summary["A4"].value, "Total")
        self.assertEqual(float(summary["B2"].value), 100)
        self.assertEqual(float(summary["B3"].value), 65)
        self.assertEqual(float(summary["B4"].value), 165)
        self.assertEqual(summary["B6"].value, "Amount")
        self.assertEqual(summary.max_column, 2)
        self.assertIn("A1:B1", {str(r) for r in summary.merged_cells.ranges})
        self.assertTrue(
            (summary["A1"].value or "").startswith(
                "Hotel Bell Elite — Purchase & Expense Ledger ("
            )
        )

        col_a = [
            details.cell(row, 1).value
            for row in range(1, (details.max_row or 1) + 1)
        ]
        purchase_banner = "Hotel Bell Elite — Purchase & Expense Ledger - Purchase"
        expense_banner = "Hotel Bell Elite — Purchase & Expense Ledger - Expense"
        self.assertIn(purchase_banner, col_a)
        self.assertIn(expense_banner, col_a)
        self.assertLess(col_a.index(purchase_banner), col_a.index(expense_banner))
        self.assertIn("HBE-PU-9", col_a)
        self.assertIn("HBE-EX-9", col_a)
        self.assertIn("HBE-EX-8", col_a)
        self.assertLess(col_a.index(purchase_banner), col_a.index("HBE-PU-9"))
        self.assertLess(col_a.index("HBE-PU-9"), col_a.index(expense_banner))
        self.assertLess(col_a.index(expense_banner), col_a.index("HBE-EX-9"))
        self.assertEqual(details.cell(2, 1).value, "ID")
        self.assertNotIn("Type", [
            details.cell(2, col).value for col in range(1, 14)
        ])

        line_headers = [
            details.cell(2, col).value for col in range(1, 13)
        ]
        self.assertEqual(
            line_headers,
            [
                "ID",
                "Date",
                "Description",
                "Category",
                "Invoice",
                "Supplier",
                "GST",
                "Payment",
                "Status",
                "Amount",
                "Paid",
                "Balance",
            ],
        )

        all_banner = "Hotel Bell Elite — Purchase & Expense Ledger"
        self.assertEqual(all_items.cell(1, 1).value, all_banner)
        self.assertEqual(all_items.cell(1, 1).fill.fgColor.rgb, "FF315A78")
        self.assertNotIn(purchase_banner, [
            all_items.cell(row, 1).value
            for row in range(1, (all_items.max_row or 1) + 1)
        ])
        self.assertNotIn(expense_banner, [
            all_items.cell(row, 1).value
            for row in range(1, (all_items.max_row or 1) + 1)
        ])
        all_headers = [
            all_items.cell(2, col).value for col in range(1, 14)
        ]
        self.assertEqual(all_headers, ["Type"] + line_headers)
        self.assertEqual(all_items.cell(2, 1).fill.fgColor.rgb, "FF315A78")
        # Flat chronological: same-day by code (EX-8 before EX-9), then later purchase.
        self.assertEqual(all_items.max_row, 5)
        self.assertEqual(
            [
                (
                    all_items.cell(row, 1).value,
                    all_items.cell(row, 2).value,
                    all_items.cell(row, 3).value,
                )
                for row in (3, 4, 5)
            ],
            [
                ("Expense", "HBE-EX-8", "2026-07-01"),
                ("Expense", "HBE-EX-9", "2026-07-01"),
                ("Purchase", "HBE-PU-9", "2026-07-02"),
            ],
        )
        types_present = {
            all_items.cell(row, 1).value for row in (3, 4, 5)
        }
        self.assertEqual(types_present, {"Purchase", "Expense"})

        purchase_only = self.client.get(
            "/accounts/purchase-ledger/report"
            "?date_from=2026-07-01&date_to=2026-07-02&kind=purchase"
        )
        self.assertEqual(purchase_only.status_code, 200)
        purchase_wb = load_workbook(BytesIO(purchase_only.data))
        self.assertEqual(
            purchase_wb.sheetnames, ["Summary", "Grouped", "All Items"]
        )
        purchase_summary = purchase_wb["Summary"]
        purchase_details = purchase_wb["Grouped"]
        purchase_all = purchase_wb["All Items"]
        self.assertEqual(purchase_summary["A2"].value, "Purchase")
        self.assertEqual(float(purchase_summary["B2"].value), 100)
        self.assertEqual(float(purchase_summary["B3"].value), 0)
        self.assertEqual(float(purchase_summary["B4"].value), 100)
        purchase_col_a = [
            purchase_details.cell(row, 1).value
            for row in range(1, (purchase_details.max_row or 1) + 1)
        ]
        self.assertIn(purchase_banner, purchase_col_a)
        self.assertNotIn(expense_banner, purchase_col_a)
        self.assertIn("HBE-PU-9", purchase_col_a)
        self.assertNotIn("HBE-EX-9", purchase_col_a)
        self.assertNotIn("HBE-EX-8", purchase_col_a)
        self.assertEqual(purchase_all.cell(3, 1).value, "Purchase")
        self.assertEqual(purchase_all.cell(3, 2).value, "HBE-PU-9")
        self.assertEqual(purchase_all.max_row, 3)


class TipsExportTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db_path = self.tmp.name
        self._orig_path = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = self.db_path
        db_mod.init_db()

        import app as app_mod

        self.app_mod = app_mod
        self.app = app_mod.app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

        conn = db_mod.get_db()
        try:
            admin = conn.execute(
                "SELECT id FROM users WHERE username = 'admin'"
            ).fetchone()
            self.admin_id = admin["id"]
            conn.execute(
                "INSERT INTO employees (emp_code, name, company, location, status) VALUES (?, ?, ?, ?, ?)",
                ("E001", "Anita", "HBE", "Hotel", "active"),
            )
            conn.execute(
                "INSERT INTO employees (emp_code, name, company, location, status) VALUES (?, ?, ?, ?, ?)",
                ("E002", "Ravi", "HBE", "Bar", "active"),
            )
            conn.executemany(
                """INSERT INTO sales_update_tips
                   (company, location, sales_date, employee_id, amount, description)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                [
                    ("HBE", "Hotel", "2026-07-10", 1, 100, ""),
                    ("HBE", "Bar", "2026-07-11", 1, 50, "Bar shift"),
                    ("HBE", "Restaurant", "2026-07-12", 1, 25, ""),
                    ("HBE", "Bar", "2026-07-12", 2, 80, ""),
                ],
            )
            conn.commit()
        finally:
            conn.close()

        self.user = {
            "id": self.admin_id,
            "username": "admin",
            "full_name": "Administrator",
            "is_admin": True,
            "is_active": True,
            "dashboard_access": set(),
            "stores_access": set(),
        }
        self._get_user_patch = mock.patch.object(
            app_mod, "get_current_user", return_value=self.user
        )
        self._get_user_patch.start()

    def tearDown(self):
        self._get_user_patch.stop()
        db_mod.DATABASE_PATH = self._orig_path
        try:
            os.unlink(self.db_path)
        except OSError:
            pass

    def test_tips_page_export_button_label_and_filename(self):
        page = self.client.get("/sales_update/tips")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn(">Export</a>", html)
        self.assertIn("/sales_update/tips/report?", html)
        fy_start, today = db_mod.indian_fiscal_year_bounds()
        expected = report_export_filename(
            "Tips",
            date_from=fy_start,
            date_to=today,
            date_filter_active=True,
        )
        self.assertTrue(expected.startswith("Hotel Bell Elite Tips"))
        export = self.client.get(
            "/sales_update/tips/report"
            f"?company=HBE&date_from={fy_start.isoformat()}&date_to={today.isoformat()}"
        )
        self.assertEqual(export.status_code, 200)
        cd = export.headers.get("Content-Disposition") or ""
        self.assertIn(expected.replace("&", "&amp;").split(".xlsx")[0], cd.replace("&", "&amp;"))

    def test_tips_export_summary_grouped_and_all_items(self):
        export = self.client.get(
            "/sales_update/tips/report"
            "?company=HBE&date_from=2026-07-01&date_to=2026-07-31"
        )
        self.assertEqual(export.status_code, 200)
        cd = export.headers.get("Content-Disposition") or ""
        self.assertIn("Hotel Bell Elite Tips", cd)
        self.assertIn("01 July 26 to 31 July 26.xlsx", cd)
        self.assertIn("no-store", (export.headers.get("Cache-Control") or "").lower())

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(export.data))
        self.assertEqual(wb.sheetnames, ["Summary", "Grouped", "All Items"])
        summary = wb["Summary"]
        grouped = wb["Grouped"]
        all_items = wb["All Items"]

        self.assertEqual(summary["A2"].value, "Hotel")
        self.assertEqual(summary["A3"].value, "Bar")
        self.assertEqual(summary["A4"].value, "Restaurant")
        self.assertEqual(summary["A5"].value, "Total")
        self.assertEqual(float(summary["B2"].value), 100)
        self.assertEqual(float(summary["B3"].value), 130)
        self.assertEqual(float(summary["B4"].value), 25)
        self.assertEqual(float(summary["B5"].value), 255)
        self.assertTrue(
            (summary["A1"].value or "").startswith("Hotel Bell Elite — Tips (")
        )
        self.assertIn("A1:B1", {str(r) for r in summary.merged_cells.ranges})
        self.assertEqual(summary["A1"].fill.fgColor.rgb, "FF315A78")

        col_a = [
            grouped.cell(row, 1).value
            for row in range(1, (grouped.max_row or 1) + 1)
        ]
        hotel_banner = "Hotel Bell Elite — Tips - Hotel"
        bar_banner = "Hotel Bell Elite — Tips - Bar"
        restaurant_banner = "Hotel Bell Elite — Tips - Restaurant"
        self.assertIn(hotel_banner, col_a)
        self.assertIn(bar_banner, col_a)
        self.assertIn(restaurant_banner, col_a)
        self.assertLess(col_a.index(hotel_banner), col_a.index(bar_banner))
        self.assertLess(col_a.index(bar_banner), col_a.index(restaurant_banner))
        self.assertIn("Anita", col_a)
        self.assertIn("Ravi", col_a)
        self.assertEqual(grouped.cell(2, 1).value, "Employee")
        self.assertEqual(
            [grouped.cell(2, col).value for col in range(1, 4)],
            ["Employee", "Emp Code", "Amount"],
        )
        self.assertEqual(grouped.cell(1, 1).fill.fgColor.rgb, "FF315A78")
        self.assertEqual(grouped.cell(2, 1).fill.fgColor.rgb, "FF315A78")

        self.assertEqual(all_items.cell(1, 1).value, "Hotel Bell Elite — Tips")
        self.assertEqual(all_items.cell(1, 1).fill.fgColor.rgb, "FF315A78")
        self.assertEqual(all_items.cell(2, 1).fill.fgColor.rgb, "FF315A78")
        all_headers = [
            all_items.cell(2, col).value for col in range(1, 8)
        ]
        self.assertEqual(
            all_headers,
            [
                "Type",
                "Date",
                "Employee",
                "Emp Code",
                "Location",
                "Amount",
                "Description",
            ],
        )
        self.assertEqual(all_items.max_row, 6)
        self.assertEqual(
            [
                (
                    all_items.cell(row, 1).value,
                    all_items.cell(row, 2).value,
                    all_items.cell(row, 3).value,
                    float(all_items.cell(row, 6).value),
                )
                for row in (3, 4, 5, 6)
            ],
            [
                ("Hotel", "2026-07-10", "Anita", 100.0),
                ("Bar", "2026-07-11", "Anita", 50.0),
                ("Restaurant", "2026-07-12", "Anita", 25.0),
                ("Bar", "2026-07-12", "Ravi", 80.0),
            ],
        )

    def test_tips_export_location_filter(self):
        export = self.client.get(
            "/sales_update/tips/report"
            "?company=HBE&location=Bar&date_from=2026-07-01&date_to=2026-07-31"
        )
        self.assertEqual(export.status_code, 200)

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(export.data))
        summary = wb["Summary"]
        grouped = wb["Grouped"]
        all_items = wb["All Items"]
        self.assertEqual(float(summary["B2"].value), 0)
        self.assertEqual(float(summary["B3"].value), 130)
        self.assertEqual(float(summary["B5"].value), 130)
        col_a = [
            grouped.cell(row, 1).value
            for row in range(1, (grouped.max_row or 1) + 1)
        ]
        self.assertIn("Hotel Bell Elite — Tips - Bar", col_a)
        self.assertNotIn("Hotel Bell Elite — Tips - Hotel", col_a)
        self.assertNotIn("Hotel Bell Elite — Tips - Restaurant", col_a)
        self.assertEqual(all_items.max_row, 4)
        self.assertTrue(
            all(all_items.cell(row, 1).value == "Bar" for row in (3, 4))
        )
