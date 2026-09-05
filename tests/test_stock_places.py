"""Warehouse vs Counter stock ledgers: inward, transfer, Store page, audit, export."""

import io
import json
import os
import tempfile
import unittest
from datetime import date
from unittest import mock

import db as db_mod


class StockPlaceTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.tmp.close()
        self.db_path = self.tmp.name
        self._orig_path = db_mod.DATABASE_PATH
        db_mod.DATABASE_PATH = self.db_path
        db_mod.init_db()

        import app as app_mod
        import stores as stores_mod

        self.app_mod = app_mod
        self.stores_mod = stores_mod
        self.app = app_mod.app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

        conn = db_mod.get_db()
        try:
            admin = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
            self.admin_id = admin["id"]
        finally:
            conn.close()

        self.user = {
            "id": self.admin_id,
            "username": "admin",
            "full_name": "Administrator",
            "is_admin": True,
            "is_active": True,
            "dashboard_access": set(),
            "stores_access": set(),
        }
        self._get_user_patch = mock.patch.object(app_mod, "get_current_user", return_value=self.user)
        self._stores_user_patch = mock.patch.object(stores_mod, "_get_user", return_value=self.user)
        self._get_user_patch.start()
        self._stores_user_patch.start()

    def tearDown(self):
        self._get_user_patch.stop()
        self._stores_user_patch.stop()
        db_mod.DATABASE_PATH = self._orig_path
        try:
            os.unlink(self.db_path)
        except OSError:
            pass

    def _qty(self, outlet, item_name, unit, place):
        conn = db_mod.get_db()
        try:
            row = conn.execute(
                """
                SELECT qty_on_hand FROM store_stock_items
                WHERE outlet = ? AND place = ?
                  AND lower(item_name) = lower(?) AND lower(unit) = lower(?)
                """,
                (outlet, place, item_name, unit),
            ).fetchone()
            return float(row["qty_on_hand"]) if row else 0.0
        finally:
            conn.close()

    def _insert_stock(self, *, outlet, place, item_name, unit, qty):
        conn = db_mod.get_db()
        try:
            db_mod.ensure_stores_schema(conn)
            conn.execute(
                """
                INSERT INTO store_stock_items
                    (outlet, place, item_name, unit, qty_on_hand, updated_at)
                VALUES (?, ?, ?, ?, ?, datetime('now','localtime'))
                """,
                (outlet, place, item_name, unit, qty),
            )
            conn.commit()
        finally:
            conn.close()

    def _supplier_id(self, name="Place Test Supplier"):
        conn = db_mod.get_db()
        try:
            conn.execute("INSERT INTO suppliers (name) VALUES (?)", (name,))
            sid = conn.execute(
                "SELECT id FROM suppliers WHERE name = ? ORDER BY id DESC LIMIT 1",
                (name,),
            ).fetchone()["id"]
            conn.commit()
            return int(sid)
        finally:
            conn.close()

    def test_demo_seed_skipped_during_pytest(self):
        conn = db_mod.get_db()
        try:
            count = conn.execute("SELECT COUNT(*) AS c FROM store_stock_items").fetchone()["c"]
            self.assertEqual(int(count), 0)
        finally:
            conn.close()

    def test_inward_credits_warehouse_only(self):
        supplier_id = self._supplier_id()
        confirm = self.client.post(
            "/stores/purchase-requests/confirm-direct-with-expense",
            json={
                "outlet": "restaurant",
                "notes": "Warehouse receive",
                "lines": [
                    {
                        "item_name": "Onion",
                        "qty": 4,
                        "unit": "kg",
                        "unit_price": 20,
                        "tax_percent": 0,
                    }
                ],
                "date": date.today().isoformat(),
                "description": "Stock inward without indent approval",
                "amount": 80,
                "payment_type": "credit",
                "category": "grocery",
                "supplier_id": supplier_id,
            },
        )
        self.assertEqual(confirm.status_code, 200, confirm.get_data(as_text=True))
        self.assertTrue(confirm.get_json().get("ok"))
        self.assertIn("warehouse", (confirm.get_json().get("message") or "").lower())
        self.assertAlmostEqual(self._qty("restaurant", "Onion", "kg", "warehouse"), 4.0)
        self.assertAlmostEqual(self._qty("restaurant", "Onion", "kg", "counter"), 0.0)
        conn = db_mod.get_db()
        try:
            move = conn.execute(
                """
                SELECT place, movement_type FROM store_stock_movements
                WHERE movement_type = 'receive' AND lower(item_name) = lower('Onion')
                ORDER BY id DESC LIMIT 1
                """
            ).fetchone()
            self.assertEqual(move["place"], "warehouse")
        finally:
            conn.close()

    def test_transfer_warehouse_to_counter(self):
        self._insert_stock(
            outlet="restaurant",
            place="warehouse",
            item_name="Tomato",
            unit="kg",
            qty=10.0,
        )
        res = self.client.post(
            "/stores/stock/transfer",
            json={
                "outlet": "restaurant",
                "item_name": "Tomato",
                "unit": "kg",
                "qty": 5,
                "direction": "to_counter",
            },
        )
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        payload = res.get_json()
        self.assertTrue(payload.get("ok"))
        self.assertAlmostEqual(self._qty("restaurant", "Tomato", "kg", "warehouse"), 5.0)
        self.assertAlmostEqual(self._qty("restaurant", "Tomato", "kg", "counter"), 5.0)
        conn = db_mod.get_db()
        try:
            moves = conn.execute(
                """
                SELECT place, qty_delta, movement_type FROM store_stock_movements
                WHERE movement_type = 'transfer' AND lower(item_name) = lower('Tomato')
                ORDER BY id ASC
                """
            ).fetchall()
            self.assertEqual(len(moves), 2)
            by_place = {row["place"]: float(row["qty_delta"]) for row in moves}
            self.assertAlmostEqual(by_place["warehouse"], -5.0)
            self.assertAlmostEqual(by_place["counter"], 5.0)
            self.assertAlmostEqual(sum(by_place.values()), 0.0)
        finally:
            conn.close()

    def test_transfer_to_other_outlet_counter(self):
        self._insert_stock(
            outlet="restaurant",
            place="warehouse",
            item_name="Onion",
            unit="kg",
            qty=10.0,
        )
        res = self.client.post(
            "/stores/stock/transfer",
            json={
                "direction": "to_counter",
                "to_outlet": "bar",
                "items": [
                    {
                        "outlet": "restaurant",
                        "item_name": "Onion",
                        "unit": "kg",
                        "qty": 4,
                    }
                ],
            },
        )
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        payload = res.get_json()
        self.assertTrue(payload.get("ok"))
        self.assertEqual(payload.get("to_outlet"), "bar")
        self.assertAlmostEqual(self._qty("restaurant", "Onion", "kg", "warehouse"), 6.0)
        self.assertAlmostEqual(self._qty("restaurant", "Onion", "kg", "counter"), 0.0)
        self.assertAlmostEqual(self._qty("bar", "Onion", "kg", "counter"), 4.0)
        self.assertAlmostEqual(self._qty("bar", "Onion", "kg", "warehouse"), 0.0)

    def test_transfer_more_than_source_returns_400(self):
        self._insert_stock(
            outlet="restaurant",
            place="warehouse",
            item_name="Tomato",
            unit="kg",
            qty=2.0,
        )
        res = self.client.post(
            "/stores/stock/transfer",
            json={
                "outlet": "restaurant",
                "item_name": "Tomato",
                "unit": "kg",
                "qty": 5,
                "direction": "to_counter",
            },
        )
        self.assertEqual(res.status_code, 400)
        self.assertFalse((res.get_json() or {}).get("ok"))
        self.assertAlmostEqual(self._qty("restaurant", "Tomato", "kg", "warehouse"), 2.0)
        self.assertAlmostEqual(self._qty("restaurant", "Tomato", "kg", "counter"), 0.0)

    def test_transfer_counter_to_warehouse(self):
        self._insert_stock(
            outlet="restaurant",
            place="counter",
            item_name="Tomato",
            unit="kg",
            qty=8.0,
        )
        self._insert_stock(
            outlet="restaurant",
            place="warehouse",
            item_name="Tomato",
            unit="kg",
            qty=1.0,
        )
        res = self.client.post(
            "/stores/stock/transfer",
            json={
                "outlet": "restaurant",
                "item_name": "Tomato",
                "unit": "kg",
                "qty": 3,
                "direction": "to_warehouse",
                "to_outlet": "restaurant",
            },
        )
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        payload = res.get_json()
        self.assertTrue(payload.get("ok"))
        self.assertEqual(payload.get("from_place"), "counter")
        self.assertEqual(payload.get("to_place"), "warehouse")
        self.assertAlmostEqual(self._qty("restaurant", "Tomato", "kg", "counter"), 5.0)
        self.assertAlmostEqual(self._qty("restaurant", "Tomato", "kg", "warehouse"), 4.0)

    def test_transfer_counter_to_warehouse_over_qty_returns_400(self):
        self._insert_stock(
            outlet="restaurant",
            place="counter",
            item_name="Cream",
            unit="liter",
            qty=2.0,
        )
        self._insert_stock(
            outlet="restaurant",
            place="warehouse",
            item_name="Cream",
            unit="liter",
            qty=50.0,
        )
        res = self.client.post(
            "/stores/stock/transfer",
            json={
                "direction": "to_warehouse",
                "to_outlet": "restaurant",
                "items": [
                    {
                        "outlet": "restaurant",
                        "item_name": "Cream",
                        "unit": "liter",
                        "qty": 5,
                    }
                ],
            },
        )
        self.assertEqual(res.status_code, 400)
        body = res.get_json() or {}
        self.assertFalse(body.get("ok"))
        self.assertIn("counter", (body.get("error") or "").lower())
        # Warehouse stock must not be returnable from the counter direction.
        self.assertAlmostEqual(self._qty("restaurant", "Cream", "liter", "counter"), 2.0)
        self.assertAlmostEqual(self._qty("restaurant", "Cream", "liter", "warehouse"), 50.0)

    def test_transfer_counter_to_warehouse_respects_pos_deduction(self):
        """Return qty is capped to counter on-hand after POS sale deduction."""
        self._insert_stock(
            outlet="restaurant",
            place="counter",
            item_name="Strawberry Ice Cream",
            unit="kg",
            qty=5.0,
        )
        self._insert_stock(
            outlet="restaurant",
            place="warehouse",
            item_name="Strawberry Ice Cream",
            unit="kg",
            qty=10.0,
        )
        # Simulate POS close deducting from counter (same place as deduct_stock_for_pos_invoice).
        conn = db_mod.get_db()
        try:
            self.stores_mod._adjust_stock(
                conn,
                outlet="restaurant",
                place=self.stores_mod.STOCK_PLACE_COUNTER,
                item_name="Strawberry Ice Cream",
                unit="kg",
                qty_delta=-0.3,
                movement_type="sale",
                ref_type="pos_invoice",
                ref_id=999001,
                notes="POS sale test",
                user_id=self.admin_id,
                allow_shortfall=True,
            )
            conn.commit()
        finally:
            conn.close()
        self.assertAlmostEqual(
            self._qty("restaurant", "Strawberry Ice Cream", "kg", "counter"), 4.7
        )
        self.assertAlmostEqual(
            self._qty("restaurant", "Strawberry Ice Cream", "kg", "warehouse"), 10.0
        )

        too_much = self.client.post(
            "/stores/stock/transfer",
            json={
                "direction": "counter_to_warehouse",
                "to_outlet": "restaurant",
                "outlet": "restaurant",
                "item_name": "Strawberry Ice Cream",
                "unit": "kg",
                "qty": 4.8,
            },
        )
        self.assertEqual(too_much.status_code, 400)
        self.assertFalse((too_much.get_json() or {}).get("ok"))
        self.assertAlmostEqual(
            self._qty("restaurant", "Strawberry Ice Cream", "kg", "counter"), 4.7
        )
        self.assertAlmostEqual(
            self._qty("restaurant", "Strawberry Ice Cream", "kg", "warehouse"), 10.0
        )

        ok = self.client.post(
            "/stores/stock/transfer",
            json={
                "direction": "to_warehouse",
                "to_outlet": "restaurant",
                "outlet": "restaurant",
                "item_name": "Strawberry Ice Cream",
                "unit": "kg",
                "qty": 4.7,
            },
        )
        self.assertEqual(ok.status_code, 200, ok.get_data(as_text=True))
        self.assertTrue((ok.get_json() or {}).get("ok"))
        self.assertAlmostEqual(
            self._qty("restaurant", "Strawberry Ice Cream", "kg", "counter"), 0.0
        )
        self.assertAlmostEqual(
            self._qty("restaurant", "Strawberry Ice Cream", "kg", "warehouse"), 14.7
        )

    def test_batch_transfer_custom_qtys(self):
        self._insert_stock(
            outlet="restaurant",
            place="warehouse",
            item_name="Onion",
            unit="kg",
            qty=25.0,
        )
        self._insert_stock(
            outlet="restaurant",
            place="warehouse",
            item_name="Potato",
            unit="kg",
            qty=40.0,
        )
        res = self.client.post(
            "/stores/stock/transfer",
            json={
                "direction": "to_counter",
                "items": [
                    {
                        "outlet": "restaurant",
                        "item_name": "Onion",
                        "unit": "kg",
                        "qty": 5,
                    },
                    {
                        "outlet": "restaurant",
                        "item_name": "Potato",
                        "unit": "kg",
                        "qty": 12,
                    },
                ],
            },
        )
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        payload = res.get_json()
        self.assertTrue(payload.get("ok"))
        self.assertEqual(payload.get("count"), 2)
        self.assertAlmostEqual(self._qty("restaurant", "Onion", "kg", "warehouse"), 20.0)
        self.assertAlmostEqual(self._qty("restaurant", "Onion", "kg", "counter"), 5.0)
        self.assertAlmostEqual(self._qty("restaurant", "Potato", "kg", "warehouse"), 28.0)
        self.assertAlmostEqual(self._qty("restaurant", "Potato", "kg", "counter"), 12.0)
        conn = db_mod.get_db()
        try:
            moves = conn.execute(
                """
                SELECT place, qty_delta FROM store_stock_movements
                WHERE movement_type = 'transfer'
                  AND lower(item_name) IN ('onion', 'potato')
                ORDER BY id ASC
                """
            ).fetchall()
            self.assertEqual(len(moves), 4)
            self.assertAlmostEqual(sum(float(row["qty_delta"]) for row in moves), 0.0)
        finally:
            conn.close()

    def test_batch_transfer_over_qty_rolls_back(self):
        self._insert_stock(
            outlet="restaurant",
            place="warehouse",
            item_name="Onion",
            unit="kg",
            qty=25.0,
        )
        self._insert_stock(
            outlet="restaurant",
            place="warehouse",
            item_name="Potato",
            unit="kg",
            qty=4.0,
        )
        res = self.client.post(
            "/stores/stock/transfer",
            json={
                "direction": "to_counter",
                "items": [
                    {
                        "outlet": "restaurant",
                        "item_name": "Onion",
                        "unit": "kg",
                        "qty": 5,
                    },
                    {
                        "outlet": "restaurant",
                        "item_name": "Potato",
                        "unit": "kg",
                        "qty": 12,
                    },
                ],
            },
        )
        self.assertEqual(res.status_code, 400)
        body = res.get_json() or {}
        self.assertFalse(body.get("ok"))
        self.assertIn("Potato", body.get("error") or "")
        self.assertAlmostEqual(self._qty("restaurant", "Onion", "kg", "warehouse"), 25.0)
        self.assertAlmostEqual(self._qty("restaurant", "Onion", "kg", "counter"), 0.0)
        self.assertAlmostEqual(self._qty("restaurant", "Potato", "kg", "warehouse"), 4.0)
        self.assertAlmostEqual(self._qty("restaurant", "Potato", "kg", "counter"), 0.0)

    def test_stock_page_filters_by_place(self):
        self._insert_stock(
            outlet="restaurant",
            place="warehouse",
            item_name="Warehouse Only Flour",
            unit="kg",
            qty=8.0,
        )
        self._insert_stock(
            outlet="restaurant",
            place="counter",
            item_name="Counter Only Cream",
            unit="liter",
            qty=2.0,
        )
        warehouse = self.client.get("/stores/stock?outlet=restaurant&place=warehouse")
        self.assertEqual(warehouse.status_code, 200)
        self.assertIn(b"Warehouse Only Flour", warehouse.data)
        self.assertNotIn(b"Counter Only Cream", warehouse.data)
        self.assertIn(b'data-place="warehouse"', warehouse.data)
        self.assertIn(b'st-stock-row-check', warehouse.data)
        self.assertIn(b'id="st-stock-transfer-selected"', warehouse.data)
        self.assertIn(b'data-st-transfer="to_counter"', warehouse.data)

        counter = self.client.get("/stores/stock?outlet=restaurant&place=counter")
        self.assertEqual(counter.status_code, 200)
        self.assertIn(b"Counter Only Cream", counter.data)
        self.assertNotIn(b"Warehouse Only Flour", counter.data)
        self.assertIn(b'data-place="counter"', counter.data)
        self.assertIn(b"Return", counter.data)
        self.assertIn(b'data-st-transfer="to_warehouse"', counter.data)
        self.assertIn(b'data-qty="2.0"', counter.data)

        empty_counter = self.client.get("/stores/stock?outlet=bar&place=counter")
        self.assertEqual(empty_counter.status_code, 200)
        self.assertIn(b"Counter is empty", empty_counter.data)
        self.assertIn(b"Return", empty_counter.data)
        self.assertIn(b'id="st-stock-transfer-lines"', counter.data)
        self.assertIn(b'id="st-stock-transfer-to-outlet"', warehouse.data)
        self.assertIn(b'id="st-stock-transfer-to-outlet" name="to_outlet" value=""', warehouse.data)
        self.assertIn(b"Restaurant Counter", warehouse.data)
        self.assertIn(b"Bar Counter", warehouse.data)
        self.assertIn(
            b'id="st-stock-transfer-dest-restaurant" aria-selected="false"',
            warehouse.data,
        )
        self.assertIn(
            b'id="st-stock-transfer-dest-bar" aria-selected="false"',
            warehouse.data,
        )
        self.assertIn(
            b'id="st-stock-transfer-submit" disabled',
            warehouse.data,
        )
        self.assertIn(b'id="st-stock-transfer-qty-mode"', warehouse.data)
        self.assertIn(b'id="st-stock-transfer-mode-pack"', warehouse.data)
        self.assertIn(b'id="st-stock-transfer-mode-total"', warehouse.data)
        self.assertIn(
            b'id="st-stock-transfer-mode-pack" aria-selected="true"',
            warehouse.data,
        )
        self.assertIn(b'data-pack-label=', counter.data)
        self.assertIn(b'data-pack-qty-in-base=', counter.data)

    def test_export_respects_place(self):
        self._insert_stock(
            outlet="restaurant",
            place="warehouse",
            item_name="Warehouse Only Flour",
            unit="kg",
            qty=8.0,
        )
        self._insert_stock(
            outlet="restaurant",
            place="counter",
            item_name="Counter Only Cream",
            unit="liter",
            qty=2.0,
        )
        from openpyxl import load_workbook

        warehouse = self.client.get("/stores/stock/export?outlet=restaurant&place=warehouse")
        self.assertEqual(warehouse.status_code, 200)
        self.assertIn(
            "Store Warehouse",
            warehouse.headers.get("Content-Disposition", ""),
        )
        wb = load_workbook(io.BytesIO(warehouse.data))
        names = [ws.cell(r, 1).value for ws in [wb.active] for r in range(2, wb.active.max_row + 1)]
        self.assertIn("Warehouse Only Flour", names)
        self.assertNotIn("Counter Only Cream", names)
        self.assertEqual(wb.active.cell(1, 9).value, "Place")

        counter = self.client.get("/stores/stock/export?outlet=restaurant&place=counter")
        self.assertEqual(counter.status_code, 200)
        self.assertIn(
            "Store Counter",
            counter.headers.get("Content-Disposition", ""),
        )
        wb2 = load_workbook(io.BytesIO(counter.data))
        names2 = [wb2.active.cell(r, 1).value for r in range(2, wb2.active.max_row + 1)]
        self.assertIn("Counter Only Cream", names2)
        self.assertNotIn("Warehouse Only Flour", names2)

    def test_audit_verify_adjusts_selected_place_only(self):
        self._insert_stock(
            outlet="restaurant",
            place="warehouse",
            item_name="Tomato",
            unit="kg",
            qty=10.0,
        )
        self._insert_stock(
            outlet="restaurant",
            place="counter",
            item_name="Tomato",
            unit="kg",
            qty=4.0,
        )
        page = self.client.get("/stores/stock-audit?outlet=restaurant&place=counter")
        self.assertEqual(page.status_code, 200)
        self.assertIn(b"Tomato", page.data)

        conn = db_mod.get_db()
        try:
            line = conn.execute(
                """
                SELECT l.id, l.system_qty, a.place
                FROM store_stock_audit_lines l
                JOIN store_stock_audits a ON a.id = l.audit_id
                WHERE a.outlet = 'restaurant' AND a.status = 'open'
                  AND a.place = 'counter' AND l.item_name = 'Tomato'
                ORDER BY l.id DESC LIMIT 1
                """
            ).fetchone()
            self.assertIsNotNone(line)
            self.assertAlmostEqual(float(line["system_qty"]), 4.0)
            line_id = int(line["id"])
        finally:
            conn.close()

        verify = self.client.post(
            "/stores/stock-audit/verify",
            json={
                "line_id": line_id,
                "actual_qty": 3.0,
                "reason": "kitchen_wastage",
                "remarks": "Prep",
            },
        )
        self.assertEqual(verify.status_code, 200, verify.get_data(as_text=True))
        self.assertTrue(verify.get_json().get("ok"))
        self.assertAlmostEqual(self._qty("restaurant", "Tomato", "kg", "counter"), 3.0)
        self.assertAlmostEqual(self._qty("restaurant", "Tomato", "kg", "warehouse"), 10.0)

    def _insert_product_with_pack(self, *, name, unit, outlet, pack_label, pack_qty_in_base):
        conn = db_mod.get_db()
        try:
            db_mod.ensure_stores_schema(conn)
            category = conn.execute(
                "SELECT id FROM store_product_categories WHERE is_active = 1 ORDER BY id LIMIT 1"
            ).fetchone()
            if category is None:
                conn.execute(
                    """
                    INSERT INTO store_product_categories (name, is_active, sort_order)
                    VALUES ('Dairy', 1, 1)
                    """
                )
                category_id = conn.execute(
                    "SELECT id FROM store_product_categories WHERE name = 'Dairy'"
                ).fetchone()["id"]
            else:
                category_id = category["id"]
            cur = conn.execute(
                """
                INSERT INTO store_products
                    (category_id, name, default_unit, outlet, approximate_price, is_active, sort_order)
                VALUES (?, ?, ?, ?, 100, 1, 10)
                """,
                (category_id, name, unit, outlet),
            )
            product_id = cur.lastrowid
            conn.execute(
                """
                INSERT INTO store_product_variants
                    (product_id, label, qty_in_base, approximate_price, sort_order, is_active)
                VALUES (?, ?, ?, 50, 10, 1)
                """,
                (product_id, pack_label, pack_qty_in_base),
            )
            conn.commit()
            return int(product_id)
        finally:
            conn.close()

    def test_stock_page_includes_default_pack_attrs(self):
        self._insert_product_with_pack(
            name="Amul Cheese",
            unit="gram",
            outlet="restaurant",
            pack_label="500 gram",
            pack_qty_in_base=500.0,
        )
        self._insert_stock(
            outlet="restaurant",
            place="warehouse",
            item_name="Amul Cheese",
            unit="gram",
            qty=1000.0,
        )
        self._insert_stock(
            outlet="restaurant",
            place="warehouse",
            item_name="Loose Spice",
            unit="kg",
            qty=3.0,
        )
        page = self.client.get("/stores/stock?outlet=restaurant&place=warehouse")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn('data-item-name="Amul Cheese"', html)
        self.assertIn('data-pack-label="500 gram"', html)
        self.assertIn('data-pack-qty-in-base="500', html)
        self.assertIn('data-item-name="Loose Spice"', html)
        # No product master pack → empty pack attrs, transfer still allowed as base qty.
        self.assertRegex(
            html,
            r'data-item-name="Loose Spice"[^>]*data-pack-label=""[^>]*data-pack-qty-in-base=""',
        )

    def test_stock_page_embeds_product_packs_json_for_transfer(self):
        """Product Master packs are embedded so Transfer works even if row attrs are stale."""
        self._insert_product_with_pack(
            name="100 Pipers",
            unit="ml",
            outlet="bar",
            pack_label="750 mL",
            pack_qty_in_base=750.0,
        )
        self._insert_stock(
            outlet="bar",
            place="warehouse",
            item_name="100 Pipers",
            unit="ml",
            qty=750.0,
        )
        page = self.client.get("/stores/stock?outlet=bar&place=warehouse")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn('id="st-stock-product-packs"', html)
        self.assertIn('data-pack-label="750 mL"', html)
        self.assertIn('data-pack-qty-in-base="750', html)
        start = html.find('id="st-stock-product-packs"')
        self.assertGreaterEqual(start, 0)
        open_tag = html.find('>', start)
        close_tag = html.find('</script>', open_tag)
        payload = html[open_tag + 1:close_tag].strip()
        packs = json.loads(payload)
        self.assertIsInstance(packs, dict)
        # name / name|unit / name|outlet|unit keys
        for key in ("100 pipers", "100 pipers|ml", "100 pipers|bar|ml"):
            self.assertIn(key, packs, msg=f"missing pack map key {key}")
            entry = packs[key]
            self.assertEqual(entry.get("label"), "750 mL")
            self.assertAlmostEqual(float(entry.get("qty_in_base")), 750.0)

    def test_default_pack_parses_qty_from_label_when_missing(self):
        label, qty = self.stores_mod._default_pack_from_product_variants(
            [{"label": "750 mL", "qty_in_base": 0}]
        )
        self.assertEqual(label, "750 mL")
        self.assertAlmostEqual(qty, 750.0)
        label2, qty2 = self.stores_mod._default_pack_from_product_variants(
            [{"label": "Half bottle", "qty_in_base": None}]
        )
        self.assertEqual(label2, "Half bottle")
        self.assertIsNone(qty2)

    def test_transfer_api_still_accepts_base_qty_for_packed_item(self):
        """UI Pack mode converts to base before POST; API remains base-unit qty."""
        self._insert_product_with_pack(
            name="Amul Cheese",
            unit="gram",
            outlet="restaurant",
            pack_label="500 gram",
            pack_qty_in_base=500.0,
        )
        self._insert_stock(
            outlet="restaurant",
            place="warehouse",
            item_name="Amul Cheese",
            unit="gram",
            qty=1000.0,
        )
        # Pack mode entering 1 pack → JS sends 500 base.
        resp = self.client.post(
            "/stores/stock/transfer",
            json={
                "direction": "to_counter",
                "to_outlet": "restaurant",
                "items": [
                    {
                        "outlet": "restaurant",
                        "item_name": "Amul Cheese",
                        "unit": "gram",
                        "qty": 500,
                    }
                ],
            },
        )
        self.assertEqual(resp.status_code, 200, resp.get_data(as_text=True))
        body = resp.get_json()
        self.assertTrue(body.get("ok"))
        self.assertAlmostEqual(float(body.get("qty")), 500.0)
        self.assertAlmostEqual(self._qty("restaurant", "Amul Cheese", "gram", "warehouse"), 500.0)
        self.assertAlmostEqual(self._qty("restaurant", "Amul Cheese", "gram", "counter"), 500.0)

        # 100 Pipers: 1 × 750 mL pack resolves to 750 base units on the wire.
        self._insert_product_with_pack(
            name="100 Pipers",
            unit="ml",
            outlet="bar",
            pack_label="750 mL",
            pack_qty_in_base=750.0,
        )
        self._insert_stock(
            outlet="bar",
            place="warehouse",
            item_name="100 Pipers",
            unit="ml",
            qty=750.0,
        )
        resp2 = self.client.post(
            "/stores/stock/transfer",
            json={
                "direction": "to_counter",
                "to_outlet": "bar",
                "items": [
                    {
                        "outlet": "bar",
                        "item_name": "100 Pipers",
                        "unit": "ml",
                        "qty": 750,
                    }
                ],
            },
        )
        self.assertEqual(resp2.status_code, 200, resp2.get_data(as_text=True))
        body2 = resp2.get_json()
        self.assertTrue(body2.get("ok"))
        self.assertAlmostEqual(float(body2.get("qty")), 750.0)
        self.assertAlmostEqual(self._qty("bar", "100 Pipers", "ml", "warehouse"), 0.0)
        self.assertAlmostEqual(self._qty("bar", "100 Pipers", "ml", "counter"), 750.0)

    def test_schema_rebuild_moves_existing_on_hand_to_warehouse(self):

        conn = db_mod.get_db()
        try:
            conn.execute("DROP TABLE store_stock_items")
            conn.execute(
                """
                CREATE TABLE store_stock_items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    outlet TEXT NOT NULL,
                    item_name TEXT NOT NULL,
                    unit TEXT NOT NULL DEFAULT 'pcs',
                    qty_on_hand REAL NOT NULL DEFAULT 0,
                    updated_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
                    UNIQUE(outlet, item_name, unit)
                )
                """
            )
            conn.execute(
                """
                INSERT INTO store_stock_items (outlet, item_name, unit, qty_on_hand, updated_at)
                VALUES ('restaurant', 'Legacy Onion', 'kg', 7.5, datetime('now','localtime'))
                """
            )
            conn.commit()
            db_mod.ensure_stores_schema(conn)
            conn.commit()
            row = conn.execute(
                """
                SELECT place, qty_on_hand FROM store_stock_items
                WHERE outlet = 'restaurant' AND item_name = 'Legacy Onion'
                """
            ).fetchone()
            self.assertIsNotNone(row)
            self.assertEqual(row["place"], "warehouse")
            self.assertAlmostEqual(float(row["qty_on_hand"]), 7.5)
            sql = conn.execute(
                "SELECT sql FROM sqlite_master WHERE type='table' AND name='store_stock_items'"
            ).fetchone()[0]
            compact = "".join(sql.lower().split())
            self.assertIn("unique(outlet,place,item_name,unit)", compact)
        finally:
            conn.close()


if __name__ == "__main__":
    unittest.main()
