#!/usr/bin/env python3
"""Generate a kitchen-fillable Excel template for menu recipe ingredients.

Reads active menu items and Product Master from bell_elite.db and writes
Menu_Recipe_Ingredients.xlsx with dropdowns for product name and unit.
"""
from __future__ import annotations

import argparse
import os
import sqlite3
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_DB = os.path.join(ROOT, "bell_elite.db")
DEFAULT_OUT = os.path.expanduser(
    "~/Downloads/Template/Menu_Recipe_Ingredients.xlsx"
)

RECIPE_UNITS = (
    "g",
    "kg",
    "ml",
    "liter",
    "pcs",
    "bunch",
    "bottle",
    "pack",
    "dozen",
    "case",
)

# Each menu gets this many blank ingredient rows so kitchen can list multiple products.
INGREDIENT_SLOTS_PER_MENU = 8
# Extra blank rows at the bottom (dropdowns still apply) if a dish needs even more.
EXTRA_BLANK_ROWS = 200

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
TITLE_FONT = Font(bold=True, name="Calibri", size=14)
BODY_FONT = Font(name="Calibri", size=11)


def connect(db_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def load_menus(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return list(
        conn.execute(
            """
            SELECT
                i.id AS menu_id,
                COALESCE(NULLIF(TRIM(i.outlet), ''), 'restaurant') AS outlet,
                COALESCE(NULLIF(TRIM(c.name), ''), '') AS category,
                TRIM(i.name) AS menu_name
            FROM pos_menu_items i
            LEFT JOIN pos_menu_categories c ON c.id = i.category_id
            WHERE COALESCE(i.is_active, 1) = 1
            ORDER BY
                CASE LOWER(COALESCE(i.outlet, 'restaurant'))
                    WHEN 'restaurant' THEN 0
                    WHEN 'bar' THEN 1
                    ELSE 2
                END,
                COALESCE(c.sort_order, 0),
                COALESCE(c.name, ''),
                COALESCE(i.sort_order, 0),
                i.name
            """
        )
    )


def load_products(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return list(
        conn.execute(
            """
            SELECT
                p.id AS product_id,
                TRIM(p.name) AS product_name,
                COALESCE(NULLIF(TRIM(p.default_unit), ''), '') AS default_unit,
                COALESCE(NULLIF(TRIM(p.outlet), ''), 'both') AS outlet,
                COALESCE(NULLIF(TRIM(c.name), ''), '') AS category_name
            FROM store_products p
            LEFT JOIN store_product_categories c ON c.id = p.category_id
            WHERE COALESCE(p.is_active, 1) = 1
            ORDER BY TRIM(p.name) COLLATE NOCASE
            """
        )
    )


def style_header(ws, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def autosize(ws, widths: dict[str, float]) -> None:
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def write_instructions(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "Menu Recipe Ingredients — Kitchen Template"
    ws["A1"].font = TITLE_FONT
    lines = [
        "",
        "Purpose",
        "Fill ingredient lines for each menu item. Kitchen can edit this Excel offline; later it can be imported into the app.",
        "",
        "How to fill (Recipes sheet)",
        "1. Each row is one ingredient for one menu item.",
        "2. Every menu is repeated on several rows (8 slots) so one dish can have multiple ingredients.",
        "3. Outlet, Category, Menu Name, and Menu ID are pre-filled — do not change Menu ID.",
        "4. On each slot row: choose Ingredient Product, Unit, and enter Qty.",
        "5. Example: CLEAR SOUP CHICKEN might use Chicken 150 g on row 1, Ginger 10 g on row 2, etc.",
        "6. Leave unused ingredient slots blank (Ingredient / Unit / Qty empty).",
        "7. If a dish needs more than 8 ingredients, copy an extra menu row at the bottom blank area.",
        "",
        "Sheets",
        "Recipes — main data entry (multiple ingredient rows per menu)",
        "Menus — full menu list (reference)",
        "Products — Product Master names used by the Ingredient dropdown",
        "Units — unit list used by the Unit dropdown",
        "",
        "Tips",
        "Filter Recipes by Outlet, Category, or Menu Name when working one dish at a time.",
        "Prefer recipe units that match the product (g/kg for weight, ml/liter for liquids, pcs for pieces).",
    ]
    for idx, text in enumerate(lines, start=2):
        cell = ws.cell(idx, 1, text)
        cell.font = BODY_FONT
        if text in ("Purpose", "How to fill (Recipes sheet)", "Sheets", "Tips"):
            cell.font = Font(bold=True, name="Calibri", size=12)
    ws.column_dimensions["A"].width = 110


def write_units(wb: Workbook) -> None:
    ws = wb.create_sheet("Units")
    ws.append(["Unit"])
    style_header(ws, 1)
    for unit in RECIPE_UNITS:
        ws.append([unit])
    autosize(ws, {"A": 14})
    ws.freeze_panes = "A2"


def write_products(wb: Workbook, products: list[sqlite3.Row]) -> int:
    ws = wb.create_sheet("Products")
    ws.append(["Product ID", "Product Name", "Default Unit", "Outlet", "Category"])
    style_header(ws, 5)
    for row in products:
        ws.append(
            [
                int(row["product_id"]),
                row["product_name"],
                row["default_unit"],
                row["outlet"],
                row["category_name"],
            ]
        )
    autosize(
        ws,
        {"A": 12, "B": 36, "C": 14, "D": 12, "E": 22},
    )
    ws.freeze_panes = "A2"
    return len(products)


def write_menus(wb: Workbook, menus: list[sqlite3.Row]) -> None:
    ws = wb.create_sheet("Menus")
    ws.append(["Outlet", "Category", "Menu Name", "Menu ID"])
    style_header(ws, 4)
    for row in menus:
        ws.append(
            [
                row["outlet"],
                row["category"],
                row["menu_name"],
                int(row["menu_id"]),
            ]
        )
    autosize(ws, {"A": 12, "B": 22, "C": 40, "D": 12})
    ws.freeze_panes = "A2"


def write_recipes(
    wb: Workbook,
    menus: list[sqlite3.Row],
    product_count: int,
    unit_count: int,
    slots_per_menu: int = INGREDIENT_SLOTS_PER_MENU,
) -> None:
    ws = wb.create_sheet("Recipes", 1)
    headers = [
        "Outlet",
        "Category",
        "Menu Name",
        "Menu ID",
        "Ingredient Product",
        "Unit",
        "Qty",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    for row in menus:
        base = [
            row["outlet"],
            row["category"],
            row["menu_name"],
            int(row["menu_id"]),
        ]
        for _ in range(max(1, int(slots_per_menu))):
            ws.append(base + ["", "", None])

    menu_rows = len(menus) * max(1, int(slots_per_menu))
    last_data_row = max(2, menu_rows + 1)
    for _ in range(EXTRA_BLANK_ROWS):
        ws.append(["", "", "", "", "", "", None])
    validate_through = 1 + menu_rows + EXTRA_BLANK_ROWS

    if product_count > 0:
        prod_dv = DataValidation(
            type="list",
            formula1=f"Products!$B$2:$B${product_count + 1}",
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Ingredient",
            error="Pick a product from the Product Master list.",
        )
        prod_dv.add(f"E2:E{validate_through}")
        ws.add_data_validation(prod_dv)

    if unit_count > 0:
        unit_dv = DataValidation(
            type="list",
            formula1=f"Units!$A$2:$A${unit_count + 1}",
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Unit",
            error="Pick a unit from the Units list.",
        )
        unit_dv.add(f"F2:F{validate_through}")
        ws.add_data_validation(unit_dv)

    autosize(
        ws,
        {
            "A": 12,
            "B": 22,
            "C": 40,
            "D": 12,
            "E": 36,
            "F": 12,
            "G": 10,
        },
    )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{last_data_row}"


def build_workbook(menus: list[sqlite3.Row], products: list[sqlite3.Row]) -> Workbook:
    wb = Workbook()
    write_instructions(wb)
    write_recipes(wb, menus, len(products), len(RECIPE_UNITS))
    write_menus(wb, menus)
    write_products(wb, products)
    write_units(wb)
    return wb


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--db", default=DEFAULT_DB, help="Path to bell_elite.db")
    parser.add_argument(
        "--out",
        default=DEFAULT_OUT,
        help="Output .xlsx path (default: ~/Downloads/Template/...)",
    )
    parser.add_argument(
        "--slots",
        type=int,
        default=INGREDIENT_SLOTS_PER_MENU,
        help=f"Ingredient rows per menu (default: {INGREDIENT_SLOTS_PER_MENU})",
    )
    args = parser.parse_args(argv)

    if not os.path.isfile(args.db):
        print(f"Database not found: {args.db}", file=sys.stderr)
        return 1

    out_dir = os.path.dirname(os.path.abspath(args.out))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    conn = connect(args.db)
    try:
        menus = load_menus(conn)
        products = load_products(conn)
    finally:
        conn.close()

    if not menus:
        print("No active menu items found.", file=sys.stderr)
        return 1

    slots = max(1, int(args.slots or INGREDIENT_SLOTS_PER_MENU))
    wb = Workbook()
    write_instructions(wb)
    # Patch instructions line with actual slot count when non-default.
    if slots != INGREDIENT_SLOTS_PER_MENU:
        wb["Instructions"]["A8"] = (
            f"2. Every menu is repeated on several rows ({slots} slots) "
            "so one dish can have multiple ingredients."
        )
    write_recipes(wb, menus, len(products), len(RECIPE_UNITS), slots_per_menu=slots)
    write_menus(wb, menus)
    write_products(wb, products)
    write_units(wb)
    wb.save(args.out)
    print(
        f"Wrote {args.out}\n"
        f"  Menus: {len(menus)}\n"
        f"  Ingredient slots per menu: {slots}\n"
        f"  Recipe data rows: {len(menus) * slots}\n"
        f"  Products: {len(products)}\n"
        f"  Units: {len(RECIPE_UNITS)}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
