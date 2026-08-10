"""Hotel Bell Elite — Sales Update application."""

import calendar
import io
import json
import os
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    pass

from flask import (
    Flask,
    Response,
    abort,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)

import auth_security
from mailer import app_base_url, send_account_unlock_email, smtp_configured
from db import (
    SQL_NOW,
    POS_INVOICE_ORDER_TYPES,
    POS_INVOICE_ORDER_TYPE_LABELS,
    POS_INVOICE_SETTLEMENT_STATUSES,
    POS_INVOICE_SETTLEMENT_STATUS_LABELS,
    POS_OUTLET_BAR,
    POS_OUTLET_RESTAURANT,
    clear_all_pos_tables,
    close_pos_invoice_and_free_table,
    delete_customer_record,
    delete_agency_record,
    ensure_cash_ledger_schema,
    ensure_customers_schema,
    ensure_agencies_schema,
    ensure_communication_hub_schema,
    ensure_hotel_rooms_schema,
    ensure_pos_schema,
    ensure_stores_schema,
    enrich_pos_floor_tables_for_display,
    get_customer,
    get_agency,
    get_db,
    get_hotel_rooms_layout,
    get_hotel_room,
    get_hotel_room_invoice,
    get_hotel_settings,
    get_hotel_tax_rates,
    get_hotel_tariff_rates,
    hotel_room_invoice_kpis,
    indian_fiscal_year_bounds,
    is_valid_agency_gst,
    list_hotel_room_invoices,
    save_hotel_room_checkin,
    save_hotel_room_reservation,
    generate_hotel_room_invoice,
    record_hotel_room_payment,
    record_hotel_room_invoice_payment,
    set_hotel_room_discount,
    append_hotel_room_folio_charge,
    update_hotel_room_charge,
    delete_hotel_room_charge,
    find_hotel_guest_by_mobile,
    clear_hotel_room_stay,
    transfer_hotel_room_stay,
    merge_hotel_room_billing,
    unmerge_hotel_rooms,
    set_hotel_merge_primary,
    enrich_hotel_room_merge_fields,
    upsert_customer,
    upsert_agency_by_name,
    get_open_pos_invoice_for_table,
    get_pos_floor_layout,
    hotel_rooms_status_counts,
    get_pos_invoice,
    get_pos_restaurant_settings,
    get_pos_tax_rates,
    init_db,
    list_customers,
    list_agencies,
    list_pos_invoices,
    list_pos_kot_pending_summary,
    list_pos_kot_tokens,
    apply_pos_kot_token_reductions,
    list_pos_menu_sales,
    list_customer_insights,
    list_pos_today_invoices,
    get_pos_menu_item_details,
    list_pos_menu_categories,
    list_pos_menu_items,
    list_store_products_lite,
    normalize_pos_outlet,
    pos_invoice_kpis,
    pos_menu_sales_kpis,
    customer_insights_kpis,
    pos_today_sales_summary,
    save_customer_record,
    save_agency_record,
    save_hotel_rooms_layout,
    save_hotel_settings,
    save_pos_floor_layout,
    update_hotel_room_status,
    save_pos_invoice,
    save_pos_menu_category,
    save_pos_menu_item,
    send_pos_invoice_pending_kot,
    settle_pos_invoice,
    sync_pos_floor_occupancy_from_open_orders,
    transfer_pos_invoice_table,
    merge_pos_invoice_tables,
    unmerge_pos_floor_tables,
    save_pos_restaurant_settings,
    search_customers,
    soft_delete_pos_invoice,
    cancel_pos_invoice,
    is_provisional_pos_order_no,
    reopen_pos_invoice_for_edit,
    soft_delete_pos_menu_category,
    soft_delete_pos_menu_item,
)
from fo_invoice_tax_parser import parse_fo_invoice_tax_report
from sales_report_parser import OUTLET_BAR, OUTLET_RESTAURANT, parse_sales_report
from workspace_access import (
    _ACCOUNTS_SUBMODULE_LABELS,
    _DASHBOARD_MODULE_LABELS,
    _DASHBOARD_MODULES,
    _PUBLIC_ENDPOINTS,
    _PAYROLL_SUBMODULE_LABELS,
    _SALES_ANALYTICS_SUBMODULE_LABELS,
    _STORES_SUBMODULE_LABELS,
    _USER_ACCESS_SUBMODULE_LABELS,
    access_module_tree,
    access_module_tree_ui,
    accounts_access_list,
    build_user_context,
    dashboard_access_list,
    delete_access_role,
    fetch_access_management_users,
    get_access_role,
    get_endpoint_accounts_submodule,
    get_endpoint_dashboard_module,
    get_endpoint_payroll_submodule,
    get_endpoint_stores_submodule,
    get_endpoint_user_access_submodule,
    SUPER_ADMINISTRATOR_ROLE_NAME,
    is_built_in_administrator_role,
    is_system_administrator,
    list_access_roles,
    normalize_username,
    payroll_access_list,
    reconcile_super_admin_only_dashboard_modules,
    role_summary_for_ui,
    sales_analytics_access_list,
    save_access_role_record,
    save_access_user_record,
    stores_access_list,
    user_access_submodule_list,
    user_can_access_accounts_submodule,
    user_can_access_dashboard,
    user_can_access_endpoint_accounts,
    user_can_access_endpoint_sales_analytics,
    user_can_access_endpoint_stores,
    user_can_access_payroll_submodule,
    user_can_access_sales_analytics_submodule,
    user_can_access_stores_submodule,
    user_can_access_supplier_master,
    user_can_access_customer_master,
    user_can_access_agency_master,
    user_can_access_user_access_submodule,
    user_can_edit_kot_sent_lines,
    user_can_approve_transactions,
    validate_access_role_form,
    validate_access_user_form,
)
from employee_payroll import register_employee_payroll
from embed_helpers import is_embed_request, is_partial_main_request
from hotel_id_documents import process_uploaded_id_document, resolve_stored_id_document
from user_photos import (
    delete_stored_user_photo,
    process_uploaded_user_photo,
    resolve_stored_user_photo,
)
from masters import build_masters_dashboard
from reports import build_reports_dashboard
from stores import register_stores
from communication_hub import register_communication_hub
from main_dashboard_data import (
    build_dow_avg,
    build_outlet_boards,
    build_sales_heatmap,
    build_sales_trend,
    build_top_selling_items,
    date_range_days,
    inr_compact,
    payment_mode_pct,
    sparkline_series_from_values,
)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "hotel-bell-elite-dev-key-change-in-production")

# Cookie session hardening for HTTPS / Android WebView (auth flow unchanged).
_app_env = (
    os.environ.get("FLASK_ENV")
    or os.environ.get("ENV")
    or os.environ.get("APP_ENV")
    or ""
).strip().lower()
_secure_cookies = (
    _app_env in ("production", "prod")
    or os.environ.get("SESSION_COOKIE_SECURE", "").strip() in ("1", "true", "True", "yes")
)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = bool(_secure_cookies)

init_db()

AUTH_USER_SESSION_KEY = "user_id"
AUTH_NOTICE_SESSION_KEY = "auth_notice"

SALES_COMPANY_LOCATIONS = {
    "HBE": {
        "label": "Hotel Bell Elite",
        "locations": ["Bar", "Restaurant"],
    }
}

SALES_ENTRY_FIELDS = (
    ("total_sales", "Total Sales"),
    ("cash", "Cash"),
    ("card", "Card"),
    ("upi", "UPI"),
    ("room_credit", "Room Transfer"),
    ("online_order", "Online Order"),
    ("tips", "Tips"),
    ("actual_cash", "Actual Cash"),
)

SALES_ENTRY_TOTAL_KEYS = (
    "cash",
    "card",
    "upi",
    "room_credit",
    "online_order",
)

MANUAL_SALES_ENTRY_KEYS = ("tips", "actual_cash")

SALES_DIGITAL_TRANSACTION_KEYS = ("card", "upi")

PETTY_CASH_DENOMINATIONS = (500, 200, 100, 50, 20, 10, 5, 2, 1)

SALES_CASH_DESTINATIONS = {
    "bank": "Bank deposit",
    "petty_cash": "Petty cash",
    "other": "Other",
}

DEFAULT_COMPANY = "HBE"
DEFAULT_LOCATION = OUTLET_BAR
OUTLET_HOTEL = "Hotel"
HOTEL_LOCATIONS = [OUTLET_HOTEL]
CASH_LEDGER_OUTLETS = (OUTLET_HOTEL, OUTLET_BAR, OUTLET_RESTAURANT)
TIP_OUTLET_LOCATIONS = CASH_LEDGER_OUTLETS
TIPS_FILTER_ALL = "All"
TIPS_FILTER_LOCATIONS = (TIPS_FILTER_ALL, *TIP_OUTLET_LOCATIONS)
CASH_LEDGER_FILTER_ALL = "All"
CASH_LEDGER_FILTER_LOCATIONS = (CASH_LEDGER_FILTER_ALL, *CASH_LEDGER_OUTLETS)
CASH_LEDGER_ENTRY_SALES = "sales_cash"
CASH_LEDGER_ENTRY_LOAD = "load_cash"
CASH_LEDGER_ENTRY_CREDIT = "credit_cash"
CASH_LEDGER_ENTRY_EXPENSE = "expense"
CASH_LEDGER_ENTRY_TRANSFER = "transfer_out"
CASH_LEDGER_ENTRY_LABELS = {
    CASH_LEDGER_ENTRY_SALES: "Actual Cash",
    CASH_LEDGER_ENTRY_LOAD: "Load Cash",
    CASH_LEDGER_ENTRY_CREDIT: "Credit Cash",
    CASH_LEDGER_ENTRY_EXPENSE: "Expense",
    CASH_LEDGER_ENTRY_TRANSFER: "Transfer Out",
}
CASH_LEDGER_ENTRY_RANK = {
    CASH_LEDGER_ENTRY_SALES: 0,
    CASH_LEDGER_ENTRY_LOAD: 1,
    CASH_LEDGER_ENTRY_CREDIT: 2,
    CASH_LEDGER_ENTRY_EXPENSE: 3,
    CASH_LEDGER_ENTRY_TRANSFER: 4,
}
CASH_LEDGER_TRANSFER_DESTINATIONS = (
    ("bank", "Bank"),
    ("owner", "Owner"),
)
CASH_LEDGER_TRANSFER_DESTINATION_LABELS = dict(CASH_LEDGER_TRANSFER_DESTINATIONS)
CASH_LEDGER_ALL_ENTRIES_FROM = date(2000, 1, 1)
HOTEL_PAYMENT_MODES = (
    ("cash", "Cash"),
    ("card", "Card"),
    ("upi", "UPI"),
    ("room_credit", "Credit"),
)

HOTEL_SALES_ENTRY_FIELDS = (
    ("total_sales", "Total Sales"),
    ("cash", "Cash"),
    ("card", "Card"),
    ("upi", "UPI"),
    ("room_credit", "Credit"),
    ("actual_cash", "Actual Cash"),
    ("tips", "Tips"),
    ("expense", "Expense"),
)

HOTEL_MANUAL_SALES_ENTRY_KEYS = ("actual_cash", "tips")

EXPENSE_PAYMENT_CASH = "cash"
EXPENSE_PAYMENT_BANK = "bank_transfer"
EXPENSE_PAYMENT_CREDIT = "credit"


def _sorted_label_choices(choices):
    """Sort (value, label) dropdown choices ascending by display label."""
    return tuple(
        sorted(
            choices,
            key=lambda item: (str(item[1] or "").casefold(), str(item[0] or "").casefold()),
        )
    )


EXPENSE_PAYMENT_TYPES = _sorted_label_choices((
    (EXPENSE_PAYMENT_CASH, "Cash"),
    (EXPENSE_PAYMENT_BANK, "Bank Transfer"),
    (EXPENSE_PAYMENT_CREDIT, "Credit"),
))
EXPENSE_CATEGORIES = _sorted_label_choices((
    ("grocery", "Grocery"),
    ("vegetables", "Vegetables"),
    ("travel", "Travel"),
    ("hardware", "Hardware"),
    ("tac", "TAC (Travel Agent commission)"),
    ("fruits", "Fruits"),
    ("snacks", "Snacks"),
    ("meat", "Meat"),
    ("sea_food", "Sea Food"),
    ("labour", "Labour"),
    ("salary", "Salary"),
    ("water_tank", "Water Tank"),
    ("liquor", "Liquor"),
    ("fuel", "Fuel"),
    ("other", "Other"),
))
EXPENSE_CATEGORY_LABELS = dict(EXPENSE_CATEGORIES)


def _slugify_expense_category_key(name):
    """Build a stable expense category key from a display name."""
    import re
    value = (name or "").strip().lower()
    value = value.replace("&", " and ")
    value = re.sub(r"[^a-z0-9]+", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    if not value:
        return ""
    if value[0].isdigit():
        value = "cat_" + value
    return value[:80]


def _expense_category_choices(conn=None):
    """Builtin + custom expense categories for dropdowns."""
    items = list(EXPENSE_CATEGORIES)
    seen = {key for key, _label in items}
    if conn is not None:
        try:
            rows = conn.execute(
                """
                SELECT category_key, name
                FROM expense_categories
                WHERE is_active = 1
                ORDER BY sort_order, lower(name), id
                """
            ).fetchall()
        except Exception:
            rows = []
        for row in rows:
            key = (row["category_key"] or "").strip()
            label = (row["name"] or "").strip()
            if not key or not label or key in seen:
                continue
            items.append((key, label))
            seen.add(key)
    return _sorted_label_choices(items)


def _expense_category_labels(conn=None):
    return dict(_expense_category_choices(conn))

HOTEL_IMPORT_FIELD_KEYS = ("total_sales", "cash", "card", "upi", "room_credit")
ROOM_TRANSFER_PAYMENT_STATUSES = _sorted_label_choices((
    ("unpaid", "Un Paid"),
    ("paid", "Paid"),
))
ROOM_TRANSFER_FILTER_ALL = "All"
ROOM_TRANSFER_FILTER_STATUSES = (
    ("unpaid", "Un Paid"),
    ("paid", "Paid"),
)
ROOM_TRANSFER_FILTER_LOCATIONS = (ROOM_TRANSFER_FILTER_ALL, OUTLET_BAR, OUTLET_RESTAURANT)
ROOM_TRANSFER_OUTLET_LOCATIONS = (OUTLET_BAR, OUTLET_RESTAURANT)
CREDIT_FILTER_LOCATIONS = (ROOM_TRANSFER_FILTER_ALL, OUTLET_HOTEL)
CREDIT_OUTLET_LOCATIONS = (OUTLET_HOTEL,)
PURCHASE_LEDGER_FILTER_ALL = "all"
EXPENSE_PAYMENT_LABELS = dict(EXPENSE_PAYMENT_TYPES)

IMPORT_FIELD_KEYS = ("total_sales", "cash", "card", "upi", "room_credit", "online_order")


def round_half_up(value, dec=0):
    try:
        quantum = Decimal("1").scaleb(-dec)
        return float(Decimal(str(value or 0)).quantize(quantum, rounding=ROUND_HALF_UP))
    except (InvalidOperation, TypeError, ValueError):
        return 0.0


def parse_money(value):
    try:
        return round_half_up(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def get_current_user():
    if getattr(g, "_auth_loaded", False):
        return getattr(g, "current_user", None)
    g._auth_loaded = True
    user_id = session.get(AUTH_USER_SESSION_KEY)
    if not user_id:
        g.current_user = None
        return None
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        user = build_user_context(conn, row) if row else None
    finally:
        conn.close()
    if user and not user.get("is_active"):
        session.pop(AUTH_USER_SESSION_KEY, None)
        user = None
    if user and user.get("is_locked"):
        session.pop(AUTH_USER_SESSION_KEY, None)
        user = None
    g.current_user = user
    return user


def _pop_auth_notice():
    return session.pop(AUTH_NOTICE_SESSION_KEY, "")


def _queue_auth_notice(message):
    session[AUTH_NOTICE_SESSION_KEY] = str(message or "").strip()


def _permission_denied_response(message):
    message = str(message or "You do not have access to this module.")
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.is_json:
        return jsonify({"error": message}), 403
    _queue_auth_notice(message)
    return redirect(url_for("home"))


register_employee_payroll(
    app,
    pop_auth_notice=_pop_auth_notice,
    permission_denied_response=_permission_denied_response,
    get_user=get_current_user,
    queue_auth_notice=_queue_auth_notice,
)
register_stores(
    app,
    pop_auth_notice=_pop_auth_notice,
    get_user=get_current_user,
)
register_communication_hub(
    app,
    pop_auth_notice=_pop_auth_notice,
    get_user=get_current_user,
)


def _access_nav_view():
    user = get_current_user()
    if user_can_access_user_access_submodule(user, "users"):
        return "users"
    if user_can_access_user_access_submodule(user, "add"):
        return "add"
    if user_can_access_user_access_submodule(user, "roles"):
        return "roles"
    return "users"


def _am_page_render(template, **kwargs):
    kwargs.setdefault("auth_notice", _pop_auth_notice())
    kwargs.setdefault("de_nav_section", "access")
    if "de_nav_access_view" not in kwargs:
        kwargs["de_nav_access_view"] = (
            "add" if kwargs.get("form_focus") else _access_nav_view()
        )
    # Edit / Add User: show < back to the Users list.
    if kwargs.get("form_focus") and kwargs.get("de_nav_access_view") != "roles":
        kwargs.setdefault("back_href", url_for("access_management"))
        kwargs.setdefault("back_label", "Back to Users")
    return render_template(template, **kwargs)


def _am_roles_page_render(template, **kwargs):
    kwargs.setdefault("de_nav_access_view", "roles")
    if kwargs.get("form_focus"):
        kwargs.setdefault("back_href", url_for("access_roles"))
        kwargs.setdefault("back_label", "Back to Roles")
    else:
        # Roles list → Users (parent User & Access page).
        kwargs.setdefault("back_href", url_for("access_management"))
        kwargs.setdefault("back_label", "Back to Users")
    return _am_page_render(template, **kwargs)


def _user_display_name(user):
    if not user:
        return "User"
    return (user.get("full_name") or user.get("username") or "User").strip()


def _user_avatar_text(user):
    name = _user_display_name(user)
    parts = [part for part in name.split() if part]
    if len(parts) >= 2:
        return (parts[0][0] + parts[1][0]).upper()
    return name[:2].upper() or "U"


@app.route("/webhook/whatsapp", methods=["GET", "POST"])
def whatsapp_webhook():
    """Meta WhatsApp Cloud API webhook (verify + indent Approve/Reject)."""
    from whatsapp_webhook import handle_events_post, handle_verification_get

    if request.method == "GET":
        body, status, headers = handle_verification_get(request)
        return body, status, headers
    return handle_events_post(request, get_db, ensure_stores_schema, ensure_communication_hub_schema)


@app.before_request
def enforce_access():
    endpoint = request.endpoint or ""
    if (
        endpoint == "service_worker"
        or request.path == "/sw.js"
        or request.path.startswith("/static/")
        or request.path.startswith("/webhook/")
        or request.path == "/communication-hub/api/mirror-export"
    ):
        return None

    user = get_current_user()
    if user and user.get("must_change_password"):
        if endpoint not in {"change_password", "logout", "favicon"}:
            return redirect(url_for("change_password"))

    if endpoint in _PUBLIC_ENDPOINTS:
        return None

    if not user:
        return redirect(url_for("index"))

    required_dashboard = get_endpoint_dashboard_module(endpoint)
    if required_dashboard and not user_can_access_dashboard(user, required_dashboard):
        agency_ok = (
            required_dashboard == "master"
            and endpoint
            in {
                "agency_master",
                "save_agency",
                "delete_agency",
                "export_agency_report",
                "create_agency",
                "list_agencies_api",
            }
            and user_can_access_agency_master(user)
        )
        customer_ok = (
            required_dashboard == "master"
            and endpoint
            in {
                "customer_master",
                "save_customer",
                "delete_customer",
                "export_customer_report",
            }
            and user_can_access_customer_master(user)
        )
        if not agency_ok and not customer_ok:
            label = _DASHBOARD_MODULE_LABELS.get(required_dashboard, "requested")
            return _permission_denied_response(f"You do not have access to {label}.")

    if not user_can_access_endpoint_sales_analytics(user, endpoint):
        return _permission_denied_response("You do not have access to this Sales Analytics section.")

    if not user_can_access_endpoint_accounts(user, endpoint):
        label = _ACCOUNTS_SUBMODULE_LABELS.get(
            get_endpoint_accounts_submodule(endpoint) or "",
            "requested Accounts section",
        )
        return _permission_denied_response(f"You do not have access to {label}.")

    required_user_access = get_endpoint_user_access_submodule(endpoint)
    if required_user_access and not user_can_access_user_access_submodule(user, required_user_access):
        label = _USER_ACCESS_SUBMODULE_LABELS.get(required_user_access, "requested User & Access section")
        return _permission_denied_response(f"You do not have access to {label}.")

    required_payroll = get_endpoint_payroll_submodule(endpoint)
    if required_payroll and not user_can_access_payroll_submodule(user, required_payroll):
        label = _PAYROLL_SUBMODULE_LABELS.get(required_payroll, "requested payroll section")
        return _permission_denied_response(f"You do not have access to the {label} payroll section.")

    if not user_can_access_endpoint_stores(user, endpoint):
        submodule = get_endpoint_stores_submodule(endpoint) or ""
        if submodule == "approvals":
            return _permission_denied_response("You do not have access to Approval.")
        label = _STORES_SUBMODULE_LABELS.get(
            submodule,
            "requested Purchase & Inventory section",
        )
        return _permission_denied_response(f"You do not have access to {label}.")

    return None


@app.context_processor
def inject_su_page_back():
    """When opened from Reports hub (?from_hub=reports), expose Back to Reports."""
    try:
        from_hub = (request.args.get("from_hub") or "").strip().lower()
    except RuntimeError:
        return {}
    if from_hub != "reports":
        return {"from_hub": from_hub} if from_hub else {}
    return {
        "from_hub": "reports",
        "back_href": url_for("reports"),
        "back_label": "Back to Reports",
    }


@app.context_processor
def inject_auth_context():
    user = get_current_user()
    return {
        "is_partial_main": is_partial_main_request(),
        "current_user": user,
        "user_can_access_dashboard": user_can_access_dashboard,
        "display_name": _user_display_name(user),
        "avatar_text": _user_avatar_text(user),
        "dashboard_modules_meta": _DASHBOARD_MODULES,
        "access_module_tree": access_module_tree(),
        "access_module_tree_ui": access_module_tree_ui(),
        "accessible_dashboard_modules": dashboard_access_list(user),
        "accessible_sales_analytics_modules": sales_analytics_access_list(user),
        "accessible_user_access_modules": user_access_submodule_list(user),
        "accessible_payroll_modules": payroll_access_list(user),
        "accessible_accounts_modules": accounts_access_list(user),
        "accessible_stores_modules": stores_access_list(user),
        "has_dashboard_access": lambda key: user_can_access_dashboard(user, key),
        "has_sales_analytics_access": lambda key: user_can_access_sales_analytics_submodule(user, key),
        "has_payroll_access": lambda key: user_can_access_payroll_submodule(user, key),
        "has_accounts_access": lambda key: user_can_access_accounts_submodule(user, key),
        "has_stores_access": lambda key: user_can_access_stores_submodule(user, key),
        "has_supplier_master_access": lambda: user_can_access_supplier_master(user),
        "has_customer_master_access": lambda: user_can_access_customer_master(user),
        "has_agency_master_access": lambda: user_can_access_agency_master(user),
        "has_user_access_submodule": lambda key: user_can_access_user_access_submodule(user, key),
        "user_can_edit_kot_sent_lines": user_can_edit_kot_sent_lines,
        "user_can_approve_transactions": user_can_approve_transactions,
        "dashboard_module_labels": _DASHBOARD_MODULE_LABELS,
        "sales_analytics_submodule_labels": _SALES_ANALYTICS_SUBMODULE_LABELS,
        "payroll_module_labels": _PAYROLL_SUBMODULE_LABELS,
        "accounts_module_labels": _ACCOUNTS_SUBMODULE_LABELS,
        "stores_module_labels": _STORES_SUBMODULE_LABELS,
        "user_access_submodule_labels": _USER_ACCESS_SUBMODULE_LABELS,
    }


def get_sales_entry_total(entries):
    return round_half_up(
        sum(parse_money(entries.get(key)) for key in SALES_ENTRY_TOTAL_KEYS),
        2,
    )


def get_denomination_total(counts_dict):
    total = 0
    for denom_str, count in (counts_dict or {}).items():
        try:
            total += int(denom_str) * int(count or 0)
        except (TypeError, ValueError):
            continue
    return total


def get_digital_transactions(entries):
    return round_half_up(
        sum(parse_money(entries.get(key)) for key in SALES_DIGITAL_TRANSACTION_KEYS),
        2,
    )


def get_difference(entries):
    return round_half_up(parse_money(entries.get("total_sales")) - get_sales_entry_total(entries), 2)


def get_cash_actual_difference(entries):
    return round_half_up(parse_money(entries.get("cash")) - parse_money(entries.get("actual_cash")), 2)


def _ledger_entry_to_dict(row):
    item = dict(row)
    for key in ("tariff", "discount", "extra_amount", "amount"):
        item[key] = round_half_up(item.get(key), 2)
    item["payment_mode"] = item.get("payment_mode") or "room_credit"
    item["invoice_number"] = item.get("invoice_number") or item.get("room") or ""
    return item


def load_hotel_ledger_entries(conn, company, location, sales_date):
    rows = conn.execute(
        """SELECT id, invoice_number, room, room_type, reserve_number, guest_name, company_name,
                  travel_agent, pax, room_plan, tariff, discount, extra_amount, amount,
                  payment_mode, sort_order, source_row
           FROM hotel_sales_ledger_entries
           WHERE company = ? AND location = ? AND sales_date = ?
           ORDER BY sort_order, id""",
        (company, location, sales_date),
    ).fetchall()
    return [_ledger_entry_to_dict(r) for r in rows]


def rollup_hotel_ledger_entries(entries):
    totals = {key: 0.0 for key in IMPORT_FIELD_KEYS}
    totals["total_sales"] = 0.0
    for entry in entries or []:
        amount = parse_money(entry.get("amount"))
        totals["total_sales"] = round_half_up(totals["total_sales"] + amount, 2)
        mode = (entry.get("payment_mode") or "").strip()
        if mode in SALES_ENTRY_TOTAL_KEYS:
            totals[mode] = round_half_up(totals[mode] + amount, 2)
    return totals


def replace_hotel_ledger_entries(conn, company, location, sales_date, lines):
    conn.execute(
        "DELETE FROM hotel_sales_ledger_entries WHERE company = ? AND location = ? AND sales_date = ?",
        (company, location, sales_date),
    )
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for line in lines:
        conn.execute(
            """INSERT INTO hotel_sales_ledger_entries
               (company, location, sales_date, invoice_number, room, room_type, reserve_number, guest_name,
                company_name, travel_agent, pax, room_plan, tariff, discount, extra_amount,
                amount, payment_mode, sort_order, source_row, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                company,
                location,
                sales_date,
                line.get("invoice_number", ""),
                line.get("room", ""),
                line.get("room_type", ""),
                line.get("reserve_number", ""),
                line.get("guest_name", ""),
                line.get("company_name", ""),
                line.get("travel_agent", ""),
                line.get("pax", ""),
                line.get("room_plan", ""),
                parse_money(line.get("tariff")),
                parse_money(line.get("discount")),
                parse_money(line.get("extra_amount")),
                parse_money(line.get("amount")),
                (line.get("payment_mode") or "").strip(),
                int(line.get("sort_order") or 0),
                line.get("source_row"),
                now,
                now,
            ),
        )


def hotel_ledger_to_credit_lines(entries, location=OUTLET_HOTEL):
    """Map hotel FO ledger room_credit rows into room_transfer_entries line dicts."""
    lines = []
    for entry in entries or []:
        mode = (entry.get("payment_mode") or "").strip()
        if mode != "room_credit":
            continue
        detail_parts = [
            (entry.get("company_name") or "").strip(),
            (entry.get("travel_agent") or "").strip(),
            (entry.get("room_type") or "").strip(),
        ]
        lines.append(
            {
                "location": location,
                "invoice_number": entry.get("invoice_number") or "",
                "outlet_name": (entry.get("company_name") or "").strip() or OUTLET_HOTEL,
                "table_room": entry.get("room") or "",
                "guest_name": entry.get("guest_name") or "",
                "ledger_detail": " · ".join(part for part in detail_parts if part),
                "amount": entry.get("amount"),
                "payment_status": "unpaid",
                "sort_order": entry.get("sort_order"),
                "source_row": entry.get("source_row"),
            }
        )
    return lines


def sync_hotel_credit_entries(conn, company, location, sales_date, ledger_entries=None):
    """Replace Hotel credit receivables for a date from FO ledger room_credit lines."""
    if ledger_entries is None:
        ledger_entries = load_hotel_ledger_entries(conn, company, location, sales_date)
    lines = hotel_ledger_to_credit_lines(ledger_entries, location=location)
    sync_room_transfer_entries(
        conn,
        company,
        sales_date,
        lines,
        locations=(location,),
    )


def sync_hotel_sales_from_ledger(conn, user, company, location, sales_date):
    entries = load_hotel_ledger_entries(conn, company, location, sales_date)
    sales_entries = rollup_hotel_ledger_entries(entries)
    existing_row = load_sales_row(company, location, sales_date)
    if existing_row:
        existing_values = existing_row.get("sales_entry_values") or {}
        for key in HOTEL_MANUAL_SALES_ENTRY_KEYS:
            sales_entries[key] = parse_money(existing_values.get(key))
    sales_entries = build_hotel_sales_entry_values(sales_entries)
    sales_entries["expense"] = _sales_expense_total(conn, company, location, sales_date)
    _apply_tip_line_total(conn, company, location, sales_date, sales_entries)
    existing_row = load_sales_row(company, location, sales_date)
    petty = (existing_row or {}).get("petty_cash_counts", {})
    cash_denoms = (existing_row or {}).get("cash_denomination_counts", {})
    upsert_sales_row(user, company, location, sales_date, sales_entries, petty, cash_denoms)
    sync_hotel_credit_entries(conn, company, location, sales_date, ledger_entries=entries)
    conn.commit()
    return {
        "entries": entries,
        "sales_entries": sales_entries,
        "sales_entry_total": get_sales_entry_total(sales_entries),
        "difference": get_difference(sales_entries),
    }


def _room_transfer_entry_paid_total(conn, entry_id):
    row = conn.execute(
        """SELECT COALESCE(SUM(amount), 0) AS total
           FROM room_transfer_payment_allocations
           WHERE room_transfer_entry_id = ?""",
        (entry_id,),
    ).fetchone()
    return round_half_up(row["total"] if row else 0, 2)


def _room_transfer_entry_balance(amount, paid_total):
    return round_half_up(max(parse_money(amount) - parse_money(paid_total), 0), 2)


def _sync_room_transfer_status_after_payment(conn, entry_id):
    entry = conn.execute(
        "SELECT id, amount FROM room_transfer_entries WHERE id = ?",
        (entry_id,),
    ).fetchone()
    if not entry:
        return
    paid_total = _room_transfer_entry_paid_total(conn, entry_id)
    balance = _room_transfer_entry_balance(entry["amount"], paid_total)
    payment_status = "paid" if balance <= 0.001 else "unpaid"
    conn.execute(
        """UPDATE room_transfer_entries
           SET payment_status = ?, updated_at = datetime('now','localtime')
           WHERE id = ?""",
        (payment_status, entry_id),
    )


def _room_transfer_entry_to_dict(row):
    item = dict(row)
    item["amount"] = round_half_up(item.get("amount"), 2)
    status = (item.get("payment_status") or "unpaid").strip().lower()
    item["payment_status"] = status if status in {"paid", "unpaid"} else "unpaid"
    paid_total = item.get("paid_amount")
    if paid_total is None:
        paid_total = 0.0
    item["paid_amount"] = round_half_up(paid_total, 2)
    item["balance"] = _room_transfer_entry_balance(item["amount"], item["paid_amount"])
    sales_date = item.get("sales_date") or ""
    try:
        parsed = date.fromisoformat(str(sales_date))
        item["sales_date_label"] = f"{parsed.day} {parsed.strftime('%b')}, {parsed.year}"
    except (TypeError, ValueError):
        item["sales_date_label"] = sales_date
    item.setdefault("payment_mode_label", "")
    return item


def _room_transfer_payment_modes_for_entries(conn, entry_ids):
    """Map entry id → display label for recorded clearance payment methods."""
    ids = []
    for raw in entry_ids or []:
        try:
            entry_id = int(raw)
        except (TypeError, ValueError):
            continue
        if entry_id and entry_id not in ids:
            ids.append(entry_id)
    if not ids:
        return {}
    placeholders = ",".join("?" for _ in ids)
    rows = conn.execute(
        f"""SELECT a.room_transfer_entry_id AS entry_id, p.payment_method
            FROM room_transfer_payment_allocations a
            JOIN room_transfer_payments p ON p.id = a.room_transfer_payment_id
            WHERE a.room_transfer_entry_id IN ({placeholders})
            ORDER BY a.room_transfer_entry_id, p.id""",
        ids,
    ).fetchall()
    by_entry = {}
    for row in rows:
        method = _normalize_room_transfer_payment_method(row["payment_method"])
        if not method:
            continue
        label = ROOM_TRANSFER_PAYMENT_METHOD_LABELS.get(method, method)
        bucket = by_entry.setdefault(int(row["entry_id"]), [])
        if label not in bucket:
            bucket.append(label)
    return {entry_id: " + ".join(labels) for entry_id, labels in by_entry.items()}


def _attach_room_transfer_payment_modes(conn, entries):
    modes = _room_transfer_payment_modes_for_entries(
        conn, [entry.get("id") for entry in (entries or [])]
    )
    for entry in entries or []:
        entry["payment_mode_label"] = modes.get(int(entry.get("id") or 0), "") or ""
    return entries


def load_room_transfer_entries(conn, company, sales_date):
    rows = conn.execute(
        """SELECT e.id, e.sales_date, e.location, e.invoice_number, e.outlet_name, e.table_room, e.guest_name,
                  e.ledger_detail, e.amount, e.payment_status, e.sort_order, e.source_row,
                  COALESCE((
                      SELECT SUM(a.amount) FROM room_transfer_payment_allocations a
                      WHERE a.room_transfer_entry_id = e.id
                  ), 0) AS paid_amount
           FROM room_transfer_entries e
           WHERE e.company = ? AND e.sales_date = ?
           ORDER BY e.sort_order, e.id""",
        (company, sales_date),
    ).fetchall()
    return [_room_transfer_entry_to_dict(r) for r in rows]


def load_pending_room_transfer_entries(conn, company, location=None):
    return load_room_transfer_entries_by_status(conn, company, "unpaid", location)


def _normalize_room_transfer_filter_status(status):
    value = (status or "unpaid").strip().lower()
    if value in {"paid", "unpaid", "all"}:
        return value
    return "unpaid"


def load_room_transfer_entries_by_status(
    conn,
    company,
    status="all",
    location=None,
    date_from=None,
    date_to=None,
    allowed_locations=None,
):
    params = [company]
    status_clause = ""
    normalized = _normalize_room_transfer_filter_status(status)
    if normalized == "paid":
        status_clause = " AND e.payment_status = 'paid'"
    elif normalized == "unpaid":
        status_clause = " AND e.payment_status = 'unpaid'"
    location_clause = ""
    if location and location != ROOM_TRANSFER_FILTER_ALL:
        location_clause = " AND e.location = ?"
        params.append(location)
    elif allowed_locations:
        scoped = [loc for loc in allowed_locations if loc and loc != ROOM_TRANSFER_FILTER_ALL]
        if scoped:
            placeholders = ",".join("?" for _ in scoped)
            location_clause = f" AND e.location IN ({placeholders})"
            params.extend(scoped)
    date_clause = ""
    if date_from:
        date_clause += " AND e.sales_date >= ?"
        params.append(date_from.isoformat() if hasattr(date_from, "isoformat") else str(date_from))
    if date_to:
        date_clause += " AND e.sales_date <= ?"
        params.append(date_to.isoformat() if hasattr(date_to, "isoformat") else str(date_to))
    rows = conn.execute(
        f"""SELECT e.id, e.sales_date, e.location, e.invoice_number, e.outlet_name, e.table_room, e.guest_name,
                  e.ledger_detail, e.amount, e.payment_status, e.sort_order, e.source_row,
                  COALESCE((
                      SELECT SUM(a.amount) FROM room_transfer_payment_allocations a
                      WHERE a.room_transfer_entry_id = e.id
                  ), 0) AS paid_amount
           FROM room_transfer_entries e
           WHERE e.company = ?{status_clause}{location_clause}{date_clause}
           ORDER BY e.sales_date DESC, e.location, e.sort_order, e.id""",
        params,
    ).fetchall()
    entries = [_room_transfer_entry_to_dict(r) for r in rows]
    if normalized == "paid":
        _attach_room_transfer_payment_modes(conn, entries)
    return entries


def rollup_room_transfer_entries(entries):
    rollup = {
        "total_amount": 0.0,
        "paid_amount": 0.0,
        "unpaid_amount": 0.0,
        "total_count": 0,
        "paid_count": 0,
        "unpaid_count": 0,
    }
    for entry in entries or []:
        amount = parse_money(entry.get("amount"))
        balance = parse_money(entry.get("balance") if entry.get("balance") is not None else amount)
        rollup["total_amount"] = round_half_up(rollup["total_amount"] + amount, 2)
        rollup["total_count"] += 1
        if entry.get("payment_status") == "paid":
            rollup["paid_amount"] = round_half_up(rollup["paid_amount"] + amount, 2)
            rollup["paid_count"] += 1
        else:
            rollup["unpaid_amount"] = round_half_up(rollup["unpaid_amount"] + balance, 2)
            rollup["unpaid_count"] += 1
    return rollup


def sync_room_transfer_entries(conn, company, sales_date, lines, locations=None):
    """Replace room_transfer_entries for the given locations/date.

    When ``locations`` is omitted, it is derived from the line locations so a
    Bar/Restaurant Collections sync cannot wipe Hotel credit rows (and vice versa).
    """
    lines = list(lines or [])
    if locations is None:
        locations = sorted(
            {
                str(line.get("location") or "").strip()
                for line in lines
                if str(line.get("location") or "").strip()
            }
        )
    else:
        locations = [str(loc).strip() for loc in locations if str(loc).strip()]
    if not locations:
        return

    loc_placeholders = ",".join("?" for _ in locations)
    existing_ids = [
        row["id"]
        for row in conn.execute(
            f"""SELECT id FROM room_transfer_entries
                WHERE company = ? AND sales_date = ? AND location IN ({loc_placeholders})""",
            (company, sales_date, *locations),
        ).fetchall()
    ]
    if existing_ids:
        placeholders = ",".join("?" for _ in existing_ids)
        payment_ids = [
            row["room_transfer_payment_id"]
            for row in conn.execute(
                f"""SELECT DISTINCT room_transfer_payment_id
                    FROM room_transfer_payment_allocations
                    WHERE room_transfer_entry_id IN ({placeholders})""",
                existing_ids,
            ).fetchall()
        ]
        conn.execute(
            f"DELETE FROM room_transfer_payment_allocations WHERE room_transfer_entry_id IN ({placeholders})",
            existing_ids,
        )
        for payment_id in payment_ids:
            remaining = conn.execute(
                """SELECT COUNT(*) AS cnt FROM room_transfer_payment_allocations
                   WHERE room_transfer_payment_id = ?""",
                (payment_id,),
            ).fetchone()
            if remaining and int(remaining["cnt"] or 0) == 0:
                conn.execute("DELETE FROM room_transfer_payments WHERE id = ?", (payment_id,))
    conn.execute(
        f"""DELETE FROM room_transfer_entries
            WHERE company = ? AND sales_date = ? AND location IN ({loc_placeholders})""",
        (company, sales_date, *locations),
    )
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for line in lines:
        payment_status = (line.get("payment_status") or "unpaid").strip().lower()
        if payment_status not in {"paid", "unpaid"}:
            payment_status = "unpaid"
        conn.execute(
            """INSERT INTO room_transfer_entries
               (company, location, sales_date, invoice_number, outlet_name, table_room,
                guest_name, ledger_detail, amount, payment_status, sort_order, source_row,
                created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                company,
                line.get("location", ""),
                sales_date,
                line.get("invoice_number", ""),
                line.get("outlet_name", ""),
                line.get("table_room", ""),
                line.get("guest_name", ""),
                line.get("ledger_detail", ""),
                parse_money(line.get("amount")),
                payment_status,
                int(line.get("sort_order") or 0),
                line.get("source_row"),
                now,
                now,
            ),
        )


def _normalize_room_transfer_payment_method(payment_method):
    value = (payment_method or ROOM_TRANSFER_PAYMENT_CASH).strip().lower()
    if value in (ROOM_TRANSFER_PAYMENT_BANK, "bank", "bank transfer"):
        return ROOM_TRANSFER_PAYMENT_BANK
    if value == ROOM_TRANSFER_PAYMENT_UPI:
        return ROOM_TRANSFER_PAYMENT_UPI
    if value in (ROOM_TRANSFER_PAYMENT_CARD, "credit card", "debit card"):
        return ROOM_TRANSFER_PAYMENT_CARD
    if value == ROOM_TRANSFER_PAYMENT_CASH:
        return ROOM_TRANSFER_PAYMENT_CASH
    return None


def _proportion_room_transfer_allocations(allocations, split_amount):
    """Split invoice allocations across one payment-mode amount (remainder on last)."""
    total = round_half_up(sum(item["amount"] for item in allocations), 2)
    if total <= 0 or split_amount <= 0:
        return []
    split_amount = round_half_up(split_amount, 2)
    result = []
    assigned = 0.0
    last_index = len(allocations) - 1
    for index, item in enumerate(allocations):
        if index == last_index:
            portion = round_half_up(split_amount - assigned, 2)
        else:
            portion = round_half_up(split_amount * (item["amount"] / total), 2)
            assigned = round_half_up(assigned + portion, 2)
        if portion <= 0:
            continue
        result.append({
            "entry_id": item["entry_id"],
            "amount": portion,
            "entry": item["entry"],
        })
    return result


def _parse_room_transfer_payment_splits(data, allocation_total, errors):
    raw_splits = data.get("payment_splits")
    if raw_splits is None:
        method = _normalize_room_transfer_payment_method(data.get("payment_method"))
        if method is None:
            errors.append("Invalid payment mode.")
            return []
        transaction_id = str(data.get("transaction_id") or "").strip()
        raw_splits = [{
            "payment_method": method,
            "amount": allocation_total,
            "transaction_id": transaction_id,
        }]
    if not isinstance(raw_splits, list) or not raw_splits:
        errors.append("Add at least one payment mode.")
        return []

    parsed = []
    seen_methods = set()
    for raw in raw_splits:
        if not isinstance(raw, dict):
            errors.append("Invalid payment mode split.")
            continue
        method = _normalize_room_transfer_payment_method(raw.get("payment_method"))
        if method is None:
            errors.append("Invalid payment mode.")
            continue
        if method in seen_methods:
            errors.append("Each payment mode can only be used once.")
            continue
        seen_methods.add(method)
        amount = parse_money(raw.get("amount"))
        if amount <= 0:
            errors.append("Each payment mode amount must be greater than zero.")
            continue
        transaction_id = str(raw.get("transaction_id") or "").strip()
        if method in ROOM_TRANSFER_PAYMENT_METHODS_REQUIRING_TXN and not transaction_id:
            errors.append("Transaction ID is required for bank transfer.")
            continue
        if method not in ROOM_TRANSFER_PAYMENT_METHODS_REQUIRING_TXN:
            transaction_id = ""
        parsed.append({
            "payment_method": method,
            "amount": round_half_up(amount, 2),
            "transaction_id": transaction_id,
        })

    if (errors):
        return []
    if not parsed:
        errors.append("Add at least one payment mode.")
        return []

    split_total = round_half_up(sum(item["amount"] for item in parsed), 2)
    if abs(split_total - allocation_total) > 0.001:
        errors.append("Modes total must equal the payment total before saving.")
        return []
    return parsed


def _validate_room_transfer_payment_payload(conn, data):
    errors = []
    payment_date = _parse_sales_date(data.get("payment_date") or date.today().isoformat())
    notes = str(data.get("notes") or "").strip()
    company = str(data.get("company") or DEFAULT_COMPANY).strip() or DEFAULT_COMPANY

    raw_allocations = data.get("allocations") or []
    if not isinstance(raw_allocations, list) or not raw_allocations:
        errors.append("Select at least one room transfer to clear.")
        return None, errors

    parsed_allocations = []
    seen_entry_ids = set()
    for raw in raw_allocations:
        try:
            entry_id = int(raw.get("entry_id") if isinstance(raw, dict) else None)
        except (TypeError, ValueError, AttributeError):
            errors.append("Invalid room transfer selection.")
            continue
        if entry_id in seen_entry_ids:
            errors.append("Duplicate room transfer in the same clearance.")
            continue
        seen_entry_ids.add(entry_id)
        amount = parse_money(raw.get("amount") if isinstance(raw, dict) else None)
        if amount <= 0:
            errors.append("Each allocation amount must be greater than zero.")
            continue
        entry = conn.execute(
            """SELECT id, company, location, sales_date, invoice_number, guest_name,
                      amount, payment_status
               FROM room_transfer_entries WHERE id = ?""",
            (entry_id,),
        ).fetchone()
        if not entry:
            errors.append("One or more selected room transfers were not found.")
            continue
        entry = dict(entry)
        if entry.get("company") != company:
            errors.append("Selected room transfers must belong to the same company.")
            continue
        paid_total = _room_transfer_entry_paid_total(conn, entry_id)
        balance = _room_transfer_entry_balance(entry.get("amount"), paid_total)
        if balance <= 0.001:
            code = entry.get("invoice_number") or f"#{entry_id}"
            errors.append(f"{code} is already fully paid.")
            continue
        if amount > balance + 0.001:
            code = entry.get("invoice_number") or f"#{entry_id}"
            errors.append(f"Allocation for {code} exceeds outstanding balance.")
            continue
        parsed_allocations.append({
            "entry_id": entry_id,
            "amount": round_half_up(amount, 2),
            "entry": entry,
        })

    if errors:
        return None, errors
    if not parsed_allocations:
        return None, ["Select at least one room transfer to clear."]

    total_amount = round_half_up(sum(item["amount"] for item in parsed_allocations), 2)
    payment_splits = _parse_room_transfer_payment_splits(data, total_amount, errors)
    if errors:
        return None, errors

    return {
        "company": company,
        "payment_date": payment_date.isoformat(),
        "notes": notes,
        "total_amount": total_amount,
        "allocations": parsed_allocations,
        "payment_splits": payment_splits,
    }, []


def _reverse_room_transfer_entry_payments(conn, entry_ids):
    ids = []
    for raw in entry_ids or []:
        try:
            entry_id = int(raw)
        except (TypeError, ValueError):
            continue
        if entry_id and entry_id not in ids:
            ids.append(entry_id)
    if not ids:
        return []
    placeholders = ",".join("?" for _ in ids)
    payment_ids = [
        row["room_transfer_payment_id"]
        for row in conn.execute(
            f"""SELECT DISTINCT room_transfer_payment_id
                FROM room_transfer_payment_allocations
                WHERE room_transfer_entry_id IN ({placeholders})""",
            ids,
        ).fetchall()
    ]
    conn.execute(
        f"DELETE FROM room_transfer_payment_allocations WHERE room_transfer_entry_id IN ({placeholders})",
        ids,
    )
    for payment_id in payment_ids:
        remaining = conn.execute(
            """SELECT COUNT(*) AS cnt FROM room_transfer_payment_allocations
               WHERE room_transfer_payment_id = ?""",
            (payment_id,),
        ).fetchone()
        if remaining and int(remaining["cnt"] or 0) == 0:
            conn.execute("DELETE FROM room_transfer_payments WHERE id = ?", (payment_id,))
    for entry_id in ids:
        _sync_room_transfer_status_after_payment(conn, entry_id)
    return ids


def _sales_expense_total(conn, company, location, sales_date):
    row = conn.execute(
        "SELECT COALESCE(SUM(amount), 0) AS total FROM sales_update_expenses WHERE company=? AND location=? AND sales_date=?",
        (company, location, sales_date),
    ).fetchone()
    return round_half_up(row["total"] if row else 0, 2)


def _sales_tip_total(conn, company, location, sales_date):
    row = conn.execute(
        """SELECT COALESCE(SUM(amount), 0) AS total FROM sales_update_tips
           WHERE company=? AND location=? AND sales_date=?""",
        (company, location, sales_date),
    ).fetchone()
    return round_half_up(row["total"] if row else 0, 2)


def _sales_tip_entries(conn, company, location, sales_date):
    rows = conn.execute(
        """SELECT t.id, t.employee_id, t.amount, t.description, t.sales_date,
                  e.name AS employee_name, e.emp_code AS employee_code
           FROM sales_update_tips t
           LEFT JOIN employees e ON e.id = t.employee_id
           WHERE t.company=? AND t.location=? AND t.sales_date=?
           ORDER BY t.created_at, t.id""",
        (company, location, sales_date),
    ).fetchall()
    entries = []
    for row in rows:
        item = dict(row)
        item["amount"] = round_half_up(item.get("amount"), 2)
        item["employee_name"] = item.get("employee_name") or "Unknown"
        item["employee_code"] = item.get("employee_code") or ""
        entries.append(item)
    return entries


def _active_employees_for_tips(conn):
    rows = conn.execute(
        """SELECT id, emp_code, name, location
           FROM employees
           WHERE status = 'active'
           ORDER BY LOWER(name), id"""
    ).fetchall()
    return [dict(row) for row in rows]


def _sales_tip_line_count(conn, company, location, sales_date):
    row = conn.execute(
        """SELECT COUNT(*) AS cnt FROM sales_update_tips
           WHERE company=? AND location=? AND sales_date=?""",
        (company, location, sales_date),
    ).fetchone()
    return int(row["cnt"] if row else 0)


def _apply_tip_line_total(conn, company, location, sales_date, sales_entries):
    """Overwrite tips from employee tip lines when any exist for the outlet/date."""
    if location not in TIP_OUTLET_LOCATIONS:
        return sales_entries
    if _sales_tip_line_count(conn, company, location, sales_date):
        sales_entries["tips"] = _sales_tip_total(conn, company, location, sales_date)
    return sales_entries


def _load_tips_analytics_bundle(conn, company, date_from, date_to, location_filter=None):
    """Employee tip rollup by outlet for the Tips analytics page (read-only)."""
    params = [company]
    sql = """
        SELECT t.employee_id, t.location, t.amount,
               e.name AS employee_name, e.emp_code AS employee_code
        FROM sales_update_tips t
        LEFT JOIN employees e ON e.id = t.employee_id
        WHERE t.company = ?
    """
    if date_from:
        sql += " AND t.sales_date >= ?"
        params.append(date_from.isoformat() if hasattr(date_from, "isoformat") else str(date_from))
    if date_to:
        sql += " AND t.sales_date <= ?"
        params.append(date_to.isoformat() if hasattr(date_to, "isoformat") else str(date_to))
    if location_filter and location_filter in TIP_OUTLET_LOCATIONS:
        sql += " AND t.location = ?"
        params.append(location_filter)
    sql += " ORDER BY LOWER(COALESCE(e.name, '')), t.employee_id, t.location"
    rows = conn.execute(sql, params).fetchall()

    by_employee = {}
    outlet_totals = {loc: 0.0 for loc in TIP_OUTLET_LOCATIONS}
    for row in rows:
        emp_id = row["employee_id"]
        loc = row["location"]
        amount = round_half_up(row["amount"], 2)
        if emp_id not in by_employee:
            by_employee[emp_id] = {
                "employee_id": emp_id,
                "employee_name": row["employee_name"] or "Unknown",
                "employee_code": row["employee_code"] or "",
                "hotel": 0.0,
                "bar": 0.0,
                "restaurant": 0.0,
                "total": 0.0,
            }
        item = by_employee[emp_id]
        if loc == OUTLET_HOTEL:
            item["hotel"] = round_half_up(item["hotel"] + amount, 2)
        elif loc == OUTLET_BAR:
            item["bar"] = round_half_up(item["bar"] + amount, 2)
        elif loc == OUTLET_RESTAURANT:
            item["restaurant"] = round_half_up(item["restaurant"] + amount, 2)
        item["total"] = round_half_up(item["total"] + amount, 2)
        if loc in outlet_totals:
            outlet_totals[loc] = round_half_up(outlet_totals[loc] + amount, 2)

    employees = sorted(
        by_employee.values(),
        key=lambda row: (-row["total"], (row["employee_name"] or "").casefold(), row["employee_id"]),
    )
    grand_total = round_half_up(sum(outlet_totals.values()), 2)
    return {
        "employees": employees,
        "outlet_totals": outlet_totals,
        "grand_total": grand_total,
        "hotel_total": outlet_totals[OUTLET_HOTEL],
        "bar_total": outlet_totals[OUTLET_BAR],
        "restaurant_total": outlet_totals[OUTLET_RESTAURANT],
    }


def _month_tip_pool_total(conn, company, year, month):
    """Sum of tip collections for a calendar payroll month (all outlets)."""
    date_from = date(int(year), int(month), 1)
    last_day = calendar.monthrange(int(year), int(month))[1]
    date_to = date(int(year), int(month), last_day)
    row = conn.execute(
        """SELECT COALESCE(SUM(amount), 0) AS total
           FROM sales_update_tips
           WHERE company=? AND sales_date >= ? AND sales_date <= ?""",
        (company, date_from.isoformat(), date_to.isoformat()),
    ).fetchone()
    return round_half_up(row["total"] if row else 0, 2)


def _available_tip_pool_total(conn, company, year=None, month=None):
    """Tips available for incentive payout: all collections minus prior payouts.

    When year/month are provided, allocations for that payroll month are excluded
    from the deduction so the current month can be edited against Remaining.
    """
    tips_row = conn.execute(
        """SELECT COALESCE(SUM(amount), 0) AS total
           FROM sales_update_tips WHERE company=?""",
        (company,),
    ).fetchone()
    tips_total = float(tips_row["total"] if tips_row else 0)

    if year is not None and month is not None:
        paid_row = conn.execute(
            """SELECT COALESCE(SUM(amount), 0) AS total
               FROM tip_incentive_payouts
               WHERE company=? AND NOT (year=? AND month=?)""",
            (company, int(year), int(month)),
        ).fetchone()
    else:
        paid_row = conn.execute(
            """SELECT COALESCE(SUM(amount), 0) AS total
               FROM tip_incentive_payouts WHERE company=?""",
            (company,),
        ).fetchone()
    paid_total = float(paid_row["total"] if paid_row else 0)
    return round_half_up(max(0.0, tips_total - paid_total), 2)


def _reconcile_tip_incentive_after_tip_delete(conn, company, sales_date, employee_id):
    """Remove incentive payout tied to a deleted tip; keep month pool consistent.

    Clears that employee's payout for the tip's payroll month. If remaining month
    allocations still exceed the available tip pool, clears all payouts for the month.
    """
    from employee_payroll import _period_from_credit_date, _upsert_month_tip_incentive

    year, month = _period_from_credit_date(sales_date)
    if year is None or month is None or not employee_id:
        return

    _upsert_month_tip_incentive(conn, company, year, month, int(employee_id), 0)

    available = _available_tip_pool_total(conn, company, year, month)
    allocated_row = conn.execute(
        """SELECT COALESCE(SUM(amount), 0) AS total
           FROM tip_incentive_payouts
           WHERE company=? AND year=? AND month=?""",
        (company, int(year), int(month)),
    ).fetchone()
    allocated = float(allocated_row["total"] if allocated_row else 0)
    if allocated > available + 0.001:
        conn.execute(
            """DELETE FROM tip_incentive_payouts
               WHERE company=? AND year=? AND month=?""",
            (company, int(year), int(month)),
        )


def _reconcile_tip_incentive_for_dates(conn, company, employee_id, sales_dates):
    """Reconcile incentive payouts for every distinct month in sales_dates."""
    from employee_payroll import _period_from_credit_date

    seen = set()
    for sales_date in sales_dates or []:
        year, month = _period_from_credit_date(sales_date)
        if year is None or month is None:
            continue
        key = (int(year), int(month))
        if key in seen:
            continue
        seen.add(key)
        _reconcile_tip_incentive_after_tip_delete(conn, company, sales_date, employee_id)


def _load_tips_detail_entries(conn, company, date_from, date_to, location_filter=None):
    """Individual tip lines for Tips Report Excel (date / employee / outlet / amount)."""
    params = [company]
    sql = """
        SELECT t.id, t.sales_date, t.location, t.amount, t.description,
               t.employee_id, e.name AS employee_name, e.emp_code AS employee_code
        FROM sales_update_tips t
        LEFT JOIN employees e ON e.id = t.employee_id
        WHERE t.company = ?
    """
    if date_from:
        sql += " AND t.sales_date >= ?"
        params.append(date_from.isoformat() if hasattr(date_from, "isoformat") else str(date_from))
    if date_to:
        sql += " AND t.sales_date <= ?"
        params.append(date_to.isoformat() if hasattr(date_to, "isoformat") else str(date_to))
    if location_filter and location_filter in TIP_OUTLET_LOCATIONS:
        sql += " AND t.location = ?"
        params.append(location_filter)
    sql += " ORDER BY t.sales_date, LOWER(COALESCE(e.name, '')), t.employee_id, t.location, t.id"
    rows = conn.execute(sql, params).fetchall()
    return [
        {
            "id": row["id"],
            "sales_date": row["sales_date"] or "",
            "location": row["location"] or "",
            "amount": round_half_up(row["amount"], 2),
            "description": row["description"] or "",
            "employee_id": row["employee_id"],
            "employee_name": row["employee_name"] or "Unknown",
            "employee_code": row["employee_code"] or "",
        }
        for row in rows
    ]


def _next_expense_code(conn, company):
    company = (company or DEFAULT_COMPANY).strip() or DEFAULT_COMPANY
    prefix = f"{company}-EX-"
    rows = conn.execute(
        """SELECT expense_code FROM sales_update_expenses
           WHERE company = ? AND expense_code IS NOT NULL AND expense_code != ''""",
        (company,),
    ).fetchall()
    max_num = 0
    for row in rows:
        code = row["expense_code"] or ""
        if not code.startswith(prefix):
            continue
        try:
            max_num = max(max_num, int(code[len(prefix):]))
        except (TypeError, ValueError):
            continue
    return f"{prefix}{max_num + 1}"


def _sales_expense_entries(conn, company, location, sales_date):
    rows = conn.execute(
        """SELECT e.id, e.expense_code, e.description, e.amount, e.payment_type, e.transaction_id,
                  e.category, e.invoice_number, e.supplier_id, s.name AS supplier_name
           FROM sales_update_expenses e
           LEFT JOIN suppliers s ON s.id = e.supplier_id
           WHERE e.company=? AND e.location=? AND e.sales_date=?
           ORDER BY e.created_at, e.id""",
        (company, location, sales_date),
    ).fetchall()
    entries = []
    for row in rows:
        item = dict(row)
        item["category"] = _normalize_expense_category(item.get("category"))
        entries.append(item)
    return entries


def _credit_expense_paid_total(conn, expense_id):
    row = conn.execute(
        "SELECT COALESCE(SUM(amount), 0) AS total FROM credit_payment_allocations WHERE expense_id = ?",
        (expense_id,),
    ).fetchone()
    return round_half_up(row["total"] if row else 0, 2)


def _credit_expense_balance(amount, paid_total):
    return round_half_up(max(round_half_up(amount, 2) - round_half_up(paid_total, 2), 0), 2)


def _credit_settlement_status(payment_type, amount, paid_total):
    amount = round_half_up(amount, 2)
    paid_total = round_half_up(paid_total, 2)
    normalized = _normalize_expense_payment_type(payment_type)
    if normalized != EXPENSE_PAYMENT_CREDIT:
        return "cleared"
    if paid_total <= 0:
        return "outstanding"
    if paid_total + 0.001 < amount:
        return "partial"
    return "cleared"


def _expense_clearance_payment_method(conn, expense_id):
    row = conn.execute(
        """SELECT p.payment_method
           FROM credit_payment_allocations a
           JOIN credit_payments p ON p.id = a.credit_payment_id
           WHERE a.expense_id = ?
           ORDER BY p.payment_date DESC, p.id DESC, a.id DESC
           LIMIT 1""",
        (expense_id,),
    ).fetchone()
    if not row:
        return None
    return _normalize_credit_payment_method(row["payment_method"])


def _clearance_method_to_expense_payment_type(payment_method):
    method = _normalize_credit_payment_method(payment_method)
    if method == CREDIT_PAYMENT_METHOD_CARD:
        return CREDIT_PAYMENT_METHOD_CARD
    return EXPENSE_PAYMENT_CASH


def _purchase_ledger_display_payment_type(payment_type, amount, paid_total, clearance_method=None):
    normalized = _normalize_expense_payment_type(payment_type)
    if normalized == EXPENSE_PAYMENT_CREDIT and clearance_method:
        if _credit_expense_balance(amount, paid_total) <= 0:
            return _clearance_method_to_expense_payment_type(clearance_method)
    return normalized


def _sync_expense_payment_after_clearance(conn, expense_id):
    expense = conn.execute(
        "SELECT id, amount, payment_type FROM sales_update_expenses WHERE id = ?",
        (expense_id,),
    ).fetchone()
    if not expense:
        return
    expense = dict(expense)
    if _normalize_expense_payment_type(expense.get("payment_type")) != EXPENSE_PAYMENT_CREDIT:
        return
    paid_total = _credit_expense_paid_total(conn, expense_id)
    if _credit_expense_balance(expense["amount"], paid_total) > 0:
        return
    clearance_method = _expense_clearance_payment_method(conn, expense_id)
    if not clearance_method:
        return
    clearance_type = _clearance_method_to_expense_payment_type(clearance_method)
    transaction_id = ""
    if clearance_type == CREDIT_PAYMENT_METHOD_CARD:
        row = conn.execute(
            """SELECT p.transaction_id
               FROM credit_payment_allocations a
               JOIN credit_payments p ON p.id = a.credit_payment_id
               WHERE a.expense_id = ?
               ORDER BY p.payment_date DESC, p.id DESC, a.id DESC
               LIMIT 1""",
            (expense_id,),
        ).fetchone()
        transaction_id = str(row["transaction_id"] or "").strip() if row else ""
    conn.execute(
        """UPDATE sales_update_expenses
           SET payment_type = ?, transaction_id = ?
           WHERE id = ?""",
        (clearance_type, transaction_id, expense_id),
    )


def _restore_expense_credit_on_payment_delete(conn, expense_id):
    expense = conn.execute(
        "SELECT amount, payment_type FROM sales_update_expenses WHERE id = ?",
        (expense_id,),
    ).fetchone()
    if not expense:
        return
    expense = dict(expense)
    paid_total = _credit_expense_paid_total(conn, expense_id)
    balance = _credit_expense_balance(expense["amount"], paid_total)
    current = _normalize_expense_payment_type(expense.get("payment_type"))
    if balance > 0 and current != EXPENSE_PAYMENT_CREDIT:
        conn.execute(
            """UPDATE sales_update_expenses
               SET payment_type = ?, transaction_id = ''
               WHERE id = ?""",
            (EXPENSE_PAYMENT_CREDIT, expense_id),
        )


CREDIT_SETTLEMENT_STATUS_LABELS = {
    "outstanding": "Outstanding",
    "partial": "Partial",
    "cleared": "Cleared",
}

CREDIT_PAYMENT_METHOD_CASH = EXPENSE_PAYMENT_CASH
CREDIT_PAYMENT_METHOD_CARD = "card"
CREDIT_PAYMENT_METHODS = _sorted_label_choices((
    (CREDIT_PAYMENT_METHOD_CASH, "Cash"),
    (CREDIT_PAYMENT_METHOD_CARD, "Bank Transfer"),
))
CREDIT_PAYMENT_METHOD_LABELS = dict(CREDIT_PAYMENT_METHODS)

ROOM_TRANSFER_PAYMENT_CASH = "cash"
ROOM_TRANSFER_PAYMENT_BANK = EXPENSE_PAYMENT_BANK
ROOM_TRANSFER_PAYMENT_UPI = "upi"
ROOM_TRANSFER_PAYMENT_CARD = "card"
# Keep user-facing order (not alpha-sorted).
ROOM_TRANSFER_PAYMENT_METHODS = (
    (ROOM_TRANSFER_PAYMENT_BANK, "Bank Transfer"),
    (ROOM_TRANSFER_PAYMENT_CASH, "Cash"),
    (ROOM_TRANSFER_PAYMENT_UPI, "UPI"),
    (ROOM_TRANSFER_PAYMENT_CARD, "Card"),
)
ROOM_TRANSFER_PAYMENT_METHOD_LABELS = dict(ROOM_TRANSFER_PAYMENT_METHODS)
ROOM_TRANSFER_PAYMENT_METHODS_REQUIRING_TXN = frozenset({ROOM_TRANSFER_PAYMENT_BANK})
PURCHASE_LEDGER_PAYMENT_LABELS = {
    **EXPENSE_PAYMENT_LABELS,
    CREDIT_PAYMENT_METHOD_CARD: "Bank Transfer",
}
CREDIT_PAYMENT_VIEW_OUTSTANDING = "outstanding"
CREDIT_PAYMENT_VIEW_HISTORY = "history"
CREDIT_SETTLEMENT_MODE_CREDIT_PAYMENT = "credit_payment"
CREDIT_SETTLEMENT_MODE_PURCHASE_VERIFICATION = "purchase_verification"
CREDIT_PAYMENT_VIEWS = _sorted_label_choices((
    (CREDIT_PAYMENT_VIEW_OUTSTANDING, "Outstanding Credit"),
    (CREDIT_PAYMENT_VIEW_HISTORY, "Payment History"),
))
PURCHASE_VERIFICATION_VIEWS = _sorted_label_choices((
    (CREDIT_PAYMENT_VIEW_OUTSTANDING, "Pending Verification"),
    (CREDIT_PAYMENT_VIEW_HISTORY, "Verified Purchase"),
))
CREDIT_SETTLEMENT_PAGE_MODES = {
    CREDIT_SETTLEMENT_MODE_CREDIT_PAYMENT: {
        "page_title": "Credit Payment",
        "page_subtitle": "Verified purchase update here for credit payment",
        "filter_aria_label": "Credit payment filters",
        "view_aria_label": "Credit payment views",
        "nav_accounts_view": "credit_payment",
        "route_endpoint": "credit_payment",
        "views": CREDIT_PAYMENT_VIEWS,
        "outstanding_summary_label": "Outstanding balance",
        "outstanding_panel_title": "Outstanding Credit",
        "outstanding_panel_aria": "Outstanding credit expenses",
        "outstanding_table_aria": "Outstanding credit expenses",
        "outstanding_empty": "No verified outstanding credit expenses found for the selected filters.",
        "history_summary_label": "Payments cleared",
        "history_summary_unit": "clearance",
        "history_panel_title": "Payment History",
        "history_panel_aria": "Credit payment history",
        "history_table_aria": "Credit payment history",
        "history_date_column": "Payment date",
        "history_empty": "No credit payments found for the selected filters.",
        "history_revert_button": "Revert",
        "history_revert_tip": "Revert to Outstanding Credit",
        "action_button": "Clear Payment",
        "row_action_button": "Pay",
        "select_modal_title": "Select Credit Items",
        "select_modal_copy": "Choose outstanding credit expenses to combine into payment clearances. Mixed suppliers are recorded as separate payments.",
        "select_table_aria": "Select credit line items",
        "select_continue": "Clear Payment",
        "clearance_modal_title": "Payment Details",
        "clearance_date_label": "Payment date *",
        "clearance_mode_label": "Payment mode *",
        "show_payment_mode": True,
        "show_verification_account": False,
        "show_history_expense_ids": False,
        "clearance_submit": "Record Payment",
        "clearance_total_label": "Payment total",
        "detail_modal_title": "Payment Detail",
        "detail_date_label": "Payment date",
        "pay_now_column": "Pay now",
        "select_error_none": "Select at least one credit expense.",
        "submit_error_record": "Unable to record payment.",
        "submit_error_network": "Network error while recording payment.",
        "delete_confirm": "Revert this payment to Outstanding Credit?",
        "delete_error": "Unable to revert payment.",
        "delete_error_network": "Network error while reverting payment.",
        "detail_error_load": "Unable to load payment.",
        "detail_error_network": "Network error while loading payment.",
    },
    CREDIT_SETTLEMENT_MODE_PURCHASE_VERIFICATION: {
        "page_title": "Purchase Verification",
        "page_subtitle": "Verify hotel purchases by combining expenses into a single supplier verification.",
        "filter_aria_label": "Purchase verification filters",
        "view_aria_label": "Purchase verification views",
        "nav_accounts_view": "purchase_verification",
        "route_endpoint": "purchase_verification",
        "views": PURCHASE_VERIFICATION_VIEWS,
        "outstanding_summary_label": "Pending balance",
        "outstanding_panel_title": "Pending Verification",
        "outstanding_panel_aria": "Purchases pending verification",
        "outstanding_table_aria": "Purchases pending verification",
        "outstanding_empty": "No purchases pending verification found for the selected filters.",
        "history_summary_label": "Purchases verified",
        "history_summary_unit": "verification",
        "history_panel_title": "Verified Purchase",
        "history_panel_aria": "Verified purchase history",
        "history_table_aria": "Verified purchase history",
        "history_date_column": "Verification date",
        "history_empty": "No verified purchases found for the selected filters.",
        "history_revert_button": "Revert",
        "history_revert_tip": "Revert to Pending Verification",
        "action_button": "Verify",
        "row_action_button": "Approve",
        "select_modal_title": "Select Items to Verify",
        "select_modal_copy": "Choose pending purchases to verify. Mixed suppliers are recorded as separate verifications.",
        "select_table_aria": "Select purchases to verify",
        "select_continue": "Verify",
        "clearance_modal_title": "Verification Details",
        "clearance_date_label": "Verification date *",
        "clearance_mode_label": "Verification mode *",
        "show_payment_mode": False,
        "show_verification_account": True,
        "show_history_expense_ids": True,
        "clearance_account_label": "Account",
        "clearance_submit": "Record Verification",
        "clearance_total_label": "Verification total",
        "detail_modal_title": "Verification Detail",
        "detail_date_label": "Verification date",
        "pay_now_column": "Verify now",
        "select_error_none": "Select at least one purchase to verify.",
        "submit_error_record": "Unable to record verification.",
        "submit_error_network": "Network error while recording verification.",
        "delete_confirm": "Revert this verification to Pending Verification?",
        "delete_error": "Unable to revert verification.",
        "delete_error_network": "Network error while reverting verification.",
        "detail_error_load": "Unable to load verification.",
        "detail_error_network": "Network error while loading verification.",
    },
}


def _credit_settlement_page_mode(value):
    if value == CREDIT_SETTLEMENT_MODE_PURCHASE_VERIFICATION:
        return CREDIT_SETTLEMENT_MODE_PURCHASE_VERIFICATION
    return CREDIT_SETTLEMENT_MODE_CREDIT_PAYMENT


def _render_credit_settlement_page(mode):
    labels = CREDIT_SETTLEMENT_PAGE_MODES[_credit_settlement_page_mode(mode)]
    today = date.today()
    selected_view = _normalize_credit_payment_view(request.args.get("view"))
    date_from, date_to, date_filter_active = _resolve_optional_filter_date_range(
        request.args, "date_from", "date_to", default_fy=True
    )
    payment_date_from, payment_date_to, payment_date_filter_active = _resolve_optional_filter_date_range(
        request.args, "payment_date_from", "payment_date_to"
    )

    selected_supplier, supplier_id = _parse_purchase_ledger_supplier(
        request.args.get("supplier")
    )

    conn = get_db()
    try:
        suppliers = _all_suppliers(conn)
        supplier_lookup = {str(s["id"]): s for s in suppliers}
        if selected_supplier != PURCHASE_LEDGER_FILTER_ALL and selected_supplier not in supplier_lookup:
            selected_supplier = PURCHASE_LEDGER_FILTER_ALL
            supplier_id = None
        page_mode = _credit_settlement_page_mode(mode)
        if page_mode == CREDIT_SETTLEMENT_MODE_PURCHASE_VERIFICATION:
            outstanding_entries = _pending_purchase_verifications(
                conn, date_from, date_to, supplier_id=supplier_id
            )
            payment_entries = _purchase_verification_entries(
                conn,
                verification_date_from=payment_date_from,
                verification_date_to=payment_date_to,
                supplier_id=supplier_id,
            )
            create_url = url_for("create_purchase_verification")
            delete_url = url_for("delete_purchase_verification")
            detail_url_template = url_for("purchase_verification_detail", verification_id=0)
        else:
            outstanding_entries = _outstanding_credit_expenses(
                conn, date_from, date_to, supplier_id=supplier_id
            )
            payment_entries = _credit_payment_entries(
                conn,
                payment_date_from=payment_date_from,
                payment_date_to=payment_date_to,
                supplier_id=supplier_id,
            )
            create_url = url_for("create_credit_payment")
            delete_url = url_for("delete_credit_payment")
            detail_url_template = url_for("credit_payment_detail", payment_id=0)
    finally:
        conn.close()

    outstanding_total = round_half_up(
        sum(entry["balance"] for entry in outstanding_entries), 2
    )
    category_totals = {}
    for entry in outstanding_entries:
        key = _normalize_expense_category(entry.get("category")) or "other"
        bucket = category_totals.setdefault(key, {"key": key, "total": 0.0, "count": 0})
        bucket["total"] = round_half_up(bucket["total"] + entry.get("balance", 0), 2)
        bucket["count"] += 1
    for bucket in category_totals.values():
        bucket["label"] = EXPENSE_CATEGORY_LABELS.get(bucket["key"], bucket["key"].replace("_", " ").title())
    meat_bucket = category_totals.get("meat") or {"key": "meat", "label": "Meat", "total": 0.0, "count": 0}
    meat_total = round_half_up(meat_bucket["total"], 2)
    meat_count = int(meat_bucket["count"])
    top_category_kpis = sorted(
        (bucket for key, bucket in category_totals.items() if key != "meat"),
        key=lambda item: (-item["total"], item["label"].lower()),
    )[:2]
    while len(top_category_kpis) < 2:
        top_category_kpis.append(
            {
                "key": "",
                "label": "—",
                "total": 0.0,
                "count": 0,
            }
        )
    payment_total = round_half_up(
        sum(entry["total_amount"] for entry in payment_entries), 2
    )
    selected_supplier_label = "All suppliers"
    if selected_supplier != PURCHASE_LEDGER_FILTER_ALL:
        match = supplier_lookup.get(selected_supplier)
        if match:
            selected_supplier_label = match["name"]

    route_endpoint = labels["route_endpoint"]
    filter_date_from = date_from.isoformat() if date_filter_active else ""
    filter_date_to = date_to.isoformat() if date_filter_active else ""
    filter_payment_date_from = (
        payment_date_from.isoformat() if payment_date_filter_active else ""
    )
    filter_payment_date_to = (
        payment_date_to.isoformat() if payment_date_filter_active else ""
    )
    active_date_filter = (
        date_filter_active
        if selected_view == CREDIT_PAYMENT_VIEW_OUTSTANDING
        else payment_date_filter_active
    )
    tab_query = {"supplier": selected_supplier}
    if date_filter_active:
        tab_query["date_from"] = filter_date_from
        tab_query["date_to"] = filter_date_to
    if payment_date_filter_active:
        tab_query["payment_date_from"] = filter_payment_date_from
        tab_query["payment_date_to"] = filter_payment_date_to

    credit_report_kwargs = {"supplier": selected_supplier}
    if date_filter_active:
        credit_report_kwargs["date_from"] = filter_date_from
        credit_report_kwargs["date_to"] = filter_date_to

    purchase_report_kwargs = {"view": selected_view, "supplier": selected_supplier}
    if date_filter_active:
        purchase_report_kwargs["date_from"] = filter_date_from
        purchase_report_kwargs["date_to"] = filter_date_to
    if payment_date_filter_active:
        purchase_report_kwargs["payment_date_from"] = filter_payment_date_from
        purchase_report_kwargs["payment_date_to"] = filter_payment_date_to

    actor = get_current_user()
    # Clear Payment / Verify / Approve / Revert require Approval module; Accounts alone is view-only.
    can_mutate_settlement = user_can_approve_transactions(actor)

    return render_template(
        "credit_settlement_page.html",
        settlement_labels=labels,
        settlement_route_endpoint=route_endpoint,
        page_title=labels["page_title"],
        page_subtitle=labels["page_subtitle"],
        filter_form_action=url_for(route_endpoint),
        selected_view=selected_view,
        credit_payment_views=labels["views"],
        date_from=filter_date_from,
        date_to=filter_date_to,
        payment_date_from=filter_payment_date_from,
        payment_date_to=filter_payment_date_to,
        date_filter_active=date_filter_active,
        payment_date_filter_active=payment_date_filter_active,
        active_date_filter=active_date_filter,
        settlement_tab_query=tab_query,
        selected_supplier=selected_supplier,
        selected_supplier_label=selected_supplier_label,
        suppliers=suppliers,
        outstanding_entries=outstanding_entries,
        outstanding_total=outstanding_total,
        meat_total=meat_total,
        meat_count=meat_count,
        top_category_kpis=top_category_kpis,
        payment_entries=payment_entries,
        payment_total=payment_total,
        credit_payment_methods=CREDIT_PAYMENT_METHODS,
        credit_payment_method_labels=CREDIT_PAYMENT_METHOD_LABELS,
        expense_category_labels=EXPENSE_CATEGORY_LABELS,
        credit_settlement_status_labels=CREDIT_SETTLEMENT_STATUS_LABELS,
        create_credit_payment_url=create_url,
        delete_credit_payment_url=delete_url,
        settlement_detail_url_template=detail_url_template,
        can_mutate_settlement=can_mutate_settlement,
        credit_payment_report_url=(
            url_for("export_credit_payment_report", **credit_report_kwargs)
            if page_mode == CREDIT_SETTLEMENT_MODE_CREDIT_PAYMENT
            else None
        ),
        purchase_verification_report_url=(
            url_for("export_purchase_verification_report", **purchase_report_kwargs)
            if page_mode == CREDIT_SETTLEMENT_MODE_PURCHASE_VERIFICATION
            else None
        ),
        today_iso=today.isoformat(),
        de_nav_section="accounts",
        de_nav_accounts_view=labels["nav_accounts_view"],
    )


def _resolve_optional_filter_date_range(args, from_key, to_key, *, default_fy=False):
    """Return (date_from, date_to, active).

    Missing both keys:
    - default_fy=False => no date filter (None, None, False)
    - default_fy=True  => current Indian FY start → today (active)
    """
    today = date.today()
    raw_from = (args.get(from_key) or "").strip()
    raw_to = (args.get(to_key) or "").strip()
    if not raw_from and not raw_to:
        if default_fy:
            fy_start, ref = indian_fiscal_year_bounds(today)
            return fy_start, ref, True
        return None, None, False
    date_from = _parse_sales_date(raw_from or today.replace(day=1).isoformat())
    date_to = _parse_sales_date(raw_to or today.isoformat())
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    return date_from, date_to, True


def _normalize_credit_payment_method(payment_method):
    value = (payment_method or CREDIT_PAYMENT_METHOD_CASH).strip().lower()
    if value in (CREDIT_PAYMENT_METHOD_CARD, "card", "credit card", "debit card"):
        return CREDIT_PAYMENT_METHOD_CARD
    if value in (EXPENSE_PAYMENT_BANK, "bank", "bank transfer", "bank_transfer"):
        return CREDIT_PAYMENT_METHOD_CARD
    return CREDIT_PAYMENT_METHOD_CASH


def _normalize_credit_payment_view(value):
    raw = (value or CREDIT_PAYMENT_VIEW_OUTSTANDING).strip().lower()
    if raw == CREDIT_PAYMENT_VIEW_HISTORY:
        return CREDIT_PAYMENT_VIEW_HISTORY
    return CREDIT_PAYMENT_VIEW_OUTSTANDING


def _purchase_ledger_entries(conn, date_from, date_to, supplier_id=None, company=None, category=None, payment_type=None):
    sql = """SELECT e.id, e.expense_code, e.sales_date, e.company, e.description, e.amount, e.payment_type,
                    e.transaction_id, e.category, e.invoice_number, e.supplier_id,
                    s.name AS supplier_name, s.gst AS supplier_gst,
                    COALESCE((
                        SELECT SUM(a.amount) FROM credit_payment_allocations a WHERE a.expense_id = e.id
                    ), 0) AS paid_amount
             FROM sales_update_expenses e
             LEFT JOIN suppliers s ON s.id = e.supplier_id
             WHERE e.location = ? AND e.sales_date >= ? AND e.sales_date <= ?"""
    params = [OUTLET_HOTEL, date_from.isoformat(), date_to.isoformat()]
    if company:
        sql += " AND e.company = ?"
        params.append(company)
    if supplier_id:
        sql += " AND e.supplier_id = ?"
        params.append(supplier_id)
    if category:
        sql += " AND e.category = ?"
        params.append(category)
    if payment_type:
        sql += " AND e.payment_type = ?"
        params.append(payment_type)
    sql += " ORDER BY e.sales_date DESC, e.created_at DESC, e.id DESC"
    rows = conn.execute(sql, params).fetchall()
    entries = []
    for row in rows:
        item = dict(row)
        item["amount"] = round_half_up(item.get("amount"), 2)
        item["paid_amount"] = round_half_up(item.get("paid_amount"), 2)
        item["payment_type"] = _normalize_expense_payment_type(item.get("payment_type"))
        item["category"] = _normalize_expense_category(item.get("category"))
        item["balance"] = _credit_expense_balance(item["amount"], item["paid_amount"])
        clearance_method = None
        if item["payment_type"] == EXPENSE_PAYMENT_CREDIT:
            clearance_method = _expense_clearance_payment_method(conn, item["id"])
        item["display_payment_type"] = _purchase_ledger_display_payment_type(
            item["payment_type"], item["amount"], item["paid_amount"], clearance_method
        )
        item["settlement_status"] = _credit_settlement_status(
            item["payment_type"], item["amount"], item["paid_amount"]
        )
        entries.append(item)
    return entries


def _outstanding_credit_expenses(conn, date_from=None, date_to=None, supplier_id=None, company=None):
    sql = """SELECT e.id, e.expense_code, e.sales_date, e.company, e.description, e.amount, e.payment_type,
                    e.category, e.supplier_id,
                    s.name AS supplier_name, s.gst AS supplier_gst,
                    COALESCE((
                        SELECT SUM(a.amount) FROM credit_payment_allocations a WHERE a.expense_id = e.id
                    ), 0) AS paid_amount,
                    COALESCE((
                        SELECT SUM(a.amount) FROM purchase_verification_allocations a WHERE a.expense_id = e.id
                    ), 0) AS verified_amount
             FROM sales_update_expenses e
             LEFT JOIN suppliers s ON s.id = e.supplier_id
             WHERE e.location = ? AND e.payment_type = ?"""
    params = [OUTLET_HOTEL, EXPENSE_PAYMENT_CREDIT]
    if date_from:
        sql += " AND e.sales_date >= ?"
        params.append(date_from.isoformat() if hasattr(date_from, "isoformat") else str(date_from))
    if date_to:
        sql += " AND e.sales_date <= ?"
        params.append(date_to.isoformat() if hasattr(date_to, "isoformat") else str(date_to))
    if company:
        sql += " AND e.company = ?"
        params.append(company)
    if supplier_id:
        sql += " AND e.supplier_id = ?"
        params.append(supplier_id)
    sql += " ORDER BY e.sales_date DESC, e.created_at DESC, e.id DESC"
    rows = conn.execute(sql, params).fetchall()
    entries = []
    for row in rows:
        item = dict(row)
        item["amount"] = round_half_up(item.get("amount"), 2)
        item["paid_amount"] = round_half_up(item.get("paid_amount"), 2)
        item["verified_amount"] = round_half_up(item.get("verified_amount"), 2)
        item["payment_type"] = EXPENSE_PAYMENT_CREDIT
        item["category"] = _normalize_expense_category(item.get("category"))
        # Credit Payment is step 3: only fully verified purchases are payable.
        if _purchase_verification_balance(item["amount"], item["verified_amount"]) > 0.001:
            continue
        item["balance"] = _credit_expense_balance(item["amount"], item["paid_amount"])
        item["settlement_status"] = _credit_settlement_status(
            EXPENSE_PAYMENT_CREDIT, item["amount"], item["paid_amount"]
        )
        if item["balance"] <= 0:
            continue
        entries.append(item)
    return entries


def _purchase_verification_verified_total(conn, expense_id):
    row = conn.execute(
        "SELECT COALESCE(SUM(amount), 0) AS total FROM purchase_verification_allocations WHERE expense_id = ?",
        (expense_id,),
    ).fetchone()
    return round_half_up(row["total"] if row else 0, 2)


def _verification_user_account(user):
    if not user:
        return ""
    return (user.get("username") or user.get("full_name") or "").strip()


def _auto_verify_expense(conn, *, expense_id, supplier_id, amount, company=None, user=None, notes=""):
    """Mark a Hotel expense fully verified so it can go straight to Credit Payment.

    Inserts a purchase_verifications header plus one full allocation.
    Caller owns commit/rollback.
    """
    try:
        expense_id = int(expense_id)
        supplier_id = int(supplier_id)
    except (TypeError, ValueError):
        return None, "Invalid expense or supplier for verification."
    total = round_half_up(amount, 2)
    if total <= 0:
        return None, "Verification amount must be greater than zero."
    company = (company or DEFAULT_COMPANY).strip() or DEFAULT_COMPANY
    account = _verification_user_account(user)
    if not account:
        return None, "You must be logged in to record a verification."
    already = _purchase_verification_verified_total(conn, expense_id)
    if already + 0.001 >= total:
        return {"verification_id": None, "already_verified": True}, None
    verification_date = date.today().isoformat()
    cursor = conn.execute(
        """INSERT INTO purchase_verifications
           (company, supplier_id, verification_date, verification_method, verification_account,
            transaction_id, total_amount, notes)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            company,
            supplier_id,
            verification_date,
            CREDIT_PAYMENT_METHOD_CASH,
            account,
            "",
            total,
            (notes or "").strip(),
        ),
    )
    verification_id = cursor.lastrowid
    conn.execute(
        """INSERT INTO purchase_verification_allocations
           (purchase_verification_id, expense_id, amount)
           VALUES (?, ?, ?)""",
        (verification_id, expense_id, total),
    )
    return {"verification_id": verification_id, "already_verified": False}, None


def _purchase_verification_balance(amount, verified_total):
    return _credit_expense_balance(amount, verified_total)


def _pending_purchase_verifications(conn, date_from=None, date_to=None, supplier_id=None, company=None):
    sql = """SELECT e.id, e.expense_code, e.sales_date, e.company, e.description, e.amount, e.payment_type,
                    e.category, e.supplier_id,
                    s.name AS supplier_name, s.gst AS supplier_gst,
                    COALESCE((
                        SELECT SUM(a.amount) FROM purchase_verification_allocations a WHERE a.expense_id = e.id
                    ), 0) AS paid_amount
             FROM sales_update_expenses e
             LEFT JOIN suppliers s ON s.id = e.supplier_id
             WHERE e.location = ?"""
    params = [OUTLET_HOTEL]
    if date_from:
        sql += " AND e.sales_date >= ?"
        params.append(date_from.isoformat() if hasattr(date_from, "isoformat") else str(date_from))
    if date_to:
        sql += " AND e.sales_date <= ?"
        params.append(date_to.isoformat() if hasattr(date_to, "isoformat") else str(date_to))
    if company:
        sql += " AND e.company = ?"
        params.append(company)
    if supplier_id:
        sql += " AND e.supplier_id = ?"
        params.append(supplier_id)
    sql += " ORDER BY e.sales_date DESC, e.created_at DESC, e.id DESC"
    rows = conn.execute(sql, params).fetchall()
    entries = []
    for row in rows:
        item = dict(row)
        item["amount"] = round_half_up(item.get("amount"), 2)
        item["paid_amount"] = round_half_up(item.get("paid_amount"), 2)
        item["payment_type"] = _normalize_expense_payment_type(item.get("payment_type"))
        item["category"] = _normalize_expense_category(item.get("category"))
        item["balance"] = _purchase_verification_balance(item["amount"], item["paid_amount"])
        if item["balance"] <= 0:
            continue
        entries.append(item)
    return entries


def _purchase_verification_entries(conn, verification_date_from=None, verification_date_to=None, supplier_id=None, company=None):
    sql = """SELECT v.id, v.company, v.supplier_id, v.verification_date AS payment_date,
                    v.verification_method AS payment_method, v.transaction_id,
                    v.verification_account, v.total_amount, v.notes, v.created_at,
                    s.name AS supplier_name, s.gst AS supplier_gst,
                    (
                        SELECT COUNT(*) FROM purchase_verification_allocations a
                        WHERE a.purchase_verification_id = v.id
                    ) AS allocation_count,
                    (
                        SELECT GROUP_CONCAT(
                            COALESCE(NULLIF(TRIM(e.expense_code), ''), '#' || e.id),
                            ', '
                        )
                        FROM purchase_verification_allocations a
                        LEFT JOIN sales_update_expenses e ON e.id = a.expense_id
                        WHERE a.purchase_verification_id = v.id
                    ) AS expense_codes
             FROM purchase_verifications v
             LEFT JOIN suppliers s ON s.id = v.supplier_id
             WHERE 1 = 1"""
    params = []
    if company:
        sql += " AND v.company = ?"
        params.append(company)
    if supplier_id:
        sql += " AND v.supplier_id = ?"
        params.append(supplier_id)
    if verification_date_from:
        sql += " AND v.verification_date >= ?"
        params.append(
            verification_date_from.isoformat()
            if hasattr(verification_date_from, "isoformat")
            else verification_date_from
        )
    if verification_date_to:
        sql += " AND v.verification_date <= ?"
        params.append(
            verification_date_to.isoformat()
            if hasattr(verification_date_to, "isoformat")
            else verification_date_to
        )
    sql += " ORDER BY v.verification_date DESC, v.created_at DESC, v.id DESC"
    rows = conn.execute(sql, params).fetchall()
    entries = []
    for row in rows:
        item = dict(row)
        item["total_amount"] = round_half_up(item.get("total_amount"), 2)
        item["payment_method"] = _normalize_credit_payment_method(item.get("payment_method"))
        item["verification_account"] = str(item.get("verification_account") or "").strip()
        item["allocation_count"] = int(item.get("allocation_count") or 0)
        item["expense_codes"] = str(item.get("expense_codes") or "").strip()
        entries.append(item)
    return entries


def _purchase_verification_detail(conn, verification_id, company=None):
    sql = """SELECT v.id, v.company, v.supplier_id, v.verification_date AS payment_date,
                    v.verification_method AS payment_method, v.transaction_id,
                    v.verification_account, v.total_amount, v.notes, v.created_at,
                    s.name AS supplier_name, s.gst AS supplier_gst
             FROM purchase_verifications v
             LEFT JOIN suppliers s ON s.id = v.supplier_id
             WHERE v.id = ?"""
    params = [verification_id]
    if company:
        sql += " AND v.company = ?"
        params.append(company)
    row = conn.execute(sql, params).fetchone()
    if not row:
        return None
    verification = dict(row)
    verification["total_amount"] = round_half_up(verification.get("total_amount"), 2)
    verification["payment_method"] = _normalize_credit_payment_method(verification.get("payment_method"))
    verification["verification_account"] = str(verification.get("verification_account") or "").strip()
    alloc_rows = conn.execute(
        """SELECT a.id, a.expense_id, a.amount,
                  e.expense_code, e.sales_date, e.description, e.amount AS expense_amount, e.category
           FROM purchase_verification_allocations a
           LEFT JOIN sales_update_expenses e ON e.id = a.expense_id
           WHERE a.purchase_verification_id = ?
           ORDER BY a.id""",
        (verification_id,),
    ).fetchall()
    allocations = []
    for alloc in alloc_rows:
        item = dict(alloc)
        item["amount"] = round_half_up(item.get("amount"), 2)
        item["expense_amount"] = round_half_up(item.get("expense_amount"), 2)
        item["category"] = _normalize_expense_category(item.get("category"))
        allocations.append(item)
    verification["allocations"] = allocations
    return verification


def _validate_purchase_verification_payload(conn, data, user=None):
    errors = []
    try:
        supplier_id = int(data.get("supplier_id"))
    except (TypeError, ValueError):
        supplier_id = None
    if not supplier_id:
        errors.append("Supplier is required.")

    verification_date = _parse_sales_date(data.get("payment_date") or date.today().isoformat())
    verification_method = CREDIT_PAYMENT_METHOD_CASH
    transaction_id = ""
    verification_account = _verification_user_account(user)
    notes = str(data.get("notes") or "").strip()
    company = str(data.get("company") or DEFAULT_COMPANY).strip() or DEFAULT_COMPANY

    if not verification_account:
        errors.append("You must be logged in to record a verification.")

    raw_allocations = data.get("allocations") or []
    if not isinstance(raw_allocations, list) or not raw_allocations:
        errors.append("Select at least one purchase to verify.")
        return None, errors

    if supplier_id:
        supplier = _get_supplier(conn, supplier_id)
        if not supplier:
            errors.append("Selected supplier was not found.")

    parsed_allocations = []
    seen_expense_ids = set()
    for raw in raw_allocations:
        try:
            expense_id = int(raw.get("expense_id"))
        except (TypeError, ValueError, AttributeError):
            errors.append("Invalid expense selection.")
            continue
        if expense_id in seen_expense_ids:
            errors.append("Duplicate expense in the same verification.")
            continue
        seen_expense_ids.add(expense_id)
        amount = parse_money(raw.get("amount") if isinstance(raw, dict) else None)
        if amount <= 0:
            errors.append("Each allocation amount must be greater than zero.")
            continue
        expense = conn.execute(
            """SELECT id, company, location, payment_type, amount, supplier_id, expense_code, description
               FROM sales_update_expenses WHERE id = ?""",
            (expense_id,),
        ).fetchone()
        if not expense:
            errors.append("One or more selected expenses were not found.")
            continue
        expense = dict(expense)
        if expense.get("location") != OUTLET_HOTEL:
            errors.append("Only hotel purchases can be verified.")
            continue
        if supplier_id and int(expense.get("supplier_id") or 0) != supplier_id:
            errors.append("All selected expenses must belong to the same supplier.")
            continue
        verified_total = _purchase_verification_verified_total(conn, expense_id)
        balance = _purchase_verification_balance(expense.get("amount"), verified_total)
        if amount > balance + 0.001:
            code = expense.get("expense_code") or f"#{expense_id}"
            errors.append(f"Allocation for {code} exceeds pending verification balance.")
            continue
        parsed_allocations.append({
            "expense_id": expense_id,
            "amount": round_half_up(amount, 2),
            "expense": expense,
        })

    if errors:
        return None, errors
    if not parsed_allocations:
        return None, ["Select at least one purchase to verify."]

    total_amount = round_half_up(sum(item["amount"] for item in parsed_allocations), 2)
    return {
        "company": company,
        "supplier_id": supplier_id,
        "verification_date": verification_date.isoformat(),
        "verification_method": verification_method,
        "verification_account": verification_account,
        "transaction_id": transaction_id,
        "notes": notes,
        "total_amount": total_amount,
        "allocations": parsed_allocations,
    }, []


def _credit_payment_entries(conn, payment_date_from=None, payment_date_to=None, supplier_id=None, company=None):
    sql = """SELECT p.id, p.company, p.supplier_id, p.payment_date, p.payment_method, p.transaction_id,
                    p.total_amount, p.notes, p.created_at,
                    s.name AS supplier_name, s.gst AS supplier_gst,
                    (
                        SELECT COUNT(*) FROM credit_payment_allocations a WHERE a.credit_payment_id = p.id
                    ) AS allocation_count
             FROM credit_payments p
             LEFT JOIN suppliers s ON s.id = p.supplier_id
             WHERE 1 = 1"""
    params = []
    if company:
        sql += " AND p.company = ?"
        params.append(company)
    if supplier_id:
        sql += " AND p.supplier_id = ?"
        params.append(supplier_id)
    if payment_date_from:
        sql += " AND p.payment_date >= ?"
        params.append(
            payment_date_from.isoformat() if hasattr(payment_date_from, "isoformat") else payment_date_from
        )
    if payment_date_to:
        sql += " AND p.payment_date <= ?"
        params.append(
            payment_date_to.isoformat() if hasattr(payment_date_to, "isoformat") else payment_date_to
        )
    sql += " ORDER BY p.payment_date DESC, p.created_at DESC, p.id DESC"
    rows = conn.execute(sql, params).fetchall()
    entries = []
    for row in rows:
        item = dict(row)
        item["total_amount"] = round_half_up(item.get("total_amount"), 2)
        item["payment_method"] = _normalize_credit_payment_method(item.get("payment_method"))
        item["allocation_count"] = int(item.get("allocation_count") or 0)
        entries.append(item)
    return entries


def _credit_payment_detail(conn, payment_id, company=None):
    sql = """SELECT p.id, p.company, p.supplier_id, p.payment_date, p.payment_method, p.transaction_id,
                    p.total_amount, p.notes, p.created_at,
                    s.name AS supplier_name, s.gst AS supplier_gst
             FROM credit_payments p
             LEFT JOIN suppliers s ON s.id = p.supplier_id
             WHERE p.id = ?"""
    params = [payment_id]
    if company:
        sql += " AND p.company = ?"
        params.append(company)
    row = conn.execute(sql, params).fetchone()
    if not row:
        return None
    payment = dict(row)
    payment["total_amount"] = round_half_up(payment.get("total_amount"), 2)
    payment["payment_method"] = _normalize_credit_payment_method(payment.get("payment_method"))
    alloc_rows = conn.execute(
        """SELECT a.id, a.expense_id, a.amount,
                  e.expense_code, e.sales_date, e.description, e.amount AS expense_amount, e.category
           FROM credit_payment_allocations a
           LEFT JOIN sales_update_expenses e ON e.id = a.expense_id
           WHERE a.credit_payment_id = ?
           ORDER BY a.id""",
        (payment_id,),
    ).fetchall()
    allocations = []
    for alloc in alloc_rows:
        item = dict(alloc)
        item["amount"] = round_half_up(item.get("amount"), 2)
        item["expense_amount"] = round_half_up(item.get("expense_amount"), 2)
        item["category"] = _normalize_expense_category(item.get("category"))
        allocations.append(item)
    payment["allocations"] = allocations
    return payment


def _validate_credit_payment_payload(conn, data):
    errors = []
    try:
        supplier_id = int(data.get("supplier_id"))
    except (TypeError, ValueError):
        supplier_id = None
    if not supplier_id:
        errors.append("Supplier is required.")

    payment_date = _parse_sales_date(data.get("payment_date") or date.today().isoformat())
    payment_method = _normalize_credit_payment_method(data.get("payment_method"))
    transaction_id = str(data.get("transaction_id") or "").strip()
    notes = str(data.get("notes") or "").strip()
    company = str(data.get("company") or DEFAULT_COMPANY).strip() or DEFAULT_COMPANY

    if payment_method == CREDIT_PAYMENT_METHOD_CARD and not transaction_id:
        errors.append("Transaction ID is required for bank transfer.")
    if payment_method != CREDIT_PAYMENT_METHOD_CARD:
        transaction_id = ""

    raw_allocations = data.get("allocations") or []
    if not isinstance(raw_allocations, list) or not raw_allocations:
        errors.append("Select at least one expense to clear.")
        return None, errors

    if supplier_id:
        supplier = _get_supplier(conn, supplier_id)
        if not supplier:
            errors.append("Selected supplier was not found.")

    parsed_allocations = []
    seen_expense_ids = set()
    for raw in raw_allocations:
        try:
            expense_id = int(raw.get("expense_id"))
        except (TypeError, ValueError, AttributeError):
            errors.append("Invalid expense selection.")
            continue
        if expense_id in seen_expense_ids:
            errors.append("Duplicate expense in the same clearance.")
            continue
        seen_expense_ids.add(expense_id)
        amount = parse_money(raw.get("amount") if isinstance(raw, dict) else None)
        if amount <= 0:
            errors.append("Each allocation amount must be greater than zero.")
            continue
        expense = conn.execute(
            """SELECT id, company, location, payment_type, amount, supplier_id, expense_code, description
               FROM sales_update_expenses WHERE id = ?""",
            (expense_id,),
        ).fetchone()
        if not expense:
            errors.append("One or more selected expenses were not found.")
            continue
        expense = dict(expense)
        if expense.get("location") != OUTLET_HOTEL:
            errors.append("Only hotel credit expenses can be cleared.")
            continue
        if _normalize_expense_payment_type(expense.get("payment_type")) != EXPENSE_PAYMENT_CREDIT:
            errors.append("Only credit expenses can be cleared.")
            continue
        verified_total = _purchase_verification_verified_total(conn, expense_id)
        if _purchase_verification_balance(expense.get("amount"), verified_total) > 0.001:
            code = expense.get("expense_code") or f"#{expense_id}"
            errors.append(f"{code} must be verified in Purchase Verification before payment.")
            continue
        if supplier_id and int(expense.get("supplier_id") or 0) != supplier_id:
            errors.append("All selected expenses must belong to the same supplier.")
            continue
        paid_total = _credit_expense_paid_total(conn, expense_id)
        balance = _credit_expense_balance(expense.get("amount"), paid_total)
        if amount > balance + 0.001:
            code = expense.get("expense_code") or f"#{expense_id}"
            errors.append(f"Allocation for {code} exceeds outstanding balance.")
            continue
        parsed_allocations.append({
            "expense_id": expense_id,
            "amount": round_half_up(amount, 2),
            "expense": expense,
        })

    if errors:
        return None, errors
    if not parsed_allocations:
        return None, ["Select at least one expense to clear."]

    total_amount = round_half_up(sum(item["amount"] for item in parsed_allocations), 2)
    return {
        "company": company,
        "supplier_id": supplier_id,
        "payment_date": payment_date.isoformat(),
        "payment_method": payment_method,
        "transaction_id": transaction_id,
        "notes": notes,
        "total_amount": total_amount,
        "allocations": parsed_allocations,
    }, []


def _parse_purchase_ledger_supplier(value):
    raw = (value or PURCHASE_LEDGER_FILTER_ALL).strip()
    if not raw or raw == PURCHASE_LEDGER_FILTER_ALL:
        return PURCHASE_LEDGER_FILTER_ALL, None
    try:
        supplier_id = int(raw)
    except (TypeError, ValueError):
        return PURCHASE_LEDGER_FILTER_ALL, None
    return str(supplier_id), supplier_id if supplier_id > 0 else None


def _parse_purchase_ledger_category(value):
    raw = (value or PURCHASE_LEDGER_FILTER_ALL).strip()
    if not raw or raw == PURCHASE_LEDGER_FILTER_ALL:
        return PURCHASE_LEDGER_FILTER_ALL, None
    normalized = _normalize_expense_category(raw)
    if normalized:
        return normalized, normalized
    return PURCHASE_LEDGER_FILTER_ALL, None


def _parse_purchase_ledger_payment(value):
    raw = (value or PURCHASE_LEDGER_FILTER_ALL).strip()
    if not raw or raw == PURCHASE_LEDGER_FILTER_ALL:
        return PURCHASE_LEDGER_FILTER_ALL, None
    normalized = _normalize_expense_payment_type(raw)
    if normalized in EXPENSE_PAYMENT_LABELS:
        return normalized, normalized
    return PURCHASE_LEDGER_FILTER_ALL, None


def _normalize_expense_payment_type(payment_type):
    value = (payment_type or EXPENSE_PAYMENT_CASH).strip().lower()
    if value in (EXPENSE_PAYMENT_BANK, "bank", "bank transfer", "bank_transfer"):
        return EXPENSE_PAYMENT_BANK
    if value in (EXPENSE_PAYMENT_CREDIT, "credit", "room credit", "room_credit"):
        return EXPENSE_PAYMENT_CREDIT
    return EXPENSE_PAYMENT_CASH


def _normalize_expense_category(category):
    import re
    value = (category or "").strip().lower().replace(" ", "_").replace("-", "_")
    value = re.sub(r"_+", "_", value).strip("_")
    aliases = {
        "grocery": "grocery",
        "vegetables": "vegetables",
        "vegetable": "vegetables",
        "travel": "travel",
        "hardware": "hardware",
        "tac": "tac",
        "travel_agent_commission": "tac",
        "tac_travel_agent_commission": "tac",
        "fruits": "fruits",
        "fruit": "fruits",
        "snacks": "snacks",
        "snack": "snacks",
        "meat": "meat",
        "sea_food": "sea_food",
        "seafood": "sea_food",
        "labour": "labour",
        "labor": "labour",
        "salary": "salary",
        "salaries": "salary",
        "water_tank": "water_tank",
        "watertank": "water_tank",
        "liquor": "liquor",
        "alcohol": "liquor",
        "fuel": "fuel",
        "petrol": "fuel",
        "diesel": "fuel",
        "other": "other",
    }
    if value in aliases:
        return aliases[value]
    if value in EXPENSE_CATEGORY_LABELS:
        return value
    # Custom categories saved via Add (or freeform keys already stored on expenses).
    if value and re.fullmatch(r"[a-z][a-z0-9_]{0,79}", value):
        return value
    return ""


def _normalize_invoice_number(value):
    return (value or "").strip()


def _duplicate_expense_invoice(conn, supplier_id, invoice_number, exclude_expense_id=None):
    invoice_number = _normalize_invoice_number(invoice_number)
    if not invoice_number or not supplier_id:
        return None
    sql = """SELECT id, expense_code FROM sales_update_expenses
             WHERE supplier_id = ? AND LOWER(TRIM(invoice_number)) = LOWER(?)
               AND TRIM(invoice_number) != ''"""
    params = [supplier_id, invoice_number]
    exclude_ids = []
    if exclude_expense_id is not None:
        if isinstance(exclude_expense_id, (list, tuple, set)):
            exclude_ids = [int(x) for x in exclude_expense_id if x is not None]
        else:
            try:
                exclude_ids = [int(exclude_expense_id)]
            except (TypeError, ValueError):
                exclude_ids = []
    if exclude_ids:
        placeholders = ",".join("?" for _ in exclude_ids)
        sql += f" AND id NOT IN ({placeholders})"
        params.extend(exclude_ids)
    return conn.execute(sql, params).fetchone()


def _normalize_gst(value):
    return "".join((value or "").upper().split())


def _supplier_row_to_dict(row):
    if not row:
        return None
    return {
        "id": row["id"],
        "name": row["name"] or "",
        "gst": row["gst"] or "",
        "address": row["address"] or "",
        "phone": row["phone"] or "",
        "bank_name": row["bank_name"] or "",
        "bank_account_number": row["bank_account_number"] or "",
        "ifsc_code": row["ifsc_code"] or "",
    }


def _all_suppliers(conn):
    rows = conn.execute(
        """SELECT id, name, gst, address, phone, bank_name, bank_account_number, ifsc_code
           FROM suppliers
           ORDER BY LOWER(name), id"""
    ).fetchall()
    return [_supplier_row_to_dict(row) for row in rows]


def _get_supplier(conn, supplier_id):
    if not supplier_id:
        return None
    row = conn.execute(
        """SELECT id, name, gst, address, phone, bank_name, bank_account_number, ifsc_code
           FROM suppliers WHERE id = ?""",
        (supplier_id,),
    ).fetchone()
    return _supplier_row_to_dict(row)


def _validate_supplier(conn, name, gst, supplier_id=None):
    errors = []
    name = (name or "").strip()
    gst = _normalize_gst(gst)
    if not name:
        errors.append("Supplier name is required.")
    if gst:
        existing = conn.execute(
            "SELECT id FROM suppliers WHERE gst = ?",
            (gst,),
        ).fetchone()
        if existing and (supplier_id is None or int(existing["id"]) != int(supplier_id)):
            errors.append("A supplier with this GST number already exists.")
    return errors, name, gst


def _supplier_form_payload(source=None):
    source = source or {}
    return {
        "name": (source.get("name") or "").strip(),
        "gst": _normalize_gst(source.get("gst")),
        "address": (source.get("address") or "").strip(),
        "phone": (source.get("phone") or "").strip(),
        "bank_name": (source.get("bank_name") or "").strip(),
        "bank_account_number": (source.get("bank_account_number") or "").strip(),
        "ifsc_code": (source.get("ifsc_code") or "").strip(),
    }


def _save_supplier_record(conn, payload, supplier_id=None):
    errors, name, gst = _validate_supplier(
        conn, payload.get("name"), payload.get("gst"), supplier_id=supplier_id
    )
    if errors:
        return None, errors
    fields = _supplier_form_payload(payload)
    if supplier_id:
        conn.execute(
            f"""UPDATE suppliers
                SET name = ?, gst = ?, address = ?, phone = ?, bank_name = ?,
                    bank_account_number = ?, ifsc_code = ?, updated_at = {SQL_NOW}
                WHERE id = ?""",
            (
                fields["name"],
                gst,
                fields["address"],
                fields["phone"],
                fields["bank_name"],
                fields["bank_account_number"],
                fields["ifsc_code"],
                supplier_id,
            ),
        )
        saved_id = supplier_id
    else:
        conn.execute(
            f"""INSERT INTO suppliers
                (name, gst, address, phone, bank_name, bank_account_number, ifsc_code, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, {SQL_NOW}, {SQL_NOW})""",
            (
                fields["name"],
                gst,
                fields["address"],
                fields["phone"],
                fields["bank_name"],
                fields["bank_account_number"],
                fields["ifsc_code"],
            ),
        )
        saved_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
    return saved_id, []


def _sales_unpaid_bill_total(conn, company, location, sales_date):
    row = conn.execute(
        """SELECT COALESCE(SUM(amount), 0) AS total FROM sales_update_pending_bills
           WHERE company=? AND location=? AND recorded_sales_date=? AND status='open'""",
        (company, location, sales_date),
    ).fetchone()
    return round_half_up(row["total"] if row else 0, 2)


def _sales_bill_payment_total(conn, company, location, sales_date):
    row = conn.execute(
        "SELECT COALESCE(SUM(amount), 0) AS total FROM sales_update_bill_payments WHERE company=? AND location=? AND sales_date=?",
        (company, location, sales_date),
    ).fetchone()
    return round_half_up(row["total"] if row else 0, 2)


def _sales_unpaid_bill_entries(conn, company, location, sales_date):
    rows = conn.execute(
        """SELECT id, invoice_number, amount FROM sales_update_pending_bills
           WHERE company=? AND location=? AND recorded_sales_date=? AND status='open'
           ORDER BY created_at, id""",
        (company, location, sales_date),
    ).fetchall()
    return [dict(r) for r in rows]


def _sales_bill_payment_entries(conn, company, location, sales_date):
    rows = conn.execute(
        """SELECT bp.id, bp.pending_bill_id, bp.amount, pb.invoice_number
           FROM sales_update_bill_payments bp
           LEFT JOIN sales_update_pending_bills pb ON pb.id = bp.pending_bill_id
           WHERE bp.company=? AND bp.location=? AND bp.sales_date=?
           ORDER BY bp.created_at, bp.id""",
        (company, location, sales_date),
    ).fetchall()
    return [dict(r) for r in rows]


def _sales_open_pending_bills(conn, company, location):
    rows = conn.execute(
        """SELECT id, invoice_number, amount, recorded_sales_date
           FROM sales_update_pending_bills
           WHERE company=? AND location=? AND status='open'
           ORDER BY recorded_sales_date DESC, id DESC""",
        (company, location),
    ).fetchall()
    return [dict(r) for r in rows]


def _sales_cash_transfer_total(conn, company, location, sales_date):
    row = conn.execute(
        "SELECT COALESCE(SUM(amount), 0) AS total FROM sales_update_cash_transfers WHERE company=? AND location=? AND sales_date=?",
        (company, location, sales_date),
    ).fetchone()
    return round_half_up(row["total"] if row else 0, 2)


def _sales_cash_transfer_entries(conn, company, location, sales_date):
    rows = conn.execute(
        "SELECT id, destination, description, amount FROM sales_update_cash_transfers WHERE company=? AND location=? AND sales_date=? ORDER BY created_at, id",
        (company, location, sales_date),
    ).fetchall()
    return [dict(r) for r in rows]


def build_sales_entry_values(conn, company, location, sales_date, submitted_values=None):
    values = dict(submitted_values or {})
    for key, _label in SALES_ENTRY_FIELDS:
        values.setdefault(key, 0.0)
        values[key] = parse_money(values.get(key))
    return values


def build_hotel_sales_entry_values(submitted_values=None):
    values = dict(submitted_values or {})
    for key, _label in HOTEL_SALES_ENTRY_FIELDS:
        values.setdefault(key, 0.0)
        values[key] = parse_money(values.get(key))
    return values


def load_sales_row(company, location, sales_date):
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT * FROM sales_updates WHERE company = ? AND location = ? AND sales_date = ?",
            (company, location, sales_date),
        ).fetchone()
        if not row:
            return None
        result = dict(row)
        result["sales_entry_values"] = json.loads(result.get("sales_entry_values") or "{}")
        result["petty_cash_counts"] = json.loads(result.get("petty_cash_counts") or "{}")
        result["cash_denomination_counts"] = json.loads(result.get("cash_denomination_counts") or "{}")
        return result
    finally:
        conn.close()


def upsert_sales_row(user, company, location, sales_date, sales_entries, petty_cash_counts=None, cash_denomination_counts=None):
    petty_cash_counts = petty_cash_counts or {}
    cash_denomination_counts = cash_denomination_counts or {}
    sales_entry_total = get_sales_entry_total(sales_entries)
    petty_cash_total = get_denomination_total(petty_cash_counts)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = get_db()
    try:
        conn.execute(
            """INSERT INTO sales_updates
               (company, location, sales_date, sales_entry_values, sales_entry_total,
                petty_cash_counts, petty_cash_total, cash_denomination_counts,
                created_by_user_id, updated_by_user_id, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(company, location, sales_date)
               DO UPDATE SET
                   sales_entry_values = excluded.sales_entry_values,
                   sales_entry_total = excluded.sales_entry_total,
                   petty_cash_counts = excluded.petty_cash_counts,
                   petty_cash_total = excluded.petty_cash_total,
                   cash_denomination_counts = excluded.cash_denomination_counts,
                   updated_by_user_id = excluded.updated_by_user_id,
                   updated_at = excluded.updated_at
            """,
            (
                company,
                location,
                sales_date,
                json.dumps(sales_entries),
                sales_entry_total,
                json.dumps(petty_cash_counts),
                petty_cash_total,
                json.dumps(cash_denomination_counts),
                user["id"],
                user["id"],
                now,
                now,
            ),
        )
        conn.commit()
    finally:
        conn.close()

    return {
        "sales_entry_total": sales_entry_total,
        "petty_cash_total": petty_cash_total,
    }


def merge_import_into_sales_values(existing_values, imported_values):
    merged = dict(existing_values or {})
    for key in IMPORT_FIELD_KEYS:
        merged[key] = parse_money(imported_values.get(key))
    return merged


def _parse_sales_date(value):
    try:
        return date.fromisoformat((value or "").strip())
    except (TypeError, ValueError):
        return date.today()


def _pct_change_vs_previous(current, previous):
    try:
        cur = float(current or 0)
        prev = float(previous or 0)
    except (TypeError, ValueError):
        return None
    if prev == 0:
        if cur == 0:
            return 0.0
        return 100.0 if cur > 0 else -100.0
    return round((cur - prev) / abs(prev) * 100, 1)


def _aggregate_sales_kpis(conn, date_from, date_to, company=None, location=None, difference_mode=None):
    sql = "SELECT sales_entry_values FROM sales_updates WHERE sales_date >= ? AND sales_date <= ?"
    params = [date_from.isoformat(), date_to.isoformat()]
    if company:
        sql += " AND company = ?"
        params.append(company)
    if location:
        sql += " AND location = ?"
        params.append(location)
    rows = conn.execute(sql, params).fetchall()

    actual = digital = cash = room_credit = tips = actual_cash = difference = 0.0
    for row in rows:
        vals = json.loads(row["sales_entry_values"] or "{}")
        actual += parse_money(vals.get("total_sales"))
        digital += get_digital_transactions(vals)
        cash += parse_money(vals.get("cash"))
        room_credit += parse_money(vals.get("room_credit"))
        tips += parse_money(vals.get("tips"))
        actual_cash += parse_money(vals.get("actual_cash"))
        if difference_mode != "cash_actual":
            difference += get_difference(vals)

    if difference_mode == "cash_actual":
        difference = round_half_up(cash - actual_cash, 2)

    expense_sql = "SELECT COALESCE(SUM(amount), 0) AS total FROM sales_update_expenses WHERE sales_date >= ? AND sales_date <= ?"
    expense_params = [date_from.isoformat(), date_to.isoformat()]
    if company:
        expense_sql += " AND company = ?"
        expense_params.append(company)
    if location:
        expense_sql += " AND location = ?"
        expense_params.append(location)
    expense_row = conn.execute(expense_sql, expense_params).fetchone()
    expense = round_half_up(expense_row["total"] if expense_row else 0, 2)

    return {
        "actual_sales": round_half_up(actual, 2),
        "digital_transactions": round_half_up(digital, 2),
        "cash": round_half_up(cash, 2),
        "room_credit": round_half_up(room_credit, 2),
        "tips": round_half_up(tips, 2),
        "expense": expense,
        "difference": round_half_up(difference, 2),
    }


def _sales_report_kpi_bundle(conn, date_from, date_to, company=None, location=None, difference_mode=None):
    current = _aggregate_sales_kpis(conn, date_from, date_to, company, location, difference_mode)
    if date_from == date_to:
        prev_to = date_from - timedelta(days=1)
        prev_from = prev_to
        vs_label = "yesterday"
    else:
        span_days = (date_to - date_from).days + 1
        prev_to = date_from - timedelta(days=1)
        prev_from = prev_to - timedelta(days=span_days - 1)
        vs_label = "previous period"
    previous = _aggregate_sales_kpis(conn, prev_from, prev_to, company, location, difference_mode)
    trends = {
        key: _pct_change_vs_previous(current[key], previous[key])
        for key in ("actual_sales", "digital_transactions", "cash", "room_credit", "tips", "expense", "difference")
    }
    return {
        "current": current,
        "trends": trends,
        "vs_label": vs_label,
        "is_single_day": date_from == date_to,
    }


def _check_sales_date_lock(user, company, location, sales_date):
    today_iso = date.today().isoformat()
    if sales_date > today_iso:
        return "Cannot save future dates."
    if not user.get("is_admin") and sales_date < today_iso:
        if load_sales_row(company, location, sales_date):
            return "This date was already saved. Only administrators can change past sales entries."
    return None


def _check_payroll_month_date_lock(conn, sales_date):
    """Block tip/payroll-linked writes when the sales date falls in a locked month."""
    from employee_payroll import (
        _is_payroll_month_locked,
        _payroll_month_frozen_message,
        _period_from_credit_date,
    )
    year, month = _period_from_credit_date(sales_date)
    if year is None:
        return None
    if _is_payroll_month_locked(conn, year, month):
        return _payroll_month_frozen_message(year, month)
    return None



@app.template_filter("inr")
def inr_format(value, dec=0):
    try:
        v = float(value or 0)
        neg = v < 0
        v = abs(v)
        if dec <= 0:
            s = f"{int(round(v)):,}"
        else:
            s = f"{v:,.{dec}f}"
        parts = s.split(".")
        int_part = parts[0]
        if len(int_part) > 4:
            raw = int_part.replace(",", "")
            if len(raw) > 3:
                last3 = raw[-3:]
                rest = raw[:-3]
                groups = []
                while len(rest) > 2:
                    groups.insert(0, rest[-2:])
                    rest = rest[:-2]
                if rest:
                    groups.insert(0, rest)
                int_part = ",".join(groups) + "," + last3
        s = int_part + ("." + parts[1] if len(parts) > 1 else "")
        return ("−" if neg else "") + "₹" + s
    except (TypeError, ValueError):
        return "₹0"


@app.route("/favicon.ico")
def favicon():
    return send_from_directory(
        app.static_folder,
        "favicon-32.png",
        mimetype="image/png",
        max_age=60 * 60 * 24 * 30,
    )


@app.route("/sw.js")
def service_worker():
    """Serve the POS offline service worker at root so scope can be '/'."""
    response = send_from_directory(
        app.static_folder,
        "sw.js",
        mimetype="application/javascript",
        max_age=0,
    )
    response.headers["Service-Worker-Allowed"] = "/"
    response.headers["Cache-Control"] = "no-cache"
    return response


@app.route("/")
def index():
    if get_current_user():
        return redirect(url_for("home"))
    notice = session.pop("login_notice", "")
    return render_template(
        "index.html",
        error="",
        notice=notice,
        username="",
        show_captcha=False,
        account_locked=False,
    )


def _login_page(
    *,
    error="",
    notice="",
    username="",
    show_captcha=False,
    account_locked=False,
):
    return render_template(
        "index.html",
        error=error,
        notice=notice,
        username=username or "",
        show_captcha=bool(show_captcha),
        account_locked=bool(account_locked),
    )


def _send_unlock_email_for_user(conn, row):
    """
    Issue unlock token and attempt email delivery.
    Returns {"ok": bool, "reason": "sent"|"no_email"|"smtp_not_configured"|"send_failed"}
    """
    email = (row["email"] if row and "email" in row.keys() else "") or ""
    email = email.strip()
    if not email:
        return {"ok": False, "reason": "no_email"}
    token = auth_security.issue_unlock_token(conn, int(row["id"]))
    base = app_base_url(request)
    unlock_url = f"{base}{url_for('unlock_account', token=token)}"
    sent = send_account_unlock_email(
        to_addr=email,
        username=row["username"],
        unlock_url=unlock_url,
    )
    if sent:
        return {"ok": True, "reason": "sent"}
    if not smtp_configured():
        return {"ok": False, "reason": "smtp_not_configured"}
    return {"ok": False, "reason": "send_failed"}


def _locked_account_message(send_result=None):
    reason = (send_result or {}).get("reason")
    if reason == "sent":
        return (
            "This account is locked after too many failed sign-in attempts. "
            "Check your email for an unlock link, or ask an administrator."
        )
    if reason == "no_email":
        return (
            "This account is locked after too many failed sign-in attempts. "
            "No email is on file — ask an administrator to unlock the account."
        )
    if reason in ("smtp_not_configured", "send_failed"):
        return (
            "This account is locked after too many failed sign-in attempts. "
            "Unlock email could not be sent — ask an administrator to unlock the account "
            "(or configure SMTP_HOST in .env)."
        )
    return (
        "This account is locked after too many failed sign-in attempts. "
        "Ask an administrator to unlock the account, or use Resend unlock email."
    )


@app.route("/login", methods=["POST"])
def login():
    username = (request.form.get("username") or "").strip()
    password = request.form.get("password") or ""
    captcha_answer = request.form.get("captcha") or ""
    ip = auth_security.client_ip_from_request(request)
    auth_security.note_ip_login_attempt(ip)

    if auth_security.ip_is_throttled(ip):
        return _login_page(
            error="Too many sign-in attempts from this network. Please wait and try again.",
            username=username,
        )

    conn = get_db()
    try:
        row = conn.execute(
            "SELECT * FROM users WHERE LOWER(username) = LOWER(?) AND is_active = 1",
            (username,),
        ).fetchone()

        if row and auth_security.is_account_locked(row):
            return _login_page(
                error=_locked_account_message(),
                username=username,
                account_locked=True,
            )

        needs_captcha = auth_security.captcha_is_required(row)
        if needs_captcha:
            if not auth_security.verify_captcha_answer(session, captcha_answer):
                state = auth_security.record_failed_login(conn, int(row["id"]))
                conn.commit()
                if state["newly_locked"]:
                    send_result = _send_unlock_email_for_user(conn, row)
                    conn.commit()
                    return _login_page(
                        error=_locked_account_message(send_result),
                        username=username,
                        account_locked=True,
                    )
                return _login_page(
                    error="CAPTCHA was incorrect. Please try again.",
                    username=username,
                    show_captcha=True,
                )

        password_ok = auth_security.verify_password_for_row(row, password)
        if not row or not password_ok:
            if row:
                state = auth_security.record_failed_login(conn, int(row["id"]))
                conn.commit()
                if state["newly_locked"]:
                    send_result = _send_unlock_email_for_user(conn, row)
                    conn.commit()
                    return _login_page(
                        error=_locked_account_message(send_result),
                        username=username,
                        account_locked=True,
                    )
                return _login_page(
                    error="Invalid username or password.",
                    username=username,
                    show_captcha=state["captcha_required"],
                )
            auth_security.record_unknown_user_failure()
            return _login_page(
                error="Invalid username or password.",
                username=username,
            )

        auth_security.clear_login_failures(conn, int(row["id"]))
        auth_security.upgrade_password_hash_if_needed(
            conn,
            int(row["id"]),
            password,
            row["password_hash"],
        )
        conn.commit()
    finally:
        conn.close()

    auth_security.clear_captcha_challenge(session)
    session.clear()
    session[AUTH_USER_SESSION_KEY] = row["id"]
    if bool(row["must_change_password"]):
        return redirect(url_for("change_password"))
    return redirect(url_for("home"))


@app.route("/change-password", methods=["GET", "POST"])
def change_password():
    user = get_current_user()
    if not user:
        return redirect(url_for("index"))
    if not user.get("must_change_password"):
        return redirect(url_for("home"))

    error = ""
    if request.method == "POST":
        new_password = request.form.get("new_password") or ""
        confirm_password = request.form.get("confirm_password") or ""
        if not new_password.strip():
            error = "Enter a new password."
        elif new_password != confirm_password:
            error = "New password and confirmation do not match."
        else:
            conn = get_db()
            try:
                row = conn.execute(
                    "SELECT id, password_hash FROM users WHERE id = ?",
                    (user["id"],),
                ).fetchone()
                if not row:
                    session.pop(AUTH_USER_SESSION_KEY, None)
                    return redirect(url_for("index"))
                if auth_security.verify_password(row["password_hash"], new_password):
                    error = "Choose a password that is different from your temporary password."
                else:
                    conn.execute(
                        f"""UPDATE users
                               SET password_hash = ?,
                                   must_change_password = 0,
                                   updated_at = {SQL_NOW}
                             WHERE id = ?""",
                        (auth_security.hash_password(new_password), user["id"]),
                    )
                    conn.commit()
                    g._auth_loaded = False
                    g.current_user = None
                    return redirect(url_for("home"))
            finally:
                conn.close()

    return render_template("change_password.html", error=error)


@app.route("/login/captcha")
def login_captcha():
    answer = auth_security.generate_captcha_text()
    auth_security.store_captcha_challenge(session, answer)
    png = auth_security.render_captcha_png(answer)
    return Response(png, mimetype="image/png")


@app.route("/login/resend-unlock", methods=["POST"])
def login_resend_unlock():
    username = (request.form.get("username") or "").strip()
    notice = "If that account exists and is locked, an unlock email has been sent."
    account_locked = False
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT * FROM users WHERE LOWER(username) = LOWER(?) AND is_active = 1",
            (username,),
        ).fetchone()
        if not row:
            # Keep generic notice to avoid account enumeration.
            account_locked = True
        elif not auth_security.is_account_locked(row):
            notice = "That account is not locked. You can sign in with your password."
            account_locked = False
        else:
            account_locked = True
            send_result = _send_unlock_email_for_user(conn, row)
            conn.commit()
            if send_result.get("ok"):
                notice = "Unlock email sent. Check your inbox (and spam folder)."
            elif send_result.get("reason") == "no_email":
                notice = (
                    "This account is locked, but no email is on file. "
                    "Ask an administrator to unlock the account."
                )
            elif send_result.get("reason") == "smtp_not_configured":
                notice = (
                    "This account is locked, but SMTP is not configured on the server "
                    "(set SMTP_HOST in .env). Ask an administrator to unlock the account."
                )
            else:
                notice = (
                    "This account is locked, but unlock email could not be sent. "
                    "Ask an administrator to unlock the account."
                )
    finally:
        conn.close()
    return _login_page(
        notice=notice,
        username=username,
        account_locked=account_locked,
    )


@app.route("/unlock-account")
def unlock_account():
    token = (request.args.get("token") or "").strip()
    conn = get_db()
    try:
        user_id = auth_security.verify_and_consume_unlock_token(conn, token)
        if user_id:
            conn.commit()
            session["login_notice"] = "Your account has been unlocked. Please sign in."
            return redirect(url_for("index"))
    finally:
        conn.close()
    return _login_page(
        error="This unlock link is invalid or has expired. Request a new unlock email or contact an administrator.",
    )


@app.route("/logout")
def logout():
    session.pop(AUTH_USER_SESSION_KEY, None)
    return redirect(url_for("index"))


@app.route("/home")
def home():
    user = get_current_user()
    return render_template(
        "home.html",
        de_nav_section="home",
        home_notifications=_home_notifications(user),
    )


DASHBOARD_PERIODS = (
    ("today", "Today"),
    ("yesterday", "Yesterday"),
    ("7d", "7 Days"),
    ("30d", "30 Days"),
    ("mtd", "MTD"),
    ("qtd", "QTD"),
    ("ytd", "YTD"),
)
DASHBOARD_PERIOD_KEYS = {key for key, _label in DASHBOARD_PERIODS}
DASHBOARD_FILTER_LOCATION_ALL = CASH_LEDGER_FILTER_ALL
DASHBOARD_FILTER_LOCATIONS = (
    (CASH_LEDGER_FILTER_ALL, "All"),
    (OUTLET_HOTEL, OUTLET_HOTEL),
    (OUTLET_RESTAURANT, OUTLET_RESTAURANT),
    (OUTLET_BAR, OUTLET_BAR),
)


def _dashboard_period_range(period, today=None):
    """Return (date_from, date_to) for a named dashboard period (inclusive)."""
    today = today or date.today()
    key = (period or "").strip().lower()
    if key == "today":
        return today, today
    if key == "yesterday":
        day = today - timedelta(days=1)
        return day, day
    if key == "7d":
        return today - timedelta(days=6), today
    if key == "30d":
        return today - timedelta(days=29), today
    if key == "mtd":
        return today.replace(day=1), today
    if key == "qtd":
        quarter_month = ((today.month - 1) // 3) * 3 + 1
        return today.replace(month=quarter_month, day=1), today
    if key == "ytd":
        return today.replace(month=1, day=1), today
    return today - timedelta(days=29), today


def _format_dashboard_date_range_label(date_from, date_to):
    months = (
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
    )
    if not date_from or not date_to:
        return "Select date range"
    if date_from == date_to:
        return f"{date_from.day} {months[date_from.month - 1]}, {date_from.year}"
    if date_from.year == date_to.year and date_from.month == date_to.month:
        return (
            f"{date_from.day} – {date_to.day} "
            f"{months[date_from.month - 1]}, {date_from.year}"
        )
    if date_from.year == date_to.year:
        return (
            f"{date_from.day} {months[date_from.month - 1]} – "
            f"{date_to.day} {months[date_to.month - 1]}, {date_to.year}"
        )
    return (
        f"{date_from.day} {months[date_from.month - 1]}, {date_from.year} – "
        f"{date_to.day} {months[date_to.month - 1]}, {date_to.year}"
    )


def _resolve_main_dashboard_filters(args):
    today = date.today()
    period = (args.get("period") or "30d").strip().lower()
    if period not in DASHBOARD_PERIOD_KEYS and period != "custom":
        period = "30d"

    location = (args.get("location") or DASHBOARD_FILTER_LOCATION_ALL).strip()
    if location not in CASH_LEDGER_FILTER_LOCATIONS:
        location = DASHBOARD_FILTER_LOCATION_ALL
    location_label = (
        "All" if location == DASHBOARD_FILTER_LOCATION_ALL else location
    )

    raw_from = (args.get("date_from") or "").strip()
    raw_to = (args.get("date_to") or "").strip()
    parsed_from = _parse_sales_date(raw_from) if raw_from else None
    parsed_to = _parse_sales_date(raw_to) if raw_to else None
    if parsed_from and parsed_to and parsed_from > parsed_to:
        parsed_from, parsed_to = parsed_to, parsed_from

    if period in DASHBOARD_PERIOD_KEYS:
        expected_from, expected_to = _dashboard_period_range(period, today)
        if parsed_from and parsed_to:
            if (parsed_from, parsed_to) == (expected_from, expected_to):
                date_from, date_to = expected_from, expected_to
            else:
                date_from, date_to = parsed_from, parsed_to
                period = "custom"
        else:
            date_from, date_to = expected_from, expected_to
    elif parsed_from and parsed_to:
        date_from, date_to = parsed_from, parsed_to
        period = "custom"
    else:
        period = "30d"
        date_from, date_to = _dashboard_period_range(period, today)

    return {
        "selected_period": period,
        "selected_period_label": (
            dict(DASHBOARD_PERIODS).get(period, "Custom")
            if period != "custom"
            else "Custom"
        ),
        "selected_location": location,
        "selected_location_label": location_label,
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "date_range_label": _format_dashboard_date_range_label(date_from, date_to),
        "today_iso": today.isoformat(),
        "dashboard_periods": DASHBOARD_PERIODS,
        "dashboard_locations": DASHBOARD_FILTER_LOCATIONS,
        "_date_from": date_from,
        "_date_to": date_to,
    }


DASHBOARD_KPI_CARDS = (
    ("actual_sales", "Total Sales"),
    ("digital_transactions", "Digital Collection"),
    ("cash", "Cash Collection"),
    ("expense", "Expense"),
    ("difference", "Difference"),
)
DASHBOARD_KPI_KEYS = tuple(key for key, _label in DASHBOARD_KPI_CARDS)


def _dashboard_kpi_spark_series(conn, date_from, date_to, company=None, location=None, difference_mode=None):
    """Daily KPI values for sparklines over the selected range (or last 14 days if single-day)."""
    spark_from = date_from
    spark_to = date_to
    if spark_from == spark_to:
        spark_from = spark_to - timedelta(days=13)

    by_day = {}
    sql = "SELECT sales_date, sales_entry_values FROM sales_updates WHERE sales_date >= ? AND sales_date <= ?"
    params = [spark_from.isoformat(), spark_to.isoformat()]
    if company:
        sql += " AND company = ?"
        params.append(company)
    if location:
        sql += " AND location = ?"
        params.append(location)
    for row in conn.execute(sql, params).fetchall():
        day_key = str(row["sales_date"] or "")[:10]
        if not day_key:
            continue
        bucket = by_day.setdefault(
            day_key,
            {
                "actual_sales": 0.0,
                "digital_transactions": 0.0,
                "cash": 0.0,
                "actual_cash": 0.0,
                "difference": 0.0,
                "expense": 0.0,
            },
        )
        vals = json.loads(row["sales_entry_values"] or "{}")
        bucket["actual_sales"] += parse_money(vals.get("total_sales"))
        bucket["digital_transactions"] += get_digital_transactions(vals)
        bucket["cash"] += parse_money(vals.get("cash"))
        bucket["actual_cash"] += parse_money(vals.get("actual_cash"))
        if difference_mode != "cash_actual":
            bucket["difference"] += get_difference(vals)

    expense_sql = """SELECT sales_date, COALESCE(SUM(amount), 0) AS total
                     FROM sales_update_expenses
                     WHERE sales_date >= ? AND sales_date <= ?"""
    expense_params = [spark_from.isoformat(), spark_to.isoformat()]
    if company:
        expense_sql += " AND company = ?"
        expense_params.append(company)
    if location:
        expense_sql += " AND location = ?"
        expense_params.append(location)
    expense_sql += " GROUP BY sales_date"
    for row in conn.execute(expense_sql, expense_params).fetchall():
        day_key = str(row["sales_date"] or "")[:10]
        if not day_key:
            continue
        bucket = by_day.setdefault(
            day_key,
            {
                "actual_sales": 0.0,
                "digital_transactions": 0.0,
                "cash": 0.0,
                "actual_cash": 0.0,
                "difference": 0.0,
                "expense": 0.0,
            },
        )
        bucket["expense"] += float(row["total"] or 0)

    series = {key: [] for key in DASHBOARD_KPI_KEYS}
    cursor = spark_from
    while cursor <= spark_to:
        bucket = by_day.get(cursor.isoformat()) or {}
        if difference_mode == "cash_actual":
            day_difference = round_half_up(
                float(bucket.get("cash") or 0) - float(bucket.get("actual_cash") or 0),
                2,
            )
        else:
            day_difference = round_half_up(float(bucket.get("difference") or 0), 2)
        series["actual_sales"].append(round_half_up(float(bucket.get("actual_sales") or 0), 2))
        series["digital_transactions"].append(
            round_half_up(float(bucket.get("digital_transactions") or 0), 2)
        )
        series["cash"].append(round_half_up(float(bucket.get("cash") or 0), 2))
        series["expense"].append(round_half_up(float(bucket.get("expense") or 0), 2))
        series["difference"].append(day_difference)
        cursor += timedelta(days=1)
    return series


def _sparkline_polyline(values, width=140, height=32, pad=2):
    """Build an SVG path for a compact sparkline."""
    if not values:
        mid = height / 2.0
        return f"M{pad},{mid:.1f} L{width - pad},{mid:.1f}"
    lo = min(values)
    hi = max(values)
    span = (hi - lo) if hi != lo else 1.0
    n = len(values)
    if n == 1:
        y = pad + (height - 2 * pad) * (1 - (values[0] - lo) / span)
        return f"M{pad},{y:.1f} L{width - pad},{y:.1f}"
    parts = []
    for i, value in enumerate(values):
        x = pad + (width - 2 * pad) * (i / (n - 1))
        y = pad + (height - 2 * pad) * (1 - (value - lo) / span)
        parts.append(f"{x:.1f},{y:.1f}")
    return "M" + " L".join(parts)


def _dashboard_outlet_names(location=None):
    if location in (OUTLET_HOTEL, OUTLET_RESTAURANT, OUTLET_BAR):
        return [location]
    return [OUTLET_HOTEL, OUTLET_RESTAURANT, OUTLET_BAR]


def _dashboard_pos_menu_outlet(location=None):
    """Map dashboard outlet filter to POS menu sales outlet (Restaurant/Bar only)."""
    if location == OUTLET_HOTEL:
        return False  # no POS menu items
    if location == OUTLET_RESTAURANT:
        return POS_OUTLET_RESTAURANT
    if location == OUTLET_BAR:
        return POS_OUTLET_BAR
    return None  # All


def _build_main_dashboard_payload(conn, date_from, date_to, location=None):
    """Retail Intelligence–style dashboard: KPIs + trend + outlets + payment + heatmap."""
    bundle = _sales_report_kpi_bundle(conn, date_from, date_to, company=None, location=location)
    sparks = _dashboard_kpi_spark_series(conn, date_from, date_to, company=None, location=location)

    if date_from == date_to:
        prev_to = date_from - timedelta(days=1)
        prev_from = prev_to
        spark_from = date_to - timedelta(days=13)
    else:
        span_days = (date_to - date_from).days + 1
        prev_to = date_from - timedelta(days=1)
        prev_from = prev_to - timedelta(days=span_days - 1)
        spark_from = date_from
    spark_to = date_to

    # Total Sales = Sales Update total_sales (Hotel + Restaurant + Bar).
    spark_dates = date_range_days(spark_from, spark_to)

    daily_series = []
    su_days = date_range_days(date_from, date_to)
    if date_from == date_to:
        day_kpi = bundle["current"]
        d = date_from.isoformat()
        daily_series.append(
            {
                "date": d,
                "actual_sales": float(day_kpi.get("actual_sales") or 0),
                "digital_transactions": float(day_kpi.get("digital_transactions") or 0),
                "cash": float(day_kpi.get("cash") or 0),
                "expense": float(day_kpi.get("expense") or 0),
                "difference": float(day_kpi.get("difference") or 0),
                "transaction_count": 0,
            }
        )
    else:
        for i, d in enumerate(su_days):
            daily_series.append(
                {
                    "date": d,
                    "actual_sales": float(
                        (sparks.get("actual_sales") or [0] * len(su_days))[i]
                    ),
                    "digital_transactions": float(
                        (sparks.get("digital_transactions") or [0] * len(su_days))[i]
                    ),
                    "cash": float((sparks.get("cash") or [0] * len(su_days))[i]),
                    "expense": float((sparks.get("expense") or [0] * len(su_days))[i]),
                    "difference": float(
                        (sparks.get("difference") or [0] * len(su_days))[i]
                    ),
                    "transaction_count": 0,
                }
            )

    kpis = []
    for key, label in DASHBOARD_KPI_CARDS:
        values = sparks.get(key) or []
        kpis.append(
            {
                "key": key,
                "label": label,
                "value": bundle["current"][key],
                "value_compact": inr_compact(bundle["current"][key]),
                "change_pct": bundle["trends"].get(key),
                "sparkline_series": sparkline_series_from_values(spark_dates, values),
            }
        )

    outlet_names = _dashboard_outlet_names(location)
    outlet_totals = {}
    prev_outlet_totals = {}
    for name in outlet_names:
        outlet_totals[name] = float(
            _aggregate_sales_kpis(
                conn, date_from, date_to, company=None, location=name
            ).get("actual_sales")
            or 0
        )
        prev_outlet_totals[name] = float(
            _aggregate_sales_kpis(
                conn, prev_from, prev_to, company=None, location=name
            ).get("actual_sales")
            or 0
        )

    grand_sales = float(bundle["current"].get("actual_sales") or 0)
    company_leaderboard, sales_contribution = build_outlet_boards(
        outlet_totals, prev_outlet_totals, grand_sales
    )

    digital_cash_stack = []
    for item in daily_series:
        total = item["digital_transactions"] + item["cash"]
        if total > 0:
            dig_pct = round(item["digital_transactions"] / total * 100, 1)
            digital_cash_stack.append(
                {
                    "date": item["date"],
                    "digital_pct": dig_pct,
                    "cash_pct": round(100.0 - dig_pct, 1),
                }
            )

    dig_pct, cash_pct = payment_mode_pct(
        bundle["current"]["digital_transactions"], bundle["current"]["cash"]
    )
    prev_kpis = _aggregate_sales_kpis(
        conn, prev_from, prev_to, company=None, location=location
    )
    prev_dig_pct, prev_cash_pct = payment_mode_pct(
        prev_kpis["digital_transactions"], prev_kpis["cash"]
    )

    pos_outlet = _dashboard_pos_menu_outlet(location)
    if pos_outlet is False:
        menu_sales_rows = []
    else:
        menu_sales_rows = list_pos_menu_sales(
            conn,
            date_from=date_from.isoformat(),
            date_to=date_to.isoformat(),
            outlet=pos_outlet,
            settlement="settled",
        )

    dashboard = {
        "kpis": kpis,
        "daily_series": daily_series,
        "sales_trend": build_sales_trend(daily_series, grand_sales),
        "company_leaderboard": company_leaderboard,
        "sales_contribution": sales_contribution,
        "digital_cash_stack": digital_cash_stack,
        "payment_mode": {
            "digital_pct": dig_pct,
            "cash_pct": cash_pct,
            "digital_trend": round(dig_pct - prev_dig_pct, 1),
            "cash_trend": round(cash_pct - prev_cash_pct, 1),
        },
        "dow_avg": build_dow_avg(daily_series),
        "heatmap": build_sales_heatmap(daily_series, date_from, date_to),
        "top_selling_items": build_top_selling_items(menu_sales_rows, limit=5, sort_by="qty"),
        "top_selling_items_by_revenue": build_top_selling_items(
            menu_sales_rows, limit=5, sort_by="revenue"
        ),
    }

    # Keep legacy keys for any partial consumers / fitters.
    kpi_cards = [
        {
            "key": k["key"],
            "label": k["label"],
            "amount": k["value"],
            "trend": k["change_pct"],
            "spark_path": _sparkline_polyline(
                [pt["value"] for pt in (k.get("sparkline_series") or [])]
            ),
        }
        for k in kpis
    ]

    return {
        "dashboard": dashboard,
        "kpi_cards": kpi_cards,
        "kpi_vs_label": bundle["vs_label"],
    }


@app.route("/main-dashboard", endpoint="main_dashboard")
def main_dashboard():
    """Top-level Dashboard — Retail Intelligence layout for Hotel / Restaurant / Bar."""
    filters = _resolve_main_dashboard_filters(request.args)
    date_from = filters.pop("_date_from")
    date_to = filters.pop("_date_to")
    location = filters["selected_location"]
    location_filter = None if location == DASHBOARD_FILTER_LOCATION_ALL else location

    conn = get_db()
    try:
        payload = _build_main_dashboard_payload(
            conn, date_from, date_to, location=location_filter
        )
    finally:
        conn.close()

    return render_template(
        "main_dashboard.html",
        de_nav_section="main_dashboard",
        filter_form_action=url_for("main_dashboard"),
        main_dashboard_clear_url=url_for("main_dashboard"),
        **filters,
        **payload,
    )


@app.route("/master")
def master():
    """Master data workspace hub."""
    conn = get_db()
    payload = build_masters_dashboard(url_for, conn)
    return render_template(
        "master.html",
        de_nav_section="master",
        de_nav_master_view="home",
        **payload,
    )


def _category_master_page_render(template, **kwargs):
    kwargs.setdefault("auth_notice", _pop_auth_notice())
    kwargs.setdefault("de_nav_section", "master")
    kwargs.setdefault("de_nav_master_view", "category_master")
    return render_template(template, **kwargs)


def _category_master_form_payload(source=None, *, category_id=""):
    source = source or {}
    outlet = normalize_pos_outlet(source.get("outlet") or POS_OUTLET_RESTAURANT)
    visible_raw = source.get("is_visible")
    if isinstance(visible_raw, bool):
        is_visible = visible_raw
    elif visible_raw is None and not source:
        is_visible = True
    else:
        is_visible = str(visible_raw or "").strip().lower() in ("1", "true", "on", "yes")
    return {
        "id": category_id or "",
        "name": " ".join(str(source.get("name") or "").split()).strip(),
        "outlet": outlet,
        "is_visible": is_visible,
    }


@app.route("/masters/categories", endpoint="category_master")
def category_master():
    """POS menu Category Master — list and manage Restaurant/Bar categories."""
    user = get_current_user()
    if not (
        user_can_access_dashboard(user, "master")
        or user_can_access_dashboard(user, "point_of_sale")
        or user_can_access_dashboard(user, "point_of_sale_bar")
    ):
        return _permission_denied_response("You do not have access to Category Master.")

    selected_id = (request.args.get("category_id") or "").strip()
    saved_flag = (request.args.get("saved") or "").strip()
    form_focus = (request.args.get("focus") or "").strip() == "form"

    conn = get_db()
    try:
        ensure_pos_schema(conn)
        categories = list_pos_menu_categories(
            conn, outlets=[POS_OUTLET_RESTAURANT, POS_OUTLET_BAR]
        )
        selected = None
        if selected_id:
            selected = next((c for c in categories if str(c["id"]) == selected_id), None)
        conn.commit()
    finally:
        conn.close()

    if selected:
        form = {
            "id": selected["id"],
            "name": selected.get("name") or "",
            "outlet": selected.get("outlet") or POS_OUTLET_RESTAURANT,
            "is_visible": bool(selected.get("is_visible", True)),
        }
    else:
        form = _category_master_form_payload()

    success_message = ""
    if saved_flag == "created":
        success_message = "Category created successfully."
    elif saved_flag == "updated":
        success_message = "Category updated successfully."
    elif saved_flag == "deleted":
        success_message = "Category deleted successfully."

    return _category_master_page_render(
        "partials/master_embed/category.html" if is_embed_request() else "category_master.html",
        categories=categories,
        form=form,
        errors=[],
        success_message=success_message,
        form_focus=form_focus or bool(selected),
        show_form=form_focus or bool(selected),
        embed_mode=is_embed_request(),
    )


@app.route("/masters/categories/save", methods=["POST"], endpoint="save_category_master")
def save_category_master():
    user = get_current_user()
    if not (
        user_can_access_dashboard(user, "master")
        or user_can_access_dashboard(user, "point_of_sale")
        or user_can_access_dashboard(user, "point_of_sale_bar")
    ):
        return _permission_denied_response("You do not have access to Category Master.")

    category_id_raw = (request.form.get("category_id") or "").strip()
    try:
        category_id = int(category_id_raw) if category_id_raw else None
    except (TypeError, ValueError):
        category_id = None
    payload = _category_master_form_payload(request.form, category_id=category_id or "")
    embed = is_embed_request() or str(request.form.get("embed") or request.args.get("embed") or "") == "1"

    conn = get_db()
    try:
        ensure_pos_schema(conn)
        try:
            save_pos_menu_category(
                conn,
                category_id=category_id,
                name=payload["name"],
                is_visible=payload["is_visible"],
                outlet=payload["outlet"],
            )
            conn.commit()
        except ValueError as exc:
            conn.rollback()
            categories = list_pos_menu_categories(
                conn, outlets=[POS_OUTLET_RESTAURANT, POS_OUTLET_BAR]
            )
            form = dict(payload)
            form["id"] = category_id or ""
            return _category_master_page_render(
                "partials/master_embed/category.html" if embed else "category_master.html",
                categories=categories,
                form=form,
                errors=[str(exc)],
                success_message="",
                form_focus=True,
                show_form=True,
                embed_mode=embed,
            ), 400
    finally:
        conn.close()

    redirect_kwargs = {"saved": "updated" if category_id else "created"}
    if embed:
        redirect_kwargs["embed"] = 1
    return redirect(url_for("category_master", **redirect_kwargs))


@app.route("/masters/categories/delete", methods=["POST"], endpoint="delete_category_master")
def delete_category_master():
    user = get_current_user()
    if not (
        user_can_access_dashboard(user, "master")
        or user_can_access_dashboard(user, "point_of_sale")
        or user_can_access_dashboard(user, "point_of_sale_bar")
    ):
        return _permission_denied_response("You do not have access to Category Master.")

    category_id_raw = (request.form.get("category_id") or "").strip()
    embed = is_embed_request() or str(request.form.get("embed") or request.args.get("embed") or "") == "1"
    redirect_kwargs = {}
    if embed:
        redirect_kwargs["embed"] = 1

    try:
        category_id = int(category_id_raw)
    except (TypeError, ValueError):
        _queue_auth_notice("Category not found.")
        return redirect(url_for("category_master", **redirect_kwargs))

    conn = get_db()
    try:
        ensure_pos_schema(conn)
        try:
            soft_delete_pos_menu_category(conn, category_id)
            conn.commit()
        except ValueError:
            conn.rollback()
            _queue_auth_notice("Category not found.")
            fail_kwargs = dict(redirect_kwargs)
            fail_kwargs["category_id"] = category_id
            return redirect(url_for("category_master", **fail_kwargs))
    finally:
        conn.close()

    redirect_kwargs["saved"] = "deleted"
    return redirect(url_for("category_master", **redirect_kwargs))


@app.route("/reports")
def reports():
    """Cross-module reports hub — view and download module reports."""
    payload = build_reports_dashboard(url_for)
    return render_template(
        "reports.html",
        de_nav_section="report",
        de_nav_report_view="home",
        **payload,
    )


_SALES_REPORT_KINDS = {
    "hotel": {
        "title": "Hotel Sales",
        "page_endpoint": "sales_report_hotel",
        "export_endpoint": "sales_report_hotel_export",
        "export_sheet": "Hotel Sales",
        "export_prefix": "hotel_sales",
    },
    "restaurant": {
        "title": "Restaurant Sales",
        "page_endpoint": "sales_report_restaurant",
        "export_endpoint": "sales_report_restaurant_export",
        "export_sheet": "Restaurant Sales",
        "export_prefix": "restaurant_sales",
        "outlet": POS_OUTLET_RESTAURANT,
    },
    "bar": {
        "title": "Bar Sales",
        "page_endpoint": "sales_report_bar",
        "export_endpoint": "sales_report_bar_export",
        "export_sheet": "Bar Sales",
        "export_prefix": "bar_sales",
        "outlet": POS_OUTLET_BAR,
    },
}


def _sales_report_filters(args):
    """Parse shared Sales Report GET filters (page + export)."""
    today = date.today()
    date_from, date_to, date_filter_active = _resolve_optional_filter_date_range(
        args, "date_from", "date_to", default_fy=True
    )
    selected_status = (args.get("status") or "all").strip().lower()
    if selected_status not in ("all", "unsettled", "settled"):
        selected_status = "all"
    return {
        "today": today,
        "date_from": date_from,
        "date_to": date_to,
        "date_filter_active": date_filter_active,
        "selected_status": selected_status,
    }


def _sales_report_status_labels():
    return {
        "all": "All statuses",
        "unsettled": "Un Settled",
        "settled": "Settled",
    }


def _sales_report_pos_kpis(invoices):
    """Invoice-count / total / settled / unsettled KPIs for POS sales reports."""
    total_sales = 0.0
    settled_count = 0
    settled_amount = 0.0
    unsettled_count = 0
    unsettled_amount = 0.0
    for inv in invoices or []:
        amount = round_half_up(inv.get("grand_total"), 2)
        total_sales += amount
        is_settled = bool(inv.get("payment_modes")) or bool(
            str(inv.get("settled_at") or "").strip()
        )
        if is_settled:
            settled_count += 1
            settled_amount += amount
        else:
            unsettled_count += 1
            unsettled_amount += amount
    return {
        "total": len(invoices or []),
        "amount_sum": round_half_up(total_sales, 2),
        "settled": settled_count,
        "settled_amount": round_half_up(settled_amount, 2),
        "open": unsettled_count,
        "outstanding": round_half_up(unsettled_amount, 2),
    }


def _sales_report_load_rows(kind, filters):
    """Load invoice rows + KPIs for a sales report kind."""
    meta = _SALES_REPORT_KINDS[kind]
    date_from = (
        filters["date_from"].isoformat()
        if filters["date_filter_active"] and filters["date_from"]
        else None
    )
    date_to = (
        filters["date_to"].isoformat()
        if filters["date_filter_active"] and filters["date_to"]
        else None
    )
    status = filters["selected_status"]

    conn = get_db()
    try:
        if kind == "hotel":
            ensure_hotel_rooms_schema(conn)
            hotel_status = ""
            if status == "settled":
                hotel_status = "settled"
            elif status == "unsettled":
                hotel_status = "open"
            rows = list_hotel_room_invoices(
                conn,
                status=hotel_status,
                date_from=date_from,
                date_to=date_to,
            )
            kpis = hotel_room_invoice_kpis(rows)
            conn.commit()
            return rows, kpis

        ensure_pos_schema(conn)
        settlement = None if status == "all" else status
        # When no date filter, match POS ledger default window.
        query_from = date_from or "2000-01-01"
        query_to = date_to or filters["today"].isoformat()
        rows = list_pos_invoices(
            conn,
            date_from=query_from,
            date_to=query_to,
            settlement=settlement,
            outlet=meta["outlet"],
        )
        kpis = _sales_report_pos_kpis(rows)
        return rows, kpis
    finally:
        conn.close()


def _sales_report_page(kind):
    meta = _SALES_REPORT_KINDS[kind]
    filters = _sales_report_filters(request.args)
    rows, kpis = _sales_report_load_rows(kind, filters)
    status_labels = _sales_report_status_labels()

    clear_kwargs = {}
    if filters["selected_status"] != "all":
        clear_kwargs["status"] = filters["selected_status"]
    from_hub = (request.args.get("from_hub") or "").strip().lower()
    if from_hub == "reports":
        clear_kwargs["from_hub"] = "reports"

    export_kwargs = dict(clear_kwargs)
    if filters["date_filter_active"]:
        if filters["date_from"]:
            export_kwargs["date_from"] = filters["date_from"].isoformat()
        if filters["date_to"]:
            export_kwargs["date_to"] = filters["date_to"].isoformat()

    filter_kwargs = {}
    if from_hub == "reports":
        filter_kwargs["from_hub"] = "reports"

    return render_template(
        "sales_report.html",
        de_nav_section="report",
        de_nav_report_view="sales",
        page_title=meta["title"],
        report_kind=kind,
        invoices=rows,
        kpis=kpis,
        today_iso=filters["today"].isoformat(),
        date_from=filters["date_from"].isoformat()
        if filters["date_filter_active"] and filters["date_from"]
        else "",
        date_to=filters["date_to"].isoformat()
        if filters["date_filter_active"] and filters["date_to"]
        else "",
        active_date_filter=filters["date_filter_active"],
        selected_status=filters["selected_status"],
        selected_status_label=status_labels.get(
            filters["selected_status"], "All statuses"
        ),
        filter_form_action=url_for(meta["page_endpoint"], **filter_kwargs),
        sales_report_clear_url=url_for(meta["page_endpoint"], **clear_kwargs),
        sales_report_export_url=url_for(meta["export_endpoint"], **export_kwargs),
        preserve_from_hub=from_hub == "reports",
    )


def _sales_report_export(kind):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from io import BytesIO

    meta = _SALES_REPORT_KINDS[kind]
    filters = _sales_report_filters(request.args)
    rows, _kpis = _sales_report_load_rows(kind, filters)

    wb = Workbook()
    ws = wb.active
    ws.title = meta["export_sheet"]
    header_font = Font(bold=True)
    ws["A1"] = f"Hotel Bell Elite — {meta['title']}"
    ws["A1"].font = Font(bold=True, size=14)

    if kind == "hotel":
        headers = (
            "Invoice No",
            "Invoice Date",
            "Room",
            "Guest",
            "Booking No",
            "Check In",
            "Check Out",
            "Amount",
            "Advance",
            "Balance",
            "Status",
        )
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=title)
            cell.font = header_font
        for idx, row in enumerate(rows, start=4):
            status_key = (row.get("status") or "").strip().lower()
            ws.cell(row=idx, column=1, value=row.get("invoice_number") or "")
            ws.cell(row=idx, column=2, value=row.get("invoice_generated_at") or "")
            ws.cell(row=idx, column=3, value=row.get("room_number") or "")
            ws.cell(row=idx, column=4, value=row.get("guest_name") or "")
            ws.cell(row=idx, column=5, value=row.get("booking_number") or "")
            ws.cell(row=idx, column=6, value=row.get("check_in_date") or "")
            ws.cell(row=idx, column=7, value=row.get("check_out_date") or "")
            ws.cell(row=idx, column=8, value=row.get("estimated_total"))
            ws.cell(row=idx, column=9, value=row.get("advance_paid"))
            ws.cell(row=idx, column=10, value=row.get("balance_amount"))
            ws.cell(
                row=idx,
                column=11,
                value="Settled" if status_key == "settled" else "Un Settled",
            )
    else:
        headers = (
            "Order No",
            "Date",
            "Saved At",
            "Customer",
            "Mobile",
            "Order Type",
            "Payment Mode",
            "Settlement",
            "Captain",
            "Items",
            "Subtotal",
            "Discount",
            "GST",
            "Service",
            "Tip",
            "Total",
        )
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=title)
            cell.font = header_font
        for idx, inv in enumerate(rows, start=4):
            is_settled = bool(inv.get("payment_modes")) or bool(
                str(inv.get("settled_at") or "").strip()
            )
            ws.cell(row=idx, column=1, value=inv.get("order_no") or "")
            ws.cell(row=idx, column=2, value=inv.get("order_date") or "")
            ws.cell(row=idx, column=3, value=inv.get("saved_at") or "")
            ws.cell(row=idx, column=4, value=inv.get("customer_name") or "")
            ws.cell(row=idx, column=5, value=inv.get("customer_mobile") or "")
            ws.cell(
                row=idx,
                column=6,
                value=inv.get("order_type_label") or inv.get("order_type") or "",
            )
            ws.cell(
                row=idx,
                column=7,
                value=inv.get("payment_mode_label") or "Un Settled",
            )
            ws.cell(
                row=idx,
                column=8,
                value="Settled" if is_settled else "Un Settled",
            )
            ws.cell(row=idx, column=9, value=inv.get("captain") or "")
            ws.cell(row=idx, column=10, value=int(inv.get("item_count") or 0))
            ws.cell(row=idx, column=11, value=round_half_up(inv.get("subtotal"), 2))
            ws.cell(row=idx, column=12, value=round_half_up(inv.get("discount"), 2))
            ws.cell(row=idx, column=13, value=round_half_up(inv.get("gst"), 2))
            ws.cell(row=idx, column=14, value=round_half_up(inv.get("service"), 2))
            ws.cell(row=idx, column=15, value=round_half_up(inv.get("tip"), 2))
            ws.cell(row=idx, column=16, value=round_half_up(inv.get("grand_total"), 2))

    for column_cells in ws.columns:
        width = 12
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, 40))
        ws.column_dimensions[column_cells[0].column_letter].width = width

    if filters["date_filter_active"]:
        stamp = (
            f"{filters['date_from'].isoformat()}_to_{filters['date_to'].isoformat()}"
        )
        fname = f"{meta['export_prefix']}_{stamp}.xlsx"
    else:
        fname = f"{meta['export_prefix']}_all.xlsx"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/reports/sales/hotel", endpoint="sales_report_hotel")
def sales_report_hotel():
    """Hotel Sales report — room invoices invoice-wise."""
    return _sales_report_page("hotel")


@app.route("/reports/sales/hotel/export", endpoint="sales_report_hotel_export")
def sales_report_hotel_export():
    """Excel export for Hotel Sales report."""
    return _sales_report_export("hotel")


@app.route("/reports/sales/restaurant", endpoint="sales_report_restaurant")
def sales_report_restaurant():
    """Restaurant Sales report — POS invoices invoice-wise."""
    return _sales_report_page("restaurant")


@app.route("/reports/sales/restaurant/export", endpoint="sales_report_restaurant_export")
def sales_report_restaurant_export():
    """Excel export for Restaurant Sales report."""
    return _sales_report_export("restaurant")


@app.route("/reports/sales/bar", endpoint="sales_report_bar")
def sales_report_bar():
    """Bar Sales report — POS invoices invoice-wise."""
    return _sales_report_page("bar")


@app.route("/reports/sales/bar/export", endpoint="sales_report_bar_export")
def sales_report_bar_export():
    """Excel export for Bar Sales report."""
    return _sales_report_export("bar")


def _menu_sales_filters(args):
    """Parse Menu Sales GET filters (page + export)."""
    base = _sales_report_filters(args)
    outlet = (args.get("outlet") or "all").strip().lower()
    if outlet not in ("all", POS_OUTLET_RESTAURANT, POS_OUTLET_BAR):
        outlet = "all"
    raw_cat = (args.get("category_id") or args.get("category") or "").strip()
    try:
        category_id = int(raw_cat) if raw_cat else 0
    except (TypeError, ValueError):
        category_id = 0
    if category_id < 0:
        category_id = 0
    base["selected_outlet"] = outlet
    base["selected_category_id"] = category_id
    return base


def _menu_sales_outlet_labels():
    return {
        "all": "All outlets",
        POS_OUTLET_RESTAURANT: "Restaurant",
        POS_OUTLET_BAR: "Bar",
    }


def _menu_sales_load(filters):
    date_from = (
        filters["date_from"].isoformat()
        if filters["date_filter_active"] and filters["date_from"]
        else None
    )
    date_to = (
        filters["date_to"].isoformat()
        if filters["date_filter_active"] and filters["date_to"]
        else None
    )
    status = filters["selected_status"]
    settlement = None if status == "all" else status
    query_from = date_from or "2000-01-01"
    query_to = date_to or filters["today"].isoformat()
    outlet = filters["selected_outlet"]
    category_id = filters["selected_category_id"] or None

    conn = get_db()
    try:
        ensure_pos_schema(conn)
        rows = list_pos_menu_sales(
            conn,
            date_from=query_from,
            date_to=query_to,
            outlet=None if outlet == "all" else outlet,
            settlement=settlement,
            category_id=category_id,
        )
        kpis = pos_menu_sales_kpis(
            rows,
            conn,
            date_from=query_from,
            date_to=query_to,
            outlet=None if outlet == "all" else outlet,
            settlement=settlement,
            category_id=category_id,
        )
        cat_outlets = (
            [POS_OUTLET_RESTAURANT, POS_OUTLET_BAR]
            if outlet == "all"
            else [outlet]
        )
        categories = list_pos_menu_categories(
            conn, outlets=cat_outlets, include_inactive=False
        )
        return rows, kpis, categories
    finally:
        conn.close()


def _menu_sales_filter_kwargs(filters, *, include_dates=True, include_from_hub=True):
    kwargs = {}
    if filters["selected_status"] != "all":
        kwargs["status"] = filters["selected_status"]
    if filters["selected_outlet"] != "all":
        kwargs["outlet"] = filters["selected_outlet"]
    if filters["selected_category_id"]:
        kwargs["category_id"] = filters["selected_category_id"]
    if include_dates and filters["date_filter_active"]:
        if filters["date_from"]:
            kwargs["date_from"] = filters["date_from"].isoformat()
        if filters["date_to"]:
            kwargs["date_to"] = filters["date_to"].isoformat()
    from_hub = (request.args.get("from_hub") or "").strip().lower()
    if include_from_hub and from_hub == "reports":
        kwargs["from_hub"] = "reports"
    return kwargs


@app.route("/reports/sales/menu", endpoint="sales_report_menu")
def sales_report_menu():
    """Menu Sales report — item-wise order count, qty, and sale value."""
    filters = _menu_sales_filters(request.args)
    rows, kpis, categories = _menu_sales_load(filters)
    status_labels = _sales_report_status_labels()
    outlet_labels = _menu_sales_outlet_labels()

    selected_category_label = "All categories"
    for cat in categories:
        if int(cat.get("id") or 0) == filters["selected_category_id"]:
            selected_category_label = cat.get("name") or "All categories"
            break

    clear_kwargs = _menu_sales_filter_kwargs(
        filters, include_dates=False, include_from_hub=True
    )
    export_kwargs = _menu_sales_filter_kwargs(
        filters, include_dates=True, include_from_hub=True
    )
    # Always keep hub deep-link so filters/export preserve Back to Reports.
    filter_kwargs = {"from_hub": "reports"}
    clear_kwargs.setdefault("from_hub", "reports")
    export_kwargs.setdefault("from_hub", "reports")

    return render_template(
        "menu_sales_report.html",
        de_nav_section="report",
        de_nav_report_view="sales",
        page_title="Menu Insights",
        rows=rows,
        kpis=kpis,
        categories=categories,
        today_iso=filters["today"].isoformat(),
        date_from=filters["date_from"].isoformat()
        if filters["date_filter_active"] and filters["date_from"]
        else "",
        date_to=filters["date_to"].isoformat()
        if filters["date_filter_active"] and filters["date_to"]
        else "",
        active_date_filter=filters["date_filter_active"],
        selected_status=filters["selected_status"],
        selected_status_label=status_labels.get(
            filters["selected_status"], "All statuses"
        ),
        selected_outlet=filters["selected_outlet"],
        selected_outlet_label=outlet_labels.get(
            filters["selected_outlet"], "All outlets"
        ),
        selected_category_id=filters["selected_category_id"],
        selected_category_label=selected_category_label,
        filter_form_action=url_for("sales_report_menu", **filter_kwargs),
        sales_report_clear_url=url_for("sales_report_menu", **clear_kwargs),
        sales_report_export_url=url_for("sales_report_menu_export", **export_kwargs),
        preserve_from_hub=True,
        back_href=url_for("reports"),
        back_label="Back to Reports",
    )


@app.route("/reports/sales/menu/export", endpoint="sales_report_menu_export")
def sales_report_menu_export():
    """Excel export for Menu Sales report (item-wise)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    filters = _menu_sales_filters(request.args)
    rows, _kpis, _categories = _menu_sales_load(filters)

    wb = Workbook()
    ws = wb.active
    ws.title = "Menu Insights"

    # All header rows use #315A78 with white text; data rows stay white.
    header_fill = PatternFill(
        fill_type="solid",
        start_color="FF315A78",
        end_color="FF315A78",
    )
    title_font = Font(bold=True, size=14, color="FFFFFFFF")
    header_font = Font(bold=True, size=11, color="FFFFFFFF")
    body_font = Font(size=11, color="FF000000")
    thin = Side(style="thin", color="FF000000")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _whole_or_float(value):
        if value is None:
            return None
        try:
            num = float(value)
            return int(num) if num == int(num) else num
        except (TypeError, ValueError):
            return value

    headers = (
        "Item",
        "Category",
        "Outlet",
        "Order Count",
        "Qty Sold",
        "Sale Value",
    )

    # Write values first, then style (merge-safe fill application).
    ws["A1"] = "Hotel Bell Elite — Menu Insights"
    for col, title in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=title)

    for idx, row in enumerate(rows, start=3):
        item = (row.get("item_name") or "").strip().upper()
        category = (row.get("category_name") or "").strip().upper()
        outlet = row.get("outlet_label") or ""
        order_count = int(row.get("order_count") or 0)
        qty = _whole_or_float(row.get("qty_sold"))
        sale = _whole_or_float(round_half_up(row.get("sale_value"), 2))
        values = (item, category, outlet, order_count, qty, sale)
        for col, value in enumerate(values, start=1):
            ws.cell(row=idx, column=col, value=value)

    ws.merge_cells("A1:F1")
    last_row = max(2, ws.max_row)

    for col in range(1, 7):
        title_cell = ws.cell(row=1, column=col)
        title_cell.fill = header_fill
        title_cell.font = title_font
        title_cell.alignment = center
        title_cell.border = grid

        header_cell = ws.cell(row=2, column=col)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = center
        header_cell.border = grid

    for row_idx in range(3, last_row + 1):
        for col in range(1, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = body_font
            cell.border = grid
            cell.alignment = left if col == 1 else center

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    col_widths = (36, 16, 12, 14, 12, 12)
    for col, width in enumerate(col_widths, start=1):
        max_len = width
        for row in range(2, last_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            max_len = max(max_len, min(len(str(value)) + 2, 42))
        ws.column_dimensions[get_column_letter(col)].width = max_len

    if filters["date_filter_active"]:
        stamp = (
            f"{filters['date_from'].isoformat()}_to_{filters['date_to'].isoformat()}"
        )
        fname = f"menu_sales_{stamp}.xlsx"
    else:
        fname = "menu_sales_all.xlsx"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    # Avoid browsers reusing a previous unstyled download of the same name.
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response


def _customer_insights_filters(args):
    """Parse Customer Insights GET filters (page + export)."""
    base = _sales_report_filters(args)
    channel = (args.get("channel") or args.get("outlet") or "all").strip().lower()
    if channel not in ("all", "restaurant", "bar", "hotel"):
        channel = "all"
    base["selected_channel"] = channel
    return base


def _customer_insights_channel_labels():
    return {
        "all": "All channels",
        "restaurant": "Restaurant",
        "bar": "Bar",
        "hotel": "Hotel",
    }


def _customer_insights_load(filters):
    date_from = (
        filters["date_from"].isoformat()
        if filters["date_filter_active"] and filters["date_from"]
        else None
    )
    date_to = (
        filters["date_to"].isoformat()
        if filters["date_filter_active"] and filters["date_to"]
        else None
    )
    status = filters["selected_status"]
    settlement = None if status == "all" else status
    query_from = date_from or "2000-01-01"
    query_to = date_to or filters["today"].isoformat()
    channel = filters["selected_channel"]

    conn = get_db()
    try:
        rows = list_customer_insights(
            conn,
            date_from=query_from,
            date_to=query_to,
            channel=channel,
            settlement=settlement,
        )
        kpis = customer_insights_kpis(rows)
        return rows, kpis
    finally:
        conn.close()


def _customer_insights_filter_kwargs(filters, *, include_dates=True, include_from_hub=True):
    kwargs = {}
    if filters["selected_status"] != "all":
        kwargs["status"] = filters["selected_status"]
    if filters["selected_channel"] != "all":
        kwargs["channel"] = filters["selected_channel"]
    if include_dates and filters["date_filter_active"]:
        if filters["date_from"]:
            kwargs["date_from"] = filters["date_from"].isoformat()
        if filters["date_to"]:
            kwargs["date_to"] = filters["date_to"].isoformat()
    from_hub = (request.args.get("from_hub") or "").strip().lower()
    if include_from_hub and from_hub == "reports":
        kwargs["from_hub"] = "reports"
    return kwargs


@app.route(
    "/reports/sales/customer-insights",
    endpoint="sales_report_customer_insights",
)
def sales_report_customer_insights():
    """Customer Insights — spend and top item across Hotel, Restaurant, and Bar."""
    filters = _customer_insights_filters(request.args)
    rows, kpis = _customer_insights_load(filters)
    status_labels = _sales_report_status_labels()
    channel_labels = _customer_insights_channel_labels()

    clear_kwargs = _customer_insights_filter_kwargs(
        filters, include_dates=False, include_from_hub=True
    )
    export_kwargs = _customer_insights_filter_kwargs(
        filters, include_dates=True, include_from_hub=True
    )
    filter_kwargs = {"from_hub": "reports"}
    clear_kwargs.setdefault("from_hub", "reports")
    export_kwargs.setdefault("from_hub", "reports")

    return render_template(
        "customer_insights_report.html",
        de_nav_section="report",
        de_nav_report_view="sales",
        page_title="Customer Insights",
        rows=rows,
        kpis=kpis,
        today_iso=filters["today"].isoformat(),
        date_from=filters["date_from"].isoformat()
        if filters["date_filter_active"] and filters["date_from"]
        else "",
        date_to=filters["date_to"].isoformat()
        if filters["date_filter_active"] and filters["date_to"]
        else "",
        active_date_filter=filters["date_filter_active"],
        selected_status=filters["selected_status"],
        selected_status_label=status_labels.get(
            filters["selected_status"], "All statuses"
        ),
        selected_channel=filters["selected_channel"],
        selected_channel_label=channel_labels.get(
            filters["selected_channel"], "All channels"
        ),
        filter_form_action=url_for("sales_report_customer_insights", **filter_kwargs),
        sales_report_clear_url=url_for("sales_report_customer_insights", **clear_kwargs),
        sales_report_export_url=url_for(
            "sales_report_customer_insights_export", **export_kwargs
        ),
        preserve_from_hub=True,
        back_href=url_for("reports"),
        back_label="Back to Reports",
    )


@app.route(
    "/reports/sales/customer-insights/export",
    endpoint="sales_report_customer_insights_export",
)
def sales_report_customer_insights_export():
    """Excel export for Customer Insights report."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    filters = _customer_insights_filters(request.args)
    rows, _kpis = _customer_insights_load(filters)

    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Insights"

    header_fill = PatternFill(
        fill_type="solid",
        start_color="FF315A78",
        end_color="FF315A78",
    )
    title_font = Font(bold=True, size=14, color="FFFFFFFF")
    header_font = Font(bold=True, size=11, color="FFFFFFFF")
    body_font = Font(size=11, color="FF000000")
    thin = Side(style="thin", color="FF000000")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _whole_or_float(value):
        if value is None:
            return None
        try:
            num = float(value)
            return int(num) if num == int(num) else num
        except (TypeError, ValueError):
            return value

    headers = (
        "Customer",
        "Mobile",
        "Orders",
        "Top Item",
        "Restaurant",
        "Bar",
        "Hotel",
        "Total",
    )

    ws["A1"] = "Hotel Bell Elite — Customer Insights"
    for col, title in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=title)

    for idx, row in enumerate(rows, start=3):
        values = (
            (row.get("customer_name") or "").strip(),
            row.get("mobile") or "",
            int(row.get("order_count") or 0),
            (row.get("top_item") or "").strip(),
            _whole_or_float(round_half_up(row.get("restaurant_value"), 2)),
            _whole_or_float(round_half_up(row.get("bar_value"), 2)),
            _whole_or_float(round_half_up(row.get("hotel_value"), 2)),
            _whole_or_float(round_half_up(row.get("total_value"), 2)),
        )
        for col, value in enumerate(values, start=1):
            ws.cell(row=idx, column=col, value=value)

    ws.merge_cells("A1:H1")
    last_row = max(2, ws.max_row)

    for col in range(1, 9):
        title_cell = ws.cell(row=1, column=col)
        title_cell.fill = header_fill
        title_cell.font = title_font
        title_cell.alignment = center
        title_cell.border = grid

        header_cell = ws.cell(row=2, column=col)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = center
        header_cell.border = grid

    for row_idx in range(3, last_row + 1):
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = body_font
            cell.border = grid
            cell.alignment = left if col in (1, 4) else center

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    col_widths = (22, 14, 10, 28, 12, 12, 12, 12)
    for col, width in enumerate(col_widths, start=1):
        max_len = width
        for row in range(2, last_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            max_len = max(max_len, min(len(str(value)) + 2, 42))
        ws.column_dimensions[get_column_letter(col)].width = max_len

    if filters["date_filter_active"]:
        stamp = (
            f"{filters['date_from'].isoformat()}_to_{filters['date_to'].isoformat()}"
        )
        fname = f"customer_insights_{stamp}.xlsx"
    else:
        fname = "customer_insights_all.xlsx"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response


@app.route("/settings")
def settings():
    """Workspace settings hub."""
    user = get_current_user()
    settings_cards = []
    if user_can_access_dashboard(user, "point_of_sale"):
        settings_cards.append(
            {
                "name": "Restaurant Settings",
                "href": url_for("point_of_sale_settings"),
                "icon": "restaurant",
                "icon_tone": "indigo",
            }
        )
    if user_can_access_dashboard(user, "point_of_sale_bar"):
        settings_cards.append(
            {
                "name": "Bar Settings",
                "href": url_for("bar_point_of_sale_settings"),
                "icon": "bar",
                "icon_tone": "rose",
            }
        )
    if user_can_access_dashboard(user, "hotel_rooms"):
        settings_cards.append(
            {
                "name": "Hotel Settings",
                "href": url_for("hotel_settings"),
                "icon": "hotel",
                "icon_tone": "teal",
            }
        )
    return render_template(
        "settings.html",
        de_nav_section="settings",
        de_nav_settings_view="home",
        settings_cards=settings_cards,
    )


def _pos_outlet_from_request():
    """Resolve restaurant|bar from path or endpoint (Bar twin routes)."""
    path = request.path or ""
    if path.startswith("/bar-point-of-sale"):
        return POS_OUTLET_BAR
    endpoint = request.endpoint or ""
    if endpoint.startswith("bar_point_of_sale") or endpoint.startswith("bar_export_pos"):
        return POS_OUTLET_BAR
    return POS_OUTLET_RESTAURANT


def _pos_menu_list_outlets(outlet):
    """
    Outlets for menu catalog GET.
    Restaurant POS may pass ?include_outlets=bar (and Bar POS ?include_outlets=restaurant)
    so each bill can search/add the other outlet's menu. Settings CRUD stays single-outlet.
    """
    outlet = normalize_pos_outlet(outlet)
    raw = (request.args.get("include_outlets") or "").strip().lower()
    if not raw:
        return [outlet]
    allowed = {
        POS_OUTLET_RESTAURANT: POS_OUTLET_BAR,
        POS_OUTLET_BAR: POS_OUTLET_RESTAURANT,
    }
    peer = allowed.get(outlet)
    extras = []
    for part in raw.replace(";", ",").split(","):
        key = part.strip()
        if key in ("bar", POS_OUTLET_BAR) and peer == POS_OUTLET_BAR:
            extras.append(POS_OUTLET_BAR)
        elif key in ("restaurant", "resto", POS_OUTLET_RESTAURANT) and peer == POS_OUTLET_RESTAURANT:
            extras.append(POS_OUTLET_RESTAURANT)
    if not extras:
        return [outlet]
    return [outlet] + extras


def _pos_api_base(outlet):
    return (
        "/bar-point-of-sale"
        if normalize_pos_outlet(outlet) == POS_OUTLET_BAR
        else "/point-of-sale"
    )


def _pos_nav_section(outlet):
    return "bar_pos" if normalize_pos_outlet(outlet) == POS_OUTLET_BAR else "pos"


def _pos_label(outlet):
    return "Bar" if normalize_pos_outlet(outlet) == POS_OUTLET_BAR else "Restaurant"


def _pos_tip_location(outlet):
    return OUTLET_BAR if normalize_pos_outlet(outlet) == POS_OUTLET_BAR else OUTLET_RESTAURANT


def _pos_endpoint(name, outlet):
    """Map shared view name to restaurant or bar Flask endpoint."""
    outlet = normalize_pos_outlet(outlet)
    if outlet != POS_OUTLET_BAR:
        return name
    if name.startswith("point_of_sale"):
        return "bar_" + name
    if name == "export_pos_invoice_ledger_report":
        return "bar_export_pos_invoice_ledger_report"
    return name


POS_RESTAURANT_RECEIPT_CONFIG = {
    "business_name": "SPICE MULTICUISINE",
    "address": "Gurudwara Lane, Aberdeen bazar, Sri Vijaya Puram, Andaman & Nicobar 744101",
    "gst": "35AANFH8592H1ZS",
    "fssai": "12922101000132",
    "logo_url": "/static/pos/spice-receipt-logo.jpg",
    "user_label": "RESTAURANT",
}

POS_BAR_RECEIPT_CONFIG = {
    "business_name": "IRISH BARREL HOUSE BAR",
    "address": "Gurudwara Lane, Aberdeen bazar, Sri Vijaya Puram, Andaman & Nicobar 744101",
    "gst": "35AANFH8592H1ZS",
    "fssai": "12922101000132",
    "logo_url": "/static/pos/irish-barrel-house-logo.png",
    "user_label": "BAR",
}


def _pos_receipt_config(outlet, user=None):
    """Thermal bill branding for restaurant vs bar outlets.

    ``user_label`` is the logged-in cashier name printed on the bill footer.
    """
    base = (
        POS_BAR_RECEIPT_CONFIG
        if normalize_pos_outlet(outlet) == POS_OUTLET_BAR
        else POS_RESTAURANT_RECEIPT_CONFIG
    )
    cfg = dict(base)
    actor = user if user is not None else get_current_user()
    label = _user_display_name(actor) if actor else ""
    if label and label != "User":
        cfg["user_label"] = label
    elif actor and (actor.get("username") or "").strip():
        cfg["user_label"] = str(actor.get("username")).strip()
    return cfg


def _pos_page_context(outlet, pos_view, **extra):
    outlet = normalize_pos_outlet(outlet)
    ctx = {
        "de_nav_section": _pos_nav_section(outlet),
        "de_nav_pos_view": pos_view,
        "pos_outlet": outlet,
        "pos_api_base": _pos_api_base(outlet),
        "pos_label": _pos_label(outlet),
        "pos_store_outlet": "bar" if outlet == POS_OUTLET_BAR else "restaurant",
        "pos_receipt_config": _pos_receipt_config(outlet),
    }
    ctx.update(extra)
    return ctx


def _pos_invoice_belongs_to_outlet(invoice, outlet):
    if not invoice:
        return False
    return normalize_pos_outlet(invoice.get("outlet")) == normalize_pos_outlet(outlet)


@app.route("/point-of-sale", endpoint="point_of_sale")
@app.route("/bar-point-of-sale", endpoint="bar_point_of_sale")
def point_of_sale():
    """Point of Sale Tables floor — counter workspace (not Sales Analytics)."""
    outlet = _pos_outlet_from_request()
    areas = []
    tables = []
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        sync_pos_floor_occupancy_from_open_orders(conn, outlet)
        layout = get_pos_floor_layout(conn, outlet)
        floor = _pos_floor_api_payload(conn, layout, outlet)
        areas = floor.get("areas") or []
        area_name_by_id = {
            str(a.get("id") or ""): (a.get("name") or a.get("id") or "Floor")
            for a in areas
            if isinstance(a, dict)
        }
        tables = []
        for raw in floor.get("tables") or []:
            if not isinstance(raw, dict):
                continue
            row = dict(raw)
            area_id = str(row.get("areaId") or "").strip()
            row["area_key"] = area_name_by_id.get(area_id) or area_id or "Floor"
            row["area"] = row["area_key"]
            tables.append(row)
        conn.commit()
    finally:
        conn.close()
    return render_template(
        "point_of_sale.html",
        tables=tables,
        areas=areas,
        **_pos_page_context(outlet, "tables"),
    )


@app.route("/hotel/rooms", endpoint="hotel_rooms")
def hotel_rooms():
    """Hotel Rooms floor board — occupancy status by floor and type."""
    conn = get_db()
    try:
        ensure_agencies_schema(conn)
        agencies = list_agencies(conn)
        conn.commit()
    finally:
        conn.close()
    return render_template(
        "hotel_rooms.html",
        de_nav_section="hotel",
        de_nav_hotel_view="rooms",
        today_iso=date.today().isoformat(),
        agencies=agencies,
    )


@app.route("/hotel/settings", endpoint="hotel_settings")
def hotel_settings():
    """Hotel Settings — floors, rooms, taxes, invoice, payment, printers."""
    return render_template(
        "hotel_settings.html",
        de_nav_section="hotel",
        de_nav_hotel_view="settings",
        hotel_room_types=[
            {"key": key, "label": label}
            for key, label in (
                ("premium_without_balcony", "Premium Room"),
                ("premium_deluxe_balcony", "Deluxe with Balcony"),
                ("premium_suite_tub", "Suite Room"),
            )
        ],
    )


@app.route("/hotel/api/settings", methods=["GET", "PUT", "POST"], endpoint="hotel_settings_api")
def hotel_settings_api():
    """Load or replace hotel settings JSON (independent from Restaurant/Bar POS)."""
    conn = get_db()
    try:
        ensure_hotel_rooms_schema(conn)
        if request.method == "GET":
            settings = get_hotel_settings(conn)
            rates = get_hotel_tax_rates(conn)
            tariff = get_hotel_tariff_rates(conn)
            conn.commit()
            return jsonify(
                {
                    "ok": True,
                    "settings": settings,
                    "taxRates": rates,
                    "tariffRates": tariff,
                }
            )

        data = request.get_json(silent=True) or {}
        if "settings" in data:
            settings = data.get("settings")
        else:
            settings = data
        if not isinstance(settings, dict):
            return jsonify({"ok": False, "error": "settings object is required."}), 400
        saved = save_hotel_settings(conn, settings)
        conn.commit()
        rates = get_hotel_tax_rates(conn)
        tariff = get_hotel_tariff_rates(conn)
        return jsonify(
            {
                "ok": True,
                "settings": saved,
                "taxRates": rates,
                "tariffRates": tariff,
            }
        )
    finally:
        conn.close()


def _hotel_invoice_ledger_filters(args):
    """Parse hotel invoice ledger GET filters (shared by page + export)."""
    today = date.today()
    date_from, date_to, date_filter_active = _resolve_optional_filter_date_range(
        args, "date_from", "date_to", default_fy=True
    )
    selected_status = (args.get("status") or "all").strip().lower()
    if selected_status not in ("all", "open", "settled"):
        selected_status = "all"
    status_filter = "" if selected_status == "all" else selected_status
    q = (args.get("q") or "").strip()
    return {
        "today": today,
        "date_from": date_from,
        "date_to": date_to,
        "date_filter_active": date_filter_active,
        "selected_status": selected_status,
        "status_filter": status_filter,
        "q": q,
    }


@app.route("/hotel/invoice-ledger", endpoint="hotel_invoice_ledger")
def hotel_invoice_ledger():
    """Hotel room invoice ledger — archived invoices with View / Print."""
    filters = _hotel_invoice_ledger_filters(request.args)
    conn = get_db()
    try:
        ensure_hotel_rooms_schema(conn)
        rows = list_hotel_room_invoices(
            conn,
            q=filters["q"],
            status=filters["status_filter"],
            date_from=filters["date_from"].isoformat()
            if filters["date_filter_active"] and filters["date_from"]
            else None,
            date_to=filters["date_to"].isoformat()
            if filters["date_filter_active"] and filters["date_to"]
            else None,
        )
        kpis = hotel_room_invoice_kpis(rows)
        conn.commit()
    finally:
        conn.close()

    status_labels = {
        "all": "All statuses",
        "open": "Un Settled",
        "settled": "Settled",
    }
    clear_kwargs = {}
    if filters["selected_status"] != "all":
        clear_kwargs["status"] = filters["selected_status"]
    if filters["q"]:
        clear_kwargs["q"] = filters["q"]

    export_kwargs = dict(clear_kwargs)
    if filters["date_filter_active"]:
        if filters["date_from"]:
            export_kwargs["date_from"] = filters["date_from"].isoformat()
        if filters["date_to"]:
            export_kwargs["date_to"] = filters["date_to"].isoformat()

    return render_template(
        "hotel_invoice_ledger.html",
        de_nav_section="hotel",
        de_nav_hotel_view="invoice_ledger",
        page_title="Invoice Ledger",
        invoices=rows,
        kpis=kpis,
        today_iso=filters["today"].isoformat(),
        date_from=filters["date_from"].isoformat()
        if filters["date_filter_active"] and filters["date_from"]
        else "",
        date_to=filters["date_to"].isoformat()
        if filters["date_filter_active"] and filters["date_to"]
        else "",
        active_date_filter=filters["date_filter_active"],
        selected_status=filters["selected_status"],
        selected_status_label=status_labels.get(
            filters["selected_status"], "All statuses"
        ),
        search_q=filters["q"],
        filter_form_action=url_for("hotel_invoice_ledger"),
        invoice_ledger_clear_url=url_for("hotel_invoice_ledger", **clear_kwargs),
        invoice_ledger_export_url=url_for("hotel_invoice_ledger_export", **export_kwargs),
    )


@app.route("/hotel/invoice-ledger/export", endpoint="hotel_invoice_ledger_export")
def hotel_invoice_ledger_export():
    """Excel download of hotel room invoices for the selected filters."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from io import BytesIO

    filters = _hotel_invoice_ledger_filters(request.args)
    conn = get_db()
    try:
        ensure_hotel_rooms_schema(conn)
        rows = list_hotel_room_invoices(
            conn,
            q=filters["q"],
            status=filters["status_filter"],
            date_from=filters["date_from"].isoformat()
            if filters["date_filter_active"] and filters["date_from"]
            else None,
            date_to=filters["date_to"].isoformat()
            if filters["date_filter_active"] and filters["date_to"]
            else None,
        )
        conn.commit()
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Hotel Invoices"
    header_font = Font(bold=True)
    ws["A1"] = "Hotel Bell Elite — Invoice Ledger"
    ws["A1"].font = Font(bold=True, size=14)
    headers = (
        "Invoice No",
        "Invoice Date",
        "Room",
        "Guest",
        "Booking No",
        "Check In",
        "Check Out",
        "Amount",
        "Advance",
        "Balance",
        "Status",
    )
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=title)
        cell.font = header_font
    for idx, row in enumerate(rows, start=4):
        ws.cell(row=idx, column=1, value=row.get("invoice_number") or "")
        ws.cell(row=idx, column=2, value=row.get("invoice_generated_at") or "")
        ws.cell(row=idx, column=3, value=row.get("room_number") or "")
        ws.cell(row=idx, column=4, value=row.get("guest_name") or "")
        ws.cell(row=idx, column=5, value=row.get("booking_number") or "")
        ws.cell(row=idx, column=6, value=row.get("check_in_date") or "")
        ws.cell(row=idx, column=7, value=row.get("check_out_date") or "")
        ws.cell(row=idx, column=8, value=row.get("estimated_total"))
        ws.cell(row=idx, column=9, value=row.get("advance_paid"))
        ws.cell(row=idx, column=10, value=row.get("balance_amount"))
        status_key = (row.get("status") or "").strip().lower()
        ws.cell(
            row=idx,
            column=11,
            value="Settled" if status_key == "settled" else "Un Settled",
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = date.today().isoformat()
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"hotel_invoice_ledger_{stamp}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route(
    "/hotel/invoice-ledger/api/<path:invoice_number>",
    endpoint="hotel_invoice_ledger_api",
)
def hotel_invoice_ledger_api(invoice_number):
    """JSON room payload for View / Print of an archived invoice."""
    conn = get_db()
    try:
        ensure_hotel_rooms_schema(conn)
        item = get_hotel_room_invoice(conn, invoice_number)
        conn.commit()
    finally:
        conn.close()
    if not item:
        return jsonify({"ok": False, "error": "Invoice not found."}), 404
    return jsonify({"ok": True, "invoice": item, "room": item.get("room")})


@app.route(
    "/hotel/invoice-ledger/api/<path:invoice_number>/settle",
    methods=["POST"],
    endpoint="hotel_invoice_ledger_settle_api",
)
def hotel_invoice_ledger_settle_api(invoice_number):
    """Record payment for an open hotel invoice from the ledger settle modal."""
    data = request.get_json(silent=True) or {}
    payment = data.get("payment") if isinstance(data.get("payment"), dict) else None
    payment_splits = data.get("payment_splits") or data.get("paymentSplits")
    note = data.get("note") or data.get("notes") or ""
    if payment is None and payment_splits is None:
        payment = {
            "amount": data.get("amount"),
            "method": data.get("method") or data.get("paymentMethod"),
            "reference": data.get("reference")
            or data.get("paymentReference")
            or data.get("payment_reference"),
            "note": note,
        }
    conn = get_db()
    try:
        ensure_hotel_rooms_schema(conn)
        try:
            result = record_hotel_room_invoice_payment(
                conn,
                invoice_number,
                payment=payment,
                payment_splits=payment_splits
                if isinstance(payment_splits, list)
                else None,
                note=note,
            )
            conn.commit()
        except ValueError as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": str(exc)}), 400
    finally:
        conn.close()
    return jsonify(
        {
            "ok": True,
            "invoice": result.get("invoice"),
            "room": result.get("room"),
            "payment": result.get("payment"),
            "payments": result.get("payments") or [],
        }
    )


@app.route("/hotel/api/rooms", methods=["GET", "PUT"], endpoint="hotel_rooms_api")
def hotel_rooms_api():
    """Load or replace hotel rooms layout JSON (floors + rooms + KPI counts)."""
    conn = get_db()
    try:
        ensure_hotel_rooms_schema(conn)
        if request.method == "GET":
            layout = get_hotel_rooms_layout(conn)
            rooms = []
            for room in layout.get("rooms") or []:
                if not isinstance(room, dict):
                    continue
                item = dict(room)
                enrich_hotel_room_merge_fields(item, layout.get("rooms"))
                rooms.append(item)
            counts = hotel_rooms_status_counts(layout)
            conn.commit()
            resp = jsonify(
                {
                    "ok": True,
                    "floors": layout.get("floors") or [],
                    "rooms": rooms,
                    "counts": counts,
                }
            )
            resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
            resp.headers["Pragma"] = "no-cache"
            return resp

        data = request.get_json(silent=True) or {}
        # Allow compact status-only updates: { roomId, status }
        room_id = data.get("roomId") or data.get("room_id") or data.get("id")
        if room_id and "status" in data and not isinstance(data.get("rooms"), list):
            try:
                saved = update_hotel_room_status(conn, room_id, data.get("status"), data)
            except ValueError as exc:
                conn.rollback()
                return jsonify({"ok": False, "error": str(exc)}), 400
            counts = hotel_rooms_status_counts(saved)
            conn.commit()
            return jsonify(
                {
                    "ok": True,
                    "floors": saved.get("floors") or [],
                    "rooms": saved.get("rooms") or [],
                    "counts": counts,
                }
            )

        floors = data.get("floors")
        rooms = data.get("rooms")
        if not isinstance(floors, list) or not isinstance(rooms, list):
            return jsonify({"ok": False, "error": "floors and rooms arrays are required."}), 400
        try:
            saved = save_hotel_rooms_layout(conn, floors, rooms)
        except ValueError as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": str(exc)}), 400
        counts = hotel_rooms_status_counts(saved)
        conn.commit()
        return jsonify(
            {
                "ok": True,
                "floors": saved.get("floors") or [],
                "rooms": saved.get("rooms") or [],
                "counts": counts,
            }
        )
    finally:
        conn.close()


@app.route("/hotel/rooms/<room_id>", endpoint="hotel_room_detail")
def hotel_room_detail(room_id):
    """Room Onboarding detail shell for a single room."""
    conn = get_db()
    try:
        room = get_hotel_room(conn, room_id)
        ensure_agencies_schema(conn)
        agencies = list_agencies(conn)
        conn.commit()
    finally:
        conn.close()
    if not room:
        abort(404)
    status = (room.get("status") or "vacant").lower()
    status_subtitles = {
        "vacant": "Ready for check-in",
        "occupied": "Occupied by Guest",
        "reserved": "Reserved for Arrival",
        "dirty": "Waiting for Cleaning",
        "out_of_order": "Out of service",
    }
    hk_dirty = status == "dirty"
    status_label = (
        room.get("statusLabel")
        or {
            "vacant": "Vacant",
            "occupied": "Occupied",
            "reserved": "Reserved",
            "dirty": "Dirty",
            "out_of_order": "Out of order",
        }.get(status, status.replace("_", " ").title())
    )
    if hk_dirty:
        hk_status = "Dirty"
        hk_subtitle = "Waiting for cleaning"
    else:
        hk_status = "Cleaned"
        hk_subtitle = "Ready for guests"
    return render_template(
        "hotel_room_detail.html",
        de_nav_section="hotel",
        de_nav_hotel_view="rooms",
        room=room,
        status_subtitle=status_subtitles.get(status, ""),
        housekeeping_status=hk_status,
        housekeeping_subtitle=hk_subtitle,
        housekeeping_dirty=hk_dirty,
        agencies=agencies,
        agency_master_url=url_for("agency_master"),
        agency_create_url=url_for("create_agency"),
        agencies_api_url=url_for("list_agencies_api"),
        today_iso=date.today().isoformat(),
    )


@app.route("/hotel/rooms/<room_id>/invoice", endpoint="hotel_room_invoice_page")
def hotel_room_invoice_page(room_id):
    """POS-style finalize workspace for a room stay invoice."""
    conn = get_db()
    try:
        ensure_hotel_rooms_schema(conn)
        room = get_hotel_room(conn, room_id)
        if room:
            enrich_hotel_room_merge_fields(room)
        conn.commit()
    finally:
        conn.close()
    if not room:
        abort(404)
    status = (room.get("status") or "vacant").lower()
    stay = room.get("stay") if isinstance(room.get("stay"), dict) else None
    guest_name = ""
    if stay:
        guest_name = (
            stay.get("guestName")
            or stay.get("guest_name")
            or " ".join(
                p
                for p in [
                    stay.get("title") or "",
                    stay.get("firstName") or stay.get("first_name") or "",
                    stay.get("lastName") or stay.get("last_name") or "",
                ]
                if p
            ).strip()
        )
    return render_template(
        "hotel_room_invoice_page.html",
        de_nav_section="hotel",
        de_nav_hotel_view="rooms",
        room=room,
        room_status=status,
        guest_name=guest_name or "Guest",
        open_settle=str(request.args.get("settle") or "").strip() in ("1", "true", "yes"),
        today_iso=date.today().isoformat(),
    )


@app.route("/hotel/api/guests/lookup", methods=["GET"], endpoint="hotel_guest_lookup_api")
def hotel_guest_lookup_api():
    """Lookup a returning hotel guest by mobile number."""
    mobile = (request.args.get("mobile") or "").strip()
    conn = get_db()
    try:
        ensure_hotel_rooms_schema(conn)
        guest = find_hotel_guest_by_mobile(conn, mobile)
        conn.commit()
    finally:
        conn.close()
    if not guest:
        return jsonify({"ok": True, "found": False, "guest": None})
    return jsonify({"ok": True, "found": True, "guest": guest})


@app.route("/hotel/api/customers", methods=["GET"], endpoint="hotel_customers_api")
def hotel_customers_api():
    """Search Customer Master for hotel reserve/check-in mobile autocomplete."""
    query = request.args.get("q") or request.args.get("mobile") or ""
    conn = get_db()
    try:
        ensure_customers_schema(conn)
        customers = search_customers(conn, query, limit=8)
        conn.commit()
        return jsonify({"ok": True, "customers": customers})
    finally:
        conn.close()


@app.route("/hotel/api/id-documents", methods=["POST"], endpoint="hotel_id_document_upload")
def hotel_id_document_upload():
    """Upload and compress a guest ID proof (image → WebP, PDF → Ghostscript)."""
    upload = request.files.get("file") or request.files.get("idDocument")
    if not upload or not upload.filename:
        return jsonify({"ok": False, "error": "Choose an ID document to upload."}), 400
    try:
        result = process_uploaded_id_document(upload, upload.filename)
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
    except RuntimeError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500
    except Exception:
        return jsonify({"ok": False, "error": "Could not compress the document."}), 500
    return jsonify({"ok": True, "document": result})


@app.route(
    "/hotel/api/id-documents/<path:stored_name>",
    methods=["GET"],
    endpoint="hotel_id_document_file",
)
def hotel_id_document_file(stored_name):
    """Serve a compressed ID document for authenticated hotel users."""
    path = resolve_stored_id_document(stored_name)
    if not path:
        abort(404)
    mime = "image/webp" if path.suffix.lower() == ".webp" else "application/pdf"
    return send_file(path, mimetype=mime, as_attachment=False, download_name=path.name)


@app.route("/hotel/api/rooms/<room_id>", methods=["GET", "PUT"], endpoint="hotel_room_detail_api")
def hotel_room_detail_api(room_id):
    """Load or update a single hotel room (status / check-in)."""
    conn = get_db()
    try:
        ensure_hotel_rooms_schema(conn)
        if request.method == "GET":
            room = get_hotel_room(conn, room_id)
            if not room:
                return jsonify({"ok": False, "error": "Room not found."}), 404
            conn.commit()
            return jsonify({"ok": True, "room": room})

        data = request.get_json(silent=True) or {}
        action = (data.get("action") or "").strip().lower()
        try:
            if action == "reserve":
                stay = data.get("stay") if isinstance(data.get("stay"), dict) else {}
                check_in = (
                    data.get("checkInDate")
                    or data.get("check_in_date")
                    or stay.get("checkInDate")
                    or stay.get("check_in_date")
                    or ""
                )
                check_out = (
                    data.get("checkOutDate")
                    or data.get("check_out_date")
                    or stay.get("checkOutDate")
                    or stay.get("check_out_date")
                    or ""
                )
                guest_name = (
                    stay.get("guestName")
                    or stay.get("guest_name")
                    or " ".join(
                        p
                        for p in (
                            (stay.get("firstName") or stay.get("first_name") or "").strip(),
                            (stay.get("lastName") or stay.get("last_name") or "").strip(),
                        )
                        if p
                    )
                ).strip()
                mobile = (stay.get("mobile") or "").strip()
                mobile_digits = "".join(ch for ch in mobile if ch.isdigit())[:10]
                if not guest_name:
                    return jsonify({"ok": False, "error": "Guest name is required."}), 400
                if len(mobile_digits) != 10:
                    return jsonify(
                        {"ok": False, "error": "Mobile number must be exactly 10 digits."}
                    ), 400
                mobile = mobile_digits
                stay["mobile"] = mobile
                if not stay.get("guestName"):
                    stay["guestName"] = guest_name
                if not stay.get("mobileCountry"):
                    stay["mobileCountry"] = "+91"
                room = save_hotel_room_reservation(
                    conn,
                    room_id,
                    check_in,
                    check_out,
                    stay_fields=stay,
                    replace=bool(data.get("replace")),
                )
                try:
                    ensure_customers_schema(conn)
                    upsert_customer(
                        conn,
                        guest_name,
                        mobile,
                        stay.get("address") or "",
                        stay.get("email") or "",
                    )
                except Exception:
                    pass
                try:
                    agency_name = (stay.get("agencyName") or stay.get("agency_name") or "").strip()
                    if agency_name:
                        upsert_agency_by_name(
                            conn,
                            agency_name,
                            stay.get("agencyGst") or stay.get("agency_gst") or "",
                            stay.get("agencyAddress") or stay.get("agency_address") or "",
                        )
                except Exception:
                    pass
                conn.commit()
                return jsonify({"ok": True, "room": room})

            if action == "checkin" or (
                action
                not in (
                    "reserve",
                    "checkout",
                    "transfer",
                    "generate_invoice",
                    "record_payment",
                    "set_discount",
                    "add_custom_charge",
                    "update_charge",
                    "delete_charge",
                    "merge_rooms",
                    "unmerge_rooms",
                    "set_merge_primary",
                )
                and isinstance(data.get("stay"), dict)
            ):
                stay = data.get("stay") or {}
                first = (stay.get("firstName") or stay.get("first_name") or "").strip()
                last = (stay.get("lastName") or stay.get("last_name") or "").strip()
                mobile = (stay.get("mobile") or "").strip()
                if not first or not last:
                    return jsonify({"ok": False, "error": "First name and last name are required."}), 400
                if not mobile:
                    return jsonify({"ok": False, "error": "Mobile number is required."}), 400
                if not (stay.get("checkInDate") or stay.get("check_in_date")):
                    return jsonify({"ok": False, "error": "Check-in date is required."}), 400
                check_in_raw = (
                    stay.get("checkInDate") or stay.get("check_in_date") or ""
                ).strip()
                try:
                    check_in_day = date.fromisoformat(check_in_raw[:10])
                except ValueError:
                    return jsonify({"ok": False, "error": "Check-in date is invalid."}), 400
                if check_in_day > date.today():
                    return jsonify(
                        {"ok": False, "error": "Future date check-in is not allowed."}
                    ), 400
                room = save_hotel_room_checkin(conn, room_id, stay, status="occupied")
                try:
                    ensure_customers_schema(conn)
                    guest_name = " ".join(p for p in (first, last) if p).strip() or first
                    upsert_customer(
                        conn,
                        guest_name,
                        mobile,
                        stay.get("address") or "",
                        stay.get("email") or "",
                    )
                except Exception:
                    pass
                try:
                    agency_name = (stay.get("agencyName") or stay.get("agency_name") or "").strip()
                    if agency_name:
                        upsert_agency_by_name(
                            conn,
                            agency_name,
                            stay.get("agencyGst") or stay.get("agency_gst") or "",
                            stay.get("agencyAddress") or stay.get("agency_address") or "",
                        )
                except Exception:
                    pass
                conn.commit()
                return jsonify({"ok": True, "room": room})

            if action == "checkout":
                clear_hotel_room_stay(conn, room_id, status="dirty")
                room = get_hotel_room(conn, room_id)
                conn.commit()
                return jsonify({"ok": True, "room": room})

            if action == "set_discount":
                result = set_hotel_room_discount(
                    conn,
                    room_id,
                    discount_type=data.get("discountType")
                    or data.get("discount_type")
                    or "pct",
                    discount_value=data.get("discountValue")
                    if data.get("discountValue") is not None
                    else data.get("discount_value"),
                    discount_reason=data.get("discountReason")
                    or data.get("discount_reason")
                    or "",
                )
                conn.commit()
                return jsonify({"ok": True, "room": result["room"]})

            if action == "add_custom_charge":
                existing = get_hotel_room(conn, room_id)
                stay = (
                    existing.get("stay")
                    if existing and isinstance(existing.get("stay"), dict)
                    else None
                )
                if stay and stay.get("invoiceGenerated") and stay.get("invoiceNumber"):
                    return (
                        jsonify(
                            {
                                "ok": False,
                                "error": "Custom charges cannot be added after the invoice is generated.",
                            }
                        ),
                        400,
                    )
                label = (
                    data.get("label")
                    or data.get("name")
                    or data.get("chargeName")
                    or data.get("charge_name")
                    or ""
                )
                amount = (
                    data.get("amount")
                    if data.get("amount") is not None
                    else data.get("rate")
                )
                result = append_hotel_room_folio_charge(
                    conn,
                    room_id,
                    amount=amount,
                    kind="other",
                    label=label,
                    source="hotel_invoice",
                    note=data.get("note") or data.get("notes") or "",
                )
                conn.commit()
                return jsonify(
                    {
                        "ok": True,
                        "room": result.get("room"),
                        "charge": result.get("charge"),
                    }
                )

            if action == "update_charge":
                result = update_hotel_room_charge(
                    conn,
                    room_id,
                    charge_key=data.get("chargeKey")
                    or data.get("charge_key")
                    or data.get("key")
                    or "",
                    label=data.get("label") or data.get("name") or "",
                    amount=data.get("amount"),
                    rate=data.get("rate"),
                )
                conn.commit()
                return jsonify({"ok": True, "room": result.get("room")})

            if action == "delete_charge":
                result = delete_hotel_room_charge(
                    conn,
                    room_id,
                    charge_key=data.get("chargeKey")
                    or data.get("charge_key")
                    or data.get("key")
                    or "",
                )
                conn.commit()
                return jsonify({"ok": True, "room": result.get("room")})

            if action == "generate_invoice":
                payment = data.get("payment") if isinstance(data.get("payment"), dict) else None
                payment_splits = data.get("payment_splits") or data.get("paymentSplits")
                note = data.get("note") or data.get("notes") or ""
                # Allow amount/method at top level for simpler clients.
                if payment is None and payment_splits is None and (
                    data.get("amount") is not None
                    or data.get("method")
                    or data.get("paymentMethod")
                ):
                    payment = {
                        "amount": data.get("amount"),
                        "method": data.get("method") or data.get("paymentMethod"),
                        "reference": data.get("reference")
                        or data.get("paymentReference")
                        or data.get("payment_reference"),
                        "note": note,
                    }
                result = generate_hotel_room_invoice(
                    conn,
                    room_id,
                    payment=payment,
                    payment_splits=payment_splits if isinstance(payment_splits, list) else None,
                    note=note,
                )
                conn.commit()
                return jsonify(
                    {
                        "ok": True,
                        "room": result["room"],
                        "minted": result.get("minted"),
                        "payment": result.get("payment"),
                        "payments": result.get("payments") or [],
                    }
                )

            if action == "record_payment":
                payment = data.get("payment") if isinstance(data.get("payment"), dict) else None
                payment_splits = data.get("payment_splits") or data.get("paymentSplits")
                note = data.get("note") or data.get("notes") or ""
                if payment is None and payment_splits is None:
                    payment = {
                        "amount": data.get("amount"),
                        "method": data.get("method") or data.get("paymentMethod"),
                        "reference": data.get("reference")
                        or data.get("paymentReference")
                        or data.get("payment_reference"),
                        "note": note,
                    }
                result = record_hotel_room_payment(
                    conn,
                    room_id,
                    payment=payment,
                    payment_splits=payment_splits if isinstance(payment_splits, list) else None,
                    note=note,
                )
                conn.commit()
                return jsonify(
                    {
                        "ok": True,
                        "room": result["room"],
                        "payment": result.get("payment"),
                        "payments": result.get("payments") or [],
                    }
                )

            if action == "transfer":
                to_room_id = (
                    data.get("toRoomId")
                    or data.get("to_room_id")
                    or data.get("destinationRoomId")
                    or ""
                )
                note = data.get("note") or data.get("reason") or ""
                result = transfer_hotel_room_stay(conn, room_id, to_room_id, note=note)
                conn.commit()
                return jsonify(
                    {
                        "ok": True,
                        "fromRoom": result["fromRoom"],
                        "toRoom": result["toRoom"],
                        "room": result["toRoom"],
                    }
                )

            if action == "merge_rooms":
                from_room_id = (
                    data.get("fromRoomId")
                    or data.get("from_room_id")
                    or room_id
                )
                to_room_id = (
                    data.get("toRoomId")
                    or data.get("to_room_id")
                    or data.get("primaryRoomId")
                    or ""
                )
                note = data.get("note") or data.get("reason") or ""
                result = merge_hotel_room_billing(
                    conn, from_room_id, to_room_id, note=note
                )
                conn.commit()
                return jsonify(
                    {
                        "ok": True,
                        "room": result.get("room"),
                        "primaryRoom": result.get("primaryRoom"),
                        "memberRoom": result.get("memberRoom"),
                    }
                )

            if action == "unmerge_rooms":
                scope = data.get("scope") or "one"
                result = unmerge_hotel_rooms(conn, room_id, scope=scope)
                conn.commit()
                return jsonify({"ok": True, "room": result.get("room")})

            if action == "set_merge_primary":
                result = set_hotel_merge_primary(conn, room_id)
                conn.commit()
                return jsonify({"ok": True, "room": result.get("room")})

            if "status" not in data:
                return jsonify({"ok": False, "error": "status is required."}), 400
            update_hotel_room_status(conn, room_id, data.get("status"), data)
            room = get_hotel_room(conn, room_id)
            conn.commit()
        except ValueError as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": str(exc)}), 400
        return jsonify({"ok": True, "room": room})
    finally:
        conn.close()


@app.route("/point-of-sale/invoice", endpoint="point_of_sale_invoice")
@app.route("/bar-point-of-sale/invoice", endpoint="bar_point_of_sale_invoice")
def point_of_sale_invoice():
    """POS Invoice workspace."""
    outlet = _pos_outlet_from_request()
    conn = get_db()
    try:
        tip_employees = _active_employees_for_tips(conn)
    finally:
        conn.close()
    tip_employee_options = [("", "Select employee…")] + [
        (
            str(emp["id"]),
            (
                f'{emp["name"]} ({emp["emp_code"]})'
                if emp.get("emp_code")
                else emp["name"]
            ),
        )
        for emp in tip_employees
    ]
    return render_template(
        "point_of_sale_invoice.html",
        tip_employees=tip_employees,
        tip_employee_options=tip_employee_options,
        tip_company=DEFAULT_COMPANY,
        tip_location=_pos_tip_location(outlet),
        tip_add_url=url_for("sales_update_add_tip"),
        tip_edit_url=url_for("sales_update_edit_tip"),
        tip_delete_url=url_for("sales_update_delete_tip"),
        **_pos_page_context(outlet, "invoice"),
    )


@app.route("/point-of-sale/settings", endpoint="point_of_sale_settings")
@app.route("/bar-point-of-sale/settings", endpoint="bar_point_of_sale_settings")
def point_of_sale_settings():
    """Restaurant/Bar Settings — floor layout and outlet configuration."""
    outlet = _pos_outlet_from_request()
    return render_template(
        "point_of_sale_settings.html",
        **_pos_page_context(outlet, "settings"),
    )


@app.route("/point-of-sale/menu", endpoint="point_of_sale_menu")
@app.route("/bar-point-of-sale/menu", endpoint="bar_point_of_sale_menu")
def point_of_sale_menu():
    """Menu catalog — categories and items (Product Master linked)."""
    outlet = _pos_outlet_from_request()
    return render_template(
        "point_of_sale_menu.html",
        **_pos_page_context(outlet, "menu"),
    )


def _pos_invoice_ledger_filters(args):
    """Parse invoice ledger GET filters (shared by page + export)."""
    today = date.today()
    date_from, date_to, date_filter_active = _resolve_optional_filter_date_range(
        args, "date_from", "date_to", default_fy=True
    )
    query_date_from = date_from if date_filter_active else date(2000, 1, 1)
    query_date_to = date_to if date_filter_active else today

    selected_order_type = (args.get("order_type") or "all").strip().lower()
    if selected_order_type not in ("all",) and selected_order_type not in POS_INVOICE_ORDER_TYPE_LABELS:
        selected_order_type = "all"
    order_type_filter = None if selected_order_type == "all" else selected_order_type

    selected_settlement = (args.get("settlement") or "all").strip().lower()
    if selected_settlement not in ("all",) and selected_settlement not in POS_INVOICE_SETTLEMENT_STATUS_LABELS:
        selected_settlement = "all"
    settlement_filter = None if selected_settlement == "all" else selected_settlement

    return {
        "today": today,
        "date_from": date_from,
        "date_to": date_to,
        "date_filter_active": date_filter_active,
        "query_date_from": query_date_from,
        "query_date_to": query_date_to,
        "selected_order_type": selected_order_type,
        "order_type_filter": order_type_filter,
        "selected_settlement": selected_settlement,
        "settlement_filter": settlement_filter,
    }


@app.route("/point-of-sale/invoice-ledger", endpoint="point_of_sale_invoice_ledger")
@app.route("/bar-point-of-sale/invoice-ledger", endpoint="bar_point_of_sale_invoice_ledger")
def point_of_sale_invoice_ledger():
    """POS Invoice Ledger — saved invoices with KPIs (Expense Ledger–style)."""
    outlet = _pos_outlet_from_request()
    filters = _pos_invoice_ledger_filters(request.args)

    user = get_current_user()
    can_cancel = user_can_edit_kot_sent_lines(user)

    conn = get_db()
    try:
        ensure_pos_schema(conn)
        invoices = list_pos_invoices(
            conn,
            date_from=filters["query_date_from"].isoformat(),
            date_to=filters["query_date_to"].isoformat(),
            order_type=filters["order_type_filter"],
            settlement=filters["settlement_filter"],
            outlet=outlet,
        )
        kpis = pos_invoice_kpis(conn, invoices, today=filters["today"].isoformat())
    finally:
        conn.close()

    for inv in invoices:
        status_key = str(inv.get("status") or "open").strip().lower() or "open"
        is_settled = (
            status_key == "closed"
            or bool(inv.get("payment_modes"))
            or bool(str(inv.get("settled_at") or "").strip())
        )
        is_cancelled = status_key == "cancelled"
        is_generated = bool(inv.get("customer_bill_sent")) or not is_provisional_pos_order_no(
            inv.get("order_no"), outlet
        )
        inv["ledger_is_settled"] = is_settled
        inv["ledger_is_cancelled"] = is_cancelled
        inv["ledger_is_generated"] = is_generated
        inv["ledger_can_delete"] = (
            can_cancel and (not is_settled) and (not is_cancelled) and (not is_generated)
        )
        inv["ledger_can_cancel"] = (
            can_cancel and (not is_settled) and (not is_cancelled) and is_generated
        )

    selected_order_type = filters["selected_order_type"]
    selected_order_type_label = "All"
    if selected_order_type != "all":
        selected_order_type_label = POS_INVOICE_ORDER_TYPE_LABELS.get(
            selected_order_type, selected_order_type
        )

    selected_settlement = filters["selected_settlement"]
    selected_settlement_label = "All"
    if selected_settlement != "all":
        selected_settlement_label = POS_INVOICE_SETTLEMENT_STATUS_LABELS.get(
            selected_settlement, selected_settlement
        )

    clear_kwargs = {}
    if selected_order_type != "all":
        clear_kwargs["order_type"] = selected_order_type
    if selected_settlement != "all":
        clear_kwargs["settlement"] = selected_settlement

    report_kwargs = {
        "order_type": selected_order_type,
        "settlement": selected_settlement,
    }
    if filters["date_filter_active"]:
        report_kwargs["date_from"] = filters["date_from"].isoformat()
        report_kwargs["date_to"] = filters["date_to"].isoformat()

    ledger_ep = _pos_endpoint("point_of_sale_invoice_ledger", outlet)
    report_ep = _pos_endpoint("export_pos_invoice_ledger_report", outlet)
    invoice_ep = _pos_endpoint("point_of_sale_invoice", outlet)

    return render_template(
        "point_of_sale_invoice_ledger.html",
        page_title="Invoice Ledger",
        invoices=invoices,
        kpis=kpis,
        can_cancel_invoices=can_cancel,
        order_types=POS_INVOICE_ORDER_TYPES,
        selected_order_type=selected_order_type,
        selected_order_type_label=selected_order_type_label,
        settlement_statuses=POS_INVOICE_SETTLEMENT_STATUSES,
        selected_settlement=selected_settlement,
        selected_settlement_label=selected_settlement_label,
        date_from=filters["date_from"].isoformat() if filters["date_from"] else "",
        date_to=filters["date_to"].isoformat() if filters["date_to"] else "",
        today_iso=filters["today"].isoformat(),
        active_date_filter=filters["date_filter_active"],
        filter_form_action=url_for(ledger_ep),
        invoice_ledger_clear_url=url_for(ledger_ep, **clear_kwargs),
        invoice_ledger_report_url=url_for(report_ep, **report_kwargs),
        new_invoice_url=url_for(invoice_ep),
        **_pos_page_context(outlet, "invoice_ledger"),
    )


@app.route("/point-of-sale/invoice-ledger/report", endpoint="export_pos_invoice_ledger_report")
@app.route("/bar-point-of-sale/invoice-ledger/report", endpoint="bar_export_pos_invoice_ledger_report")
def export_pos_invoice_ledger_report():
    """Excel download of saved POS invoices for the selected filters."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    outlet = _pos_outlet_from_request()
    filters = _pos_invoice_ledger_filters(request.args)

    conn = get_db()
    try:
        ensure_pos_schema(conn)
        invoices = list_pos_invoices(
            conn,
            date_from=filters["query_date_from"].isoformat(),
            date_to=filters["query_date_to"].isoformat(),
            order_type=filters["order_type_filter"],
            settlement=filters["settlement_filter"],
            outlet=outlet,
        )
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Ledger"
    header_font = Font(bold=True)
    headers = [
        "Order No",
        "Date",
        "Saved At",
        "Customer",
        "Mobile",
        "Order Type",
        "Payment Mode",
        "Captain",
        "Items",
        "Subtotal",
        "Discount",
        "GST",
        "Service",
        "Tip",
        "Total",
    ]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font

    for idx, inv in enumerate(invoices, start=2):
        ws.cell(row=idx, column=1, value=inv.get("order_no") or "")
        ws.cell(row=idx, column=2, value=inv.get("order_date") or "")
        ws.cell(row=idx, column=3, value=inv.get("saved_at") or "")
        ws.cell(row=idx, column=4, value=inv.get("customer_name") or "")
        ws.cell(row=idx, column=5, value=inv.get("customer_mobile") or "")
        ws.cell(
            row=idx,
            column=6,
            value=inv.get("order_type_label") or inv.get("order_type") or "",
        )
        ws.cell(row=idx, column=7, value=inv.get("payment_mode_label") or "Unsettled")
        ws.cell(row=idx, column=8, value=inv.get("captain") or "")
        ws.cell(row=idx, column=9, value=int(inv.get("item_count") or 0))
        ws.cell(row=idx, column=10, value=round_half_up(inv.get("subtotal"), 2))
        ws.cell(row=idx, column=11, value=round_half_up(inv.get("discount"), 2))
        ws.cell(row=idx, column=12, value=round_half_up(inv.get("gst"), 2))
        ws.cell(row=idx, column=13, value=round_half_up(inv.get("service"), 2))
        ws.cell(row=idx, column=14, value=round_half_up(inv.get("tip"), 2))
        ws.cell(row=idx, column=15, value=round_half_up(inv.get("grand_total"), 2))

    for column_cells in ws.columns:
        width = 12
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, 40))
        ws.column_dimensions[column_cells[0].column_letter].width = width

    if filters["date_filter_active"]:
        fname = (
            f"invoice_ledger_{filters['query_date_from'].isoformat()}_to_"
            f"{filters['query_date_to'].isoformat()}.xlsx"
        )
    else:
        fname = "invoice_ledger_all.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/point-of-sale/api/invoices", methods=["POST"], endpoint="point_of_sale_api_invoices_save")
@app.route("/bar-point-of-sale/api/invoices", methods=["POST"], endpoint="bar_point_of_sale_api_invoices_save")
def point_of_sale_api_invoices_save():
    """Create or update a POS invoice from the Create Invoice workspace."""
    outlet = _pos_outlet_from_request()
    data = request.get_json(silent=True) or {}
    if isinstance(data, dict):
        data = dict(data)
        data["outlet"] = outlet
    user = get_current_user()
    created_by = ""
    if user:
        created_by = (
            str(user.get("full_name") or user.get("username") or user.get("id") or "").strip()
        )
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        try:
            saved = save_pos_invoice(
                conn,
                data,
                created_by=created_by,
                allow_kot_cancel=user_can_edit_kot_sent_lines(user),
            )
            sync_pos_floor_occupancy_from_open_orders(conn, outlet)
            conn.commit()
        except ValueError as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": str(exc)}), 400
        except Exception as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": f"Could not save invoice: {exc}"}), 500
        return jsonify({"ok": True, "invoice": saved})
    finally:
        conn.close()


@app.route("/point-of-sale/api/invoices/<int:invoice_id>", methods=["GET"], endpoint="point_of_sale_api_invoice_detail")
@app.route("/bar-point-of-sale/api/invoices/<int:invoice_id>", methods=["GET"], endpoint="bar_point_of_sale_api_invoice_detail")
def point_of_sale_api_invoice_detail(invoice_id):
    """Return one saved POS invoice with line items."""
    outlet = _pos_outlet_from_request()
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        invoice = get_pos_invoice(conn, invoice_id)
        if not invoice or not _pos_invoice_belongs_to_outlet(invoice, outlet):
            return jsonify({"ok": False, "error": "Invoice not found."}), 404
        return jsonify({"ok": True, "invoice": invoice})
    finally:
        conn.close()


@app.route("/point-of-sale/api/invoices/by-table", methods=["GET"], endpoint="point_of_sale_api_invoice_by_table")
@app.route("/bar-point-of-sale/api/invoices/by-table", methods=["GET"], endpoint="bar_point_of_sale_api_invoice_by_table")
def point_of_sale_api_invoice_by_table():
    """Look up the open dine-in order for a table, if any — shared by the Tables
    page tile tap and the Create Invoice table picker to resume a bill instead
    of blocking on 'occupied'."""
    outlet = _pos_outlet_from_request()
    table = (request.args.get("table") or "").strip()
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        invoice = get_open_pos_invoice_for_table(conn, table, outlet) if table else None
        return jsonify({"ok": True, "invoice": invoice})
    finally:
        conn.close()


@app.route("/point-of-sale/api/invoices/transfer-table", methods=["POST"], endpoint="point_of_sale_api_invoice_transfer_table")
@app.route("/bar-point-of-sale/api/invoices/transfer-table", methods=["POST"], endpoint="bar_point_of_sale_api_invoice_transfer_table")
def point_of_sale_api_invoice_transfer_table():
    """Move an open dine-in bill from one table to another available table."""
    outlet = _pos_outlet_from_request()
    payload = request.get_json(silent=True) or {}
    from_table = (payload.get("from_table") or payload.get("fromTable") or "").strip()
    to_table = (payload.get("to_table") or payload.get("toTable") or "").strip()
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        try:
            invoice = transfer_pos_invoice_table(conn, from_table, to_table, outlet)
            sync_pos_floor_occupancy_from_open_orders(conn, outlet)
            layout = get_pos_floor_layout(conn, outlet)
            conn.commit()
        except ValueError as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": str(exc)}), 400
        return jsonify(
            {
                "ok": True,
                "invoice": invoice,
                "areas": layout.get("areas") or [],
                "tables": layout.get("tables") or [],
            }
        )
    finally:
        conn.close()


@app.route("/point-of-sale/api/invoices/merge-tables", methods=["POST"], endpoint="point_of_sale_api_invoice_merge_tables")
@app.route("/bar-point-of-sale/api/invoices/merge-tables", methods=["POST"], endpoint="bar_point_of_sale_api_invoice_merge_tables")
def point_of_sale_api_invoice_merge_tables():
    """Combine two open dine-in bills into one invoice on the destination table."""
    outlet = _pos_outlet_from_request()
    payload = request.get_json(silent=True) or {}
    from_table = (payload.get("from_table") or payload.get("fromTable") or "").strip()
    to_table = (payload.get("to_table") or payload.get("toTable") or "").strip()
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        try:
            invoice = merge_pos_invoice_tables(conn, from_table, to_table, outlet)
            sync_pos_floor_occupancy_from_open_orders(conn, outlet)
            layout = get_pos_floor_layout(conn, outlet)
            conn.commit()
        except ValueError as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": str(exc)}), 400
        floor = _pos_floor_api_payload(conn, layout)
        return jsonify(
            {
                "ok": True,
                "invoice": invoice,
                "areas": floor["areas"],
                "tables": floor["tables"],
            }
        )
    finally:
        conn.close()


@app.route("/point-of-sale/api/invoices/<int:invoice_id>/close", methods=["POST"], endpoint="point_of_sale_api_invoice_close")
@app.route("/bar-point-of-sale/api/invoices/<int:invoice_id>/close", methods=["POST"], endpoint="bar_point_of_sale_api_invoice_close")
def point_of_sale_api_invoice_close(invoice_id):
    """Close a bill and free its table — legacy path without payment details.
    Prefer /settle for staff Settle Bill (payment + optional split)."""
    outlet = _pos_outlet_from_request()
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        try:
            existing = get_pos_invoice(conn, invoice_id)
            if not existing or not _pos_invoice_belongs_to_outlet(existing, outlet):
                return jsonify({"ok": False, "error": "Invoice not found."}), 404
            user = get_current_user()
            user_id = user.get("id") if user else None
            invoice = close_pos_invoice_and_free_table(conn, invoice_id, user_id=user_id)
            conn.commit()
        except ValueError as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": str(exc)}), 400
        return jsonify({"ok": True, "invoice": invoice})
    finally:
        conn.close()


@app.route("/point-of-sale/api/invoices/<int:invoice_id>/settle", methods=["POST"], endpoint="point_of_sale_api_invoice_settle")
@app.route("/bar-point-of-sale/api/invoices/<int:invoice_id>/settle", methods=["POST"], endpoint="bar_point_of_sale_api_invoice_settle")
def point_of_sale_api_invoice_settle(invoice_id):
    """Settle a bill with payment mode(s) — same split rules as Room Transfer
    Clear Payment — then close the order and free the table when dine-in."""
    outlet = _pos_outlet_from_request()
    payload = request.get_json(silent=True) or {}
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        try:
            existing = get_pos_invoice(conn, invoice_id)
            if not existing or not _pos_invoice_belongs_to_outlet(existing, outlet):
                return jsonify({"ok": False, "error": "Invoice not found."}), 404
            user = get_current_user()
            user_id = user.get("id") if user else None
            invoice = settle_pos_invoice(
                conn,
                invoice_id,
                payment_splits=payload.get("payment_splits"),
                payment_date=payload.get("payment_date"),
                notes=payload.get("notes") or "",
                user_id=user_id,
                hotel_room_id=payload.get("hotel_room_id")
                or payload.get("hotelRoomId")
                or "",
            )
            conn.commit()
        except ValueError as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": str(exc)}), 400
        return jsonify({"ok": True, "invoice": invoice})
    finally:
        conn.close()


@app.route("/point-of-sale/api/invoices/<int:invoice_id>/send-kot", methods=["POST"], endpoint="point_of_sale_api_invoice_send_kot")
@app.route("/bar-point-of-sale/api/invoices/<int:invoice_id>/send-kot", methods=["POST"], endpoint="bar_point_of_sale_api_invoice_send_kot")
def point_of_sale_api_invoice_send_kot(invoice_id):
    """Send pending (unsent) lines for one open invoice to the kitchen."""
    outlet = _pos_outlet_from_request()
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        try:
            existing = get_pos_invoice(conn, invoice_id)
            if not existing or not _pos_invoice_belongs_to_outlet(existing, outlet):
                return jsonify({"ok": False, "error": "Invoice not found."}), 404
            invoice = send_pos_invoice_pending_kot(conn, invoice_id)
            kot_pending = list_pos_kot_pending_summary(conn, outlet)
            conn.commit()
        except ValueError as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": str(exc)}), 400
        return jsonify({"ok": True, "invoice": invoice, "kot_pending": kot_pending})
    finally:
        conn.close()


@app.route("/point-of-sale/api/kot-pending/send-all", methods=["POST"], endpoint="point_of_sale_api_kot_pending_send_all")
@app.route("/bar-point-of-sale/api/kot-pending/send-all", methods=["POST"], endpoint="bar_point_of_sale_api_kot_pending_send_all")
def point_of_sale_api_kot_pending_send_all():
    """Send pending KOT lines for every open dine-in table listed in the summary."""
    outlet = _pos_outlet_from_request()
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        summary = list_pos_kot_pending_summary(conn, outlet)
        sent = []
        errors = []
        for entry in summary.get("tables") or []:
            invoice_id = entry.get("invoice_id")
            try:
                invoice = send_pos_invoice_pending_kot(conn, invoice_id)
                sent.append({"invoice_id": invoice_id, "table": entry.get("name") or "", "order_no": invoice.get("order_no")})
            except ValueError as exc:
                errors.append({"invoice_id": invoice_id, "error": str(exc)})
        kot_pending = list_pos_kot_pending_summary(conn, outlet)
        conn.commit()
        return jsonify(
            {
                "ok": True,
                "sent_count": len(sent),
                "sent": sent,
                "errors": errors,
                "kot_pending": kot_pending,
            }
        )
    finally:
        conn.close()


@app.route("/point-of-sale/api/kot-tokens", methods=["GET"], endpoint="point_of_sale_api_kot_tokens")
@app.route("/bar-point-of-sale/api/kot-tokens", methods=["GET"], endpoint="bar_point_of_sale_api_kot_tokens")
def point_of_sale_api_kot_tokens():
    """List open dine-in KOTs already sent to kitchen (for resend / reprint)."""
    outlet = _pos_outlet_from_request()
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        payload = list_pos_kot_tokens(conn, outlet)
        conn.commit()
        return jsonify({"ok": True, **payload})
    finally:
        conn.close()


@app.route(
    "/point-of-sale/api/kot-tokens/reduce",
    methods=["POST"],
    endpoint="point_of_sale_api_kot_tokens_reduce",
)
@app.route(
    "/bar-point-of-sale/api/kot-tokens/reduce",
    methods=["POST"],
    endpoint="bar_point_of_sale_api_kot_tokens_reduce",
)
def point_of_sale_api_kot_tokens_reduce():
    """Persist kitchen-sent qty reductions from the Tables KOT hub (Cancellation Access)."""
    outlet = _pos_outlet_from_request()
    user = get_current_user()
    if not user_can_edit_kot_sent_lines(user):
        return jsonify(
            {"ok": False, "error": "Cancellation Access is required to reduce kitchen-sent items."}
        ), 403
    data = request.get_json(silent=True) or {}
    changes = data.get("changes") if isinstance(data, dict) else None
    created_by = ""
    if user:
        created_by = str(
            user.get("full_name") or user.get("username") or user.get("id") or ""
        ).strip()
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        try:
            result = apply_pos_kot_token_reductions(
                conn,
                changes or [],
                allow_kot_cancel=True,
                created_by=created_by,
            )
            # Drop any invoice that no longer belongs to this outlet from the response.
            invoices = [
                inv
                for inv in (result.get("invoices") or [])
                if _pos_invoice_belongs_to_outlet(inv, outlet)
            ]
            sync_pos_floor_occupancy_from_open_orders(conn, outlet)
            conn.commit()
        except ValueError as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": str(exc)}), 400
        except Exception as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": f"Could not save KOT changes: {exc}"}), 500
        tokens = list_pos_kot_tokens(conn, outlet)
        cancelled_count = sum(1 for inv in invoices if inv.get("cancelled"))
        return jsonify(
            {
                "ok": True,
                "updated_count": len(invoices),
                "cancelled_count": cancelled_count,
                "invoices": invoices,
                **tokens,
            }
        )
    finally:
        conn.close()


@app.route("/point-of-sale/api/today-invoices", methods=["GET"], endpoint="point_of_sale_api_today_invoices")
@app.route("/bar-point-of-sale/api/today-invoices", methods=["GET"], endpoint="bar_point_of_sale_api_today_invoices")
def point_of_sale_api_today_invoices():
    """List today's active POS invoices for the Tables Invoice hub (newest first)."""
    outlet = _pos_outlet_from_request()
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        payload = list_pos_today_invoices(conn, outlet=outlet)
        conn.commit()
        return jsonify({"ok": True, **payload})
    finally:
        conn.close()


@app.route("/point-of-sale/api/invoices/<int:invoice_id>/reopen-edit", methods=["POST"], endpoint="point_of_sale_api_invoice_reopen_edit")
@app.route("/bar-point-of-sale/api/invoices/<int:invoice_id>/reopen-edit", methods=["POST"], endpoint="bar_point_of_sale_api_invoice_reopen_edit")
def point_of_sale_api_invoice_reopen_edit(invoice_id):
    """Unlock an unsettled generated invoice for editing (Cancellation Access)."""
    outlet = _pos_outlet_from_request()
    user = get_current_user()
    if not user_can_edit_kot_sent_lines(user):
        return jsonify(
            {"ok": False, "error": "Cancellation Access is required to edit unsettled invoices."}
        ), 403
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        try:
            existing = get_pos_invoice(conn, invoice_id)
            if not existing or not _pos_invoice_belongs_to_outlet(existing, outlet):
                return jsonify({"ok": False, "error": "Invoice not found."}), 404
            invoice = reopen_pos_invoice_for_edit(conn, invoice_id)
            sync_pos_floor_occupancy_from_open_orders(conn, outlet)
            conn.commit()
        except ValueError as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": str(exc)}), 400
        return jsonify({"ok": True, "invoice": invoice})
    finally:
        conn.close()


@app.route("/point-of-sale/api/invoices/<int:invoice_id>/delete", methods=["POST", "DELETE"], endpoint="point_of_sale_api_invoice_delete")
@app.route("/bar-point-of-sale/api/invoices/<int:invoice_id>/delete", methods=["POST", "DELETE"], endpoint="bar_point_of_sale_api_invoice_delete")
def point_of_sale_api_invoice_delete(invoice_id):
    """Cancel an unsettled POS invoice (Cancellation Access).

    Issued official numbers are kept as status=cancelled; provisional drafts
    are soft-deleted. Requires a cancellation reason.
    """
    outlet = _pos_outlet_from_request()
    user = get_current_user()
    if not user_can_edit_kot_sent_lines(user):
        return jsonify(
            {"ok": False, "error": "Cancellation Access is required to cancel invoices."}
        ), 403
    data = request.get_json(silent=True) or {}
    reason = ""
    if isinstance(data, dict):
        reason = data.get("reason") or data.get("cancelReason") or data.get("cancel_reason") or ""
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        try:
            existing = get_pos_invoice(conn, invoice_id)
            if not existing or not _pos_invoice_belongs_to_outlet(existing, outlet):
                return jsonify({"ok": False, "error": "Invoice not found."}), 404
            result = cancel_pos_invoice(conn, invoice_id, reason=reason)
            sync_pos_floor_occupancy_from_open_orders(conn, outlet)
            conn.commit()
        except ValueError as exc:
            conn.rollback()
            msg = str(exc)
            code = 404 if "not found" in msg.lower() else 400
            return jsonify({"ok": False, "error": msg}), code
        return jsonify(
            {
                "ok": True,
                "mode": result.get("mode"),
                "invoice": result.get("invoice"),
            }
        )
    finally:
        conn.close()


@app.route("/point-of-sale/api/customers", methods=["GET"], endpoint="point_of_sale_api_customers")
@app.route("/bar-point-of-sale/api/customers", methods=["GET"], endpoint="bar_point_of_sale_api_customers")
def point_of_sale_api_customers():
    """Search Customer Master for POS Customer Details autocomplete."""
    query = request.args.get("q") or request.args.get("mobile") or ""
    conn = get_db()
    try:
        ensure_customers_schema(conn)
        customers = search_customers(conn, query, limit=8)
        conn.commit()
        return jsonify({"ok": True, "customers": customers})
    finally:
        conn.close()


@app.route(
    "/point-of-sale/api/hotel-rooms/occupied",
    methods=["GET"],
    endpoint="point_of_sale_api_hotel_rooms_occupied",
)
@app.route(
    "/bar-point-of-sale/api/hotel-rooms/occupied",
    methods=["GET"],
    endpoint="bar_point_of_sale_api_hotel_rooms_occupied",
)
def point_of_sale_api_hotel_rooms_occupied():
    """Occupied hotel rooms for POS Room Transfer settle (folio post)."""
    conn = get_db()
    try:
        ensure_hotel_rooms_schema(conn)
        layout = get_hotel_rooms_layout(conn)
        rooms = []
        for room in layout.get("rooms") or []:
            status = str(room.get("status") or "").strip().lower()
            stay = room.get("stay")
            if status != "occupied" or not isinstance(stay, dict) or not stay:
                continue
            guest = " ".join(
                p
                for p in (
                    str(stay.get("firstName") or "").strip(),
                    str(stay.get("lastName") or "").strip(),
                )
                if p
            ).strip() or str(stay.get("guestName") or "").strip() or "Guest"
            rooms.append(
                {
                    "id": room.get("id"),
                    "number": room.get("number"),
                    "guestName": guest,
                    "roomType": room.get("roomType") or room.get("room_type") or "",
                }
            )
        rooms.sort(key=lambda item: str(item.get("number") or ""))
        conn.commit()
        return jsonify({"ok": True, "rooms": rooms})
    finally:
        conn.close()


def _pos_floor_api_payload(conn, layout=None, outlet=POS_OUTLET_RESTAURANT):
    """Floor JSON for API responses with merged-table display helpers."""
    layout = layout or get_pos_floor_layout(conn, outlet)
    tables = enrich_pos_floor_tables_for_display(layout.get("tables") or [])
    return {
        "areas": layout.get("areas") or [],
        "tables": tables,
    }


@app.route("/point-of-sale/api/floor", methods=["GET", "PUT", "POST"], endpoint="point_of_sale_api_floor")
@app.route("/bar-point-of-sale/api/floor", methods=["GET", "PUT", "POST"], endpoint="bar_point_of_sale_api_floor")
def point_of_sale_api_floor():
    """Load or replace outlet floor layout (areas + tables) in SQLite."""
    outlet = _pos_outlet_from_request()
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        if request.method == "GET":
            sync_pos_floor_occupancy_from_open_orders(conn, outlet)
            payload = _pos_floor_api_payload(conn, outlet=outlet)
            kot_pending = list_pos_kot_pending_summary(conn, outlet)
            sales = pos_today_sales_summary(conn, outlet=outlet)
            conn.commit()
            resp = jsonify({
                "ok": True,
                **payload,
                "kot_pending": kot_pending,
                "sales_total": sales["sales_total"],
                "sales_count": sales["sales_count"],
                "unsettled_count": sales["unsettled_count"],
                "unsettled_total": sales["unsettled_total"],
            })
            resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
            resp.headers["Pragma"] = "no-cache"
            return resp

        data = request.get_json(silent=True) or {}
        areas = data.get("areas")
        tables = data.get("tables")
        if not isinstance(areas, list) or not isinstance(tables, list):
            return jsonify({"ok": False, "error": "areas and tables arrays are required."}), 400
        saved = save_pos_floor_layout(conn, areas, tables, outlet)
        conn.commit()
        return jsonify({"ok": True, **_pos_floor_api_payload(conn, saved, outlet)})
    finally:
        conn.close()


@app.route("/point-of-sale/api/floor/unmerge-tables", methods=["POST"], endpoint="point_of_sale_api_floor_unmerge_tables")
@app.route("/bar-point-of-sale/api/floor/unmerge-tables", methods=["POST"], endpoint="bar_point_of_sale_api_floor_unmerge_tables")
def point_of_sale_api_floor_unmerge_tables():
    """Split a visually merged table group back into separate floor tiles."""
    outlet = _pos_outlet_from_request()
    payload = request.get_json(silent=True) or {}
    table = (payload.get("table") or payload.get("table_label") or payload.get("from_table") or "").strip()
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        try:
            result = unmerge_pos_floor_tables(conn, table, outlet)
            sync_pos_floor_occupancy_from_open_orders(conn, outlet)
            floor = _pos_floor_api_payload(conn, result.get("layout"), outlet)
            conn.commit()
        except ValueError as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": str(exc)}), 400
        return jsonify(
            {
                "ok": True,
                "primary_name": result.get("primary_name"),
                "label": result.get("label"),
                "group_names": result.get("group_names") or [],
                "areas": floor["areas"],
                "tables": floor["tables"],
            }
        )
    finally:
        conn.close()


@app.route("/point-of-sale/api/floor/clear-all", methods=["POST"], endpoint="point_of_sale_api_floor_clear_all")
@app.route("/bar-point-of-sale/api/floor/clear-all", methods=["POST"], endpoint="bar_point_of_sale_api_floor_clear_all")
def point_of_sale_api_floor_clear_all():
    """Free every table back to available in one shot (Tables page 'Clear all
    tables' — day close / demo reset). Also closes any dangling open bills."""
    outlet = _pos_outlet_from_request()
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        user = get_current_user()
        user_id = user.get("id") if user else None
        payload = clear_all_pos_tables(conn, user_id=user_id, outlet=outlet)
        kot_pending = list_pos_kot_pending_summary(conn, outlet)
        conn.commit()
        floor = _pos_floor_api_payload(conn, payload, outlet)
        return jsonify({"ok": True, **floor, "kot_pending": kot_pending})
    finally:
        conn.close()


@app.route("/point-of-sale/api/settings", methods=["GET", "PUT", "POST"], endpoint="point_of_sale_api_settings")
@app.route("/bar-point-of-sale/api/settings", methods=["GET", "PUT", "POST"], endpoint="bar_point_of_sale_api_settings")
def point_of_sale_api_settings():
    """Load or replace outlet settings JSON blob in SQLite."""
    outlet = _pos_outlet_from_request()
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        if request.method == "GET":
            settings = get_pos_restaurant_settings(conn, outlet)
            rates = get_pos_tax_rates(conn, outlet)
            return jsonify({"ok": True, "settings": settings, "taxRates": rates})

        data = request.get_json(silent=True) or {}
        if "settings" in data:
            settings = data.get("settings")
        else:
            settings = data
        if not isinstance(settings, dict):
            return jsonify({"ok": False, "error": "settings object is required."}), 400
        saved = save_pos_restaurant_settings(conn, settings, outlet)
        conn.commit()
        rates = get_pos_tax_rates(conn, outlet)
        return jsonify({"ok": True, "settings": saved, "taxRates": rates})
    finally:
        conn.close()


@app.route("/point-of-sale/api/menu/categories", methods=["GET", "POST", "PUT"], endpoint="point_of_sale_api_menu_categories")
@app.route("/bar-point-of-sale/api/menu/categories", methods=["GET", "POST", "PUT"], endpoint="bar_point_of_sale_api_menu_categories")
def point_of_sale_api_menu_categories():
    """List or create/update outlet menu categories."""
    outlet = _pos_outlet_from_request()
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        if request.method == "GET":
            categories = list_pos_menu_categories(
                conn, outlets=_pos_menu_list_outlets(outlet)
            )
            conn.commit()
            return jsonify({"ok": True, "categories": categories})

        data = request.get_json(silent=True) or {}
        category_id = data.get("id")
        try:
            category_id = int(category_id) if category_id not in (None, "") else None
        except (TypeError, ValueError):
            return jsonify({"ok": False, "error": "Invalid category id."}), 400
        try:
            saved = save_pos_menu_category(
                conn,
                category_id=category_id,
                name=data.get("name"),
                is_visible=bool(data.get("is_visible", True)),
                sort_order=data.get("sort_order"),
                outlet=outlet,
            )
            conn.commit()
        except ValueError as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": str(exc)}), 400
        return jsonify({"ok": True, "category": saved, "categories": list_pos_menu_categories(conn, outlet=outlet)})
    finally:
        conn.close()


@app.route("/point-of-sale/api/menu/categories/<int:category_id>/delete", methods=["POST", "DELETE"], endpoint="point_of_sale_api_menu_category_delete")
@app.route("/bar-point-of-sale/api/menu/categories/<int:category_id>/delete", methods=["POST", "DELETE"], endpoint="bar_point_of_sale_api_menu_category_delete")
def point_of_sale_api_menu_category_delete(category_id):
    """Soft-delete a menu category."""
    outlet = _pos_outlet_from_request()
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        try:
            soft_delete_pos_menu_category(conn, category_id)
            conn.commit()
        except ValueError as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": str(exc)}), 404
        return jsonify({"ok": True, "categories": list_pos_menu_categories(conn, outlet=outlet)})
    finally:
        conn.close()


@app.route("/point-of-sale/api/menu/items", methods=["GET", "POST", "PUT"], endpoint="point_of_sale_api_menu_items")
@app.route("/bar-point-of-sale/api/menu/items", methods=["GET", "POST", "PUT"], endpoint="bar_point_of_sale_api_menu_items")
def point_of_sale_api_menu_items():
    """List or create/update outlet menu items with optional recipe ingredients."""
    outlet = _pos_outlet_from_request()
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        ensure_stores_schema(conn)
        if request.method == "GET":
            category_id = request.args.get("category_id")
            try:
                category_id = int(category_id) if category_id not in (None, "") else None
            except (TypeError, ValueError):
                return jsonify({"ok": False, "error": "Invalid category id."}), 400
            items = list_pos_menu_items(
                conn,
                category_id=category_id,
                outlets=_pos_menu_list_outlets(outlet),
            )
            conn.commit()
            return jsonify({"ok": True, "items": items, "category_id": category_id})

        data = request.get_json(silent=True) or {}
        item_id = data.get("id")
        try:
            item_id = int(item_id) if item_id not in (None, "") else None
        except (TypeError, ValueError):
            return jsonify({"ok": False, "error": "Invalid item id."}), 400
        user = get_current_user() or {}
        updated_by = (
            (user.get("full_name") or user.get("username") or "").strip()
            if isinstance(user, dict)
            else ""
        )
        try:
            saved = save_pos_menu_item(
                conn,
                item_id=item_id,
                category_id=data.get("category_id"),
                product_id=data.get("product_id"),
                name=data.get("name"),
                code=data.get("code"),
                barcode=data.get("barcode"),
                variant=data.get("variant"),
                rate=data.get("rate"),
                sort_order=data.get("sort_order"),
                recipe=data["recipe"] if "recipe" in data else None,
                menu_type=data["menu_type"] if "menu_type" in data else None,
                item_kind=data["item_kind"] if "item_kind" in data else None,
                portion_size=data["portion_size"] if "portion_size" in data else None,
                prep_time_mins=data["prep_time_mins"] if "prep_time_mins" in data else None,
                shelf_life=data["shelf_life"] if "shelf_life" in data else None,
                notes=data["notes"] if "notes" in data else None,
                target_margin_pct=data["target_margin_pct"] if "target_margin_pct" in data else None,
                updated_by=updated_by or None,
                price_change_reason=data.get("price_change_reason") or "",
                outlet=outlet,
            )
            conn.commit()
        except ValueError as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": str(exc)}), 400
        except Exception as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": f"Could not save menu item: {exc}"}), 500
        category_id = saved.get("category_id")
        return jsonify(
            {
                "ok": True,
                "item": saved,
                "items": list_pos_menu_items(conn, category_id=category_id, outlet=outlet),
                "categories": list_pos_menu_categories(conn, outlet=outlet),
            }
        )
    finally:
        conn.close()


@app.route("/point-of-sale/api/menu/items/<int:item_id>", methods=["GET"], endpoint="point_of_sale_api_menu_item_detail")
@app.route("/bar-point-of-sale/api/menu/items/<int:item_id>", methods=["GET"], endpoint="bar_point_of_sale_api_menu_item_detail")
def point_of_sale_api_menu_item_detail(item_id):
    """Menu Details popup payload: recipe, FIFO costing, price history, analysis."""
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        ensure_stores_schema(conn)
        detail = get_pos_menu_item_details(conn, item_id)
        if not detail:
            return jsonify({"ok": False, "error": "Menu item not found."}), 404
        conn.commit()
        return jsonify({"ok": True, "item": detail})
    finally:
        conn.close()


@app.route("/point-of-sale/api/menu/items/<int:item_id>/delete", methods=["POST", "DELETE"], endpoint="point_of_sale_api_menu_item_delete")
@app.route("/bar-point-of-sale/api/menu/items/<int:item_id>/delete", methods=["POST", "DELETE"], endpoint="bar_point_of_sale_api_menu_item_delete")
def point_of_sale_api_menu_item_delete(item_id):
    """Soft-delete a menu item."""
    conn = get_db()
    try:
        ensure_pos_schema(conn)
        try:
            soft_delete_pos_menu_item(conn, item_id)
            conn.commit()
        except ValueError as exc:
            conn.rollback()
            return jsonify({"ok": False, "error": str(exc)}), 404
        category_id = request.args.get("category_id") or (request.get_json(silent=True) or {}).get(
            "category_id"
        )
        try:
            category_id = int(category_id) if category_id not in (None, "") else None
        except (TypeError, ValueError):
            category_id = None
        outlet = _pos_outlet_from_request()
        payload = {
            "ok": True,
            "categories": list_pos_menu_categories(conn, outlet=outlet),
        }
        if category_id is not None:
            payload["items"] = list_pos_menu_items(conn, category_id=category_id, outlet=outlet)
            payload["category_id"] = category_id
        return jsonify(payload)
    finally:
        conn.close()


@app.route("/point-of-sale/api/menu/products", methods=["GET"], endpoint="point_of_sale_api_menu_products")
@app.route("/bar-point-of-sale/api/menu/products", methods=["GET"], endpoint="bar_point_of_sale_api_menu_products")
@app.route("/stores/api/products-lite", methods=["GET"])
def point_of_sale_api_menu_products():
    """Lite Product Master list for menu item pickers."""
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        q = (request.args.get("q") or "").strip()
        outlet = (request.args.get("outlet") or "").strip().lower()
        if not outlet and (request.path or "").startswith("/bar-point-of-sale"):
            outlet = "bar"
        elif not outlet and (request.path or "").startswith("/point-of-sale"):
            outlet = "restaurant"
        if outlet == "restaurant":
            outlets = ["restaurant", "both"]
        elif outlet == "bar":
            outlets = ["bar", "both"]
        elif outlet == "both":
            outlets = ["both"]
        elif outlet in ("all", "*"):
            outlets = None
        else:
            outlets = None
        products = list_store_products_lite(conn, outlets=outlets, q=q)
        return jsonify({"ok": True, "products": products})
    finally:
        conn.close()


def _home_notifications(user):
    """Build home-page bell items for modules the current user can access."""
    notifications = []
    if not user:
        return notifications
    if user_can_access_stores_submodule(user, "approvals"):
        conn = get_db()
        try:
            ensure_stores_schema(conn)
            pending_count = conn.execute(
                "SELECT COUNT(*) AS c FROM store_indents WHERE status = 'pending'"
            ).fetchone()["c"]
        finally:
            conn.close()
        pending_count = int(pending_count or 0)
        if pending_count > 0:
            label = "indent" if pending_count == 1 else "indents"
            notifications.append({
                "id": "stores-approvals-pending",
                "title": "Indents awaiting approval",
                "body": f"{pending_count} {label} waiting for your review.",
                "href": url_for("stores_approvals"),
            })
    if user_can_access_dashboard(user, "communication_hub"):
        from communication_hub import build_hub_home_notification, pull_hub_mirror_into

        conn = get_db()
        try:
            pull_hub_mirror_into(conn)
            hub_item = build_hub_home_notification(conn)
        finally:
            conn.close()
        if hub_item:
            hub_item["href"] = url_for("communication_hub")
            notifications.append(hub_item)
    return notifications


@app.route("/home/api/notifications")
def home_api_notifications():
    """JSON feed for the home bell (polled so new WhatsApp messages surface live)."""
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "Unauthorized"}), 401
    items = _home_notifications(user)
    return jsonify({"ok": True, "notifications": items, "unread": len(items) > 0})


@app.route("/dashboard")
def dashboard():
    return render_template(
        "dashboard.html",
        current_user=get_current_user(),
        de_nav_section="analytics",
        de_nav_sales_view="dashboard",
    )


def _resolve_cash_ledger_date_range(args):
    """Return (date_from, date_to, date_filter_active).

    With no date query params, default to the current Indian FY → today.
    """
    return _resolve_optional_filter_date_range(
        args, "date_from", "date_to", default_fy=True
    )


def _normalize_cash_ledger_location(location):
    value = (location or CASH_LEDGER_FILTER_ALL).strip()
    if value in CASH_LEDGER_FILTER_LOCATIONS:
        return value
    return CASH_LEDGER_FILTER_ALL


def _cash_ledger_outlet_scope(location=None):
    """Return outlet names to include for sales/expense rows."""
    normalized = _normalize_cash_ledger_location(location)
    if normalized == CASH_LEDGER_FILTER_ALL:
        return CASH_LEDGER_OUTLETS
    return (normalized,)


def _normalize_cash_ledger_transfer_destination(value):
    key = (value or "").strip().lower().replace(" ", "_").replace("-", "_")
    if key in CASH_LEDGER_TRANSFER_DESTINATION_LABELS:
        return key
    return ""


def _cash_ledger_sales_rows(conn, company, date_from, date_to, location=None):
    outlets = _cash_ledger_outlet_scope(location)
    placeholders = ",".join("?" for _ in outlets)
    rows = conn.execute(
        f"""SELECT id, location, sales_date, sales_entry_values
            FROM sales_updates
            WHERE company = ?
              AND location IN ({placeholders})
              AND sales_date >= ? AND sales_date <= ?
            ORDER BY sales_date, location, id""",
        (company, *outlets, date_from.isoformat(), date_to.isoformat()),
    ).fetchall()
    entries = []
    for row in rows:
        item = dict(row)
        try:
            values = json.loads(item.get("sales_entry_values") or "{}")
        except (TypeError, ValueError, json.JSONDecodeError):
            values = {}
        amount = parse_money(values.get("actual_cash"))
        if amount <= 0:
            continue
        entries.append(
            {
                "id": f"sales-{item['id']}",
                "source_id": item["id"],
                "entry_type": CASH_LEDGER_ENTRY_SALES,
                "entry_date": item["sales_date"],
                "location": item["location"] or "",
                "detail": item["location"] or "",
                "description": f"Actual cash — {item['location']}",
                "amount": amount,
                "signed_amount": amount,
                "can_delete": False,
            }
        )
    return entries


def _cash_ledger_expense_rows(conn, company, date_from, date_to, location=None):
    outlets = _cash_ledger_outlet_scope(location)
    placeholders = ",".join("?" for _ in outlets)
    rows = conn.execute(
        f"""SELECT e.id, e.location, e.sales_date, e.description, e.amount, e.expense_code,
                   e.category, s.name AS supplier_name
            FROM sales_update_expenses e
            LEFT JOIN suppliers s ON s.id = e.supplier_id
            WHERE e.company = ?
              AND e.location IN ({placeholders})
              AND e.payment_type = ?
              AND e.sales_date >= ? AND e.sales_date <= ?
            ORDER BY e.sales_date, e.id""",
        (
            company,
            *outlets,
            EXPENSE_PAYMENT_CASH,
            date_from.isoformat(),
            date_to.isoformat(),
        ),
    ).fetchall()
    entries = []
    for row in rows:
        item = dict(row)
        amount = round_half_up(item.get("amount"), 2)
        if amount <= 0:
            continue
        desc = (item.get("description") or "").strip() or "Cash expense"
        code = (item.get("expense_code") or "").strip()
        if code:
            desc = f"{code} · {desc}"
        entries.append(
            {
                "id": f"expense-{item['id']}",
                "source_id": item["id"],
                "entry_type": CASH_LEDGER_ENTRY_EXPENSE,
                "entry_date": item["sales_date"],
                "location": item.get("location") or "",
                "detail": item.get("location") or "",
                "description": desc,
                "amount": amount,
                "signed_amount": -amount,
                "can_delete": False,
                "supplier_name": item.get("supplier_name") or "",
            }
        )
    return entries


def _cash_ledger_load_rows(conn, company, date_from, date_to):
    rows = conn.execute(
        """SELECT id, load_date, description, amount
           FROM cash_ledger_loads
           WHERE company = ? AND load_date >= ? AND load_date <= ?
           ORDER BY load_date, id""",
        (company, date_from.isoformat(), date_to.isoformat()),
    ).fetchall()
    entries = []
    for row in rows:
        item = dict(row)
        amount = round_half_up(item.get("amount"), 2)
        if amount <= 0:
            continue
        entries.append(
            {
                "id": f"load-{item['id']}",
                "source_id": item["id"],
                "entry_type": CASH_LEDGER_ENTRY_LOAD,
                "entry_date": item["load_date"],
                "location": "",
                "detail": "Cash load",
                "description": (item.get("description") or "").strip() or "Cash load",
                "amount": amount,
                "signed_amount": amount,
                "can_delete": True,
            }
        )
    return entries


def _cash_ledger_split_concat(value):
    """Split sqlite GROUP_CONCAT output into unique non-empty parts."""
    parts = []
    seen = set()
    for raw in str(value or "").split(","):
        item = raw.strip()
        if not item or item in seen:
            continue
        seen.add(item)
        parts.append(item)
    return parts


def _cash_ledger_credit_description(invoices, guests, notes=""):
    invoices = [item for item in (invoices or []) if item]
    guests = [item for item in (guests or []) if item]
    if len(invoices) == 1:
        desc = f"Credit collection — {invoices[0]}"
        if guests:
            desc = f"{desc} — {guests[0]}"
    elif 1 < len(invoices) <= 3:
        desc = f"Credit collection — {' + '.join(invoices)}"
    elif len(invoices) > 3:
        desc = f"Credit collection — {len(invoices)} invoices"
    elif guests:
        desc = f"Credit collection — {guests[0]}"
    else:
        desc = "Credit collection"
    note = (notes or "").strip()
    if note:
        desc = f"{desc} — {note}"
    return desc


def _cash_ledger_credit_rows(conn, company, date_from, date_to, location=None):
    """Cash repayments from Credit / Room Transfer clearance."""
    has_table = conn.execute(
        """SELECT 1 FROM sqlite_master
           WHERE type = 'table' AND name = 'room_transfer_payments'"""
    ).fetchone()
    if not has_table:
        return []

    location = _normalize_cash_ledger_location(location)
    params = [company, ROOM_TRANSFER_PAYMENT_CASH, date_from.isoformat(), date_to.isoformat()]
    location_sql = ""
    if location != CASH_LEDGER_FILTER_ALL:
        location_sql = "AND a.location = ?"
        params.append(location)

    rows = conn.execute(
        f"""SELECT p.id AS payment_id,
                   p.payment_date,
                   p.notes,
                   SUM(a.amount) AS amount,
                   GROUP_CONCAT(a.location) AS locations,
                   GROUP_CONCAT(a.invoice_number) AS invoices,
                   GROUP_CONCAT(a.guest_name) AS guests
            FROM room_transfer_payments p
            JOIN room_transfer_payment_allocations a
              ON a.room_transfer_payment_id = p.id
            WHERE p.company = ?
              AND p.payment_method = ?
              AND p.payment_date >= ? AND p.payment_date <= ?
              {location_sql}
            GROUP BY p.id, p.payment_date, p.notes
            ORDER BY p.payment_date, p.id""",
        params,
    ).fetchall()

    entries = []
    for row in rows:
        item = dict(row)
        amount = round_half_up(item.get("amount"), 2)
        if amount <= 0:
            continue
        locations = _cash_ledger_split_concat(item.get("locations"))
        invoices = _cash_ledger_split_concat(item.get("invoices"))
        guests = _cash_ledger_split_concat(item.get("guests"))
        if location != CASH_LEDGER_FILTER_ALL:
            detail = location
        elif len(locations) == 1:
            detail = locations[0]
        elif locations:
            detail = " + ".join(locations)
        else:
            detail = "Credit"
        entries.append(
            {
                "id": f"credit-{item['payment_id']}",
                "source_id": item["payment_id"],
                "entry_type": CASH_LEDGER_ENTRY_CREDIT,
                "entry_date": item["payment_date"],
                "location": detail,
                "detail": detail,
                "description": _cash_ledger_credit_description(
                    invoices, guests, item.get("notes") or ""
                ),
                "amount": amount,
                "signed_amount": amount,
                "can_delete": False,
            }
        )
    return entries


def _cash_ledger_transfer_rows(conn, company, date_from, date_to):
    rows = conn.execute(
        """SELECT id, transfer_date, destination, description, amount
           FROM cash_ledger_transfers
           WHERE company = ? AND transfer_date >= ? AND transfer_date <= ?
           ORDER BY transfer_date, id""",
        (company, date_from.isoformat(), date_to.isoformat()),
    ).fetchall()
    entries = []
    for row in rows:
        item = dict(row)
        amount = round_half_up(item.get("amount"), 2)
        if amount <= 0:
            continue
        destination = _normalize_cash_ledger_transfer_destination(item.get("destination")) or "bank"
        dest_label = CASH_LEDGER_TRANSFER_DESTINATION_LABELS.get(destination, destination)
        entries.append(
            {
                "id": f"transfer-{item['id']}",
                "source_id": item["id"],
                "entry_type": CASH_LEDGER_ENTRY_TRANSFER,
                "entry_date": item["transfer_date"],
                "location": "",
                "detail": dest_label,
                "destination": destination,
                "description": (item.get("description") or "").strip() or f"Transfer to {dest_label}",
                "amount": amount,
                "signed_amount": -amount,
                "can_delete": True,
            }
        )
    return entries


def _build_cash_ledger_entries(conn, company, date_from, date_to, location=None):
    ensure_cash_ledger_schema(conn)
    location = _normalize_cash_ledger_location(location)
    entries = []
    entries.extend(_cash_ledger_sales_rows(conn, company, date_from, date_to, location=location))
    # Loads/transfers are company-level; show them only when viewing all locations.
    if location == CASH_LEDGER_FILTER_ALL:
        entries.extend(_cash_ledger_load_rows(conn, company, date_from, date_to))
    entries.extend(_cash_ledger_credit_rows(conn, company, date_from, date_to, location=location))
    entries.extend(_cash_ledger_expense_rows(conn, company, date_from, date_to, location=location))
    if location == CASH_LEDGER_FILTER_ALL:
        entries.extend(_cash_ledger_transfer_rows(conn, company, date_from, date_to))
    entries.sort(
        key=lambda row: (
            row.get("entry_date") or "",
            CASH_LEDGER_ENTRY_RANK.get(row.get("entry_type"), 99),
            row.get("source_id") or 0,
            row.get("id") or "",
        )
    )
    running = 0.0
    for entry in entries:
        running = round_half_up(running + entry.get("signed_amount", 0), 2)
        entry["running_balance"] = running
    return entries


def _cash_ledger_totals(entries):
    sales_total = 0.0
    load_total = 0.0
    credit_total = 0.0
    expense_total = 0.0
    transfer_total = 0.0
    sales_count = load_count = credit_count = expense_count = transfer_count = 0
    for entry in entries:
        kind = entry.get("entry_type")
        amount = round_half_up(entry.get("amount"), 2)
        if kind == CASH_LEDGER_ENTRY_SALES:
            sales_total += amount
            sales_count += 1
        elif kind == CASH_LEDGER_ENTRY_LOAD:
            load_total += amount
            load_count += 1
        elif kind == CASH_LEDGER_ENTRY_CREDIT:
            credit_total += amount
            credit_count += 1
        elif kind == CASH_LEDGER_ENTRY_EXPENSE:
            expense_total += amount
            expense_count += 1
        elif kind == CASH_LEDGER_ENTRY_TRANSFER:
            transfer_total += amount
            transfer_count += 1
    available = round_half_up(
        sales_total + load_total + credit_total - expense_total - transfer_total, 2
    )
    return {
        "sales_total": round_half_up(sales_total, 2),
        "sales_count": sales_count,
        "load_total": round_half_up(load_total, 2),
        "load_count": load_count,
        "credit_total": round_half_up(credit_total, 2),
        "credit_count": credit_count,
        "expense_total": round_half_up(expense_total, 2),
        "expense_count": expense_count,
        "transfer_total": round_half_up(transfer_total, 2),
        "transfer_count": transfer_count,
        "available_total": available,
    }


def _cash_ledger_available_as_of(conn, company, as_of_date, *, exclude_expense_id=None):
    """Available Cash through as_of_date using the Cash Ledger formula.

    Optionally excludes an existing cash expense (used when editing) so its
    amount is treated as still available for re-save / amount changes.
    """
    as_of = as_of_date
    if isinstance(as_of, str):
        as_of = _parse_sales_date(as_of)
    if not as_of:
        as_of = date.today()
    company = company or DEFAULT_COMPANY
    entries = _build_cash_ledger_entries(conn, company, date(2000, 1, 1), as_of)
    available = _cash_ledger_totals(entries)["available_total"]
    if exclude_expense_id:
        try:
            exclude_id = int(exclude_expense_id)
        except (TypeError, ValueError):
            exclude_id = None
        if exclude_id:
            row = conn.execute(
                """SELECT amount, payment_type, sales_date
                   FROM sales_update_expenses WHERE id = ? AND company = ?""",
                (exclude_id, company),
            ).fetchone()
            if (
                row
                and _normalize_expense_payment_type(row["payment_type"]) == EXPENSE_PAYMENT_CASH
                and (row["sales_date"] or "") <= as_of.isoformat()
            ):
                available = round_half_up(available + round_half_up(row["amount"], 2), 2)
    return available


def _validate_cash_expense_against_available(
    conn, company, sales_date, amount, payment_type, *, exclude_expense_id=None
):
    """Reject cash expenses that exceed Cash Ledger Available Cash."""
    if _normalize_expense_payment_type(payment_type) != EXPENSE_PAYMENT_CASH:
        return None
    available = _cash_ledger_available_as_of(
        conn, company, sales_date, exclude_expense_id=exclude_expense_id
    )
    if round_half_up(amount, 2) - available > 0.001:
        return (
            "Cash expense cannot be more than available cash "
            f"(₹{available:,.2f})."
        )
    return None


@app.route("/accounts")
def accounts():
    user = get_current_user()
    preferred = (
        ("purchase_ledger", "purchase_ledger"),
        ("cash_ledger", "cash_ledger"),
        ("purchase_verification", "purchase_verification"),
        ("credit_payment", "credit_payment"),
        ("supplier_master", "supplier_master"),
    )
    for key, endpoint in preferred:
        if user_can_access_accounts_submodule(user, key):
            return redirect(url_for(endpoint))
    return redirect(url_for("home"))


@app.route("/accounts/purchase-ledger")
def purchase_ledger():
    today = date.today()
    date_from, date_to, date_filter_active = _resolve_optional_filter_date_range(
        request.args, "date_from", "date_to", default_fy=True
    )
    query_date_from = date_from if date_filter_active else date(2000, 1, 1)
    query_date_to = date_to if date_filter_active else today

    selected_supplier, supplier_id = _parse_purchase_ledger_supplier(
        request.args.get("supplier")
    )
    selected_category, category = _parse_purchase_ledger_category(
        request.args.get("category")
    )
    selected_payment, payment_type = _parse_purchase_ledger_payment(
        request.args.get("payment")
    )

    conn = get_db()
    expense_categories = EXPENSE_CATEGORIES
    expense_category_labels = EXPENSE_CATEGORY_LABELS
    try:
        suppliers = _all_suppliers(conn)
        supplier_lookup = {str(s["id"]): s for s in suppliers}
        if selected_supplier != PURCHASE_LEDGER_FILTER_ALL and selected_supplier not in supplier_lookup:
            selected_supplier = PURCHASE_LEDGER_FILTER_ALL
            supplier_id = None
        if selected_category != PURCHASE_LEDGER_FILTER_ALL and not _normalize_expense_category(selected_category):
            selected_category = PURCHASE_LEDGER_FILTER_ALL
            category = None
        if selected_payment != PURCHASE_LEDGER_FILTER_ALL and selected_payment not in EXPENSE_PAYMENT_LABELS:
            selected_payment = PURCHASE_LEDGER_FILTER_ALL
            payment_type = None
        entries = _purchase_ledger_entries(
            conn,
            query_date_from,
            query_date_to,
            supplier_id,
            category=category,
            payment_type=payment_type,
        )
        available_cash = _cash_ledger_available_as_of(conn, DEFAULT_COMPANY, today)
        expense_categories = _expense_category_choices(conn)
        expense_category_labels = _expense_category_labels(conn)
    finally:
        conn.close()

    total_amount = round_half_up(sum(entry["amount"] for entry in entries), 2)
    outstanding_entries = [
        entry for entry in entries
        if entry.get("settlement_status") in ("outstanding", "partial")
    ]
    cleared_entries = [
        entry for entry in entries
        if entry.get("settlement_status") == "cleared"
    ]
    cash_entries = [
        entry for entry in entries
        if entry.get("display_payment_type") == EXPENSE_PAYMENT_CASH
    ]
    outstanding_total = round_half_up(sum(entry["balance"] for entry in outstanding_entries), 2)
    cleared_total = round_half_up(sum(entry["amount"] for entry in cleared_entries), 2)
    cash_total = round_half_up(sum(entry["amount"] for entry in cash_entries), 2)
    selected_supplier_label = "All suppliers"
    if selected_supplier != PURCHASE_LEDGER_FILTER_ALL:
        match = supplier_lookup.get(selected_supplier)
        if match:
            selected_supplier_label = match["name"]
    selected_category_label = "All categories"
    if selected_category != PURCHASE_LEDGER_FILTER_ALL:
        selected_category_label = expense_category_labels.get(
            selected_category, selected_category.replace("_", " ").title()
        )
    selected_payment_label = "All payments"
    if selected_payment != PURCHASE_LEDGER_FILTER_ALL:
        selected_payment_label = EXPENSE_PAYMENT_LABELS.get(selected_payment, selected_payment_label)

    filter_date_from = date_from.isoformat() if date_filter_active else ""
    filter_date_to = date_to.isoformat() if date_filter_active else ""
    report_kwargs = {
        "supplier": selected_supplier,
        "category": selected_category,
        "payment": selected_payment,
    }
    if date_filter_active:
        report_kwargs["date_from"] = filter_date_from
        report_kwargs["date_to"] = filter_date_to
    clear_kwargs = {
        "supplier": selected_supplier,
        "category": selected_category,
        "payment": selected_payment,
    }

    return render_template(
        "purchase_ledger.html",
        page_title="Expense Ledger",
        page_subtitle="",
        filter_form_action=url_for("purchase_ledger"),
        date_from=filter_date_from,
        date_to=filter_date_to,
        active_date_filter=date_filter_active,
        selected_supplier=selected_supplier,
        selected_supplier_label=selected_supplier_label,
        selected_category=selected_category,
        selected_category_label=selected_category_label,
        selected_payment=selected_payment,
        selected_payment_label=selected_payment_label,
        suppliers=suppliers,
        purchase_entries=entries,
        purchase_total=total_amount,
        outstanding_total=outstanding_total,
        outstanding_count=len(outstanding_entries),
        cleared_total=cleared_total,
        cleared_count=len(cleared_entries),
        cash_total=cash_total,
        cash_count=len(cash_entries),
        expense_payment_types=EXPENSE_PAYMENT_TYPES,
        expense_payment_labels=EXPENSE_PAYMENT_LABELS,
        purchase_ledger_payment_labels=PURCHASE_LEDGER_PAYMENT_LABELS,
        expense_categories=expense_categories,
        expense_category_labels=expense_category_labels,
        credit_settlement_status_labels=CREDIT_SETTLEMENT_STATUS_LABELS,
        purchase_add_url=url_for("purchase_ledger_add"),
        purchase_edit_url=url_for("purchase_ledger_edit"),
        purchase_delete_url=url_for("purchase_ledger_delete"),
        purchase_ledger_report_url=url_for("export_purchase_ledger_report", **report_kwargs),
        purchase_ledger_clear_url=url_for("purchase_ledger", **clear_kwargs),
        supplier_create_url=url_for("create_supplier"),
        available_cash=available_cash,
        available_cash_url=url_for("cash_ledger_available"),
        default_company=DEFAULT_COMPANY,
        default_location=OUTLET_HOTEL,
        today_iso=today.isoformat(),
        de_nav_section="accounts",
        de_nav_accounts_view="purchase_ledger",
    )


@app.route("/accounts/purchase-ledger/report")
def export_purchase_ledger_report():
    """Excel download of purchase ledger entries for the selected filters."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    today = date.today()
    date_from, date_to, date_filter_active = _resolve_optional_filter_date_range(
        request.args, "date_from", "date_to", default_fy=True
    )
    query_date_from = date_from if date_filter_active else date(2000, 1, 1)
    query_date_to = date_to if date_filter_active else today

    selected_supplier, supplier_id = _parse_purchase_ledger_supplier(
        request.args.get("supplier")
    )
    selected_category, category = _parse_purchase_ledger_category(
        request.args.get("category")
    )
    selected_payment, payment_type = _parse_purchase_ledger_payment(
        request.args.get("payment")
    )
    if selected_category != PURCHASE_LEDGER_FILTER_ALL and not _normalize_expense_category(selected_category):
        category = None
        selected_category = PURCHASE_LEDGER_FILTER_ALL
    if selected_payment != PURCHASE_LEDGER_FILTER_ALL and selected_payment not in EXPENSE_PAYMENT_LABELS:
        payment_type = None

    conn = get_db()
    try:
        category_labels = _expense_category_labels(conn)
        entries = _purchase_ledger_entries(
            conn,
            query_date_from,
            query_date_to,
            supplier_id,
            category=category,
            payment_type=payment_type,
        )
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Ledger"
    header_font = Font(bold=True)
    headers = [
        "Expense ID",
        "Date",
        "Expense",
        "Category",
        "Invoice",
        "Supplier",
        "GST",
        "Payment",
        "Status",
        "Amount",
        "Paid",
        "Balance",
    ]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font

    for idx, entry in enumerate(entries, start=2):
        category_key = entry.get("category") or ""
        payment_key = entry.get("display_payment_type") or entry.get("payment_type") or ""
        status_key = entry.get("settlement_status") or ""
        ws.cell(row=idx, column=1, value=entry.get("expense_code") or "")
        ws.cell(row=idx, column=2, value=entry.get("sales_date") or "")
        ws.cell(row=idx, column=3, value=entry.get("description") or "")
        ws.cell(
            row=idx,
            column=4,
            value=category_labels.get(
                category_key,
                EXPENSE_CATEGORY_LABELS.get(category_key, category_key),
            ),
        )
        ws.cell(row=idx, column=5, value=entry.get("invoice_number") or "")
        ws.cell(row=idx, column=6, value=entry.get("supplier_name") or "")
        ws.cell(row=idx, column=7, value=entry.get("supplier_gst") or "")
        ws.cell(
            row=idx,
            column=8,
            value=PURCHASE_LEDGER_PAYMENT_LABELS.get(payment_key, payment_key),
        )
        ws.cell(
            row=idx,
            column=9,
            value=CREDIT_SETTLEMENT_STATUS_LABELS.get(status_key, status_key),
        )
        ws.cell(row=idx, column=10, value=round_half_up(entry.get("amount"), 2))
        ws.cell(row=idx, column=11, value=round_half_up(entry.get("paid_amount"), 2))
        ws.cell(row=idx, column=12, value=round_half_up(entry.get("balance"), 2))

    for column_cells in ws.columns:
        width = 12
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, 40))
        ws.column_dimensions[column_cells[0].column_letter].width = width

    if date_filter_active:
        fname = f"purchase_ledger_{query_date_from.isoformat()}_to_{query_date_to.isoformat()}.xlsx"
    else:
        fname = "purchase_ledger_all.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/accounts/purchase-ledger/add", methods=["POST"])
def purchase_ledger_add():
    user = get_current_user()
    data = request.get_json(silent=True) or {}
    conn = get_db()
    try:
        result, error = _create_sales_expense(
            conn,
            user,
            data,
            default_location=OUTLET_HOTEL,
            include_sales_totals=False,
        )
        if error:
            status = 403 if "Cannot save" in error or "already saved" in error else 400
            return jsonify({"ok": False, "error": error}), status
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True, **result})


def _update_purchase_ledger_expense(conn, user, data):
    """Update a hotel purchase only when it is still outstanding credit."""
    expense_id = data.get("id") or data.get("expense_id")
    try:
        expense_id = int(expense_id)
    except (TypeError, ValueError):
        return None, "Purchase not found."

    existing = conn.execute(
        """SELECT id, company, location, sales_date, description, amount, payment_type,
                  transaction_id, supplier_id, category, invoice_number, expense_code
           FROM sales_update_expenses WHERE id = ?""",
        (expense_id,),
    ).fetchone()
    if not existing:
        return None, "Purchase not found."
    existing = dict(existing)
    if existing.get("location") != OUTLET_HOTEL:
        return None, "Only hotel purchases can be edited here."

    paid_total = _credit_expense_paid_total(conn, expense_id)
    status = _credit_settlement_status(
        existing.get("payment_type"), existing.get("amount"), paid_total
    )
    if status != "outstanding":
        return None, "Only outstanding credit purchases can be edited."

    company = existing.get("company") or data.get("company", DEFAULT_COMPANY)
    location = OUTLET_HOTEL
    sales_date = (data.get("date") or existing.get("sales_date") or "").strip()
    description = (data.get("description") or "").strip()
    amount = parse_money(data.get("amount"))
    raw_payment_type = (data.get("payment_type") or "").strip()
    if not raw_payment_type:
        return None, "Please select a payment type."
    payment_type = _normalize_expense_payment_type(raw_payment_type)
    category = _normalize_expense_category(data.get("category"))
    transaction_id = (data.get("transaction_id") or "").strip()
    invoice_number = (data.get("invoice_number") or "").strip()
    supplier_id = data.get("supplier_id")

    lock_error = _check_sales_date_lock(user, company, location, sales_date)
    if lock_error:
        return None, lock_error
    if sales_date != existing.get("sales_date"):
        prior_lock = _check_sales_date_lock(user, company, location, existing.get("sales_date"))
        if prior_lock:
            return None, prior_lock

    if not description or amount <= 0:
        return None, "Description and positive amount are required."
    if not supplier_id:
        return None, "Please select a supplier."
    if not category:
        return None, "Please select a category."
    if payment_type == EXPENSE_PAYMENT_BANK and not transaction_id:
        return None, "Transaction ID is required for bank transfer."
    if payment_type != EXPENSE_PAYMENT_BANK:
        transaction_id = ""

    supplier = _get_supplier(conn, supplier_id)
    if not supplier:
        return None, "Selected supplier was not found."

    # Multi-category stock inward can share one supplier+invoice across expenses.
    # When editing one row, exclude siblings that already share the prior invoice.
    exclude_ids = [expense_id]
    prior_invoice = _normalize_invoice_number(existing.get("invoice_number"))
    prior_supplier = existing.get("supplier_id")
    if prior_invoice and prior_supplier:
        sibling_rows = conn.execute(
            """SELECT id FROM sales_update_expenses
               WHERE supplier_id = ? AND LOWER(TRIM(invoice_number)) = LOWER(?)
                 AND TRIM(invoice_number) != ''""",
            (prior_supplier, prior_invoice),
        ).fetchall()
        exclude_ids = [int(row["id"]) for row in sibling_rows] or [expense_id]

    duplicate = _duplicate_expense_invoice(
        conn, supplier_id, invoice_number, exclude_expense_id=exclude_ids
    )
    if duplicate:
        code = duplicate["expense_code"] or f"#{duplicate['id']}"
        return None, f"An expense with this supplier and invoice number already exists ({code})."

    conn.execute(
        f"""UPDATE sales_update_expenses
           SET sales_date = ?, description = ?, amount = ?, payment_type = ?,
               transaction_id = ?, supplier_id = ?, category = ?, invoice_number = ?,
               updated_at = {SQL_NOW}
           WHERE id = ? AND location = ?""",
        (
            sales_date,
            description,
            amount,
            payment_type,
            transaction_id,
            supplier_id,
            category,
            invoice_number,
            expense_id,
            OUTLET_HOTEL,
        ),
    )
    return {
        "expense_id": expense_id,
        "expense_code": existing.get("expense_code") or "",
        "sales_date": sales_date,
    }, None


@app.route("/accounts/purchase-ledger/edit", methods=["POST"])
def purchase_ledger_edit():
    user = get_current_user()
    data = request.get_json(silent=True) or {}
    conn = get_db()
    try:
        result, error = _update_purchase_ledger_expense(conn, user, data)
        if error:
            status = 403 if "Cannot save" in error or "already saved" in error else 400
            if "not found" in error.lower():
                status = 404
            return jsonify({"ok": False, "error": error}), status
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True, **result})


def _delete_purchase_ledger_expense(conn, user, data):
    """Delete a hotel purchase only when it is still outstanding credit."""
    expense_id = data.get("id") or data.get("expense_id")
    try:
        expense_id = int(expense_id)
    except (TypeError, ValueError):
        return None, "Purchase not found."

    existing = conn.execute(
        """SELECT id, company, location, sales_date, amount, payment_type, expense_code
           FROM sales_update_expenses WHERE id = ?""",
        (expense_id,),
    ).fetchone()
    if not existing:
        return None, "Purchase not found."
    existing = dict(existing)
    if existing.get("location") != OUTLET_HOTEL:
        return None, "Only hotel purchases can be deleted here."

    paid_total = _credit_expense_paid_total(conn, expense_id)
    status = _credit_settlement_status(
        existing.get("payment_type"), existing.get("amount"), paid_total
    )
    if status != "outstanding":
        return None, "Only outstanding credit purchases can be deleted."

    company = existing.get("company") or DEFAULT_COMPANY
    sales_date = existing.get("sales_date") or ""
    lock_error = _check_sales_date_lock(user, company, OUTLET_HOTEL, sales_date)
    if lock_error:
        return None, lock_error

    tables = {
        row[0]
        for row in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()
    }
    if "credit_payment_allocations" in tables:
        conn.execute(
            "DELETE FROM credit_payment_allocations WHERE expense_id = ?",
            (expense_id,),
        )
    if "purchase_verification_allocations" in tables:
        conn.execute(
            "DELETE FROM purchase_verification_allocations WHERE expense_id = ?",
            (expense_id,),
        )
    conn.execute(
        "DELETE FROM sales_update_expenses WHERE id = ? AND location = ?",
        (expense_id, OUTLET_HOTEL),
    )
    return {
        "expense_id": expense_id,
        "expense_code": existing.get("expense_code") or "",
        "sales_date": sales_date,
    }, None


@app.route("/accounts/purchase-ledger/delete", methods=["POST"])
def purchase_ledger_delete():
    user = get_current_user()
    data = request.get_json(silent=True) or {}
    conn = get_db()
    try:
        result, error = _delete_purchase_ledger_expense(conn, user, data)
        if error:
            status = 403 if "Cannot save" in error or "already saved" in error else 400
            if "not found" in error.lower():
                status = 404
            return jsonify({"ok": False, "error": error}), status
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True, **result})


@app.route("/accounts/cash-ledger/available")
def cash_ledger_available():
    """JSON Available Cash as of a date (Cash Ledger formula)."""
    company = request.args.get("company") or DEFAULT_COMPANY
    as_of = _parse_sales_date(request.args.get("date") or date.today().isoformat())
    exclude_expense_id = request.args.get("exclude_expense_id")
    conn = get_db()
    try:
        available = _cash_ledger_available_as_of(
            conn, company, as_of, exclude_expense_id=exclude_expense_id
        )
    finally:
        conn.close()
    return jsonify({"ok": True, "available_cash": available, "date": as_of.isoformat()})


@app.route("/accounts/cash-ledger")
def cash_ledger():
    today = date.today()
    date_from, date_to, date_filter_active = _resolve_cash_ledger_date_range(request.args)
    selected_location = _normalize_cash_ledger_location(request.args.get("location"))

    company = DEFAULT_COMPANY
    conn = get_db()
    try:
        entries = _build_cash_ledger_entries(
            conn, company, date_from, date_to, location=selected_location
        )
    finally:
        conn.close()

    filter_date_from = date_from.isoformat() if date_filter_active else ""
    filter_date_to = date_to.isoformat() if date_filter_active else ""
    report_kwargs = {}
    if selected_location != CASH_LEDGER_FILTER_ALL:
        report_kwargs["location"] = selected_location
    if date_filter_active:
        report_kwargs["date_from"] = date_from.isoformat()
        report_kwargs["date_to"] = date_to.isoformat()

    clear_kwargs = {}
    if selected_location != CASH_LEDGER_FILTER_ALL:
        clear_kwargs["location"] = selected_location

    totals = _cash_ledger_totals(entries)
    # Chronological order is required for running balance; show newest first in the UI.
    display_entries = list(reversed(entries))
    return render_template(
        "cash_ledger.html",
        page_title="Cash Ledger",
        page_subtitle="",
        filter_form_action=url_for("cash_ledger"),
        date_from=filter_date_from,
        date_to=filter_date_to,
        date_filter_active=date_filter_active,
        selected_location=selected_location,
        cash_ledger_filter_locations=CASH_LEDGER_FILTER_LOCATIONS,
        ledger_entries=display_entries,
        sales_total=totals["sales_total"],
        sales_count=totals["sales_count"],
        load_total=totals["load_total"],
        load_count=totals["load_count"],
        credit_total=totals["credit_total"],
        credit_count=totals["credit_count"],
        expense_total=totals["expense_total"],
        expense_count=totals["expense_count"],
        transfer_total=totals["transfer_total"],
        transfer_count=totals["transfer_count"],
        available_total=totals["available_total"],
        cash_ledger_entry_labels=CASH_LEDGER_ENTRY_LABELS,
        cash_ledger_transfer_destinations=CASH_LEDGER_TRANSFER_DESTINATIONS,
        cash_ledger_transfer_destination_labels=CASH_LEDGER_TRANSFER_DESTINATION_LABELS,
        load_url=url_for("cash_ledger_load"),
        transfer_url=url_for("cash_ledger_transfer"),
        delete_load_url=url_for("cash_ledger_delete_load"),
        delete_transfer_url=url_for("cash_ledger_delete_transfer"),
        cash_ledger_report_url=url_for("export_cash_ledger_report", **report_kwargs),
        cash_ledger_clear_url=url_for("cash_ledger", **clear_kwargs),
        default_company=company,
        today_iso=today.isoformat(),
        de_nav_section="accounts",
        de_nav_accounts_view="cash_ledger",
    )


@app.route("/accounts/cash-ledger/load", methods=["POST"])
def cash_ledger_load():
    data = request.get_json(silent=True) or {}
    company = (data.get("company") or DEFAULT_COMPANY).strip() or DEFAULT_COMPANY
    raw_date = (data.get("date") or data.get("load_date") or "").strip()
    description = (data.get("description") or "").strip()
    amount = parse_money(data.get("amount"))
    if not raw_date:
        return jsonify({"ok": False, "error": "Date is required."}), 400
    try:
        load_date = date.fromisoformat(raw_date)
    except ValueError:
        return jsonify({"ok": False, "error": "Enter a valid date."}), 400
    if amount <= 0:
        return jsonify({"ok": False, "error": "Enter a positive amount."}), 400
    if not description:
        return jsonify({"ok": False, "error": "Description is required."}), 400

    conn = get_db()
    try:
        ensure_cash_ledger_schema(conn)
        cursor = conn.execute(
            """INSERT INTO cash_ledger_loads (company, load_date, description, amount)
               VALUES (?, ?, ?, ?)""",
            (company, load_date.isoformat(), description, amount),
        )
        conn.commit()
        load_id = cursor.lastrowid
    finally:
        conn.close()
    return jsonify({"ok": True, "id": load_id})


@app.route("/accounts/cash-ledger/transfer", methods=["POST"])
def cash_ledger_transfer():
    data = request.get_json(silent=True) or {}
    company = (data.get("company") or DEFAULT_COMPANY).strip() or DEFAULT_COMPANY
    raw_date = (data.get("date") or data.get("transfer_date") or "").strip()
    description = (data.get("description") or "").strip()
    amount = parse_money(data.get("amount"))
    destination = _normalize_cash_ledger_transfer_destination(data.get("destination"))
    if not raw_date:
        return jsonify({"ok": False, "error": "Date is required."}), 400
    try:
        transfer_date = date.fromisoformat(raw_date)
    except ValueError:
        return jsonify({"ok": False, "error": "Enter a valid date."}), 400
    if amount <= 0:
        return jsonify({"ok": False, "error": "Enter a positive amount."}), 400
    if not destination:
        return jsonify({"ok": False, "error": "Select Bank or Owner."}), 400
    if not description:
        return jsonify({"ok": False, "error": "Description is required."}), 400

    conn = get_db()
    try:
        ensure_cash_ledger_schema(conn)
        cursor = conn.execute(
            """INSERT INTO cash_ledger_transfers
               (company, transfer_date, destination, description, amount)
               VALUES (?, ?, ?, ?, ?)""",
            (company, transfer_date.isoformat(), destination, description, amount),
        )
        conn.commit()
        transfer_id = cursor.lastrowid
    finally:
        conn.close()
    return jsonify({"ok": True, "id": transfer_id})


@app.route("/accounts/cash-ledger/report")
def export_cash_ledger_report():
    """Excel download of cash ledger movements for the selected date range."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    date_from, date_to, date_filter_active = _resolve_cash_ledger_date_range(request.args)
    selected_location = _normalize_cash_ledger_location(request.args.get("location"))

    company = DEFAULT_COMPANY
    conn = get_db()
    try:
        entries = _build_cash_ledger_entries(
            conn, company, date_from, date_to, location=selected_location
        )
    finally:
        conn.close()

    totals = _cash_ledger_totals(entries)
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    header_font = Font(bold=True)
    summary_headers = ["Metric", "Amount", "Count"]
    for col, title in enumerate(summary_headers, start=1):
        cell = summary.cell(row=1, column=col, value=title)
        cell.font = header_font
    summary_rows = [
        ("Actual Cash", totals["sales_total"], totals["sales_count"]),
        ("Load Cash", totals["load_total"], totals["load_count"]),
        ("Credit Cash", totals["credit_total"], totals["credit_count"]),
        ("Expense", totals["expense_total"], totals["expense_count"]),
        ("Transfer Out", totals["transfer_total"], totals["transfer_count"]),
        ("Available Cash", totals["available_total"], len(entries)),
        ("Location", selected_location, ""),
        (
            "Date From",
            date_from.isoformat() if date_filter_active else "All",
            "",
        ),
        (
            "Date To",
            date_to.isoformat() if date_filter_active else "All",
            "",
        ),
    ]
    for idx, (label, amount, count) in enumerate(summary_rows, start=2):
        summary.cell(row=idx, column=1, value=label)
        summary.cell(row=idx, column=2, value=amount if isinstance(amount, (int, float)) else amount)
        summary.cell(row=idx, column=3, value=count)

    movements = wb.create_sheet("Cash Movements")
    headers = ["Date", "Type", "Detail", "Description", "Amount", "Balance"]
    for col, title in enumerate(headers, start=1):
        cell = movements.cell(row=1, column=col, value=title)
        cell.font = header_font
    for idx, entry in enumerate(entries, start=2):
        entry_type = entry.get("entry_type") or ""
        movements.cell(row=idx, column=1, value=entry.get("entry_date") or "")
        movements.cell(
            row=idx,
            column=2,
            value=CASH_LEDGER_ENTRY_LABELS.get(entry_type, entry_type),
        )
        movements.cell(row=idx, column=3, value=entry.get("detail") or "")
        movements.cell(row=idx, column=4, value=entry.get("description") or "")
        movements.cell(row=idx, column=5, value=round_half_up(entry.get("signed_amount"), 2))
        movements.cell(row=idx, column=6, value=round_half_up(entry.get("running_balance"), 2))

    for ws in (summary, movements):
        for column_cells in ws.columns:
            width = 12
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(value) + 2, 48))
            ws.column_dimensions[column_cells[0].column_letter].width = width

    fname = (
        f"cash_ledger_{date_from.isoformat()}_to_{date_to.isoformat()}.xlsx"
        if date_filter_active
        else "cash_ledger_all.xlsx"
    )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/accounts/cash-ledger/load/delete", methods=["POST"])
def cash_ledger_delete_load():
    data = request.get_json(silent=True) or {}
    try:
        load_id = int(data.get("id") or data.get("load_id"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Load entry not found."}), 404

    conn = get_db()
    try:
        row = conn.execute(
            "SELECT id FROM cash_ledger_loads WHERE id = ?",
            (load_id,),
        ).fetchone()
        if not row:
            return jsonify({"ok": False, "error": "Load entry not found."}), 404
        conn.execute("DELETE FROM cash_ledger_loads WHERE id = ?", (load_id,))
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True, "id": load_id})


@app.route("/accounts/cash-ledger/transfer/delete", methods=["POST"])
def cash_ledger_delete_transfer():
    data = request.get_json(silent=True) or {}
    try:
        transfer_id = int(data.get("id") or data.get("transfer_id"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Transfer entry not found."}), 404

    conn = get_db()
    try:
        row = conn.execute(
            "SELECT id FROM cash_ledger_transfers WHERE id = ?",
            (transfer_id,),
        ).fetchone()
        if not row:
            return jsonify({"ok": False, "error": "Transfer entry not found."}), 404
        conn.execute("DELETE FROM cash_ledger_transfers WHERE id = ?", (transfer_id,))
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True, "id": transfer_id})


@app.route("/accounts/credit-payment")
def credit_payment():
    return _render_credit_settlement_page(CREDIT_SETTLEMENT_MODE_CREDIT_PAYMENT)


_VENDOR_PAYMENT_TEMPLATE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "templates",
    "accounts",
    "vendor_payment_template.xlsx",
)
_VENDOR_DEBIT_ACC_NO = "387905000829"
_VENDOR_MOBILE_NUM = 9933226086
_VENDOR_EMAIL_ID = "mithra.varma@gmail.com"
# ICICI CIB PAB_VENDOR upload headers (A–S) when the local .xlsx template is absent.
_VENDOR_PAYMENT_HEADERS = (
    "PYMT_PROD_TYPE_CODE",
    "PYMT_MODE",
    "DEBIT_ACC_NO",
    "BNF_NAME",
    "BENE_ACC_NO",
    "BENE_IFSC",
    "AMOUNT",
    "CREDIT_NARR",
    "PYMT_REF_NO",
    "MOBILE_NUM",
    "EMAIL_ID",
    "REMARK",
    "PYMT_DATE",
    "REF1",
    "REF2",
    "REF3",
    "REF4",
    "REF5",
    "REF6",
)


def _vendor_payment_category_narration(category):
    """H/I narration for the expense's category only."""
    key = _normalize_expense_category(category)
    label = EXPENSE_CATEGORY_LABELS.get(key) or (category or "").strip() or "OTHER"
    return label.upper()


def _credit_payment_report_rows(conn, date_from, date_to, supplier_id=None):
    """One ICICI vendor-payment row per supplier + category with outstanding credit."""
    entries = _outstanding_credit_expenses(
        conn, date_from, date_to, supplier_id=supplier_id
    )
    grouped = {}
    for entry in entries:
        sid = entry.get("supplier_id")
        if not sid:
            continue
        category = _normalize_expense_category(entry.get("category")) or "other"
        key = (sid, category)
        bucket = grouped.get(key)
        if not bucket:
            bucket = {
                "supplier_id": sid,
                "category": category,
                "amount": 0.0,
            }
            grouped[key] = bucket
        bucket["amount"] = round_half_up(bucket["amount"] + entry.get("balance", 0), 2)

    rows = []
    for bucket in grouped.values():
        if bucket["amount"] <= 0:
            continue
        supplier = _get_supplier(conn, bucket["supplier_id"])
        if not supplier:
            continue
        account = (supplier.get("bank_account_number") or "").strip()
        ifsc = (supplier.get("ifsc_code") or "").strip().upper()
        if not account or not ifsc:
            continue
        rows.append({
            "name": (supplier.get("name") or "").strip(),
            "account": account,
            "ifsc": ifsc,
            "amount": bucket["amount"],
            "narration": _vendor_payment_category_narration(bucket["category"]),
            "mode": "FT" if ifsc.startswith("ICIC") else "NEFT",
        })
    rows.sort(key=lambda item: (item["name"].lower(), item["narration"]))
    return rows


@app.route("/accounts/credit-payment/report")
def export_credit_payment_report():
    """ICICI vendor payment Excel for outstanding credit suppliers."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font

    today = date.today()
    date_from, date_to, _date_filter_active = _resolve_optional_filter_date_range(
        request.args, "date_from", "date_to", default_fy=True
    )
    _, supplier_id = _parse_purchase_ledger_supplier(request.args.get("supplier"))

    conn = get_db()
    try:
        rows = _credit_payment_report_rows(
            conn, date_from, date_to, supplier_id=supplier_id
        )
    finally:
        conn.close()

    if os.path.isfile(_VENDOR_PAYMENT_TEMPLATE):
        wb = load_workbook(_VENDOR_PAYMENT_TEMPLATE)
        ws = wb.active
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Credit Payment"
        for col, title in enumerate(_VENDOR_PAYMENT_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = Font(bold=True)

    payment_date = today
    for item in rows:
        ws.append([
            "PAB_VENDOR",                 # A
            item["mode"],                 # B
            _VENDOR_DEBIT_ACC_NO,         # C
            item["name"],                 # D
            item["account"],              # E
            item["ifsc"],                 # F
            item["amount"],               # G
            item["narration"],            # H
            item["narration"],            # I
            _VENDOR_MOBILE_NUM,           # J
            _VENDOR_EMAIL_ID,             # K
            "NIL",                        # L
            payment_date,                 # M
            "NIL",                        # N
            "NIL",                        # O
            "NIL",                        # P
            "NIL",                        # Q
            "NIL",                        # R
            "NIL",                        # S
        ])
        ws.cell(row=ws.max_row, column=13).value = payment_date

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"credit_payment_report_{today.isoformat()}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/accounts/purchase-verification")
def purchase_verification():
    return _render_credit_settlement_page(CREDIT_SETTLEMENT_MODE_PURCHASE_VERIFICATION)


@app.route("/accounts/purchase-verification/report")
def export_purchase_verification_report():
    """Excel report for pending or verified purchases based on page filters."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    today = date.today()
    selected_view = _normalize_credit_payment_view(request.args.get("view"))
    date_from, date_to, date_filter_active = _resolve_optional_filter_date_range(
        request.args, "date_from", "date_to", default_fy=True
    )
    payment_date_from, payment_date_to, payment_date_filter_active = _resolve_optional_filter_date_range(
        request.args, "payment_date_from", "payment_date_to"
    )
    _, supplier_id = _parse_purchase_ledger_supplier(request.args.get("supplier"))

    wb = Workbook()
    ws = wb.active
    header_font = Font(bold=True)

    conn = get_db()
    try:
        if selected_view == CREDIT_PAYMENT_VIEW_HISTORY:
            ws.title = "Verified Purchases"
            headers = [
                "Verification Date",
                "Supplier",
                "GST",
                "Method",
                "Account",
                "Transaction ID",
                "Expense IDs",
                "Amount",
                "Notes",
            ]
            for col, title in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=title)
                cell.font = header_font
            entries = _purchase_verification_entries(
                conn,
                verification_date_from=payment_date_from,
                verification_date_to=payment_date_to,
                supplier_id=supplier_id,
            )
            for idx, entry in enumerate(entries, start=2):
                method = entry.get("payment_method") or ""
                method_label = CREDIT_PAYMENT_METHOD_LABELS.get(method, method)
                ws.cell(row=idx, column=1, value=entry.get("payment_date") or "")
                ws.cell(row=idx, column=2, value=entry.get("supplier_name") or "")
                ws.cell(row=idx, column=3, value=entry.get("supplier_gst") or "")
                ws.cell(row=idx, column=4, value=method_label)
                ws.cell(row=idx, column=5, value=entry.get("verification_account") or "")
                ws.cell(row=idx, column=6, value=entry.get("transaction_id") or "")
                ws.cell(row=idx, column=7, value=entry.get("expense_codes") or "")
                ws.cell(row=idx, column=8, value=round_half_up(entry.get("total_amount"), 2))
                ws.cell(row=idx, column=9, value=entry.get("notes") or "")
            fname = (
                f"purchase_verification_history_"
                f"{payment_date_from.isoformat() if payment_date_filter_active else 'All'}_to_"
                f"{payment_date_to.isoformat() if payment_date_filter_active else 'All'}.xlsx"
            )
        else:
            ws.title = "Pending Verification"
            headers = [
                "Expense ID",
                "Date",
                "Expense",
                "Category",
                "Supplier",
                "GST",
                "Payment Type",
                "Amount",
                "Verified",
                "Balance",
            ]
            for col, title in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=title)
                cell.font = header_font
            entries = _pending_purchase_verifications(
                conn, date_from, date_to, supplier_id=supplier_id
            )
            for idx, entry in enumerate(entries, start=2):
                category = entry.get("category") or ""
                category_label = EXPENSE_CATEGORY_LABELS.get(category, category)
                payment_type = entry.get("payment_type") or ""
                payment_label = EXPENSE_PAYMENT_LABELS.get(payment_type, payment_type)
                ws.cell(row=idx, column=1, value=entry.get("expense_code") or "")
                ws.cell(row=idx, column=2, value=entry.get("sales_date") or "")
                ws.cell(row=idx, column=3, value=entry.get("description") or "")
                ws.cell(row=idx, column=4, value=category_label)
                ws.cell(row=idx, column=5, value=entry.get("supplier_name") or "")
                ws.cell(row=idx, column=6, value=entry.get("supplier_gst") or "")
                ws.cell(row=idx, column=7, value=payment_label)
                ws.cell(row=idx, column=8, value=round_half_up(entry.get("amount"), 2))
                ws.cell(row=idx, column=9, value=round_half_up(entry.get("paid_amount"), 2))
                ws.cell(row=idx, column=10, value=round_half_up(entry.get("balance"), 2))
            fname = (
                f"purchase_verification_pending_"
                f"{date_from.isoformat() if date_filter_active else 'All'}_to_"
                f"{date_to.isoformat() if date_filter_active else 'All'}.xlsx"
            )
    finally:
        conn.close()

    for column_cells in ws.columns:
        width = 12
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, 40))
        ws.column_dimensions[column_cells[0].column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/accounts/purchase-verification/create", methods=["POST"])
def create_purchase_verification():
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "You must be logged in to record a verification."}), 401
    if not user_can_approve_transactions(user):
        return jsonify({
            "ok": False,
            "error": "You do not have Approval access to verify purchases.",
        }), 403
    data = request.get_json(silent=True) or {}
    conn = get_db()
    try:
        payload, errors = _validate_purchase_verification_payload(conn, data, user=user)
        if errors:
            return jsonify({"ok": False, "error": errors[0], "errors": errors}), 400
        cursor = conn.execute(
            """INSERT INTO purchase_verifications
               (company, supplier_id, verification_date, verification_method, verification_account,
                transaction_id, total_amount, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                payload["company"],
                payload["supplier_id"],
                payload["verification_date"],
                payload["verification_method"],
                payload["verification_account"],
                payload["transaction_id"],
                payload["total_amount"],
                payload["notes"],
            ),
        )
        verification_id = cursor.lastrowid
        for allocation in payload["allocations"]:
            conn.execute(
                """INSERT INTO purchase_verification_allocations
                   (purchase_verification_id, expense_id, amount)
                   VALUES (?, ?, ?)""",
                (verification_id, allocation["expense_id"], allocation["amount"]),
            )
        conn.commit()
        verification = _purchase_verification_detail(conn, verification_id)
    finally:
        conn.close()

    return jsonify({"ok": True, "payment": verification})


@app.route("/accounts/purchase-verification/delete", methods=["POST"])
def delete_purchase_verification():
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "You must be logged in to revert a verification."}), 401
    if not user_can_approve_transactions(user):
        return jsonify({
            "ok": False,
            "error": "You do not have Approval access to revert verifications.",
        }), 403
    data = request.get_json(silent=True) or {}
    try:
        verification_id = int(data.get("payment_id"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Verification id is required."}), 400

    conn = get_db()
    try:
        verification = conn.execute(
            "SELECT id FROM purchase_verifications WHERE id = ?",
            (verification_id,),
        ).fetchone()
        if not verification:
            return jsonify({"ok": False, "error": "Verification was not found."}), 404
        conn.execute(
            "DELETE FROM purchase_verification_allocations WHERE purchase_verification_id = ?",
            (verification_id,),
        )
        conn.execute("DELETE FROM purchase_verifications WHERE id = ?", (verification_id,))
        conn.commit()
    finally:
        conn.close()

    return jsonify({"ok": True})


@app.route("/accounts/purchase-verification/<int:verification_id>")
def purchase_verification_detail(verification_id):
    conn = get_db()
    try:
        verification = _purchase_verification_detail(conn, verification_id)
    finally:
        conn.close()
    if not verification:
        return jsonify({"ok": False, "error": "Verification was not found."}), 404
    return jsonify({"ok": True, "payment": verification})


@app.route("/accounts/credit-payment/create", methods=["POST"])
def create_credit_payment():
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "You must be logged in to record a payment."}), 401
    if not user_can_approve_transactions(user):
        return jsonify({
            "ok": False,
            "error": "You do not have Approval access to clear payments.",
        }), 403
    data = request.get_json(silent=True) or {}
    conn = get_db()
    try:
        payload, errors = _validate_credit_payment_payload(conn, data)
        if errors:
            return jsonify({"ok": False, "error": errors[0], "errors": errors}), 400
        cursor = conn.execute(
            """INSERT INTO credit_payments
               (company, supplier_id, payment_date, payment_method, transaction_id, total_amount, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (
                payload["company"],
                payload["supplier_id"],
                payload["payment_date"],
                payload["payment_method"],
                payload["transaction_id"],
                payload["total_amount"],
                payload["notes"],
            ),
        )
        payment_id = cursor.lastrowid
        affected_expense_ids = []
        for allocation in payload["allocations"]:
            conn.execute(
                """INSERT INTO credit_payment_allocations (credit_payment_id, expense_id, amount)
                   VALUES (?, ?, ?)""",
                (payment_id, allocation["expense_id"], allocation["amount"]),
            )
            affected_expense_ids.append(allocation["expense_id"])
        for expense_id in affected_expense_ids:
            _sync_expense_payment_after_clearance(conn, expense_id)
        conn.commit()
        payment = _credit_payment_detail(conn, payment_id)
    finally:
        conn.close()

    return jsonify({"ok": True, "payment": payment})


@app.route("/accounts/credit-payment/delete", methods=["POST"])
def delete_credit_payment():
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "You must be logged in to revert a payment."}), 401
    if not user_can_approve_transactions(user):
        return jsonify({
            "ok": False,
            "error": "You do not have Approval access to revert payments.",
        }), 403
    data = request.get_json(silent=True) or {}
    try:
        payment_id = int(data.get("payment_id"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Payment id is required."}), 400

    conn = get_db()
    try:
        payment = conn.execute(
            "SELECT id FROM credit_payments WHERE id = ?",
            (payment_id,),
        ).fetchone()
        if not payment:
            return jsonify({"ok": False, "error": "Payment was not found."}), 404
        allocation_rows = conn.execute(
            "SELECT expense_id FROM credit_payment_allocations WHERE credit_payment_id = ?",
            (payment_id,),
        ).fetchall()
        affected_expense_ids = [row["expense_id"] for row in allocation_rows]
        conn.execute(
            "DELETE FROM credit_payment_allocations WHERE credit_payment_id = ?",
            (payment_id,),
        )
        conn.execute("DELETE FROM credit_payments WHERE id = ?", (payment_id,))
        for expense_id in affected_expense_ids:
            _restore_expense_credit_on_payment_delete(conn, expense_id)
        conn.commit()
    finally:
        conn.close()

    return jsonify({"ok": True})


@app.route("/accounts/credit-payment/<int:payment_id>")
def credit_payment_detail(payment_id):
    conn = get_db()
    try:
        payment = _credit_payment_detail(conn, payment_id)
    finally:
        conn.close()
    if not payment:
        return jsonify({"ok": False, "error": "Payment was not found."}), 404
    return jsonify({"ok": True, "payment": payment})


@app.route("/sales_update/hotel")
def sales_update_hotel():
    user = get_current_user()
    selected_company = request.args.get("company", DEFAULT_COMPANY)
    selected_location = request.args.get("location", OUTLET_HOTEL)
    selected_date = request.args.get("date", date.today().isoformat())
    today_iso = date.today().isoformat()

    if selected_company not in SALES_COMPANY_LOCATIONS:
        selected_company = DEFAULT_COMPANY
    if selected_location not in HOTEL_LOCATIONS:
        selected_location = OUTLET_HOTEL

    conn = get_db()
    try:
        entry_date = _parse_sales_date(selected_date)
        outlet_records = {
            OUTLET_HOTEL: _load_outlet_entry_bundle(
                conn, user, selected_company, OUTLET_HOTEL, selected_date, today_iso
            )
        }
        kpi_bundle = _sales_report_kpi_bundle(conn, entry_date, entry_date, selected_company, selected_location)
        suppliers = _all_suppliers(conn)
        tip_employees = _active_employees_for_tips(conn)
        available_cash = _cash_ledger_available_as_of(conn, selected_company, entry_date)
    finally:
        conn.close()

    hotel_outlet = outlet_records[OUTLET_HOTEL]
    return render_template(
        "sales_update_hotel.html",
        selected_company=selected_company,
        selected_company_label=SALES_COMPANY_LOCATIONS[selected_company]["label"],
        selected_location=selected_location,
        selected_locations=HOTEL_LOCATIONS,
        selected_date=selected_date,
        max_sales_date=today_iso,
        sales_company_locations=SALES_COMPANY_LOCATIONS,
        petty_cash_denominations=PETTY_CASH_DENOMINATIONS,
        outlet_records=outlet_records,
        sales_entry_locked=hotel_outlet["sales_entry_locked"],
        sales_update_is_admin=user.get("is_admin", False),
        hotel_sales_entry_fields=HOTEL_SALES_ENTRY_FIELDS,
        hotel_manual_sales_entry_keys=HOTEL_MANUAL_SALES_ENTRY_KEYS,
        expense_payment_types=EXPENSE_PAYMENT_TYPES,
        expense_categories=EXPENSE_CATEGORIES,
        suppliers=suppliers,
        tip_employees=tip_employees,
        available_cash=available_cash,
        cash_date_from=selected_date,
        cash_date_to=selected_date,
        cash_panel=False,
        kpi=kpi_bundle["current"],
        kpi_trends=kpi_bundle["trends"],
        kpi_vs_label=kpi_bundle["vs_label"],
        de_nav_section="analytics",
        de_nav_sales_view="hotel",
    )


@app.route("/sales_update/hotel/upload_report", methods=["POST"])
def upload_hotel_occupancy_report():
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "Not authenticated"}), 401

    company = request.form.get("company", DEFAULT_COMPANY)
    location = request.form.get("location", OUTLET_HOTEL)
    sales_date_str = (request.form.get("date") or date.today().isoformat()).strip()
    if location not in HOTEL_LOCATIONS:
        location = OUTLET_HOTEL

    upload = request.files.get("report_file")
    if not upload or not upload.filename:
        return jsonify({"ok": False, "error": "Please choose an FO Invoice Tax report."}), 400

    try:
        parsed = parse_fo_invoice_tax_report(upload.stream)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"Could not read report: {exc}"}), 400

    lines_by_date = parsed.get("lines_by_date") or {}
    if not lines_by_date:
        return jsonify({"ok": False, "error": "No invoice lines found in the FO Invoice Tax report."}), 400

    for report_date in sorted(lines_by_date):
        lock_error = _check_sales_date_lock(user, company, location, report_date)
        if lock_error:
            return jsonify({"ok": False, "error": f"{report_date}: {lock_error}"}), 403

    conn = get_db()
    results_by_date = {}
    try:
        for report_date, lines in sorted(lines_by_date.items()):
            replace_hotel_ledger_entries(conn, company, location, report_date, lines)
            conn.commit()
            results_by_date[report_date] = sync_hotel_sales_from_ledger(conn, user, company, location, report_date)
    finally:
        conn.close()

    meta = parsed.get("meta", {})
    imported_dates = meta.get("imported_dates") or sorted(lines_by_date)
    response_date = sales_date_str if sales_date_str in results_by_date else imported_dates[0]
    result = results_by_date[response_date]
    return jsonify({
        "ok": True,
        "date": response_date,
        "imported_dates": imported_dates,
        "message": f"Imported {meta.get('line_count', 0)} invoice lines for {', '.join(imported_dates)}",
        "sales_entries": result["sales_entries"],
        "ledger_rollup": rollup_hotel_ledger_entries(result["entries"]),
        "meta": meta,
    })


@app.route("/sales_update/hotel/save_ledger", methods=["POST"])
def save_hotel_ledger():
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "Not authenticated"}), 401

    data = request.get_json(silent=True) or {}
    company = data.get("company", DEFAULT_COMPANY)
    location = data.get("location", OUTLET_HOTEL)
    sales_date = data.get("date", date.today().isoformat())
    updates = data.get("updates") or []

    if location not in HOTEL_LOCATIONS:
        location = OUTLET_HOTEL

    lock_error = _check_sales_date_lock(user, company, location, sales_date)
    if lock_error:
        return jsonify({"ok": False, "error": lock_error}), 403

    allowed_modes = {mode for mode, _ in HOTEL_PAYMENT_MODES}
    conn = get_db()
    try:
        for item in updates:
            entry_id = item.get("id")
            if not entry_id:
                continue
            payment_mode = (item.get("payment_mode") or "").strip()
            if payment_mode not in allowed_modes:
                return jsonify({"ok": False, "error": "Invalid payment mode."}), 400
            conn.execute(
                """UPDATE hotel_sales_ledger_entries
                   SET payment_mode = ?, updated_at = datetime('now','localtime')
                   WHERE id = ? AND company = ? AND location = ? AND sales_date = ?""",
                (payment_mode, entry_id, company, location, sales_date),
            )
        conn.commit()
        result = sync_hotel_sales_from_ledger(conn, user, company, location, sales_date)
    finally:
        conn.close()

    return jsonify({
        "ok": True,
        "date": sales_date,
        "sales_entries": result["sales_entries"],
        "ledger_rollup": rollup_hotel_ledger_entries(result["entries"]),
        "difference": result["difference"],
    })


@app.route("/sales_update/hotel/clear_ledger", methods=["POST"])
def clear_hotel_ledger():
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "Not authenticated"}), 401

    data = request.get_json(silent=True) or {}
    company = data.get("company", DEFAULT_COMPANY)
    location = data.get("location", OUTLET_HOTEL)
    sales_date = data.get("date", date.today().isoformat())
    if location not in HOTEL_LOCATIONS:
        location = OUTLET_HOTEL

    lock_error = _check_sales_date_lock(user, company, location, sales_date)
    if lock_error:
        return jsonify({"ok": False, "error": lock_error}), 403

    conn = get_db()
    try:
        conn.execute(
            "DELETE FROM hotel_sales_ledger_entries WHERE company = ? AND location = ? AND sales_date = ?",
            (company, location, sales_date),
        )
        conn.commit()
        result = sync_hotel_sales_from_ledger(conn, user, company, location, sales_date)
    finally:
        conn.close()

    return jsonify({
        "ok": True,
        "date": sales_date,
        "sales_entries": result["sales_entries"],
        "ledger_rollup": rollup_hotel_ledger_entries([]),
    })


def _load_outlet_entry_bundle(conn, user, company, location, sales_date, today_iso):
    is_future = sales_date > today_iso
    row = None if is_future else load_sales_row(company, location, sales_date)
    sales_entry_locked = bool(row and sales_date < today_iso and not user.get("is_admin"))
    sales_entries = row.get("sales_entry_values", {}) if row else {}
    tip_entries = []
    tip_total = 0.0
    if location in HOTEL_LOCATIONS:
        sales_entries = build_hotel_sales_entry_values(sales_entries)
        expense_total = _sales_expense_total(conn, company, location, sales_date)
        sales_entries["expense"] = expense_total
        expense_entries = _sales_expense_entries(conn, company, location, sales_date)
    else:
        sales_entries = build_sales_entry_values(conn, company, location, sales_date, sales_entries)
        expense_entries = []
        expense_total = 0.0
    if location in TIP_OUTLET_LOCATIONS:
        tip_entries = _sales_tip_entries(conn, company, location, sales_date)
        tip_total = _sales_tip_total(conn, company, location, sales_date)
        _apply_tip_line_total(conn, company, location, sales_date, sales_entries)
    petty_cash_counts = row.get("petty_cash_counts", {}) if row else {}
    bundle = {
        "sales_entry_values": sales_entries,
        "sales_entry_total": get_sales_entry_total(sales_entries),
        "sales_entry_locked": sales_entry_locked,
        "petty_cash_counts": petty_cash_counts,
        "petty_cash_total": get_denomination_total(petty_cash_counts),
        "tip_entries": tip_entries,
        "tip_total": tip_total,
    }
    if location in HOTEL_LOCATIONS:
        bundle["expense_entries"] = expense_entries
        bundle["expense_total"] = expense_total
    return bundle


def _render_sales_update_outlet(user, outlet, sales_view, filter_endpoint):
    selected_company = request.args.get("company", DEFAULT_COMPANY)
    selected_date = request.args.get("date", date.today().isoformat())
    today_iso = date.today().isoformat()

    if selected_company not in SALES_COMPANY_LOCATIONS:
        selected_company = DEFAULT_COMPANY
    locations = SALES_COMPANY_LOCATIONS[selected_company]["locations"]
    if outlet not in locations:
        outlet = locations[0]

    selected_location = outlet
    selected_locations = [outlet]

    conn = get_db()
    try:
        outlet_records = {
            outlet: _load_outlet_entry_bundle(
                conn, user, selected_company, outlet, selected_date, today_iso
            )
        }
        cash_transfer_entries = _sales_cash_transfer_entries(conn, selected_company, selected_location, selected_date)
        cash_transfer_total = _sales_cash_transfer_total(conn, selected_company, selected_location, selected_date)
        tip_employees = _active_employees_for_tips(conn)
        entry_date = _parse_sales_date(selected_date)
        kpi_bundle = _sales_report_kpi_bundle(
            conn, entry_date, entry_date, selected_company, selected_location, difference_mode="cash_actual"
        )
    finally:
        conn.close()

    selected_outlet = outlet_records[selected_location]
    sales_entries = selected_outlet["sales_entry_values"]
    sales_entry_locked = selected_outlet["sales_entry_locked"]
    sales_entry_total = selected_outlet["sales_entry_total"]

    row = None if selected_date > today_iso else load_sales_row(selected_company, selected_location, selected_date)
    petty_cash_counts = row.get("petty_cash_counts", {}) if row else {}
    cash_denomination_counts = row.get("cash_denomination_counts", {}) if row else {}
    cash_available = parse_money(sales_entries.get("cash"))
    cash_unallocated = round_half_up(max(0.0, cash_available - cash_transfer_total), 2)

    sales_record = {
        "sales_entry_values": sales_entries,
        "petty_cash_counts": petty_cash_counts,
        "cash_denomination_counts": cash_denomination_counts,
        "cash_available": cash_available,
        "cash_transfer_total": cash_transfer_total,
        "cash_unallocated": cash_unallocated,
        "cash_transfer_entries": cash_transfer_entries,
        "staff_account_entries": [],
    }

    return render_template(
        "sales_update.html",
        page_title=f"Sales Update - {outlet}",
        page_subtitle=f"Upload reports and record daily {outlet} sales.",
        filter_form_action=url_for(filter_endpoint),
        hide_location_filter=True,
        selected_company=selected_company,
        selected_company_label=SALES_COMPANY_LOCATIONS[selected_company]["label"],
        selected_location=selected_location,
        selected_date=selected_date,
        selected_locations=selected_locations,
        max_sales_date=today_iso,
        sales_company_locations=SALES_COMPANY_LOCATIONS,
        sales_entry_fields=SALES_ENTRY_FIELDS,
        petty_cash_denominations=PETTY_CASH_DENOMINATIONS,
        sales_entry_total_keys=SALES_ENTRY_TOTAL_KEYS,
        sales_digital_transaction_keys=SALES_DIGITAL_TRANSACTION_KEYS,
        sales_cash_destinations=SALES_CASH_DESTINATIONS,
        sales_entry_locked=sales_entry_locked,
        sales_update_is_admin=user.get("is_admin", False),
        sales_record=sales_record,
        outlet_records=outlet_records,
        credit_employees=[],
        tip_employees=tip_employees,
        kpi=kpi_bundle["current"],
        kpi_trends=kpi_bundle["trends"],
        kpi_vs_label=kpi_bundle["vs_label"],
        kpi_is_single_day=kpi_bundle["is_single_day"],
        cash_panel=False,
        cash_date_from=selected_date,
        cash_date_to=selected_date,
        cash_transfer_day_collected=cash_available,
        cash_transfer_day_available=cash_unallocated,
        whatsapp_sales_report_configured=False,
        whatsapp_sales_report_company=None,
        sales_entry_total=sales_entry_total,
        de_nav_section="analytics",
        de_nav_sales_view=sales_view,
        kpi_fourth_metric="room_transfer",
        manual_sales_entry_keys=MANUAL_SALES_ENTRY_KEYS,
    )


@app.route("/sales_update")
@app.route("/sales_update/entry")
def sales_update_entry():
    return redirect(url_for("sales_update_bar", **request.args))


@app.route("/sales_update/bar")
def sales_update_bar():
    user = get_current_user()
    return _render_sales_update_outlet(user, OUTLET_BAR, "bar", "sales_update_bar")


@app.route("/sales_update/restaurant")
def sales_update_restaurant():
    user = get_current_user()
    return _render_sales_update_outlet(user, OUTLET_RESTAURANT, "restaurant", "sales_update_restaurant")


def _render_room_transfer_receivables_page(
    user,
    *,
    page_endpoint,
    create_payment_endpoint,
    reverse_payment_endpoint,
    page_title,
    page_subtitle,
    filter_locations,
    allowed_locations,
    nav_sales_view,
    template_name,
    receivables_panel_title,
    receivables_empty_noun,
    receivables_empty_import_hint,
    fixed_location=None,
):
    today = date.today()
    selected_company = request.args.get("company", DEFAULT_COMPANY)
    selected_payment_status = _normalize_room_transfer_filter_status(request.args.get("status"))
    if selected_payment_status == "all":
        selected_payment_status = "unpaid"
    selected_location = (
        fixed_location
        if fixed_location is not None
        else request.args.get("location", ROOM_TRANSFER_FILTER_ALL)
    )
    date_from, date_to, date_filter_active = _resolve_optional_filter_date_range(
        request.args, "date_from", "date_to", default_fy=True
    )

    if selected_company not in SALES_COMPANY_LOCATIONS:
        selected_company = DEFAULT_COMPANY
    if selected_location not in filter_locations:
        selected_location = ROOM_TRANSFER_FILTER_ALL

    filter_date_from = date_from.isoformat() if date_filter_active else ""
    filter_date_to = date_to.isoformat() if date_filter_active else ""

    conn = get_db()
    try:
        entries = load_room_transfer_entries_by_status(
            conn,
            selected_company,
            selected_payment_status,
            selected_location,
            date_from=date_from if date_filter_active else None,
            date_to=date_to if date_filter_active else None,
            allowed_locations=allowed_locations,
        )
        rollup = rollup_room_transfer_entries(entries)
        summary_entries = load_room_transfer_entries_by_status(
            conn,
            selected_company,
            "all",
            selected_location,
            date_from=date_from if date_filter_active else None,
            date_to=date_to if date_filter_active else None,
            allowed_locations=allowed_locations,
        )
        summary_rollup = rollup_room_transfer_entries(summary_entries)
    finally:
        conn.close()

    status_tab_query = {
        "company": selected_company,
    }
    if fixed_location is None:
        status_tab_query["location"] = selected_location
    if date_filter_active:
        status_tab_query["date_from"] = filter_date_from
        status_tab_query["date_to"] = filter_date_to

    return render_template(
        template_name,
        page_title=page_title,
        page_subtitle=page_subtitle,
        filter_form_action=url_for(page_endpoint),
        create_room_transfer_payment_url=url_for(create_payment_endpoint),
        reverse_room_transfer_payment_url=url_for(reverse_payment_endpoint),
        selected_company=selected_company,
        selected_company_label=SALES_COMPANY_LOCATIONS[selected_company]["label"],
        selected_payment_status=selected_payment_status,
        selected_location=selected_location,
        date_from=filter_date_from,
        date_to=filter_date_to,
        active_date_filter=date_filter_active,
        today_iso=today.isoformat(),
        room_transfer_filter_statuses=ROOM_TRANSFER_FILTER_STATUSES,
        room_transfer_filter_locations=filter_locations,
        room_transfer_status_tab_query=status_tab_query,
        room_transfer_entries=entries,
        room_transfer_rollup=rollup,
        room_transfer_summary_rollup=summary_rollup,
        room_transfer_payment_statuses=ROOM_TRANSFER_PAYMENT_STATUSES,
        room_transfer_payment_methods=ROOM_TRANSFER_PAYMENT_METHODS,
        receivables_page_endpoint=page_endpoint,
        receivables_panel_title=receivables_panel_title,
        receivables_empty_noun=receivables_empty_noun,
        receivables_empty_import_hint=receivables_empty_import_hint,
        sales_update_is_admin=user.get("is_admin", False),
        de_nav_section="analytics",
        de_nav_sales_view=nav_sales_view,
    )


def _create_room_transfer_payment_response():
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "Not authenticated"}), 401

    data = request.get_json(silent=True) or {}
    conn = get_db()
    payment_ids = []
    try:
        payload, errors = _validate_room_transfer_payment_payload(conn, data)
        if errors:
            return jsonify({"ok": False, "error": errors[0], "errors": errors}), 400
        touched_entry_ids = set()
        for split in payload["payment_splits"]:
            cursor = conn.execute(
                """INSERT INTO room_transfer_payments
                   (company, payment_date, payment_method, transaction_id, total_amount, notes)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (
                    payload["company"],
                    payload["payment_date"],
                    split["payment_method"],
                    split["transaction_id"],
                    split["amount"],
                    payload["notes"],
                ),
            )
            payment_id = cursor.lastrowid
            payment_ids.append(payment_id)
            split_allocations = _proportion_room_transfer_allocations(
                payload["allocations"],
                split["amount"],
            )
            for allocation in split_allocations:
                entry = allocation["entry"]
                conn.execute(
                    """INSERT INTO room_transfer_payment_allocations
                       (room_transfer_payment_id, room_transfer_entry_id, amount,
                        invoice_number, guest_name, location, sales_date)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (
                        payment_id,
                        allocation["entry_id"],
                        allocation["amount"],
                        entry.get("invoice_number") or "",
                        entry.get("guest_name") or "",
                        entry.get("location") or "",
                        entry.get("sales_date") or "",
                    ),
                )
                touched_entry_ids.add(allocation["entry_id"])
        for entry_id in touched_entry_ids:
            _sync_room_transfer_status_after_payment(conn, entry_id)
        conn.commit()
    finally:
        conn.close()

    return jsonify({
        "ok": True,
        "payment_id": payment_ids[0] if payment_ids else None,
        "payment_ids": payment_ids,
    })


def _reverse_room_transfer_payment_response(empty_selection_error, none_paid_error):
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "Not authenticated"}), 401

    data = request.get_json(silent=True) or {}
    company = str(data.get("company") or DEFAULT_COMPANY).strip() or DEFAULT_COMPANY
    raw_ids = data.get("entry_ids") or []
    if not isinstance(raw_ids, list) or not raw_ids:
        return jsonify({"ok": False, "error": empty_selection_error}), 400

    conn = get_db()
    try:
        valid_ids = []
        for raw in raw_ids:
            try:
                entry_id = int(raw)
            except (TypeError, ValueError):
                continue
            row = conn.execute(
                "SELECT id, company, payment_status FROM room_transfer_entries WHERE id = ?",
                (entry_id,),
            ).fetchone()
            if not row or row["company"] != company:
                continue
            if row["payment_status"] != "paid" and _room_transfer_entry_paid_total(conn, entry_id) <= 0:
                continue
            valid_ids.append(entry_id)
        if not valid_ids:
            return jsonify({"ok": False, "error": none_paid_error}), 400
        reversed_ids = _reverse_room_transfer_entry_payments(conn, valid_ids)
        conn.commit()
    finally:
        conn.close()

    return jsonify({"ok": True, "entry_ids": reversed_ids})


@app.route("/sales_update/room_transfer")
def sales_update_room_transfer():
    user = get_current_user()
    return _render_room_transfer_receivables_page(
        user,
        page_endpoint="sales_update_room_transfer",
        create_payment_endpoint="create_room_transfer_payment",
        reverse_payment_endpoint="reverse_room_transfer_payment",
        page_title="Room Transfer",
        page_subtitle="Room credit lines from Collections reports. Clear payment to record how each settlement was made.",
        filter_locations=ROOM_TRANSFER_FILTER_LOCATIONS,
        allowed_locations=ROOM_TRANSFER_OUTLET_LOCATIONS,
        nav_sales_view="room_transfer",
        template_name="sales_update_room_transfer.html",
        receivables_panel_title="Room transfers",
        receivables_empty_noun="room transfers",
        receivables_empty_import_hint="Upload a Collections report on Sales Update - Bar or Restaurant to import room credit lines.",
        fixed_location=ROOM_TRANSFER_FILTER_ALL,
    )


@app.route("/sales_update/credit")
def sales_update_credit():
    user = get_current_user()
    return _render_room_transfer_receivables_page(
        user,
        page_endpoint="sales_update_credit",
        create_payment_endpoint="create_sales_credit_payment",
        reverse_payment_endpoint="reverse_sales_credit_payment",
        page_title="Credit",
        page_subtitle="Hotel FO Invoice Tax credit lines from Sales Update. Clear payment to record how each settlement was made.",
        filter_locations=CREDIT_FILTER_LOCATIONS,
        allowed_locations=CREDIT_OUTLET_LOCATIONS,
        nav_sales_view="credit",
        template_name="sales_update_credit.html",
        receivables_panel_title="Credit",
        receivables_empty_noun="credits",
        receivables_empty_import_hint="Upload an FO Invoice Tax report on Sales Update - Hotel to import credit lines.",
    )


@app.route("/sales_update/room_transfer/create_payment", methods=["POST"])
def create_room_transfer_payment():
    return _create_room_transfer_payment_response()


@app.route("/sales_update/credit/create_payment", methods=["POST"])
def create_sales_credit_payment():
    return _create_room_transfer_payment_response()


@app.route("/sales_update/room_transfer/reverse_payment", methods=["POST"])
def reverse_room_transfer_payment():
    return _reverse_room_transfer_payment_response(
        "Select at least one room transfer.",
        "No paid room transfers found to reverse.",
    )


@app.route("/sales_update/credit/reverse_payment", methods=["POST"])
def reverse_sales_credit_payment():
    return _reverse_room_transfer_payment_response(
        "Select at least one credit entry.",
        "No paid credit entries found to reverse.",
    )


@app.route("/sales_update/room_transfer/save_status", methods=["POST"])
def save_room_transfer_status():
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "Not authenticated"}), 401

    data = request.get_json(silent=True) or {}
    company = data.get("company", DEFAULT_COMPANY)
    status_filter = _normalize_room_transfer_filter_status(data.get("status"))
    location_filter = data.get("location", ROOM_TRANSFER_FILTER_ALL)
    if location_filter not in ROOM_TRANSFER_FILTER_LOCATIONS:
        location_filter = ROOM_TRANSFER_FILTER_ALL
    today = date.today()
    default_from = today.replace(day=1)
    date_from = _parse_sales_date(data.get("date_from") or default_from.isoformat())
    date_to = _parse_sales_date(data.get("date_to") or today.isoformat())
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    updates = data.get("updates") or []
    allowed = {status for status, _ in ROOM_TRANSFER_PAYMENT_STATUSES}

    conn = get_db()
    try:
        for item in updates:
            entry_id = item.get("id")
            if not entry_id:
                continue
            payment_status = (item.get("payment_status") or "unpaid").strip().lower()
            if payment_status not in allowed:
                return jsonify({"ok": False, "error": "Invalid payment status."}), 400
            conn.execute(
                """UPDATE room_transfer_entries
                   SET payment_status = ?, updated_at = datetime('now','localtime')
                   WHERE id = ? AND company = ?""",
                (payment_status, entry_id, company),
            )
        conn.commit()
        entries = load_room_transfer_entries_by_status(
            conn,
            company,
            status_filter,
            location_filter,
            date_from=date_from,
            date_to=date_to,
        )
        rollup = rollup_room_transfer_entries(entries)
        summary_rollup = rollup_room_transfer_entries(
            load_room_transfer_entries_by_status(
                conn,
                company,
                "all",
                location_filter,
                date_from=date_from,
                date_to=date_to,
            )
        )
    finally:
        conn.close()

    return jsonify({
        "ok": True,
        "entries": entries,
        "rollup": rollup,
        "summary_rollup": summary_rollup,
        "status": status_filter,
        "location": location_filter,
    })


@app.route("/sales_update/save", methods=["POST"])
def save_sales_update():
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "Not authenticated"}), 401

    data = request.get_json(silent=True) or {}
    company = data.get("company", DEFAULT_COMPANY)
    location = data.get("location", DEFAULT_LOCATION)
    sales_date = data.get("date", date.today().isoformat())
    sales_entries = data.get("sales_entries", {})
    petty_cash_counts = data.get("petty_cash_counts", {})
    cash_denomination_counts = data.get("cash_denomination_counts", {})
    sales_only = bool(data.get("sales_only"))

    lock_error = _check_sales_date_lock(user, company, location, sales_date)
    if lock_error:
        return jsonify({"ok": False, "error": lock_error}), 403 if "administrator" in lock_error else 400

    conn = get_db()
    try:
        if location in HOTEL_LOCATIONS:
            sales_entries = build_hotel_sales_entry_values(sales_entries)
            sales_entries["expense"] = _sales_expense_total(conn, company, location, sales_date)
        else:
            sales_entries = build_sales_entry_values(conn, company, location, sales_date, sales_entries)
        _apply_tip_line_total(conn, company, location, sales_date, sales_entries)
    finally:
        conn.close()

    existing_row = load_sales_row(company, location, sales_date)
    if sales_only:
        petty_cash_counts = (existing_row or {}).get("petty_cash_counts", {})
        cash_denomination_counts = (existing_row or {}).get("cash_denomination_counts", {})
    elif not cash_denomination_counts:
        cash_denomination_counts = (existing_row or {}).get("cash_denomination_counts", {})

    if cash_denomination_counts:
        cash_total = get_denomination_total(cash_denomination_counts)
        if cash_total > 0:
            sales_entries["cash"] = round_half_up(cash_total, 2)

    totals = upsert_sales_row(user, company, location, sales_date, sales_entries, petty_cash_counts, cash_denomination_counts)

    return jsonify({
        "ok": True,
        "company": company,
        "location": location,
        "date": sales_date,
        "sales_entries": sales_entries,
        "sales_entry_total": totals["sales_entry_total"],
        "petty_cash_total": totals["petty_cash_total"],
    })


def _pick_report_fallback_date(available_dates, requested):
    """Choose the report date closest to the page selection when that day has no rows."""
    parsed_dates = []
    for item in available_dates or []:
        try:
            parsed_dates.append(date.fromisoformat(str(item).strip()))
        except (TypeError, ValueError):
            continue
    if not parsed_dates:
        return None
    return min(parsed_dates, key=lambda d: (abs((d - requested).days), -d.toordinal()))


@app.route("/sales_update/upload_report", methods=["POST"])
def upload_sales_report():
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "Not authenticated"}), 401

    sales_date_str = (request.form.get("date") or date.today().isoformat()).strip()
    requested_date = _parse_sales_date(sales_date_str)
    sales_date = requested_date
    active_location = (request.form.get("location") or DEFAULT_LOCATION).strip()
    upload = request.files.get("report_file")

    if not upload or not upload.filename:
        return jsonify({"ok": False, "error": "Please choose an Excel report file."}), 400

    file_bytes = upload.read()
    if not file_bytes:
        return jsonify({"ok": False, "error": "Please choose an Excel report file."}), 400

    def _parse_for(target_date):
        return parse_sales_report(io.BytesIO(file_bytes), target_date)

    try:
        parsed = _parse_for(sales_date)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"Could not read report: {exc}"}), 400

    meta = parsed.get("meta", {})
    imported_rows = int(meta.get("rows_bar") or 0) + int(meta.get("rows_restaurant") or 0)
    date_adjusted = False
    if imported_rows == 0:
        available = meta.get("available_dates") or []
        fallback = _pick_report_fallback_date(available, requested_date)
        if fallback is None:
            error = f"No sales rows found in the report for {requested_date.isoformat()}."
            error += " Check that the file is a Collections report with invoice lines."
            return jsonify({"ok": False, "error": error, "meta": meta}), 400
        sales_date = fallback
        date_adjusted = sales_date != requested_date
        try:
            parsed = _parse_for(sales_date)
        except Exception as exc:
            return jsonify({"ok": False, "error": f"Could not read report: {exc}"}), 400
        meta = parsed.get("meta", {})
        imported_rows = int(meta.get("rows_bar") or 0) + int(meta.get("rows_restaurant") or 0)
        if imported_rows == 0:
            error = f"No sales rows found in the report for {sales_date.isoformat()}."
            if available:
                error += f" Report contains data for: {', '.join(available)}."
            return jsonify({"ok": False, "error": error, "meta": meta}), 400

    company = DEFAULT_COMPANY
    results = {}

    for outlet in (OUTLET_BAR, OUTLET_RESTAURANT):
        existing_row = load_sales_row(company, outlet, sales_date.isoformat())
        existing_values = (existing_row or {}).get("sales_entry_values", {})
        merged = merge_import_into_sales_values(existing_values, parsed[outlet])

        conn = get_db()
        try:
            merged = build_sales_entry_values(conn, company, outlet, sales_date.isoformat(), merged)
            _apply_tip_line_total(conn, company, outlet, sales_date.isoformat(), merged)
        finally:
            conn.close()

        petty = (existing_row or {}).get("petty_cash_counts", {})
        cash_denoms = (existing_row or {}).get("cash_denomination_counts", {})
        upsert_sales_row(user, company, outlet, sales_date.isoformat(), merged, petty, cash_denoms)
        results[outlet.lower()] = merged

    room_lines = parsed.get("room_transfer_lines") or []
    if room_lines:
        conn = get_db()
        try:
            sync_room_transfer_entries(conn, company, sales_date.isoformat(), room_lines)
            conn.commit()
        finally:
            conn.close()

    if date_adjusted:
        message = (
            f"Report is for {sales_date.isoformat()} — "
            f"Bar and Restaurant updated for that day (page was on {requested_date.isoformat()})."
        )
    else:
        message = f"Report imported — Bar and Restaurant updated for {sales_date.isoformat()}"

    return jsonify({
        "ok": True,
        "date": sales_date.isoformat(),
        "requested_date": requested_date.isoformat(),
        "date_adjusted": date_adjusted,
        "bar": results.get("bar", {}),
        "restaurant": results.get("restaurant", {}),
        "active_location": active_location,
        "meta": parsed.get("meta", {}),
        "message": message,
    })


@app.route("/sales_update/add_expense", methods=["POST"])
def sales_update_add_expense():
    user = get_current_user()
    data = request.get_json(silent=True) or {}
    conn = get_db()
    try:
        result, error = _create_sales_expense(
            conn,
            user,
            data,
            include_sales_totals=True,
        )
        if error:
            status = 403 if "Cannot save" in error or "already saved" in error else 400
            return jsonify({"ok": False, "error": error}), status
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True, **result})


def _create_sales_expense(
    conn,
    user,
    data,
    *,
    default_location=None,
    include_sales_totals=False,
    allow_shared_invoice=False,
    skip_cash_check=False,
):
    company = data.get("company", DEFAULT_COMPANY)
    location = data.get("location", default_location or DEFAULT_LOCATION)
    sales_date = data.get("date", "")
    description = (data.get("description") or "").strip()
    amount = parse_money(data.get("amount"))
    raw_payment_type = (data.get("payment_type") or "").strip()
    if not raw_payment_type:
        return None, "Please select a payment type."
    payment_type = _normalize_expense_payment_type(raw_payment_type)
    category = _normalize_expense_category(data.get("category"))
    transaction_id = (data.get("transaction_id") or "").strip()
    invoice_number = (data.get("invoice_number") or "").strip()
    supplier_id = data.get("supplier_id")

    lock_error = _check_sales_date_lock(user, company, location, sales_date)
    if lock_error:
        return None, lock_error

    if not description or amount <= 0:
        return None, "Description and positive amount are required."
    if not supplier_id:
        return None, "Please select a supplier."
    if not category:
        return None, "Please select a category."
    if payment_type == EXPENSE_PAYMENT_BANK and not transaction_id:
        return None, "Transaction ID is required for bank transfer."
    if payment_type != EXPENSE_PAYMENT_BANK:
        transaction_id = ""

    supplier = _get_supplier(conn, supplier_id)
    if not supplier:
        return None, "Selected supplier was not found."

    if not skip_cash_check:
        cash_error = _validate_cash_expense_against_available(
            conn, company, sales_date, amount, payment_type
        )
        if cash_error:
            return None, cash_error

    if not allow_shared_invoice:
        duplicate = _duplicate_expense_invoice(conn, supplier_id, invoice_number)
        if duplicate:
            code = duplicate["expense_code"] or f"#{duplicate['id']}"
            return None, f"An expense with this supplier and invoice number already exists ({code})."

    expense_code = _next_expense_code(conn, company)
    cursor = conn.execute(
        """INSERT INTO sales_update_expenses
           (company, location, sales_date, description, amount, payment_type, transaction_id, supplier_id, category, expense_code, invoice_number)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (company, location, sales_date, description, amount, payment_type, transaction_id, supplier_id, category, expense_code, invoice_number),
    )
    expense_id = cursor.lastrowid
    result = {
        "expense_id": expense_id,
        "expense_code": expense_code,
        "sales_date": sales_date,
        "category": category,
        "amount": amount,
    }
    if include_sales_totals:
        result["expense_total"] = _sales_expense_total(conn, company, location, sales_date)
        result["expense_entries"] = _sales_expense_entries(conn, company, location, sales_date)
    return result, None


@app.route("/sales_update/edit_expense", methods=["POST"])
def sales_update_edit_expense():
    user = get_current_user()
    data = request.get_json(silent=True) or {}
    expense_id = data.get("id") or data.get("expense_id")
    company = data.get("company", DEFAULT_COMPANY)
    location = data.get("location", DEFAULT_LOCATION)
    sales_date = data.get("date", "")
    description = (data.get("description") or "").strip()
    amount = parse_money(data.get("amount"))
    raw_payment_type = (data.get("payment_type") or "").strip()
    if not raw_payment_type:
        return jsonify({"ok": False, "error": "Please select a payment type."}), 400
    payment_type = _normalize_expense_payment_type(raw_payment_type)
    category = _normalize_expense_category(data.get("category"))
    transaction_id = (data.get("transaction_id") or "").strip()
    invoice_number = (data.get("invoice_number") or "").strip()
    supplier_id = data.get("supplier_id")

    lock_error = _check_sales_date_lock(user, company, location, sales_date)
    if lock_error:
        return jsonify({"ok": False, "error": lock_error}), 403

    if not description or amount <= 0:
        return jsonify({"ok": False, "error": "Description and positive amount are required."}), 400
    if not supplier_id:
        return jsonify({"ok": False, "error": "Please select a supplier."}), 400
    if not category:
        return jsonify({"ok": False, "error": "Please select a category."}), 400
    if payment_type == EXPENSE_PAYMENT_BANK and not transaction_id:
        return jsonify({"ok": False, "error": "Transaction ID is required for bank transfer."}), 400
    if payment_type != EXPENSE_PAYMENT_BANK:
        transaction_id = ""

    conn = get_db()
    try:
        supplier = _get_supplier(conn, supplier_id)
        if not supplier:
            return jsonify({"ok": False, "error": "Selected supplier was not found."}), 400
        cash_error = _validate_cash_expense_against_available(
            conn,
            company,
            sales_date,
            amount,
            payment_type,
            exclude_expense_id=expense_id,
        )
        if cash_error:
            return jsonify({"ok": False, "error": cash_error}), 400
        duplicate = _duplicate_expense_invoice(
            conn, supplier_id, invoice_number, exclude_expense_id=expense_id
        )
        if duplicate:
            code = duplicate["expense_code"] or f"#{duplicate['id']}"
            return jsonify({
                "ok": False,
                "error": f"An expense with this supplier and invoice number already exists ({code}).",
            }), 400
        conn.execute(
            """UPDATE sales_update_expenses
               SET description=?, amount=?, payment_type=?, transaction_id=?, supplier_id=?, category=?,
                   invoice_number=?, updated_at=datetime('now','localtime')
               WHERE id=? AND company=? AND location=? AND sales_date=?""",
            (
                description, amount, payment_type, transaction_id, supplier_id, category,
                invoice_number, expense_id, company, location, sales_date,
            ),
        )
        conn.commit()
        expense_total = _sales_expense_total(conn, company, location, sales_date)
        expense_entries = _sales_expense_entries(conn, company, location, sales_date)
    finally:
        conn.close()

    return jsonify({"ok": True, "expense_total": expense_total, "expense_entries": expense_entries})


@app.route("/sales_update/delete_expense", methods=["POST"])
def sales_update_delete_expense():
    user = get_current_user()
    data = request.get_json(silent=True) or {}
    expense_id = data.get("id") or data.get("expense_id")
    company = data.get("company", DEFAULT_COMPANY)
    location = data.get("location", DEFAULT_LOCATION)
    sales_date = data.get("date", "")

    lock_error = _check_sales_date_lock(user, company, location, sales_date)
    if lock_error:
        return jsonify({"ok": False, "error": lock_error}), 403

    conn = get_db()
    try:
        conn.execute(
            "DELETE FROM sales_update_expenses WHERE id=? AND company=? AND location=? AND sales_date=?",
            (expense_id, company, location, sales_date),
        )
        conn.commit()
        expense_total = _sales_expense_total(conn, company, location, sales_date)
        expense_entries = _sales_expense_entries(conn, company, location, sales_date)
    finally:
        conn.close()

    return jsonify({"ok": True, "expense_total": expense_total, "expense_entries": expense_entries})


@app.route("/sales_update/add_tip", methods=["POST"])
def sales_update_add_tip():
    user = get_current_user()
    data = request.get_json(silent=True) or {}
    company = data.get("company", DEFAULT_COMPANY)
    location = data.get("location", OUTLET_HOTEL)
    sales_date = data.get("date", "")
    description = (data.get("description") or "").strip()
    amount = parse_money(data.get("amount"))
    try:
        employee_id = int(data.get("employee_id") or 0)
    except (TypeError, ValueError):
        employee_id = 0

    if location not in TIP_OUTLET_LOCATIONS:
        return jsonify({"ok": False, "error": "Tips can only be recorded for Hotel, Bar, or Restaurant."}), 400

    lock_error = _check_sales_date_lock(user, company, location, sales_date)
    if lock_error:
        return jsonify({"ok": False, "error": lock_error}), 403

    if employee_id <= 0:
        return jsonify({"ok": False, "error": "Please select an employee."}), 400
    if amount <= 0:
        return jsonify({"ok": False, "error": "Please enter a tip amount greater than 0."}), 400

    conn = get_db()
    try:
        payroll_lock_error = _check_payroll_month_date_lock(conn, sales_date)
        if payroll_lock_error:
            return jsonify({"ok": False, "error": payroll_lock_error, "locked": True}), 403
        employee = conn.execute(
            "SELECT id FROM employees WHERE id = ? AND status = 'active'",
            (employee_id,),
        ).fetchone()
        if not employee:
            return jsonify({"ok": False, "error": "Selected employee was not found."}), 400
        cursor = conn.execute(
            """INSERT INTO sales_update_tips
               (company, location, sales_date, employee_id, amount, description)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (company, location, sales_date, employee_id, amount, description),
        )
        tip_id = cursor.lastrowid
        conn.commit()
        tip_total = _sales_tip_total(conn, company, location, sales_date)
        tip_entries = _sales_tip_entries(conn, company, location, sales_date)
    finally:
        conn.close()

    return jsonify({
        "ok": True,
        "tip_id": tip_id,
        "tip_total": tip_total,
        "tip_entries": tip_entries,
    })


@app.route("/sales_update/edit_tip", methods=["POST"])
def sales_update_edit_tip():
    user = get_current_user()
    data = request.get_json(silent=True) or {}
    tip_id = data.get("id") or data.get("tip_id")
    company = data.get("company", DEFAULT_COMPANY)
    location = data.get("location", OUTLET_HOTEL)
    sales_date = data.get("date", "")
    description = (data.get("description") or "").strip()
    amount = parse_money(data.get("amount"))
    try:
        employee_id = int(data.get("employee_id") or 0)
    except (TypeError, ValueError):
        employee_id = 0

    if location not in TIP_OUTLET_LOCATIONS:
        return jsonify({"ok": False, "error": "Tips can only be recorded for Hotel, Bar, or Restaurant."}), 400

    lock_error = _check_sales_date_lock(user, company, location, sales_date)
    if lock_error:
        return jsonify({"ok": False, "error": lock_error}), 403

    if not tip_id:
        return jsonify({"ok": False, "error": "Missing tip id."}), 400
    if employee_id <= 0:
        return jsonify({"ok": False, "error": "Please select an employee."}), 400
    if amount <= 0:
        return jsonify({"ok": False, "error": "Please enter a tip amount greater than 0."}), 400

    conn = get_db()
    try:
        payroll_lock_error = _check_payroll_month_date_lock(conn, sales_date)
        if payroll_lock_error:
            return jsonify({"ok": False, "error": payroll_lock_error, "locked": True}), 403
        employee = conn.execute(
            "SELECT id FROM employees WHERE id = ? AND status = 'active'",
            (employee_id,),
        ).fetchone()
        if not employee:
            return jsonify({"ok": False, "error": "Selected employee was not found."}), 400
        updated = conn.execute(
            """UPDATE sales_update_tips
               SET employee_id=?, amount=?, description=?, updated_at=datetime('now','localtime')
               WHERE id=? AND company=? AND location=? AND sales_date=?""",
            (employee_id, amount, description, tip_id, company, location, sales_date),
        )
        if updated.rowcount == 0:
            return jsonify({"ok": False, "error": "Tip entry was not found."}), 404
        conn.commit()
        tip_total = _sales_tip_total(conn, company, location, sales_date)
        tip_entries = _sales_tip_entries(conn, company, location, sales_date)
    finally:
        conn.close()

    return jsonify({
        "ok": True,
        "tip_id": tip_id,
        "tip_total": tip_total,
        "tip_entries": tip_entries,
    })


@app.route("/sales_update/delete_tip", methods=["POST"])
def sales_update_delete_tip():
    user = get_current_user()
    data = request.get_json(silent=True) or {}
    tip_id = data.get("id") or data.get("tip_id")
    company = data.get("company", DEFAULT_COMPANY)
    location = data.get("location", OUTLET_HOTEL)
    sales_date = data.get("date", "")

    if location not in TIP_OUTLET_LOCATIONS:
        return jsonify({"ok": False, "error": "Tips can only be recorded for Hotel, Bar, or Restaurant."}), 400

    lock_error = _check_sales_date_lock(user, company, location, sales_date)
    if lock_error:
        return jsonify({"ok": False, "error": lock_error}), 403

    conn = get_db()
    try:
        payroll_lock_error = _check_payroll_month_date_lock(conn, sales_date)
        if payroll_lock_error:
            return jsonify({"ok": False, "error": payroll_lock_error, "locked": True}), 403
        tip_row = conn.execute(
            """SELECT id, company, employee_id, sales_date
               FROM sales_update_tips
               WHERE id=? AND company=? AND location=? AND sales_date=?""",
            (tip_id, company, location, sales_date),
        ).fetchone()
        if not tip_row:
            return jsonify({"ok": False, "error": "Tip entry not found."}), 404
        conn.execute(
            "DELETE FROM sales_update_tips WHERE id=? AND company=? AND location=? AND sales_date=?",
            (tip_id, company, location, sales_date),
        )
        _reconcile_tip_incentive_after_tip_delete(
            conn,
            tip_row["company"],
            tip_row["sales_date"],
            tip_row["employee_id"],
        )
        conn.commit()
        tip_total = _sales_tip_total(conn, company, location, sales_date)
        tip_entries = _sales_tip_entries(conn, company, location, sales_date)
    finally:
        conn.close()

    return jsonify({"ok": True, "tip_total": tip_total, "tip_entries": tip_entries})


@app.route("/sales_update/tips/employee_lines", methods=["POST"])
def sales_update_tips_employee_lines():
    """List tip lines for one employee under the current Tips filters (for edit modal)."""
    data = request.get_json(silent=True) or {}
    company = data.get("company", DEFAULT_COMPANY)
    if company not in SALES_COMPANY_LOCATIONS:
        company = DEFAULT_COMPANY
    try:
        employee_id = int(data.get("employee_id") or 0)
    except (TypeError, ValueError):
        employee_id = 0
    if employee_id <= 0:
        return jsonify({"ok": False, "error": "Employee is required."}), 400

    location_filter = (data.get("location") or TIPS_FILTER_ALL).strip()
    if location_filter not in TIPS_FILTER_LOCATIONS:
        location_filter = TIPS_FILTER_ALL
    outlet_filter = location_filter if location_filter in TIP_OUTLET_LOCATIONS else None
    date_from_raw = (data.get("date_from") or "").strip()
    date_to_raw = (data.get("date_to") or "").strip()
    date_filter_active = bool(date_from_raw and date_to_raw)

    conn = get_db()
    try:
        emp = conn.execute(
            "SELECT id, name, emp_code FROM employees WHERE id=?",
            (employee_id,),
        ).fetchone()
        if not emp:
            return jsonify({"ok": False, "error": "Employee not found."}), 404

        params = [company, employee_id]
        sql = """
            SELECT id, company, location, sales_date, employee_id, amount, description
            FROM sales_update_tips
            WHERE company=? AND employee_id=?
        """
        if date_filter_active:
            sql += " AND sales_date >= ? AND sales_date <= ?"
            params.extend([date_from_raw, date_to_raw])
        if outlet_filter:
            sql += " AND location = ?"
            params.append(outlet_filter)
        sql += " ORDER BY sales_date DESC, id DESC"
        rows = conn.execute(sql, params).fetchall()
        lines = [{
            "id": int(row["id"]),
            "company": row["company"],
            "location": row["location"],
            "sales_date": row["sales_date"],
            "employee_id": int(row["employee_id"]),
            "amount": round_half_up(row["amount"], 2),
            "description": row["description"] or "",
        } for row in rows]
    finally:
        conn.close()

    return jsonify({
        "ok": True,
        "employee": {
            "id": int(emp["id"]),
            "name": emp["name"] or "Unknown",
            "emp_code": emp["emp_code"] or "",
        },
        "lines": lines,
    })


@app.route("/sales_update/tips/delete_employee", methods=["POST"])
def sales_update_tips_delete_employee():
    """Delete all tip lines for one employee in the current Tips filters; clear payouts."""
    user = get_current_user()
    data = request.get_json(silent=True) or {}
    company = data.get("company", DEFAULT_COMPANY)
    if company not in SALES_COMPANY_LOCATIONS:
        company = DEFAULT_COMPANY
    try:
        employee_id = int(data.get("employee_id") or 0)
    except (TypeError, ValueError):
        employee_id = 0
    if employee_id <= 0:
        return jsonify({"ok": False, "error": "Employee is required."}), 400

    location_filter = (data.get("location") or TIPS_FILTER_ALL).strip()
    if location_filter not in TIPS_FILTER_LOCATIONS:
        location_filter = TIPS_FILTER_ALL
    outlet_filter = location_filter if location_filter in TIP_OUTLET_LOCATIONS else None

    date_from_raw = (data.get("date_from") or "").strip()
    date_to_raw = (data.get("date_to") or "").strip()
    date_filter_active = bool(date_from_raw and date_to_raw)

    conn = get_db()
    try:
        params = [company, employee_id]
        sql = """
            SELECT id, sales_date
            FROM sales_update_tips
            WHERE company=? AND employee_id=?
        """
        if date_filter_active:
            sql += " AND sales_date >= ? AND sales_date <= ?"
            params.extend([date_from_raw, date_to_raw])
        if outlet_filter:
            sql += " AND location = ?"
            params.append(outlet_filter)
        tip_rows = conn.execute(sql, params).fetchall()
        if not tip_rows:
            return jsonify({"ok": False, "error": "No tip lines found for this employee."}), 404

        for tip in tip_rows:
            payroll_lock_error = _check_payroll_month_date_lock(conn, tip["sales_date"])
            if payroll_lock_error:
                return jsonify({"ok": False, "error": payroll_lock_error, "locked": True}), 403

        tip_ids = [int(tip["id"]) for tip in tip_rows]
        sales_dates = [tip["sales_date"] for tip in tip_rows]
        placeholders = ",".join("?" for _ in tip_ids)
        conn.execute(
            f"DELETE FROM sales_update_tips WHERE id IN ({placeholders})",
            tip_ids,
        )
        _reconcile_tip_incentive_for_dates(conn, company, employee_id, sales_dates)
        conn.commit()
        deleted = len(tip_ids)
    finally:
        conn.close()

    return jsonify({
        "ok": True,
        "deleted": deleted,
        "employee_id": employee_id,
        "message": f"Deleted {deleted} tip line{'s' if deleted != 1 else ''} and cleared related incentive payout.",
    })


@app.route("/sales_update/tips")
def sales_update_tips_page():
    user = get_current_user()
    selected_company = request.args.get("company", DEFAULT_COMPANY)
    if selected_company not in SALES_COMPANY_LOCATIONS:
        selected_company = DEFAULT_COMPANY

    today = date.today()
    date_from, date_to, date_filter_active = _resolve_optional_filter_date_range(
        request.args, "date_from", "date_to", default_fy=True
    )

    location_filter = (request.args.get("location") or TIPS_FILTER_ALL).strip()
    if location_filter not in TIPS_FILTER_LOCATIONS:
        location_filter = TIPS_FILTER_ALL
    outlet_filter = location_filter if location_filter in TIP_OUTLET_LOCATIONS else None

    conn = get_db()
    try:
        tips_bundle = _load_tips_analytics_bundle(
            conn,
            selected_company,
            date_from if date_filter_active else None,
            date_to if date_filter_active else None,
            outlet_filter,
        )
    finally:
        conn.close()

    filter_date_from = date_from.isoformat() if date_filter_active else ""
    filter_date_to = date_to.isoformat() if date_filter_active else ""
    clear_query = {"company": selected_company, "location": location_filter}
    report_kwargs = {"company": selected_company, "location": location_filter}
    if date_filter_active:
        report_kwargs["date_from"] = filter_date_from
        report_kwargs["date_to"] = filter_date_to

    payout_year, payout_month = today.year, today.month

    from_hub = (request.args.get("from_hub") or "").strip().lower()
    if from_hub == "reports":
        clear_query["from_hub"] = "reports"
        de_nav_section = "report"
        de_nav_payroll_view = ""
        de_nav_report_view = "home"
        tips_back_href = url_for("reports")
        tips_back_label = "Back to Reports"
    else:
        from_hub = ""
        de_nav_section = "payroll"
        de_nav_payroll_view = "tips"
        de_nav_report_view = ""
        tips_back_href = url_for("employees")
        tips_back_label = "Back to Employee Payroll"

    return render_template(
        "sales_update_tips.html",
        page_title="Tips",
        selected_company=selected_company,
        selected_company_label=SALES_COMPANY_LOCATIONS[selected_company]["label"],
        selected_location=location_filter,
        tips_filter_locations=TIPS_FILTER_LOCATIONS,
        date_from=filter_date_from,
        date_to=filter_date_to,
        date_filter_active=date_filter_active,
        today_iso=today.isoformat(),
        tips_employees=tips_bundle["employees"],
        tips_grand_total=tips_bundle["grand_total"],
        tips_hotel_total=tips_bundle["hotel_total"],
        tips_bar_total=tips_bundle["bar_total"],
        tips_restaurant_total=tips_bundle["restaurant_total"],
        filter_form_action=url_for("sales_update_tips_page"),
        tips_clear_url=url_for("sales_update_tips_page", **clear_query),
        tips_report_url=url_for("export_tips_report", **report_kwargs),
        tip_incentive_payout_url=url_for("tips_incentive_payout"),
        tips_delete_employee_url=url_for("sales_update_tips_delete_employee"),
        tips_employee_lines_url=url_for("sales_update_tips_employee_lines"),
        tips_edit_tip_url=url_for("sales_update_edit_tip"),
        payout_year=payout_year,
        payout_month=payout_month,
        back_href=tips_back_href,
        back_label=tips_back_label,
        from_hub=from_hub,
        de_nav_section=de_nav_section,
        de_nav_payroll_view=de_nav_payroll_view,
        de_nav_report_view=de_nav_report_view,
    )


@app.route("/sales_update/tips/report")
def export_tips_report():
    """Excel download of tip details and employee rollup for the selected filters."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    selected_company = request.args.get("company", DEFAULT_COMPANY)
    if selected_company not in SALES_COMPANY_LOCATIONS:
        selected_company = DEFAULT_COMPANY

    date_from, date_to, date_filter_active = _resolve_optional_filter_date_range(
        request.args, "date_from", "date_to", default_fy=True
    )
    location_filter = (request.args.get("location") or TIPS_FILTER_ALL).strip()
    if location_filter not in TIPS_FILTER_LOCATIONS:
        location_filter = TIPS_FILTER_ALL
    outlet_filter = location_filter if location_filter in TIP_OUTLET_LOCATIONS else None

    conn = get_db()
    try:
        tips_bundle = _load_tips_analytics_bundle(
            conn,
            selected_company,
            date_from if date_filter_active else None,
            date_to if date_filter_active else None,
            outlet_filter,
        )
        detail_entries = _load_tips_detail_entries(
            conn,
            selected_company,
            date_from if date_filter_active else None,
            date_to if date_filter_active else None,
            outlet_filter,
        )
    finally:
        conn.close()

    wb = Workbook()
    header_font = Font(bold=True)

    summary = wb.active
    summary.title = "Summary"
    summary_headers = ["Metric", "Amount"]
    for col, title in enumerate(summary_headers, start=1):
        cell = summary.cell(row=1, column=col, value=title)
        cell.font = header_font
    summary_rows = [
        ("Total Tips", tips_bundle["grand_total"]),
        ("Hotel", tips_bundle["hotel_total"]),
        ("Bar", tips_bundle["bar_total"]),
        ("Restaurant", tips_bundle["restaurant_total"]),
        ("Location", location_filter),
        ("Date From", date_from.isoformat() if date_filter_active else "All"),
        ("Date To", date_to.isoformat() if date_filter_active else "All"),
        ("Employees", len(tips_bundle["employees"])),
        ("Tip Lines", len(detail_entries)),
    ]
    for idx, (label, value) in enumerate(summary_rows, start=2):
        summary.cell(row=idx, column=1, value=label)
        summary.cell(row=idx, column=2, value=value)

    by_employee = wb.create_sheet("By Employee")
    emp_headers = ["Employee", "Emp Code", "Hotel", "Bar", "Restaurant", "Total"]
    for col, title in enumerate(emp_headers, start=1):
        cell = by_employee.cell(row=1, column=col, value=title)
        cell.font = header_font
    for idx, row in enumerate(tips_bundle["employees"], start=2):
        by_employee.cell(row=idx, column=1, value=row.get("employee_name") or "")
        by_employee.cell(row=idx, column=2, value=row.get("employee_code") or "")
        by_employee.cell(row=idx, column=3, value=round_half_up(row.get("hotel"), 2))
        by_employee.cell(row=idx, column=4, value=round_half_up(row.get("bar"), 2))
        by_employee.cell(row=idx, column=5, value=round_half_up(row.get("restaurant"), 2))
        by_employee.cell(row=idx, column=6, value=round_half_up(row.get("total"), 2))

    detail = wb.create_sheet("Tip Details")
    detail_headers = [
        "Date",
        "Employee",
        "Emp Code",
        "Location",
        "Amount",
        "Description",
    ]
    for col, title in enumerate(detail_headers, start=1):
        cell = detail.cell(row=1, column=col, value=title)
        cell.font = header_font
    for idx, entry in enumerate(detail_entries, start=2):
        detail.cell(row=idx, column=1, value=entry.get("sales_date") or "")
        detail.cell(row=idx, column=2, value=entry.get("employee_name") or "")
        detail.cell(row=idx, column=3, value=entry.get("employee_code") or "")
        detail.cell(row=idx, column=4, value=entry.get("location") or "")
        detail.cell(row=idx, column=5, value=round_half_up(entry.get("amount"), 2))
        detail.cell(row=idx, column=6, value=entry.get("description") or "")

    for ws in (summary, by_employee, detail):
        for column_cells in ws.columns:
            width = 12
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(value) + 2, 40))
            ws.column_dimensions[column_cells[0].column_letter].width = width

    if date_filter_active:
        fname = f"tips_report_{date_from.isoformat()}_to_{date_to.isoformat()}.xlsx"
    else:
        fname = "tips_report_all.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _tips_incentive_payout_payload(conn, company, year, month):
    """Build GET payload for Incentive Payout modal (pool + active employees)."""
    from employee_payroll import (
        _get_month_tip_incentive_map,
        _get_payroll_month_state,
        _period_label,
    )

    pool = _available_tip_pool_total(conn, company, year, month)
    payout_map = _get_month_tip_incentive_map(conn, year, month, company=company)
    employees = []
    allocated = 0.0
    for emp in _active_employees_for_tips(conn):
        amount = round_half_up(payout_map.get(emp["id"], 0), 2)
        allocated = round_half_up(allocated + amount, 2)
        employees.append({
            "id": emp["id"],
            "emp_code": emp.get("emp_code") or "",
            "name": emp.get("name") or "",
            "location": emp.get("location") or "",
            "amount": amount,
        })
    payroll_state = _get_payroll_month_state(conn, year, month)
    remaining = round_half_up(max(0.0, pool - allocated), 2)
    return {
        "ok": True,
        "company": company,
        "year": year,
        "month": month,
        "month_label": _period_label(year, month),
        "total_tips": pool,
        "allocated": allocated,
        "remaining": remaining,
        "can_edit": bool(payroll_state["can_edit"]),
        "locked": bool(payroll_state["locked"]),
        "message": payroll_state.get("message") or "",
        "employees": employees,
    }


@app.route("/sales_update/tips/incentive-payout", methods=["GET", "POST"])
def tips_incentive_payout():
    """Load or save monthly tip incentive allocations for active employees."""
    from employee_payroll import (
        _get_payroll_month_state,
        _parse_period_value,
        _upsert_month_tip_incentive,
    )

    get_current_user()
    if request.method == "GET":
        source = request.args
    else:
        source = request.get_json(silent=True) or {}

    selected_company = (source.get("company") or DEFAULT_COMPANY).strip() or DEFAULT_COMPANY
    if selected_company not in SALES_COMPANY_LOCATIONS:
        selected_company = DEFAULT_COMPANY

    from employee_payroll import _default_reporting_period

    default_year, default_month = _default_reporting_period()
    try:
        year, month = _parse_period_value(
            source.get("year", default_year),
            source.get("month", default_month),
        )
    except (TypeError, ValueError):
        year, month = default_year, default_month

    if request.method == "GET":
        conn = get_db()
        try:
            payload = _tips_incentive_payout_payload(conn, selected_company, year, month)
        finally:
            conn.close()
        return jsonify(payload)

    allocations_raw = source.get("allocations")
    if allocations_raw is None:
        allocations_raw = source.get("employees") or []
    if not isinstance(allocations_raw, list):
        return jsonify({"ok": False, "error": "Allocations must be a list."}), 400

    parsed = []
    for item in allocations_raw:
        if not isinstance(item, dict):
            return jsonify({"ok": False, "error": "Invalid allocation row."}), 400
        try:
            emp_id = int(item.get("employee_id", item.get("id")))
        except (TypeError, ValueError):
            return jsonify({"ok": False, "error": "Invalid employee id."}), 400
        amount = parse_money(item.get("amount"))
        if amount < 0:
            return jsonify({"ok": False, "error": "Incentive cannot be negative."}), 400
        parsed.append((emp_id, round_half_up(amount, 2)))

    allocated_total = round_half_up(sum(amount for _, amount in parsed), 2)

    conn = get_db()
    try:
        payroll_state = _get_payroll_month_state(conn, year, month)
        if payroll_state["locked"] or not payroll_state["can_edit"]:
            return jsonify({
                "ok": False,
                "error": payroll_state.get("message") or "This payroll month is read-only.",
                "locked": bool(payroll_state["locked"]),
            }), 403

        pool = _available_tip_pool_total(conn, selected_company, year, month)
        if allocated_total > pool + 1e-9:
            return jsonify({
                "ok": False,
                "error": (
                    f"Allocated ₹{allocated_total:,.2f} exceeds available tip pool "
                    f"₹{pool:,.2f}."
                ),
                "total_tips": pool,
                "allocated": allocated_total,
                "remaining": round_half_up(max(0.0, pool - allocated_total), 2),
            }), 400

        active_ids = {emp["id"] for emp in _active_employees_for_tips(conn)}
        for emp_id, amount in parsed:
            if emp_id not in active_ids:
                return jsonify({
                    "ok": False,
                    "error": f"Employee {emp_id} is not an active employee.",
                }), 400
            _upsert_month_tip_incentive(
                conn, selected_company, year, month, emp_id, amount
            )
        conn.commit()
        payload = _tips_incentive_payout_payload(conn, selected_company, year, month)
        payload["saved"] = True
        return jsonify(payload)
    finally:
        conn.close()


@app.route("/sales_update/add_unpaid_bill", methods=["POST"])
def sales_update_add_unpaid_bill():
    user = get_current_user()
    data = request.get_json(silent=True) or {}
    company = data.get("company", DEFAULT_COMPANY)
    location = data.get("location", DEFAULT_LOCATION)
    sales_date = data.get("date", "")
    invoice_number = (data.get("invoice_number") or "").strip()
    amount = parse_money(data.get("amount"))

    if not invoice_number or amount <= 0:
        return jsonify({"ok": False, "error": "Invoice number and positive amount are required."}), 400

    conn = get_db()
    try:
        conn.execute(
            """INSERT INTO sales_update_pending_bills
               (company, location, recorded_sales_date, invoice_number, amount, status)
               VALUES (?, ?, ?, ?, ?, 'open')""",
            (company, location, sales_date, invoice_number, amount),
        )
        conn.commit()
        total = _sales_unpaid_bill_total(conn, company, location, sales_date)
        entries = _sales_unpaid_bill_entries(conn, company, location, sales_date)
    finally:
        conn.close()

    return jsonify({"ok": True, "unpaid_pending_bill_total": total, "unpaid_bill_entries": entries})


@app.route("/sales_update/delete_unpaid_bill", methods=["POST"])
def sales_update_delete_unpaid_bill():
    user = get_current_user()
    data = request.get_json(silent=True) or {}
    bill_id = data.get("id")
    company = data.get("company", DEFAULT_COMPANY)
    location = data.get("location", DEFAULT_LOCATION)
    sales_date = data.get("date", "")

    conn = get_db()
    try:
        conn.execute(
            "DELETE FROM sales_update_pending_bills WHERE id=? AND company=? AND location=? AND recorded_sales_date=?",
            (bill_id, company, location, sales_date),
        )
        conn.commit()
        total = _sales_unpaid_bill_total(conn, company, location, sales_date)
        entries = _sales_unpaid_bill_entries(conn, company, location, sales_date)
    finally:
        conn.close()

    return jsonify({"ok": True, "unpaid_pending_bill_total": total, "unpaid_bill_entries": entries})


@app.route("/sales_update/open_pending_bills", methods=["GET"])
def sales_update_open_pending_bills():
    company = request.args.get("company", DEFAULT_COMPANY)
    location = request.args.get("location", DEFAULT_LOCATION)
    conn = get_db()
    try:
        bills = _sales_open_pending_bills(conn, company, location)
    finally:
        conn.close()
    return jsonify({"ok": True, "open_pending_bills": bills})


@app.route("/sales_update/add_bill_payment", methods=["POST"])
def sales_update_add_bill_payment():
    user = get_current_user()
    data = request.get_json(silent=True) or {}
    company = data.get("company", DEFAULT_COMPANY)
    location = data.get("location", DEFAULT_LOCATION)
    sales_date = data.get("date", "")
    pending_bill_id = data.get("pending_bill_id")
    amount = parse_money(data.get("amount"))

    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO sales_update_bill_payments (company, location, sales_date, pending_bill_id, amount) VALUES (?, ?, ?, ?, ?)",
            (company, location, sales_date, pending_bill_id, amount),
        )
        conn.execute(
            "UPDATE sales_update_pending_bills SET status='cleared', cleared_sales_date=? WHERE id=?",
            (sales_date, pending_bill_id),
        )
        conn.commit()
        total = _sales_bill_payment_total(conn, company, location, sales_date)
        entries = _sales_bill_payment_entries(conn, company, location, sales_date)
    finally:
        conn.close()

    return jsonify({"ok": True, "previous_bill_payment_total": total, "bill_payment_entries": entries})


@app.route("/sales_update/add_cash_transfer", methods=["POST"])
def sales_update_add_cash_transfer():
    user = get_current_user()
    data = request.get_json(silent=True) or {}
    company = data.get("company", DEFAULT_COMPANY)
    location = data.get("location", DEFAULT_LOCATION)
    sales_date = data.get("date", "")
    destination = (data.get("destination") or "bank").strip().lower()
    description = (data.get("description") or "").strip()
    amount = parse_money(data.get("amount"))

    if amount <= 0:
        return jsonify({"ok": False, "error": "Amount must be greater than zero."}), 400

    conn = get_db()
    try:
        conn.execute(
            """INSERT INTO sales_update_cash_transfers
               (company, location, sales_date, destination, description, amount)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (company, location, sales_date, destination, description, amount),
        )
        conn.commit()
        entries = _sales_cash_transfer_entries(conn, company, location, sales_date)
        total = _sales_cash_transfer_total(conn, company, location, sales_date)
    finally:
        conn.close()

    return jsonify({"ok": True, "cash_transfer_total": total, "cash_transfer_entries": entries})


@app.route("/sales_update/delete_cash_transfer", methods=["POST"])
def sales_update_delete_cash_transfer():
    user = get_current_user()
    data = request.get_json(silent=True) or {}
    transfer_id = data.get("id")
    company = data.get("company", DEFAULT_COMPANY)
    location = data.get("location", DEFAULT_LOCATION)
    sales_date = data.get("date", "")

    conn = get_db()
    try:
        conn.execute(
            "DELETE FROM sales_update_cash_transfers WHERE id=? AND company=? AND location=? AND sales_date=?",
            (transfer_id, company, location, sales_date),
        )
        conn.commit()
        entries = _sales_cash_transfer_entries(conn, company, location, sales_date)
        total = _sales_cash_transfer_total(conn, company, location, sales_date)
    finally:
        conn.close()

    return jsonify({"ok": True, "cash_transfer_total": total, "cash_transfer_entries": entries})


@app.route("/sales_update/delete_bill_payment", methods=["POST"])
def sales_update_delete_bill_payment():
    user = get_current_user()
    data = request.get_json(silent=True) or {}
    payment_id = data.get("id")
    company = data.get("company", DEFAULT_COMPANY)
    location = data.get("location", DEFAULT_LOCATION)
    sales_date = data.get("date", "")

    conn = get_db()
    try:
        row = conn.execute(
            "SELECT pending_bill_id FROM sales_update_bill_payments WHERE id=? AND company=? AND location=? AND sales_date=?",
            (payment_id, company, location, sales_date),
        ).fetchone()
        if row:
            conn.execute(
                "UPDATE sales_update_pending_bills SET status='open', cleared_sales_date=NULL WHERE id=?",
                (row["pending_bill_id"],),
            )
        conn.execute(
            "DELETE FROM sales_update_bill_payments WHERE id=? AND company=? AND location=? AND sales_date=?",
            (payment_id, company, location, sales_date),
        )
        conn.commit()
        total = _sales_bill_payment_total(conn, company, location, sales_date)
        entries = _sales_bill_payment_entries(conn, company, location, sales_date)
    finally:
        conn.close()

    return jsonify({"ok": True, "previous_bill_payment_total": total, "bill_payment_entries": entries})


@app.route("/sales_update/send_whatsapp_report", methods=["POST"])
def sales_update_send_whatsapp_report():
    return jsonify({"ok": False, "error": "WhatsApp report is not configured."}), 501


def _supplier_page_render(template, **kwargs):
    kwargs.setdefault("auth_notice", _pop_auth_notice())
    kwargs.setdefault("de_nav_section", "master")
    kwargs.setdefault("de_nav_master_view", "home")
    return render_template(template, **kwargs)


@app.route("/suppliers")
def supplier_master():
    user = get_current_user()
    if not user_can_access_supplier_master(user):
        return _permission_denied_response("You do not have access to Supplier Master.")

    selected_supplier_id = request.args.get("supplier_id", "").strip()
    saved_flag = request.args.get("saved", "").strip()
    form_focus = request.args.get("focus", "").strip() == "form"

    conn = get_db()
    try:
        suppliers = _all_suppliers(conn)
        selected_supplier = None
        if selected_supplier_id:
            selected_supplier = _get_supplier(conn, selected_supplier_id)
    finally:
        conn.close()

    form = selected_supplier or _supplier_form_payload()
    if selected_supplier:
        form = dict(form)
        form["id"] = selected_supplier["id"]
    else:
        form = {"id": "", **_supplier_form_payload()}

    success_message = ""
    if saved_flag == "created":
        success_message = "Supplier created successfully."
    elif saved_flag == "updated":
        success_message = "Supplier updated successfully."
    elif saved_flag == "deleted":
        success_message = "Supplier deleted successfully."

    return _supplier_page_render(
        "partials/master_embed/supplier.html" if is_embed_request() else "supplier_master.html",
        suppliers=suppliers,
        form=form,
        selected_supplier=selected_supplier,
        errors=[],
        success_message=success_message,
        form_focus=form_focus or bool(selected_supplier),
        show_form=form_focus or bool(selected_supplier),
        supplier_report_url=url_for("export_supplier_report"),
        embed_mode=is_embed_request(),
    )


@app.route("/suppliers/report")
def export_supplier_report():
    """Excel download of all suppliers from Supplier Master."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    user = get_current_user()
    if not user_can_access_supplier_master(user):
        return _permission_denied_response("You do not have access to Supplier Master.")

    conn = get_db()
    try:
        suppliers = _all_suppliers(conn)
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Suppliers"
    header_font = Font(bold=True)
    headers = [
        "Name",
        "GST",
        "Phone",
        "Address",
        "Bank",
        "Account Number",
        "IFSC",
    ]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font

    for idx, supplier in enumerate(suppliers, start=2):
        ws.cell(row=idx, column=1, value=supplier.get("name") or "")
        ws.cell(row=idx, column=2, value=supplier.get("gst") or "")
        ws.cell(row=idx, column=3, value=supplier.get("phone") or "")
        ws.cell(row=idx, column=4, value=supplier.get("address") or "")
        ws.cell(row=idx, column=5, value=supplier.get("bank_name") or "")
        ws.cell(row=idx, column=6, value=supplier.get("bank_account_number") or "")
        ws.cell(row=idx, column=7, value=supplier.get("ifsc_code") or "")

    for column_cells in ws.columns:
        width = 12
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, 40))
        ws.column_dimensions[column_cells[0].column_letter].width = width

    fname = f"supplier_report_{date.today().isoformat()}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/suppliers/save", methods=["POST"])
def save_supplier():
    user = get_current_user()
    if not user_can_access_supplier_master(user):
        return _permission_denied_response("You do not have access to Supplier Master.")

    supplier_id_raw = request.form.get("supplier_id", "").strip()
    supplier_id = int(supplier_id_raw) if supplier_id_raw else None
    payload = _supplier_form_payload(request.form)
    embed = is_embed_request() or str(request.form.get("embed") or request.args.get("embed") or "") == "1"

    conn = get_db()
    try:
        saved_id, errors = _save_supplier_record(conn, payload, supplier_id=supplier_id)
        if errors:
            suppliers = _all_suppliers(conn)
            selected_supplier = _get_supplier(conn, supplier_id) if supplier_id else None
            form = dict(payload)
            form["id"] = supplier_id or ""
            return _supplier_page_render(
                "partials/master_embed/supplier.html" if embed else "supplier_master.html",
                suppliers=suppliers,
                form=form,
                selected_supplier=selected_supplier,
                errors=errors,
                success_message="",
                form_focus=True,
                show_form=True,
                supplier_report_url=url_for("export_supplier_report"),
                embed_mode=embed,
            ), 400
        conn.commit()
    finally:
        conn.close()

    result_flag = "updated" if supplier_id else "created"
    redirect_kwargs = {"saved": result_flag}
    if embed:
        redirect_kwargs["embed"] = 1
    return redirect(url_for("supplier_master", **redirect_kwargs))


@app.route("/suppliers/delete", methods=["POST"])
def delete_supplier():
    user = get_current_user()
    if not user_can_access_supplier_master(user):
        return _permission_denied_response("You do not have access to Supplier Master.")

    supplier_id = request.form.get("supplier_id", "").strip()
    embed = is_embed_request() or str(request.form.get("embed") or request.args.get("embed") or "") == "1"
    redirect_kwargs = {}
    if embed:
        redirect_kwargs["embed"] = 1

    if not supplier_id:
        _queue_auth_notice("Supplier not found.")
        return redirect(url_for("supplier_master", **redirect_kwargs))

    conn = get_db()
    try:
        in_use = conn.execute(
            "SELECT COUNT(*) AS total FROM sales_update_expenses WHERE supplier_id = ?",
            (supplier_id,),
        ).fetchone()["total"]
        if in_use:
            _queue_auth_notice("This supplier cannot be deleted because it is linked to existing expenses.")
            fail_kwargs = dict(redirect_kwargs)
            fail_kwargs["supplier_id"] = supplier_id
            return redirect(url_for("supplier_master", **fail_kwargs))
        conn.execute("DELETE FROM suppliers WHERE id = ?", (supplier_id,))
        conn.commit()
    finally:
        conn.close()

    redirect_kwargs["saved"] = "deleted"
    return redirect(url_for("supplier_master", **redirect_kwargs))


@app.route("/suppliers/create", methods=["POST"])
def create_supplier():
    user = get_current_user()
    can_add = (
        user_can_access_supplier_master(user)
        or user_can_access_sales_analytics_submodule(user, "hotel")
        or user_can_access_dashboard(user, "accounts")
    )
    if not can_add:
        return jsonify({"ok": False, "error": "You do not have access to add suppliers."}), 403

    data = request.get_json(silent=True) or {}
    payload = _supplier_form_payload(data)

    conn = get_db()
    try:
        saved_id, errors = _save_supplier_record(conn, payload)
        if errors:
            return jsonify({"ok": False, "error": errors[0], "errors": errors}), 400
        conn.commit()
        supplier = _get_supplier(conn, saved_id)
        suppliers = _all_suppliers(conn)
    finally:
        conn.close()

    return jsonify({"ok": True, "supplier": supplier, "suppliers": suppliers})


def _customer_form_payload(source=None):
    source = source or {}
    mobile = "".join(ch for ch in str(source.get("mobile") or "") if ch.isdigit())[:10]
    return {
        "first_name": " ".join(str(source.get("first_name") or "").split()).strip(),
        "mobile": mobile,
        "email": " ".join(str(source.get("email") or "").split()).strip().lower(),
        "address": " ".join(str(source.get("address") or "").split()).strip(),
    }


def _customer_page_render(template, **kwargs):
    kwargs.setdefault("auth_notice", _pop_auth_notice())
    kwargs.setdefault("de_nav_section", "master")
    kwargs.setdefault("de_nav_master_view", "customer_master")
    return render_template(template, **kwargs)


@app.route("/customers")
def customer_master():
    user = get_current_user()
    if not user_can_access_customer_master(user):
        return _permission_denied_response("You do not have access to Customer Master.")

    selected_customer_id = request.args.get("customer_id", "").strip()
    saved_flag = request.args.get("saved", "").strip()
    form_focus = request.args.get("focus", "").strip() == "form"

    conn = get_db()
    try:
        ensure_customers_schema(conn)
        customers = list_customers(conn)
        selected_customer = get_customer(conn, selected_customer_id) if selected_customer_id else None
        conn.commit()
    finally:
        conn.close()

    if selected_customer:
        form = {
            "id": selected_customer["id"],
            "first_name": selected_customer.get("first_name") or "",
            "mobile": selected_customer.get("mobile") or "",
            "email": selected_customer.get("email") or "",
            "address": selected_customer.get("address") or "",
        }
    else:
        form = {"id": "", **_customer_form_payload()}

    success_message = ""
    if saved_flag == "created":
        success_message = "Customer created successfully."
    elif saved_flag == "updated":
        success_message = "Customer updated successfully."
    elif saved_flag == "deleted":
        success_message = "Customer deleted successfully."

    return _customer_page_render(
        "partials/master_embed/customer.html" if is_embed_request() else "customer_master.html",
        customers=customers,
        form=form,
        selected_customer=selected_customer,
        errors=[],
        success_message=success_message,
        form_focus=form_focus or bool(selected_customer),
        show_form=form_focus or bool(selected_customer),
        customer_report_url=url_for("export_customer_report"),
        embed_mode=is_embed_request(),
    )


@app.route("/customers/report")
def export_customer_report():
    """Excel download of all customers from Customer Master."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    user = get_current_user()
    if not user_can_access_customer_master(user):
        return _permission_denied_response("You do not have access to Customer Master.")

    conn = get_db()
    try:
        ensure_customers_schema(conn)
        customers = list_customers(conn)
        conn.commit()
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    header_font = Font(bold=True)
    headers = ["First Name", "Mobile", "Email", "Address"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font

    for idx, customer in enumerate(customers, start=2):
        ws.cell(row=idx, column=1, value=customer.get("first_name") or "")
        ws.cell(row=idx, column=2, value=customer.get("mobile") or "")
        ws.cell(row=idx, column=3, value=customer.get("email") or "")
        ws.cell(row=idx, column=4, value=customer.get("address") or "")

    for column_cells in ws.columns:
        width = 12
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, 40))
        ws.column_dimensions[column_cells[0].column_letter].width = width

    fname = f"customer_report_{date.today().isoformat()}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/customers/save", methods=["POST"])
def save_customer():
    user = get_current_user()
    if not user_can_access_customer_master(user):
        return _permission_denied_response("You do not have access to Customer Master.")

    customer_id_raw = request.form.get("customer_id", "").strip()
    customer_id = int(customer_id_raw) if customer_id_raw else None
    payload = _customer_form_payload(request.form)

    conn = get_db()
    try:
        saved_id, errors = save_customer_record(
            conn,
            payload.get("first_name"),
            payload.get("mobile"),
            customer_id=customer_id,
            address=payload.get("address") or "",
            email=payload.get("email") or "",
        )
        if errors:
            customers = list_customers(conn)
            selected_customer = get_customer(conn, customer_id) if customer_id else None
            form = dict(payload)
            form["id"] = customer_id or ""
            return _customer_page_render(
                "partials/master_embed/customer.html" if is_embed_request() else "customer_master.html",
                customers=customers,
                form=form,
                selected_customer=selected_customer,
                errors=errors,
                success_message="",
                form_focus=True,
                show_form=True,
                customer_report_url=url_for("export_customer_report"),
                embed_mode=is_embed_request(),
            ), 400
        conn.commit()
    finally:
        conn.close()

    result_flag = "updated" if customer_id else "created"
    redirect_kwargs = {"saved": result_flag}
    if is_embed_request():
        redirect_kwargs["embed"] = 1
    return redirect(url_for("customer_master", **redirect_kwargs))


@app.route("/customers/delete", methods=["POST"])
def delete_customer():
    user = get_current_user()
    if not user_can_access_customer_master(user):
        return _permission_denied_response("You do not have access to Customer Master.")

    customer_id = request.form.get("customer_id", "").strip()
    if not customer_id:
        _queue_auth_notice("Customer not found.")
        return redirect(url_for("customer_master"))

    conn = get_db()
    try:
        deleted = delete_customer_record(conn, customer_id)
        if not deleted:
            _queue_auth_notice("Customer not found.")
            return redirect(url_for("customer_master"))
        conn.commit()
    finally:
        conn.close()

    return redirect(url_for("customer_master", saved="deleted"))


def _agency_form_payload(source=None):
    source = source or {}
    return {
        "name": " ".join(str(source.get("name") or "").split()).strip(),
        "gst": " ".join(str(source.get("gst") or "").split()).strip().upper(),
        "address": " ".join(str(source.get("address") or "").split()).strip(),
    }


def _agency_page_render(template, **kwargs):
    kwargs.setdefault("auth_notice", _pop_auth_notice())
    kwargs.setdefault("de_nav_section", "master")
    kwargs.setdefault("de_nav_master_view", "agency_master")
    return render_template(template, **kwargs)


@app.route("/agencies")
def agency_master():
    user = get_current_user()
    if not user_can_access_agency_master(user):
        return _permission_denied_response("You do not have access to Agency Master.")

    selected_agency_id = request.args.get("agency_id", "").strip()
    saved_flag = request.args.get("saved", "").strip()
    form_focus = request.args.get("focus", "").strip() == "form"

    conn = get_db()
    try:
        ensure_agencies_schema(conn)
        agencies = list_agencies(conn)
        selected_agency = get_agency(conn, selected_agency_id) if selected_agency_id else None
        conn.commit()
    finally:
        conn.close()

    if selected_agency:
        form = {
            "id": selected_agency["id"],
            "name": selected_agency.get("name") or "",
            "gst": selected_agency.get("gst") or "",
            "address": selected_agency.get("address") or "",
        }
    else:
        form = {"id": "", **_agency_form_payload()}

    success_message = ""
    if saved_flag == "created":
        success_message = "Agency created successfully."
    elif saved_flag == "updated":
        success_message = "Agency updated successfully."
    elif saved_flag == "deleted":
        success_message = "Agency deleted successfully."

    return _agency_page_render(
        "partials/master_embed/agency.html" if is_embed_request() else "agency_master.html",
        agencies=agencies,
        form=form,
        selected_agency=selected_agency,
        errors=[],
        success_message=success_message,
        form_focus=form_focus or bool(selected_agency),
        show_form=form_focus or bool(selected_agency),
        agency_report_url=url_for("export_agency_report"),
        embed_mode=is_embed_request(),
    )


@app.route("/agencies/report")
def export_agency_report():
    """Excel download of all agencies from Agency Master."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    user = get_current_user()
    if not user_can_access_agency_master(user):
        return _permission_denied_response("You do not have access to Agency Master.")

    conn = get_db()
    try:
        ensure_agencies_schema(conn)
        agencies = list_agencies(conn)
        conn.commit()
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Agencies"
    header_font = Font(bold=True)
    headers = ["Agency Name", "GST", "Address"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font

    for idx, agency in enumerate(agencies, start=2):
        ws.cell(row=idx, column=1, value=agency.get("name") or "")
        ws.cell(row=idx, column=2, value=agency.get("gst") or "")
        ws.cell(row=idx, column=3, value=agency.get("address") or "")

    for column_cells in ws.columns:
        width = 12
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, 50))
        ws.column_dimensions[column_cells[0].column_letter].width = width

    fname = f"agency_report_{date.today().isoformat()}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/agencies/save", methods=["POST"])
def save_agency():
    user = get_current_user()
    if not user_can_access_agency_master(user):
        return _permission_denied_response("You do not have access to Agency Master.")

    agency_id_raw = request.form.get("agency_id", "").strip()
    agency_id = int(agency_id_raw) if agency_id_raw else None
    payload = _agency_form_payload(request.form)

    conn = get_db()
    try:
        saved_id, errors = save_agency_record(
            conn,
            payload.get("name"),
            payload.get("gst"),
            payload.get("address"),
            agency_id=agency_id,
        )
        if errors:
            agencies = list_agencies(conn)
            selected_agency = get_agency(conn, agency_id) if agency_id else None
            form = dict(payload)
            form["id"] = agency_id or ""
            return _agency_page_render(
                "partials/master_embed/agency.html" if is_embed_request() else "agency_master.html",
                agencies=agencies,
                form=form,
                selected_agency=selected_agency,
                errors=errors,
                success_message="",
                form_focus=True,
                show_form=True,
                agency_report_url=url_for("export_agency_report"),
                embed_mode=is_embed_request(),
            ), 400
        conn.commit()
    finally:
        conn.close()

    result_flag = "updated" if agency_id else "created"
    redirect_kwargs = {"saved": result_flag}
    if is_embed_request():
        redirect_kwargs["embed"] = 1
    return redirect(url_for("agency_master", **redirect_kwargs))


@app.route("/agencies/delete", methods=["POST"])
def delete_agency():
    user = get_current_user()
    if not user_can_access_agency_master(user):
        return _permission_denied_response("You do not have access to Agency Master.")

    agency_id = request.form.get("agency_id", "").strip()
    if not agency_id:
        _queue_auth_notice("Agency not found.")
        return redirect(url_for("agency_master"))

    conn = get_db()
    try:
        deleted = delete_agency_record(conn, agency_id)
        if not deleted:
            _queue_auth_notice("Agency not found.")
            return redirect(url_for("agency_master"))
        conn.commit()
    finally:
        conn.close()

    return redirect(url_for("agency_master", saved="deleted"))


@app.route("/agencies/create", methods=["POST"], endpoint="create_agency")
def create_agency():
    """JSON create/update agency for check-in and other embeds."""
    user = get_current_user()
    if not user_can_access_agency_master(user):
        return jsonify({"ok": False, "error": "You do not have access to Agency Master."}), 403

    data = request.get_json(silent=True) or {}
    payload = _agency_form_payload(data)
    conn = get_db()
    try:
        agency = upsert_agency_by_name(
            conn,
            payload.get("name"),
            payload.get("gst"),
            payload.get("address"),
        )
        if not agency:
            err_gst = (payload.get("gst") or "").strip()
            if err_gst and not is_valid_agency_gst(err_gst):
                return (
                    jsonify(
                        {
                            "ok": False,
                            "error": "GST must be a valid 15-character GSTIN (e.g. 35AANFH8592H1ZS).",
                        }
                    ),
                    400,
                )
            return jsonify({"ok": False, "error": "Agency name is required."}), 400
        agencies = list_agencies(conn)
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True, "agency": agency, "agencies": agencies})


@app.route("/agencies/api", methods=["GET"], endpoint="list_agencies_api")
def list_agencies_api():
    user = get_current_user()
    if not user_can_access_agency_master(user):
        return jsonify({"ok": False, "error": "You do not have access to Agency Master."}), 403
    conn = get_db()
    try:
        ensure_agencies_schema(conn)
        agencies = list_agencies(conn)
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True, "agencies": agencies})


@app.route("/access-management")
def access_management():
    user = get_current_user()
    selected_user_id = request.args.get("user_id", "").strip()
    saved_flag = request.args.get("saved", "").strip()
    form_focus = request.args.get("focus", "").strip() == "form"
    can_users = user_can_access_user_access_submodule(user, "users")
    can_add = user_can_access_user_access_submodule(user, "add")

    if form_focus:
        if not can_add and not (selected_user_id and can_users):
            if can_users:
                return redirect(url_for("access_management"))
            return _permission_denied_response("You do not have access to Add User.")
    elif not can_users:
        if can_add:
            return redirect(url_for("access_management", focus="form"))
        if user_can_access_user_access_submodule(user, "roles"):
            return redirect(url_for("access_roles"))
        return _permission_denied_response("You do not have access to Users.")

    conn = get_db()
    try:
        users, selected_user = fetch_access_management_users(conn, selected_user_id or None)
        roles = list_access_roles(conn, active_only=True)
    finally:
        conn.close()

    form = {
        "id": selected_user["id"] if selected_user else "",
        "username": selected_user["username"] if selected_user else "",
        "full_name": selected_user.get("full_name", "") if selected_user else "",
        "email": selected_user.get("email", "") if selected_user else "",
        "role_id": selected_user.get("role_id") if selected_user else "",
        "is_admin": bool(selected_user["is_admin"]) if selected_user else False,
        "photo_path": selected_user.get("photo_path", "") if selected_user else "",
        "avatar_tone": selected_user.get("avatar_tone", 0) if selected_user else 0,
    }
    success_message = ""
    if saved_flag == "created":
        success_message = "User created successfully."
    elif saved_flag == "updated":
        success_message = "User access updated successfully."

    return _am_page_render(
        "access_management.html",
        users=users,
        roles=roles,
        roles_summary=[role_summary_for_ui(role) for role in roles],
        form=form,
        selected_user=selected_user,
        errors=[],
        success_message=success_message,
        form_focus=form_focus,
    )


@app.route("/access-management/photos/<path:stored_name>", endpoint="access_user_photo")
def access_user_photo(stored_name):
    """Serve a compressed user profile photo for authenticated users."""
    path = resolve_stored_user_photo(stored_name)
    if not path:
        abort(404)
    return send_file(path, mimetype="image/webp", as_attachment=False, download_name=path.name)


@app.route("/access-management/save", methods=["POST"])
def save_access_user():
    actor = get_current_user()
    user_id_raw = request.form.get("user_id", "").strip()
    username = normalize_username(request.form.get("username"))
    full_name = (request.form.get("full_name") or "").strip()
    email = (request.form.get("email") or "").strip()
    password = request.form.get("password", "")
    role_id_raw = request.form.get("role_id", "").strip()
    remove_photo = bool(request.form.get("remove_photo"))
    photo_upload = request.files.get("photo")

    try:
        user_id = int(user_id_raw) if user_id_raw else None
    except (TypeError, ValueError):
        user_id = None
    try:
        role_id = int(role_id_raw) if role_id_raw else None
    except (TypeError, ValueError):
        role_id = None

    conn = get_db()
    try:
        errors, original, _role = validate_access_user_form(
            conn,
            actor=actor,
            user_id=user_id,
            username=username,
            password=password,
            role_id=role_id,
            email=email,
        )
        new_photo_path = None
        clear_photo = False
        previous_photo = ""
        if original is not None:
            try:
                previous_photo = str(original["photo_path"] or "").strip()
            except (KeyError, IndexError, TypeError):
                previous_photo = ""

        if photo_upload and photo_upload.filename:
            try:
                new_photo_path = process_uploaded_user_photo(photo_upload, photo_upload.filename)
            except ValueError as exc:
                errors.append(str(exc))
            except RuntimeError as exc:
                errors.append(str(exc))
        elif remove_photo:
            clear_photo = True

        if errors:
            users, selected_user = fetch_access_management_users(conn, user_id)
            roles = list_access_roles(conn, active_only=True)
            form = {
                "id": user_id or "",
                "username": username,
                "full_name": full_name,
                "email": email,
                "role_id": role_id or "",
                "is_admin": bool(_role and _role.get("is_admin")),
                "photo_path": "" if clear_photo else (previous_photo if not new_photo_path else new_photo_path),
                "avatar_tone": selected_user.get("avatar_tone", 0) if selected_user else 0,
            }
            if new_photo_path:
                delete_stored_user_photo(new_photo_path)
            return _am_page_render(
                "access_management.html",
                users=users,
                roles=roles,
                roles_summary=[role_summary_for_ui(role) for role in roles],
                form=form,
                selected_user=selected_user,
                errors=errors,
                success_message="",
                form_focus=True,
            ), 400

        photo_arg = None
        if new_photo_path is not None:
            photo_arg = new_photo_path
        elif clear_photo:
            photo_arg = ""

        saved_user_id, result_flag = save_access_user_record(
            conn,
            user_id=user_id,
            username=username,
            full_name=full_name,
            email=email,
            password=password,
            role_id=role_id,
            photo_path=photo_arg,
            sql_now=SQL_NOW,
        )
        if (new_photo_path or clear_photo) and previous_photo and previous_photo != new_photo_path:
            delete_stored_user_photo(previous_photo)
        conn.commit()
    finally:
        conn.close()

    if user_id and actor and int(actor["id"]) == int(saved_user_id):
        g._auth_loaded = False
        get_current_user()

    return redirect(
        url_for("access_management", user_id=saved_user_id, saved=result_flag),
        code=303,
    )


@app.route("/access-management/roles", methods=["GET", "POST"])
def access_roles():
    # Soft-nav / some clients re-POST the redirect target; accept POST here too.
    if request.method == "POST":
        return save_access_role()

    user = get_current_user()
    if not user_can_access_user_access_submodule(user, "roles"):
        return _permission_denied_response("You do not have access to Roles.")

    selected_role_id = request.args.get("role_id", "").strip()
    saved_flag = request.args.get("saved", "").strip()
    form_focus = request.args.get("focus", "").strip() == "form"

    conn = get_db()
    try:
        roles = list_access_roles(conn)
        selected_role = None
        if selected_role_id:
            try:
                selected_role = get_access_role(conn, int(selected_role_id))
            except (TypeError, ValueError):
                selected_role = None
    finally:
        conn.close()

    form = {
        "id": selected_role["id"] if selected_role else "",
        "name": selected_role["name"] if selected_role else "",
        "description": selected_role.get("description", "") if selected_role else "",
        "is_admin": bool(selected_role["is_admin"]) if selected_role else False,
        "is_super_admin_role": is_built_in_administrator_role(selected_role),
        "is_active": bool(selected_role["is_active"]) if selected_role else True,
        "dashboard_modules": dashboard_access_list(selected_role) if selected_role else [],
        "sales_analytics_modules": sales_analytics_access_list(selected_role) if selected_role else [],
        "user_access_modules": user_access_submodule_list(selected_role) if selected_role else [],
        "payroll_modules": payroll_access_list(selected_role) if selected_role else [],
        "accounts_modules": accounts_access_list(selected_role) if selected_role else [],
        "stores_modules": stores_access_list(selected_role) if selected_role else [],
    }
    success_message = ""
    if saved_flag == "created":
        success_message = "Role created successfully."
    elif saved_flag == "updated":
        success_message = "Role updated successfully."

    return _am_roles_page_render(
        "access_roles.html",
        roles=roles,
        form=form,
        selected_role=selected_role,
        errors=[],
        success_message=success_message,
        form_focus=form_focus,
    )


@app.route("/access-management/roles/save", methods=["GET", "POST"])
def save_access_role():
    # Stale soft-nav fallbacks sometimes GET this URL after a failed fetch.
    if request.method == "GET":
        return redirect(url_for("access_roles"), code=303)
    actor = get_current_user()
    if not user_can_access_user_access_submodule(actor, "roles"):
        return _permission_denied_response("You do not have access to Roles.")

    role_id_raw = request.form.get("role_id", "").strip()
    name = (request.form.get("name") or "").strip()
    description = (request.form.get("description") or "").strip()
    is_admin = bool(request.form.get("is_admin"))
    try:
        role_id = int(role_id_raw) if role_id_raw else None
    except (TypeError, ValueError):
        role_id = None
    # Active by default for new roles; checkbox posts when checked on edit.
    is_active = True if role_id is None else bool(request.form.get("is_active"))
    dashboard_modules = request.form.getlist("dashboard_modules")
    sales_analytics_modules = request.form.getlist("sales_analytics_modules")
    user_access_modules = request.form.getlist("user_access_modules")
    payroll_modules = request.form.getlist("payroll_modules")
    accounts_modules = request.form.getlist("accounts_modules")
    stores_modules = request.form.getlist("stores_modules")

    if sales_analytics_modules and not is_admin and "sales_analytics" not in dashboard_modules:
        dashboard_modules = list(dashboard_modules) + ["sales_analytics"]
    if user_access_modules and not is_admin and "access_management" not in dashboard_modules:
        dashboard_modules = list(dashboard_modules) + ["access_management"]
    if payroll_modules and not is_admin and "employee_payroll" not in dashboard_modules:
        dashboard_modules = list(dashboard_modules) + ["employee_payroll"]
    if accounts_modules and not is_admin and "accounts" not in dashboard_modules:
        dashboard_modules = list(dashboard_modules) + ["accounts"]
    if stores_modules and not is_admin and "stores" not in dashboard_modules:
        dashboard_modules = list(dashboard_modules) + ["stores"]

    conn = get_db()
    try:
        existing_role = get_access_role(conn, role_id) if role_id else None
        # Full authority may only be toggled on the built-in Super Administrator row.
        # Do not rewrite other roles' names to "Super Administrator" (causes name clashes).
        if existing_role and is_built_in_administrator_role(existing_role):
            name = SUPER_ADMINISTRATOR_ROLE_NAME
            is_admin = bool(request.form.get("is_admin"))
        else:
            is_admin = False

        dashboard_modules = reconcile_super_admin_only_dashboard_modules(
            actor,
            dashboard_modules,
            dashboard_access_list(existing_role) if existing_role else [],
        )

        errors, _original = validate_access_role_form(
            conn,
            actor=actor,
            role_id=role_id,
            name=name,
            is_admin=is_admin,
            dashboard_modules=dashboard_modules,
            sales_analytics_modules=sales_analytics_modules,
            user_access_modules=user_access_modules,
            payroll_modules=payroll_modules,
            accounts_modules=accounts_modules,
            stores_modules=stores_modules,
        )
        if errors:
            roles = list_access_roles(conn)
            selected_role = get_access_role(conn, role_id) if role_id else None
            form = {
                "id": role_id or "",
                "name": name,
                "description": description,
                "is_admin": is_admin,
                "is_super_admin_role": is_built_in_administrator_role(
                    selected_role, name=name, role_id=role_id
                ),
                "is_active": is_active,
                "dashboard_modules": dashboard_modules,
                "sales_analytics_modules": sales_analytics_modules,
                "user_access_modules": user_access_modules,
                "payroll_modules": payroll_modules,
                "accounts_modules": accounts_modules,
                "stores_modules": stores_modules,
            }
            return _am_roles_page_render(
                "access_roles.html",
                roles=roles,
                form=form,
                selected_role=selected_role,
                errors=errors,
                success_message="",
                form_focus=True,
            ), 400

        saved_role_id, result_flag = save_access_role_record(
            conn,
            role_id=role_id,
            name=name,
            description=description,
            is_admin=is_admin,
            is_active=is_active,
            dashboard_modules=dashboard_modules,
            sales_analytics_modules=sales_analytics_modules,
            user_access_modules=user_access_modules,
            payroll_modules=payroll_modules,
            accounts_modules=accounts_modules,
            stores_modules=stores_modules,
            sql_now=SQL_NOW,
        )
        conn.commit()
    finally:
        conn.close()

    if actor and actor.get("role_id") and int(actor["role_id"]) == int(saved_role_id):
        g._auth_loaded = False
        get_current_user()

    # 303 so soft-nav / browsers follow with GET (POST→302→re-POST hit Method Not Allowed).
    return redirect(
        url_for("access_roles", role_id=saved_role_id, saved=result_flag),
        code=303,
    )


@app.route("/access-management/roles/delete/<int:role_id>", methods=["POST"], endpoint="delete_access_role")
def delete_access_role_view(role_id):
    actor = get_current_user()
    if not user_can_access_user_access_submodule(actor, "roles"):
        return _permission_denied_response("You do not have access to Roles.")

    conn = get_db()
    try:
        ok, message = delete_access_role(conn, role_id)
        if not ok:
            _queue_auth_notice(message)
            return redirect(url_for("access_roles"), code=303)
        conn.commit()
        _queue_auth_notice("Role deleted.")
    finally:
        conn.close()
    return redirect(url_for("access_roles"), code=303)


@app.route("/access-management/unlock/<int:user_id>", methods=["POST"])
def unlock_access_user(user_id):
    actor = get_current_user()
    if not user_can_access_user_access_submodule(actor, "users"):
        return _permission_denied_response("You do not have access to unlock users.")

    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not row:
            _queue_auth_notice("User not found.")
            return redirect(url_for("access_management"))
        auth_security.admin_unlock_user(conn, user_id)
        conn.commit()
        _queue_auth_notice(f"Unlocked account for {row['username']}.")
    finally:
        conn.close()
    return redirect(url_for("access_management"))


@app.route("/access-management/active/<int:user_id>", methods=["POST"], endpoint="toggle_access_user_active")
def toggle_access_user_active(user_id):
    actor = get_current_user()
    if not user_can_access_user_access_submodule(actor, "users"):
        return _permission_denied_response("You do not have access to update users.")

    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not row:
            _queue_auth_notice("User not found.")
            return redirect(url_for("access_management"), code=303)

        user = build_user_context(conn, row)
        if actor and int(actor["id"]) == int(user_id):
            _queue_auth_notice("You cannot change active status for the account you are currently using.")
            return redirect(url_for("access_management"), code=303)

        currently_active = bool(user.get("is_active"))
        if currently_active and user.get("is_admin"):
            active_admin_count = int(
                conn.execute(
                    "SELECT COUNT(*) FROM users WHERE is_admin = 1 AND is_active = 1"
                ).fetchone()[0]
            )
            if active_admin_count <= 1:
                _queue_auth_notice("At least one active administrator must remain in the system.")
                return redirect(url_for("access_management"), code=303)

        new_active = 0 if currently_active else 1
        conn.execute(
            f"""UPDATE users
                   SET is_active = ?, updated_at = {SQL_NOW}
                 WHERE id = ?""",
            (new_active, user_id),
        )
        conn.commit()
        label = user.get("display_name") or user.get("username") or "User"
        if new_active:
            _queue_auth_notice(f"Activated {label}.")
        else:
            _queue_auth_notice(f"Marked {label} inactive.")
    finally:
        conn.close()
    return redirect(url_for("access_management"), code=303)


@app.route("/access-management/delete/<int:user_id>", methods=["POST"])
def delete_access_user(user_id):
    actor = get_current_user()
    if not user_can_access_user_access_submodule(actor, "users"):
        return _permission_denied_response("You do not have access to delete users.")

    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not row:
            _queue_auth_notice("User not found.")
            return redirect(url_for("access_management"))

        user = build_user_context(conn, row)
        if is_system_administrator(user):
            _queue_auth_notice("The default administrator account cannot be deleted.")
            return redirect(url_for("access_management"))

        if actor and int(actor["id"]) == int(user_id):
            _queue_auth_notice("You cannot delete the account you are currently using.")
            return redirect(url_for("access_management"))

        active_admin_count = int(
            conn.execute(
                "SELECT COUNT(*) FROM users WHERE is_admin = 1 AND is_active = 1"
            ).fetchone()[0]
        )
        if user.get("is_admin") and user.get("is_active") and active_admin_count <= 1:
            _queue_auth_notice("At least one active administrator must remain in the system.")
            return redirect(url_for("access_management"))

        photo_path = (user.get("photo_path") or "").strip()
        conn.execute("DELETE FROM user_permissions WHERE user_id = ?", (user_id,))
        conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
        conn.commit()
        if photo_path:
            delete_stored_user_photo(photo_path)
    finally:
        conn.close()

    return redirect(url_for("access_management"))


# ── Print Agent (Windows silent printing bridge) ─────────────────────────────

@app.route("/api/print-agent/register", methods=["POST"], endpoint="print_agent_register")
def print_agent_register():
    """First-launch registration from Hotel Print Agent desktop app."""
    from print_agent_store import register_print_agent

    payload = request.get_json(silent=True) or {}
    conn = get_db()
    try:
        result = register_print_agent(conn, payload, request_host_url=request.host_url)
        status = 200 if result.get("ok") else 400
        return jsonify(result), status
    finally:
        conn.close()


@app.route("/api/print-agent/heartbeat", methods=["POST"], endpoint="print_agent_heartbeat")
def print_agent_heartbeat():
    from print_agent_store import heartbeat_print_agent

    payload = request.get_json(silent=True) or {}
    auth = request.headers.get("Authorization") or ""
    bearer = auth[7:].strip() if auth.lower().startswith("bearer ") else ""
    conn = get_db()
    try:
        result = heartbeat_print_agent(conn, payload, bearer)
        status = 200 if result.get("ok") else 401
        return jsonify(result), status
    finally:
        conn.close()


@app.route("/api/print-agent/updates/latest", methods=["GET"], endpoint="print_agent_updates_latest")
def print_agent_updates_latest():
    from print_agent_store import print_agent_latest_update

    current = (request.args.get("current") or "").strip()
    return jsonify(print_agent_latest_update(current))


@app.route("/api/print-agent/config", methods=["GET"], endpoint="print_agent_browser_config")
def print_agent_browser_config():
    """Browser-facing hint: local agent URL + cloud origins."""
    from print_agent_store import default_print_agent_origins

    return jsonify(
        {
            "ok": True,
            "localBaseUrl": "http://127.0.0.1:4567",
            "roles": [
                "kitchen1",
                "billing",
                "bar",
                "hotel_folio",
                "hotel_invoice",
                "kitchen2",
                "label",
            ],
            "cloudOrigin": (os.environ.get("APP_BASE_URL") or "https://belleliteaccounts.com").rstrip(
                "/"
            ),
            "allowedOrigins": default_print_agent_origins(request.host_url),
        }
    )


@app.route("/api/print-agent/browser-pair", methods=["GET"], endpoint="print_agent_browser_pair")
def print_agent_browser_pair():
    """Logged-in cloud tab fetches API key to talk to the local agent on this PC."""
    from print_agent_store import browser_pair_print_agent

    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "Sign in required."}), 401

    business_id = (
        (request.args.get("businessId") or "").strip()
        or str(user.get("business_id") or user.get("businessId") or "").strip()
        or "default"
    )
    agent_id = (request.args.get("agentId") or "").strip()
    conn = get_db()
    try:
        result = browser_pair_print_agent(
            conn, business_id=business_id, agent_id=agent_id or None
        )
        # Also try empty / any agent if business-specific miss (single-tenant installs)
        if not result.get("ok") and not agent_id and business_id != "default":
            result = browser_pair_print_agent(conn, business_id=None)
        status = 200 if result.get("ok") else 404
        return jsonify(result), status
    finally:
        conn.close()


if __name__ == "__main__":
    init_db()
    app.run(debug=True, host="127.0.0.1", port=8002)
