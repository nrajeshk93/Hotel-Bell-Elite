"""Promotion — Meta WhatsApp template blasts from Communication Hub."""

from __future__ import annotations

import io
import logging
import re
import time
from datetime import datetime

from flask import jsonify, render_template, request, send_file
from openpyxl import Workbook, load_workbook

import whatsapp_client as wa
from db import ensure_communication_hub_schema, get_db

log = logging.getLogger(__name__)

PROMO_ROW_CAP = 500
PROMO_SEND_DELAY_SEC = 0.15

_pop_auth_notice = None
_get_user = None


def _bind_helpers(*, pop_auth_notice, get_user):
    global _pop_auth_notice, _get_user
    _pop_auth_notice = pop_auth_notice
    _get_user = get_user


def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    return text


def parse_promotion_excel(file_storage) -> tuple[list[dict], list[dict]]:
    """Parse Excel: column A = name, column B = mobile.

    Returns (valid_rows, skipped_rows). Each valid row:
    ``{row_number, name, phone, phone_display}``.
    """
    raw = file_storage.read()
    if not raw:
        raise ValueError("Upload an Excel file with customer name and mobile.")
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Could not read Excel file. Upload a valid .xlsx.") from exc
    try:
        ws = wb.active
        if ws is None:
            raise ValueError("Excel workbook has no sheets.")
        valid: list[dict] = []
        skipped: list[dict] = []
        seen_phones: set[str] = set()
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_col=2, values_only=True), start=1):
            name = _cell_str(row[0] if row else "")
            mobile_raw = _cell_str(row[1] if row and len(row) > 1 else "")
            # Skip blank rows and a common header row.
            if not name and not mobile_raw:
                continue
            if idx == 1 and name.lower() in {"name", "customer", "customer name"} and (
                not mobile_raw or mobile_raw.lower() in {"mobile", "phone", "whatsapp", "number"}
            ):
                continue
            if not mobile_raw:
                skipped.append(
                    {
                        "row_number": idx,
                        "name": name,
                        "phone": "",
                        "reason": "Missing mobile number",
                    }
                )
                continue
            phone = wa.normalise_whatsapp_number(mobile_raw)
            if not phone:
                skipped.append(
                    {
                        "row_number": idx,
                        "name": name,
                        "phone": mobile_raw,
                        "reason": "Invalid mobile number",
                    }
                )
                continue
            if phone in seen_phones:
                skipped.append(
                    {
                        "row_number": idx,
                        "name": name,
                        "phone": phone,
                        "reason": "Duplicate mobile in file",
                    }
                )
                continue
            seen_phones.add(phone)
            valid.append(
                {
                    "row_number": idx,
                    "name": name or "Customer",
                    "phone": phone,
                    "phone_display": mobile_raw,
                }
            )
        return valid, skipped
    finally:
        wb.close()


def build_promotion_sample_xlsx() -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Recipients"
    ws.append(["Name", "Mobile"])
    ws.append(["Anita Sharma", "9876543210"])
    ws.append(["Ravi Kumar", "919876543210"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _find_template(name: str, language: str) -> dict | None:
    ok, err, items = wa.list_approved_message_templates()
    if not ok:
        raise RuntimeError(err or "Could not load templates.")
    name_l = (name or "").strip().lower()
    lang_l = (language or "").strip().lower()
    for item in items:
        if (item.get("name") or "").strip().lower() != name_l:
            continue
        if lang_l and (item.get("language") or "").strip().lower() != lang_l:
            continue
        return item
    return None


def run_promotion_campaign(
    conn,
    *,
    template_name: str,
    template_language: str,
    rows: list[dict],
    user_id=None,
    send_fn=None,
    delay_sec: float = PROMO_SEND_DELAY_SEC,
) -> dict:
    """Create campaign rows and send paced template messages."""
    ensure_communication_hub_schema(conn)
    template = _find_template(template_name, template_language)
    if not template:
        raise ValueError("Selected template was not found among approved Meta templates.")
    if not template.get("sendable"):
        raise ValueError(
            template.get("block_reason")
            or "This template is not supported for Promotion sends."
        )
    if not rows:
        raise ValueError("No valid recipients to send.")
    if len(rows) > PROMO_ROW_CAP:
        raise ValueError(f"Campaign exceeds the {PROMO_ROW_CAP}-row limit.")

    body_params_needed = int(template.get("body_param_count") or 0)
    send = send_fn or wa.send_template_message
    now = _now()
    cur = conn.execute(
        """INSERT INTO wa_promo_campaigns
           (template_name, template_language, created_by, created_at, status,
            total_rows, sent_count, failed_count, skipped_count)
           VALUES (?, ?, ?, ?, 'running', ?, 0, 0, 0)""",
        (
            template["name"],
            template["language"],
            user_id,
            now,
            len(rows),
        ),
    )
    campaign_id = int(cur.lastrowid)
    recipient_ids: list[tuple[int, dict]] = []
    for row in rows:
        cur = conn.execute(
            """INSERT INTO wa_promo_recipients
               (campaign_id, row_number, customer_name, phone_e164, status)
               VALUES (?, ?, ?, ?, 'pending')""",
            (
                campaign_id,
                int(row.get("row_number") or 0),
                str(row.get("name") or "")[:200],
                str(row.get("phone") or ""),
            ),
        )
        recipient_ids.append((int(cur.lastrowid), row))
    conn.commit()

    sent = failed = skipped = 0
    outcomes: list[dict] = []
    for i, (recipient_id, row) in enumerate(recipient_ids):
        phone = str(row.get("phone") or "")
        name = str(row.get("name") or "Customer")
        body_parameters = [name] if body_params_needed >= 1 else None
        ok, err, payload = send(
            phone,
            template["name"],
            template["language"],
            body_parameters,
        )
        wa_id = wa.first_message_id(payload) if ok else ""
        if ok:
            status = "sent"
            sent += 1
            error = ""
        else:
            status = "failed"
            failed += 1
            error = (err or "Send failed")[:500]
        sent_at = _now()
        conn.execute(
            """UPDATE wa_promo_recipients
               SET status = ?, error = ?, wa_message_id = ?, sent_at = ?
               WHERE id = ?""",
            (status, error, wa_id, sent_at, recipient_id),
        )
        outcomes.append(
            {
                "id": recipient_id,
                "row_number": int(row.get("row_number") or 0),
                "name": name,
                "phone": phone,
                "status": status,
                "error": error,
                "wa_message_id": wa_id,
            }
        )
        if delay_sec > 0 and i < len(recipient_ids) - 1:
            time.sleep(delay_sec)

    finished = _now()
    conn.execute(
        """UPDATE wa_promo_campaigns
           SET status = 'completed', finished_at = ?,
               sent_count = ?, failed_count = ?, skipped_count = ?
           WHERE id = ?""",
        (finished, sent, failed, skipped, campaign_id),
    )
    conn.commit()
    return {
        "campaign_id": campaign_id,
        "template_name": template["name"],
        "template_language": template["language"],
        "total": len(rows),
        "sent": sent,
        "failed": failed,
        "skipped": skipped,
        "outcomes": outcomes,
    }


def register_promotion(app, *, pop_auth_notice, get_user):
    _bind_helpers(pop_auth_notice=pop_auth_notice, get_user=get_user)

    def _page_render(**kwargs):
        kwargs.setdefault("auth_notice", _pop_auth_notice() if _pop_auth_notice else None)
        kwargs.setdefault("de_nav_section", "communication_hub")
        kwargs.setdefault("de_nav_communication_hub_view", "promotion")
        return render_template("communication_hub_promotion.html", **kwargs)

    @app.route("/communication-hub/promotion")
    def communication_hub_promotion():
        user = _get_user() if _get_user else None
        return _page_render(
            page_title="Promotion",
            whatsapp_configured=wa.whatsapp_configured(),
            whatsapp_templates_configured=wa.whatsapp_templates_configured(),
            whatsapp_dry_run=not wa.whatsapp_live_sends_allowed(),
            promo_row_cap=PROMO_ROW_CAP,
            current_user_name=(
                (user.get("full_name") or user.get("username") or "User").strip()
                if user
                else "User"
            ),
        )

    @app.route("/communication-hub/api/promotion/templates", methods=["GET"])
    def communication_hub_api_promotion_templates():
        force = (request.args.get("refresh") or "").strip() in {"1", "true", "yes"}
        ok, err, items = wa.list_approved_message_templates(force_refresh=force)
        if not ok:
            return jsonify({"ok": False, "error": err or "Could not load templates.", "templates": []}), 400
        return jsonify({"ok": True, "templates": items})

    @app.route("/communication-hub/api/promotion/sample.xlsx", methods=["GET"])
    def communication_hub_api_promotion_sample():
        buf = build_promotion_sample_xlsx()
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="promotion_recipients_sample.xlsx",
        )

    @app.route("/communication-hub/api/promotion/preview", methods=["POST"])
    def communication_hub_api_promotion_preview():
        upload = request.files.get("file") or request.files.get("excel")
        if not upload or not getattr(upload, "filename", None):
            return jsonify({"ok": False, "error": "Choose an Excel (.xlsx) file."}), 400
        filename = str(upload.filename or "").lower()
        if not filename.endswith(".xlsx"):
            return jsonify({"ok": False, "error": "Upload an .xlsx Excel file."}), 400
        try:
            valid, skipped = parse_promotion_excel(upload)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400
        if len(valid) > PROMO_ROW_CAP:
            return (
                jsonify(
                    {
                        "ok": False,
                        "error": f"File has {len(valid)} valid rows; max is {PROMO_ROW_CAP}.",
                        "valid_count": len(valid),
                        "skipped_count": len(skipped),
                    }
                ),
                400,
            )
        return jsonify(
            {
                "ok": True,
                "valid_count": len(valid),
                "skipped_count": len(skipped),
                "rows": valid,
                "skipped": skipped[:50],
                "row_cap": PROMO_ROW_CAP,
            }
        )

    @app.route("/communication-hub/api/promotion/send", methods=["POST"])
    def communication_hub_api_promotion_send():
        user = _get_user() if _get_user else None
        user_id = user.get("id") if user else None
        template_name = ""
        template_language = ""
        rows: list[dict] = []

        if request.content_type and "multipart/form-data" in request.content_type:
            template_name = (request.form.get("template_name") or "").strip()
            template_language = (request.form.get("template_language") or "").strip()
            upload = request.files.get("file") or request.files.get("excel")
            if not upload or not getattr(upload, "filename", None):
                return jsonify({"ok": False, "error": "Choose an Excel (.xlsx) file."}), 400
            try:
                rows, _skipped = parse_promotion_excel(upload)
            except ValueError as exc:
                return jsonify({"ok": False, "error": str(exc)}), 400
        else:
            data = request.get_json(silent=True) or {}
            template_name = (data.get("template_name") or "").strip()
            template_language = (data.get("template_language") or "").strip()
            raw_rows = data.get("rows") or []
            if not isinstance(raw_rows, list):
                return jsonify({"ok": False, "error": "rows must be a list."}), 400
            for item in raw_rows:
                if not isinstance(item, dict):
                    continue
                phone = wa.normalise_whatsapp_number(item.get("phone") or item.get("mobile") or "")
                if not phone:
                    continue
                rows.append(
                    {
                        "row_number": int(item.get("row_number") or 0),
                        "name": _cell_str(item.get("name") or "Customer")[:200],
                        "phone": phone,
                    }
                )

        if not template_name:
            return jsonify({"ok": False, "error": "Select a WhatsApp template."}), 400
        if not template_language:
            return jsonify({"ok": False, "error": "Template language is required."}), 400
        if not rows:
            return jsonify({"ok": False, "error": "No valid recipients to send."}), 400
        if len(rows) > PROMO_ROW_CAP:
            return (
                jsonify({"ok": False, "error": f"Campaign exceeds the {PROMO_ROW_CAP}-row limit."}),
                400,
            )

        conn = get_db()
        try:
            result = run_promotion_campaign(
                conn,
                template_name=template_name,
                template_language=template_language,
                rows=rows,
                user_id=user_id,
            )
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400
        except RuntimeError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400
        finally:
            conn.close()
        return jsonify({"ok": True, **result})
