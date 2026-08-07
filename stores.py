"""Stores module — simple Bar/Kitchen indent-to-stock flow."""

from __future__ import annotations

import io
import logging
import re
import sqlite3
import uuid
from datetime import date, datetime, timedelta
from typing import Any

from flask import Blueprint, flash, jsonify, redirect, render_template, request, send_file, url_for

from db import (
    _normalize_pos_menu_unit,
    _qty_in_product_units,
    ensure_pos_schema,
    ensure_stores_schema,
    get_db,
    indian_fiscal_year_label,
    list_pos_menu_recipe_lines,
)
from embed_helpers import is_embed_request
from whatsapp_indent import (
    assign_fresh_approval_token,
    notify_indent_pending_whatsapp,
    supersede_indent_whatsapp_sends,
)
from workspace_access import user_can_access_stores_submodule

logger = logging.getLogger(__name__)
STORES_OUTLETS = (
    {"key": "bar", "label": "Bar"},
    {"key": "restaurant", "label": "Restaurant"},
)
OUTLET_KEYS = {item["key"] for item in STORES_OUTLETS}
# Indent filter can also use "All" (key stays "both" for existing URLs/data).
STORES_FILTER_OUTLETS = (
    {"key": "both", "label": "All"},
) + STORES_OUTLETS
PRODUCT_OUTLETS = (
    {"key": "bar", "label": "Bar"},
    {"key": "restaurant", "label": "Restaurant"},
)
PRODUCT_OUTLET_KEYS = {item["key"] for item in PRODUCT_OUTLETS}
FILTER_OUTLET_KEYS = {item["key"] for item in STORES_FILTER_OUTLETS}
DEFAULT_UNITS = ("kg", "gram", "pcs", "liter", "mL", "bunch", "bottle", "pack")

STATUS_LABELS = {
    "draft": "Draft",
    "pending": "Waiting approval",
    "approved": "Approved",
    "rejected": "Rejected",
    "stocked": "Stocked",
    "open": "Open",
    "received": "Received in stock",
    "cancelled": "Cancelled",
}

INDENT_LIST_VIEWS = (
    ("pending", "Pending Approval"),
    ("approved", "Approved"),
    ("rejected", "Rejected"),
)
INDENT_LIST_VIEW_STATUSES = {
    "pending": ("draft", "pending"),
    "approved": ("approved",),
    "rejected": ("rejected",),
}
EDITABLE_INDENT_STATUSES = ("draft", "pending", "rejected")


def _parse_indent_list_view(raw: str | None) -> str:
    key = (raw or "").strip().lower()
    if key in INDENT_LIST_VIEW_STATUSES:
        return key
    return "pending"


PAGE_META = {
    "product_master": {
        "title": "Products",
        "subtitle": "Categories and products used when raising indents for Bar and Restaurant.",
        "step": "Master",
        "list_endpoint": "stores_product_master",
        "cta": "Add product",
        "cta_endpoint": "stores_product_master",
        "cta_args": {"focus": "form"},
        "show_outlet_tabs": True,
    },
    "indent": {
        "title": "Indent",
        "subtitle": "",
        "step": "1 · Indent",
        "list_endpoint": "stores_indent",
        "cta": "New Indent",
        "cta_endpoint": "stores_indent",
        "cta_args": {"focus": "form"},
    },
    "purchase_orders": {
        "title": "Purchase Order",
        "subtitle": "",
        "step": "Purchase Order",
        "list_endpoint": "stores_orders",
        "cta": None,
    },
    "approvals": {
        "title": "Approvals",
        "subtitle": "Review waiting indents. Approve to buy, or reject.",
        "step": "2 · Approvals",
        "list_endpoint": "stores_approvals",
        "cta": None,
    },
    "purchase_requests": {
        "title": "Stock Inward",
        "subtitle": "",
        "step": "3 · Stock Inward",
        "list_endpoint": "stores_purchase_requests",
        "cta": None,
    },
    "stock": {
        "title": "Stock",
        "subtitle": "What is currently in the store for this outlet.",
        "step": "4 · Stock",
        "list_endpoint": "stores_stock",
        "cta": None,
    },
    "stock_audit": {
        "title": "Stock Audit",
        "subtitle": "Verify and adjust your stock accuracy, one item at a time.",
        "step": "5 · Stock Audit",
        "list_endpoint": "stores_stock_audit",
        "cta": None,
        "show_outlet_tabs": True,
    },
    "stock_audit_report": {
        "title": "Stock Audit Report",
        "subtitle": "Stock adjustments from verified audits.",
        "step": "Reports · Stock Audit",
        "list_endpoint": "stores_stock_audit_report",
        "cta": None,
        "show_outlet_tabs": True,
    },
}

stores_bp = Blueprint("stores", __name__)

_pop_auth_notice = None
_get_user = None


def _bind_helpers(pop_auth_notice, get_user):
    global _pop_auth_notice, _get_user
    _pop_auth_notice = pop_auth_notice
    _get_user = get_user


def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _whatsapp_indent_host_allowed() -> tuple[bool, str]:
    """Block live WhatsApp sends when this app host is not the webhook host.

    Meta delivers Approve/Reject webhooks to production only. Submitting from a
    local preview with live credentials creates tokens in the local DB that
    production cannot find (unknown_token). Set ``WHATSAPP_INDENT_PUBLIC_HOST``
    (e.g. belleliteaccounts.com) so off-host live sends fail with a clear flash.
    """
    import os

    expected = (os.environ.get("WHATSAPP_INDENT_PUBLIC_HOST") or "").strip().lower()
    if not expected:
        return True, ""
    try:
        from whatsapp_client import whatsapp_live_sends_allowed

        if not whatsapp_live_sends_allowed():
            return True, ""
    except Exception:
        pass
    host = (request.host or "").split(":")[0].strip().lower()
    if not host:
        return True, ""
    if host == expected or host.endswith("." + expected):
        return True, ""
    return (
        False,
        f"WhatsApp approval must be sent from {expected} (where webhooks arrive). "
        "Open Indent on that site and Send for Approval again — local preview "
        "cannot receive Approve/Reject clicks.",
    )


def _notify_indent_pending_whatsapp(conn, indent_id: int, outlet: str) -> None:
    """Best-effort WhatsApp indent_approval notify; never blocks indent save.

    Success is silent in the UI (indent flash already says sent for approval).
    Failures still surface so staff know WhatsApp did not go out.
    """
    allowed, host_msg = _whatsapp_indent_host_allowed()
    if not allowed:
        flash(host_msg, "error")
        return
    try:
        ok, message = notify_indent_pending_whatsapp(
            conn,
            int(indent_id),
            outlet_label=_outlet_label(outlet),
        )
        conn.commit()
        if not ok and message:
            flash(message, "error")
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        flash("Indent saved, but WhatsApp approval notify failed.", "error")


def _format_stores_dt(value: Any) -> str:
    """Display datetimes as ``19-July 10.05 AM`` (date-only → ``19-July``)."""
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text in {"—", "-", "never"}:
        return text
    parsed = _parse_stores_dt(text)
    if parsed is None:
        return text
    day_month = f"{parsed.day}-{parsed.strftime('%B')}"
    if len(text) <= 10:
        return day_month
    hour12 = parsed.hour % 12 or 12
    ampm = "AM" if parsed.hour < 12 else "PM"
    return f"{day_month} {hour12}.{parsed.minute:02d} {ampm}"


def _parse_stores_dt(text: str) -> datetime | None:
    for fmt, length in (
        ("%Y-%m-%d %H:%M:%S", 19),
        ("%Y-%m-%d %H:%M", 16),
        ("%Y-%m-%d", 10),
    ):
        chunk = text[:length]
        if len(chunk) < length:
            continue
        try:
            return datetime.strptime(chunk, fmt)
        except ValueError:
            continue
    return None


def _format_stores_date_line(value: Any) -> str:
    """``19 July`` for multi-line submitted cells."""
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    parsed = _parse_stores_dt(text)
    if parsed is None:
        return text
    return f"{parsed.day} {parsed.strftime('%B')}"


def _format_stores_time_line(value: Any) -> str:
    """``10:05 AM`` for multi-line submitted cells."""
    if value is None:
        return ""
    text = str(value).strip()
    if not text or len(text) <= 10:
        return ""
    parsed = _parse_stores_dt(text)
    if parsed is None:
        return ""
    hour12 = parsed.hour % 12 or 12
    ampm = "AM" if parsed.hour < 12 else "PM"
    return f"{hour12}:{parsed.minute:02d} {ampm}"


def _parse_optional_price(raw: str | None) -> tuple[float | None, str | None]:
    text = str(raw or "").strip().replace(",", "")
    if not text:
        return None, None
    try:
        value = float(text)
    except (TypeError, ValueError):
        return None, "Approximate price must be a number."
    if value < 0:
        return None, "Approximate price cannot be negative."
    return round(value, 2), None


def _format_optional_price(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return ""
    if number != number:  # NaN
        return ""
    if number == int(number):
        return str(int(number))
    return f"{number:.2f}".rstrip("0").rstrip(".")


def _normalize_outlet_key(raw: str | None) -> str:
    key = (raw or "").strip().lower()
    # Legacy stores key — keep old bookmarks working.
    if key == "kitchen":
        return "restaurant"
    return key


def _parse_outlet(raw: str | None) -> str:
    """Operational outlet used for stock / saved indents (never 'both')."""
    key = _normalize_outlet_key(raw or "bar")
    return key if key in OUTLET_KEYS else "bar"


def _parse_outlet_filter(raw: str | None) -> str:
    """Outlet filter for Stores list UI — All, Bar, or Restaurant. Defaults to All."""
    if raw is None or not str(raw).strip():
        return "both"
    key = _normalize_outlet_key(raw)
    return key if key in FILTER_OUTLET_KEYS else "both"


def _outlet_label(outlet: str) -> str:
    for item in STORES_FILTER_OUTLETS:
        if item["key"] == outlet:
            return item["label"]
    return outlet.title()


def _outlet_match_sql(column: str, outlet: str) -> tuple[str, tuple[Any, ...]]:
    """SQL fragment + params for list filters (supports All)."""
    if outlet == "both":
        return f"{column} IN ('bar', 'restaurant')", ()
    return f"{column} = ?", (outlet,)


def _parse_product_outlet(raw: str | None) -> str:
    key = _normalize_outlet_key(raw or "restaurant")
    return key if key in PRODUCT_OUTLET_KEYS else "restaurant"


def _parse_optional_supplier_id(raw) -> int | None:
    text = str(raw or "").strip()
    if not text:
        return None
    try:
        supplier_id = int(text)
    except (TypeError, ValueError):
        return None
    return supplier_id if supplier_id > 0 else None


def _product_preferred_supplier_ids_from_form(form) -> tuple[int | None, int | None, int | None]:
    return (
        _parse_optional_supplier_id(form.get("preferred_supplier_1_id")),
        _parse_optional_supplier_id(form.get("preferred_supplier_2_id")),
        _parse_optional_supplier_id(form.get("preferred_supplier_3_id")),
    )


def _product_outlet_label(outlet: str) -> str:
    key = _normalize_outlet_key(outlet or "restaurant")
    for item in PRODUCT_OUTLETS:
        if item["key"] == key:
            return item["label"]
    # Legacy products may still be stored as "both".
    if key == "both":
        return "Both"
    return "Restaurant"


def _title_case_words(name: str) -> str:
    """Title-case labels: each word Capital + rest lowercase.

    Preserves short codes like T.D. / B/L. Collapses extra spaces.
    """
    cleaned = " ".join(str(name or "").split())
    if not cleaned:
        return cleaned

    def _fix_atom(atom: str) -> str:
        if not atom:
            return atom
        # Keep letter codes like T.D. or B/L uppercase.
        if re.fullmatch(r"[A-Za-z]([./][A-Za-z])+\.?", atom):
            return atom.upper()
        return atom[:1].upper() + atom[1:].lower()

    def _fix_word(word: str) -> str:
        if "-" in word:
            return "-".join(_fix_atom(part) for part in word.split("-"))
        return _fix_atom(word)

    return " ".join(_fix_word(word) for word in cleaned.split(" "))


def _title_case_product_name(name: str) -> str:
    return _title_case_words(name)

def _rename_store_item_name_refs(conn, old_name: str, new_name: str) -> None:
    """Keep free-text item_name columns in sync when a product is renamed."""
    if not old_name or not new_name or old_name == new_name:
        return
    tables = (
        "store_indent_lines",
        "store_purchase_request_lines",
        "store_stock_items",
        "store_stock_movements",
        "store_counter_transfer_lines",
        "store_stock_issue_lines",
        "store_stock_verification_lines",
    )
    existing = {
        str(row[0])
        for row in conn.execute(
            "SELECT name FROM sqlite_master WHERE type = 'table'"
        ).fetchall()
    }
    for table in tables:
        if table not in existing:
            continue
        conn.execute(
            f"UPDATE {table} SET item_name = ? WHERE item_name = ?",
            (new_name, old_name),
        )


def _status_label(status: str) -> str:
    return STATUS_LABELS.get(status, status.replace("_", " ").title())


def _next_doc_no(conn, table: str, column: str, prefix: str, outlet: str) -> str:
    day = date.today().strftime("%Y%m%d")
    outlet_code = outlet.upper()[:3]
    like = f"{prefix}-{outlet_code}-{day}-%"
    row = conn.execute(
        f"SELECT {column} AS doc_no FROM {table} WHERE {column} LIKE ? ORDER BY id DESC LIMIT 1",
        (like,),
    ).fetchone()
    seq = 1
    if row and row["doc_no"]:
        try:
            seq = int(str(row["doc_no"]).rsplit("-", 1)[-1]) + 1
        except ValueError:
            seq = 1
    return f"{prefix}-{outlet_code}-{day}-{seq:03d}"


_OUTLET_INDENT_CODES = {
    "bar": "BAR",
    "restaurant": "RES",
}

# Legacy: IND/Restaurant/1/2026-27 — still counted so FY sequences do not reset mid-year.
_IND_FY_NO_RE = re.compile(
    r"^IND/(Bar|Restaurant)/(\d+)/(\d{4}-\d{2})$",
    re.IGNORECASE,
)
# Current: IND/RES/26-27/1
_IND_SHORT_FY_NO_RE = re.compile(
    r"^IND/(BAR|RES)/(\d{2}-\d{2})/(\d+)$",
    re.IGNORECASE,
)


def _indent_outlet_code(outlet: str) -> str:
    key = _parse_outlet(outlet)
    return _OUTLET_INDENT_CODES.get(key, str(key or "OUT").upper()[:3])


def _short_fiscal_year_label(when=None) -> str:
    """Short Indian FY label, e.g. 2026-27 → 26-27."""
    fy = indian_fiscal_year_label(when)
    parts = str(fy or "").split("-")
    if len(parts) == 2 and len(parts[0]) >= 2:
        return f"{parts[0][-2:]}-{parts[1][-2:]}"
    return fy


def _next_indent_no(conn, outlet: str, when=None) -> str:
    """Allocate IND/{BAR|RES}/{YY-YY}/{n}, series per outlet + fiscal year from 1."""
    outlet_key = _parse_outlet(outlet)
    outlet_code = _indent_outlet_code(outlet_key)
    outlet_label = _outlet_label(outlet_key)
    fy = indian_fiscal_year_label(when)
    short_fy = _short_fiscal_year_label(when)
    rows = conn.execute(
        """
        SELECT indent_no
        FROM store_indents
        WHERE outlet = ?
          AND upper(indent_no) LIKE 'IND/%'
        """,
        (outlet_key,),
    ).fetchall()
    max_n = 0
    for row in rows:
        text = str(row["indent_no"] or "").strip()
        match_new = _IND_SHORT_FY_NO_RE.match(text)
        if match_new:
            if match_new.group(1).upper() != outlet_code:
                continue
            if match_new.group(2) != short_fy:
                continue
            max_n = max(max_n, int(match_new.group(3)))
            continue
        match_old = _IND_FY_NO_RE.match(text)
        if not match_old:
            continue
        if match_old.group(1).lower() != outlet_label.lower():
            continue
        if match_old.group(3) != fy:
            continue
        max_n = max(max_n, int(match_old.group(2)))
    return f"IND/{outlet_code}/{short_fy}/{max_n + 1}"


def _parse_lines_from_form(form) -> list[dict[str, Any]]:
    names = form.getlist("item_name")
    qtys = form.getlist("quantity")
    units = form.getlist("unit")
    prices = form.getlist("approximate_price")
    pack_labels = form.getlist("pack_label")
    pack_qtys = form.getlist("pack_qty_in_base")
    lines = []
    for idx, name in enumerate(names):
        item_name = (name or "").strip()
        if not item_name:
            continue
        try:
            qty = float(qtys[idx] if idx < len(qtys) else 0)
        except (TypeError, ValueError, IndexError):
            qty = 0
        if qty <= 0:
            continue
        unit = (units[idx] if idx < len(units) else "pcs") or "pcs"
        unit = unit.strip() or "pcs"
        price_raw = prices[idx] if idx < len(prices) else ""
        approx_price, _price_err = _parse_optional_price(price_raw)
        pack_label = (pack_labels[idx] if idx < len(pack_labels) else "") or ""
        pack_label = pack_label.strip()
        pack_qty_in_base = None
        if pack_label:
            try:
                pack_qty_raw = pack_qtys[idx] if idx < len(pack_qtys) else ""
                pack_qty_in_base = float(pack_qty_raw or 0)
            except (TypeError, ValueError, IndexError):
                pack_qty_in_base = 0.0
            if pack_qty_in_base <= 0:
                pack_label = ""
                pack_qty_in_base = None
        lines.append({
            "item_name": item_name,
            "quantity": qty,
            "unit": unit,
            "notes": "",
            "approximate_price": approx_price,
            "pack_label": pack_label,
            "pack_qty_in_base": pack_qty_in_base,
        })
    return lines


def _row_pack_qty_in_base(row: Any) -> float | None:
    try:
        keys = row.keys() if hasattr(row, "keys") else ()
        if "pack_qty_in_base" not in keys:
            return None
        raw = row["pack_qty_in_base"]
    except (KeyError, TypeError, IndexError):
        return None
    if raw is None or raw == "":
        return None
    try:
        qty = float(raw)
    except (TypeError, ValueError):
        return None
    if qty <= 0:
        return None
    return qty


def _line_stock_qty_delta(line: Any, received_qty: float) -> float:
    """Convert received packs (or base qty) into stock delta in the line's base unit."""
    pack_qty = _row_pack_qty_in_base(line)
    if pack_qty is None:
        return float(received_qty)
    return float(received_qty) * pack_qty


def _line_stock_unit_cost(line: Any, unit_cost: float | None) -> float | None:
    """Normalize pack unit cost into cost per base-unit stock quantity."""
    if unit_cost is None:
        return None
    pack_qty = _row_pack_qty_in_base(line)
    if pack_qty is None:
        return unit_cost
    return round(float(unit_cost) / pack_qty, 4)


def _format_indent_line_item(line: Any) -> str:
    name = str(line["item_name"] if hasattr(line, "keys") else line.get("item_name") or "").strip()
    try:
        pack_label = ""
        if hasattr(line, "keys") and "pack_label" in line.keys():
            pack_label = (line["pack_label"] or "").strip()
        elif isinstance(line, dict):
            pack_label = (line.get("pack_label") or "").strip()
    except (KeyError, TypeError):
        pack_label = ""
    if pack_label:
        return f"{name} — {pack_label}"
    return name


def _row_pack_label(row: Any) -> str:
    try:
        if hasattr(row, "keys") and "pack_label" in row.keys():
            return (row["pack_label"] or "").strip()
        if isinstance(row, dict):
            return (row.get("pack_label") or "").strip()
    except (KeyError, TypeError):
        return ""
    return ""


def _update_product_master_price_from_inward(
    conn,
    *,
    item_name: str,
    pack_label: str,
    unit_price: float | None,
) -> None:
    """Persist the inward rate as the latest Product Master / pack price (ex-tax)."""
    if unit_price is None:
        return
    try:
        price = float(unit_price)
    except (TypeError, ValueError):
        return
    if price < 0 or price != price:
        return
    name = (item_name or "").strip()
    if not name:
        return
    product = conn.execute(
        """
        SELECT id, default_unit FROM store_products
        WHERE lower(name) = lower(?) AND is_active = 1
        ORDER BY id
        LIMIT 1
        """,
        (name,),
    ).fetchone()
    if not product:
        return
    pid = int(product["id"])
    pack = (pack_label or "").strip()
    if pack:
        updated = conn.execute(
            """
            UPDATE store_product_variants
            SET approximate_price = ?
            WHERE product_id = ? AND is_active = 1 AND lower(label) = lower(?)
            """,
            (price, pid, pack),
        )
        if updated.rowcount:
            variants = _load_variants_by_product_ids(conn, [pid]).get(pid, [])
            derived = _approximate_price_from_variants(variants)
            if derived is not None:
                conn.execute(
                    """
                    UPDATE store_products
                    SET approximate_price = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (derived, _now(), pid),
                )
                return
            # Pack row updated but no derivable unit rate — still stamp the product.
    conn.execute(
        """
        UPDATE store_products
        SET approximate_price = ?, updated_at = ?
        WHERE id = ?
        """,
        (price, _now(), pid),
    )
    # Fill blank pack-variant ₹ prices from the base-unit inward rate.
    _fill_blank_pack_prices_from_unit_rate(conn, pid, price)


def _fill_blank_pack_prices_from_unit_rate(
    conn, product_id: int, unit_price: float
) -> None:
    """Set missing pack prices to unit_price × qty_in_base (base-unit rate)."""
    try:
        rate = float(unit_price)
    except (TypeError, ValueError):
        return
    if rate < 0 or rate != rate:
        return
    rows = conn.execute(
        """
        SELECT id, qty_in_base, approximate_price
        FROM store_product_variants
        WHERE product_id = ? AND is_active = 1
        """,
        (product_id,),
    ).fetchall()
    for row in rows:
        if row["approximate_price"] is not None:
            continue
        try:
            qty = float(row["qty_in_base"] or 0)
        except (TypeError, ValueError):
            continue
        if qty <= 0:
            continue
        conn.execute(
            """
            UPDATE store_product_variants
            SET approximate_price = ?
            WHERE id = ?
            """,
            (round(rate * qty, 2), int(row["id"])),
        )


def _last_inward_unit_price(conn, item_name: str) -> float | None:
    """Latest stock-inward unit cost for a product name (movement / entered rate)."""
    name = (item_name or "").strip()
    if not name:
        return None
    row = conn.execute(
        """
        SELECT unit_cost
        FROM store_stock_movements
        WHERE lower(item_name) = lower(?)
          AND ref_type IN ('stock_inward', 'stock_inward_direct')
          AND unit_cost IS NOT NULL
        ORDER BY id DESC
        LIMIT 1
        """,
        (name,),
    ).fetchone()
    if not row:
        return None
    try:
        price = float(row["unit_cost"])
    except (TypeError, ValueError):
        return None
    if price < 0 or price != price:
        return None
    return price


def _heal_product_prices_from_last_inward(conn) -> int:
    """Fill blank Product Master approx prices from the newest stock inward.

    Covers inwards that ran before auto-update, or where the rate only landed
    on the stock movement.
    """
    rows = conn.execute(
        """
        SELECT p.id, p.name, p.approximate_price
        FROM store_products p
        WHERE p.is_active = 1
        """
    ).fetchall()
    healed = 0
    for row in rows:
        price = _last_inward_unit_price(conn, row["name"])
        if price is None:
            continue
        needs_product = row["approximate_price"] is None
        if needs_product:
            conn.execute(
                """
                UPDATE store_products
                SET approximate_price = ?, updated_at = ?
                WHERE id = ?
                """,
                (price, _now(), int(row["id"])),
            )
            healed += 1
        blank_packs = conn.execute(
            """
            SELECT COUNT(*) AS c FROM store_product_variants
            WHERE product_id = ? AND is_active = 1 AND approximate_price IS NULL
            """,
            (int(row["id"]),),
        ).fetchone()["c"]
        if needs_product or int(blank_packs or 0):
            unit_rate = price
            if not needs_product:
                try:
                    unit_rate = float(row["approximate_price"])
                except (TypeError, ValueError):
                    unit_rate = price
            before_blank = int(blank_packs or 0)
            _fill_blank_pack_prices_from_unit_rate(conn, int(row["id"]), unit_rate)
            if before_blank:
                healed += 1
    return healed


def _product_supplier_best_prices(conn, item_name: str) -> dict[int, float]:
    """Lowest known rate per supplier for a product (from issued PO lines)."""
    name = (item_name or "").strip()
    if not name:
        return {}
    rows = conn.execute(
        """
        SELECT pl.supplier_id AS supplier_id, MIN(pl.rate) AS best_rate
        FROM store_po_lines pl
        JOIN store_indent_lines il ON il.id = pl.line_id
        WHERE lower(il.item_name) = lower(?)
          AND pl.supplier_id IS NOT NULL
          AND pl.supplier_id > 0
          AND pl.rate IS NOT NULL
        GROUP BY pl.supplier_id
        """,
        (name,),
    ).fetchall()
    out: dict[int, float] = {}
    for row in rows:
        try:
            sid = int(row["supplier_id"] or 0)
            rate = float(row["best_rate"])
        except (TypeError, ValueError):
            continue
        if sid <= 0 or rate < 0 or rate != rate:
            continue
        out[sid] = rate
    return out


def _last_inward_supplier_and_price(
    conn, item_name: str
) -> tuple[int | None, float | None]:
    """Resolve supplier + unit cost from the newest stock inward movement."""
    name = (item_name or "").strip()
    if not name:
        return None, None
    mov = conn.execute(
        """
        SELECT ref_type, ref_id, unit_cost
        FROM store_stock_movements
        WHERE lower(item_name) = lower(?)
          AND ref_type IN ('stock_inward', 'stock_inward_direct')
        ORDER BY id DESC
        LIMIT 1
        """,
        (name,),
    ).fetchone()
    if not mov:
        return None, None
    try:
        unit_cost = float(mov["unit_cost"]) if mov["unit_cost"] is not None else None
    except (TypeError, ValueError):
        unit_cost = None
    if unit_cost is not None and (unit_cost < 0 or unit_cost != unit_cost):
        unit_cost = None

    ref_type = str(mov["ref_type"] or "")
    try:
        ref_id = int(mov["ref_id"] or 0)
    except (TypeError, ValueError):
        ref_id = 0
    supplier_id = None
    if ref_type == "stock_inward_direct" and ref_id > 0:
        exp = conn.execute(
            "SELECT supplier_id FROM sales_update_expenses WHERE id = ?",
            (ref_id,),
        ).fetchone()
        if exp:
            try:
                supplier_id = int(exp["supplier_id"] or 0) or None
            except (TypeError, ValueError):
                supplier_id = None
    elif ref_type == "stock_inward" and ref_id > 0:
        indent = conn.execute(
            "SELECT indent_no FROM store_indents WHERE id = ?",
            (ref_id,),
        ).fetchone()
        indent_no = str(indent["indent_no"] or "").strip() if indent else ""
        if indent_no:
            exp = conn.execute(
                """
                SELECT supplier_id
                FROM sales_update_expenses
                WHERE supplier_id IS NOT NULL
                  AND supplier_id > 0
                  AND instr(lower(coalesce(description, '')), lower(?)) > 0
                ORDER BY id DESC
                LIMIT 1
                """,
                (indent_no,),
            ).fetchone()
            if exp:
                try:
                    supplier_id = int(exp["supplier_id"] or 0) or None
                except (TypeError, ValueError):
                    supplier_id = None
    return supplier_id, unit_cost


def _sync_product_preferred_suppliers_from_history(conn, *, item_name: str) -> bool:
    """Backfill preferred suppliers from the latest inward for this product."""
    supplier_id, unit_price = _last_inward_supplier_and_price(conn, item_name)
    if not supplier_id:
        return False
    before = conn.execute(
        """
        SELECT preferred_supplier_1_id
        FROM store_products
        WHERE lower(name) = lower(?) AND is_active = 1
        ORDER BY id LIMIT 1
        """,
        ((item_name or "").strip(),),
    ).fetchone()
    _update_product_preferred_suppliers_from_inward(
        conn,
        item_name=item_name,
        supplier_id=supplier_id,
        unit_price=unit_price,
    )
    after = conn.execute(
        """
        SELECT preferred_supplier_1_id
        FROM store_products
        WHERE lower(name) = lower(?) AND is_active = 1
        ORDER BY id LIMIT 1
        """,
        ((item_name or "").strip(),),
    ).fetchone()
    before_id = int(before["preferred_supplier_1_id"] or 0) if before else 0
    after_id = int(after["preferred_supplier_1_id"] or 0) if after else 0
    return after_id > 0 and after_id != before_id


def _update_product_preferred_suppliers_from_inward(
    conn,
    *,
    item_name: str,
    supplier_id: Any,
    unit_price: float | None,
) -> None:
    """Refresh Product Master preferred suppliers after stock inward.

    Supplier 1 = last inward supplier. Suppliers 2–3 = next-best known prices
    for this product (PO history + this inward rate).
    """
    try:
        sid = int(supplier_id or 0)
    except (TypeError, ValueError):
        sid = 0
    if sid <= 0:
        return
    name = (item_name or "").strip()
    if not name:
        return
    product = conn.execute(
        """
        SELECT id,
               preferred_supplier_1_id,
               preferred_supplier_2_id,
               preferred_supplier_3_id
        FROM store_products
        WHERE lower(name) = lower(?) AND is_active = 1
        ORDER BY id
        LIMIT 1
        """,
        (name,),
    ).fetchone()
    if not product:
        return
    pid = int(product["id"])

    prices = _product_supplier_best_prices(conn, name)
    if unit_price is not None:
        try:
            price = float(unit_price)
        except (TypeError, ValueError):
            price = None
        else:
            if price >= 0 and price == price:
                prev = prices.get(sid)
                prices[sid] = price if prev is None else min(prev, price)

    # Keep any existing preferred slots that have no price history yet so we
    # don't wipe manual picks until a cheaper known rate appears.
    for raw in (
        product["preferred_supplier_1_id"],
        product["preferred_supplier_2_id"],
        product["preferred_supplier_3_id"],
    ):
        try:
            existing_id = int(raw or 0)
        except (TypeError, ValueError):
            existing_id = 0
        if existing_id > 0 and existing_id not in prices:
            prices[existing_id] = float("inf")

    ranked = sorted(
        prices.items(),
        key=lambda item: (item[1], item[0]),
    )
    chosen: list[int] = [sid]
    for other_id, _rate in ranked:
        if other_id == sid:
            continue
        if other_id in chosen:
            continue
        chosen.append(other_id)
        if len(chosen) >= 3:
            break

    while len(chosen) < 3:
        chosen.append(0)

    conn.execute(
        """
        UPDATE store_products
        SET preferred_supplier_1_id = ?,
            preferred_supplier_2_id = ?,
            preferred_supplier_3_id = ?,
            updated_at = ?
        WHERE id = ?
        """,
        (
            chosen[0] or None,
            chosen[1] or None,
            chosen[2] or None,
            _now(),
            pid,
        ),
    )


def _load_variants_by_product_ids(conn, product_ids: list[int]) -> dict[int, list[dict[str, Any]]]:
    if not product_ids:
        return {}
    placeholders = ",".join("?" for _ in product_ids)
    rows = conn.execute(
        f"""
        SELECT id, product_id, label, qty_in_base, approximate_price, sort_order
        FROM store_product_variants
        WHERE is_active = 1 AND product_id IN ({placeholders})
        ORDER BY sort_order, id
        """,
        product_ids,
    ).fetchall()
    by_product: dict[int, list[dict[str, Any]]] = {int(pid): [] for pid in product_ids}
    for row in rows:
        pid = int(row["product_id"])
        by_product.setdefault(pid, []).append({
            "id": row["id"],
            "label": row["label"],
            "qty_in_base": float(row["qty_in_base"] or 0),
            "qty_in_base_display": _format_ledger_qty(row["qty_in_base"]),
            "approximate_price": row["approximate_price"],
            "approximate_price_display": _format_optional_price(row["approximate_price"]),
        })
    return by_product


def _split_variant_label(label: str) -> tuple[str, str]:
    """Split stored pack label like '500 gram' or 'Half kg' into (qty, unit)."""
    text = (label or "").strip()
    if not text:
        return "", ""
    match = re.match(r"^([\d]+(?:\.[\d]+)?)\s+(.+)$", text)
    if match:
        return match.group(1), match.group(2).strip()
    parts = text.rsplit(None, 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    return text, ""


def _parse_pack_qty_token(raw: Any) -> tuple[float | None, str, str | None]:
    """Parse pack qty token.

    Numeric tokens keep float conversion. Alphanumeric tokens are kept as the
    display qty and count as 1 of the chosen pack unit for stock math.
    Returns (numeric_or_none, display, error).
    """
    text = str(raw or "").strip()
    if not text:
        return None, "", "Pack quantity is required."
    try:
        value = float(text)
    except (TypeError, ValueError):
        return None, text, None
    if value != value:  # NaN
        return None, text, "Pack quantity must be a valid number."
    if value <= 0:
        return None, text, "Pack quantity must be greater than zero."
    return value, text, None


def _infer_product_unit_from_pack_units(pack_units: list[str]) -> tuple[str | None, str | None]:
    """Infer Product Master stock unit from pack variant units.

    Returns (unit, error). Weight packs stock in kg; volume in liter; count in pcs.
    Mixed incompatible families return an error.
    """
    cleaned = [(u or "").strip() for u in pack_units if (u or "").strip()]
    if not cleaned:
        return None, None
    norms = [_normalize_pos_menu_unit(u) for u in cleaned]
    weight = [n for n in norms if n in ("kg", "g")]
    volume = [n for n in norms if n in ("liter", "ml")]
    count = [n for n in norms if n in ("pcs", "dozen")]
    other = [n for n in norms if n not in ("kg", "g", "liter", "ml", "pcs", "dozen")]
    families = sum(1 for group in (weight, volume, count, other) if group)
    if families > 1:
        return None, "Pack units must be in the same family (weight, volume, or count)."
    if weight:
        return "kg", None
    if volume:
        return "liter", None
    if count:
        return "pcs", None
    # Same custom unit (bunch, bottle, …) — keep the first display form.
    first_norm = norms[0]
    for raw, norm in zip(cleaned, norms):
        if norm == first_norm:
            return raw, None
    return cleaned[0], None


def _pack_units_from_form(form) -> list[str]:
    """Collect pack units from filled pack rows (qty present)."""
    qtys = form.getlist("variant_qty")
    units = form.getlist("variant_unit")
    # Legacy forms posted label instead of qty/unit.
    if not qtys and form.getlist("variant_label"):
        labels = form.getlist("variant_label")
        out: list[str] = []
        for raw_label in labels:
            label = (raw_label or "").strip()
            if not label:
                continue
            _qty, unit = _split_variant_label(label)
            if unit:
                out.append(unit)
            else:
                # Unparseable label — still count as a filled pack row for inference skip.
                out.append("")
        return [u for u in out if u]
    out = []
    row_count = max(len(qtys), len(units))
    for idx in range(row_count):
        qty_raw = qtys[idx] if idx < len(qtys) else ""
        unit = ((units[idx] if idx < len(units) else "") or "").strip()
        price_raw = ""
        prices = form.getlist("variant_approximate_price")
        if idx < len(prices):
            price_raw = prices[idx]
        if (not str(qty_raw or "").strip()) and (not str(price_raw or "").strip()):
            continue
        if unit:
            out.append(unit)
    return out


def _approximate_price_from_variants(variants: list[dict[str, Any]]) -> float | None:
    """Derive product unit price from pack prices (₹ per base unit).

    Prefers the pack with the largest qty_in_base so a 1 kg pack anchors the
    implied kg rate when smaller gram packs are also present.
    """
    best: tuple[float, float] | None = None
    for variant in variants or []:
        price = variant.get("approximate_price")
        qty = variant.get("qty_in_base")
        if price is None or qty is None:
            continue
        try:
            price_f = float(price)
            qty_f = float(qty)
        except (TypeError, ValueError):
            continue
        if price_f < 0 or qty_f <= 0 or price_f != price_f or qty_f != qty_f:
            continue
        unit_price = round(price_f / qty_f, 4)
        if best is None or qty_f > best[0]:
            best = (qty_f, unit_price)
    return best[1] if best else None


def _parse_variants_from_form(
    form,
    *,
    product_unit: str,
) -> tuple[list[dict[str, Any]], list[str]]:
    qtys = form.getlist("variant_qty")
    units = form.getlist("variant_unit")
    prices = form.getlist("variant_approximate_price")
    # Legacy field names (older forms) — keep parse tolerant.
    if not qtys and form.getlist("variant_label"):
        labels = form.getlist("variant_label")
        legacy_qtys = form.getlist("variant_qty_in_base")
        variants: list[dict[str, Any]] = []
        errors: list[str] = []
        seen_labels: set[str] = set()
        for idx, raw_label in enumerate(labels):
            label = (raw_label or "").strip()
            qty_raw = legacy_qtys[idx] if idx < len(legacy_qtys) else ""
            price_raw = prices[idx] if idx < len(prices) else ""
            if not label and (not str(qty_raw or "").strip()) and (not str(price_raw or "").strip()):
                continue
            if not label:
                errors.append("Each pack variant needs a quantity and unit.")
                continue
            try:
                qty_in_base = float(qty_raw)
            except (TypeError, ValueError):
                errors.append(f"Pack “{label}” needs a valid quantity.")
                continue
            if qty_in_base <= 0:
                errors.append(f"Pack “{label}” quantity must be greater than zero.")
                continue
            label_key = label.casefold()
            if label_key in seen_labels:
                errors.append(f"Duplicate pack “{label}”.")
                continue
            seen_labels.add(label_key)
            approx_price, price_error = _parse_optional_price(price_raw)
            if price_error:
                errors.append(f"Pack “{label}”: {price_error}")
                continue
            pack_qty, pack_unit = _split_variant_label(label)
            variants.append({
                "label": label,
                "qty_in_base": qty_in_base,
                "approximate_price": approx_price,
                "pack_qty": pack_qty,
                "pack_unit": pack_unit,
            })
        return variants, errors

    variants = []
    errors: list[str] = []
    seen_labels: set[str] = set()
    base_unit = (product_unit or "kg").strip() or "kg"
    row_count = max(len(qtys), len(units), len(prices))
    for idx in range(row_count):
        qty_raw = qtys[idx] if idx < len(qtys) else ""
        unit = ((units[idx] if idx < len(units) else "") or "").strip()
        price_raw = prices[idx] if idx < len(prices) else ""
        # Unit alone does not count — empty rows keep a default unit selected.
        if (not str(qty_raw or "").strip()) and (not str(price_raw or "").strip()):
            continue
        if not str(qty_raw or "").strip() or not unit:
            errors.append("Each pack needs a quantity and unit.")
            continue
        pack_qty_num, qty_display_raw, qty_error = _parse_pack_qty_token(qty_raw)
        if qty_error:
            errors.append(qty_error)
            continue
        if pack_qty_num is not None:
            qty_in_base = _qty_in_product_units(pack_qty_num, unit, base_unit)
            qty_display = _format_ledger_qty(pack_qty_num)
        else:
            # Alphanumeric pack size (e.g. Half, BoxA) — treat as 1 of the pack unit.
            qty_in_base = _qty_in_product_units(1.0, unit, base_unit)
            qty_display = qty_display_raw
        if qty_in_base is None or qty_in_base <= 0:
            errors.append(
                f"Pack unit “{unit}” is not compatible with product unit “{base_unit}”."
            )
            continue
        label = f"{qty_display} {unit}".strip()
        label_key = label.casefold()
        if label_key in seen_labels:
            errors.append(f"Duplicate pack “{label}”.")
            continue
        seen_labels.add(label_key)
        approx_price, price_error = _parse_optional_price(price_raw)
        if price_error:
            errors.append(f"Pack “{label}”: {price_error}")
            continue
        variants.append({
            "label": label,
            "qty_in_base": float(qty_in_base),
            "approximate_price": approx_price,
            "pack_qty": qty_display,
            "pack_unit": unit,
        })
    return variants, errors


def _save_product_variants(conn, product_id: int, variants: list[dict[str, Any]]) -> None:
    existing = conn.execute(
        """
        SELECT id, label FROM store_product_variants
        WHERE product_id = ? AND is_active = 1
        """,
        (product_id,),
    ).fetchall()
    by_label = {
        (row["label"] or "").strip().casefold(): int(row["id"])
        for row in existing
    }
    keep_ids: set[int] = set()
    for sort_idx, variant in enumerate(variants):
        label = variant["label"]
        key = label.casefold()
        existing_id = by_label.get(key)
        if existing_id:
            conn.execute(
                """
                UPDATE store_product_variants
                SET label = ?, qty_in_base = ?, approximate_price = ?, sort_order = ?, is_active = 1
                WHERE id = ?
                """,
                (
                    label,
                    variant["qty_in_base"],
                    variant.get("approximate_price"),
                    (sort_idx + 1) * 10,
                    existing_id,
                ),
            )
            keep_ids.add(existing_id)
        else:
            cur = conn.execute(
                """
                INSERT INTO store_product_variants
                    (product_id, label, qty_in_base, approximate_price, sort_order, is_active)
                VALUES (?, ?, ?, ?, ?, 1)
                """,
                (
                    product_id,
                    label,
                    variant["qty_in_base"],
                    variant.get("approximate_price"),
                    (sort_idx + 1) * 10,
                ),
            )
            keep_ids.add(int(cur.lastrowid))
    for row in existing:
        rid = int(row["id"])
        if rid not in keep_ids:
            conn.execute(
                "UPDATE store_product_variants SET is_active = 0 WHERE id = ?",
                (rid,),
            )


def _unit_cost_with_tax(unit_price: Any, tax_percent: Any) -> float | None:
    """Unit cost including tax: price × (1 + tax%/100)."""
    try:
        price = float(unit_price)
    except (TypeError, ValueError):
        return None
    if price <= 0 or price != price:
        return None
    try:
        tax = float(tax_percent or 0)
    except (TypeError, ValueError):
        tax = 0.0
    if tax < 0:
        tax = 0.0
    return round(price * (1.0 + tax / 100.0), 4)


def _adjust_stock(
    conn,
    *,
    outlet,
    item_name,
    unit,
    qty_delta,
    movement_type,
    ref_type,
    ref_id,
    notes,
    user_id,
    unit_cost=None,
    allow_shortfall=False,
):
    """Adjust on-hand qty and write a stock movement.

    Returns the qty_delta actually applied (may be clamped when allow_shortfall
    is True). Returns 0.0 when nothing was written (no row / zero delta).
    """
    try:
        qty_delta = float(qty_delta)
    except (TypeError, ValueError):
        qty_delta = 0.0
    if abs(qty_delta) < 0.0001:
        return 0.0

    existing = conn.execute(
        """
        SELECT id, qty_on_hand FROM store_stock_items
        WHERE outlet = ? AND lower(item_name) = lower(?) AND lower(unit) = lower(?)
        """,
        (outlet, item_name, unit),
    ).fetchone()
    if existing:
        on_hand = float(existing["qty_on_hand"] or 0)
        new_qty = on_hand + qty_delta
        if new_qty < -0.0001:
            if not allow_shortfall:
                raise ValueError(f"Not enough stock for {item_name} ({unit}).")
            # Deduct only what exists; do not block the caller (e.g. POS sale).
            qty_delta = -on_hand
            new_qty = 0.0
            if abs(qty_delta) < 0.0001:
                return 0.0
        conn.execute(
            """
            UPDATE store_stock_items
            SET qty_on_hand = ?, item_name = ?, unit = ?, updated_at = ?
            WHERE id = ?
            """,
            (round(new_qty, 3), item_name, unit, _now(), existing["id"]),
        )
    else:
        if qty_delta < 0:
            if not allow_shortfall:
                raise ValueError(f"Not enough stock for {item_name} ({unit}).")
            return 0.0
        conn.execute(
            """
            INSERT INTO store_stock_items (outlet, item_name, unit, qty_on_hand, updated_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (outlet, item_name, unit, round(float(qty_delta), 3), _now()),
        )
    cost_value = None
    if unit_cost is not None:
        try:
            cost_value = float(unit_cost)
            if cost_value <= 0 or cost_value != cost_value:
                cost_value = None
        except (TypeError, ValueError):
            cost_value = None
    conn.execute(
        """
        INSERT INTO store_stock_movements
            (outlet, item_name, unit, qty_delta, movement_type, ref_type, ref_id,
             notes, unit_cost, created_by, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            outlet,
            item_name,
            unit,
            float(qty_delta),
            movement_type,
            ref_type,
            ref_id,
            notes or "",
            cost_value,
            user_id,
            _now(),
        ),
    )
    return float(qty_delta)


def _resolve_pos_sale_stock_outlet(conn, product_outlet, item_name, unit) -> str:
    """Pick bar/restaurant stock outlet for a Product Master row."""
    po = _normalize_outlet_key(product_outlet or "restaurant")
    if po in OUTLET_KEYS:
        return po
    # Product outlet is "both" (or unknown): prefer an existing stock row.
    rows = conn.execute(
        """
        SELECT outlet, qty_on_hand
        FROM store_stock_items
        WHERE lower(item_name) = lower(?) AND lower(unit) = lower(?)
        """,
        (item_name, unit),
    ).fetchall()
    if not rows:
        return "restaurant"
    by_outlet = {
        _normalize_outlet_key(r["outlet"]): float(r["qty_on_hand"] or 0) for r in rows
    }
    if by_outlet.get("restaurant", 0) > 0.0001:
        return "restaurant"
    if by_outlet.get("bar", 0) > 0.0001:
        return "bar"
    if "restaurant" in by_outlet:
        return "restaurant"
    if "bar" in by_outlet:
        return "bar"
    return "restaurant"


def deduct_stock_for_pos_invoice(conn, invoice_id, *, user_id=None):
    """Deduct recipe ingredients for a closed POS invoice (idempotent).

    Matches ingredients to ``store_stock_items`` by outlet + product name +
    product default unit (after converting recipe qty). Shortfalls and missing
    recipes/stock rows are skipped with logging — never raises for business
    cases. Marks ``pos_invoices.stock_deducted_at`` so re-close / reprint does
    not double-deduct.
    """
    ensure_pos_schema(conn)
    ensure_stores_schema(conn)
    try:
        invoice_id = int(invoice_id)
    except (TypeError, ValueError):
        logger.warning("POS stock deduct skipped: invalid invoice_id=%r", invoice_id)
        return {"ok": False, "reason": "invalid_id"}

    inv = conn.execute(
        """
        SELECT id, order_no, status, is_active,
               COALESCE(stock_deducted_at, '') AS stock_deducted_at
        FROM pos_invoices
        WHERE id = ?
        """,
        (invoice_id,),
    ).fetchone()
    if not inv or not int(inv["is_active"] or 0):
        return {"ok": False, "reason": "not_found"}
    if (inv["stock_deducted_at"] or "").strip():
        return {"ok": True, "skipped": True, "reason": "already_deducted"}

    existing_mv = conn.execute(
        """
        SELECT id FROM store_stock_movements
        WHERE ref_type = 'pos_invoice' AND ref_id = ?
        LIMIT 1
        """,
        (invoice_id,),
    ).fetchone()
    if existing_mv:
        conn.execute(
            """
            UPDATE pos_invoices
            SET stock_deducted_at = ?, updated_at = ?
            WHERE id = ? AND COALESCE(stock_deducted_at, '') = ''
            """,
            (_now(), _now(), invoice_id),
        )
        return {"ok": True, "skipped": True, "reason": "movements_exist"}

    order_no = (inv["order_no"] or "").strip() or f"#{invoice_id}"
    lines = conn.execute(
        """
        SELECT menu_item_id, name, qty
        FROM pos_invoice_lines
        WHERE invoice_id = ?
        ORDER BY sort_order ASC, id ASC
        """,
        (invoice_id,),
    ).fetchall()

    menu_qty: dict[int, float] = {}
    skipped: list[dict[str, Any]] = []
    for line in lines:
        mid = line["menu_item_id"]
        try:
            qty = float(line["qty"] or 0)
        except (TypeError, ValueError):
            qty = 0.0
        if mid is None:
            skipped.append(
                {
                    "reason": "no_menu_item",
                    "name": line["name"] or "",
                    "qty": qty,
                }
            )
            continue
        if qty <= 0:
            continue
        menu_qty[int(mid)] = menu_qty.get(int(mid), 0.0) + qty

    needs: dict[tuple[str, str, str], float] = {}
    if menu_qty:
        recipes = list_pos_menu_recipe_lines(conn, list(menu_qty.keys()))
        recipes_by_menu: dict[int, list] = {}
        for recipe in recipes:
            recipes_by_menu.setdefault(int(recipe["menu_item_id"]), []).append(recipe)

        for mid, sold_qty in menu_qty.items():
            recipe_lines = recipes_by_menu.get(mid) or []
            if not recipe_lines:
                skipped.append({"reason": "no_recipe", "menu_item_id": mid, "qty": sold_qty})
                logger.info(
                    "POS stock deduct: no recipe for menu_item_id=%s on invoice %s",
                    mid,
                    order_no,
                )
                continue
            for recipe in recipe_lines:
                product_name = (recipe.get("product_name") or "").strip()
                product_unit = (recipe.get("product_unit") or "").strip() or "pcs"
                if not product_name:
                    skipped.append(
                        {
                            "reason": "missing_product",
                            "menu_item_id": mid,
                            "product_id": recipe.get("product_id"),
                        }
                    )
                    logger.info(
                        "POS stock deduct: missing product for recipe on menu_item_id=%s invoice %s",
                        mid,
                        order_no,
                    )
                    continue
                per_portion = _qty_in_product_units(
                    recipe.get("qty"), recipe.get("unit"), product_unit
                )
                if per_portion is None:
                    skipped.append(
                        {
                            "reason": "unit_mismatch",
                            "menu_item_id": mid,
                            "product_name": product_name,
                            "recipe_unit": recipe.get("unit"),
                            "product_unit": product_unit,
                        }
                    )
                    logger.info(
                        "POS stock deduct: unit mismatch %s (%s→%s) invoice %s",
                        product_name,
                        recipe.get("unit"),
                        product_unit,
                        order_no,
                    )
                    continue
                need = float(per_portion) * float(sold_qty)
                if need <= 0:
                    continue
                outlet = _resolve_pos_sale_stock_outlet(
                    conn, recipe.get("product_outlet"), product_name, product_unit
                )
                key = (outlet, product_name, product_unit)
                needs[key] = needs.get(key, 0.0) + need

    deducted: list[dict[str, Any]] = []
    for (outlet, name, unit), need_qty in needs.items():
        applied = _adjust_stock(
            conn,
            outlet=outlet,
            item_name=name,
            unit=unit,
            qty_delta=-abs(need_qty),
            movement_type="sale",
            ref_type="pos_invoice",
            ref_id=invoice_id,
            notes=f"POS sale {order_no}",
            user_id=user_id,
            allow_shortfall=True,
        )
        if abs(applied) < 0.0001:
            skipped.append(
                {
                    "reason": "no_stock",
                    "outlet": outlet,
                    "item_name": name,
                    "unit": unit,
                    "needed": round(need_qty, 4),
                }
            )
            logger.info(
                "POS stock deduct: no on-hand stock for %s (%s) outlet=%s needed=%.4f invoice %s",
                name,
                unit,
                outlet,
                need_qty,
                order_no,
            )
            continue
        shortfall = round(need_qty - abs(applied), 4)
        if shortfall > 0.0001:
            skipped.append(
                {
                    "reason": "partial",
                    "outlet": outlet,
                    "item_name": name,
                    "unit": unit,
                    "needed": round(need_qty, 4),
                    "applied": round(abs(applied), 4),
                    "shortfall": shortfall,
                }
            )
            logger.info(
                "POS stock deduct: partial %s (%s) outlet=%s applied=%.4f needed=%.4f invoice %s",
                name,
                unit,
                outlet,
                abs(applied),
                need_qty,
                order_no,
            )
        deducted.append(
            {
                "outlet": outlet,
                "item_name": name,
                "unit": unit,
                "qty_delta": round(applied, 4),
            }
        )

    conn.execute(
        """
        UPDATE pos_invoices
        SET stock_deducted_at = ?, updated_at = ?
        WHERE id = ?
        """,
        (_now(), _now(), invoice_id),
    )
    return {
        "ok": True,
        "invoice_id": invoice_id,
        "order_no": order_no,
        "deducted": deducted,
        "skipped": skipped,
    }

def _load_product_catalog(conn, stores_outlet: str | None = None):
    """Load product master.

    When stores_outlet is Bar/Restaurant, include that outlet's products plus Both.
    When stores_outlet is Both (or omitted), include every active product.
    """
    filter_outlet = _parse_outlet_filter(stores_outlet) if stores_outlet else None
    params: list[Any] = []
    outlet_sql = ""
    if filter_outlet and filter_outlet != "both":
        outlet_sql = " AND lower(coalesce(p.outlet, '')) IN (?, 'both')"
        params.append(filter_outlet)
    rows = conn.execute(
        f"""
        SELECT c.id AS category_id, c.name AS category_name, c.sort_order AS category_sort,
               p.id AS product_id, p.name AS product_name, p.default_unit, p.outlet,
               p.approximate_price, p.is_active, p.sort_order
        FROM store_product_categories c
        LEFT JOIN store_products p
          ON p.category_id = c.id AND p.is_active = 1{outlet_sql}
        WHERE c.is_active = 1
        ORDER BY c.sort_order, c.name, p.sort_order, p.name
        """,
        params,
    ).fetchall()
    categories = []
    by_id = {}
    product_ids: list[int] = []
    for row in rows:
        cat_id = row["category_id"]
        if cat_id not in by_id:
            node = {
                "id": cat_id,
                "name": row["category_name"],
                "products": [],
            }
            by_id[cat_id] = node
            categories.append(node)
        if row["product_id"]:
            product_ids.append(int(row["product_id"]))
            by_id[cat_id]["products"].append({
                "id": row["product_id"],
                "name": row["product_name"],
                "default_unit": row["default_unit"],
                "outlet": _parse_product_outlet(row["outlet"]),
                "outlet_label": _product_outlet_label(row["outlet"]),
                "approximate_price": row["approximate_price"],
                "approximate_price_display": _format_optional_price(row["approximate_price"]),
                "variants": [],
            })
    variants_by_product = _load_variants_by_product_ids(conn, product_ids)
    for cat in categories:
        for product in cat["products"]:
            product["variants"] = variants_by_product.get(int(product["id"]), [])
    if filter_outlet:
        categories = [cat for cat in categories if cat["products"]]
    return categories


def _product_names_for_outlet(conn, stores_outlet: str) -> set[str]:
    catalog = _load_product_catalog(conn, stores_outlet=stores_outlet)
    names: set[str] = set()
    for cat in catalog:
        for product in cat["products"]:
            names.add(str(product["name"]).strip().lower())
    return names


def _product_category_by_item_name(conn, stores_outlet: str | None = None) -> dict[str, str]:
    """Map product name (casefold) → Product Master category display name."""
    catalog = _load_product_catalog(conn, stores_outlet=stores_outlet)
    mapping: dict[str, str] = {}
    for cat in catalog:
        cat_name = str(cat.get("name") or "").strip()
        if not cat_name:
            continue
        for product in cat.get("products") or []:
            pname = str(product.get("name") or "").strip()
            if pname:
                mapping[pname.casefold()] = cat_name
    return mapping


def _resolve_expense_category_from_product_category(
    product_category_name: str,
    expense_choices: list[tuple[str, str]] | None = None,
) -> tuple[str, str] | None:
    """Map a Product Master category to an expense category key + label."""
    import app as app_module

    raw = (product_category_name or "").strip()
    if not raw:
        return None
    choices = list(expense_choices or app_module.EXPENSE_CATEGORIES)
    # Exact / casefold label match first (e.g. Dairy Products).
    for key, label in choices:
        if label.casefold() == raw.casefold():
            return key, label
    # Builtin aliases (Vegetable → vegetables).
    normalized = app_module._normalize_expense_category(raw)
    if normalized:
        for key, label in choices:
            if key == normalized:
                return key, label
        # Normalized custom key with original label preserved when possible.
        return normalized, raw
    key = app_module._slugify_expense_category_key(raw)
    if not key:
        return None
    return key, raw


def _ensure_expense_category(conn, category_key: str, category_label: str) -> tuple[str, str]:
    """Ensure expense category exists (builtin or custom table); return key, label."""
    import app as app_module
    import re

    key = (category_key or "").strip()
    label = (category_label or "").strip() or key.replace("_", " ").title()
    if not key:
        return "", ""
    for builtin_key, builtin_label in app_module.EXPENSE_CATEGORIES:
        if builtin_key == key or builtin_label.casefold() == label.casefold():
            return builtin_key, builtin_label
    if not re.fullmatch(r"[a-z][a-z0-9_]{0,79}", key):
        key = app_module._slugify_expense_category_key(label) or key
    if not key or not re.fullmatch(r"[a-z][a-z0-9_]{0,79}", key):
        return "", ""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS expense_categories (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            category_key  TEXT    NOT NULL UNIQUE,
            name          TEXT    NOT NULL COLLATE NOCASE,
            sort_order    INTEGER NOT NULL DEFAULT 0,
            is_active     INTEGER NOT NULL DEFAULT 1,
            created_at    TEXT    NOT NULL DEFAULT (datetime('now','localtime'))
        )
    """)
    by_key = conn.execute(
        "SELECT category_key, name, is_active FROM expense_categories WHERE category_key = ?",
        (key,),
    ).fetchone()
    by_name = conn.execute(
        "SELECT category_key, name, is_active FROM expense_categories WHERE lower(name) = lower(?)",
        (label,),
    ).fetchone()
    existing = by_name or by_key
    if existing:
        if int(existing["is_active"] or 0) != 1:
            conn.execute(
                "UPDATE expense_categories SET is_active = 1, name = ? WHERE category_key = ?",
                (label, existing["category_key"]),
            )
        return existing["category_key"], existing["name"] if int(existing["is_active"] or 0) == 1 else label
    max_sort = conn.execute(
        "SELECT COALESCE(MAX(sort_order), 0) AS m FROM expense_categories"
    ).fetchone()["m"]
    conn.execute(
        """
        INSERT INTO expense_categories (category_key, name, sort_order, is_active)
        VALUES (?, ?, ?, 1)
        """,
        (key, label, int(max_sort) + 10),
    )
    return key, label


def _sync_product_categories_into_expense_categories(conn) -> list[tuple[str, str]]:
    """Ensure every Product Master category is selectable as an expense category."""
    import app as app_module

    rows = conn.execute(
        """
        SELECT name FROM store_product_categories
        WHERE is_active = 1
        ORDER BY sort_order, lower(name), id
        """
    ).fetchall()
    for row in rows:
        resolved = _resolve_expense_category_from_product_category(row["name"])
        if not resolved:
            continue
        key, label = resolved
        # Prefer Product Master display name for custom categories.
        if key not in dict(app_module.EXPENSE_CATEGORIES):
            label = (row["name"] or "").strip() or label
        _ensure_expense_category(conn, key, label)
    return app_module._expense_category_choices(conn)


def _pick_expense_category_from_product_categories(
    category_names: list[str],
    amounts: list[float] | None = None,
    expense_choices: list[tuple[str, str]] | None = None,
) -> tuple[str, str] | None:
    """Pick one expense category from product categories (dominant by amount, else first)."""
    pairs: list[tuple[str, float]] = []
    for idx, name in enumerate(category_names):
        label = (name or "").strip()
        if not label:
            continue
        amt = 0.0
        if amounts and idx < len(amounts):
            try:
                amt = float(amounts[idx] or 0)
            except (TypeError, ValueError):
                amt = 0.0
        pairs.append((label, amt))
    if not pairs:
        return None
    totals: dict[str, float] = {}
    order: list[str] = []
    for label, amt in pairs:
        key = label.casefold()
        if key not in totals:
            totals[key] = 0.0
            order.append(label)
        totals[key] += max(amt, 0.0)
    # Prefer highest amount; ties keep first-seen product category.
    best_label = max(order, key=lambda lab: (totals[lab.casefold()], -order.index(lab)))
    return _resolve_expense_category_from_product_category(best_label, expense_choices)


def _inward_line_amount(qty: Any, unit_price: Any, tax_percent: Any = 0) -> float:
    """Line total including tax, rounded like expense amounts."""
    import app as app_module

    unit_cost = _unit_cost_with_tax(unit_price, tax_percent)
    if unit_cost is None:
        return 0.0
    try:
        q = float(qty or 0)
    except (TypeError, ValueError):
        q = 0.0
    if q <= 0:
        return 0.0
    return app_module.round_half_up(q * float(unit_cost), 2)


def _group_inward_lines_by_expense_category(
    conn,
    *,
    stores_outlet: str,
    lines: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], str | None]:
    """Group inward lines by Product Master → expense category.

    Each input line must include item_name, qty, unit_price, and optionally
    tax_percent. Extra keys are preserved on the line within the group.

    Returns (groups, error) where each group is:
      {category_key, category_label, amount, lines[]}
    """
    import app as app_module

    if not lines:
        return [], "No lines to group."

    product_cat_map = _product_category_by_item_name(conn, stores_outlet=None)
    expense_choices = _sync_product_categories_into_expense_categories(conn)
    groups_by_key: dict[str, dict[str, Any]] = {}
    order: list[str] = []

    for line in lines:
        item_name = str(line.get("item_name") or "").strip()
        if not item_name:
            return [], "Every line needs a product name."
        pm_cat = product_cat_map.get(item_name.casefold(), "")
        if not pm_cat:
            return [], f"{item_name} has no Product Master category."
        resolved = _resolve_expense_category_from_product_category(pm_cat, expense_choices)
        if not resolved:
            return [], f"Could not resolve expense category for {item_name}."
        raw_key, raw_label = resolved
        # Prefer Product Master display name for custom categories.
        if raw_key not in dict(app_module.EXPENSE_CATEGORIES):
            raw_label = pm_cat or raw_label
        ensured_key, ensured_label = _ensure_expense_category(conn, raw_key, raw_label)
        category_key = ensured_key or raw_key
        category_label = ensured_label or raw_label or pm_cat
        if not category_key:
            return [], f"Could not resolve expense category for {item_name}."

        amount = _inward_line_amount(
            line.get("qty"),
            line.get("unit_price"),
            line.get("tax_percent"),
        )
        if category_key not in groups_by_key:
            groups_by_key[category_key] = {
                "category_key": category_key,
                "category_label": category_label,
                "amount": 0.0,
                "lines": [],
            }
            order.append(category_key)
        group = groups_by_key[category_key]
        group["lines"].append(line)
        group["amount"] = app_module.round_half_up(float(group["amount"]) + amount, 2)

    groups = [groups_by_key[key] for key in order]
    for group in groups:
        if float(group["amount"] or 0) <= 0:
            return [], (
                f"Category {group['category_label']} needs a positive amount."
            )
    return groups, None


def _create_inward_category_expenses(
    conn,
    user,
    *,
    base_expense_data: dict[str, Any],
    groups: list[dict[str, Any]],
    description_suffix: str = "",
) -> tuple[list[dict[str, Any]], str | None]:
    """Create one Hotel expense per category group; share supplier/invoice/payment.

    Cash is validated once against the grand total. Invoice uniqueness is
    checked once against existing expenses; batch rows may share the invoice.
    """
    import app as app_module

    if not groups:
        return [], "No expense categories to create."

    grand = app_module.round_half_up(
        sum(float(g.get("amount") or 0) for g in groups),
        2,
    )
    posted_raw = base_expense_data.get("amount")
    if posted_raw not in (None, ""):
        posted_amount = app_module.parse_money(posted_raw)
        if posted_amount > 0 and abs(posted_amount - grand) > 1.0:
            return [], "Expense amount does not match invoice line totals."

    company = base_expense_data.get("company") or app_module.DEFAULT_COMPANY
    sales_date = base_expense_data.get("date") or ""
    payment_type = app_module._normalize_expense_payment_type(
        base_expense_data.get("payment_type")
    )
    cash_error = app_module._validate_cash_expense_against_available(
        conn, company, sales_date, grand, payment_type
    )
    if cash_error:
        return [], cash_error

    duplicate = app_module._duplicate_expense_invoice(
        conn,
        base_expense_data.get("supplier_id"),
        base_expense_data.get("invoice_number"),
    )
    if duplicate:
        code = duplicate["expense_code"] or f"#{duplicate['id']}"
        return [], (
            f"An expense with this supplier and invoice number already exists ({code})."
        )

    base_desc = (base_expense_data.get("description") or "").strip()
    results: list[dict[str, Any]] = []
    for group in groups:
        cat_label = str(group.get("category_label") or group.get("category_key") or "").strip()
        if base_desc:
            desc = f"{base_desc} · {cat_label}" if cat_label else base_desc
        elif description_suffix and cat_label:
            desc = f"Stock inward {description_suffix} · {cat_label}"
        elif cat_label:
            desc = f"Stock inward · {cat_label}"
        else:
            desc = "Stock inward"

        expense_data = dict(base_expense_data)
        expense_data["category"] = group["category_key"]
        expense_data["amount"] = group["amount"]
        expense_data["description"] = desc
        result, err = app_module._create_sales_expense(
            conn,
            user,
            expense_data,
            default_location=app_module.OUTLET_HOTEL,
            allow_shared_invoice=True,
            skip_cash_check=True,
        )
        if err:
            return [], err
        entry = {
            "expense_id": result["expense_id"],
            "expense_code": result.get("expense_code"),
            "category": group["category_key"],
            "category_label": group.get("category_label") or "",
            "amount": group["amount"],
        }
        results.append(entry)
        group["expense_id"] = entry["expense_id"]
        group["expense_code"] = entry.get("expense_code")
    return results, None


def _format_ledger_qty(value: Any) -> str:
    try:
        n = float(value or 0)
    except (TypeError, ValueError):
        return "0"
    if abs(n - round(n)) < 0.0001:
        return str(int(round(n)))
    return ("%g" % (round(n * 1000) / 1000))


def _stores_ledger_payload(conn, outlet: str) -> dict[str, Any]:
    """Indent → inward progress ledger for the Indent page popup."""
    outlet_key = _parse_outlet_filter(outlet)
    outlet_sql, outlet_params = _outlet_match_sql("i.outlet", outlet_key)
    rows = conn.execute(
        f"""
        SELECT i.id, i.indent_no, i.outlet, i.status, i.created_at,
               COALESCE((
                   SELECT COUNT(*) FROM store_indent_lines l WHERE l.indent_id = i.id
               ), 0) AS line_count,
               COALESCE((
                   SELECT SUM(COALESCE(l.quantity, 0)) FROM store_indent_lines l WHERE l.indent_id = i.id
               ), 0) AS qty_ordered,
               COALESCE((
                   SELECT SUM(COALESCE(l.quantity_received, 0)) FROM store_indent_lines l WHERE l.indent_id = i.id
               ), 0) AS qty_received,
               COALESCE((
                   SELECT SUM(
                       CASE
                         WHEN COALESCE(l.quantity, 0) - COALESCE(l.quantity_received, 0) > 0.0001
                         THEN COALESCE(l.quantity, 0) - COALESCE(l.quantity_received, 0)
                         ELSE 0
                       END
                   )
                   FROM store_indent_lines l WHERE l.indent_id = i.id
               ), 0) AS qty_pending
        FROM store_indents i
        WHERE {outlet_sql}
        ORDER BY i.created_at DESC, i.id DESC
        LIMIT 100
        """,
        outlet_params,
    ).fetchall()

    indent_ids = [int(row["id"]) for row in rows]
    pending_lines_by_id: dict[int, list[dict[str, Any]]] = {iid: [] for iid in indent_ids}
    received_lines_by_id: dict[int, list[dict[str, Any]]] = {iid: [] for iid in indent_ids}
    item_names_by_id: dict[int, list[str]] = {iid: [] for iid in indent_ids}
    if indent_ids:
        placeholders = ",".join("?" for _ in indent_ids)
        line_rows = conn.execute(
            f"""
            SELECT indent_id, item_name, quantity, quantity_received, unit
            FROM store_indent_lines
            WHERE indent_id IN ({placeholders})
            ORDER BY id
            """,
            indent_ids,
        ).fetchall()
        for line in line_rows:
            try:
                ordered = float(line["quantity"] or 0)
            except (TypeError, ValueError):
                ordered = 0.0
            try:
                received = float(line["quantity_received"] or 0)
            except (TypeError, ValueError, KeyError):
                received = 0.0
            pending = ordered - received
            if pending < 0:
                pending = 0.0
            iid = int(line["indent_id"])
            item_name = (line["item_name"] or "").strip()
            if item_name:
                item_names_by_id.setdefault(iid, []).append(item_name)
            line_payload = {
                "item_name": item_name,
                "unit": line["unit"] or "",
                "qty_ordered": ordered,
                "qty_received": received,
                "qty_pending": pending if pending > 0.0001 else 0.0,
                "qty_ordered_display": _format_ledger_qty(ordered),
                "qty_received_display": _format_ledger_qty(received),
                "qty_pending_display": _format_ledger_qty(pending if pending > 0.0001 else 0.0),
            }
            if pending > 0.0001:
                pending_lines_by_id.setdefault(iid, []).append(line_payload)
            if received > 0.0001:
                received_lines_by_id.setdefault(iid, []).append(line_payload)

    ledger_rows: list[dict[str, Any]] = []
    indents_created = 0
    qty_ordered_sum = 0.0
    qty_received_sum = 0.0
    qty_pending_sum = 0.0
    for row in rows:
        indents_created += 1
        try:
            ordered = float(row["qty_ordered"] or 0)
        except (TypeError, ValueError):
            ordered = 0.0
        try:
            received = float(row["qty_received"] or 0)
        except (TypeError, ValueError):
            received = 0.0
        try:
            pending = float(row["qty_pending"] or 0)
        except (TypeError, ValueError):
            pending = 0.0
        if pending < 0:
            pending = 0.0
        qty_ordered_sum += ordered
        qty_received_sum += received
        qty_pending_sum += pending
        outlet_val = _normalize_outlet_key(row["outlet"] or "restaurant")
        if outlet_val not in ("bar", "restaurant"):
            outlet_val = "restaurant"
        iid = int(row["id"])
        status = row["status"] or ""
        pending_lines = pending_lines_by_id.get(iid, [])
        received_lines = received_lines_by_id.get(iid, [])
        item_names = item_names_by_id.get(iid, [])
        can_view_pending = status == "approved" and pending > 0.0001 and bool(pending_lines)
        can_view_received = received > 0.0001 and bool(received_lines)
        indent_no = row["indent_no"] or ""
        status_label = _status_label(status)
        created_at = _format_stores_dt(row["created_at"] or "")
        outlet_label = _outlet_label(outlet_val)
        qty_ordered_display = _format_ledger_qty(ordered)
        qty_received_display = _format_ledger_qty(received)
        qty_pending_display = _format_ledger_qty(pending)
        search_parts = [
            indent_no,
            outlet_val,
            outlet_label,
            status,
            status_label,
            created_at,
            qty_ordered_display,
            qty_received_display,
            qty_pending_display,
            *item_names,
        ]
        search_text = " ".join(str(part or "").lower() for part in search_parts if part)
        ledger_rows.append({
            "id": iid,
            "indent_no": indent_no,
            "outlet": outlet_val,
            "outlet_label": outlet_label,
            "status": status,
            "status_label": status_label,
            "created_at": created_at,
            "line_count": int(row["line_count"] or 0),
            "qty_ordered": ordered,
            "qty_received": received,
            "qty_pending": pending,
            "qty_ordered_display": qty_ordered_display,
            "qty_received_display": qty_received_display,
            "qty_pending_display": qty_pending_display,
            "can_view_pending": can_view_pending,
            "can_view_received": can_view_received,
            "pending_lines": pending_lines if can_view_pending else [],
            "received_lines": received_lines if can_view_received else [],
            "item_names": item_names,
            "search_text": search_text,
            "inward_url": (
                url_for("stores_purchase_requests", outlet=outlet_val, indent=iid)
                if can_view_pending
                else ""
            ),
        })

    return {
        "summary": {
            "indents_created": indents_created,
            "qty_ordered": qty_ordered_sum,
            "qty_received": qty_received_sum,
            "qty_pending": qty_pending_sum,
            "qty_ordered_display": _format_ledger_qty(qty_ordered_sum),
            "qty_received_display": _format_ledger_qty(qty_received_sum),
            "qty_pending_display": _format_ledger_qty(qty_pending_sum),
        },
        "rows": ledger_rows,
    }


def _indent_view_payload(conn, indents: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Serialize indent rows + lines for the list-page view modal."""
    if not indents:
        return []
    indent_ids = [int(row["id"]) for row in indents if row.get("id") is not None]
    lines_by_id: dict[int, list[dict[str, Any]]] = {iid: [] for iid in indent_ids}
    if indent_ids:
        placeholders = ",".join("?" for _ in indent_ids)
        line_rows = conn.execute(
            f"""
            SELECT indent_id, item_name, quantity, unit, notes, approximate_price,
                   pack_label, pack_qty_in_base
            FROM store_indent_lines
            WHERE indent_id IN ({placeholders})
            ORDER BY id
            """,
            indent_ids,
        ).fetchall()
        for line in line_rows:
            approx = line["approximate_price"] if "approximate_price" in line.keys() else None
            pack_label = ""
            try:
                pack_label = (line["pack_label"] or "").strip()
            except (KeyError, TypeError):
                pack_label = ""
            pack_qty = _row_pack_qty_in_base(line)
            base_unit = line["unit"] or ""
            if pack_label and pack_qty is not None:
                display_unit = f"{_format_ledger_qty(pack_qty)} {base_unit}".strip()
            else:
                display_unit = base_unit
            lines_by_id.setdefault(int(line["indent_id"]), []).append({
                "item_name": line["item_name"],
                "display_name": _format_indent_line_item(line),
                "quantity": line["quantity"],
                "unit": base_unit,
                "display_unit": display_unit,
                "notes": line["notes"] or "",
                "approximate_price": approx,
                "approximate_price_display": _format_optional_price(approx),
                "pack_label": pack_label,
                "pack_qty_in_base": pack_qty,
            })
    payload = []
    for row in indents:
        iid = int(row["id"])
        outlet_key = _parse_outlet(row.get("outlet"))
        payload.append({
            "id": iid,
            "indent_no": row.get("indent_no") or "",
            "outlet": outlet_key,
            "outlet_label": _outlet_label(outlet_key),
            "status": row.get("status") or "",
            "status_label": _status_label(row.get("status") or ""),
            "notes": row.get("notes") or "",
            "decision_note": row.get("decision_note") or "",
            "created_at": _format_stores_dt(row.get("created_at") or ""),
            "created_by_name": row.get("created_by_name") or "",
            "decided_at": _format_stores_dt(row.get("decided_at") or ""),
            "decided_by_name": row.get("decided_by_name") or "",
            "decided_by_username": row.get("decided_by_username") or "",
            "line_count": int(row.get("line_count") or 0),
            "total_qty": row.get("total_qty") or 0,
            "lines": lines_by_id.get(iid, []),
            "can_mutate": (row.get("status") or "") in EDITABLE_INDENT_STATUSES,
            "can_download_po": (row.get("status") or "") == "approved",
            "po_url": url_for("stores_indent_purchase_order", indent_id=iid)
            if (row.get("status") or "") == "approved"
            else "",
            "edit_url": url_for(
                "stores_indent",
                outlet=outlet_key,
                edit=iid,
                focus="form",
            ),
        })
    return payload


def _load_flat_products(conn, stores_outlet: str | None = None) -> list[dict[str, Any]]:
    filter_outlet = _parse_outlet_filter(stores_outlet) if stores_outlet else None
    params: list[Any] = []
    outlet_sql = ""
    if filter_outlet and filter_outlet != "both":
        outlet_sql = " AND lower(coalesce(p.outlet, '')) IN (?, 'both')"
        params.append(filter_outlet)
    rows = conn.execute(
        f"""
        SELECT p.id, p.name, p.default_unit, p.outlet, p.approximate_price, p.category_id,
               c.name AS category_name, c.sort_order AS category_sort, p.sort_order
        FROM store_products p
        JOIN store_product_categories c ON c.id = p.category_id
        WHERE p.is_active = 1 AND c.is_active = 1{outlet_sql}
        ORDER BY c.sort_order, c.name, p.sort_order, p.name
        """,
        params,
    ).fetchall()
    products = []
    for row in rows:
        item = dict(row)
        item["outlet"] = _parse_product_outlet(item.get("outlet"))
        item["outlet_label"] = _product_outlet_label(item["outlet"])
        item["approximate_price_display"] = _format_optional_price(item.get("approximate_price"))
        products.append(item)
    variants_by_product = _load_variants_by_product_ids(
        conn, [int(p["id"]) for p in products]
    )
    for item in products:
        variants = variants_by_product.get(int(item["id"]), [])
        item["variants"] = variants
        item["variant_count"] = len(variants)
    return products


def _first_stores_endpoint(user) -> str | None:
    preferred = (
        "product_master",
        "indent",
        "approvals",
        "purchase_requests",
        "stock",
        "stock_audit",
    )
    endpoint_map = {
        "product_master": "stores_product_master",
        "indent": "stores_indent",
        "approvals": "stores_approvals",
        "purchase_requests": "stores_purchase_requests",
        "stock": "stores_stock",
        "stock_audit": "stores_stock_audit",
    }
    for key in preferred:
        if user_can_access_stores_submodule(user, key):
            return endpoint_map[key]
    return None


def _page_render(page_key: str, **kwargs):
    user = _get_user() if _get_user else None
    raw_outlet = kwargs.pop("outlet", None) or request.args.get("outlet")
    # List filters use All/Bar/Restaurant across Stores pages (including Product Master).
    outlet = _parse_outlet_filter(raw_outlet)
    outlets_for_ui = STORES_FILTER_OUTLETS
    meta = PAGE_META[page_key]
    cta_url = None
    if meta.get("cta_endpoint"):
        args = dict(meta.get("cta_args") or {})
        if args.get("focus") == "form":
            # Create forms need a concrete outlet — only carry Bar/Restaurant.
            if outlet in OUTLET_KEYS:
                args["outlet"] = outlet
        else:
            args["outlet"] = outlet
        cta_url = url_for(meta["cta_endpoint"], **args)
    kwargs.setdefault("auth_notice", _pop_auth_notice() if _pop_auth_notice else None)
    if page_key == "indent":
        kwargs.setdefault("indent_write_outlets", STORES_OUTLETS)
    indent_form_unset = bool(kwargs.pop("indent_form_unset", False))
    selected_outlet = "" if indent_form_unset else outlet
    selected_outlet_label = "Select outlet" if indent_form_unset else _outlet_label(outlet)
    if page_key == "purchase_requests":
        kwargs.setdefault(
            "back_href",
            url_for("stores_stock", outlet=outlet if outlet else "both"),
        )
        kwargs.setdefault("back_label", "Back to Stock")
    nav_stores_view = kwargs.pop("de_nav_stores_view", page_key)
    return render_template(
        "stores_page.html",
        de_nav_section="stores",
        de_nav_stores_view=nav_stores_view,
        stores_outlets=outlets_for_ui,
        selected_outlet=selected_outlet,
        selected_outlet_label=selected_outlet_label,
        indent_form_unset=indent_form_unset,
        page_key=page_key,
        page_title=meta["title"],
        page_subtitle=meta["subtitle"],
        page_list_endpoint=meta["list_endpoint"],
        page_cta=meta.get("cta"),
        page_cta_url=cta_url,
        show_outlet_tabs=meta.get("show_outlet_tabs", True),
        status_label=_status_label,
        stores_dt=_format_stores_dt,
        stores_date=_format_stores_date_line,
        stores_time=_format_stores_time_line,
        default_units=kwargs.pop("default_units", DEFAULT_UNITS),
        product_outlets=PRODUCT_OUTLETS,
        current_user=user,
        **kwargs,
    )


@stores_bp.route("/stores/product-master", methods=["GET", "POST"])
def stores_product_master():
    user = _get_user()
    outlet = _parse_outlet_filter(
        request.args.get("outlet") or request.form.get("list_outlet")
    )
    edit_id = request.args.get("edit") or request.form.get("product_id") or ""
    try:
        edit_id_int = int(edit_id) if str(edit_id).strip() else 0
    except (TypeError, ValueError):
        edit_id_int = 0
    focus = (
        request.args.get("focus") == "form"
        or request.method == "POST"
        or bool(edit_id_int)
    )
    errors: list[str] = []
    show_category_modal = False
    show_unit_modal = False
    category_form_name = ""
    unit_form_name = ""
    form = {
        "product_id": "",
        "category_id": "",
        "name": "",
        "default_unit": "kg",
        "outlet": "",
        "approximate_price": "",
        "preferred_supplier_1_id": "",
        "preferred_supplier_2_id": "",
        "preferred_supplier_3_id": "",
        "variants": [],
    }

    def _pm_redirect(**extra):
        args = {"outlet": outlet}
        if str(request.args.get("embed") or request.form.get("embed") or "") == "1":
            args["embed"] = 1
        args.update(extra)
        return redirect(url_for("stores_product_master", **args))

    conn = get_db()
    try:
        ensure_stores_schema(conn)
        if request.method == "POST":
            action = (request.form.get("action") or "save_product").strip()
            if action == "save_category":
                cat_name = _title_case_words(request.form.get("category_name") or "")
                category_form_name = cat_name
                if not cat_name:
                    errors.append("Category name is required.")
                    show_category_modal = True
                else:
                    exists = conn.execute(
                        "SELECT id FROM store_product_categories WHERE lower(name) = lower(?)",
                        (cat_name,),
                    ).fetchone()
                    if exists:
                        errors.append("That category already exists.")
                        show_category_modal = True
                    else:
                        max_sort = conn.execute(
                            "SELECT COALESCE(MAX(sort_order), 0) AS m FROM store_product_categories"
                        ).fetchone()["m"]
                        cursor = conn.execute(
                            """
                            INSERT INTO store_product_categories (name, sort_order, is_active)
                            VALUES (?, ?, 1)
                            """,
                            (cat_name, int(max_sort) + 10),
                        )
                        conn.commit()
                        flash("Category added.", "ok")
                        return _pm_redirect(focus="form", category=cursor.lastrowid)
            elif action == "save_unit":
                unit_name = (request.form.get("unit_name") or "").strip()
                unit_form_name = unit_name
                if not unit_name:
                    errors.append("Unit name is required.")
                    show_unit_modal = True
                else:
                    exists = conn.execute(
                        "SELECT id FROM store_product_units WHERE lower(name) = lower(?)",
                        (unit_name,),
                    ).fetchone()
                    if exists:
                        errors.append("That unit already exists.")
                        show_unit_modal = True
                    else:
                        max_sort = conn.execute(
                            "SELECT COALESCE(MAX(sort_order), 0) AS m FROM store_product_units"
                        ).fetchone()["m"]
                        conn.execute(
                            """
                            INSERT INTO store_product_units (name, sort_order, is_active)
                            VALUES (?, ?, 1)
                            """,
                            (unit_name, int(max_sort) + 10),
                        )
                        conn.commit()
                        flash("Unit added.", "ok")
                        return _pm_redirect(focus="form", unit=unit_name)
            else:
                form["name"] = _title_case_product_name(request.form.get("name") or "")
                raw_outlet = (request.form.get("outlet") or "").strip().lower()
                form["outlet"] = raw_outlet if raw_outlet in PRODUCT_OUTLET_KEYS else ""
                form["category_id"] = (request.form.get("category_id") or "").strip()
                form["product_id"] = (request.form.get("product_id") or "").strip()
                pref1, pref2, pref3 = _product_preferred_supplier_ids_from_form(request.form)
                form["preferred_supplier_1_id"] = str(pref1) if pref1 else ""
                form["preferred_supplier_2_id"] = str(pref2) if pref2 else ""
                form["preferred_supplier_3_id"] = str(pref3) if pref3 else ""
                try:
                    category_id = int(form["category_id"])
                except (TypeError, ValueError):
                    category_id = 0
                try:
                    product_id = int(form["product_id"]) if form["product_id"] else 0
                except (TypeError, ValueError):
                    product_id = 0

                existing_unit = ""
                if product_id:
                    existing_row = conn.execute(
                        "SELECT default_unit FROM store_products WHERE id = ? AND is_active = 1",
                        (product_id,),
                    ).fetchone()
                    if existing_row:
                        existing_unit = (existing_row["default_unit"] or "").strip()

                pack_units = _pack_units_from_form(request.form)
                inferred_unit, unit_infer_error = _infer_product_unit_from_pack_units(pack_units)
                # Prefer an existing compatible stock unit so edits don't orphan kg stock
                # when packs stay in gram.
                if existing_unit and pack_units:
                    existing_norm = _normalize_pos_menu_unit(existing_unit)
                    pack_norms = {_normalize_pos_menu_unit(u) for u in pack_units}
                    compatible = False
                    if existing_norm in ("kg", "g") and pack_norms <= {"kg", "g"}:
                        compatible = True
                    elif existing_norm in ("liter", "ml") and pack_norms <= {"liter", "ml"}:
                        compatible = True
                    elif existing_norm in ("pcs", "dozen") and pack_norms <= {"pcs", "dozen"}:
                        compatible = True
                    elif existing_norm in pack_norms:
                        compatible = True
                    if compatible:
                        form["default_unit"] = existing_unit
                    else:
                        form["default_unit"] = inferred_unit or existing_unit or "kg"
                else:
                    form["default_unit"] = inferred_unit or existing_unit or "kg"

                variants, variant_errors = _parse_variants_from_form(
                    request.form,
                    product_unit=form["default_unit"],
                )
                approx_price = _approximate_price_from_variants(variants)
                # Pack variants are the only price UI — don't wipe an inward-backed
                # product rate when the editor submits no pack rows.
                if approx_price is None and product_id:
                    keep = conn.execute(
                        """
                        SELECT approximate_price FROM store_products
                        WHERE id = ? AND is_active = 1
                        """,
                        (product_id,),
                    ).fetchone()
                    if keep is not None and keep["approximate_price"] is not None:
                        try:
                            approx_price = float(keep["approximate_price"])
                        except (TypeError, ValueError):
                            approx_price = None
                form["approximate_price"] = _format_optional_price(approx_price) if approx_price is not None else ""
                form["variants"] = [
                    {
                        "qty": v.get("pack_qty") or "",
                        "unit": v.get("pack_unit") or "",
                        "approximate_price": _format_optional_price(v.get("approximate_price")),
                    }
                    for v in variants
                ]
                if not form["name"]:
                    errors.append("Product name is required.")
                if not category_id:
                    errors.append("Choose a category.")
                if not form["outlet"]:
                    errors.append("Choose an outlet.")
                if unit_infer_error:
                    errors.append(unit_infer_error)
                errors.extend(variant_errors)
                for slot, supplier_id in ((1, pref1), (2, pref2), (3, pref3)):
                    if not supplier_id:
                        continue
                    exists_supplier = conn.execute(
                        "SELECT id FROM suppliers WHERE id = ?",
                        (supplier_id,),
                    ).fetchone()
                    if not exists_supplier:
                        errors.append(f"Supplier {slot} was not found.")
                if not errors:
                    exists = conn.execute(
                        """
                        SELECT id FROM store_products
                        WHERE category_id = ? AND lower(name) = lower(?) AND is_active = 1
                          AND (? = 0 OR id != ?)
                        """,
                        (category_id, form["name"], product_id, product_id),
                    ).fetchone()
                    if exists:
                        errors.append("That product already exists in this category.")
                    elif product_id:
                        row = conn.execute(
                            "SELECT id, name FROM store_products WHERE id = ? AND is_active = 1",
                            (product_id,),
                        ).fetchone()
                        if not row:
                            errors.append("Product not found.")
                        else:
                            old_name = row["name"] or ""
                            conn.execute(
                                """
                                UPDATE store_products
                                SET category_id = ?, name = ?, default_unit = ?, outlet = ?,
                                    approximate_price = ?,
                                    preferred_supplier_1_id = ?,
                                    preferred_supplier_2_id = ?,
                                    preferred_supplier_3_id = ?,
                                    updated_at = ?
                                WHERE id = ?
                                """,
                                (
                                    category_id,
                                    form["name"],
                                    form["default_unit"],
                                    form["outlet"],
                                    approx_price,
                                    pref1,
                                    pref2,
                                    pref3,
                                    _now(),
                                    product_id,
                                ),
                            )
                            _rename_store_item_name_refs(conn, old_name, form["name"])
                            _save_product_variants(conn, product_id, variants)
                            conn.commit()
                            flash("Product updated.", "ok")
                            return _pm_redirect()
                    else:
                        max_sort = conn.execute(
                            """
                            SELECT COALESCE(MAX(sort_order), 0) AS m
                            FROM store_products WHERE category_id = ?
                            """,
                            (category_id,),
                        ).fetchone()["m"]
                        cur = conn.execute(
                            """
                            INSERT INTO store_products
                                (category_id, name, default_unit, outlet, approximate_price,
                                 preferred_supplier_1_id, preferred_supplier_2_id, preferred_supplier_3_id,
                                 is_active, sort_order, updated_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                            """,
                            (
                                category_id,
                                form["name"],
                                form["default_unit"],
                                form["outlet"],
                                approx_price,
                                pref1,
                                pref2,
                                pref3,
                                int(max_sort) + 10,
                                _now(),
                            ),
                        )
                        _save_product_variants(conn, int(cur.lastrowid), variants)
                        conn.commit()
                        flash("Product added to master.", "ok")
                        return _pm_redirect()

        if edit_id_int and request.method == "GET":
            row = conn.execute(
                """
                SELECT id, category_id, name, default_unit, outlet, approximate_price,
                       preferred_supplier_1_id, preferred_supplier_2_id, preferred_supplier_3_id
                FROM store_products
                WHERE id = ? AND is_active = 1
                """,
                (edit_id_int,),
            ).fetchone()
            if row:
                # Heal preferred suppliers from the latest stock inward when missing
                # or stale (covers inwards that ran before auto-update shipped).
                try:
                    current_pref = int(row["preferred_supplier_1_id"] or 0)
                except (TypeError, ValueError):
                    current_pref = 0
                last_sid, last_price = _last_inward_supplier_and_price(conn, row["name"])
                healed = False
                if last_sid and last_sid != current_pref:
                    _update_product_preferred_suppliers_from_inward(
                        conn,
                        item_name=row["name"],
                        supplier_id=last_sid,
                        unit_price=last_price,
                    )
                    healed = True
                # Fill blank approx price from the same latest inward.
                try:
                    current_price = (
                        float(row["approximate_price"])
                        if row["approximate_price"] is not None
                        else None
                    )
                except (TypeError, ValueError):
                    current_price = None
                if current_price is None and last_price is not None:
                    _update_product_master_price_from_inward(
                        conn,
                        item_name=row["name"],
                        pack_label="",
                        unit_price=last_price,
                    )
                    healed = True
                elif current_price is None:
                    fallback_price = _last_inward_unit_price(conn, row["name"])
                    if fallback_price is not None:
                        _update_product_master_price_from_inward(
                            conn,
                            item_name=row["name"],
                            pack_label="",
                            unit_price=fallback_price,
                        )
                        healed = True
                if healed:
                    conn.commit()
                    row = conn.execute(
                        """
                        SELECT id, category_id, name, default_unit, outlet, approximate_price,
                               preferred_supplier_1_id, preferred_supplier_2_id, preferred_supplier_3_id
                        FROM store_products
                        WHERE id = ? AND is_active = 1
                        """,
                        (edit_id_int,),
                    ).fetchone()
                form["product_id"] = str(row["id"])
                form["category_id"] = str(row["category_id"])
                form["name"] = row["name"]
                form["default_unit"] = row["default_unit"] or "kg"
                form["outlet"] = _parse_product_outlet(row["outlet"] if "outlet" in row.keys() else None)
                form["approximate_price"] = _format_optional_price(
                    row["approximate_price"] if "approximate_price" in row.keys() else None
                )
                for slot in (1, 2, 3):
                    key = f"preferred_supplier_{slot}_id"
                    val = row[key] if key in row.keys() else None
                    form[key] = str(val) if val else ""
                form["variants"] = []
                for v in _load_variants_by_product_ids(conn, [edit_id_int]).get(edit_id_int, []):
                    pack_qty, pack_unit = _split_variant_label(v["label"] or "")
                    pack_price = v["approximate_price_display"] or ""
                    # Derive blank pack ₹ from the product unit rate (latest inward).
                    if not pack_price and form["approximate_price"]:
                        try:
                            unit_rate = float(form["approximate_price"])
                            qty_base = float(v.get("qty_in_base") or 0)
                        except (TypeError, ValueError):
                            unit_rate = None
                            qty_base = 0.0
                        if unit_rate is not None and qty_base > 0:
                            pack_price = _format_optional_price(round(unit_rate * qty_base, 2))
                    form["variants"].append({
                        "qty": pack_qty,
                        "unit": pack_unit,
                        "approximate_price": pack_price,
                    })
                # No packs yet — show the inward/product rate on the placeholder row
                # (qty 1 × default unit) so Pack variants isn't an empty ₹ Price.
                if not form["variants"] and form["approximate_price"]:
                    form["variants"] = [{
                        "qty": "1",
                        "unit": form["default_unit"] or "kg",
                        "approximate_price": form["approximate_price"],
                    }]
            else:
                flash("Product not found.", "error")
                return _pm_redirect()
        elif request.method == "GET" and not form["category_id"]:
            preselect = (request.args.get("category") or "").strip()
            if preselect.isdigit():
                form["category_id"] = preselect
        if request.method == "GET":
            preselect_unit = (request.args.get("unit") or "").strip()
            if preselect_unit:
                form["default_unit"] = preselect_unit

        catalog = _load_product_catalog(conn, stores_outlet=outlet)
        if _heal_product_prices_from_last_inward(conn):
            conn.commit()
        products = _load_flat_products(conn, stores_outlet=outlet)
        categories = conn.execute(
            """
            SELECT id, name FROM store_product_categories
            WHERE is_active = 1
            ORDER BY sort_order, name
            """
        ).fetchall()
        unit_rows = conn.execute(
            """
            SELECT name FROM store_product_units
            WHERE is_active = 1
            ORDER BY sort_order, name
            """
        ).fetchall()
        product_units = [row["name"] for row in unit_rows]
        if not product_units:
            product_units = list(DEFAULT_UNITS)
        if form.get("default_unit") and form["default_unit"] not in product_units:
            product_units = list(product_units) + [form["default_unit"]]
        product_count = len(products)
        import app as app_module
        suppliers = app_module._all_suppliers(conn)
    finally:
        conn.close()

    if request.method == "GET" and is_embed_request():
        return render_template(
            "partials/master_embed/product.html",
            stores_outlets=STORES_FILTER_OUTLETS,
            selected_outlet=outlet,
            selected_outlet_label=_outlet_label(outlet),
            products=products,
            product_count=product_count,
            categories=[dict(row) for row in categories],
            default_units=product_units,
            product_outlets=PRODUCT_OUTLETS,
            suppliers=suppliers,
            show_form=focus or bool(errors) or show_category_modal or show_unit_modal,
            show_category_modal=show_category_modal,
            show_unit_modal=show_unit_modal,
            category_form_name=category_form_name,
            unit_form_name=unit_form_name,
            form=form,
            errors=errors,
            editing=bool(form.get("product_id")),
        )

    return _page_render(
        "product_master",
        outlet=outlet,
        catalog=catalog,
        products=products,
        categories=[dict(row) for row in categories],
        default_units=product_units,
        product_count=product_count,
        suppliers=suppliers,
        show_form=focus or bool(errors) or show_category_modal or show_unit_modal,
        show_category_modal=show_category_modal,
        show_unit_modal=show_unit_modal,
        category_form_name=category_form_name,
        unit_form_name=unit_form_name,
        form=form,
        errors=errors,
        editing=bool(form.get("product_id")),
    )


@stores_bp.route("/stores/product-master/<int:product_id>/delete", methods=["GET", "POST"])
def stores_product_delete(product_id: int):
    _get_user()
    outlet = _parse_outlet_filter(request.args.get("outlet"))
    wants_json = (
        request.headers.get("X-Requested-With") == "XMLHttpRequest"
        or "application/json" in (request.headers.get("Accept") or "").lower()
    )
    deleted_name = ""
    error = ""
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        row = conn.execute(
            "SELECT id, name FROM store_products WHERE id = ? AND is_active = 1",
            (product_id,),
        ).fetchone()
        if not row:
            error = "Product not found."
            if not wants_json:
                flash(error, "error")
        else:
            deleted_name = row["name"] or ""
            conn.execute(
                """
                UPDATE store_products
                SET is_active = 0, updated_at = ?
                WHERE id = ?
                """,
                (_now(), product_id),
            )
            conn.execute(
                "UPDATE store_product_variants SET is_active = 0 WHERE product_id = ?",
                (product_id,),
            )
            conn.commit()
            if not wants_json:
                flash(f"Deleted {deleted_name}.", "ok")
    finally:
        conn.close()
    if wants_json:
        if error:
            return jsonify({"ok": False, "error": error, "product_id": product_id}), 404
        return jsonify({
            "ok": True,
            "product_id": product_id,
            "name": deleted_name,
            "message": f"Deleted {deleted_name}." if deleted_name else "Deleted.",
        })
    return redirect(
        url_for(
            "stores_product_master",
            outlet=outlet,
            **({"embed": 1} if str(request.args.get("embed") or "") == "1" else {}),
        )
    )

@stores_bp.route("/stores")
def stores():
    user = _get_user()
    endpoint = _first_stores_endpoint(user)
    if not endpoint:
        flash("No Purchase & Inventory pages are available for this account.", "error")
        return redirect(url_for("home"))
    return redirect(
        url_for(endpoint, outlet=request.args.get("outlet") or "both")
    )


@stores_bp.route("/stores/indent", methods=["GET", "POST"])
def stores_indent():
    outlet = _parse_outlet_filter(request.args.get("outlet") or request.form.get("outlet"))
    list_view = _parse_indent_list_view(request.args.get("view") or request.form.get("view"))
    user = _get_user()
    edit_raw = request.args.get("edit") or request.form.get("indent_id") or ""
    try:
        edit_id = int(edit_raw) if str(edit_raw).strip() else 0
    except (TypeError, ValueError):
        edit_id = 0
    # Edit opens in a list-page modal; full-page form is for New Indent / POST errors.
    focus = (
        request.args.get("focus") == "form"
        or request.method == "POST"
    )
    open_edit_id = 0
    errors: list[str] = []
    form = {
        "indent_id": "",
        "notes": "",
        "submission_token": "",
        "lines": [{
            "item_name": "",
            "quantity": "",
            "unit": "kg",
            "notes": "",
            "approximate_price": "",
            "pack_label": "",
            "pack_qty_in_base": "",
        }],
    }

    conn = get_db()
    try:
        ensure_stores_schema(conn)
        if request.method == "POST":
            form["notes"] = (request.form.get("notes") or "").strip()
            form["indent_id"] = (request.form.get("indent_id") or "").strip()
            form["submission_token"] = (request.form.get("submission_token") or "").strip()
            lines = _parse_lines_from_form(request.form)
            form["lines"] = lines or [{
                "item_name": "",
                "quantity": "",
                "unit": "kg",
                "notes": "",
                "approximate_price": "",
                "pack_label": "",
                "pack_qty_in_base": "",
            }]
            for line in form["lines"]:
                line["approximate_price_display"] = _format_optional_price(line.get("approximate_price"))
                line["display_name"] = _format_indent_line_item(line)
            action = (request.form.get("action") or "save").strip()
            form_outlet_raw = (request.form.get("outlet") or "").strip()
            if not form_outlet_raw or _parse_outlet_filter(form_outlet_raw) == "both":
                errors.append("Choose Bar or Restaurant before saving this indent.")
                write_outlet = ""
            else:
                write_outlet = _parse_outlet(form_outlet_raw)
            # Edit modal may send indent_id in the body and/or ?edit= on the action URL.
            indent_id_raw = (
                form.get("indent_id")
                or request.form.get("indent_id")
                or request.args.get("edit")
                or ""
            ).strip()
            try:
                indent_id = int(indent_id_raw) if indent_id_raw else 0
            except (TypeError, ValueError):
                indent_id = 0
            form["indent_id"] = str(indent_id) if indent_id else ""
            existing = None
            if indent_id:
                existing = conn.execute(
                    "SELECT * FROM store_indents WHERE id = ?",
                    (indent_id,),
                ).fetchone()
                if not existing:
                    errors.append("Indent not found.")
                    indent_id = 0
                elif existing["status"] not in EDITABLE_INDENT_STATUSES:
                    errors.append("Only draft, waiting, or rejected indents can be edited.")
                else:
                    write_outlet = _parse_outlet(existing["outlet"])
                    outlet = write_outlet
            if not lines:
                errors.append("Add at least one item with a quantity.")
            else:
                missing_price = [
                    line["item_name"]
                    for line in lines
                    if line.get("approximate_price") is None
                    or float(line.get("approximate_price") or 0) <= 0
                ]
                if missing_price:
                    errors.append("Enter an approximate price greater than 0 for each item.")
            if not errors and lines and write_outlet:
                allowed = _product_names_for_outlet(conn, write_outlet)
                if allowed:
                    bad = sorted({
                        line["item_name"]
                        for line in lines
                        if str(line.get("item_name") or "").strip().lower() not in allowed
                    })
                    if bad:
                        errors.append(
                            "These items are not in the "
                            f"{_outlet_label(write_outlet)} product master: {', '.join(bad)}."
                        )
            if not errors and write_outlet:
                # Create form: save=draft, submit=pending.
                # Edit modal Save is the final save → always Waiting approval (pending).
                if indent_id and existing:
                    status = "pending"
                else:
                    status = "pending" if action == "submit" else "draft"
                if indent_id and existing:
                    prior_status = (existing["status"] or "")
                    # Fresh submitted_at only when (re)entering pending so WhatsApp
                    # idempotency treats reject→resubmit as a new approval round.
                    if status == "pending" and prior_status != "pending":
                        submitted_at = _now()
                    elif status == "pending":
                        submitted_at = existing["submitted_at"] or _now()
                    else:
                        submitted_at = None
                    # Bind the WHERE to the status we read earlier so a concurrent duplicate
                    # request (e.g. an overlapping double-submit) can't both win a "new
                    # approval round" transition and each fire off their own WhatsApp send.
                    update_cur = conn.execute(
                        """
                        UPDATE store_indents
                        SET notes = ?,
                            status = ?,
                            decided_by = NULL,
                            decided_at = NULL,
                            decision_note = '',
                            submitted_at = ?
                        WHERE id = ? AND status = ?
                        """,
                        (
                            form["notes"],
                            status,
                            submitted_at,
                            indent_id,
                            prior_status,
                        ),
                    )
                    won_transition = update_cur.rowcount > 0
                    if not won_transition:
                        # Someone else changed this indent between our read and write
                        # (rare race). Still apply this request's edits so they aren't
                        # silently dropped, just without re-triggering the approval flow.
                        conn.execute(
                            """
                            UPDATE store_indents
                            SET notes = ?, status = ?, submitted_at = ?
                            WHERE id = ?
                            """,
                            (form["notes"], status, submitted_at, indent_id),
                        )
                    conn.execute(
                        "DELETE FROM store_indent_lines WHERE indent_id = ?",
                        (indent_id,),
                    )
                    for line in lines:
                        conn.execute(
                            """
                            INSERT INTO store_indent_lines
                                (indent_id, item_name, quantity, unit, notes, approximate_price,
                                 pack_label, pack_qty_in_base)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                indent_id,
                                line["item_name"],
                                line["quantity"],
                                line["unit"],
                                line.get("notes") or "",
                                line.get("approximate_price"),
                                line.get("pack_label") or "",
                                line.get("pack_qty_in_base"),
                            ),
                        )
                    # New approval round: allow WhatsApp notify again after reject/draft.
                    is_new_round = won_transition and status == "pending" and prior_status != "pending"
                    if is_new_round:
                        supersede_indent_whatsapp_sends(conn, indent_id)
                        assign_fresh_approval_token(conn, indent_id)
                    conn.commit()
                    if prior_status == "rejected":
                        msg = "Indent updated and sent for approval."
                    elif prior_status != "pending":
                        msg = "Indent sent for approval."
                    else:
                        msg = "Indent updated."
                    flash(msg, "ok")
                    # Only notify when entering pending (draft/rejected → pending).
                    # Editing an already-pending indent must not re-spam WhatsApp.
                    if is_new_round:
                        _notify_indent_pending_whatsapp(conn, indent_id, write_outlet)
                    return redirect(url_for("stores_indent", outlet=write_outlet, view="pending"))

                # Guard against duplicate indents from a double form submit (double-click,
                # soft-nav retry, browser resubmit): the same rendered form carries a
                # one-time token, so a repeat POST is recognised and short-circuited here
                # instead of creating a second indent + sending a second approval request.
                submission_token = form["submission_token"]
                dup_indent = None
                if submission_token:
                    dup_indent = conn.execute(
                        "SELECT id, outlet, status FROM store_indents WHERE submission_token = ?",
                        (submission_token,),
                    ).fetchone()
                if dup_indent:
                    flash(
                        "Indent sent for approval." if status == "pending" else "Indent saved as draft.",
                        "ok",
                    )
                    return redirect(
                        url_for(
                            "stores_indent",
                            outlet=dup_indent["outlet"] or write_outlet,
                            view="pending" if dup_indent["status"] == "pending" else list_view,
                        )
                    )

                indent_no = _next_indent_no(conn, write_outlet)
                try:
                    cur = conn.execute(
                        """
                        INSERT INTO store_indents
                            (outlet, indent_no, status, notes, created_by, created_at, submitted_at, submission_token)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            write_outlet,
                            indent_no,
                            status,
                            form["notes"],
                            user["id"] if user else None,
                            _now(),
                            _now() if status == "pending" else None,
                            submission_token,
                        ),
                    )
                except sqlite3.IntegrityError:
                    # Lost the race to a concurrent duplicate submit with the same token.
                    conn.rollback()
                    dup_indent = conn.execute(
                        "SELECT id, outlet, status FROM store_indents WHERE submission_token = ?",
                        (submission_token,),
                    ).fetchone()
                    flash(
                        "Indent sent for approval." if status == "pending" else "Indent saved as draft.",
                        "ok",
                    )
                    return redirect(
                        url_for(
                            "stores_indent",
                            outlet=(dup_indent["outlet"] if dup_indent else write_outlet),
                            view="pending" if (dup_indent and dup_indent["status"] == "pending") else list_view,
                        )
                    )
                new_id = cur.lastrowid
                for line in lines:
                    conn.execute(
                        """
                        INSERT INTO store_indent_lines
                            (indent_id, item_name, quantity, unit, notes, approximate_price,
                             pack_label, pack_qty_in_base)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            new_id,
                            line["item_name"],
                            line["quantity"],
                            line["unit"],
                            line.get("notes") or "",
                            line.get("approximate_price"),
                            line.get("pack_label") or "",
                            line.get("pack_qty_in_base"),
                        ),
                    )
                if status == "pending":
                    assign_fresh_approval_token(conn, new_id)
                conn.commit()
                msg = "Indent sent for approval." if status == "pending" else "Indent saved as draft."
                flash(msg, "ok")
                if status == "pending":
                    _notify_indent_pending_whatsapp(conn, new_id, write_outlet)
                return redirect(
                    url_for(
                        "stores_indent",
                        outlet=write_outlet,
                        view="pending" if status == "pending" else list_view,
                    )
                )

        if edit_id and request.method == "GET":
            row = conn.execute(
                """
                SELECT * FROM store_indents WHERE id = ?
                """,
                (edit_id,),
            ).fetchone()
            if not row:
                flash("Indent not found.", "error")
                return redirect(url_for("stores_indent", outlet=outlet, view=list_view))
            if row["status"] not in EDITABLE_INDENT_STATUSES:
                flash("Only draft, waiting, or rejected indents can be edited.", "error")
                if row["status"] == "rejected":
                    return redirect(url_for("stores_indent", outlet=row["outlet"], view="rejected"))
                if row["status"] == "approved":
                    return redirect(url_for("stores_indent", outlet=row["outlet"], view="approved"))
                return redirect(url_for("stores_indent", outlet=row["outlet"]))
            outlet = _parse_outlet(row["outlet"])
            open_edit_id = edit_id
            list_view = "rejected" if row["status"] == "rejected" else "pending"

        # Create form: no default outlet — user must pick Bar or Restaurant.
        indent_form_unset = bool(
            ((focus and not open_edit_id) or request.method == "POST")
            and outlet == "both"
            and not open_edit_id
        )
        if indent_form_unset:
            catalog = []
        else:
            catalog = _load_product_catalog(conn, stores_outlet=outlet)
        status_keys = INDENT_LIST_VIEW_STATUSES[list_view]
        status_placeholders = ",".join("?" for _ in status_keys)
        stores_ledger_data = {
            "summary": {
                "indents_created": 0,
                "qty_ordered": 0,
                "qty_received": 0,
                "qty_pending": 0,
                "qty_ordered_display": "0",
                "qty_received_display": "0",
                "qty_pending_display": "0",
            },
            "rows": [],
        }
        if outlet == "both":
            rows = conn.execute(
                f"""
                SELECT i.*, u.full_name AS created_by_name,
                       d.full_name AS decided_by_name,
                       d.username AS decided_by_username,
                       (SELECT COUNT(*) FROM store_indent_lines l WHERE l.indent_id = i.id) AS line_count,
                       (SELECT COALESCE(SUM(l.quantity), 0) FROM store_indent_lines l WHERE l.indent_id = i.id) AS total_qty
                FROM store_indents i
                LEFT JOIN users u ON u.id = i.created_by
                LEFT JOIN users d ON d.id = i.decided_by
                WHERE i.outlet IN ('bar', 'restaurant')
                  AND i.status IN ({status_placeholders})
                ORDER BY i.created_at DESC, i.id DESC
                """,
                status_keys,
            ).fetchall()
        else:
            rows = conn.execute(
                f"""
                SELECT i.*, u.full_name AS created_by_name,
                       d.full_name AS decided_by_name,
                       d.username AS decided_by_username,
                       (SELECT COUNT(*) FROM store_indent_lines l WHERE l.indent_id = i.id) AS line_count,
                       (SELECT COALESCE(SUM(l.quantity), 0) FROM store_indent_lines l WHERE l.indent_id = i.id) AS total_qty
                FROM store_indents i
                LEFT JOIN users u ON u.id = i.created_by
                LEFT JOIN users d ON d.id = i.decided_by
                WHERE i.outlet = ?
                  AND i.status IN ({status_placeholders})
                ORDER BY i.created_at DESC, i.id DESC
                """,
                (outlet, *status_keys),
            ).fetchall()
        indents = [dict(row) for row in rows]
        indent_view_data = _indent_view_payload(conn, indents)
        stores_ledger_data = _stores_ledger_payload(conn, "both")
    finally:
        conn.close()

    show_form = (focus and not open_edit_id) or bool(errors)
    indent_form_unset = bool(show_form and outlet == "both" and not open_edit_id)
    # Fresh "New Indent" form: mint a one-time token so a resubmitted POST (double
    # click, soft-nav retry) can be recognised server-side and de-duplicated.
    if show_form and not form.get("indent_id") and not form.get("submission_token"):
        form["submission_token"] = uuid.uuid4().hex

    return _page_render(
        "indent",
        outlet=outlet,
        indents=indents,
        indent_view_data=indent_view_data,
        stores_ledger_data=stores_ledger_data,
        product_catalog=catalog,
        show_form=show_form,
        open_edit_id=open_edit_id,
        indent_form_unset=indent_form_unset,
        form=form,
        errors=errors,
        editing=bool(form.get("indent_id")),
        indent_list_views=INDENT_LIST_VIEWS,
        selected_indent_view=list_view,
        de_nav_stores_view="indent",
    )


def _load_indent_list_for_view(conn, outlet: str, list_view: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    """Return (indents, indent_view_data, stores_ledger_data) for a list view."""
    status_keys = INDENT_LIST_VIEW_STATUSES[list_view]
    status_placeholders = ",".join("?" for _ in status_keys)
    if outlet == "both":
        rows = conn.execute(
            f"""
            SELECT i.*, u.full_name AS created_by_name,
                   d.full_name AS decided_by_name,
                   d.username AS decided_by_username,
                   (SELECT COUNT(*) FROM store_indent_lines l WHERE l.indent_id = i.id) AS line_count,
                   (SELECT COALESCE(SUM(l.quantity), 0) FROM store_indent_lines l WHERE l.indent_id = i.id) AS total_qty
            FROM store_indents i
            LEFT JOIN users u ON u.id = i.created_by
            LEFT JOIN users d ON d.id = i.decided_by
            WHERE i.outlet IN ('bar', 'restaurant')
              AND i.status IN ({status_placeholders})
            ORDER BY i.created_at DESC, i.id DESC
            """,
            status_keys,
        ).fetchall()
    else:
        rows = conn.execute(
            f"""
            SELECT i.*, u.full_name AS created_by_name,
                   d.full_name AS decided_by_name,
                   d.username AS decided_by_username,
                   (SELECT COUNT(*) FROM store_indent_lines l WHERE l.indent_id = i.id) AS line_count,
                   (SELECT COALESCE(SUM(l.quantity), 0) FROM store_indent_lines l WHERE l.indent_id = i.id) AS total_qty
            FROM store_indents i
            LEFT JOIN users u ON u.id = i.created_by
            LEFT JOIN users d ON d.id = i.decided_by
            WHERE i.outlet = ?
              AND i.status IN ({status_placeholders})
            ORDER BY i.created_at DESC, i.id DESC
            """,
            (outlet, *status_keys),
        ).fetchall()
    indents = [dict(row) for row in rows]
    indent_view_data = _indent_view_payload(conn, indents)
    stores_ledger_data = _stores_ledger_payload(conn, "both")
    return indents, indent_view_data, stores_ledger_data


def _supplier_initials(name: str) -> str:
    parts = [p for p in re.split(r"\s+", str(name or "").strip()) if p]
    if not parts:
        return "?"
    if len(parts) == 1:
        return parts[0][:2].upper()
    return (parts[0][0] + parts[-1][0]).upper()


def _po_line_display_unit(line: dict[str, Any]) -> str:
    pack_label = (line.get("pack_label") or "").strip()
    if pack_label:
        return pack_label
    return str(line.get("unit") or "").strip()


def _po_line_total_display(line: dict[str, Any]) -> str:
    """Base quantity with unit (pack qty × line qty when packed), matching the PO PDF."""
    try:
        qty = float(line.get("quantity") or 0)
    except (TypeError, ValueError):
        qty = 0.0
    unit = str(line.get("unit") or "").strip()
    pack_label = (line.get("pack_label") or "").strip()
    pack_raw = line.get("pack_qty_in_base")
    pack_qty = None
    if pack_label and pack_raw not in (None, ""):
        try:
            pack_qty = float(pack_raw)
        except (TypeError, ValueError):
            pack_qty = None
        if pack_qty is not None and pack_qty <= 0:
            pack_qty = None
    total = qty * pack_qty if pack_qty is not None else qty
    qty_label = _format_ledger_qty(total)
    return f"{qty_label} {unit}".strip() if unit else qty_label


def _parse_selected_supplier_ids(*raw_values) -> list[int]:
    """Parse selected supplier ids from form getlist values and/or a CSV query string."""
    out: list[int] = []
    seen: set[int] = set()
    for raw in raw_values:
        if raw is None:
            continue
        if isinstance(raw, (list, tuple)):
            parts = raw
        else:
            parts = str(raw).split(",")
        for part in parts:
            text = str(part or "").strip()
            if not text or text in ("0", "none", "null"):
                continue
            try:
                sid = int(text)
            except (TypeError, ValueError):
                continue
            if sid <= 0 or sid in seen:
                continue
            seen.add(sid)
            out.append(sid)
    return out


_PO_FY_NO_RE = re.compile(r"^HBE/PO/(\d+)/(\d{4}-\d{2})$", re.IGNORECASE)
_PO_SHORT_FY_NO_RE = re.compile(
    r"^PO/(BAR|RES)/(\d{2}-\d{2})/(\d+)$",
    re.IGNORECASE,
)


def _next_po_no(conn, outlet: str, when=None) -> str:
    """Allocate PO/{BAR|RES}/{YY-YY}/{n} — series per outlet + fiscal year."""
    outlet_key = _parse_outlet(outlet)
    outlet_code = _indent_outlet_code(outlet_key)
    fy = indian_fiscal_year_label(when)
    short_fy = _short_fiscal_year_label(when)
    series_key = f"{outlet_code}/{short_fy}"
    row = conn.execute(
        "SELECT last_seq FROM store_po_seq WHERE fiscal_year = ?",
        (series_key,),
    ).fetchone()
    current = int(row["last_seq"]) if row else 0
    # Issued numbers win, so a missing/reset counter can never hand out a duplicate.
    for issued in conn.execute(
        """
        SELECT po.po_no, i.outlet
        FROM store_purchase_orders po
        JOIN store_indents i ON i.id = po.indent_id
        WHERE i.outlet = ?
        """,
        (outlet_key,),
    ).fetchall():
        text = str(issued["po_no"] or "").strip()
        match_new = _PO_SHORT_FY_NO_RE.match(text)
        if match_new:
            if match_new.group(1).upper() != outlet_code:
                continue
            if match_new.group(2) != short_fy:
                continue
            current = max(current, int(match_new.group(3)))
            continue
        match_old = _PO_FY_NO_RE.match(text)
        if match_old and match_old.group(2) == fy:
            current = max(current, int(match_old.group(1)))
    nxt = current + 1
    conn.execute(
        """
        INSERT INTO store_po_seq (fiscal_year, last_seq)
        VALUES (?, ?)
        ON CONFLICT(fiscal_year) DO UPDATE SET last_seq = excluded.last_seq
        """,
        (series_key, nxt),
    )
    return f"PO/{outlet_code}/{short_fy}/{nxt}"


def _find_po_no(conn, indent_id: int, supplier_id: int) -> str:
    """Most recent PO number for indent × supplier (multiple batches allowed)."""
    row = conn.execute(
        """
        SELECT po_no FROM store_purchase_orders
        WHERE indent_id = ? AND supplier_id = ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (indent_id, supplier_id),
    ).fetchone()
    return str(row["po_no"] or "").strip() if row else ""


def _get_or_create_po_no(conn, indent_id: int, supplier_id: int, when=None) -> str:
    """Return the latest PO for this supplier, or allocate one if none exist."""
    if not indent_id or not supplier_id:
        return ""
    existing = _find_po_no(conn, indent_id, supplier_id)
    if existing:
        return existing
    return _allocate_po_no(conn, indent_id, supplier_id, when=when)


def _allocate_po_no(conn, indent_id: int, supplier_id: int, when=None) -> str:
    """Always allocate a fresh PO number (partial re-orders need a new PO)."""
    po_id, po_no = _allocate_po_row(conn, indent_id, supplier_id, when=when)
    return po_no if po_id else ""


def _allocate_po_row(conn, indent_id: int, supplier_id: int, when=None) -> tuple[int, str]:
    """Allocate a fresh PO row; returns ``(purchase_order_id, po_no)`` or ``(0, "")``."""
    if not indent_id or not supplier_id:
        return 0, ""
    indent = conn.execute(
        "SELECT outlet FROM store_indents WHERE id = ?",
        (indent_id,),
    ).fetchone()
    outlet = (indent["outlet"] if indent else "") or "restaurant"
    for _ in range(5):
        po_no = _next_po_no(conn, outlet, when)
        try:
            cur = conn.execute(
                """
                INSERT INTO store_purchase_orders (indent_id, supplier_id, po_no, created_at)
                VALUES (?, ?, ?, datetime('now','localtime'))
                """,
                (indent_id, supplier_id, po_no),
            )
            conn.commit()
            return int(cur.lastrowid), po_no
        except sqlite3.IntegrityError:
            conn.rollback()
    return 0, ""


def _save_purchase_order_lines(
    conn, purchase_order_id: int, lines: list[dict[str, Any]]
) -> None:
    """Freeze the lines that were on a PO at generate time (for deferred send/PDF)."""
    if not purchase_order_id:
        return
    conn.execute(
        "DELETE FROM store_purchase_order_lines WHERE purchase_order_id = ?",
        (purchase_order_id,),
    )
    for line in lines or []:
        try:
            qty = float(line.get("quantity") or 0)
        except (TypeError, ValueError):
            qty = 0.0
        if qty <= 0.0001:
            continue
        line_id = line.get("line_id")
        try:
            line_id_int = int(line_id) if line_id not in (None, "") else None
        except (TypeError, ValueError):
            line_id_int = None
        rate = line.get("rate")
        try:
            rate_num = float(rate) if rate not in (None, "") else None
        except (TypeError, ValueError):
            rate_num = None
        pack_raw = line.get("pack_qty_in_base")
        try:
            pack_qty = float(pack_raw) if pack_raw not in (None, "") else None
        except (TypeError, ValueError):
            pack_qty = None
        conn.execute(
            """
            INSERT INTO store_purchase_order_lines (
                purchase_order_id, line_id, item_name, display_name,
                quantity, unit, pack_label, pack_qty_in_base, rate
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                purchase_order_id,
                line_id_int,
                str(line.get("item_name") or "").strip(),
                str(line.get("display_name") or line.get("item_name") or "").strip(),
                qty,
                str(line.get("unit") or line.get("display_unit") or "").strip(),
                str(line.get("pack_label") or "").strip(),
                pack_qty,
                rate_num,
            ),
        )
    conn.commit()


def _find_purchase_order(
    conn, *, indent_id: int, supplier_id: int, po_no: str = ""
) -> dict[str, Any] | None:
    """Resolve a purchase order row by po_no, or the latest for indent×supplier."""
    text = str(po_no or "").strip()
    if text:
        row = conn.execute(
            """
            SELECT id, indent_id, supplier_id, po_no, created_at
            FROM store_purchase_orders
            WHERE indent_id = ? AND supplier_id = ? AND po_no = ?
            LIMIT 1
            """,
            (indent_id, supplier_id, text),
        ).fetchone()
        if row:
            return dict(row)
    row = conn.execute(
        """
        SELECT id, indent_id, supplier_id, po_no, created_at
        FROM store_purchase_orders
        WHERE indent_id = ? AND supplier_id = ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (indent_id, supplier_id),
    ).fetchone()
    return dict(row) if row else None


def _load_purchase_order_lines(conn, purchase_order_id: int) -> list[dict[str, Any]]:
    """Frozen lines for one issued PO (empty if none stored)."""
    if not purchase_order_id:
        return []
    rows = conn.execute(
        """
        SELECT line_id, item_name, display_name, quantity, unit,
               pack_label, pack_qty_in_base, rate,
               COALESCE(quantity_received, 0) AS quantity_received
        FROM store_purchase_order_lines
        WHERE purchase_order_id = ?
        ORDER BY id
        """,
        (purchase_order_id,),
    ).fetchall()
    out: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        item["display_unit"] = _po_line_display_unit(item)
        out.append(item)
    return out


def _po_inward_available_qty(
    *, po_qty: float, po_received: float, indent_remaining: float
) -> float:
    """Qty still receivable on a PO line, capped by indent remaining."""
    try:
        ordered = float(po_qty or 0)
    except (TypeError, ValueError):
        ordered = 0.0
    try:
        already = float(po_received or 0)
    except (TypeError, ValueError):
        already = 0.0
    try:
        indent_left = float(indent_remaining or 0)
    except (TypeError, ValueError):
        indent_left = 0.0
    if ordered <= 0.0001:
        # Legacy / reconstructed lines without a frozen PO qty: fall back to indent.
        return indent_left if indent_left > 0.0001 else 0.0
    po_left = ordered - already
    if po_left <= 0.0001 or indent_left <= 0.0001:
        return 0.0
    return po_left if po_left <= indent_left else indent_left


def _apply_po_line_received(
    conn, *, purchase_order_id: int, indent_line_id: int, received_qty: float
) -> None:
    """Increment quantity_received on the matching frozen PO line."""
    if not purchase_order_id or not indent_line_id:
        return
    try:
        qty = float(received_qty or 0)
    except (TypeError, ValueError):
        return
    if qty <= 0.0001:
        return
    row = conn.execute(
        """
        SELECT id, COALESCE(quantity, 0) AS quantity,
               COALESCE(quantity_received, 0) AS quantity_received
        FROM store_purchase_order_lines
        WHERE purchase_order_id = ? AND line_id = ?
        ORDER BY id
        LIMIT 1
        """,
        (int(purchase_order_id), int(indent_line_id)),
    ).fetchone()
    if not row:
        return
    try:
        already = float(row["quantity_received"] or 0)
        ordered = float(row["quantity"] or 0)
    except (TypeError, ValueError):
        already = 0.0
        ordered = 0.0
    new_received = already + qty
    if ordered > 0.0001 and new_received > ordered + 0.0001:
        new_received = ordered
    conn.execute(
        """
        UPDATE store_purchase_order_lines
        SET quantity_received = ?
        WHERE id = ?
        """,
        (new_received, int(row["id"])),
    )


def _reconstruct_po_lines_from_assignment(
    conn, indent_id: int, supplier_id: int
) -> list[dict[str, Any]]:
    """Build PO lines from indent + supplier assignment when frozen rows are missing.

    Used for older purchase orders created before ``store_purchase_order_lines``
    was populated. Prefers ``quantity_ordered``, then override qty, then indent qty.
    """
    try:
        supplier_id_int = int(supplier_id)
    except (TypeError, ValueError):
        return []
    if supplier_id_int <= 0:
        return []
    rows = conn.execute(
        """
        SELECT l.id AS line_id,
               l.item_name,
               l.quantity,
               l.unit,
               l.notes,
               l.approximate_price,
               l.pack_label,
               l.pack_qty_in_base,
               COALESCE(l.quantity_ordered, 0) AS quantity_ordered,
               pl.rate AS override_rate,
               pl.quantity AS override_qty
        FROM store_indent_lines l
        INNER JOIN store_po_lines pl
          ON pl.indent_id = l.indent_id AND pl.line_id = l.id
        WHERE l.indent_id = ?
          AND pl.supplier_id = ?
        ORDER BY l.id
        """,
        (indent_id, supplier_id_int),
    ).fetchall()
    out: list[dict[str, Any]] = []
    for row in rows:
        line = dict(row)
        try:
            ordered_qty = float(line.get("quantity_ordered") or 0)
        except (TypeError, ValueError):
            ordered_qty = 0.0
        try:
            indent_qty = float(line.get("quantity") or 0)
        except (TypeError, ValueError):
            indent_qty = 0.0
        override_qty = line.get("override_qty")
        try:
            override_num = (
                float(override_qty) if override_qty not in (None, "") else None
            )
        except (TypeError, ValueError):
            override_num = None
        if ordered_qty > 0.0001:
            qty = ordered_qty
        elif override_num is not None and override_num > 0.0001:
            qty = override_num
        else:
            qty = indent_qty
        if qty <= 0.0001:
            continue
        rate = line.get("override_rate")
        if rate in (None, ""):
            rate = line.get("approximate_price")
        try:
            rate_num = float(rate) if rate not in (None, "") else None
        except (TypeError, ValueError):
            rate_num = None
        display_name = _format_indent_line_item(line)
        line_for_display = dict(line)
        line_for_display["quantity"] = qty
        out.append(
            {
                "line_id": int(line["line_id"]),
                "item_name": line.get("item_name") or "",
                "display_name": display_name,
                "quantity": qty,
                "unit": line.get("unit") or "",
                "display_unit": _po_line_display_unit(line),
                "pack_label": (line.get("pack_label") or "").strip(),
                "pack_qty_in_base": line.get("pack_qty_in_base"),
                "total_display": _po_line_total_display(line_for_display),
                "rate": rate_num,
                "approximate_price": line.get("approximate_price"),
                "notes": line.get("notes") or "",
                "supplier_id": supplier_id_int,
            }
        )
    return out


def _po_lines_for_send_or_pdf(
    conn, *, indent_id: int, supplier_id: int, po_row: dict | None, group_lines: list | None = None
) -> list[dict[str, Any]]:
    """Prefer frozen PO lines; otherwise reconstruct and backfill for older POs."""
    frozen: list[dict[str, Any]] = []
    if po_row and po_row.get("id"):
        frozen = _load_purchase_order_lines(conn, int(po_row["id"]))
    if frozen:
        return frozen
    reconstructed = _reconstruct_po_lines_from_assignment(conn, indent_id, supplier_id)
    if reconstructed:
        if po_row and po_row.get("id"):
            _save_purchase_order_lines(conn, int(po_row["id"]), reconstructed)
        return reconstructed
    return list(group_lines or [])


def _format_qty_display(value: float) -> str:
    try:
        num = float(value or 0)
    except (TypeError, ValueError):
        num = 0.0
    if abs(num - round(num)) < 0.0001:
        return str(int(round(num)))
    return "%g" % num


def _build_inward_lines_for_po(
    conn, po: dict[str, Any], *, outlet: str
) -> tuple[dict[str, Any] | None, list[dict[str, Any]]]:
    """Resolve indent + remaining inward rows for one generated purchase order."""
    try:
        indent_id = int(po.get("indent_id") or 0)
        supplier_id = int(po.get("supplier_id") or 0)
        po_id = int(po.get("po_id") or po.get("id") or 0)
    except (TypeError, ValueError):
        return None, []
    if indent_id <= 0:
        return None, []

    indent = conn.execute(
        """
        SELECT i.*, u.full_name AS created_by_name,
               d.full_name AS decided_by_name,
               d.username AS decided_by_username
        FROM store_indents i
        LEFT JOIN users u ON u.id = i.created_by
        LEFT JOIN users d ON d.id = i.decided_by
        WHERE i.id = ? AND i.status = 'approved'
        """,
        (indent_id,),
    ).fetchone()
    if not indent:
        return None, []

    po_row = {
        "id": po_id,
        "indent_id": indent_id,
        "supplier_id": supplier_id,
        "po_no": po.get("po_no") or "",
    }
    po_lines = _po_lines_for_send_or_pdf(
        conn,
        indent_id=indent_id,
        supplier_id=supplier_id,
        po_row=po_row if po_id else None,
        group_lines=[],
    )
    if not po_lines:
        selected = dict(indent)
        selected["outlet"] = _parse_outlet(selected.get("outlet"))
        selected["outlet_label"] = _outlet_label(selected["outlet"])
        selected["po_id"] = po_id
        selected["po_no"] = po.get("po_no") or ""
        selected["supplier_name"] = po.get("supplier_name") or ""
        return selected, []

    indent_line_ids: list[int] = []
    for row in po_lines:
        try:
            lid = int(row.get("line_id") or 0)
        except (TypeError, ValueError):
            lid = 0
        if lid > 0:
            indent_line_ids.append(lid)
    indent_lines_by_id: dict[int, dict[str, Any]] = {}
    if indent_line_ids:
        placeholders = ",".join("?" for _ in indent_line_ids)
        for row in conn.execute(
            f"""
            SELECT id, item_name, quantity, quantity_received, unit, notes, approximate_price,
                   pack_label, pack_qty_in_base
            FROM store_indent_lines
            WHERE indent_id = ? AND id IN ({placeholders})
            """,
            (indent_id, *indent_line_ids),
        ).fetchall():
            indent_lines_by_id[int(row["id"])] = dict(row)

    product_cat_map = _product_category_by_item_name(
        conn, stores_outlet=indent["outlet"] or outlet
    )
    selected_lines: list[dict[str, Any]] = []
    for po_line in po_lines:
        try:
            line_id = int(po_line.get("line_id") or 0)
        except (TypeError, ValueError):
            line_id = 0
        indent_line = indent_lines_by_id.get(line_id) if line_id else None
        if not indent_line:
            continue
        try:
            po_qty = float(po_line.get("quantity") or 0)
        except (TypeError, ValueError):
            po_qty = 0.0
        try:
            po_received = float(po_line.get("quantity_received") or 0)
        except (TypeError, ValueError):
            po_received = 0.0
        try:
            qty_val = float(indent_line.get("quantity") or 0)
        except (TypeError, ValueError):
            qty_val = 0.0
        try:
            received_val = float(indent_line.get("quantity_received") or 0)
        except (TypeError, ValueError):
            received_val = 0.0
        remaining_val = qty_val - received_val
        if remaining_val <= 0.0001:
            continue
        # Cap available inward to what THIS PO still has left to receive
        # (not just indent remaining — a partial PO can be fully done while
        # the indent still has qty for other POs).
        available = _po_inward_available_qty(
            po_qty=po_qty,
            po_received=po_received,
            indent_remaining=remaining_val,
        )
        if available <= 0.0001:
            continue
        approx = indent_line.get("approximate_price")
        if po_line.get("rate") not in (None, ""):
            approx = po_line.get("rate")
        pack_label = (indent_line.get("pack_label") or po_line.get("pack_label") or "").strip()
        pack_qty = _row_pack_qty_in_base(indent_line)
        if pack_qty is None:
            pack_qty = po_line.get("pack_qty_in_base")
        base_unit = indent_line.get("unit") or po_line.get("unit") or ""
        raw_item_name = (indent_line.get("item_name") or po_line.get("item_name") or "").strip()
        display_name = (
            (po_line.get("display_name") or "").strip()
            or _format_indent_line_item(indent_line)
            or raw_item_name
        )
        notes = indent_line.get("notes") or ""
        line_pk = int(indent_line["id"])
        try:
            rate_val = float(approx) if approx is not None and approx != "" else 0.0
        except (TypeError, ValueError):
            rate_val = 0.0
        if pack_label and pack_qty is not None:
            try:
                display_unit = f"{_format_ledger_qty(float(pack_qty))} {base_unit}".strip()
            except (TypeError, ValueError):
                display_unit = base_unit
        else:
            display_unit = base_unit
        product_category = product_cat_map.get(raw_item_name.casefold(), "")
        selected_lines.append(
            {
                "id": line_pk,
                "item_name": display_name,
                "product_category": product_category,
                "quantity": qty_val,
                "quantity_display": _format_qty_display(qty_val),
                "quantity_received": received_val,
                "quantity_received_display": _format_qty_display(received_val),
                "remaining": available,
                "remaining_display": _format_qty_display(available),
                "unit": display_unit,
                "notes": notes,
                "approximate_price": approx,
                "approximate_price_display": _format_optional_price(approx),
                "rate_value": rate_val,
                "initial": (raw_item_name or "?")[:1].upper(),
                "pack_label": pack_label,
                "pack_qty_in_base": pack_qty,
            }
        )

    selected = dict(indent)
    selected["outlet"] = _parse_outlet(selected.get("outlet"))
    selected["outlet_label"] = _outlet_label(selected["outlet"])
    selected["po_id"] = po_id
    selected["po_no"] = po.get("po_no") or ""
    selected["supplier_name"] = po.get("supplier_name") or ""
    return selected, selected_lines


def _po_line_remaining_qty(indent_qty: float, ordered_qty: float) -> float:
    try:
        remaining = float(indent_qty or 0) - float(ordered_qty or 0)
    except (TypeError, ValueError):
        remaining = float(indent_qty or 0)
    return remaining if remaining > 0.0001 else 0.0


def _commit_po_group_quantities(conn, indent_id: int, lines: list[dict[str, Any]]) -> None:
    """Mark generated line quantities as ordered and leave remaining on the draft."""
    for line in lines or []:
        try:
            line_id = int(line.get("line_id"))
        except (TypeError, ValueError):
            continue
        try:
            qty = float(line.get("quantity") or 0)
        except (TypeError, ValueError):
            qty = 0.0
        if line_id <= 0 or qty <= 0.0001:
            continue
        conn.execute(
            """
            UPDATE store_indent_lines
            SET quantity_ordered = MIN(
                COALESCE(quantity, 0),
                COALESCE(quantity_ordered, 0) + ?
            )
            WHERE id = ? AND indent_id = ?
            """,
            (qty, line_id, indent_id),
        )
        row = conn.execute(
            """
            SELECT quantity, COALESCE(quantity_ordered, 0) AS quantity_ordered
            FROM store_indent_lines
            WHERE id = ? AND indent_id = ?
            """,
            (line_id, indent_id),
        ).fetchone()
        if not row:
            continue
        remaining = _po_line_remaining_qty(row["quantity"], row["quantity_ordered"])
        if remaining > 0.0001:
            conn.execute(
                """
                UPDATE store_po_lines
                SET quantity = ?, updated_at = datetime('now','localtime')
                WHERE line_id = ? AND indent_id = ?
                """,
                (remaining, line_id, indent_id),
            )
        else:
            conn.execute(
                """
                UPDATE store_po_lines
                SET quantity = NULL, updated_at = datetime('now','localtime')
                WHERE line_id = ? AND indent_id = ?
                """,
                (line_id, indent_id),
            )
    conn.commit()


def _po_default_message(
    supplier_name: str,
    lines: list[dict[str, Any]],
    indent_no: str,
    po_no: str = "",
) -> str:
    name = (supplier_name or "Supplier").strip() or "Supplier"
    bullets = []
    for line in lines:
        qty = line.get("quantity") or 0
        try:
            qty_num = float(qty)
            qty_label = str(int(qty_num)) if abs(qty_num - round(qty_num)) < 0.0001 else f"{qty_num:g}"
        except (TypeError, ValueError):
            qty_label = str(qty)
        unit = _po_line_display_unit(line)
        item = line.get("display_name") or line.get("item_name") or "Item"
        bullets.append(f"• {item} — {qty_label} {unit}".rstrip())
    body = "\n".join(bullets) if bullets else "• (no items)"
    reference = (po_no or "").strip() or (indent_no or "").strip()
    return (
        f"Hello {name},\n"
        f"Please find our purchase order {reference} as below.\n\n"
        f"{body}\n\n"
        f"Please confirm:\n"
        f"✓ Availability\n"
        f"✓ Price\n"
        f"✓ Expected delivery date\n\n"
        f"Thank you,\n"
        f"— Hotel Bell Elite\n"
        f"({reference})"
    )


def _po_is_outside_session_error(err: str) -> bool:
    text = str(err or "").lower()
    return (
        "131047" in text
        or ("outside" in text and "window" in text)
        or "24 hour" in text
        or "24-hour" in text
    )


def _load_pending_inward_indents(conn, outlet: str) -> list[dict[str, Any]]:
    """Approved indents that still have quantity left to stock inward."""
    outlet_sql, outlet_params = _outlet_match_sql("i.outlet", outlet)
    rows = conn.execute(
        f"""
        SELECT i.id, i.indent_no, i.outlet, i.decided_at, i.created_at,
               (SELECT COUNT(*) FROM store_indent_lines l WHERE l.indent_id = i.id) AS line_count
        FROM store_indents i
        WHERE {outlet_sql} AND i.status = 'approved'
          AND EXISTS (
            SELECT 1 FROM store_indent_lines l
            WHERE l.indent_id = i.id
              AND COALESCE(l.quantity, 0) - COALESCE(l.quantity_received, 0) > 0.0001
          )
        ORDER BY i.decided_at DESC, i.id DESC
        """,
        outlet_params,
    ).fetchall()
    return [dict(row) for row in rows]


def _indent_needs_po_generation(conn, indent_id: int) -> bool:
    """True when the indent still has remaining qty to put on a purchase order."""
    payload = _load_po_supplier_groups(conn, indent_id)
    if not payload or payload.get("status") != "approved":
        return False
    return bool(_pending_po_groups(payload.get("groups") or []))


def _load_pending_po_indents(conn, outlet: str) -> list[dict[str, Any]]:
    """Approved indents that still need purchase orders generated.

    Indents where every line is fully covered by generated PO quantities are
    omitted from the Generate PO indent picker.
    """
    outlet_sql, outlet_params = _outlet_match_sql("i.outlet", outlet)
    rows = conn.execute(
        f"""
        SELECT i.id, i.indent_no, i.outlet, i.decided_at, i.created_at,
               (SELECT COUNT(*) FROM store_indent_lines l WHERE l.indent_id = i.id) AS line_count
        FROM store_indents i
        WHERE {outlet_sql} AND i.status = 'approved'
          AND EXISTS (
            SELECT 1 FROM store_indent_lines l WHERE l.indent_id = i.id
          )
        ORDER BY i.decided_at DESC, i.id DESC
        """,
        outlet_params,
    ).fetchall()
    out: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        if _indent_needs_po_generation(conn, int(item["id"])):
            out.append(item)
    return out


def _default_pack_from_product_variants(
    variants: list[dict[str, Any]] | None,
) -> tuple[str, float | None]:
    """First active Product Master pack label + qty_in_base, or empty."""
    for variant in variants or []:
        label = str((variant or {}).get("label") or "").strip()
        if not label:
            continue
        qty = (variant or {}).get("qty_in_base")
        try:
            qty_num = float(qty) if qty not in (None, "") else None
        except (TypeError, ValueError):
            qty_num = None
        if qty_num is not None and qty_num <= 0:
            qty_num = None
        return label, qty_num
    return "", None


def _load_po_supplier_groups(conn, indent_id: int) -> dict[str, Any]:
    """Group approved indent lines by preferred / overridden supplier."""
    indent = conn.execute(
        """
        SELECT i.*, u.full_name AS created_by_name, d.full_name AS decided_by_name
        FROM store_indents i
        LEFT JOIN users u ON u.id = i.created_by
        LEFT JOIN users d ON d.id = i.decided_by
        WHERE i.id = ?
        """,
        (indent_id,),
    ).fetchone()
    if not indent:
        return {}

    lines = [
        dict(row)
        for row in conn.execute(
            """
            SELECT id, item_name, quantity, unit, notes, approximate_price,
                   pack_label, pack_qty_in_base, COALESCE(quantity_ordered, 0) AS quantity_ordered
            FROM store_indent_lines
            WHERE indent_id = ?
            ORDER BY id
            """,
            (indent_id,),
        ).fetchall()
    ]
    overrides = {
        int(row["line_id"]): dict(row)
        for row in conn.execute(
            """
            SELECT line_id, supplier_id, rate, quantity
            FROM store_po_lines
            WHERE indent_id = ?
            """,
            (indent_id,),
        ).fetchall()
    }
    # Inactive products still carry the supplier mapping for already-approved indents.
    product_rows = conn.execute(
        """
        SELECT id, name, preferred_supplier_1_id, preferred_supplier_2_id, preferred_supplier_3_id
        FROM store_products
        ORDER BY is_active DESC, id
        """
    ).fetchall()
    product_by_name: dict[str, dict[str, Any]] = {}
    for row in product_rows:
        key = str(row["name"] or "").strip().lower()
        if key:
            product_by_name.setdefault(key, dict(row))
    variants_by_product = _load_variants_by_product_ids(
        conn, [int(row["id"]) for row in product_by_name.values()]
    )
    suppliers = {
        int(row["id"]): dict(row)
        for row in conn.execute(
            """
            SELECT id, name, gst, address, phone
            FROM suppliers
            ORDER BY LOWER(name), id
            """
        ).fetchall()
    }

    groups_map: dict[int | None, dict[str, Any]] = {}
    healed_pack = False
    for line in lines:
        line_id = int(line["id"])
        override = overrides.get(line_id) or {}
        product = product_by_name.get(str(line.get("item_name") or "").strip().lower())
        supplier_id = override.get("supplier_id")
        if supplier_id is not None:
            try:
                supplier_id = int(supplier_id) if supplier_id else None
            except (TypeError, ValueError):
                supplier_id = None
        if not supplier_id and product:
            for key in (
                "preferred_supplier_1_id",
                "preferred_supplier_2_id",
                "preferred_supplier_3_id",
            ):
                raw = product.get(key)
                if raw:
                    try:
                        supplier_id = int(raw)
                        break
                    except (TypeError, ValueError):
                        continue
        if supplier_id and supplier_id not in suppliers:
            supplier_id = None

        rate = override.get("rate")
        if rate is None or rate == "":
            rate = line.get("approximate_price")
        try:
            rate_num = float(rate) if rate is not None and rate != "" else None
        except (TypeError, ValueError):
            rate_num = None
        try:
            indent_qty = float(line.get("quantity") or 0)
        except (TypeError, ValueError):
            indent_qty = 0.0
        try:
            ordered_qty = float(line.get("quantity_ordered") or 0)
        except (TypeError, ValueError):
            ordered_qty = 0.0
        remaining_qty = _po_line_remaining_qty(indent_qty, ordered_qty)
        if remaining_qty <= 0.0001:
            continue
        qty = remaining_qty
        override_qty = override.get("quantity")
        if override_qty not in (None, ""):
            try:
                qty = float(override_qty)
            except (TypeError, ValueError):
                qty = remaining_qty
        if qty <= 0:
            qty = remaining_qty
        if remaining_qty > 0 and qty > remaining_qty:
            qty = remaining_qty
        amount = round(qty * rate_num, 2) if rate_num is not None else None
        display_name = _format_indent_line_item(line)
        line_for_display = dict(line)
        line_for_display["quantity"] = qty
        pack_label = (line.get("pack_label") or "").strip()
        pack_qty = line.get("pack_qty_in_base")
        if not pack_label and product:
            pack_label, pack_qty = _default_pack_from_product_variants(
                variants_by_product.get(int(product["id"]))
            )
            if pack_label:
                # Persist so PO PDF / inward / later edits keep Product Master pack.
                conn.execute(
                    """
                    UPDATE store_indent_lines
                    SET pack_label = ?, pack_qty_in_base = ?
                    WHERE id = ?
                      AND (pack_label IS NULL OR TRIM(COALESCE(pack_label, '')) = '')
                    """,
                    (pack_label, pack_qty, line_id),
                )
                line["pack_label"] = pack_label
                line["pack_qty_in_base"] = pack_qty
                line_for_display["pack_label"] = pack_label
                line_for_display["pack_qty_in_base"] = pack_qty
                display_name = _format_indent_line_item(line)
                healed_pack = True
        item = {
            "line_id": line_id,
            "item_name": line.get("item_name") or "",
            "display_name": display_name,
            "quantity": qty,
            "quantity_display": _format_ledger_qty(qty),
            "indent_quantity": indent_qty,
            "indent_quantity_display": _format_ledger_qty(indent_qty),
            "remaining_quantity": remaining_qty,
            "remaining_quantity_display": _format_ledger_qty(remaining_qty),
            "ordered_quantity": ordered_qty,
            "quantity_is_partial": remaining_qty > 0 and qty + 1e-9 < remaining_qty,
            "unit": line.get("unit") or "",
            "display_unit": _po_line_display_unit(line_for_display),
            "pack_label": pack_label,
            "pack_qty_in_base": pack_qty,
            "total_display": _po_line_total_display(line_for_display),
            "approximate_price": line.get("approximate_price"),
            "rate": rate_num,
            "amount": amount,
            "supplier_id": supplier_id,
            "notes": line.get("notes") or "",
        }
        if supplier_id not in groups_map:
            supplier = suppliers.get(supplier_id) if supplier_id else None
            groups_map[supplier_id] = {
                "supplier_id": supplier_id,
                "supplier_name": (supplier or {}).get("name") or "Unassigned",
                "phone": (supplier or {}).get("phone") or "",
                "gst": (supplier or {}).get("gst") or "",
                "address": (supplier or {}).get("address") or "",
                "initials": _supplier_initials((supplier or {}).get("name") or "Unassigned"),
                "is_unassigned": supplier_id is None,
                "lines": [],
                "item_count": 0,
                "estimated_value": 0.0,
                "can_send": False,
            }
        groups_map[supplier_id]["lines"].append(item)
        groups_map[supplier_id]["item_count"] += 1
        if amount is not None:
            groups_map[supplier_id]["estimated_value"] = round(
                groups_map[supplier_id]["estimated_value"] + amount, 2
            )

    groups = []
    unassigned = groups_map.pop(None, None)
    if unassigned:
        unassigned["can_send"] = False
        unassigned["po_no"] = ""
        groups.append(unassigned)
    for supplier_id in sorted(
        groups_map.keys(),
        key=lambda sid: (str(groups_map[sid]["supplier_name"] or "").lower(), sid or 0),
    ):
        group = groups_map[supplier_id]
        group["can_send"] = bool(group["supplier_id"] and (group["phone"] or "").strip())
        # Remaining lines still need a (new) PO — don't stamp the previous batch number.
        group["po_no"] = ""
        groups.append(group)

    grouped_counts = {
        int(g["supplier_id"]): g["item_count"] for g in groups if g.get("supplier_id")
    }
    supplier_options = [
        {
            "id": int(row["id"]),
            "name": row["name"] or "",
            "phone": row.get("phone") or "",
            "gst": row.get("gst") or "",
            "address": row.get("address") or "",
            "initials": _supplier_initials(row["name"] or ""),
            "item_count": grouped_counts.get(int(row["id"]), 0),
        }
        for row in sorted(suppliers.values(), key=lambda s: (str(s.get("name") or "").lower(), s["id"]))
    ]

    indent_data = dict(indent)
    if healed_pack:
        conn.commit()
    return {
        "indent": indent_data,
        "indent_id": int(indent_data["id"]),
        "indent_no": indent_data.get("indent_no") or "",
        "outlet": indent_data.get("outlet") or "",
        "outlet_label": _outlet_label(_parse_outlet(indent_data.get("outlet"))),
        "status": indent_data.get("status") or "",
        "groups": groups,
        "supplier_options": supplier_options,
        "supplier_count": sum(1 for g in groups if not g.get("is_unassigned")),
    }


def _pending_po_groups(groups: list | None) -> list[dict[str, Any]]:
    """Supplier groups that still need a PO (any lines with remaining indent qty)."""
    return [group for group in (groups or []) if group.get("lines")]


def _record_po_send(
    conn,
    *,
    indent_id: int,
    supplier_id: int | None,
    phone: str,
    message: str,
    pdf_name: str,
    include_pdf: bool,
    conversation_id: int | None,
    wa_message_id: str,
    status: str,
    error: str,
    sent_by: int | None,
    po_no: str = "",
) -> int:
    cur = conn.execute(
        """
        INSERT INTO store_po_sends (
            indent_id, supplier_id, po_no, phone, message, pdf_name, include_pdf,
            conversation_id, wa_message_id, status, error, sent_by, sent_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now','localtime'))
        """,
        (
            indent_id,
            supplier_id,
            po_no or "",
            phone or "",
            message or "",
            pdf_name or "",
            1 if include_pdf else 0,
            conversation_id,
            wa_message_id or "",
            status or "failed",
            error or "",
            sent_by,
        ),
    )
    return int(cur.lastrowid)


def _load_po_send_history(conn, *, limit: int = 100) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT s.*,
               i.indent_no,
               i.outlet,
               sp.name AS supplier_name,
               u.full_name AS sent_by_name,
               COALESCE(NULLIF(s.po_no, ''), po.po_no, '') AS po_no_display
        FROM store_po_sends s
        LEFT JOIN store_indents i ON i.id = s.indent_id
        LEFT JOIN suppliers sp ON sp.id = s.supplier_id
        LEFT JOIN users u ON u.id = s.sent_by
        LEFT JOIN store_purchase_orders po
               ON po.indent_id = s.indent_id AND po.supplier_id = s.supplier_id
        ORDER BY s.sent_at DESC, s.id DESC
        LIMIT ?
        """,
        (max(1, int(limit or 100)),),
    ).fetchall()
    out = []
    for row in rows:
        item = dict(row)
        item["outlet_label"] = _outlet_label(_parse_outlet(item.get("outlet")))
        item["sent_at_display"] = _format_stores_dt(item.get("sent_at"))
        item["supplier_name"] = item.get("supplier_name") or "—"
        item["po_no"] = item.get("po_no_display") or item.get("po_no") or ""
        out.append(item)
    return out


def _load_generated_purchase_orders(
    conn, outlet: str, *, limit: int = 200, send_queue: bool = False, pending_inward: bool = False
) -> list[dict[str, Any]]:
    """Issued PO numbers, newest first.

    When ``send_queue`` is True (Send to Supplier tab), only return POs that are
    not successfully sent, still have stock-inward remaining on the indent, and
    have sendable lines (frozen PO lines or a current supplier assignment).

    When ``pending_inward`` is True (Stock Inward PO picker), only return POs
    that still have at least one line with remaining qty to receive.
    """
    outlet_sql, outlet_params = _outlet_match_sql("i.outlet", outlet)
    # Prefer matching by po_no; fall back to indent×supplier for legacy send rows.
    send_match = """
               (
                   SELECT s.status
                   FROM store_po_sends s
                   WHERE s.indent_id = po.indent_id
                     AND s.supplier_id = po.supplier_id
                     AND (
                       s.po_no = po.po_no
                       OR (
                         COALESCE(TRIM(s.po_no), '') = ''
                         AND NOT EXISTS (
                           SELECT 1 FROM store_po_sends sx
                           WHERE sx.indent_id = po.indent_id
                             AND sx.supplier_id = po.supplier_id
                             AND sx.po_no = po.po_no
                         )
                       )
                     )
                   ORDER BY
                     CASE WHEN s.po_no = po.po_no THEN 0 ELSE 1 END,
                     s.sent_at DESC, s.id DESC
                   LIMIT 1
               ) AS last_send_status,
               (
                   SELECT s.sent_at
                   FROM store_po_sends s
                   WHERE s.indent_id = po.indent_id
                     AND s.supplier_id = po.supplier_id
                     AND (
                       s.po_no = po.po_no
                       OR (
                         COALESCE(TRIM(s.po_no), '') = ''
                         AND NOT EXISTS (
                           SELECT 1 FROM store_po_sends sx
                           WHERE sx.indent_id = po.indent_id
                             AND sx.supplier_id = po.supplier_id
                             AND sx.po_no = po.po_no
                         )
                       )
                     )
                   ORDER BY
                     CASE WHEN s.po_no = po.po_no THEN 0 ELSE 1 END,
                     s.sent_at DESC, s.id DESC
                   LIMIT 1
               ) AS last_sent_at
    """
    queue_sql = ""
    if send_queue:
        queue_sql = """
          AND i.status = 'approved'
          AND EXISTS (
            SELECT 1 FROM store_indent_lines l
            WHERE l.indent_id = i.id
              AND COALESCE(l.quantity, 0) - COALESCE(l.quantity_received, 0) > 0.0001
          )
          AND COALESCE((
                   SELECT s.status
                   FROM store_po_sends s
                   WHERE s.indent_id = po.indent_id
                     AND s.supplier_id = po.supplier_id
                     AND (
                       s.po_no = po.po_no
                       OR (
                         COALESCE(TRIM(s.po_no), '') = ''
                         AND NOT EXISTS (
                           SELECT 1 FROM store_po_sends sx
                           WHERE sx.indent_id = po.indent_id
                             AND sx.supplier_id = po.supplier_id
                             AND sx.po_no = po.po_no
                         )
                       )
                     )
                   ORDER BY
                     CASE WHEN s.po_no = po.po_no THEN 0 ELSE 1 END,
                     s.sent_at DESC, s.id DESC
                   LIMIT 1
               ), '') != 'sent'
          AND (
            EXISTS (
              SELECT 1 FROM store_purchase_order_lines pol
              WHERE pol.purchase_order_id = po.id
            )
            OR EXISTS (
              SELECT 1 FROM store_po_lines pl
              WHERE pl.indent_id = po.indent_id
                AND pl.supplier_id = po.supplier_id
            )
          )
        """
    elif pending_inward:
        # Match _build_inward_lines_for_po: remaining on this PO's lines only
        # (PO qty − PO received), also requiring indent remaining.
        queue_sql = """
          AND i.status = 'approved'
          AND (
            EXISTS (
              SELECT 1
              FROM store_purchase_order_lines pol
              JOIN store_indent_lines l
                ON l.id = pol.line_id AND l.indent_id = po.indent_id
              WHERE pol.purchase_order_id = po.id
                AND COALESCE(pol.quantity, 0) - COALESCE(pol.quantity_received, 0) > 0.0001
                AND COALESCE(l.quantity, 0) - COALESCE(l.quantity_received, 0) > 0.0001
            )
            OR (
              NOT EXISTS (
                SELECT 1 FROM store_purchase_order_lines pol
                WHERE pol.purchase_order_id = po.id
              )
              AND EXISTS (
                SELECT 1
                FROM store_po_lines pl
                JOIN store_indent_lines l
                  ON l.id = pl.line_id AND l.indent_id = po.indent_id
                WHERE pl.indent_id = po.indent_id
                  AND pl.supplier_id = po.supplier_id
                  AND COALESCE(l.quantity, 0) - COALESCE(l.quantity_received, 0) > 0.0001
              )
            )
          )
        """
    rows = conn.execute(
        f"""
        SELECT po.id AS po_id,
               po.po_no,
               po.indent_id,
               po.supplier_id,
               po.created_at,
               i.indent_no,
               i.outlet,
               i.status AS indent_status,
               sp.name AS supplier_name,
               sp.phone AS supplier_phone,
               {send_match}
        FROM store_purchase_orders po
        JOIN store_indents i ON i.id = po.indent_id
        LEFT JOIN suppliers sp ON sp.id = po.supplier_id
        WHERE {outlet_sql}
          {queue_sql}
        ORDER BY po.created_at DESC, po.id DESC
        LIMIT ?
        """,
        (*outlet_params, max(1, int(limit or 200))),
    ).fetchall()
    out: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        item["outlet_label"] = _outlet_label(_parse_outlet(item.get("outlet")))
        item["supplier_name"] = item.get("supplier_name") or "—"
        item["supplier_phone"] = item.get("supplier_phone") or ""
        item["created_at_display"] = _format_stores_dt(item.get("created_at"))
        item["last_sent_at_display"] = _format_stores_dt(item.get("last_sent_at"))
        status = str(item.get("last_send_status") or "").strip().lower()
        if status == "sent":
            item["status"] = "sent"
            item["status_label"] = "Sent"
        elif status == "failed":
            item["status"] = "failed"
            item["status_label"] = "Failed"
        elif status:
            item["status"] = status
            item["status_label"] = status.replace("_", " ").title()
        else:
            item["status"] = "created"
            item["status_label"] = "Created"
        out.append(item)
    return out


@stores_bp.route("/stores/orders", endpoint="stores_orders")
def stores_orders():
    """Purchase Order page — register of generated POs (PO/RES|BAR/…).

    Path is /stores/orders (not …/purchase-order) so soft-nav does not treat it
    as an Excel download. ``?tab=send`` opens the Send to Supplier view of the
    same register.
    """
    _get_user()
    outlet = _parse_outlet_filter(request.args.get("outlet"))
    list_view = "approved"
    po_tab = "send" if (request.args.get("tab") or "").strip().lower() == "send" else "orders"

    conn = get_db()
    try:
        ensure_stores_schema(conn)
        indents, indent_view_data, stores_ledger_data = _load_indent_list_for_view(
            conn, outlet, list_view
        )
        purchase_orders = _load_generated_purchase_orders(
            conn, outlet, send_queue=(po_tab == "send")
        )
        pending_inward_indents = _load_pending_po_indents(conn, outlet)
        catalog = _load_product_catalog(conn, stores_outlet=outlet) if outlet != "both" else []
    finally:
        conn.close()

    empty_form = {
        "indent_id": "",
        "notes": "",
        "submission_token": "",
        "lines": [{
            "item_name": "",
            "quantity": "",
            "unit": "kg",
            "notes": "",
            "approximate_price": "",
            "pack_label": "",
            "pack_qty_in_base": "",
        }],
    }

    return _page_render(
        "purchase_orders",
        outlet=outlet,
        indents=indents,
        indent_view_data=indent_view_data,
        stores_ledger_data=stores_ledger_data,
        product_catalog=catalog,
        show_form=False,
        open_edit_id=0,
        indent_form_unset=False,
        form=empty_form,
        errors=[],
        editing=False,
        indent_list_views=(),
        selected_indent_view=list_view,
        de_nav_stores_view="purchase_order",
        po_tab=po_tab,
        po_send_data=None,
        po_history=[],
        purchase_orders=purchase_orders,
        pending_inward_indents=pending_inward_indents,
    )


@stores_bp.route("/stores/orders/history", endpoint="stores_orders_history")
def stores_orders_history():
    """Purchase Order history of WhatsApp sends."""
    _get_user()
    outlet = _parse_outlet_filter(request.args.get("outlet"))
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        history = _load_po_send_history(conn)
        indents, indent_view_data, stores_ledger_data = _load_indent_list_for_view(
            conn, outlet, "approved"
        )
        pending_inward_indents = _load_pending_po_indents(conn, outlet)
    finally:
        conn.close()

    return _page_render(
        "purchase_orders",
        outlet=outlet,
        indents=indents,
        indent_view_data=indent_view_data,
        stores_ledger_data=stores_ledger_data,
        product_catalog=[],
        show_form=False,
        open_edit_id=0,
        indent_form_unset=False,
        form={
            "indent_id": "",
            "notes": "",
            "submission_token": "",
            "lines": [],
        },
        errors=[],
        editing=False,
        indent_list_views=(),
        selected_indent_view="approved",
        de_nav_stores_view="purchase_order",
        po_tab="history",
        po_send_data=None,
        po_history=history,
        pending_inward_indents=pending_inward_indents,
    )


@stores_bp.route("/stores/orders/<int:indent_id>", endpoint="stores_orders_send")
def stores_orders_send(indent_id: int):
    """Send-to-supplier workspace for one approved indent.

    Reviews grouped lines and generates/sends purchase orders. The old compose
    message-preview step is retired — WhatsApp goes out on Generate.
    """
    _get_user()
    outlet = _parse_outlet_filter(request.args.get("outlet"))
    # Legacy ?step=compose bookmarks / soft-nav must never show the message preview.
    if request.args.get("step") == "compose":
        return redirect(
            url_for(
                "stores_orders_send",
                indent_id=indent_id,
                outlet=outlet if outlet and outlet != "both" else None,
            )
        )
    po_step = "items"
    po_selected_supplier_id = None
    po_selected_group = None
    po_message = ""
    po_pdf_name = "PO.pdf"
    po_no = ""
    po_selected_supplier_ids = _parse_selected_supplier_ids(request.args.get("suppliers"))
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        payload = _load_po_supplier_groups(conn, indent_id)
        if not payload:
            flash("Indent not found.", "error")
            return redirect(url_for("stores_orders", outlet=outlet))
        if payload.get("status") != "approved":
            flash("Purchase orders can only be sent for approved indents.", "error")
            return redirect(url_for("stores_orders", outlet=outlet or payload.get("outlet")))
        if outlet == "both" or not outlet:
            outlet = _parse_outlet(payload.get("outlet"))
        # Hide supplier groups that already have a generated PO number.
        pending = _pending_po_groups(payload.get("groups") or [])
        payload = dict(payload)
        payload["groups"] = pending
        payload["supplier_count"] = sum(
            1 for g in pending if not g.get("is_unassigned")
        )
        indents, indent_view_data, stores_ledger_data = _load_indent_list_for_view(
            conn, outlet, "approved"
        )
        pending_inward_indents = _load_pending_po_indents(conn, outlet)
        # Fully generated indents leave the Generate PO picker — bounce to the next
        # pending indent or the Purchase Orders register.
        if not pending:
            if pending_inward_indents:
                nxt = pending_inward_indents[0]
                return redirect(
                    url_for(
                        "stores_orders_send",
                        indent_id=int(nxt["id"]),
                        outlet=outlet,
                    )
                )
            flash("All purchase orders for this indent are already generated.", "ok")
            return redirect(url_for("stores_orders", outlet=outlet))
    finally:
        conn.close()

    return _page_render(
        "purchase_orders",
        outlet=outlet,
        indents=indents,
        indent_view_data=indent_view_data,
        stores_ledger_data=stores_ledger_data,
        product_catalog=[],
        show_form=False,
        open_edit_id=0,
        indent_form_unset=False,
        form={
            "indent_id": "",
            "notes": "",
            "submission_token": "",
            "lines": [],
        },
        errors=[],
        editing=False,
        indent_list_views=(),
        selected_indent_view="approved",
        de_nav_stores_view="purchase_order",
        po_tab="generate",
        po_step=po_step,
        po_send_data=payload,
        po_history=[],
        pending_inward_indents=pending_inward_indents,
        po_selected_supplier_id=po_selected_supplier_id,
        po_selected_group=po_selected_group,
        po_selected_supplier_ids=po_selected_supplier_ids,
        po_message=po_message,
        po_pdf_name=po_pdf_name,
        po_no=po_no,
    )



def _save_po_line_overrides(conn, indent_id: int, rows: list) -> int:
    """Persist supplier / rate / qty overrides for PO lines. Returns the row count saved."""
    indent_lines = {
        int(row["id"]): {
            "quantity": float(row["quantity"] or 0),
            "ordered": float(row["quantity_ordered"] or 0),
        }
        for row in conn.execute(
            """
            SELECT id, quantity, COALESCE(quantity_ordered, 0) AS quantity_ordered
            FROM store_indent_lines
            WHERE indent_id = ?
            """,
            (indent_id,),
        ).fetchall()
    }
    valid_line_ids = set(indent_lines.keys())
    supplier_ids = {
        int(row["id"]) for row in conn.execute("SELECT id FROM suppliers").fetchall()
    }

    saved = 0
    for row in rows:
        if not isinstance(row, dict):
            continue
        try:
            line_id = int(row.get("line_id"))
        except (TypeError, ValueError):
            continue
        if line_id not in valid_line_ids:
            continue

        supplier_raw = row.get("supplier_id")
        supplier_id = None
        if supplier_raw not in (None, "", 0, "0"):
            try:
                supplier_id = int(supplier_raw)
            except (TypeError, ValueError):
                supplier_id = None
            if supplier_id not in supplier_ids:
                supplier_id = None

        rate_raw = row.get("rate")
        rate = None
        if rate_raw not in (None, ""):
            try:
                rate = float(rate_raw)
                if rate < 0:
                    rate = None
            except (TypeError, ValueError):
                rate = None

        quantity = None
        if "quantity" in row:
            qty_raw = row.get("quantity")
            if qty_raw not in (None, ""):
                try:
                    quantity = float(qty_raw)
                except (TypeError, ValueError):
                    quantity = None
                if quantity is not None:
                    if quantity <= 0:
                        quantity = None
                    else:
                        meta = indent_lines.get(line_id) or {}
                        max_qty = _po_line_remaining_qty(
                            meta.get("quantity") or 0, meta.get("ordered") or 0
                        )
                        if max_qty > 0 and quantity > max_qty:
                            quantity = max_qty

        conn.execute(
            """
            INSERT INTO store_po_lines (indent_id, line_id, supplier_id, rate, quantity, updated_at)
            VALUES (?, ?, ?, ?, ?, datetime('now','localtime'))
            ON CONFLICT(line_id) DO UPDATE SET
                supplier_id = excluded.supplier_id,
                rate = excluded.rate,
                quantity = excluded.quantity,
                updated_at = datetime('now','localtime')
            """,
            (indent_id, line_id, supplier_id, rate, quantity),
        )
        saved += 1
    conn.commit()
    return saved


@stores_bp.route("/stores/orders/<int:indent_id>/lines", methods=["POST"], endpoint="stores_orders_lines")
def stores_orders_lines(indent_id: int):
    """Save supplier / rate overrides for PO lines (JSON, used by Edit items)."""
    _get_user()
    payload = request.get_json(silent=True) or {}
    rows = payload.get("lines") if isinstance(payload, dict) else None
    if not isinstance(rows, list):
        return jsonify({"ok": False, "error": "Expected lines array."}), 400

    conn = get_db()
    try:
        ensure_stores_schema(conn)
        indent = conn.execute(
            "SELECT id, status FROM store_indents WHERE id = ?",
            (indent_id,),
        ).fetchone()
        if not indent:
            return jsonify({"ok": False, "error": "Indent not found."}), 404
        if indent["status"] != "approved":
            return jsonify({"ok": False, "error": "Only approved indents can be edited for PO."}), 400

        saved = _save_po_line_overrides(conn, indent_id, rows)
        groups = _load_po_supplier_groups(conn, indent_id)
    finally:
        conn.close()

    return jsonify({"ok": True, "saved": saved, "groups": groups.get("groups") if groups else []})


@stores_bp.route(
    "/stores/orders/<int:indent_id>/lines/next",
    methods=["POST"],
    endpoint="stores_orders_lines_next",
)
def stores_orders_lines_next(indent_id: int):
    """Save line suppliers and generate PO numbers (WhatsApp send is user-controlled)."""
    _get_user()
    outlet = _parse_outlet_filter(request.form.get("outlet"))
    issued: list[dict[str, Any]] = []
    selected_ids: list[int] = []
    wants_json = (
        request.headers.get("X-Requested-With") == "XMLHttpRequest"
        or "application/json" in (request.headers.get("Accept") or "").lower()
    )
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        indent = conn.execute(
            "SELECT id, status FROM store_indents WHERE id = ?",
            (indent_id,),
        ).fetchone()
        if not indent:
            if wants_json:
                return jsonify({"ok": False, "error": "Indent not found."}), 404
            flash("Indent not found.", "error")
            return redirect(url_for("stores_orders", outlet=outlet))
        if indent["status"] != "approved":
            if wants_json:
                return jsonify(
                    {"ok": False, "error": "Purchase orders can only be sent for approved indents."}
                ), 400
            flash("Purchase orders can only be sent for approved indents.", "error")
            return redirect(url_for("stores_orders", outlet=outlet))

        rows = []
        for key, value in request.form.items():
            if not key.startswith("line_supplier_"):
                continue
            try:
                line_id = int(key.rsplit("_", 1)[1])
            except (IndexError, ValueError):
                continue
            rows.append(
                {
                    "line_id": line_id,
                    "supplier_id": (value or "").strip() or None,
                    "rate": (request.form.get(f"line_rate_{line_id}") or "").strip() or None,
                    "quantity": (request.form.get(f"line_qty_{line_id}") or "").strip() or None,
                }
            )
        _save_po_line_overrides(conn, indent_id, rows)
        payload = _load_po_supplier_groups(conn, indent_id)

        groups = (payload or {}).get("groups") or []
        # Only suppliers that still need a PO can be generated from this page.
        pending_ids = {
            int(g["supplier_id"])
            for g in groups
            if not g.get("is_unassigned")
            and g.get("supplier_id")
            and not str(g.get("po_no") or "").strip()
        }
        if not pending_ids:
            has_assigned = any(
                not g.get("is_unassigned") and g.get("supplier_id") for g in groups
            )
            msg = (
                "Every remaining item on this indent already has a purchase order."
                if has_assigned
                else "Assign a supplier to at least one item before generating."
            )
            if wants_json:
                return jsonify({"ok": False, "error": msg}), 400
            flash(msg, "ok" if has_assigned else "error")
            return redirect(url_for("stores_orders_send", indent_id=indent_id, outlet=outlet))

        selected_ids = _parse_selected_supplier_ids(request.form.getlist("selected_supplier"))
        selected_ids = [sid for sid in selected_ids if sid in pending_ids]
        # Suppliers that only became assigned during this save had no checkbox — include them.
        selectable_ids = set(_parse_selected_supplier_ids(request.form.getlist("selectable_supplier")))
        newly_assigned = sorted(pending_ids - selectable_ids)
        for sid in newly_assigned:
            if sid not in selected_ids:
                selected_ids.append(sid)
        if not selected_ids:
            msg = "Select at least one supplier before generating."
            if wants_json:
                return jsonify({"ok": False, "error": msg}), 400
            flash(msg, "error")
            return redirect(url_for("stores_orders_send", indent_id=indent_id, outlet=outlet))

        group_by_supplier = {
            int(g["supplier_id"]): g
            for g in groups
            if g.get("supplier_id") and not g.get("is_unassigned")
        }
        for sid in selected_ids:
            group = group_by_supplier.get(sid)
            if not group:
                continue
            lines_snapshot = [dict(line) for line in (group.get("lines") or [])]
            if not lines_snapshot:
                continue
            po_id, po_no = _allocate_po_row(conn, indent_id, sid)
            if not po_id or not po_no:
                continue
            _save_purchase_order_lines(conn, po_id, lines_snapshot)
            _commit_po_group_quantities(conn, indent_id, lines_snapshot)
            phone = str(group.get("phone") or "").strip()
            issued.append(
                {
                    "indent_id": indent_id,
                    "purchase_order_id": po_id,
                    "supplier_id": sid,
                    "po_no": po_no,
                    "supplier_name": group.get("supplier_name") or "Supplier",
                    "phone": phone,
                    "can_send": bool(phone),
                    "item_count": len(lines_snapshot),
                }
            )
    finally:
        conn.close()

    # Always land on Send to Supplier after generate so the new PO is ready to send.
    # Remaining indent qty (if any) stays available on Generate PO for a later batch.
    orders_url = url_for("stores_orders", outlet=outlet, tab="send")
    redirect_url = orders_url

    if wants_json:
        return jsonify(
            {
                "ok": True,
                "issued": issued,
                "redirect": redirect_url,
                "continue_url": redirect_url,
            }
        )

    if issued:
        flash(
            f"{len(issued)} purchase order{'s' if len(issued) != 1 else ''} generated. "
            "Send to suppliers from Send to Supplier when ready.",
            "ok",
        )
    return redirect(redirect_url)

@stores_bp.route(
    "/stores/orders/<int:indent_id>/pdf/<int:supplier_id>",
    endpoint="stores_orders_pdf",
)
def stores_orders_pdf(indent_id: int, supplier_id: int):
    """Inline PDF purchase order for one supplier (frozen issued lines when available)."""
    from purchase_order_pdf import build_purchase_order_pdf, po_pdf_filename

    conn = get_db()
    try:
        ensure_stores_schema(conn)
        payload = _load_po_supplier_groups(conn, indent_id)
        if not payload:
            flash("Indent not found.", "error")
            return redirect(url_for("stores_orders"))
        if payload.get("status") != "approved":
            flash("Purchase orders are available for approved indents only.", "error")
            return redirect(url_for("stores_orders"))

        po_no_q = str(request.args.get("po_no") or "").strip()
        po_row = _find_purchase_order(
            conn, indent_id=indent_id, supplier_id=supplier_id, po_no=po_no_q
        )

        group = next(
            (g for g in payload.get("groups") or [] if g.get("supplier_id") == supplier_id),
            None,
        )
        supplier_row = conn.execute(
            "SELECT id, name, phone, gst, address FROM suppliers WHERE id = ?",
            (supplier_id,),
        ).fetchone()
        supplier = {
            "name": (group or {}).get("supplier_name")
            or (supplier_row["name"] if supplier_row else "")
            or "",
            "phone": (group or {}).get("phone")
            or (supplier_row["phone"] if supplier_row else "")
            or "",
            "gst": (group or {}).get("gst")
            or (supplier_row["gst"] if supplier_row else "")
            or "",
            "address": (group or {}).get("address")
            or (supplier_row["address"] if supplier_row else "")
            or "",
        }
        lines = _po_lines_for_send_or_pdf(
            conn,
            indent_id=indent_id,
            supplier_id=supplier_id,
            po_row=po_row,
            group_lines=list((group or {}).get("lines") or []),
        )
        if not lines:
            flash("No purchase order lines found for this supplier.", "error")
            return redirect(url_for("stores_orders", tab="orders"))
        if po_row and po_row.get("id"):
            conn.commit()
        po_no = str((po_row or {}).get("po_no") or "").strip() or _get_or_create_po_no(
            conn, indent_id, supplier_id
        )
        pdf_bytes = build_purchase_order_pdf(
            payload.get("indent") or {},
            supplier,
            lines,
            outlet_label=payload.get("outlet_label") or "",
            po_no=po_no,
        )
        fname = po_pdf_filename(
            supplier.get("name") or "Supplier",
            po_no or payload.get("indent_no") or str(indent_id),
        )
    finally:
        conn.close()

    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=False,
        download_name=fname,
    )


def _send_po_whatsapp(
    indent_id: int,
    supplier_id: int,
    *,
    user=None,
    include_pdf: bool = True,
    custom_message: str = "",
    po_no: str = "",
    lines: list | None = None,
    group_snapshot: dict | None = None,
) -> dict[str, Any]:
    """Send one supplier group's purchase order via WhatsApp.

    Returns a dict with at least ``ok`` (bool). On failure also includes ``error``.
    Does not raise for business failures — callers inspect ``ok``.

    Prefers frozen ``store_purchase_order_lines`` for ``po_no`` so deferred sends
    match what was generated (not remaining indent qty).
    """
    import os

    import communication_hub as hub
    import whatsapp_client as wa
    from purchase_order_pdf import build_purchase_order_pdf, po_pdf_filename

    include_pdf = bool(include_pdf)
    custom_message = str(custom_message or "").strip()
    snapshot = group_snapshot if isinstance(group_snapshot, dict) else {}

    conn = get_db()
    try:
        ensure_stores_schema(conn)
        groups_payload = _load_po_supplier_groups(conn, indent_id)
        if not groups_payload:
            return {"ok": False, "error": "Indent not found.", "status": 404}
        if groups_payload.get("status") != "approved":
            return {"ok": False, "error": "Only approved indents can be sent.", "status": 400}

        po_row = _find_purchase_order(
            conn,
            indent_id=indent_id,
            supplier_id=supplier_id,
            po_no=str(po_no or snapshot.get("po_no") or "").strip(),
        )
        frozen_lines = _po_lines_for_send_or_pdf(
            conn,
            indent_id=indent_id,
            supplier_id=supplier_id,
            po_row=po_row,
            group_lines=[],
        )

        group = next(
            (g for g in groups_payload.get("groups") or [] if g.get("supplier_id") == supplier_id),
            None,
        )
        supplier_row = conn.execute(
            "SELECT id, name, phone, gst, address FROM suppliers WHERE id = ?",
            (supplier_id,),
        ).fetchone()
        if snapshot:
            group = {
                "supplier_id": supplier_id,
                "supplier_name": snapshot.get("supplier_name")
                or (group or {}).get("supplier_name")
                or (supplier_row["name"] if supplier_row else "")
                or "Supplier",
                "phone": snapshot.get("phone")
                or (group or {}).get("phone")
                or (supplier_row["phone"] if supplier_row else "")
                or "",
                "gst": snapshot.get("gst")
                or (group or {}).get("gst")
                or (supplier_row["gst"] if supplier_row else "")
                or "",
                "address": snapshot.get("address")
                or (group or {}).get("address")
                or (supplier_row["address"] if supplier_row else "")
                or "",
                "lines": lines if lines is not None else snapshot.get("lines") or [],
                "is_unassigned": False,
            }
        elif not group:
            if not supplier_row and not frozen_lines:
                return {"ok": False, "error": "Assign a supplier before sending.", "status": 400}
            group = {
                "supplier_id": supplier_id,
                "supplier_name": (supplier_row["name"] if supplier_row else "") or "Supplier",
                "phone": (supplier_row["phone"] if supplier_row else "") or "",
                "gst": (supplier_row["gst"] if supplier_row else "") or "",
                "address": (supplier_row["address"] if supplier_row else "") or "",
                "lines": frozen_lines,
                "is_unassigned": False,
            }
        if group.get("is_unassigned"):
            return {"ok": False, "error": "Assign a supplier before sending.", "status": 400}
        phone = wa.normalise_whatsapp_number(group.get("phone") or "")
        if not phone:
            return {"ok": False, "error": "Supplier phone number is missing or invalid.", "status": 400}

        supplier_name = group.get("supplier_name") or "Supplier"
        if lines is not None:
            send_lines = list(lines)
        elif frozen_lines:
            send_lines = list(frozen_lines)
        else:
            send_lines = list(group.get("lines") or [])
        if not send_lines:
            return {
                "ok": False,
                "error": (
                    "This purchase order has no items to send. "
                    "The supplier assignment may have changed — "
                    "generate a new PO from Generate PO."
                ),
                "status": 400,
            }
        indent_no = groups_payload.get("indent_no") or str(indent_id)
        send_po_no = (
            str(po_no or "").strip()
            or str((po_row or {}).get("po_no") or "").strip()
            or _get_or_create_po_no(conn, indent_id, supplier_id)
        )
        po_ref = send_po_no or indent_no
        message = custom_message or _po_default_message(
            supplier_name, send_lines, indent_no, send_po_no
        )
        pdf_name = po_pdf_filename(supplier_name, po_ref)
        pdf_bytes = b""
        if include_pdf:
            pdf_bytes = build_purchase_order_pdf(
                groups_payload.get("indent") or {},
                {
                    "name": supplier_name,
                    "phone": group.get("phone") or "",
                    "gst": group.get("gst") or "",
                    "address": group.get("address") or "",
                },
                send_lines,
                outlet_label=groups_payload.get("outlet_label") or "",
                po_no=send_po_no,
            )

        conversation = hub.get_or_create_conversation(
            conn, phone, supplier_name, revive=True
        )
        if not conversation:
            return {"ok": False, "error": "Could not open WhatsApp conversation.", "status": 400}
        conversation_id = int(conversation["id"])
        user_id = int(user["id"]) if user and user.get("id") else None

        live = wa.whatsapp_live_sends_allowed()
        wa_message_id = ""
        status = "sent"
        error = ""

        if not live:
            # Dry-run / testing: still create Hub history + PO send row.
            hub.append_message(
                conn,
                conversation_id,
                direction="out",
                body=message if not include_pdf else ((message[:180] + "…") if len(message) > 180 else message) or pdf_name,
                message_type="document" if include_pdf else "text",
                media_mime="application/pdf" if include_pdf else "",
                media_filename=pdf_name if include_pdf else "",
                media_size=len(pdf_bytes) if include_pdf else 0,
                wa_message_id="",
                status="sent",
                created_by=user_id,
            )
            send_id = _record_po_send(
                conn,
                indent_id=indent_id,
                supplier_id=supplier_id,
                phone=phone,
                message=message,
                pdf_name=pdf_name if include_pdf else "",
                include_pdf=include_pdf,
                conversation_id=conversation_id,
                wa_message_id="",
                status="sent",
                error="",
                sent_by=user_id,
                po_no=send_po_no,
            )
            conn.commit()
            return {
                "ok": True,
                "dry_run": True,
                "po_no": send_po_no,
                "send_id": send_id,
                "conversation_id": conversation_id,
                "pdf_name": pdf_name if include_pdf else "",
                "supplier_name": supplier_name,
            }

        if not wa.whatsapp_configured():
            return {"ok": False, "error": "WhatsApp API is not configured.", "status": 400}

        if include_pdf and pdf_bytes:
            ok, err, result = hub.send_conversation_document_bytes(
                conn,
                conversation_id,
                data=pdf_bytes,
                filename=pdf_name,
                mime="application/pdf",
                caption=message,
                user_id=user_id,
            )
            if not ok and _po_is_outside_session_error(err):
                template_name = (os.environ.get("WHATSAPP_PO_TEMPLATE") or "").strip()
                template_lang = (os.environ.get("WHATSAPP_PO_TEMPLATE_LANGUAGE") or "en").strip() or "en"
                if template_name:
                    # Upload media then send as template document header.
                    import tempfile
                    tmp_path = ""
                    try:
                        fd, tmp_path = tempfile.mkstemp(suffix=".pdf")
                        with os.fdopen(fd, "wb") as handle:
                            handle.write(pdf_bytes)
                        ok_up, err_up, body_up = wa.upload_media_file(tmp_path, "application/pdf")
                        media_id = ""
                        if isinstance(body_up, dict):
                            media_id = str(body_up.get("id") or "").strip()
                        if ok_up and media_id:
                            ok_tpl, err_tpl, payload_tpl = wa.send_template_message(
                                phone,
                                template_name,
                                template_lang,
                                body_parameters=[supplier_name, po_ref],
                                header_document_id=media_id,
                                header_document_filename=pdf_name,
                            )
                            if ok_tpl:
                                wa_message_id = wa.first_message_id(payload_tpl)
                                hub.append_message(
                                    conn,
                                    conversation_id,
                                    direction="out",
                                    body=message,
                                    message_type="template",
                                    media_mime="application/pdf",
                                    media_filename=pdf_name,
                                    media_size=len(pdf_bytes),
                                    wa_message_id=wa_message_id,
                                    status="sent",
                                    created_by=user_id,
                                )
                                ok, err = True, ""
                            else:
                                ok, err = False, err_tpl or err
                        else:
                            ok, err = False, err_up or err
                    finally:
                        if tmp_path:
                            try:
                                os.unlink(tmp_path)
                            except OSError:
                                pass
                else:
                    err = (
                        "WhatsApp session expired (outside 24-hour window). "
                        "Ask the supplier to message first, or set WHATSAPP_PO_TEMPLATE."
                    )
            if ok:
                msg = (result or {}).get("message") or {}
                wa_message_id = str(msg.get("wa_message_id") or wa_message_id or "")
            else:
                status = "failed"
                error = err or "Send failed"
        else:
            ok, err, result = hub.send_conversation_text(
                conn, conversation_id, message, user_id=user_id
            )
            if not ok and _po_is_outside_session_error(err):
                template_name = (os.environ.get("WHATSAPP_PO_TEMPLATE") or "").strip()
                template_lang = (os.environ.get("WHATSAPP_PO_TEMPLATE_LANGUAGE") or "en").strip() or "en"
                if template_name:
                    ok_tpl, err_tpl, payload_tpl = wa.send_template_message(
                        phone,
                        template_name,
                        template_lang,
                        body_parameters=[supplier_name, po_ref],
                    )
                    if ok_tpl:
                        wa_message_id = wa.first_message_id(payload_tpl)
                        hub.append_message(
                            conn,
                            conversation_id,
                            direction="out",
                            body=message,
                            message_type="template",
                            wa_message_id=wa_message_id,
                            status="sent",
                            created_by=user_id,
                        )
                        ok, err = True, ""
                    else:
                        ok, err = False, err_tpl or err
                else:
                    err = (
                        "WhatsApp session expired (outside 24-hour window). "
                        "Ask the supplier to message first, or set WHATSAPP_PO_TEMPLATE."
                    )
            if ok:
                msg = (result or {}).get("message") or {}
                wa_message_id = str(msg.get("wa_message_id") or wa_message_id or "")
            else:
                status = "failed"
                error = err or "Send failed"

        send_id = _record_po_send(
            conn,
            indent_id=indent_id,
            supplier_id=supplier_id,
            phone=phone,
            message=message,
            pdf_name=pdf_name if include_pdf else "",
            include_pdf=include_pdf,
            conversation_id=conversation_id,
            wa_message_id=wa_message_id,
            status=status,
            error=error,
            sent_by=user_id,
            po_no=send_po_no,
        )
        conn.commit()
        if status != "sent":
            return {
                "ok": False,
                "error": error or "Send failed",
                "send_id": send_id,
                "status": 400,
                "supplier_name": supplier_name,
                "po_no": send_po_no,
            }
        return {
            "ok": True,
            "send_id": send_id,
            "po_no": send_po_no,
            "conversation_id": conversation_id,
            "pdf_name": pdf_name if include_pdf else "",
            "wa_message_id": wa_message_id,
            "supplier_name": supplier_name,
        }
    finally:
        conn.close()


@stores_bp.route("/stores/orders/<int:indent_id>/send", methods=["POST"], endpoint="stores_orders_send_wa")
def stores_orders_send_wa(indent_id: int):
    """Send a supplier group's purchase order via WhatsApp (Communication Hub)."""
    user = _get_user() if _get_user else None
    payload = request.get_json(silent=True) or {}
    try:
        supplier_id = int(payload.get("supplier_id"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Supplier is required."}), 400
    include_pdf = bool(payload.get("include_pdf", True))
    custom_message = str(payload.get("message") or "").strip()
    po_no = str(payload.get("po_no") or "").strip()
    result = _send_po_whatsapp(
        indent_id,
        supplier_id,
        user=user,
        include_pdf=include_pdf,
        custom_message=custom_message,
        po_no=po_no,
    )
    status = int(result.pop("status", 200 if result.get("ok") else 400))
    return jsonify(result), status



def _build_indent_purchase_order_xlsx(indent: dict[str, Any], lines: list[dict[str, Any]]) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Order"

    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True)
    label_font = Font(bold=True)

    ws["A1"] = "Hotel Bell Elite"
    ws["A1"].font = title_font
    ws["A2"] = "Purchase Order"
    ws["A2"].font = Font(bold=True, size=13)

    meta = [
        ("Indent No", indent.get("indent_no") or ""),
        ("Outlet", _outlet_label(_parse_outlet(indent.get("outlet")))),
        ("Status", _status_label(indent.get("status") or "")),
        ("Created", _format_stores_dt(indent.get("created_at"))),
        ("Created by", indent.get("created_by_name") or ""),
        ("Notes", indent.get("notes") or ""),
    ]
    row_idx = 4
    for label, value in meta:
        ws.cell(row=row_idx, column=1, value=label).font = label_font
        ws.cell(row=row_idx, column=2, value=value)
        row_idx += 1

    row_idx += 1
    headers = ("#", "Item", "Qty", "Unit", "Approx. price", "Amount")
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col, value=title)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if col != 2 else "left")

    total_amount = 0.0
    has_amount = False
    for idx, line in enumerate(lines, start=1):
        row_idx += 1
        qty = float(line.get("quantity") or 0)
        price = line.get("approximate_price")
        try:
            price_num = float(price) if price is not None and price != "" else None
        except (TypeError, ValueError):
            price_num = None
        amount = None
        if price_num is not None:
            amount = round(qty * price_num, 2)
            total_amount += amount
            has_amount = True
        ws.cell(row=row_idx, column=1, value=idx)
        item_label = _format_indent_line_item(line)
        pack_label = (line.get("pack_label") or "").strip()
        ws.cell(row=row_idx, column=2, value=item_label)
        ws.cell(row=row_idx, column=3, value=qty)
        ws.cell(row=row_idx, column=4, value=pack_label or (line.get("unit") or ""))
        ws.cell(row=row_idx, column=5, value=price_num if price_num is not None else "")
        ws.cell(row=row_idx, column=6, value=amount if amount is not None else "")

    if has_amount:
        row_idx += 1
        ws.cell(row=row_idx, column=5, value="Total").font = header_font
        ws.cell(row=row_idx, column=6, value=round(total_amount, 2)).font = header_font

    widths = (6, 32, 10, 10, 14, 12)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@stores_bp.route("/stores/indent/<int:indent_id>/purchase-order")
def stores_indent_purchase_order(indent_id: int):
    """Excel purchase order for an approved indent."""
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        indent = conn.execute(
            """
            SELECT i.*, u.full_name AS created_by_name
            FROM store_indents i
            LEFT JOIN users u ON u.id = i.created_by
            WHERE i.id = ?
            """,
            (indent_id,),
        ).fetchone()
        if not indent:
            flash("Indent not found.", "error")
            return redirect(url_for("stores_orders"))
        if indent["status"] != "approved":
            flash("Purchase orders are available for approved indents only.", "error")
            return redirect(url_for("stores_indent", outlet=indent["outlet"], view="pending"))
        lines = [
            dict(row)
            for row in conn.execute(
                """
                SELECT item_name, quantity, unit, approximate_price, notes
                FROM store_indent_lines
                WHERE indent_id = ?
                ORDER BY id
                """,
                (indent_id,),
            ).fetchall()
        ]
        indent_data = dict(indent)
    finally:
        conn.close()

    buf = _build_indent_purchase_order_xlsx(indent_data, lines)
    safe_no = re.sub(r"[^\w.-]+", "_", str(indent_data.get("indent_no") or indent_id))
    fname = f"PO_{safe_no}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _delete_approved_indent_non_inwarded(conn, indent_id: int) -> tuple[bool, str]:
    """Admin-only: cancel non-inwarded qty on an approved indent.

    Already inwarded quantities (``quantity_received``) and related stock
    movements are left untouched. Lines with no receipt are removed; partially
    received lines are reduced to ``quantity_received``. If nothing remains,
    the indent header is deleted.
    """
    indent = conn.execute(
        "SELECT id, indent_no, status FROM store_indents WHERE id = ?",
        (indent_id,),
    ).fetchone()
    if not indent:
        return False, "Indent not found."
    if (indent["status"] or "") != "approved":
        return False, "Only approved indents support this delete."

    lines = conn.execute(
        """
        SELECT id, item_name, quantity, COALESCE(quantity_received, 0) AS quantity_received
        FROM store_indent_lines
        WHERE indent_id = ?
        ORDER BY id
        """,
        (indent_id,),
    ).fetchall()
    if not lines:
        conn.execute("DELETE FROM store_indents WHERE id = ?", (indent_id,))
        return True, f"Deleted {indent['indent_no']}."

    removed_lines = 0
    trimmed_lines = 0
    kept_lines = 0
    for line in lines:
        try:
            ordered = float(line["quantity"] or 0)
        except (TypeError, ValueError):
            ordered = 0.0
        try:
            received = float(line["quantity_received"] or 0)
        except (TypeError, ValueError):
            received = 0.0
        if received < 0:
            received = 0.0
        if received <= 0.0001:
            conn.execute("DELETE FROM store_indent_lines WHERE id = ?", (int(line["id"]),))
            removed_lines += 1
            continue
        if ordered - received > 0.0001:
            conn.execute(
                "UPDATE store_indent_lines SET quantity = ? WHERE id = ?",
                (received, int(line["id"])),
            )
            trimmed_lines += 1
            kept_lines += 1
        else:
            # Fully inwarded — leave as-is.
            kept_lines += 1

    remaining = conn.execute(
        "SELECT COUNT(*) AS c FROM store_indent_lines WHERE indent_id = ?",
        (indent_id,),
    ).fetchone()["c"]
    indent_no = indent["indent_no"]
    if not remaining:
        conn.execute("DELETE FROM store_indents WHERE id = ?", (indent_id,))
        return True, f"Deleted {indent_no} (no inwarded stock)."

    if removed_lines == 0 and trimmed_lines == 0:
        return False, f"{indent_no} is fully inwarded — nothing left to delete."

    parts = []
    if removed_lines:
        parts.append(f"removed {removed_lines} pending line{'s' if removed_lines != 1 else ''}")
    if trimmed_lines:
        parts.append(f"trimmed {trimmed_lines} partially inwarded line{'s' if trimmed_lines != 1 else ''}")
    return True, f"Updated {indent_no}: {', '.join(parts)}. Inwarded stock unchanged."


@stores_bp.route("/stores/indent/<int:indent_id>/delete", methods=["GET", "POST"])
def stores_indent_delete(indent_id: int):
    outlet = _parse_outlet_filter(request.args.get("outlet") or request.form.get("outlet"))
    user = _get_user() if _get_user else None
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        indent = conn.execute(
            "SELECT id, indent_no, outlet, status FROM store_indents WHERE id = ?",
            (indent_id,),
        ).fetchone()
        if not indent:
            flash("Indent not found.", "error")
            return redirect(url_for("stores_indent", outlet=outlet))
        outlet = _parse_outlet(indent["outlet"])
        status = (indent["status"] or "").strip().lower()

        if status == "approved":
            if not user or not user.get("is_admin"):
                flash("Only administrators can delete approved indents.", "error")
                return redirect(url_for("stores_orders", outlet=outlet))
            ok, message = _delete_approved_indent_non_inwarded(conn, indent_id)
            if ok:
                conn.commit()
                flash(message, "ok")
            else:
                conn.rollback()
                flash(message, "error")
            return redirect(url_for("stores_orders", outlet=outlet))

        if status not in ("draft", "pending"):
            flash("Only draft or waiting indents can be deleted.", "error")
            return redirect(url_for("stores_indent", outlet=outlet))
        conn.execute("DELETE FROM store_indent_lines WHERE indent_id = ?", (indent_id,))
        conn.execute("DELETE FROM store_indents WHERE id = ?", (indent_id,))
        conn.commit()
        flash(f"Deleted {indent['indent_no']}.", "ok")
    finally:
        conn.close()
    return redirect(url_for("stores_indent", outlet=outlet))


@stores_bp.route("/stores/indent/<int:indent_id>")
def stores_indent_detail(indent_id: int):
    outlet = _parse_outlet(request.args.get("outlet"))
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        indent = conn.execute(
            """
            SELECT i.*, u.full_name AS created_by_name,
                   d.full_name AS decided_by_name,
                   d.username AS decided_by_username
            FROM store_indents i
            LEFT JOIN users u ON u.id = i.created_by
            LEFT JOIN users d ON d.id = i.decided_by
            WHERE i.id = ?
            """,
            (indent_id,),
        ).fetchone()
        if not indent:
            flash("Indent not found.", "error")
            return redirect(url_for("stores_indent", outlet=outlet))
        outlet = indent["outlet"]
        lines = conn.execute(
            "SELECT * FROM store_indent_lines WHERE indent_id = ? ORDER BY id",
            (indent_id,),
        ).fetchall()
        detail_lines = []
        for line in lines:
            item = dict(line)
            item["approximate_price_display"] = _format_optional_price(item.get("approximate_price"))
            item["display_name"] = _format_indent_line_item(item)
            detail_lines.append(item)
    finally:
        conn.close()
    return _page_render(
        "indent",
        outlet=outlet,
        indents=[],
        show_form=False,
        detail=dict(indent),
        detail_lines=detail_lines,
        form=None,
        errors=[],
    )


@stores_bp.route("/stores/indent/<int:indent_id>/submit", methods=["POST"])
def stores_indent_submit(indent_id: int):
    outlet = _parse_outlet(request.form.get("outlet") or request.args.get("outlet"))
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        indent = conn.execute("SELECT * FROM store_indents WHERE id = ?", (indent_id,)).fetchone()
        if not indent:
            flash("Indent not found.", "error")
            return redirect(url_for("stores_indent", outlet=outlet))
        outlet = indent["outlet"]
        if indent["status"] != "draft":
            flash("Only draft indents can be sent for approval.", "error")
            return redirect(url_for("stores_indent", outlet=outlet))
        line_count = conn.execute(
            "SELECT COUNT(*) AS c FROM store_indent_lines WHERE indent_id = ?",
            (indent_id,),
        ).fetchone()["c"]
        if not line_count:
            flash("Add items before sending for approval.", "error")
            return redirect(url_for("stores_indent_detail", indent_id=indent_id, outlet=outlet))
        # Bind the WHERE to status='draft' so a concurrent duplicate submit (double
        # click, retried request) can't both flip the row and each send an approval.
        update_cur = conn.execute(
            "UPDATE store_indents SET status = 'pending', submitted_at = ? WHERE id = ? AND status = 'draft'",
            (_now(), indent_id),
        )
        if update_cur.rowcount == 0:
            conn.commit()
            flash("Indent already sent for approval.", "ok")
            return redirect(url_for("stores_indent", outlet=outlet))
        supersede_indent_whatsapp_sends(conn, indent_id)
        assign_fresh_approval_token(conn, indent_id)
        conn.commit()
        flash("Indent sent for approval.", "ok")
        _notify_indent_pending_whatsapp(conn, indent_id, outlet)
    finally:
        conn.close()
    return redirect(url_for("stores_indent", outlet=outlet))


@stores_bp.route("/stores/approvals")
def stores_approvals():
    outlet = _parse_outlet_filter(request.args.get("outlet"))
    outlet_sql, outlet_params = _outlet_match_sql("i.outlet", outlet)
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        pending = conn.execute(
            f"""
            SELECT i.*, u.full_name AS created_by_name,
                   (SELECT COUNT(*) FROM store_indent_lines l WHERE l.indent_id = i.id) AS line_count,
                   (SELECT COALESCE(SUM(
                        COALESCE(l.quantity, 0) * COALESCE(l.approximate_price, 0)
                    ), 0)
                    FROM store_indent_lines l WHERE l.indent_id = i.id) AS approximate_total
            FROM store_indents i
            LEFT JOIN users u ON u.id = i.created_by
            WHERE {outlet_sql} AND i.status = 'pending'
            ORDER BY i.submitted_at ASC, i.id ASC
            """,
            outlet_params,
        ).fetchall()
        recent = conn.execute(
            f"""
            SELECT i.*, u.full_name AS created_by_name, d.full_name AS decided_by_name
            FROM store_indents i
            LEFT JOIN users u ON u.id = i.created_by
            LEFT JOIN users d ON d.id = i.decided_by
            WHERE {outlet_sql} AND i.status IN ('approved', 'rejected')
            ORDER BY i.decided_at DESC, i.id DESC
            LIMIT 20
            """,
            outlet_params,
        ).fetchall()
    finally:
        conn.close()
    pending_rows = []
    for row in pending:
        item = dict(row)
        total = item.get("approximate_total")
        try:
            total_num = float(total or 0)
        except (TypeError, ValueError):
            total_num = 0.0
        item["approximate_total"] = total_num
        item["approximate_total_display"] = (
            _format_optional_price(total_num) if total_num > 0 else ""
        )
        pending_rows.append(item)
    return _page_render(
        "approvals",
        outlet=outlet,
        pending=pending_rows,
        recent=[dict(row) for row in recent],
    )


@stores_bp.route("/stores/indent/<int:indent_id>/decide", methods=["POST"])
def stores_indent_decide(indent_id: int):
    user = _get_user()
    decision = (request.form.get("decision") or "").strip().lower()
    note = (request.form.get("decision_note") or "").strip()
    outlet = _parse_outlet(request.form.get("outlet"))
    if decision not in {"approved", "rejected"}:
        flash("Choose approve or reject.", "error")
        return redirect(url_for("stores_approvals", outlet=outlet))
    if decision == "rejected" and not note:
        flash("Add a short reason when rejecting.", "error")
        return redirect(url_for("stores_approvals", outlet=outlet))

    conn = get_db()
    try:
        ensure_stores_schema(conn)
        indent = conn.execute("SELECT * FROM store_indents WHERE id = ?", (indent_id,)).fetchone()
        if not indent or indent["status"] != "pending":
            flash("This indent is not waiting for approval.", "error")
            return redirect(url_for("stores_approvals", outlet=outlet))
        outlet = indent["outlet"]
        conn.execute(
            """
            UPDATE store_indents
            SET status = ?, decided_by = ?, decided_at = ?, decision_note = ?
            WHERE id = ?
            """,
            (decision, user["id"] if user else None, _now(), note, indent_id),
        )
        conn.commit()
    finally:
        conn.close()
    flash("Indent approved." if decision == "approved" else "Indent rejected.", "ok")
    return redirect(url_for("stores_approvals", outlet=outlet))


@stores_bp.route("/stores/indent/<int:indent_id>/reopen", methods=["POST"])
def stores_indent_reopen(indent_id: int):
    """Return a rejected indent to Waiting for approval."""
    outlet = _parse_outlet_filter(request.form.get("outlet") or request.args.get("outlet"))
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        indent = conn.execute(
            "SELECT id, outlet, status, indent_no FROM store_indents WHERE id = ?",
            (indent_id,),
        ).fetchone()
        if not indent:
            flash("Indent not found.", "error")
            return redirect(url_for("stores_approvals", outlet=outlet))
        outlet = _parse_outlet(indent["outlet"])
        if indent["status"] != "rejected":
            flash("Only rejected indents can be returned to waiting.", "error")
            return redirect(url_for("stores_approvals", outlet=outlet))
        conn.execute(
            """
            UPDATE store_indents
            SET status = 'pending',
                decided_by = NULL,
                decided_at = NULL,
                decision_note = '',
                submitted_at = COALESCE(submitted_at, ?)
            WHERE id = ?
            """,
            (_now(), indent_id),
        )
        conn.commit()
        flash(f"{indent['indent_no']} returned to waiting approval.", "ok")
    finally:
        conn.close()
    return redirect(url_for("stores_approvals", outlet=outlet))


def _parse_inward_view(raw: Any) -> str:
    value = str(raw or "").strip().lower()
    if value == "direct":
        return "direct"
    return "approved"


@stores_bp.route("/stores/purchase-requests", methods=["GET", "POST"])
def stores_purchase_requests():
    outlet = _parse_outlet_filter(request.args.get("outlet") or request.form.get("outlet"))
    inward_view = _parse_inward_view(request.args.get("view") or request.form.get("view"))
    user = _get_user()

    if request.method == "POST" and request.form.get("action") == "create_from_indent":
        try:
            indent_id = int(request.form.get("indent_id") or 0)
        except (TypeError, ValueError):
            indent_id = 0
        conn = get_db()
        try:
            ensure_stores_schema(conn)
            indent = conn.execute(
                "SELECT * FROM store_indents WHERE id = ? AND status = 'approved'",
                (indent_id,),
            ).fetchone()
            if not indent:
                flash("Select an approved indent.", "error")
                return redirect(url_for("stores_purchase_requests", outlet=outlet, view=inward_view))
            write_outlet = _parse_outlet(indent["outlet"])
            existing = conn.execute(
                "SELECT id FROM store_purchase_requests WHERE indent_id = ?",
                (indent_id,),
            ).fetchone()
            if existing:
                flash("A purchase request already exists for this indent.", "error")
                return redirect(url_for("stores_purchase_requests", outlet=write_outlet, view="approved"))
            lines = conn.execute(
                "SELECT * FROM store_indent_lines WHERE indent_id = ? ORDER BY id",
                (indent_id,),
            ).fetchall()
            if not lines:
                flash("This indent has no items.", "error")
                return redirect(url_for("stores_purchase_requests", outlet=write_outlet, view="approved"))
            pr_no = _next_doc_no(conn, "store_purchase_requests", "pr_no", "PR", write_outlet)
            cur = conn.execute(
                """
                INSERT INTO store_purchase_requests
                    (indent_id, outlet, pr_no, status, notes, created_by, created_at)
                VALUES (?, ?, ?, 'open', ?, ?, ?)
                """,
                (
                    indent_id,
                    write_outlet,
                    pr_no,
                    (request.form.get("notes") or "").strip(),
                    user["id"] if user else None,
                    _now(),
                ),
            )
            pr_id = cur.lastrowid
            for line in lines:
                conn.execute(
                    """
                    INSERT INTO store_purchase_request_lines
                        (pr_id, item_name, quantity, unit, notes, pack_label, pack_qty_in_base)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        pr_id,
                        line["item_name"],
                        line["quantity"],
                        line["unit"],
                        line["notes"] or "",
                        (line["pack_label"] if "pack_label" in line.keys() else "") or "",
                        line["pack_qty_in_base"] if "pack_qty_in_base" in line.keys() else None,
                    ),
                )
            conn.commit()
        finally:
            conn.close()
        flash("Purchase request created.", "ok")
        return redirect(url_for("stores_purchase_requests", outlet=write_outlet, view="approved"))

    if request.method == "POST" and request.form.get("action") == "confirm_stock_inward":
        # Stock + expense must go through the expense modal / JSON endpoint.
        flash("Confirm stock inward from the expense popup.", "error")
        try:
            indent_id = int(request.form.get("indent_id") or 0)
        except (TypeError, ValueError):
            indent_id = 0
        redirect_kwargs = {"outlet": outlet, "view": inward_view}
        if indent_id:
            redirect_kwargs["indent"] = indent_id
        return redirect(url_for("stores_purchase_requests", **redirect_kwargs))

    # Lazy import avoids circular import with app.register_stores
    import app as app_module

    approved_indents: list[dict[str, Any]] = []
    generated_purchase_orders: list[dict[str, Any]] = []
    inward_supplier_options: list[dict[str, Any]] = []
    selected_inward_supplier_id = 0
    selected_indent = None
    selected_po = None
    selected_lines: list[dict[str, Any]] = []
    indent_view_data: list[dict[str, Any]] = []
    product_catalog: list[dict[str, Any]] = []
    direct_outlet_unset = False

    conn = get_db()
    expense_categories = app_module.EXPENSE_CATEGORIES
    try:
        ensure_stores_schema(conn)
        suppliers = app_module._all_suppliers(conn)
        today = date.today()
        available_cash = app_module._cash_ledger_available_as_of(
            conn, app_module.DEFAULT_COMPANY, today
        )
        expense_categories = app_module._expense_category_choices(conn)
        expense_categories = _sync_product_categories_into_expense_categories(conn)
        conn.commit()

        if inward_view == "direct":
            write_outlet = _parse_outlet(outlet) if outlet and outlet != "both" else ""
            direct_outlet_unset = not bool(write_outlet)
            if write_outlet:
                product_catalog = _load_product_catalog(conn, stores_outlet=write_outlet)
            inward_confirm_url = url_for("stores_confirm_direct_stock_inward_expense")
        else:
            # Filter POs by outlet (All shows Bar + Restaurant).
            # Hide fully received POs from the inward picker.
            all_generated_pos = _load_generated_purchase_orders(
                conn, outlet, limit=300, pending_inward=True
            )
            inward_supplier_options = []
            seen_supplier_ids: set[int] = set()
            for po in all_generated_pos:
                try:
                    sid = int(po.get("supplier_id") or 0)
                except (TypeError, ValueError):
                    sid = 0
                if sid <= 0 or sid in seen_supplier_ids:
                    continue
                seen_supplier_ids.add(sid)
                name = str(po.get("supplier_name") or "").strip() or f"Supplier #{sid}"
                if name == "—":
                    name = f"Supplier #{sid}"
                inward_supplier_options.append({"id": sid, "name": name})
            inward_supplier_options.sort(key=lambda row: str(row.get("name") or "").lower())

            selected_inward_supplier_id = 0
            try:
                selected_inward_supplier_id = int(
                    request.args.get("supplier_id")
                    or request.form.get("supplier_id")
                    or 0
                )
            except (TypeError, ValueError):
                selected_inward_supplier_id = 0
            if selected_inward_supplier_id and selected_inward_supplier_id not in seen_supplier_ids:
                selected_inward_supplier_id = 0

            if selected_inward_supplier_id:
                generated_purchase_orders = [
                    po
                    for po in all_generated_pos
                    if int(po.get("supplier_id") or 0) == selected_inward_supplier_id
                ]
            else:
                generated_purchase_orders = list(all_generated_pos)
            # Keep approved_indents for any legacy templates / view-indent helpers.
            approved_indents = []
            seen_indent_ids: set[int] = set()
            for po in generated_purchase_orders:
                try:
                    iid = int(po.get("indent_id") or 0)
                except (TypeError, ValueError):
                    iid = 0
                if iid <= 0 or iid in seen_indent_ids:
                    continue
                seen_indent_ids.add(iid)
                approved_indents.append(
                    {
                        "id": iid,
                        "indent_no": po.get("indent_no") or f"#{iid}",
                        "outlet": po.get("outlet") or outlet,
                    }
                )

            selected_po_id = 0
            selected_po_no = str(request.args.get("po") or request.form.get("po") or "").strip()
            try:
                selected_po_id = int(
                    request.args.get("po_id") or request.form.get("po_id") or 0
                )
            except (TypeError, ValueError):
                selected_po_id = 0
            # Legacy ?indent= still opens the newest PO for that indent.
            legacy_indent_id = 0
            try:
                legacy_indent_id = int(request.args.get("indent") or 0)
            except (TypeError, ValueError):
                legacy_indent_id = 0

            if selected_po_id:
                for row in generated_purchase_orders:
                    if int(row.get("po_id") or 0) == selected_po_id:
                        selected_po = row
                        break
            if selected_po is None and selected_po_no:
                for row in generated_purchase_orders:
                    if str(row.get("po_no") or "") == selected_po_no:
                        selected_po = row
                        break
            if selected_po is None and legacy_indent_id:
                for row in generated_purchase_orders:
                    if int(row.get("indent_id") or 0) == legacy_indent_id:
                        selected_po = row
                        break
            if selected_po is None and len(generated_purchase_orders) == 1:
                selected_po = generated_purchase_orders[0]

            # Keep a concrete outlet filter aligned with the selected PO.
            # When filter is All, leave it as All — the indent carries write outlet.
            if selected_po is not None:
                po_outlet = _parse_outlet(selected_po.get("outlet"))
                if (
                    outlet in OUTLET_KEYS
                    and po_outlet in OUTLET_KEYS
                    and po_outlet != outlet
                ):
                    redirect_kwargs = {
                        "outlet": po_outlet,
                        "view": inward_view,
                        "po_id": int(selected_po.get("po_id") or 0),
                    }
                    if selected_inward_supplier_id:
                        redirect_kwargs["supplier_id"] = selected_inward_supplier_id
                    return redirect(
                        url_for("stores_purchase_requests", **redirect_kwargs)
                    )

            if selected_po is not None:
                selected_indent, selected_lines = _build_inward_lines_for_po(
                    conn, selected_po, outlet=outlet
                )
            indent_view_data = _indent_view_payload(
                conn,
                [selected_indent] if selected_indent else [],
            )
            inward_confirm_url = url_for("stores_confirm_stock_inward_expense")
    finally:
        conn.close()

    return _page_render(
        "purchase_requests",
        outlet=outlet,
        inward_view=inward_view,
        inward_list_views=[
            ("approved", "Indent Approved"),
            ("direct", "Without Indent Approval"),
        ],
        approved_indents=approved_indents,
        generated_purchase_orders=generated_purchase_orders,
        inward_supplier_options=inward_supplier_options,
        selected_inward_supplier_id=selected_inward_supplier_id,
        selected_po=selected_po,
        selected_indent=selected_indent,
        selected_lines=selected_lines,
        indent_view_data=indent_view_data,
        product_catalog=product_catalog,
        direct_outlet_unset=direct_outlet_unset,
        suppliers=suppliers,
        expense_categories=expense_categories,
        expense_payment_types=app_module.EXPENSE_PAYMENT_TYPES,
        available_cash=available_cash,
        available_cash_url=url_for("cash_ledger_available"),
        supplier_create_url=url_for("create_supplier"),
        default_company=app_module.DEFAULT_COMPANY,
        default_location=app_module.OUTLET_HOTEL,
        today_iso=today.isoformat(),
        inward_confirm_url=inward_confirm_url,
        inward_save_category_url=url_for("stores_save_expense_category"),
        back_href=url_for("stores_stock", outlet=outlet if outlet else "both"),
    )


@stores_bp.route("/stores/purchase-requests/expense-category", methods=["POST"])
def stores_save_expense_category():
    """Add a custom expense category for Inward / Expense Ledger dropdowns."""
    user = _get_user()
    if not user:
        return jsonify({"ok": False, "error": "You must be logged in."}), 401

    import app as app_module
    import re

    data = request.get_json(silent=True) or {}
    name = (data.get("category_name") or data.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "Category name is required."}), 400
    if len(name) > 80:
        return jsonify({"ok": False, "error": "Category name must be 80 characters or fewer."}), 400

    key = app_module._slugify_expense_category_key(name)
    if not key or not re.fullmatch(r"[a-z][a-z0-9_]{0,79}", key):
        return jsonify({"ok": False, "error": "Enter a valid category name."}), 400

    # Prefer builtin key when the name matches an existing label (case-insensitive).
    for builtin_key, builtin_label in app_module.EXPENSE_CATEGORIES:
        if builtin_label.casefold() == name.casefold() or builtin_key == key:
            return jsonify({
                "ok": True,
                "category_key": builtin_key,
                "category_label": builtin_label,
                "existing": True,
            })

    conn = get_db()
    try:
        ensure_stores_schema(conn)
        # Schema for expense_categories lives in init_db path; ensure via pragma/create.
        conn.execute("""
            CREATE TABLE IF NOT EXISTS expense_categories (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                category_key  TEXT    NOT NULL UNIQUE,
                name          TEXT    NOT NULL COLLATE NOCASE,
                sort_order    INTEGER NOT NULL DEFAULT 0,
                is_active     INTEGER NOT NULL DEFAULT 1,
                created_at    TEXT    NOT NULL DEFAULT (datetime('now','localtime'))
            )
        """)
        by_key = conn.execute(
            "SELECT category_key, name, is_active FROM expense_categories WHERE category_key = ?",
            (key,),
        ).fetchone()
        by_name = conn.execute(
            "SELECT category_key, name, is_active FROM expense_categories WHERE lower(name) = lower(?)",
            (name,),
        ).fetchone()
        existing = by_name or by_key
        if existing:
            if int(existing["is_active"] or 0) != 1:
                conn.execute(
                    "UPDATE expense_categories SET is_active = 1, name = ? WHERE category_key = ?",
                    (name, existing["category_key"]),
                )
                conn.commit()
            return jsonify({
                "ok": True,
                "category_key": existing["category_key"],
                "category_label": name if int(existing["is_active"] or 0) != 1 else existing["name"],
                "existing": True,
            })

        max_sort = conn.execute(
            "SELECT COALESCE(MAX(sort_order), 0) AS m FROM expense_categories"
        ).fetchone()["m"]
        conn.execute(
            """
            INSERT INTO expense_categories (category_key, name, sort_order, is_active)
            VALUES (?, ?, ?, 1)
            """,
            (key, name, int(max_sort) + 10),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        return jsonify({"ok": False, "error": "Could not save category."}), 500
    finally:
        conn.close()

    return jsonify({
        "ok": True,
        "category_key": key,
        "category_label": name,
        "existing": False,
    })


@stores_bp.route("/stores/purchase-requests/confirm-with-expense", methods=["POST"])
def stores_confirm_stock_inward_expense():
    """Confirm stock inward and record Hotel expense in one transaction."""
    user = _get_user()
    if not user:
        return jsonify({"ok": False, "error": "You must be logged in."}), 401

    import app as app_module

    data = request.get_json(silent=True) or {}
    try:
        indent_id = int(data.get("indent_id") or 0)
    except (TypeError, ValueError):
        indent_id = 0
    try:
        purchase_order_id = int(data.get("po_id") or data.get("purchase_order_id") or 0)
    except (TypeError, ValueError):
        purchase_order_id = 0
    notes = (data.get("notes") or "").strip()[:500]
    raw_lines = data.get("lines") or []
    if not isinstance(raw_lines, list):
        raw_lines = []

    # line_id -> (received_qty, unit_price, tax_percent)
    selected: dict[int, tuple[float, float | None, float | None]] = {}
    for raw in raw_lines:
        if not isinstance(raw, dict):
            continue
        try:
            line_id = int(raw.get("line_id") or raw.get("id") or 0)
            qty = float(raw.get("received_qty") or raw.get("quantity") or 0)
        except (TypeError, ValueError):
            continue
        if line_id <= 0 or qty <= 0:
            continue
        unit_price = None
        tax_percent = None
        try:
            if raw.get("unit_price") not in (None, ""):
                unit_price = float(raw.get("unit_price"))
        except (TypeError, ValueError):
            unit_price = None
        try:
            if raw.get("tax_percent") not in (None, ""):
                tax_percent = float(raw.get("tax_percent"))
        except (TypeError, ValueError):
            tax_percent = None
        selected[line_id] = (qty, unit_price, tax_percent)

    if not indent_id:
        return jsonify({"ok": False, "error": "Select an approved indent."}), 400
    if not selected:
        return jsonify({"ok": False, "error": "Select at least one item with a received quantity."}), 400

    conn = get_db()
    write_outlet = "bar"
    expenses: list[dict[str, Any]] = []
    try:
        ensure_stores_schema(conn)
        indent = conn.execute(
            "SELECT * FROM store_indents WHERE id = ?",
            (indent_id,),
        ).fetchone()
        if not indent or indent["status"] != "approved":
            return jsonify({"ok": False, "error": "Select an approved indent."}), 400
        write_outlet = _parse_outlet(indent["outlet"])
        lines = conn.execute(
            "SELECT * FROM store_indent_lines WHERE indent_id = ? ORDER BY id",
            (indent_id,),
        ).fetchall()
        if not lines:
            return jsonify({"ok": False, "error": "This indent has no items."}), 400

        po_lines_by_indent_line: dict[int, dict[str, Any]] = {}
        if purchase_order_id:
            po_row = conn.execute(
                """
                SELECT id, indent_id, supplier_id, po_no
                FROM store_purchase_orders
                WHERE id = ? AND indent_id = ?
                """,
                (purchase_order_id, indent_id),
            ).fetchone()
            if not po_row:
                return jsonify({"ok": False, "error": "Purchase order was not found for this indent."}), 400
            for pol in _load_purchase_order_lines(conn, purchase_order_id):
                try:
                    lid = int(pol.get("line_id") or 0)
                except (TypeError, ValueError):
                    lid = 0
                if lid > 0:
                    po_lines_by_indent_line[lid] = pol

        lines_by_id = {int(row["id"]): row for row in lines}
        group_input: list[dict[str, Any]] = []
        for line_id, (received_qty, unit_price, tax_percent) in selected.items():
            line = lines_by_id.get(line_id)
            if not line:
                return jsonify({"ok": False, "error": "One or more selected lines were not found."}), 400
            ordered = float(line["quantity"] or 0)
            try:
                already = float(line["quantity_received"] or 0)
            except (KeyError, TypeError, ValueError):
                already = 0.0
            remaining = ordered - already
            if remaining <= 0.0001:
                return jsonify({
                    "ok": False,
                    "error": f"{line['item_name']} is already fully received.",
                }), 400
            if purchase_order_id:
                pol = po_lines_by_indent_line.get(line_id)
                if not pol:
                    return jsonify({
                        "ok": False,
                        "error": f"{line['item_name']} is not on this purchase order.",
                    }), 400
                try:
                    po_qty = float(pol.get("quantity") or 0)
                    po_received = float(pol.get("quantity_received") or 0)
                except (TypeError, ValueError):
                    po_qty = 0.0
                    po_received = 0.0
                po_available = _po_inward_available_qty(
                    po_qty=po_qty,
                    po_received=po_received,
                    indent_remaining=remaining,
                )
                if po_available <= 0.0001:
                    return jsonify({
                        "ok": False,
                        "error": f"{line['item_name']} is already fully received on this purchase order.",
                    }), 400
                if received_qty - po_available > 0.0001:
                    return jsonify({
                        "ok": False,
                        "error": (
                            f"Received quantity for {line['item_name']} cannot exceed "
                            f"remaining PO qty ({po_available:g})."
                        ),
                    }), 400
            elif received_qty - remaining > 0.0001:
                return jsonify({
                    "ok": False,
                    "error": (
                        f"Received quantity for {line['item_name']} cannot exceed "
                        f"remaining qty ({remaining:g})."
                    ),
                }), 400
            unit_cost = _unit_cost_with_tax(unit_price, tax_percent)
            if unit_cost is None:
                # Fall back to approved indent price (ex-tax) when UI omits entered rate.
                unit_cost = _unit_cost_with_tax(line["approximate_price"], 0)
            price_for_amount = unit_price
            tax_for_amount = tax_percent
            if price_for_amount is None:
                try:
                    price_for_amount = float(line["approximate_price"] or 0)
                except (TypeError, ValueError):
                    price_for_amount = 0.0
                tax_for_amount = 0.0
            group_input.append({
                "item_name": line["item_name"],
                "qty": received_qty,
                "unit_price": price_for_amount,
                "tax_percent": tax_for_amount,
                "_line": line,
                "_unit_cost": unit_cost,
                "_entered_price": unit_price,
                "_received_qty": received_qty,
            })

        groups, group_error = _group_inward_lines_by_expense_category(
            conn,
            stores_outlet=indent["outlet"],
            lines=group_input,
        )
        if group_error:
            conn.rollback()
            return jsonify({"ok": False, "error": group_error}), 400

        expense_data = {
            "company": data.get("company") or app_module.DEFAULT_COMPANY,
            "location": app_module.OUTLET_HOTEL,
            "date": data.get("date") or date.today().isoformat(),
            "description": (data.get("description") or "").strip()
            or f"Stock inward {indent['indent_no']}",
            "amount": data.get("amount"),
            "payment_type": data.get("payment_type"),
            "transaction_id": data.get("transaction_id"),
            "invoice_number": data.get("invoice_number"),
            "supplier_id": data.get("supplier_id"),
        }
        expenses, expense_error = _create_inward_category_expenses(
            conn,
            user,
            base_expense_data=expense_data,
            groups=groups,
            description_suffix=str(indent["indent_no"] or ""),
        )
        if expense_error:
            conn.rollback()
            return jsonify({"ok": False, "error": expense_error}), 400

        payment_type = app_module._normalize_expense_payment_type(expense_data.get("payment_type"))
        # Credit ≤ that group's approved-rate subtotal auto-verifies that expense.
        if payment_type == app_module.EXPENSE_PAYMENT_CREDIT:
            for group in groups:
                approved_total = 0.0
                for payload in group["lines"]:
                    line = payload["_line"]
                    received_qty = float(payload["_received_qty"])
                    try:
                        approx = float(line["approximate_price"] or 0)
                    except (TypeError, ValueError):
                        approx = 0.0
                    approved_total += received_qty * approx
                approved_total = app_module.round_half_up(approved_total, 2)
                group_amount = app_module.parse_money(group["amount"])
                if group_amount - approved_total > 0.001:
                    continue
                verify_notes = (
                    f"Auto-verified from stock inward {indent['indent_no']}"
                    f" · {group.get('category_label') or group['category_key']}"
                )
                if notes:
                    verify_notes = f"{verify_notes}: {notes}"
                _, verify_error = app_module._auto_verify_expense(
                    conn,
                    expense_id=group["expense_id"],
                    supplier_id=expense_data["supplier_id"],
                    amount=group_amount,
                    company=expense_data["company"],
                    user=user,
                    notes=verify_notes,
                )
                if verify_error:
                    conn.rollback()
                    return jsonify({"ok": False, "error": verify_error}), 400

        movement_note = f"Stock inward from {indent['indent_no']}"
        if notes:
            movement_note = f"{movement_note}: {notes}"
        for group in groups:
            for payload in group["lines"]:
                line = payload["_line"]
                received_qty = float(payload["_received_qty"])
                unit_cost = payload["_unit_cost"]
                entered_price = payload["_entered_price"]
                stock_qty = _line_stock_qty_delta(line, received_qty)
                stock_unit_cost = _line_stock_unit_cost(line, unit_cost)
                _adjust_stock(
                    conn,
                    outlet=write_outlet,
                    item_name=line["item_name"],
                    unit=line["unit"] or "",
                    qty_delta=stock_qty,
                    movement_type="receive",
                    ref_type="stock_inward",
                    ref_id=indent_id,
                    notes=movement_note,
                    user_id=user["id"] if user else None,
                    unit_cost=stock_unit_cost,
                )
                master_price = entered_price
                if master_price is None:
                    try:
                        if line["approximate_price"] is not None and line["approximate_price"] != "":
                            master_price = float(line["approximate_price"])
                    except (KeyError, TypeError, ValueError):
                        master_price = None
                if master_price is None and unit_cost is not None:
                    try:
                        master_price = float(unit_cost)
                    except (TypeError, ValueError):
                        master_price = None
                _update_product_master_price_from_inward(
                    conn,
                    item_name=line["item_name"],
                    pack_label=_row_pack_label(line),
                    unit_price=master_price,
                )
                _update_product_preferred_suppliers_from_inward(
                    conn,
                    item_name=line["item_name"],
                    supplier_id=expense_data.get("supplier_id"),
                    unit_price=master_price,
                )
                try:
                    already = float(line["quantity_received"] or 0)
                except (KeyError, TypeError, ValueError):
                    already = 0.0
                conn.execute(
                    """
                    UPDATE store_indent_lines
                    SET quantity_received = ?
                    WHERE id = ?
                    """,
                    (already + float(received_qty), int(line["id"])),
                )
                if purchase_order_id:
                    _apply_po_line_received(
                        conn,
                        purchase_order_id=purchase_order_id,
                        indent_line_id=int(line["id"]),
                        received_qty=received_qty,
                    )

        # Refresh remaining after this confirm.
        remaining_rows = conn.execute(
            """
            SELECT COALESCE(quantity, 0) - COALESCE(quantity_received, 0) AS remaining
            FROM store_indent_lines
            WHERE indent_id = ?
            """,
            (indent_id,),
        ).fetchall()
        still_open = any(float(row["remaining"] or 0) > 0.0001 for row in remaining_rows)
        po_still_open = False
        if purchase_order_id:
            po_left = conn.execute(
                """
                SELECT 1
                FROM store_purchase_order_lines pol
                JOIN store_indent_lines l
                  ON l.id = pol.line_id AND l.indent_id = ?
                WHERE pol.purchase_order_id = ?
                  AND COALESCE(pol.quantity, 0) - COALESCE(pol.quantity_received, 0) > 0.0001
                  AND COALESCE(l.quantity, 0) - COALESCE(l.quantity_received, 0) > 0.0001
                LIMIT 1
                """,
                (indent_id, purchase_order_id),
            ).fetchone()
            po_still_open = bool(po_left)
        if still_open:
            # Keep approved so the indent stays on Stock Inward for other POs.
            conn.execute(
                "UPDATE store_indents SET status = 'approved' WHERE id = ?",
                (indent_id,),
            )
            redirect_kwargs: dict[str, Any] = {
                "outlet": write_outlet,
                "view": "approved",
            }
            if po_still_open:
                redirect_kwargs["po_id"] = purchase_order_id
                message = "Partial stock inward recorded. Remaining items stay on Stock Inward."
            else:
                message = (
                    "Stock inward recorded for this purchase order. "
                    "Remaining indent qty stays available for other POs."
                )
            redirect_url = url_for("stores_purchase_requests", **redirect_kwargs)
        else:
            conn.execute(
                "UPDATE store_indents SET status = 'stocked' WHERE id = ?",
                (indent_id,),
            )
            redirect_url = url_for("stores_stock", outlet=write_outlet)
            message = "Stock inward and expense recorded."

        conn.commit()
    except ValueError as exc:
        conn.rollback()
        return jsonify({"ok": False, "error": str(exc)}), 400
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return jsonify({
        "ok": True,
        "redirect": redirect_url,
        "expense_id": expenses[0]["expense_id"] if expenses else None,
        "expense_code": expenses[0].get("expense_code") if expenses else None,
        "expenses": expenses,
        "message": message,
        "partial": still_open,
    })


@stores_bp.route("/stores/purchase-requests/confirm-direct-with-expense", methods=["POST"])
def stores_confirm_direct_stock_inward_expense():
    """Confirm without-indent stock inward; expense always awaits Purchase Verification."""
    user = _get_user()
    if not user:
        return jsonify({"ok": False, "error": "You must be logged in."}), 401

    import app as app_module

    data = request.get_json(silent=True) or {}
    outlet_filter = _parse_outlet_filter(data.get("outlet"))
    if outlet_filter not in OUTLET_KEYS:
        return jsonify({"ok": False, "error": "Choose Bar or Restaurant."}), 400
    write_outlet = outlet_filter
    notes = (data.get("notes") or "").strip()[:500]
    raw_lines = data.get("lines") or []
    if not isinstance(raw_lines, list):
        raw_lines = []

    parsed_lines: list[dict[str, Any]] = []
    for raw in raw_lines:
        if not isinstance(raw, dict):
            continue
        item_name = str(raw.get("item_name") or "").strip()
        if not item_name:
            continue
        try:
            qty = float(raw.get("qty") or raw.get("received_qty") or raw.get("quantity") or 0)
        except (TypeError, ValueError):
            continue
        if qty <= 0:
            continue
        unit_price = None
        tax_percent = None
        try:
            if raw.get("unit_price") not in (None, ""):
                unit_price = float(raw.get("unit_price"))
        except (TypeError, ValueError):
            unit_price = None
        if unit_price is None or unit_price <= 0:
            continue
        try:
            if raw.get("tax_percent") not in (None, ""):
                tax_percent = float(raw.get("tax_percent"))
        except (TypeError, ValueError):
            tax_percent = None
        unit = str(raw.get("unit") or "").strip() or "kg"
        pack_label = str(raw.get("pack_label") or "").strip()
        pack_qty = None
        try:
            if raw.get("pack_qty_in_base") not in (None, ""):
                pack_qty = float(raw.get("pack_qty_in_base"))
                if pack_qty <= 0:
                    pack_qty = None
        except (TypeError, ValueError):
            pack_qty = None
        parsed_lines.append({
            "item_name": item_name,
            "qty": qty,
            "unit": unit,
            "unit_price": unit_price,
            "tax_percent": tax_percent,
            "pack_label": pack_label,
            "pack_qty_in_base": pack_qty,
        })

    if not parsed_lines:
        return jsonify({
            "ok": False,
            "error": "Add at least one product with quantity and price.",
        }), 400

    conn = get_db()
    expenses: list[dict[str, Any]] = []
    try:
        ensure_stores_schema(conn)
        allowed_names = _product_names_for_outlet(conn, write_outlet)
        for line in parsed_lines:
            if line["item_name"].casefold() not in allowed_names:
                return jsonify({
                    "ok": False,
                    "error": f"{line['item_name']} is not in Product Master for this outlet.",
                }), 400

        groups, group_error = _group_inward_lines_by_expense_category(
            conn,
            stores_outlet=write_outlet,
            lines=parsed_lines,
        )
        if group_error:
            conn.rollback()
            return jsonify({"ok": False, "error": group_error}), 400

        invoice_number = (data.get("invoice_number") or "").strip()
        description = (data.get("description") or "").strip()
        if not description:
            description = "Stock inward without indent approval"
            if invoice_number:
                description = f"{description} · Inv {invoice_number}"

        expense_data = {
            "company": data.get("company") or app_module.DEFAULT_COMPANY,
            "location": app_module.OUTLET_HOTEL,
            "date": data.get("date") or date.today().isoformat(),
            "description": description,
            "amount": data.get("amount"),
            "payment_type": data.get("payment_type"),
            "transaction_id": data.get("transaction_id"),
            "invoice_number": invoice_number,
            "supplier_id": data.get("supplier_id"),
        }
        expenses, expense_error = _create_inward_category_expenses(
            conn,
            user,
            base_expense_data=expense_data,
            groups=groups,
            description_suffix="without indent approval",
        )
        if expense_error:
            conn.rollback()
            return jsonify({"ok": False, "error": expense_error}), 400

        # Never auto-verify — without an approved indent, Purchase Verification is required.
        movement_note = "Stock inward without indent approval"
        if invoice_number:
            movement_note = f"{movement_note} · Inv {invoice_number}"
        if notes:
            movement_note = f"{movement_note}: {notes}"

        for group in groups:
            expense_id = int(group["expense_id"])
            for line in group["lines"]:
                unit_cost = _unit_cost_with_tax(line["unit_price"], line["tax_percent"])
                fake_line = {
                    "item_name": line["item_name"],
                    "unit": line["unit"],
                    "pack_label": line["pack_label"],
                    "pack_qty_in_base": line["pack_qty_in_base"],
                }
                stock_qty = _line_stock_qty_delta(fake_line, line["qty"])
                stock_unit_cost = _line_stock_unit_cost(fake_line, unit_cost)
                _adjust_stock(
                    conn,
                    outlet=write_outlet,
                    item_name=line["item_name"],
                    unit=line["unit"] or "",
                    qty_delta=stock_qty,
                    movement_type="receive",
                    ref_type="stock_inward_direct",
                    ref_id=expense_id,
                    notes=movement_note,
                    user_id=user["id"] if user else None,
                    unit_cost=stock_unit_cost,
                )
                _update_product_master_price_from_inward(
                    conn,
                    item_name=line["item_name"],
                    pack_label=line["pack_label"],
                    unit_price=line["unit_price"],
                )
                _update_product_preferred_suppliers_from_inward(
                    conn,
                    item_name=line["item_name"],
                    supplier_id=expense_data.get("supplier_id"),
                    unit_price=line["unit_price"],
                )

        conn.commit()
    except ValueError as exc:
        conn.rollback()
        return jsonify({"ok": False, "error": str(exc)}), 400
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return jsonify({
        "ok": True,
        "redirect": url_for("stores_stock", outlet=write_outlet),
        "expense_id": expenses[0]["expense_id"] if expenses else None,
        "expense_code": expenses[0].get("expense_code") if expenses else None,
        "expenses": expenses,
        "message": "Stock inward recorded. Expense awaits Purchase Verification.",
        "partial": False,
    })


@stores_bp.route("/stores/purchase-requests/<int:pr_id>")
def stores_pr_detail(pr_id: int):
    outlet = _parse_outlet(request.args.get("outlet"))
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        pr = conn.execute(
            """
            SELECT p.*, i.indent_no
            FROM store_purchase_requests p
            LEFT JOIN store_indents i ON i.id = p.indent_id
            WHERE p.id = ?
            """,
            (pr_id,),
        ).fetchone()
        if not pr:
            flash("Purchase request not found.", "error")
            return redirect(url_for("stores_purchase_requests", outlet=outlet))
        outlet = pr["outlet"]
        lines = conn.execute(
            "SELECT * FROM store_purchase_request_lines WHERE pr_id = ? ORDER BY id",
            (pr_id,),
        ).fetchall()
    finally:
        conn.close()
    return _page_render(
        "purchase_requests",
        outlet=outlet,
        detail=dict(pr),
        detail_lines=[dict(line) for line in lines],
    )


@stores_bp.route("/stores/purchase-requests/<int:pr_id>/receive", methods=["POST"])
def stores_pr_receive(pr_id: int):
    user = _get_user()
    outlet = _parse_outlet(request.form.get("outlet"))
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        pr = conn.execute("SELECT * FROM store_purchase_requests WHERE id = ?", (pr_id,)).fetchone()
        if not pr:
            flash("Purchase request not found.", "error")
            return redirect(url_for("stores_purchase_requests", outlet=outlet))
        outlet = pr["outlet"]
        if pr["status"] != "open":
            flash("This purchase request was already received.", "error")
            return redirect(url_for("stores_purchase_requests", outlet=outlet))
        lines = conn.execute(
            "SELECT * FROM store_purchase_request_lines WHERE pr_id = ? ORDER BY id",
            (pr_id,),
        ).fetchall()
        for line in lines:
            _adjust_stock(
                conn,
                outlet=outlet,
                item_name=line["item_name"],
                unit=line["unit"],
                qty_delta=_line_stock_qty_delta(line, float(line["quantity"])),
                movement_type="receive",
                ref_type="purchase_request",
                ref_id=pr_id,
                notes=f"Received from {pr['pr_no']}",
                user_id=user["id"] if user else None,
            )
        conn.execute(
            "UPDATE store_purchase_requests SET status = 'received', received_at = ? WHERE id = ?",
            (_now(), pr_id),
        )
        conn.commit()
    except ValueError as exc:
        conn.rollback()
        flash(str(exc), "error")
        return redirect(url_for("stores_purchase_requests", outlet=outlet))
    finally:
        conn.close()
    flash("Items received into stock.", "ok")
    return redirect(url_for("stores_stock", outlet=outlet))


def _inward_weighted_unit_costs(
    conn,
    outlet_sql: str,
    outlet_params: tuple[Any, ...],
) -> dict[tuple[str, str, str], float]:
    """Weighted-average unit cost (incl. tax when stored) from receive movements.

    Prefers ``store_stock_movements.unit_cost`` (inward price + tax). Falls back to
    the matching indent line ``approximate_price`` for older receipts without cost.
    """
    rows = conn.execute(
        f"""
        SELECT m.outlet,
               lower(m.item_name) AS name_key,
               lower(m.unit) AS unit_key,
               SUM(
                 m.qty_delta * COALESCE(
                   m.unit_cost,
                   (
                     SELECT l.approximate_price
                     FROM store_indent_lines l
                     WHERE m.ref_type = 'stock_inward'
                       AND l.indent_id = m.ref_id
                       AND lower(l.item_name) = lower(m.item_name)
                     ORDER BY l.id
                     LIMIT 1
                   )
                 )
               ) AS cost_total,
               SUM(
                 CASE
                   WHEN COALESCE(
                     m.unit_cost,
                     (
                       SELECT l.approximate_price
                       FROM store_indent_lines l
                       WHERE m.ref_type = 'stock_inward'
                         AND l.indent_id = m.ref_id
                         AND lower(l.item_name) = lower(m.item_name)
                       ORDER BY l.id
                       LIMIT 1
                     )
                   ) IS NOT NULL THEN m.qty_delta
                   ELSE 0
                 END
               ) AS qty_priced
        FROM store_stock_movements m
        WHERE m.movement_type = 'receive'
          AND m.qty_delta > 0
          AND {outlet_sql}
        GROUP BY m.outlet, lower(m.item_name), lower(m.unit)
        """,
        outlet_params,
    ).fetchall()
    costs: dict[tuple[str, str, str], float] = {}
    for row in rows:
        try:
            qty_priced = float(row["qty_priced"] or 0)
            cost_total = float(row["cost_total"] or 0)
        except (TypeError, ValueError):
            continue
        if qty_priced <= 0.0001 or cost_total <= 0:
            continue
        key = (
            (row["outlet"] or "").strip().lower(),
            (row["name_key"] or "").strip().lower(),
            (row["unit_key"] or "").strip().lower(),
        )
        costs[key] = round(cost_total / qty_priced, 4)
    return costs


def _enrich_stock_items(
    conn,
    items: list[dict[str, Any]],
    *,
    inward_costs: dict[tuple[str, str, str], float] | None = None,
) -> list[dict[str, Any]]:
    """Attach category + unit price for Stock display (inward WAC preferred)."""
    if not items:
        return items
    products = conn.execute(
        """
        SELECT p.name, p.outlet, p.default_unit, p.approximate_price,
               c.name AS category_name
        FROM store_products p
        LEFT JOIN store_product_categories c
          ON c.id = p.category_id AND c.is_active = 1
        WHERE p.is_active = 1
        """
    ).fetchall()
    by_name: dict[str, list[dict[str, Any]]] = {}
    for row in products:
        key = (row["name"] or "").strip().lower()
        if not key:
            continue
        by_name.setdefault(key, []).append(dict(row))

    costs = inward_costs or {}
    for item in items:
        name_key = (item.get("item_name") or "").strip().lower()
        unit_key = (item.get("unit") or "").strip().lower()
        outlet_key = (item.get("outlet") or "").strip().lower()
        candidates = by_name.get(name_key) or []
        match = None
        for preferred in (
            lambda p: (p.get("outlet") or "").strip().lower() == outlet_key
            and (p.get("default_unit") or "").strip().lower() == unit_key,
            lambda p: (p.get("outlet") or "").strip().lower() == outlet_key,
            lambda p: (p.get("outlet") or "").strip().lower() == "both"
            and (p.get("default_unit") or "").strip().lower() == unit_key,
            lambda p: (p.get("outlet") or "").strip().lower() == "both",
            lambda p: True,
        ):
            for cand in candidates:
                if preferred(cand):
                    match = cand
                    break
            if match:
                break
        if match:
            item["category_name"] = match.get("category_name") or ""
        else:
            item.setdefault("category_name", "")

        inward_price = costs.get((outlet_key, name_key, unit_key))
        if inward_price is not None and inward_price > 0:
            item["approximate_price"] = inward_price
            item["approximate_price_display"] = _format_optional_price(inward_price)
            item["price_source"] = "inward"
        elif match and match.get("approximate_price") is not None:
            price = match.get("approximate_price")
            item["approximate_price"] = price
            item["approximate_price_display"] = _format_optional_price(price)
            item["price_source"] = "product"
        else:
            item.setdefault("approximate_price", None)
            item.setdefault("approximate_price_display", "")
            item.setdefault("price_source", None)
    return items


STOCK_REPORT_LOW_THRESHOLD = 5.0


def _stock_item_status(qty: float, low_threshold: float = STOCK_REPORT_LOW_THRESHOLD) -> tuple[str, str]:
    if qty <= 0:
        return "out", "Out"
    if qty <= low_threshold:
        return "low", "Low"
    return "healthy", "Healthy"


def _load_stock_report_items(
    conn,
    outlet: str,
    *,
    category: str = "",
    status: str = "",
    q: str = "",
) -> list[dict[str, Any]]:
    outlet_sql, outlet_params = _outlet_match_sql("outlet", outlet)
    outlet_sql_m, outlet_params_m = _outlet_match_sql("m.outlet", outlet)
    items = conn.execute(
        f"""
        SELECT * FROM store_stock_items
        WHERE {outlet_sql}
        ORDER BY lower(item_name), lower(unit)
        """,
        outlet_params,
    ).fetchall()
    inward_costs = _inward_weighted_unit_costs(conn, outlet_sql_m, outlet_params_m)
    stock_items = _enrich_stock_items(
        conn,
        [dict(row) for row in items],
        inward_costs=inward_costs,
    )
    cat_filter = (category or "").strip().lower()
    if cat_filter in ("", "all"):
        cat_filter = ""
    status_filter = (status or "").strip().lower()
    if status_filter in ("", "all"):
        status_filter = ""
    needle = (q or "").strip().lower()
    out: list[dict[str, Any]] = []
    for item in stock_items:
        try:
            qty = float(item.get("qty_on_hand") or 0)
        except (TypeError, ValueError):
            qty = 0.0
        status_key, status_label = _stock_item_status(qty)
        category_name = (item.get("category_name") or "").strip()
        if cat_filter and category_name.lower() != cat_filter:
            continue
        if status_filter and status_key != status_filter:
            continue
        hay = " ".join(
            [
                str(item.get("item_name") or ""),
                str(item.get("unit") or ""),
                category_name,
                str(item.get("qty_on_hand") or ""),
                status_label,
            ]
        ).lower()
        if needle and needle not in hay:
            continue
        unit_price = item.get("approximate_price")
        try:
            unit_price_f = float(unit_price) if unit_price is not None else None
        except (TypeError, ValueError):
            unit_price_f = None
        line_value = (
            round(qty * unit_price_f, 2) if unit_price_f is not None else None
        )
        out.append(
            {
                "item_name": item.get("item_name") or "",
                "category_name": category_name or "Uncategorised",
                "qty_on_hand": round(qty, 3),
                "unit": item.get("unit") or "",
                "status": status_label,
                "unit_price": unit_price_f,
                "value": line_value,
                "outlet": item.get("outlet") or "",
                "outlet_label": _outlet_label(_normalize_outlet_key(item.get("outlet"))),
            }
        )
    return out


def _excel_number(value: Any) -> int | float | None:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if abs(number - round(number)) < 1e-9:
        return int(round(number))
    return round(number, 3)


def _build_stock_report_xlsx(rows: list[dict[str, Any]]) -> io.BytesIO:
    """Build Stock Report Excel matching the bordered table layout users expect."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Report"
    header_font = Font(name="Calibri", size=12, bold=True, color="000000")
    body_font = Font(name="Calibri", size=12, color="000000")
    # Light blue header (Excel theme accent ≈ tinted)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = [
        "Product",
        "Category",
        "On hand",
        "Unit",
        "Status",
        "Unit price",
        "Value",
        "Outlet",
    ]
    widths = (22, 18, 12, 10, 12, 12, 12, 14)
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]
    for idx, row in enumerate(rows, start=2):
        values = [
            row.get("item_name") or "",
            row.get("category_name") or "",
            _excel_number(row.get("qty_on_hand")),
            row.get("unit") or "",
            row.get("status") or "",
            _excel_number(row.get("unit_price")),
            _excel_number(row.get("value")),
            row.get("outlet_label") or row.get("outlet") or "",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col, value=value)
            cell.font = body_font
            cell.alignment = center
            cell.border = border
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:H{len(rows) + 1}"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _stock_export_filter_args(
    outlet: str,
    *,
    category: str = "",
    status: str = "",
    q: str = "",
) -> dict[str, str]:
    args: dict[str, str] = {"outlet": outlet or "both"}
    cat = (category or "").strip()
    if cat and cat.lower() != "all":
        args["category"] = cat
    st = (status or "").strip().lower()
    if st and st != "all":
        args["status"] = st
    needle = (q or "").strip()
    if needle:
        args["q"] = needle
    return args


@stores_bp.route("/stores/stock")
def stores_stock():
    outlet = _parse_outlet_filter(request.args.get("outlet"))
    outlet_sql, outlet_params = _outlet_match_sql("outlet", outlet)
    outlet_sql_m, outlet_params_m = _outlet_match_sql("m.outlet", outlet)
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        items = conn.execute(
            f"""
            SELECT * FROM store_stock_items
            WHERE {outlet_sql}
            ORDER BY lower(item_name), lower(unit)
            """,
            outlet_params,
        ).fetchall()
        inward_costs = _inward_weighted_unit_costs(conn, outlet_sql_m, outlet_params_m)
        stock_items = _enrich_stock_items(
            conn,
            [dict(row) for row in items],
            inward_costs=inward_costs,
        )
        movements = conn.execute(
            f"""
            SELECT m.*, u.full_name AS created_by_name
            FROM store_stock_movements m
            LEFT JOIN users u ON u.id = m.created_by
            WHERE {outlet_sql_m}
            ORDER BY m.created_at DESC, m.id DESC
            LIMIT 25
            """,
            outlet_params_m,
        ).fetchall()
    finally:
        conn.close()
    categories = sorted(
        {
            (item.get("category_name") or "").strip()
            for item in stock_items
            if (item.get("category_name") or "").strip()
        },
        key=lambda name: name.lower(),
    )
    has_prices = any(item.get("approximate_price") is not None for item in stock_items)
    has_inward_prices = any(item.get("price_source") == "inward" for item in stock_items)
    stock_export_url = url_for(
        "stores_stock_export", **_stock_export_filter_args(outlet)
    )
    return _page_render(
        "stock",
        outlet=outlet,
        stock_items=stock_items,
        stock_categories=categories,
        stock_has_prices=has_prices,
        stock_has_inward_prices=has_inward_prices,
        stock_export_url=stock_export_url,
        movements=[dict(row) for row in movements],
    )


@stores_bp.route("/stores/stock/export")
def stores_stock_export():
    """Excel download of current stock on hand (Stock Report)."""
    outlet = _parse_outlet_filter(request.args.get("outlet"))
    category = str(request.args.get("category") or "").strip()
    status = str(request.args.get("status") or "").strip().lower()
    q = str(request.args.get("q") or "").strip()
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        rows = _load_stock_report_items(
            conn,
            outlet,
            category=category,
            status=status,
            q=q,
        )
    finally:
        conn.close()
    buf = _build_stock_report_xlsx(rows)
    stamp = datetime.now().strftime("%Y%m%d")
    outlet_slug = re.sub(r"[^\w.-]+", "_", outlet or "all")
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"stock_report_{outlet_slug}_{stamp}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


STOCK_AUDIT_REASONS_NEGATIVE = (
    ("kitchen_wastage", "Kitchen Wastage"),
    ("spillage", "Spillage"),
    ("theft_loss", "Theft/Loss"),
    ("counting_error", "Counting Error"),
    ("other", "Other"),
)
STOCK_AUDIT_REASONS_POSITIVE = (
    ("supplier_excess_delivery", "Supplier Excess Delivery"),
    ("stock_found_during_audit", "Stock Found During Audit"),
    ("transfer_in_not_recorded", "Transfer In Not Recorded"),
    ("production_return", "Production Return"),
    ("recipe_underconsumption", "Recipe Underconsumption"),
    ("inventory_correction", "Inventory Correction"),
    ("other", "Other"),
)
STOCK_AUDIT_REASONS = tuple(
    dict(STOCK_AUDIT_REASONS_NEGATIVE + STOCK_AUDIT_REASONS_POSITIVE).items()
)
STOCK_AUDIT_REASON_KEYS = {key for key, _ in STOCK_AUDIT_REASONS}
STOCK_AUDIT_REASON_KEYS_NEGATIVE = {key for key, _ in STOCK_AUDIT_REASONS_NEGATIVE}
STOCK_AUDIT_REASON_KEYS_POSITIVE = {key for key, _ in STOCK_AUDIT_REASONS_POSITIVE}
STOCK_AUDIT_EPS = 0.0001
STOCK_AUDIT_VERIFY_TTL_DAYS = 7


def _parse_audit_ts(raw: Any) -> datetime | None:
    text = str(raw or "").strip()
    if not text:
        return None
    for fmt, size in (
        ("%Y-%m-%d %H:%M:%S", 19),
        ("%Y-%m-%dT%H:%M:%S", 19),
        ("%Y-%m-%d", 10),
    ):
        try:
            return datetime.strptime(text[:size], fmt)
        except ValueError:
            continue
    return None


def _current_audit_stock_qty(
    conn, outlet: str, stock_item_id: Any, item_name: str, unit: str
) -> float | None:
    try:
        sid = int(stock_item_id) if stock_item_id is not None else None
    except (TypeError, ValueError):
        sid = None
    if sid is not None:
        row = conn.execute(
            "SELECT qty_on_hand FROM store_stock_items WHERE id = ?",
            (sid,),
        ).fetchone()
        if row is not None:
            try:
                return float(row["qty_on_hand"] or 0)
            except (TypeError, ValueError):
                return 0.0
    row = conn.execute(
        """
        SELECT qty_on_hand FROM store_stock_items
        WHERE outlet = ? AND lower(item_name) = lower(?) AND lower(unit) = lower(?)
        """,
        (outlet, item_name or "", unit or "pcs"),
    ).fetchone()
    if row is None:
        return None
    try:
        return float(row["qty_on_hand"] or 0)
    except (TypeError, ValueError):
        return 0.0


def _audit_refresh_line_statuses(conn, audit_id: int, outlet: str) -> bool:
    """Convert skipped→pending and expire verified lines older than one week."""
    changed = False
    skipped = conn.execute(
        """
        UPDATE store_stock_audit_lines
        SET status = 'pending', verified_at = NULL, verified_by = NULL
        WHERE audit_id = ? AND lower(status) = 'skipped'
        """,
        (audit_id,),
    )
    if skipped.rowcount:
        changed = True

    cutoff = datetime.now() - timedelta(days=STOCK_AUDIT_VERIFY_TTL_DAYS)
    rows = conn.execute(
        """
        SELECT id, stock_item_id, item_name, unit, system_qty, verified_at
        FROM store_stock_audit_lines
        WHERE audit_id = ? AND lower(status) = 'verified'
        """,
        (audit_id,),
    ).fetchall()
    for row in rows:
        when = _parse_audit_ts(row["verified_at"])
        if when is not None and when > cutoff:
            continue
        live_qty = _current_audit_stock_qty(
            conn,
            outlet,
            row["stock_item_id"],
            row["item_name"] or "",
            row["unit"] or "pcs",
        )
        try:
            system_qty = (
                round(float(live_qty), 3)
                if live_qty is not None
                else round(float(row["system_qty"] or 0), 3)
            )
        except (TypeError, ValueError):
            system_qty = 0.0
        conn.execute(
            """
            UPDATE store_stock_audit_lines
            SET status = 'pending',
                system_qty = ?,
                actual_qty = NULL,
                variance_qty = NULL,
                variance_value = NULL,
                reason = '',
                remarks = '',
                verified_at = NULL,
                verified_by = NULL
            WHERE id = ?
            """,
            (system_qty, int(row["id"])),
        )
        changed = True
    return changed


def _audit_concrete_outlet(outlet: str) -> str:
    """Audits are per concrete outlet; All defaults to Restaurant."""
    key = (outlet or "").strip().lower()
    if key in OUTLET_KEYS:
        return key
    return "restaurant"


def _audit_week_label(when: datetime | None = None) -> str:
    dt = when or datetime.now()
    week_start = dt.date() - timedelta(days=dt.weekday())
    return f"Week of {week_start.day} {week_start.strftime('%b %Y')}"


def _audit_initials(name: str) -> str:
    parts = [p for p in str(name or "").strip().split() if p]
    if not parts:
        return "?"
    if len(parts) == 1:
        return parts[0][:2].upper()
    return (parts[0][0] + parts[-1][0]).upper()


def _audit_line_dict(row) -> dict[str, Any]:
    item = dict(row)
    try:
        system_qty = float(item.get("system_qty") or 0)
    except (TypeError, ValueError):
        system_qty = 0.0
    actual_raw = item.get("actual_qty")
    try:
        actual_qty = float(actual_raw) if actual_raw is not None else None
    except (TypeError, ValueError):
        actual_qty = None
    try:
        variance_qty = (
            float(item["variance_qty"])
            if item.get("variance_qty") is not None
            else None
        )
    except (TypeError, ValueError):
        variance_qty = None
    if variance_qty is None and actual_qty is not None:
        variance_qty = round(actual_qty - system_qty, 3)
    try:
        unit_cost = float(item["unit_cost"]) if item.get("unit_cost") is not None else None
    except (TypeError, ValueError):
        unit_cost = None
    variance_value = item.get("variance_value")
    if variance_value is None and variance_qty is not None and unit_cost is not None:
        variance_value = round(variance_qty * unit_cost, 2)
    status = (item.get("status") or "pending").strip().lower()
    item["system_qty"] = system_qty
    item["actual_qty"] = actual_qty
    item["variance_qty"] = variance_qty
    item["variance_value"] = variance_value
    item["unit_cost"] = unit_cost
    item["status"] = status
    item["initials"] = _audit_initials(item.get("item_name") or "")
    item["category_name"] = (item.get("category_name") or "").strip()
    return item


def _audit_kpis(lines: list[dict[str, Any]]) -> dict[str, Any]:
    total = len(lines)
    pending = sum(
        1 for line in lines if line.get("status") in ("pending", "skipped")
    )
    verified = sum(1 for line in lines if line.get("status") == "verified")
    skipped = 0
    with_variance = 0
    variance_value = 0.0
    for line in lines:
        if line.get("status") != "verified":
            continue
        vq = line.get("variance_qty")
        try:
            vq_f = float(vq) if vq is not None else 0.0
        except (TypeError, ValueError):
            vq_f = 0.0
        if abs(vq_f) > STOCK_AUDIT_EPS:
            with_variance += 1
        vv = line.get("variance_value")
        try:
            if vv is not None:
                variance_value += float(vv)
            elif line.get("unit_cost") is not None:
                variance_value += vq_f * float(line["unit_cost"])
        except (TypeError, ValueError):
            pass
    return {
        "total": total,
        "pending": pending,
        "verified": verified,
        "skipped": skipped,
        "with_variance": with_variance,
        "variance_value": round(variance_value, 2),
        "remaining": pending,
        "progress_pct": int(round((verified / total) * 100)) if total else 0,
    }


def _last_purchase_dates(
    conn, outlet: str, lines: list[dict[str, Any]]
) -> dict[tuple[str, str], str]:
    if not lines:
        return {}
    rows = conn.execute(
        """
        SELECT lower(item_name) AS name_key, lower(unit) AS unit_key,
               MAX(created_at) AS last_at
        FROM store_stock_movements
        WHERE outlet = ? AND movement_type = 'receive' AND qty_delta > 0
        GROUP BY lower(item_name), lower(unit)
        """,
        (outlet,),
    ).fetchall()
    return {
        ((row["name_key"] or "").strip(), (row["unit_key"] or "").strip()): row["last_at"]
        for row in rows
    }


def _seed_audit_lines(conn, audit_id: int, outlet: str) -> None:
    items = conn.execute(
        """
        SELECT * FROM store_stock_items
        WHERE outlet = ?
        ORDER BY lower(item_name), lower(unit)
        """,
        (outlet,),
    ).fetchall()
    stock_items = _enrich_stock_items(conn, [dict(row) for row in items])
    for item in stock_items:
        try:
            system_qty = float(item.get("qty_on_hand") or 0)
        except (TypeError, ValueError):
            system_qty = 0.0
        unit_cost = item.get("approximate_price")
        try:
            unit_cost = float(unit_cost) if unit_cost is not None else None
        except (TypeError, ValueError):
            unit_cost = None
        conn.execute(
            """
            INSERT INTO store_stock_audit_lines (
                audit_id, stock_item_id, item_name, unit, category_name,
                system_qty, status, unit_cost
            ) VALUES (?, ?, ?, ?, ?, ?, 'pending', ?)
            """,
            (
                audit_id,
                item.get("id"),
                item.get("item_name") or "",
                item.get("unit") or "pcs",
                item.get("category_name") or "",
                round(system_qty, 3),
                unit_cost,
            ),
        )


def _sync_audit_lines_from_stock(conn, audit_id: int, outlet: str) -> bool:
    """Add stock items that appeared after the open audit was seeded (e.g. new inwards).

    Also refreshes ``system_qty`` / category / unit cost for still-pending lines so
    recent receives show the live on-hand figure before verification.
    """
    existing = conn.execute(
        """
        SELECT id, stock_item_id, item_name, unit, status, system_qty
        FROM store_stock_audit_lines
        WHERE audit_id = ?
        """,
        (audit_id,),
    ).fetchall()
    by_stock_id: dict[int, Any] = {}
    by_name_unit: dict[tuple[str, str], Any] = {}
    for row in existing:
        try:
            sid = int(row["stock_item_id"]) if row["stock_item_id"] is not None else 0
        except (TypeError, ValueError):
            sid = 0
        if sid > 0:
            by_stock_id[sid] = row
        key = (
            str(row["item_name"] or "").strip().lower(),
            str(row["unit"] or "pcs").strip().lower(),
        )
        by_name_unit[key] = row

    items = conn.execute(
        """
        SELECT * FROM store_stock_items
        WHERE outlet = ?
        ORDER BY lower(item_name), lower(unit)
        """,
        (outlet,),
    ).fetchall()
    stock_items = _enrich_stock_items(conn, [dict(row) for row in items])
    changed = False
    for item in stock_items:
        try:
            stock_id = int(item.get("id") or 0)
        except (TypeError, ValueError):
            stock_id = 0
        name = str(item.get("item_name") or "").strip()
        unit = str(item.get("unit") or "pcs").strip() or "pcs"
        key = (name.lower(), unit.lower())
        try:
            system_qty = round(float(item.get("qty_on_hand") or 0), 3)
        except (TypeError, ValueError):
            system_qty = 0.0
        unit_cost = item.get("approximate_price")
        try:
            unit_cost = float(unit_cost) if unit_cost is not None else None
        except (TypeError, ValueError):
            unit_cost = None
        category_name = str(item.get("category_name") or "").strip()

        match = by_stock_id.get(stock_id) if stock_id > 0 else None
        if match is None:
            match = by_name_unit.get(key)
        if match is None:
            conn.execute(
                """
                INSERT INTO store_stock_audit_lines (
                    audit_id, stock_item_id, item_name, unit, category_name,
                    system_qty, status, unit_cost
                ) VALUES (?, ?, ?, ?, ?, ?, 'pending', ?)
                """,
                (
                    audit_id,
                    stock_id or None,
                    name,
                    unit,
                    category_name,
                    system_qty,
                    unit_cost,
                ),
            )
            changed = True
            continue

        status = str(match["status"] or "").strip().lower()
        if status not in ("pending", "skipped"):
            continue
        try:
            prev_qty = float(match["system_qty"] or 0)
        except (TypeError, ValueError):
            prev_qty = 0.0
        needs_qty = abs(prev_qty - system_qty) > STOCK_AUDIT_EPS
        try:
            prev_sid = int(match["stock_item_id"] or 0)
        except (TypeError, ValueError):
            prev_sid = 0
        needs_link = stock_id > 0 and prev_sid != stock_id
        if not needs_qty and not needs_link:
            continue
        conn.execute(
            """
            UPDATE store_stock_audit_lines
            SET stock_item_id = COALESCE(?, stock_item_id),
                category_name = CASE
                    WHEN ? != '' THEN ?
                    ELSE category_name
                END,
                system_qty = ?,
                unit_cost = COALESCE(?, unit_cost)
            WHERE id = ?
            """,
            (
                stock_id or None,
                category_name,
                category_name,
                system_qty,
                unit_cost,
                int(match["id"]),
            ),
        )
        changed = True
    return changed


def _get_or_create_open_audit(conn, outlet: str, user_id: int | None) -> dict[str, Any]:
    row = conn.execute(
        """
        SELECT * FROM store_stock_audits
        WHERE outlet = ? AND status = 'open'
        ORDER BY id DESC
        LIMIT 1
        """,
        (outlet,),
    ).fetchone()
    if row:
        return dict(row)
    label = _audit_week_label()
    cur = conn.execute(
        """
        INSERT INTO store_stock_audits (outlet, status, label, started_at, started_by)
        VALUES (?, 'open', ?, ?, ?)
        """,
        (outlet, label, _now(), user_id),
    )
    audit_id = int(cur.lastrowid)
    _seed_audit_lines(conn, audit_id, outlet)
    conn.commit()
    created = conn.execute(
        "SELECT * FROM store_stock_audits WHERE id = ?", (audit_id,)
    ).fetchone()
    return dict(created)


def _load_audit_lines(conn, audit_id: int) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT * FROM store_stock_audit_lines
        WHERE audit_id = ?
        ORDER BY id ASC
        """,
        (audit_id,),
    ).fetchall()
    return [_audit_line_dict(row) for row in rows]


def _serialize_audit_line(line: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": line.get("id"),
        "item_name": line.get("item_name") or "",
        "unit": line.get("unit") or "",
        "category_name": line.get("category_name") or "",
        "system_qty": line.get("system_qty"),
        "actual_qty": line.get("actual_qty"),
        "variance_qty": line.get("variance_qty"),
        "variance_value": line.get("variance_value"),
        "status": line.get("status") or "pending",
        "reason": line.get("reason") or "",
        "remarks": line.get("remarks") or "",
        "unit_cost": line.get("unit_cost"),
        "initials": line.get("initials") or "?",
        "last_purchase_at": line.get("last_purchase_at") or "",
        "stock_updated_at": line.get("stock_updated_at") or "",
    }


@stores_bp.route("/stores/stock-audit")
def stores_stock_audit():
    filter_outlet = _parse_outlet_filter(request.args.get("outlet"))
    outlet = _audit_concrete_outlet(filter_outlet)
    line_id_raw = request.args.get("line_id")
    try:
        selected_line_id = int(line_id_raw) if line_id_raw else None
    except (TypeError, ValueError):
        selected_line_id = None
    user = _get_user() if _get_user else None
    user_id = user.get("id") if user else None
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        audit = _get_or_create_open_audit(conn, outlet, user_id)
        changed = False
        if _audit_refresh_line_statuses(conn, int(audit["id"]), outlet):
            changed = True
        if _sync_audit_lines_from_stock(conn, int(audit["id"]), outlet):
            changed = True
        if changed:
            conn.commit()
        lines = _load_audit_lines(conn, audit["id"])
        if not lines:
            _seed_audit_lines(conn, audit["id"], outlet)
            conn.commit()
            lines = _load_audit_lines(conn, audit["id"])
        purchases = _last_purchase_dates(conn, outlet, lines)
        stock_meta = {
            int(row["id"]): row["updated_at"]
            for row in conn.execute(
                "SELECT id, updated_at FROM store_stock_items WHERE outlet = ?",
                (outlet,),
            ).fetchall()
            if row["id"] is not None
        }
        for line in lines:
            name_key = (line.get("item_name") or "").strip().lower()
            unit_key = (line.get("unit") or "").strip().lower()
            line["last_purchase_at"] = purchases.get((name_key, unit_key), "") or ""
            sid = line.get("stock_item_id")
            try:
                sid_i = int(sid) if sid is not None else None
            except (TypeError, ValueError):
                sid_i = None
            line["stock_updated_at"] = stock_meta.get(sid_i, "") if sid_i else ""
        kpis = _audit_kpis(lines)
        selected = None
        if selected_line_id:
            for line in lines:
                if int(line.get("id") or 0) == selected_line_id:
                    selected = line
                    break
        if selected is None:
            for line in lines:
                if line.get("status") == "pending":
                    selected = line
                    break
        if selected is None and lines:
            selected = lines[0]
        selected_index = 0
        if selected:
            for idx, line in enumerate(lines):
                if int(line.get("id") or 0) == int(selected.get("id") or 0):
                    selected_index = idx
                    break
        history = conn.execute(
            """
            SELECT a.*, u.full_name AS started_by_name,
                   (SELECT COUNT(*) FROM store_stock_audit_lines l WHERE l.audit_id = a.id) AS line_count,
                   (SELECT COUNT(*) FROM store_stock_audit_lines l
                    WHERE l.audit_id = a.id AND l.status = 'verified') AS verified_count
            FROM store_stock_audits a
            LEFT JOIN users u ON u.id = a.started_by
            WHERE a.outlet = ? AND a.status = 'completed'
            ORDER BY a.completed_at DESC, a.id DESC
            LIMIT 20
            """,
            (outlet,),
        ).fetchall()
    finally:
        conn.close()
    audit_categories = sorted(
        {
            ((line.get("category_name") or "").strip() or "Uncategorised")
            for line in lines
        },
        key=lambda name: name.lower(),
    )
    return _page_render(
        "stock_audit",
        outlet=outlet,
        audit=audit,
        audit_lines=lines,
        audit_kpis=kpis,
        audit_selected=selected,
        audit_selected_index=selected_index,
        audit_reasons=STOCK_AUDIT_REASONS,
        audit_reasons_negative=STOCK_AUDIT_REASONS_NEGATIVE,
        audit_reasons_positive=STOCK_AUDIT_REASONS_POSITIVE,
        audit_categories=audit_categories,
        audit_history=[dict(row) for row in history],
        audit_verify_url=url_for("stores_stock_audit_verify"),
        audit_skip_url=url_for("stores_stock_audit_skip"),
        audit_new_url=url_for("stores_stock_audit_new"),
        audit_history_url=url_for("stores_stock_audit_history"),
    )


@stores_bp.route("/stores/stock-audit/verify", methods=["POST"])
def stores_stock_audit_verify():
    user = _get_user() if _get_user else None
    if not user:
        return jsonify({"ok": False, "error": "You must be logged in."}), 401
    data = request.get_json(silent=True) or request.form
    try:
        line_id = int(data.get("line_id"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Select an audit line."}), 400
    try:
        actual_qty = float(data.get("actual_qty"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Enter a valid actual count."}), 400
    if actual_qty < 0:
        return jsonify({"ok": False, "error": "Actual count cannot be negative."}), 400
    reason = str(data.get("reason") or "").strip().lower()
    remarks = str(data.get("remarks") or "").strip()[:200]
    go_next = str(data.get("go_next") or "").strip().lower() in ("1", "true", "yes")
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        line_row = conn.execute(
            """
            SELECT l.*, a.outlet, a.status AS audit_status, a.id AS audit_pk
            FROM store_stock_audit_lines l
            JOIN store_stock_audits a ON a.id = l.audit_id
            WHERE l.id = ?
            """,
            (line_id,),
        ).fetchone()
        if not line_row:
            return jsonify({"ok": False, "error": "Audit line not found."}), 404
        if (line_row["audit_status"] or "") != "open":
            return jsonify({"ok": False, "error": "This audit is already completed."}), 400
        if (line_row["status"] or "") == "verified":
            return jsonify({"ok": False, "error": "This item is already verified."}), 400
        system_qty = float(line_row["system_qty"] or 0)
        variance = round(actual_qty - system_qty, 3)
        if abs(variance) > STOCK_AUDIT_EPS:
            allowed = (
                STOCK_AUDIT_REASON_KEYS_POSITIVE
                if variance > 0
                else STOCK_AUDIT_REASON_KEYS_NEGATIVE
            )
            if reason not in allowed:
                return jsonify(
                    {"ok": False, "error": "Please select reason to save"}
                ), 400
        if abs(variance) <= STOCK_AUDIT_EPS:
            reason = reason if reason in STOCK_AUDIT_REASON_KEYS else ""
        unit_cost = line_row["unit_cost"]
        try:
            unit_cost_f = float(unit_cost) if unit_cost is not None else None
        except (TypeError, ValueError):
            unit_cost_f = None
        variance_value = (
            round(variance * unit_cost_f, 2) if unit_cost_f is not None else None
        )
        notes = reason.replace("_", " ").title() if reason else "Stock audit"
        if remarks:
            notes = f"{notes}: {remarks}" if reason else remarks
        # Reconcile live stock to the counted quantity (not only snapshot delta).
        stock_row = None
        try:
            sid = int(line_row["stock_item_id"]) if line_row["stock_item_id"] is not None else None
        except (TypeError, ValueError):
            sid = None
        if sid is not None:
            stock_row = conn.execute(
                "SELECT id, outlet, qty_on_hand FROM store_stock_items WHERE id = ?",
                (sid,),
            ).fetchone()
        if stock_row is None:
            stock_row = conn.execute(
                """
                SELECT id, outlet, qty_on_hand FROM store_stock_items
                WHERE outlet = ? AND lower(item_name) = lower(?) AND lower(unit) = lower(?)
                """,
                (line_row["outlet"], line_row["item_name"], line_row["unit"]),
            ).fetchone()
        current_qty = float(stock_row["qty_on_hand"] or 0) if stock_row else 0.0
        stock_outlet = (stock_row["outlet"] if stock_row else None) or line_row["outlet"]
        live_delta = round(actual_qty - current_qty, 3)
        adjusted = False
        if abs(live_delta) > STOCK_AUDIT_EPS or stock_row is None:
            applied = _adjust_stock(
                conn,
                outlet=stock_outlet,
                item_name=line_row["item_name"],
                unit=line_row["unit"],
                qty_delta=live_delta if stock_row is not None else round(actual_qty, 3),
                movement_type="adjustment",
                ref_type="stock_audit",
                ref_id=int(line_row["audit_pk"]),
                notes=notes,
                user_id=user.get("id"),
                unit_cost=unit_cost_f,
                allow_shortfall=True,
            )
            adjusted = abs(float(applied or 0)) > STOCK_AUDIT_EPS
        conn.execute(
            """
            UPDATE store_stock_audit_lines
            SET actual_qty = ?, variance_qty = ?, variance_value = ?,
                status = 'verified', reason = ?, remarks = ?,
                verified_at = ?, verified_by = ?
            WHERE id = ?
            """,
            (
                round(actual_qty, 3),
                variance,
                variance_value,
                reason,
                remarks,
                _now(),
                user.get("id"),
                line_id,
            ),
        )
        conn.commit()
        lines = _load_audit_lines(conn, int(line_row["audit_pk"]))
        kpis = _audit_kpis(lines)
        next_line = None
        if go_next:
            for line in lines:
                if line.get("status") == "pending":
                    next_line = line
                    break
        current = None
        for line in lines:
            if int(line.get("id") or 0) == line_id:
                current = line
                break
        serialized = _serialize_audit_line(current or {})
        serialized["qty_on_hand"] = round(actual_qty, 3)
        return jsonify(
            {
                "ok": True,
                "line": serialized,
                "next_line_id": next_line.get("id") if next_line else None,
                "kpis": kpis,
                "message": "Stock verified"
                + (" and adjusted." if adjusted or abs(variance) > STOCK_AUDIT_EPS else "."),
            }
        )
    except ValueError as exc:
        conn.rollback()
        return jsonify({"ok": False, "error": str(exc)}), 400
    except Exception as exc:
        conn.rollback()
        logger.exception("stock audit verify failed")
        return jsonify({"ok": False, "error": str(exc) or "Could not verify."}), 500
    finally:
        conn.close()


@stores_bp.route("/stores/stock-audit/skip", methods=["POST"])
def stores_stock_audit_skip():
    user = _get_user() if _get_user else None
    if not user:
        return jsonify({"ok": False, "error": "You must be logged in."}), 401
    data = request.get_json(silent=True) or request.form
    try:
        line_id = int(data.get("line_id"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Select an audit line."}), 400
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        line_row = conn.execute(
            """
            SELECT l.*, a.status AS audit_status, a.id AS audit_pk
            FROM store_stock_audit_lines l
            JOIN store_stock_audits a ON a.id = l.audit_id
            WHERE l.id = ?
            """,
            (line_id,),
        ).fetchone()
        if not line_row:
            return jsonify({"ok": False, "error": "Audit line not found."}), 404
        if (line_row["audit_status"] or "") != "open":
            return jsonify({"ok": False, "error": "This audit is already completed."}), 400
        # Skip for Now leaves the item pending so it stays in the queue.
        conn.execute(
            """
            UPDATE store_stock_audit_lines
            SET status = 'pending', verified_at = NULL, verified_by = NULL
            WHERE id = ?
            """,
            (line_id,),
        )
        conn.commit()
        lines = _load_audit_lines(conn, int(line_row["audit_pk"]))
        kpis = _audit_kpis(lines)
        next_line = None
        for line in lines:
            if line.get("status") == "pending" and int(line.get("id") or 0) != line_id:
                next_line = line
                break
        return jsonify(
            {
                "ok": True,
                "kpis": kpis,
                "next_line_id": next_line.get("id") if next_line else None,
                "message": "Left as pending.",
            }
        )
    finally:
        conn.close()


@stores_bp.route("/stores/stock-audit/history")
def stores_stock_audit_history():
    filter_outlet = _parse_outlet_filter(request.args.get("outlet"))
    outlet = _audit_concrete_outlet(filter_outlet)
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        rows = conn.execute(
            """
            SELECT a.*, u.full_name AS started_by_name,
                   (SELECT COUNT(*) FROM store_stock_audit_lines l WHERE l.audit_id = a.id) AS line_count,
                   (SELECT COUNT(*) FROM store_stock_audit_lines l
                    WHERE l.audit_id = a.id AND l.status = 'verified') AS verified_count
            FROM store_stock_audits a
            LEFT JOIN users u ON u.id = a.started_by
            WHERE a.outlet = ? AND a.status = 'completed'
            ORDER BY a.completed_at DESC, a.id DESC
            LIMIT 50
            """,
            (outlet,),
        ).fetchall()
    finally:
        conn.close()
    history = []
    for row in rows:
        item = dict(row)
        history.append(
            {
                "id": item.get("id"),
                "label": item.get("label") or "",
                "started_at": item.get("started_at") or "",
                "completed_at": item.get("completed_at") or "",
                "started_by_name": item.get("started_by_name") or "",
                "line_count": int(item.get("line_count") or 0),
                "verified_count": int(item.get("verified_count") or 0),
            }
        )
    return jsonify({"ok": True, "outlet": outlet, "history": history})


@stores_bp.route("/stores/stock-audit/new", methods=["POST"])
def stores_stock_audit_new():
    user = _get_user() if _get_user else None
    if not user:
        return jsonify({"ok": False, "error": "You must be logged in."}), 401
    data = request.get_json(silent=True) or request.form
    filter_outlet = _parse_outlet_filter(
        data.get("outlet") if data.get("outlet") is not None else request.args.get("outlet")
    )
    outlet = _audit_concrete_outlet(filter_outlet)
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        open_row = conn.execute(
            """
            SELECT id FROM store_stock_audits
            WHERE outlet = ? AND status = 'open'
            ORDER BY id DESC LIMIT 1
            """,
            (outlet,),
        ).fetchone()
        if open_row:
            conn.execute(
                """
                UPDATE store_stock_audits
                SET status = 'completed', completed_at = ?
                WHERE id = ?
                """,
                (_now(), open_row["id"]),
            )
        audit = _get_or_create_open_audit(conn, outlet, user.get("id"))
        if _sync_audit_lines_from_stock(conn, int(audit["id"]), outlet):
            conn.commit()
        lines = _load_audit_lines(conn, audit["id"])
        if not lines:
            _seed_audit_lines(conn, audit["id"], outlet)
            conn.commit()
            lines = _load_audit_lines(conn, audit["id"])
        return jsonify(
            {
                "ok": True,
                "audit_id": audit["id"],
                "redirect": url_for("stores_stock_audit", outlet=outlet),
                "line_count": len(lines),
            }
        )
    except Exception as exc:
        conn.rollback()
        logger.exception("stock audit new failed")
        return jsonify({"ok": False, "error": str(exc) or "Could not start audit."}), 500
    finally:
        conn.close()


def _parse_report_date(raw: Any) -> date | None:
    text = str(raw or "").strip()
    if not text:
        return None
    try:
        return datetime.strptime(text[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _audit_reason_label(reason_key: str) -> str:
    key = str(reason_key or "").strip().lower()
    for item_key, label in STOCK_AUDIT_REASONS:
        if item_key == key:
            return label
    if not key:
        return ""
    return key.replace("_", " ").title()


def _load_stock_audit_adjustment_rows(
    conn,
    *,
    outlet: str,
    date_from: date | None = None,
    date_to: date | None = None,
    reason: str = "",
    category: str = "",
    q: str = "",
) -> list[dict[str, Any]]:
    outlet_sql, outlet_params = _outlet_match_sql("a.outlet", outlet)
    clauses = [
        outlet_sql,
        "lower(coalesce(l.status, '')) = 'verified'",
        f"abs(coalesce(l.variance_qty, 0)) > {STOCK_AUDIT_EPS}",
    ]
    params: list[Any] = list(outlet_params)
    if date_from is not None:
        clauses.append("date(l.verified_at) >= date(?)")
        params.append(date_from.isoformat())
    if date_to is not None:
        clauses.append("date(l.verified_at) <= date(?)")
        params.append(date_to.isoformat())
    reason_key = str(reason or "").strip().lower()
    if reason_key and reason_key != "all":
        clauses.append("lower(coalesce(l.reason, '')) = ?")
        params.append(reason_key)
    category_key = str(category or "").strip().lower()
    if category_key and category_key != "all":
        if category_key == "uncategorised":
            clauses.append(
                "(trim(coalesce(l.category_name, '')) = '' OR lower(trim(l.category_name)) = 'uncategorised')"
            )
        else:
            clauses.append("lower(trim(coalesce(l.category_name, ''))) = ?")
            params.append(category_key)
    needle = str(q or "").strip().lower()
    if needle:
        clauses.append(
            """
            (
              lower(coalesce(l.item_name, '')) LIKE ?
              OR lower(coalesce(l.category_name, '')) LIKE ?
              OR lower(coalesce(l.unit, '')) LIKE ?
              OR lower(coalesce(l.reason, '')) LIKE ?
              OR lower(coalesce(l.remarks, '')) LIKE ?
              OR lower(coalesce(a.label, '')) LIKE ?
            )
            """
        )
        like = f"%{needle}%"
        params.extend([like, like, like, like, like, like])

    rows = conn.execute(
        f"""
        SELECT l.id, l.item_name, l.unit, l.category_name, l.system_qty, l.actual_qty,
               l.variance_qty, l.variance_value, l.reason, l.remarks, l.unit_cost,
               l.verified_at, l.verified_by,
               a.id AS audit_id, a.outlet, a.label AS audit_label, a.status AS audit_status,
               u.full_name AS verified_by_name
        FROM store_stock_audit_lines l
        JOIN store_stock_audits a ON a.id = l.audit_id
        LEFT JOIN users u ON u.id = l.verified_by
        WHERE {' AND '.join(clauses)}
        ORDER BY coalesce(l.verified_at, '') DESC, l.id DESC
        """,
        tuple(params),
    ).fetchall()

    out: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        try:
            system_qty = float(item.get("system_qty") or 0)
        except (TypeError, ValueError):
            system_qty = 0.0
        try:
            actual_qty = float(item.get("actual_qty") or 0)
        except (TypeError, ValueError):
            actual_qty = 0.0
        try:
            variance_qty = float(item.get("variance_qty") or 0)
        except (TypeError, ValueError):
            variance_qty = round(actual_qty - system_qty, 3)
        try:
            variance_value = (
                float(item["variance_value"])
                if item.get("variance_value") is not None
                else None
            )
        except (TypeError, ValueError):
            variance_value = None
        reason_key = str(item.get("reason") or "").strip().lower()
        out.append(
            {
                "id": item.get("id"),
                "item_name": item.get("item_name") or "",
                "unit": item.get("unit") or "",
                "category_name": (item.get("category_name") or "").strip() or "Uncategorised",
                "system_qty": round(system_qty, 3),
                "actual_qty": round(actual_qty, 3),
                "variance_qty": round(variance_qty, 3),
                "variance_value": (
                    round(variance_value, 2) if variance_value is not None else None
                ),
                "reason": reason_key,
                "reason_label": _audit_reason_label(reason_key),
                "remarks": (item.get("remarks") or "").strip(),
                "unit_cost": item.get("unit_cost"),
                "verified_at": item.get("verified_at") or "",
                "verified_by_name": item.get("verified_by_name") or "",
                "outlet": item.get("outlet") or "",
                "outlet_label": _outlet_label(_normalize_outlet_key(item.get("outlet"))),
                "audit_id": item.get("audit_id"),
                "audit_label": item.get("audit_label") or "",
                "audit_status": item.get("audit_status") or "",
            }
        )
    return out


def _stock_audit_adjustment_kpis(rows: list[dict[str, Any]]) -> dict[str, Any]:
    count = len(rows)
    positive = 0
    negative = 0
    value_sum = 0.0
    value_known = False
    for row in rows:
        try:
            vq = float(row.get("variance_qty") or 0)
        except (TypeError, ValueError):
            vq = 0.0
        if vq > STOCK_AUDIT_EPS:
            positive += 1
        elif vq < -STOCK_AUDIT_EPS:
            negative += 1
        vv = row.get("variance_value")
        if vv is not None:
            try:
                value_sum += float(vv)
                value_known = True
            except (TypeError, ValueError):
                pass
    return {
        "count": count,
        "positive": positive,
        "negative": negative,
        "variance_value": round(value_sum, 2) if value_known else None,
    }


def _parse_stock_audit_report_filters() -> dict[str, Any]:
    outlet = _parse_outlet_filter(request.args.get("outlet"))
    date_from = _parse_report_date(request.args.get("date_from"))
    date_to = _parse_report_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from
    reason = str(request.args.get("reason") or "").strip().lower()
    if reason == "all":
        reason = ""
    category = str(request.args.get("category") or "").strip()
    if category.lower() == "all":
        category = ""
    q = str(request.args.get("q") or "").strip()
    return {
        "outlet": outlet,
        "date_from": date_from,
        "date_to": date_to,
        "reason": reason,
        "category": category,
        "q": q,
    }


def _stock_audit_report_filter_args(filters: dict[str, Any]) -> dict[str, str]:
    args: dict[str, str] = {"outlet": filters.get("outlet") or "both"}
    if filters.get("date_from"):
        args["date_from"] = filters["date_from"].isoformat()
    if filters.get("date_to"):
        args["date_to"] = filters["date_to"].isoformat()
    if filters.get("reason"):
        args["reason"] = filters["reason"]
    if filters.get("category"):
        args["category"] = filters["category"]
    if filters.get("q"):
        args["q"] = filters["q"]
    return args


def _build_stock_audit_report_xlsx(rows: list[dict[str, Any]]) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Audit"
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    ws["A1"] = "Hotel Bell Elite — Stock Audit Report"
    ws["A1"].font = title_font
    ws["A2"] = f"Generated {_now()}"
    headers = (
        "Verified at",
        "Outlet",
        "Audit",
        "Product",
        "Category",
        "Unit",
        "System qty",
        "Actual qty",
        "Variance qty",
        "Variance value",
        "Reason",
        "Remarks",
        "Verified by",
    )
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=title)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if col >= 7 else "left")
    for idx, row in enumerate(rows, start=5):
        ws.cell(row=idx, column=1, value=row.get("verified_at") or "")
        ws.cell(row=idx, column=2, value=row.get("outlet_label") or "")
        ws.cell(row=idx, column=3, value=row.get("audit_label") or "")
        ws.cell(row=idx, column=4, value=row.get("item_name") or "")
        ws.cell(row=idx, column=5, value=row.get("category_name") or "")
        ws.cell(row=idx, column=6, value=row.get("unit") or "")
        ws.cell(row=idx, column=7, value=row.get("system_qty"))
        ws.cell(row=idx, column=8, value=row.get("actual_qty"))
        ws.cell(row=idx, column=9, value=row.get("variance_qty"))
        vv = row.get("variance_value")
        ws.cell(row=idx, column=10, value=vv if vv is not None else "")
        ws.cell(row=idx, column=11, value=row.get("reason_label") or "")
        ws.cell(row=idx, column=12, value=row.get("remarks") or "")
        ws.cell(row=idx, column=13, value=row.get("verified_by_name") or "")
    from openpyxl.utils import get_column_letter

    widths = (18, 12, 18, 22, 16, 8, 12, 12, 12, 14, 22, 24, 18)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@stores_bp.route("/stores/stock-audit/report")
def stores_stock_audit_report():
    filters = _parse_stock_audit_report_filters()
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        rows = _load_stock_audit_adjustment_rows(
            conn,
            outlet=filters["outlet"],
            date_from=filters["date_from"],
            date_to=filters["date_to"],
            reason=filters["reason"],
            category=filters["category"],
            q=filters["q"],
        )
        cat_rows = conn.execute(
            """
            SELECT DISTINCT trim(coalesce(l.category_name, '')) AS cat
            FROM store_stock_audit_lines l
            JOIN store_stock_audits a ON a.id = l.audit_id
            WHERE lower(coalesce(l.status, '')) = 'verified'
              AND abs(coalesce(l.variance_qty, 0)) > ?
            ORDER BY 1
            """,
            (STOCK_AUDIT_EPS,),
        ).fetchall()
        categories = sorted(
            {((row["cat"] or "").strip() or "Uncategorised") for row in cat_rows},
            key=lambda name: name.lower(),
        )
    finally:
        conn.close()
    kpis = _stock_audit_adjustment_kpis(rows)
    filter_args = _stock_audit_report_filter_args(filters)
    return _page_render(
        "stock_audit_report",
        outlet=filters["outlet"],
        audit_report_rows=rows,
        audit_report_kpis=kpis,
        audit_report_reasons=STOCK_AUDIT_REASONS,
        audit_report_categories=categories,
        audit_report_today_iso=date.today().isoformat(),
        audit_report_filters={
            "date_from": filters["date_from"].isoformat() if filters["date_from"] else "",
            "date_to": filters["date_to"].isoformat() if filters["date_to"] else "",
            "reason": filters["reason"] or "all",
            "category": filters["category"] or "all",
            "q": filters["q"] or "",
        },
        audit_report_export_url=url_for(
            "stores_stock_audit_report_export", **filter_args
        ),
    )


@stores_bp.route("/stores/stock-audit/report/export")
def stores_stock_audit_report_export():
    filters = _parse_stock_audit_report_filters()
    conn = get_db()
    try:
        ensure_stores_schema(conn)
        rows = _load_stock_audit_adjustment_rows(
            conn,
            outlet=filters["outlet"],
            date_from=filters["date_from"],
            date_to=filters["date_to"],
            reason=filters["reason"],
            category=filters["category"],
            q=filters["q"],
        )
    finally:
        conn.close()
    buf = _build_stock_audit_report_xlsx(rows)
    stamp = datetime.now().strftime("%Y%m%d")
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"stock_audit_report_{stamp}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def register_stores(app, *, pop_auth_notice, get_user):
    _bind_helpers(pop_auth_notice, get_user)
    app.register_blueprint(stores_bp)
    app.jinja_env.filters["stores_dt"] = _format_stores_dt
    for rule in list(app.url_map.iter_rules()):
        if not rule.endpoint.startswith("stores."):
            continue
        bare = rule.endpoint.split(".", 1)[1]
        if bare in app.view_functions:
            continue
        app.add_url_rule(
            rule.rule,
            endpoint=bare,
            view_func=app.view_functions[rule.endpoint],
            methods=rule.methods,
        )
