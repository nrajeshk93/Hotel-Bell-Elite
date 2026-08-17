"""Excel template + import for bulk POS menu items."""

from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from db import (
    POS_OUTLET_BAR,
    POS_OUTLET_RESTAURANT,
    list_pos_menu_categories,
    list_pos_menu_items,
    normalize_pos_outlet,
    save_pos_menu_item,
)

BLANK_ROWS = 200
HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
TITLE_FONT = Font(bold=True, name="Calibri", size=14)
BODY_FONT = Font(name="Calibri", size=11)
REQUIRED_FILL = PatternFill("solid", fgColor="DBEAFE")

MENU_TYPE_LABELS = ("Veg", "Non-Veg", "Liquor")
ITEM_TYPE_LABELS = ("Food", "Liquor")
OUTLET_LABELS = ("Restaurant", "Bar")
PLACEHOLDER_CATEGORY_NAMES = frozenset({"test", "tes"})

UNIFIED_HEADERS = (
    "Outlet *",
    "Menu *",
    "Category *",
    "Menu Type",
    "Rate (₹) *",
    "Code",
)


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("₹", "rs")
    return re.sub(r"[^a-z0-9]+", "", text)


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()


def _is_bar(outlet: str) -> bool:
    return normalize_pos_outlet(outlet) == POS_OUTLET_BAR


def _is_placeholder_category(name: str) -> bool:
    return (name or "").strip().casefold() in PLACEHOLDER_CATEGORY_NAMES


def _outlet_label(outlet: str) -> str:
    return "Bar" if _is_bar(outlet) else "Restaurant"


def menu_bulk_headers(outlet: str = "") -> tuple[str, ...]:
    return UNIFIED_HEADERS


def _style_header(ws, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _add_list_validation(ws, formula: str, cells: str, title: str, error: str) -> None:
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle=title,
        error=error,
    )
    dv.add(cells)
    ws.add_data_validation(dv)


def _col_range(letter: str, count: int) -> str:
    end = 1 + max(int(count or 0), 1)
    return f"Dropdowns!${letter}$2:${letter}${end}"


def _dependent_list_formula(bar_ref: str, rest_ref: str) -> str:
    return f'=INDIRECT(IF($A2="Bar","{bar_ref}","{rest_ref}"))'


def _category_names_for_outlet(conn, outlet: str) -> list[str]:
    names: list[str] = []
    seen: set[str] = set()
    for cat in list_pos_menu_categories(conn, outlet=outlet):
        name = (cat.get("name") or "").strip()
        if not name or _is_placeholder_category(name):
            continue
        key = name.casefold()
        if key in seen:
            continue
        seen.add(key)
        names.append(name)
    return names


def build_pos_menu_bulk_template(conn, outlet: str = "restaurant") -> io.BytesIO:
    """Workbook matching Single: Outlet drives Category and Menu Type dropdowns."""
    outlet = normalize_pos_outlet(outlet)
    headers = menu_bulk_headers(outlet)
    rest_cats = _category_names_for_outlet(conn, POS_OUTLET_RESTAURANT)
    bar_cats = _category_names_for_outlet(conn, POS_OUTLET_BAR)
    default_outlet = _outlet_label(outlet)

    wb = Workbook()
    info = wb.active
    info.title = "Instructions"
    info["A1"] = "Bulk menu upload"
    info["A1"].font = TITLE_FONT
    lines = [
        "",
        "This file matches Add menu item → Single.",
        "Set Outlet on each row. Category and Menu Type then match that outlet.",
        "",
        "Restaurant: Category is the restaurant list. Menu Type is Veg / Non-Veg / Liquor.",
        "Bar: Category is the bar list. Menu Type is Food / Liquor (Item Type on Single).",
        "Add a new category in the app, then download again — it will appear in the list.",
        "",
        "1. Open the Menu Items sheet.",
        "2. Enter one dish per row.",
        "3. Required columns: Outlet, Menu, Category, Rate.",
        "4. Pick Outlet first, then Category and Menu Type from the dropdowns.",
        "5. Code is optional.",
        "6. Save the file and upload it from Add menu item → Bulk.",
        "",
        "Do not rename the header row. Extra blank rows are included for more dishes.",
    ]
    for idx, text in enumerate(lines, start=2):
        cell = info.cell(idx, 1, text)
        cell.font = BODY_FONT
    info.column_dimensions["A"].width = 92

    lists = wb.create_sheet("Dropdowns")
    lists["A1"] = "Restaurant"
    lists["B1"] = "Bar"
    lists["C1"] = "RestaurantType"
    lists["D1"] = "BarType"
    lists["E1"] = "Outlet"
    _style_header(lists, 5)
    for i, name in enumerate(rest_cats, start=2):
        lists.cell(i, 1, name)
    for i, name in enumerate(bar_cats, start=2):
        lists.cell(i, 2, name)
    for i, label in enumerate(MENU_TYPE_LABELS, start=2):
        lists.cell(i, 3, label)
    for i, label in enumerate(ITEM_TYPE_LABELS, start=2):
        lists.cell(i, 4, label)
    for i, label in enumerate(OUTLET_LABELS, start=2):
        lists.cell(i, 5, label)
    lists.column_dimensions["A"].width = 22
    lists.column_dimensions["B"].width = 22
    lists.column_dimensions["C"].width = 16
    lists.column_dimensions["D"].width = 14
    lists.column_dimensions["E"].width = 16
    lists.freeze_panes = "A2"
    lists.sheet_state = "hidden"

    ws = wb.create_sheet("Menu Items", 1)
    ws.append(list(headers))
    _style_header(ws, len(headers))
    for col in (1, 2, 3, 5):
        ws.cell(1, col).fill = REQUIRED_FILL
        ws.cell(1, col).font = Font(color="1E3A5F", bold=True, name="Calibri", size=11)
    last_row = 1 + BLANK_ROWS
    for _ in range(BLANK_ROWS):
        ws.append([default_outlet, "", "", "", None, ""])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16
    for row in range(2, last_row + 1):
        ws.cell(row, 5).number_format = "0.00"

    _add_list_validation(
        ws,
        f"Dropdowns!$E$2:$E${1 + len(OUTLET_LABELS)}",
        f"A2:A{last_row}",
        "Outlet",
        "Pick Restaurant or Bar.",
    )
    _add_list_validation(
        ws,
        _dependent_list_formula(_col_range("B", len(bar_cats)), _col_range("A", len(rest_cats))),
        f"C2:C{last_row}",
        "Category",
        "Pick a category for that outlet (same list as Single).",
    )
    _add_list_validation(
        ws,
        _dependent_list_formula(
            _col_range("D", len(ITEM_TYPE_LABELS)),
            _col_range("C", len(MENU_TYPE_LABELS)),
        ),
        f"D2:D{last_row}",
        "Menu Type",
        "Restaurant: Veg, Non-Veg, or Liquor. Bar: Food or Liquor.",
    )

    for sheet in wb.worksheets:
        sheet.protection.sheet = False

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _first_row_values(ws, max_col: int | None = None):
    kwargs = {"min_row": 1, "max_row": 1, "values_only": True}
    if max_col is not None:
        kwargs["max_col"] = max_col
    return next(ws.iter_rows(**kwargs), None)


def _header_key_tuple(row, count: int) -> tuple[str, ...]:
    cells = list(row or ())
    keys = []
    for idx in range(count):
        keys.append(_norm_header(cells[idx] if idx < len(cells) else ""))
    return tuple(keys)


def _sheet_named(wb, title: str):
    want = str(title or "").strip().casefold()
    for name in wb.sheetnames:
        if str(name).strip().casefold() == want:
            return wb[name]
    return None


def _assert_matches_template(wb, outlet: str = "") -> None:
    """Reject workbooks that are not the Download template."""
    items = _sheet_named(wb, "Menu Items")
    if items is None or _sheet_named(wb, "Dropdowns") is None:
        raise ValueError(
            "This file is not the menu template. Download the template and upload that file."
        )
    instructions = _sheet_named(wb, "Instructions")
    title_row = _first_row_values(instructions, 1) if instructions is not None else None
    title = _clean_cell(title_row[0] if title_row else "")
    if title.casefold() != "bulk menu upload":
        raise ValueError(
            "This file is not the menu template. Download the template and upload that file."
        )

    header_row = _first_row_values(items)
    want = tuple(_norm_header(h) for h in UNIFIED_HEADERS)
    got = _header_key_tuple(header_row, len(UNIFIED_HEADERS))
    if got == want:
        return
    raise ValueError(
        "This file does not match the menu template. Download the template and upload that file."
    )


def _parse_outlet_label(raw: str) -> str:
    v = raw.strip().casefold()
    if v in ("bar",):
        return POS_OUTLET_BAR
    if v in ("restaurant", "resto"):
        return POS_OUTLET_RESTAURANT
    return ""


def _parse_menu_type(raw: str) -> str:
    v = raw.strip().lower().replace("-", "_").replace(" ", "_")
    if v in ("veg",):
        return "veg"
    if v in ("non_veg", "nonveg"):
        return "non_veg"
    if v in ("liquor", "liquour", "alcohol"):
        return "liquor"
    return ""


def _parse_item_kind(raw: str) -> str:
    v = raw.strip().lower()
    if v in ("liquor", "liquour", "alcohol", "bar"):
        return "liquor"
    if v in ("food", "", "none"):
        return "food"
    return ""


def _parse_rate(raw: Any) -> float | None:
    if raw in (None, ""):
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _row_types(raw: str, bar: bool) -> tuple[str, str] | None:
    """Map the Single-style Menu Type cell to (menu_type, item_kind). None = invalid."""
    text = _clean_cell(raw)
    if bar:
        if not text:
            return "", "food"
        kind = _parse_item_kind(text)
        if kind:
            return "", kind
        menu_type = _parse_menu_type(text)
        if menu_type == "liquor":
            return "", "liquor"
        if menu_type:
            return "", "food"
        return None
    if not text:
        return "", "food"
    menu_type = _parse_menu_type(text)
    if menu_type:
        return menu_type, "liquor" if menu_type == "liquor" else "food"
    kind = _parse_item_kind(text)
    if kind == "liquor":
        return "liquor", "liquor"
    if kind == "food":
        return "", "food"
    return None


def import_pos_menu_bulk(conn, file_stream, *, outlet: str = "restaurant", updated_by: str = "") -> dict[str, Any]:
    """Create menu items from an uploaded workbook. Skips blanks and duplicates."""
    fallback_outlet = normalize_pos_outlet(outlet)
    try:
        wb = load_workbook(file_stream, data_only=True)
    except Exception as exc:
        raise ValueError("Could not read the Excel file. Upload an .xlsx template.") from exc

    try:
        _assert_matches_template(wb, fallback_outlet)
    except ValueError:
        try:
            wb.close()
        except Exception:
            pass
        raise

    ws = _sheet_named(wb, "Menu Items")
    rows_iter = ws.iter_rows(values_only=True)
    try:
        next(rows_iter)
    except StopIteration as exc:
        raise ValueError("The Menu Items sheet is empty.") from exc

    col_index: dict[str, int] = {
        "outlet": 0,
        "menu": 1,
        "category": 2,
        "menu_type": 3,
        "rate": 4,
        "code": 5,
    }

    categories = list_pos_menu_categories(
        conn, outlets=[POS_OUTLET_RESTAURANT, POS_OUTLET_BAR]
    )
    cat_by_outlet_name: dict[tuple[str, str], dict[str, Any]] = {}
    for cat in categories:
        name = (cat.get("name") or "").strip()
        if not name or _is_placeholder_category(name):
            continue
        cat_by_outlet_name[(normalize_pos_outlet(cat.get("outlet")), name.casefold())] = cat

    existing = list_pos_menu_items(
        conn, outlets=[POS_OUTLET_RESTAURANT, POS_OUTLET_BAR]
    )
    seen = {
        (
            (it.get("name") or "").strip().casefold(),
            int(it.get("category_id") or 0),
        )
        for it in existing
        if (it.get("name") or "").strip()
    }

    created: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for offset, values in enumerate(rows_iter, start=2):
        values = values or ()

        def col(key: str) -> Any:
            idx = col_index.get(key)
            if idx is None or idx >= len(values):
                return None
            return values[idx]

        name = _clean_cell(col("menu"))
        category_name = _clean_cell(col("category"))
        rate_raw = col("rate")
        if not name and not category_name and rate_raw in (None, ""):
            continue
        if not name:
            errors.append({"row": offset, "error": "Menu name is required."})
            continue
        if not category_name:
            errors.append({"row": offset, "error": f"{name}: Category is required."})
            continue

        outlet_raw = _clean_cell(col("outlet"))
        row_outlet = _parse_outlet_label(outlet_raw) if outlet_raw else fallback_outlet
        if outlet_raw and not _parse_outlet_label(outlet_raw):
            errors.append({"row": offset, "error": f"{name}: Outlet must be Restaurant or Bar."})
            continue
        if not row_outlet:
            errors.append({"row": offset, "error": f"{name}: Outlet is required."})
            continue

        cat = cat_by_outlet_name.get((row_outlet, category_name.casefold()))
        if not cat:
            errors.append(
                {
                    "row": offset,
                    "error": (
                        f"{name}: Category “{category_name}” was not found on the "
                        f"{_outlet_label(row_outlet)} menu."
                    ),
                }
            )
            continue
        rate_val = _parse_rate(rate_raw)
        if rate_val is None:
            errors.append({"row": offset, "error": f"{name}: Rate must be a number."})
            continue
        if rate_val < 0:
            errors.append({"row": offset, "error": f"{name}: Rate cannot be negative."})
            continue

        cat_id = int(cat["id"])
        key = (name.casefold(), cat_id)
        if key in seen:
            skipped.append({"row": offset, "name": name, "reason": "Already on the menu."})
            continue

        bar = _is_bar(row_outlet)
        parsed = _row_types(_clean_cell(col("menu_type")), bar)
        if parsed is None:
            if bar:
                errors.append({"row": offset, "error": f"{name}: Menu Type must be Food or Liquor."})
            else:
                errors.append(
                    {"row": offset, "error": f"{name}: Menu Type must be Veg, Non-Veg, or Liquor."}
                )
            continue
        menu_type, item_kind = parsed

        try:
            saved = save_pos_menu_item(
                conn,
                category_id=cat_id,
                name=name,
                code=_clean_cell(col("code")),
                rate=rate_val,
                menu_type=menu_type if not bar else None,
                item_kind=item_kind,
                recipe=[],
                updated_by=updated_by or None,
                outlet=row_outlet,
            )
        except ValueError as exc:
            errors.append({"row": offset, "error": f"{name}: {exc}"})
            continue
        seen.add(key)
        created.append({"row": offset, "id": saved.get("id"), "name": saved.get("name")})

    try:
        wb.close()
    except Exception:
        pass

    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "created_count": len(created),
        "skipped_count": len(skipped),
        "error_count": len(errors),
    }
