"""Employee Payroll module ported from Neeraj Textile."""

import calendar
import io
import os
import re
import tempfile
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from flask import Blueprint, has_request_context, jsonify, redirect, render_template, request, send_file, session, url_for

from db import SQL_NOW, get_db
from workspace_access import user_can_access_payroll_submodule

payroll_bp = Blueprint("payroll", __name__)

_EMPLOYEE_DISPLAY_ORDER = "LOWER(location), LOWER(name), id DESC"
_EMPLOYEE_SORT_ORDERS = {
    'id': "LENGTH(emp_code), LOWER(emp_code), id",
    'name': "LOWER(name), LENGTH(emp_code), LOWER(emp_code), id",
    'department': "LOWER(location), LOWER(name), LENGTH(emp_code), LOWER(emp_code), id",
}
_DEFAULT_COMPANY = "Hotel Bell Elite"
_EMP_CODE_PREFIX = "HBE"
_EMP_CODE_RE = re.compile(rf"^{re.escape(_EMP_CODE_PREFIX)}(\d+)$", re.IGNORECASE)
_EPF_MAX = 1800.0
_ESIC_RATE = 0.0075  # employee share 0.75% of actual gross when <= wage limit
_ESIC_WAGE_LIMIT = 21000.0
_ESIC_FIXED_ABOVE_LIMIT = 158.0  # fixed ESI when actual gross > wage limit
_PAYROLL_DEPARTMENTS = (
    "OM",
    "FO",
    "F&B",
    "KITCHEN",
    "UTILITY",
    "BAR",
    "HK",
    "MAINTENANCE",
    "SECURITY",
)
_PAYROLL_LOCK_START = (2026, 7)
REPORTING_PERIOD_SESSION_KEY = "reporting_period"

# Injected from app at registration time
_pop_auth_notice = None
_permission_denied_response = None
get_current_user = None


def _bind_app_helpers(pop_auth_notice, permission_denied_response, get_user):
    global _pop_auth_notice, _permission_denied_response, get_current_user
    _pop_auth_notice = pop_auth_notice
    _permission_denied_response = permission_denied_response
    get_current_user = get_user


def _round_half_up(value, dec=0):
    """Round using half-up rules instead of banker's rounding."""
    try:
        quantum = Decimal('1').scaleb(-dec)
        return float(Decimal(str(value or 0)).quantize(quantum, rounding=ROUND_HALF_UP))
    except (InvalidOperation, TypeError, ValueError):
        return 0.0


def _round_rupee(value):
    """Round a value to the nearest rupee using half-up rules."""
    return int(_round_half_up(value, 0))


def _inr_format(value, dec=0):
    """Format a number as Indian-rupee text (₹1,23,456.78)."""
    try:
        v = float(value or 0)
        neg = v < 0
        v = abs(v)
        if dec <= 0:
            s = f"{int(round(v)):,}"
        else:
            s = f"{v:,.{dec}f}"
        parts = s.split(".")
        int_part = parts[0]
        if len(int_part) > 4:
            raw = int_part.replace(",", "")
            if len(raw) > 3:
                last3 = raw[-3:]
                rest = raw[:-3]
                groups = []
                while len(rest) > 2:
                    groups.insert(0, rest[-2:])
                    rest = rest[:-2]
                if rest:
                    groups.insert(0, rest)
                int_part = ",".join(groups) + "," + last3
        s = int_part + ("." + parts[1] if len(parts) > 1 else "")
        return ("−" if neg else "") + "₹" + s
    except (TypeError, ValueError):
        return "₹0"


def _next_emp_code(conn):
    """Return the next unique Employee ID in the HBE{n} sequence."""
    rows = conn.execute(
        "SELECT emp_code FROM employees WHERE emp_code LIKE ?",
        (f"{_EMP_CODE_PREFIX}%",),
    ).fetchall()
    max_n = 0
    for row in rows:
        match = _EMP_CODE_RE.match((row['emp_code'] or '').strip())
        if match:
            max_n = max(max_n, int(match.group(1)))
    return f"{_EMP_CODE_PREFIX}{max_n + 1}"


def _emp_code_taken(conn, emp_code, exclude_id=None):
    sql = "SELECT id FROM employees WHERE UPPER(TRIM(emp_code))=UPPER(TRIM(?))"
    params = [emp_code]
    if exclude_id is not None:
        sql += " AND id<>?"
        params.append(exclude_id)
    return conn.execute(sql, params).fetchone() is not None


def _default_reporting_period(today=None):
    """Use previous month until the 10th; otherwise use the current month."""
    today = today or date.today()
    if today.day <= 10:
        if today.month == 1:
            return today.year - 1, 12
        return today.year, today.month - 1
    return today.year, today.month


def _parse_period_value(raw_year, raw_month):
    if raw_year in (None, '') or raw_month in (None, ''):
        raise ValueError('Missing year or month')

    year = int(raw_year)
    month = int(raw_month)
    if month < 1 or month > 12:
        raise ValueError('Month out of range')
    return year, month


def _session_reporting_period():
    raw = session.get(REPORTING_PERIOD_SESSION_KEY) or {}
    if not isinstance(raw, dict):
        return None
    try:
        return _parse_period_value(raw.get('year'), raw.get('month'))
    except (TypeError, ValueError):
        return None


def _remember_reporting_period(year, month):
    session[REPORTING_PERIOD_SESSION_KEY] = {'year': int(year), 'month': int(month)}


def _period_anchor_date(year, month, today=None):
    """Pick a representative date inside the selected reporting month."""
    today = today or date.today()
    if year == today.year and month == today.month:
        day = today.day
    else:
        day = calendar.monthrange(year, month)[1]
    return date(year, month, day)


def _default_reporting_date(today=None):
    today = today or date.today()
    year, month = _default_reporting_period(today)
    return _period_anchor_date(year, month, today)


def _period_from_source(source, year_key='year', month_key='month'):
    """Read year/month from args, form, or JSON with reporting defaults."""
    default_year, default_month = _default_reporting_period()
    session_period = _session_reporting_period()
    session_year = session_month = None
    if session_period:
        session_year, session_month = session_period

    raw_year = source.get(year_key)
    raw_month = source.get(month_key)
    candidate_year = raw_year if raw_year not in (None, '') else (
        session_year if session_year is not None else default_year
    )
    candidate_month = raw_month if raw_month not in (None, '') else (
        session_month if session_month is not None else default_month
    )

    try:
        year, month = _parse_period_value(candidate_year, candidate_month)
    except (TypeError, ValueError):
        fallback = session_period or (default_year, default_month)
        year, month = fallback

    _remember_reporting_period(year, month)
    return year, month


def _period_label(year, month):
    return f'{calendar.month_name[month]} {year}'


def _previous_period(year, month):
    if month == 1:
        return year - 1, 12
    return year, month - 1


def _parse_attendance_date(value):
    try:
        return date.fromisoformat((value or '').strip())
    except (TypeError, ValueError, AttributeError):
        return None


def _is_future_attendance_date(value, today=None):
    att_dt = _parse_attendance_date(value)
    if not att_dt:
        return False
    return att_dt > (today or date.today())


def _parse_db_datetime(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace(' ', 'T')).replace(tzinfo=None)
    except ValueError:
        return None


def _can_modify_attendance_record(user, att_dt, record=None, today=None, now=None, payroll_locked=False):
    if not att_dt:
        return False
    if payroll_locked:
        return False
    today = today or date.today()
    if att_dt > today:
        return False
    if user and user.get('is_admin'):
        return True
    if att_dt == today:
        return True
    if not record:
        return True
    updated_at = _parse_db_datetime(record.get('updated_at') if hasattr(record, 'get') else record['updated_at'])
    if not updated_at:
        return False
    now = now or datetime.now()
    return timedelta(0) <= (now - updated_at) <= timedelta(minutes=30)


def _attendance_date_lock_message(today=None, payroll_locked=False, year=None, month=None):
    if payroll_locked and year and month:
        return _payroll_month_frozen_message(year, month)
    today = today or date.today()
    return (
        f'Only administrators can modify locked attendance. Standard users can edit today, '
        f'or fill any unmarked previous date until 30 minutes after it is marked. Today is {today.strftime("%d/%m/%y")}.'
    )


def _payroll_landing_redirect(user, year, month):
    if not user:
        return None
    if user_can_access_payroll_submodule(user, 'employee'):
        return None
    if user_can_access_payroll_submodule(user, 'attendance'):
        return redirect(url_for('attendance_overview', year=year, month=month))
    if user_can_access_payroll_submodule(user, 'report'):
        return redirect(url_for('report', year=year, month=month))
    if user_can_access_payroll_submodule(user, 'credit'):
        return redirect(url_for('credits_dashboard', year=year, month=month))
    return _permission_denied_response('You do not have access to Employee Payroll.')


def _attendance_scope_for_user(user=None):
    user = user or get_current_user()
    if not user or user.get('is_admin'):
        return {'companies': set(), 'locations': set(), 'restricted': False}
    companies = set(user.get('attendance_companies') or set())
    locations = set(user.get('attendance_locations') or set())
    return {
        'companies': companies,
        'locations': locations,
        'restricted': bool(companies or locations),
    }


def _append_attendance_scope_conditions(conditions, params, user=None, table_alias=''):
    scope = _attendance_scope_for_user(user)
    prefix = f'{table_alias}.' if table_alias else ''
    if scope['companies']:
        placeholders = ','.join('?' for _ in scope['companies'])
        conditions.append(f'{prefix}company IN ({placeholders})')
        params.extend(sorted(scope['companies']))
    if scope['locations']:
        placeholders = ','.join('?' for _ in scope['locations'])
        conditions.append(f'{prefix}location IN ({placeholders})')
        params.extend(sorted(scope['locations']))
    return scope


def _attendance_filter_options(conn, user=None):
    scope = _attendance_scope_for_user(user)
    departments = list(_PAYROLL_DEPARTMENTS)
    if scope['locations']:
        allowed = set(scope['locations'])
        departments = [name for name in departments if name in allowed]
    return departments


def _user_can_access_attendance_employee(conn, emp_id, user=None):
    user = user or get_current_user()
    if user and user.get('is_admin'):
        return True
    scope = _attendance_scope_for_user(user)
    if not scope['restricted']:
        return True
    row = conn.execute("SELECT company, location FROM employees WHERE id=?", (emp_id,)).fetchone()
    if not row:
        return False
    if scope['companies'] and row['company'] not in scope['companies']:
        return False
    if scope['locations'] and row['location'] not in scope['locations']:
        return False
    return True
def _is_payroll_month_locked(conn, year, month):
    row = conn.execute(
        "SELECT 1 FROM payroll_month_locks WHERE year=? AND month=?",
        (year, month)
    ).fetchone()
    return bool(row)


def _payroll_month_frozen_message(year, month):
    label = _period_label(year, month)
    return (
        f'{label} is locked. No attendance, credit, repayment, tip incentive, '
        f'or payroll edits are allowed for this month — including for administrators.'
    )


def _period_from_credit_date(cr_date):
    """Return (year, month) for a credit entry date, or (None, None) if invalid."""
    try:
        d = datetime.strptime(str(cr_date)[:10], '%Y-%m-%d').date()
        return d.year, d.month
    except (TypeError, ValueError):
        return None, None


def _is_credit_date_locked(conn, cr_date):
    """True when the credit date falls in a locked payroll month."""
    year, month = _period_from_credit_date(cr_date)
    if year is None:
        return False
    return _is_payroll_month_locked(conn, year, month)


def _annotate_credit_editability(conn, items):
    """Attach can_edit to each credit dict based on payroll month lock."""
    annotated = []
    for item in items:
        row = dict(item)
        row['can_edit'] = not _is_credit_date_locked(conn, row.get('date'))
        annotated.append(row)
    return annotated


def _employee_has_locked_month_data(conn, emp_id):
    """True when the employee has attendance/credits/tips tied to a locked month."""
    if conn.execute(
        """SELECT 1
           FROM attendance a
           JOIN payroll_month_locks l
             ON CAST(substr(a.date, 1, 4) AS INTEGER) = l.year
            AND CAST(substr(a.date, 6, 2) AS INTEGER) = l.month
           WHERE a.employee_id = ?
           LIMIT 1""",
        (emp_id,),
    ).fetchone():
        return True
    if conn.execute(
        """SELECT 1
           FROM credits c
           JOIN payroll_month_locks l
             ON CAST(substr(c.date, 1, 4) AS INTEGER) = l.year
            AND CAST(substr(c.date, 6, 2) AS INTEGER) = l.month
           WHERE c.employee_id = ?
           LIMIT 1""",
        (emp_id,),
    ).fetchone():
        return True
    if conn.execute(
        """SELECT 1
           FROM tip_incentive_payouts t
           JOIN payroll_month_locks l ON t.year = l.year AND t.month = l.month
           WHERE t.employee_id = ?
           LIMIT 1""",
        (emp_id,),
    ).fetchone():
        return True
    return False


def _wage_fields_changed(existing, new_vals):
    """True when payroll-affecting master fields would change."""
    def _f(key, default=0):
        try:
            return float(existing[key] if existing[key] is not None else default)
        except (TypeError, ValueError, KeyError, IndexError):
            return float(default)

    def _i(key, default=0):
        try:
            return int(existing[key] if existing[key] is not None else default)
        except (TypeError, ValueError, KeyError, IndexError):
            return int(default)

    return (
        abs(_f('gross_salary') - float(new_vals.get('gross_salary') or 0)) > 1e-9
        or abs(_f('basic_salary') - float(new_vals.get('basic_salary') or 0)) > 1e-9
        or abs(_f('epf_amount') - float(new_vals.get('epf_amount') or 0)) > 1e-9
        or abs(_f('esic_amount') - float(new_vals.get('esic_amount') or 0)) > 1e-9
        or _i('epf_exempt') != int(new_vals.get('epf_exempt') or 0)
        or _i('esic_exempt') != int(new_vals.get('esic_exempt') or 0)
        or _i('total_off') != int(new_vals.get('total_off') or 0)
    )


def _get_payroll_month_state(conn, year, month):
    start_year, start_month = _PAYROLL_LOCK_START
    label = _period_label(year, month)
    start_label = _period_label(start_year, start_month)
    supported = (year, month) >= _PAYROLL_LOCK_START
    locked = supported and _is_payroll_month_locked(conn, year, month)
    prev_year, prev_month = _previous_period(year, month)
    prev_label = _period_label(prev_year, prev_month)
    prev_locked = supported and ((year, month) == _PAYROLL_LOCK_START or _is_payroll_month_locked(conn, prev_year, prev_month))

    if not supported:
        return {
            'label': label,
            'status_label': 'Legacy',
            'status_badge': 'bg-gray',
            'supported': False,
            'locked': False,
            'can_edit': False,
            'can_lock': False,
            'message': f'Month locking starts from {start_label}. Earlier months stay read-only in the new repayment flow.',
            'button_label': 'Lock Not Available',
            'reason': f'Month locking starts from {start_label}.',
            'previous_label': prev_label,
            'previous_locked': False,
        }

    if locked:
        return {
            'label': label,
            'status_label': 'Locked',
            'status_badge': 'bg-green',
            'supported': True,
            'locked': True,
            'can_edit': False,
            'can_lock': False,
            'message': _payroll_month_frozen_message(year, month),
            'button_label': 'Locked',
            'reason': f'{label} is already locked.',
            'previous_label': prev_label,
            'previous_locked': prev_locked,
        }

    if (year, month) == _PAYROLL_LOCK_START:
        return {
            'label': label,
            'status_label': 'Open',
            'status_badge': 'bg-orange',
            'supported': True,
            'locked': False,
            'can_edit': True,
            'can_lock': True,
            'message': f'{label} is the starting month for the new repayment flow. Save repayments, then lock it to open the next month.',
            'button_label': f'Lock {label}',
            'reason': '',
            'previous_label': prev_label,
            'previous_locked': True,
        }

    if prev_locked:
        return {
            'label': label,
            'status_label': 'Open',
            'status_badge': 'bg-orange',
            'supported': True,
            'locked': False,
            'can_edit': True,
            'can_lock': True,
            'message': f'{label} is open for repayment updates because {prev_label} is already locked.',
            'button_label': f'Lock {label}',
            'reason': '',
            'previous_label': prev_label,
            'previous_locked': True,
        }

    # Earlier months are not locked yet — still allow repayment edits for this
    # unlocked month; only the Lock action waits on the previous month.
    return {
        'label': label,
        'status_label': 'Open',
        'status_badge': 'bg-orange',
        'supported': True,
        'locked': False,
        'can_edit': True,
        'can_lock': False,
        'message': f'{label} is open for repayment updates. Lock {prev_label} first before locking {label}.',
        'button_label': f'Lock {label}',
        'reason': f'Lock {prev_label} first.',
        'previous_label': prev_label,
        'previous_locked': False,
    }


def _salary_repayment_descriptor(year, month):
    label = _period_label(year, month)
    return f'Salary Repayment – {label}'


def _salary_repayment_date(year, month):
    last_day = calendar.monthrange(year, month)[1]
    return f'{year}-{month:02d}-{last_day:02d}'


def _get_month_credit_repayment(conn, emp_id, year, month):
    row = conn.execute(
        """
        SELECT amount
        FROM credits
        WHERE employee_id=?
          AND (
                (entry_type='salary_repayment' AND payroll_year=? AND payroll_month=?)
             OR description=?
          )
        ORDER BY id DESC
        LIMIT 1
        """,
        (emp_id, year, month, _salary_repayment_descriptor(year, month))
    ).fetchone()
    return _round_half_up(abs(float(row['amount'] or 0)), 2) if row else 0.0


def _get_month_tip_incentive(conn, emp_id, year, month, company=None):
    """Tip incentive payout for one employee in a payroll month."""
    if company:
        row = conn.execute(
            """SELECT amount FROM tip_incentive_payouts
               WHERE company=? AND year=? AND month=? AND employee_id=?""",
            (company, year, month, emp_id),
        ).fetchone()
    else:
        row = conn.execute(
            """SELECT COALESCE(SUM(amount), 0) AS amount FROM tip_incentive_payouts
               WHERE year=? AND month=? AND employee_id=?""",
            (year, month, emp_id),
        ).fetchone()
    return _round_half_up(float(row['amount'] or 0), 2) if row else 0.0


def _get_month_tip_incentive_map(conn, year, month, company=None):
    """Map employee_id -> tip incentive for a payroll month."""
    if company:
        rows = conn.execute(
            """SELECT employee_id, amount FROM tip_incentive_payouts
               WHERE company=? AND year=? AND month=?""",
            (company, year, month),
        ).fetchall()
    else:
        rows = conn.execute(
            """SELECT employee_id, COALESCE(SUM(amount), 0) AS amount
               FROM tip_incentive_payouts
               WHERE year=? AND month=?
               GROUP BY employee_id""",
            (year, month),
        ).fetchall()
    return {
        row['employee_id']: _round_half_up(float(row['amount'] or 0), 2)
        for row in rows
        if float(row['amount'] or 0) > 0
    }


def _upsert_month_tip_incentive(conn, company, year, month, emp_id, amount):
    """Save tip incentive for one employee; zero amount deletes the row."""
    amount = _round_half_up(max(0.0, float(amount or 0)), 2)
    if amount <= 0:
        conn.execute(
            """DELETE FROM tip_incentive_payouts
               WHERE company=? AND year=? AND month=? AND employee_id=?""",
            (company, year, month, emp_id),
        )
        return 0.0
    conn.execute(
        """INSERT INTO tip_incentive_payouts (company, year, month, employee_id, amount, updated_at)
           VALUES (?, ?, ?, ?, ?, datetime('now','localtime'))
           ON CONFLICT(company, year, month, employee_id)
           DO UPDATE SET amount=excluded.amount, updated_at=datetime('now','localtime')""",
        (company, year, month, emp_id, amount),
    )
    return amount


def _get_month_credit_repayment_map(conn, year, month):
    rows = conn.execute(
        """
        SELECT employee_id, amount
        FROM credits
        WHERE (entry_type='salary_repayment' AND payroll_year=? AND payroll_month=?)
           OR description=?
        ORDER BY employee_id, id DESC
        """,
        (year, month, _salary_repayment_descriptor(year, month))
    ).fetchall()
    result = {}
    for row in rows:
        emp_id = row['employee_id']
        if emp_id in result:
            continue
        result[emp_id] = _round_half_up(abs(float(row['amount'] or 0)), 2)
    return result


def _upsert_month_credit_repayment(conn, emp_id, year, month, amount):
    desc = _salary_repayment_descriptor(year, month)
    repayment_date = _salary_repayment_date(year, month)
    amount = _round_half_up(amount, 2)
    rows = conn.execute(
        """
        SELECT id
        FROM credits
        WHERE employee_id=?
          AND (
                (entry_type='salary_repayment' AND payroll_year=? AND payroll_month=?)
             OR description=?
          )
        ORDER BY id DESC
        """,
        (emp_id, year, month, desc)
    ).fetchall()

    if amount > 0:
        keep_id = rows[0]['id'] if rows else None
        if keep_id:
            conn.execute(
                """
                UPDATE credits
                SET amount=?, date=?, description=?, entry_type='salary_repayment',
                    payroll_year=?, payroll_month=?
                WHERE id=?
                """,
                (-amount, repayment_date, desc, year, month, keep_id)
            )
            stale_rows = rows[1:]
        else:
            conn.execute(
                """
                INSERT INTO credits (employee_id, date, description, amount, entry_type, payroll_year, payroll_month)
                VALUES (?,?,?,?,?,?,?)
                """,
                (emp_id, repayment_date, desc, -amount, 'salary_repayment', year, month)
            )
            stale_rows = []

        for row in stale_rows:
            conn.execute("DELETE FROM credits WHERE id=?", (row['id'],))
        return

    for row in rows:
        conn.execute("DELETE FROM credits WHERE id=?", (row['id'],))


def _employee_month_salary(conn, employee, year, month, credit_repayment=None, tip_incentive=None):
    emp = dict(employee) if not isinstance(employee, dict) else dict(employee)
    credit_repayment = (
        _get_month_credit_repayment(conn, emp['id'], year, month)
        if credit_repayment is None
        else _round_half_up(credit_repayment, 2)
    )
    tip_incentive = (
        _get_month_tip_incentive(conn, emp['id'], year, month)
        if tip_incentive is None
        else _round_half_up(tip_incentive, 2)
    )
    att = _get_month_attendance(conn, emp['id'], year, month)
    calendar_days = int(att.get('num_days', 0) or 0)
    total_off = int(emp.get('total_off') or 0)
    if att['tracked']:
        salary = _calc_salary(
            emp['gross_salary'],
            calendar_days=calendar_days,
            weekday_leave_days=att.get('weekday_leave_days', att.get('absent', 0)),
            total_off=total_off,
            tracked=True,
            custom_basic=emp.get('basic_salary', 0),
            custom_epf=emp.get('epf_amount', 0),
            custom_esic=emp.get('esic_amount', 0),
            credit_repayment=credit_repayment,
            sunday_incentive_days=0,
            sunday_shift='',
            epf_exempt=bool(emp.get('epf_exempt', 0)),
            esic_exempt=bool(emp.get('esic_exempt', 0)),
            tip_incentive=tip_incentive,
        )
    else:
        salary = _calc_salary(
            emp['gross_salary'],
            calendar_days=calendar_days,
            tracked=False,
            custom_basic=emp.get('basic_salary', 0),
            custom_epf=emp.get('epf_amount', 0),
            custom_esic=emp.get('esic_amount', 0),
            credit_repayment=credit_repayment,
            sunday_incentive_days=0,
            sunday_shift='',
            epf_exempt=bool(emp.get('epf_exempt', 0)),
            esic_exempt=bool(emp.get('esic_exempt', 0)),
            tip_incentive=tip_incentive,
        )
    salary['credit_repayment'] = _round_half_up(credit_repayment, 2)
    salary['tip_incentive'] = _round_half_up(tip_incentive, 2)
    return att, salary


def _get_month_repayment_limits(conn, employee, year, month):
    current_repayment = _get_month_credit_repayment(conn, employee['id'], year, month)
    _, base_salary = _employee_month_salary(conn, employee, year, month, credit_repayment=0)
    outstanding_before_month = max(0.0, _get_total_credits(conn, employee['id']) + current_repayment)
    max_allowed = min(max(0.0, float(base_salary.get('net', 0) or 0)), outstanding_before_month)
    return {
        'credit_repayment': _round_half_up(current_repayment, 2),
        'credit_total_before_month': _round_half_up(outstanding_before_month, 2),
        'repayment_max': _round_half_up(max_allowed, 2),
        'salary_cap': _round_half_up(base_salary.get('net', 0) or 0, 2),
    }


def _apply_total_off_to_attendance_view(att, total_off=0):
    """Adjust attendance display so Total Off entitlement is not counted as Absent.

    Marked leave within Total Off is treated as paid leave; only leave beyond
    entitlement remains chargeable Absent (same basis as LOP days).
    """
    view = dict(att or {})
    leave_days = float(view.get('weekday_leave_days', 0) or 0)
    if leave_days.is_integer():
        leave_days = int(leave_days)
    off = max(0, int(total_off or 0))
    covered = min(float(leave_days), float(off))
    if covered.is_integer():
        covered = int(covered)
    chargeable = max(0.0, float(leave_days) - float(off))
    if chargeable.is_integer():
        chargeable = int(chargeable)
    view['absent_marked'] = view.get('absent', 0)
    view['half_day_marked'] = view.get('half_day', 0)
    view['leave_days'] = leave_days
    view['leave_covered_by_off'] = covered
    view['total_off'] = off
    # Absent shown in UI = unpaid leave after Total Off (not raw marked A days).
    view['absent'] = chargeable
    return view


def _attach_employee_month_context(conn, employee, year, month, payroll_state=None):
    emp = dict(employee) if not isinstance(employee, dict) else dict(employee)
    att, salary = _employee_month_salary(conn, emp, year, month)
    emp.update(salary)
    total_off = int(salary.get('total_off', emp.get('total_off') or 0) or 0)
    att = _apply_total_off_to_attendance_view(att, total_off)
    if att.get('tracked'):
        # ATTEND. badge stays as marked present days (P + ½H) from _get_month_attendance.
        # Do not overwrite with paid_calendar_days — unmarked days are not attendance
        # and must not display as a full-month preset like 31/31.
        att['lop_days'] = salary.get('lop_days', 0)
        att['total_off'] = total_off
    emp['att'] = att
    emp['credit_total'] = _get_total_credits(conn, emp['id'])
    emp.update(_get_month_repayment_limits(conn, emp, year, month))
    can_edit = bool(payroll_state['can_edit']) if payroll_state else True
    emp['repayment_input_visible'] = (emp['credit_total_before_month'] > 0) or (emp['credit_repayment'] > 0)
    emp['repayment_input_enabled'] = can_edit and emp['repayment_max'] > 0
    return emp

def _send_xlsx(buf: io.BytesIO, filename: str):
    resp = send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    return resp


def _get_locations():
    return list(_PAYROLL_DEPARTMENTS)


def _payroll_nav_view(mode) -> str:
    if mode == 'report':
        return 'report'
    if mode in ('attendance', 'attendance_list', 'attendance_date'):
        return 'attendance'
    if mode in ('credits', 'credits_dashboard'):
        return 'credit'
    return 'employee'


def _sales_nav_view(endpoint=None) -> str:
    ep = endpoint or request.endpoint
    if ep == 'sales_update_entry':
        return 'entry'
    if ep == 'sales_report':
        return 'report'
    return 'dashboard'


def _su_render(template, **kwargs):
    """Render sales module pages with unified workspace sidebar context."""
    kwargs.setdefault('auth_notice', _pop_auth_notice())
    kwargs.setdefault('de_nav_host', 'sales')
    kwargs.setdefault('de_nav_section', 'sales')
    kwargs.setdefault('de_nav_sales_view', _sales_nav_view())
    return render_template(template, **kwargs)


def _emp_render(template, **kwargs):
    """Render employees.html with location lists always injected."""
    default_year, default_month = _default_reporting_period()
    kwargs.setdefault('locations', _get_locations())
    kwargs.setdefault('today_year', default_year)
    kwargs.setdefault('today_month', default_month)
    kwargs.setdefault('auth_notice', _pop_auth_notice())
    kwargs.setdefault('de_nav_host', 'payroll')
    kwargs.setdefault('de_nav_section', 'payroll')
    kwargs.setdefault('de_nav_payroll_view', _payroll_nav_view(kwargs.get('mode')))
    sel_year = kwargs.get('sel_year')
    sel_month = kwargs.get('sel_month')
    if sel_year and sel_month and 'payroll_state' not in kwargs:
        conn = get_db()
        try:
            kwargs['payroll_state'] = _get_payroll_month_state(conn, int(sel_year), int(sel_month))
        finally:
            conn.close()
    return render_template(template, **kwargs)


_SUNDAY_SHIFT_RATES = {'shift1': 550.0, 'shift2': 600.0, 'shift3': 650.0}


def _calc_total_off_lop(weekday_leave_days, total_off, gross, calendar_days):
    """Compute LOP days/deduction from Total Off paid-leave entitlement."""
    weekday_leave_days = max(0.0, float(weekday_leave_days or 0))
    total_off = max(0, int(total_off or 0))
    cal_days = int(calendar_days or 0)
    lop_days = max(0.0, weekday_leave_days - float(total_off))
    if cal_days <= 0:
        return {
            'lop_days': lop_days,
            'lop_deduction': 0.0,
            'pay_ratio': 0.0,
            'daily_rate': 0.0,
            'total_off': total_off,
            'weekday_leave_days': weekday_leave_days,
            'paid_calendar_days': 0.0,
        }
    daily_rate = round(float(gross) / cal_days, 2)
    pay_ratio = max(0.0, min(1.0, (cal_days - lop_days) / cal_days))
    gross_actual = round(float(gross) * pay_ratio, 2)
    lop_deduction = round(float(gross) - gross_actual, 2)
    paid_calendar_days = cal_days - lop_days
    if paid_calendar_days.is_integer():
        paid_calendar_days = int(paid_calendar_days)
    if lop_days.is_integer():
        lop_days = int(lop_days)
    return {
        'lop_days': lop_days,
        'lop_deduction': lop_deduction,
        'pay_ratio': pay_ratio,
        'daily_rate': daily_rate,
        'total_off': total_off,
        'weekday_leave_days': weekday_leave_days,
        'paid_calendar_days': paid_calendar_days,
    }


def _calc_salary(gross, calendar_days=0, weekday_leave_days=0, total_off=0,
                  tracked=False, custom_basic=0, custom_epf=0, custom_esic=0,
                  credit_repayment=0, sunday_incentive_days=0,
                  sunday_shift='', epf_exempt=False, esic_exempt=False,
                  tip_incentive=0):
    """Calculate salary components from gross / actual gross.

    Deductions (per Actual Gross for the month):
      EPF = min(₹1,800, 12% of Actual Gross) unless exempt / custom EPF
      ESI = 0.75% of Actual Gross when Actual Gross <= ₹21,000
          = ₹158 fixed when Actual Gross > ₹21,000
          (unless exempt / custom ESI)
      Basic = Actual Gross − EPF − ESI (residual)
    Leave: weekday leave beyond Total Off is LOP at gross / calendar_days.
    Tip incentive is added to net (manual monthly payout from Tips pool).
    """
    gross = max(0.0, float(gross))
    credit_repayment = float(credit_repayment or 0)
    tip_incentive = max(0.0, float(tip_incentive or 0))
    custom_basic = float(custom_basic or 0)
    custom_epf = float(custom_epf or 0)
    custom_esic = float(custom_esic or 0)

    def _components(actual_gross, custom_scale=1.0):
        actual_gross = max(0.0, float(actual_gross or 0))
        custom_scale = max(0.0, float(custom_scale or 0))

        if esic_exempt:
            esic = 0.0
            esic_applicable = False
        elif custom_esic > 0:
            esic = round(custom_esic * custom_scale, 2)
            esic_applicable = True
        elif actual_gross > _ESIC_WAGE_LIMIT:
            esic = _ESIC_FIXED_ABOVE_LIMIT
            esic_applicable = True
        elif actual_gross > 0:
            esic = round(actual_gross * _ESIC_RATE, 2)
            esic_applicable = True
        else:
            esic = 0.0
            esic_applicable = False

        if epf_exempt:
            epf = 0.0
        elif custom_epf > 0:
            epf = min(_EPF_MAX, round(custom_epf * custom_scale, 2))
        else:
            epf = min(_EPF_MAX, round(actual_gross * 0.12, 2))

        basic = max(0.0, round(actual_gross - epf - esic, 2))
        return basic, max(0.0, epf), max(0.0, esic), esic_applicable

    basic_full, epf_full, esic_full, esic_applicable = _components(gross, 1.0)
    net_full = max(0.0, round(
        gross - epf_full - esic_full - credit_repayment + tip_incentive, 2))

    result = {
        'basic_full': basic_full, 'epf_full': epf_full,
        'esic_full': esic_full, 'net_full': net_full,
        'esic_applicable': esic_applicable,
        'credit_repayment': credit_repayment,
        'tip_incentive': tip_incentive,
        'sunday_incentive_days': 0.0,
        'sunday_incentive': 0.0,
        'lop_deduction': 0.0,
        'lop_days': 0.0,
        'total_off': int(total_off or 0),
        'weekday_leave_days': 0.0,
        'paid_calendar_days': 0.0,
        'daily_rate': 0.0,
    }

    cal_days = int(calendar_days or 0)
    if tracked and cal_days > 0:
        lop_info = _calc_total_off_lop(weekday_leave_days, total_off, gross, cal_days)
        ratio = lop_info['pay_ratio']
        daily_rate = lop_info['daily_rate']
        sunday_days = max(0.0, float(sunday_incentive_days or 0))
        fixed_rate = _SUNDAY_SHIFT_RATES.get(sunday_shift or '', None)
        if fixed_rate is not None:
            sunday_incentive = round(fixed_rate * sunday_days, 2)
        else:
            sunday_incentive = round(daily_rate * sunday_days, 2)

        gross_actual = round(gross * ratio, 2)
        custom_scale = ratio if (custom_epf > 0 or custom_esic > 0 or custom_basic > 0) else 1.0
        basic_a, epf_a, esic_a, esic_app_a = _components(gross_actual, custom_scale)

        result['present_days'] = lop_info['paid_calendar_days']
        result['total_days'] = cal_days
        result['daily_rate'] = daily_rate
        result['sunday_incentive_days'] = sunday_days
        result['sunday_incentive'] = sunday_incentive
        result['sunday_shift_rate'] = fixed_rate or daily_rate
        result['lop_deduction'] = lop_info['lop_deduction']
        result['lop_days'] = lop_info['lop_days']
        result['total_off'] = lop_info['total_off']
        result['weekday_leave_days'] = lop_info['weekday_leave_days']
        result['paid_calendar_days'] = lop_info['paid_calendar_days']
        result['gross_actual'] = gross_actual
        result['basic'] = basic_a
        result['epf'] = epf_a
        result['esic'] = esic_a
        result['esic_applicable'] = esic_app_a
        result['tip_incentive'] = tip_incentive
        result['net'] = max(0.0, round(
            gross_actual - epf_a - esic_a
            - credit_repayment + sunday_incentive + tip_incentive, 2))
    elif not tracked and cal_days > 0:
        result['present_days'] = 0
        result['total_days'] = cal_days
        result['gross_actual'] = 0.0
        result['basic'] = 0.0
        result['epf'] = 0.0
        result['esic'] = 0.0
        result['tip_incentive'] = tip_incentive
        result['net'] = max(0.0, round(tip_incentive, 2))
    else:
        result['present_days'] = cal_days
        result['total_days'] = cal_days
        result['gross_actual'] = gross
        result['basic'] = basic_full
        result['epf'] = epf_full
        result['esic'] = esic_full
        result['tip_incentive'] = tip_incentive
        result['net'] = net_full

    for key in (
        'basic_full', 'epf_full', 'esic_full', 'net_full',
        'gross_actual', 'basic', 'epf', 'esic', 'net',
        'sunday_incentive', 'lop_deduction', 'tip_incentive',
    ):
        result[key] = _round_rupee(result.get(key, 0))

    return result


def _get_month_attendance(conn, emp_id, year, month):
    """Return attendance dict for an employee in a given month.

    Returns: {present: int, absent: int, half_day: int, total_working: int,
              effective: float, tracked: bool, records: {date_str: status}}
    """
    _, num_days = calendar.monthrange(year, month)
    sunday_dates = set()
    for day in range(1, num_days + 1):
        if calendar.weekday(year, month, day) == 6:  # Sunday
            sunday_dates.add(f'{year}-{month:02d}-{day:02d}')
    rows = conn.execute(
        "SELECT date, status, updated_at FROM attendance WHERE employee_id=? AND date LIKE ?",
        (emp_id, f'{year}-{month:02d}-%')
    ).fetchall()
    today = date.today()
    records = {}
    record_meta = {}
    for r in rows:
        if _is_future_attendance_date(r['date'], today=today):
            continue
        records[r['date']] = r['status']
        record_meta[r['date']] = dict(r)
    present = half_day = absent = 0
    sunday_present = sunday_half_day = 0
    for d, s in records.items():
        if s == 'present':
            present += 1
            if d in sunday_dates:
                sunday_present += 1
        elif s == 'half_day':
            half_day += 1
            if d in sunday_dates:
                sunday_half_day += 1
        elif s == 'absent':
            absent += 1
    # Effective present days: full days + half days * 0.5
    effective = present + half_day * 0.5
    weekday_leave_days = absent + half_day * 0.5
    present_effective = effective
    if present_effective.is_integer():
        present_effective = int(present_effective)
    sunday_effective = sunday_present + sunday_half_day * 0.5
    # tracked = True if any attendance record exists for this month
    tracked = len(records) > 0
    # Sundays are working days for all employees (not auto-holidays).
    holiday = 0
    badge_num = float(effective)
    if badge_num.is_integer():
        badge_num = int(badge_num)
    badge_den = num_days
    return {
        'present': present, 'absent': absent, 'half_day': half_day,
        'weekday_leave_days': weekday_leave_days,
        'present_effective': present_effective,
        'effective': effective,
        'holiday': holiday,
        'sunday_present': sunday_present,
        'sunday_half_day': sunday_half_day,
        'sunday_effective': sunday_effective,
        'total_working': num_days,
        'tracked': tracked, 'records': records, 'record_meta': record_meta, 'num_days': num_days,
        'sunday_dates': sunday_dates,
        'badge_num': badge_num,
        'display_badge_num': badge_num,
        'badge_den': badge_den,
    }
@payroll_bp.route('/employees')
def employees():
    q        = request.args.get('q', '').strip()
    status   = request.args.get('status', 'active').strip()
    location = request.args.get('location', '').strip()
    sort_by  = request.args.get('sort', 'id').strip().lower()
    if sort_by not in _EMPLOYEE_SORT_ORDERS:
        sort_by = 'id'
    order_by = _EMPLOYEE_SORT_ORDERS[sort_by]
    year, month = _period_from_source(request.args)
    payroll_redirect = _payroll_landing_redirect(get_current_user(), year, month)
    if payroll_redirect is not None:
        return payroll_redirect
    conn     = get_db()

    # Build query with all active filters
    conditions, params = [], []
    if status in ('active', 'inactive'):
        conditions.append("status = ?")
        params.append(status)
    if q:
        conditions.append("(name LIKE ? OR emp_code LIKE ? OR location LIKE ? OR mobile LIKE ?)")
        params.extend([f'%{q}%', f'%{q}%', f'%{q}%', f'%{q}%'])
    if location:
        conditions.append("location = ?")
        params.append(location)
    where = (" WHERE " + " AND ".join(conditions)) if conditions else ""
    rows = conn.execute(
        f"SELECT * FROM employees{where} ORDER BY {order_by}",
        tuple(params)
    ).fetchall()
    payroll_state = _get_payroll_month_state(conn, year, month)

    emps = []
    kpi_net = 0.0
    kpi_att_tracked = 0
    for r in rows:
        e = _attach_employee_month_context(conn, r, year, month, payroll_state=payroll_state)
        if e['att']['tracked']:
            kpi_att_tracked += 1
        kpi_net += e['net']
        emps.append(e)

    total_employees = conn.execute("SELECT COUNT(*) FROM employees WHERE status='active'").fetchone()[0]
    kpi_credits = float(conn.execute("SELECT COALESCE(SUM(amount),0) FROM credits").fetchone()[0])
    kpi_att_pct = round(kpi_att_tracked / total_employees * 100) if total_employees else 0
    # All active employees for search autocomplete
    ac_rows = conn.execute(
        "SELECT name, emp_code, location FROM employees WHERE status='active' ORDER BY name"
    ).fetchall()
    autocomplete_emps = [{'name': r['name'], 'emp_code': r['emp_code'] or '',
                           'location': r['location'] or ''}
                         for r in ac_rows]
    conn.close()
    return _emp_render('employees.html', employees=emps, search=q,
                       sel_year=year, sel_month=month,
                       month_name=calendar.month_name[month],
                       payroll_state=payroll_state,
                       sel_status=status, sel_location=location,
                       sel_sort=sort_by,
                       total_employees=total_employees,
                       kpi_att_tracked=kpi_att_tracked,
                       kpi_att_pct=kpi_att_pct,
                       kpi_net=round(kpi_net, 2),
                       kpi_credits=round(kpi_credits, 2),
                       autocomplete_emps=autocomplete_emps)


@payroll_bp.route('/report')
def report():
    year, month = _period_from_source(request.args)
    conn  = get_db()
    payroll_state = _get_payroll_month_state(conn, year, month)

    active_rows = conn.execute(
        f"SELECT * FROM employees WHERE status='active' ORDER BY {_EMPLOYEE_DISPLAY_ORDER}"
    ).fetchall()
    inactive_count = conn.execute("SELECT COUNT(*) FROM employees WHERE status='inactive'").fetchone()[0]

    total_gross = total_net = total_epf = total_esic = total_incentive = 0.0
    total_present = total_absent = total_half = tracked_count = 0
    emp_list = []

    for r in active_rows:
        e = _attach_employee_month_context(conn, r, year, month, payroll_state=payroll_state)
        if e['att']['tracked']:
            tracked_count += 1
            total_present += e['att']['present']
            total_absent  += float(e['att'].get('absent', 0) or 0)
            total_half    += e['att'].get('half_day_marked', e['att'].get('half_day', 0))
        total_gross += e['gross_salary']
        total_net   += e['net']
        total_epf   += e['epf']
        total_esic  += e['esic']
        total_incentive += float(e.get('tip_incentive') or 0)
        emp_list.append(e)

    active_count = len(active_rows)
    avg_gross = round(total_gross / active_count, 2) if active_count else 0
    avg_net   = round(total_net   / active_count, 2) if active_count else 0

    cr_row = conn.execute("SELECT COALESCE(SUM(amount),0), COUNT(*) FROM credits").fetchone()
    total_credits = float(cr_row[0])
    credit_count  = int(cr_row[1])
    conn.close()

    report_data = dict(
        total_count=active_count + inactive_count,
        active_count=active_count,
        inactive_count=inactive_count,
        avg_gross=avg_gross,
        avg_net=avg_net,
        total_gross=round(total_gross, 2),
        total_net=round(total_net, 2),
        total_epf=round(total_epf, 2),
        total_esic=round(total_esic, 2),
        total_incentive=round(total_incentive, 2),
        total_credits=round(total_credits, 2),
        credit_count=credit_count,
        total_present=total_present,
        total_absent=total_absent,
        total_half=total_half,
        tracked_count=tracked_count,
        employees=emp_list,
    )
    return _emp_render('employees.html', mode='report',
                       sel_year=year, sel_month=month,
                       month_name=calendar.month_name[month],
                       payroll_state=payroll_state,
                       report=report_data)


@payroll_bp.route('/add_employee', methods=['GET', 'POST'])
def add_employee():
    if request.method == 'POST':
        name        = request.form.get('name', '').strip()
        company     = request.form.get('company', '').strip() or _DEFAULT_COMPANY
        location    = request.form.get('location', '').strip()
        mobile           = request.form.get('mobile', '').strip()
        guardian_mobile  = request.form.get('guardian_mobile', '').strip()
        sex          = request.form.get('sex', '').strip()
        address     = request.form.get('address', '').strip()
        aadhar      = request.form.get('aadhar', '').strip()
        pan         = request.form.get('pan', '').strip()
        epf_number  = request.form.get('epf_number', '').strip()
        esic_number  = request.form.get('esic_number', '').strip()
        salary       = request.form.get('gross_salary', '0').strip()
        basic_salary = request.form.get('basic_salary', '0').strip()
        epf_amount   = request.form.get('epf_amount', '0').strip()
        esic_amount  = request.form.get('esic_amount', '0').strip()
        credit_repayment = request.form.get('credit_repayment', '0').strip()
        epf_exempt    = 1 if request.form.get('epf_exempt') else 0
        esic_exempt   = 1 if request.form.get('esic_exempt') else 0
        weekday_shift = request.form.get('weekday_shift', '').strip()
        sunday_shift  = request.form.get('sunday_shift', '').strip()
        bank_name           = request.form.get('bank_name', '').strip()
        account_holder_name = request.form.get('account_holder_name', '').strip()
        account_number      = request.form.get('account_number', '').strip()
        ifsc_code           = request.form.get('ifsc_code', '').strip().upper()
        status       = request.form.get('status', 'active').strip()
        total_off_raw = request.form.get('total_off', '0').strip()

        errors = []
        if not name:
            errors.append('Employee Name is required.')
        if not re.match(r'^\d{10}$', mobile):
            errors.append('Mobile number must be exactly 10 digits.')
        if guardian_mobile and not re.match(r'^\d{10}$', guardian_mobile):
            errors.append('Guardian mobile number must be exactly 10 digits.')
        if aadhar and not re.match(r'^\d{12}$', aadhar):
            errors.append('Aadhar number must be exactly 12 digits.')
        if pan and not re.match(r'^[A-Z]{5}\d{4}[A-Z]$', pan.upper()):
            errors.append('PAN must be in format ABCDE1234F.')
        if ifsc_code and not re.match(r'^[A-Z]{4}0[A-Z0-9]{6}$', ifsc_code):
            errors.append('IFSC Code must be in format ABCD0123456 (4 letters, 0, 6 alphanumeric).')
        try:
            salary_val = float(salary)
            if salary_val < 0:
                errors.append('Salary must be a positive number.')
        except ValueError:
            errors.append('Salary must be a valid number.')
            salary_val = 0
        try:
            basic_val = float(basic_salary) if basic_salary else 0
        except ValueError:
            basic_val = 0
        try:
            epf_val = float(epf_amount) if epf_amount and not epf_exempt else 0
        except ValueError:
            epf_val = 0
        if epf_val > _EPF_MAX:
            epf_val = _EPF_MAX
        try:
            esic_val = float(esic_amount) if esic_amount and not esic_exempt else 0
        except ValueError:
            esic_val = 0
        try:
            cr_val = float(credit_repayment) if credit_repayment else 0
        except ValueError:
            cr_val = 0
        try:
            total_off_val = int(total_off_raw) if total_off_raw else 0
            if total_off_val < 0 or total_off_val > 31:
                errors.append('Total Off must be between 0 and 31.')
        except ValueError:
            errors.append('Total Off must be a whole number.')
            total_off_val = 0
        if status not in ('active', 'inactive'):
            status = 'active'
        if sex not in ('Male', 'Female', ''):
            sex = ''
        if weekday_shift not in ('shift1', 'shift2', 'shift3', 'shift4', ''):
            weekday_shift = ''
        if sunday_shift not in ('shift1', 'shift2', 'shift3', ''):
            sunday_shift = ''

        conn = get_db()
        emp_code = _next_emp_code(conn)
        if _emp_code_taken(conn, emp_code):
            errors.append('Could not assign a unique Employee ID. Please try again.')

        if errors:
            form_data = dict(request.form)
            form_data['emp_code'] = emp_code
            conn.close()
            return _emp_render('employees.html', errors=errors, form=form_data,
                                   mode='add', employees=[], search='')

        conn.execute(
            "INSERT INTO employees (emp_code, name, company, location, mobile, guardian_mobile, sex, address, aadhar, pan, epf_number, esic_number, gross_salary, basic_salary, epf_amount, esic_amount, credit_repayment, epf_exempt, esic_exempt, weekday_shift, sunday_shift, bank_name, account_holder_name, account_number, ifsc_code, total_off, status) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (emp_code, name, company, location, mobile, guardian_mobile, sex, address, aadhar, pan.upper(), epf_number, esic_number, salary_val, basic_val, epf_val, esic_val, cr_val, epf_exempt, esic_exempt, weekday_shift, sunday_shift, bank_name, account_holder_name, account_number, ifsc_code, total_off_val, status)
        )
        conn.commit()
        conn.close()
        return redirect(url_for('employees'))

    conn = get_db()
    next_code = _next_emp_code(conn)
    conn.close()
    return _emp_render('employees.html', mode='add', employees=[], search='',
                       form={'emp_code': next_code})


@payroll_bp.route('/edit_employee/<int:emp_id>', methods=['GET', 'POST'])
def edit_employee(emp_id):
    conn = get_db()
    existing = conn.execute("SELECT * FROM employees WHERE id=?", (emp_id,)).fetchone()
    if not existing:
        conn.close()
        return redirect(url_for('employees'))

    if request.method == 'POST':
        # Employee ID is system-assigned and cannot be changed after create.
        emp_code    = (existing['emp_code'] or '').strip()
        name        = request.form.get('name', '').strip()
        company     = request.form.get('company', '').strip() or _DEFAULT_COMPANY
        location    = request.form.get('location', '').strip()
        mobile           = request.form.get('mobile', '').strip()
        guardian_mobile  = request.form.get('guardian_mobile', '').strip()
        sex          = request.form.get('sex', '').strip()
        address     = request.form.get('address', '').strip()
        aadhar      = request.form.get('aadhar', '').strip()
        pan         = request.form.get('pan', '').strip()
        epf_number  = request.form.get('epf_number', '').strip()
        esic_number  = request.form.get('esic_number', '').strip()
        salary       = request.form.get('gross_salary', '0').strip()
        basic_salary = request.form.get('basic_salary', '0').strip()
        epf_amount   = request.form.get('epf_amount', '0').strip()
        esic_amount  = request.form.get('esic_amount', '0').strip()
        if 'credit_repayment' in request.form:
            credit_repayment = request.form.get('credit_repayment', '0').strip()
        else:
            credit_repayment = str(existing['credit_repayment'] or 0)
        if 'weekday_shift' in request.form:
            weekday_shift = request.form.get('weekday_shift', '').strip()
        else:
            weekday_shift = existing['weekday_shift'] or ''
        if 'sunday_shift' in request.form:
            sunday_shift = request.form.get('sunday_shift', '').strip()
        else:
            sunday_shift = existing['sunday_shift'] or ''
        epf_exempt    = 1 if request.form.get('epf_exempt') else 0
        esic_exempt   = 1 if request.form.get('esic_exempt') else 0
        bank_name           = request.form.get('bank_name', '').strip()
        account_holder_name = request.form.get('account_holder_name', '').strip()
        account_number      = request.form.get('account_number', '').strip()
        ifsc_code           = request.form.get('ifsc_code', '').strip().upper()
        status       = request.form.get('status', 'active').strip()
        total_off_raw = request.form.get('total_off', '0').strip()

        errors = []
        if not name:
            errors.append('Employee Name is required.')
        if not re.match(r'^\d{10}$', mobile):
            errors.append('Mobile number must be exactly 10 digits.')
        if guardian_mobile and not re.match(r'^\d{10}$', guardian_mobile):
            errors.append('Guardian mobile number must be exactly 10 digits.')
        if aadhar and not re.match(r'^\d{12}$', aadhar):
            errors.append('Aadhar number must be exactly 12 digits.')
        if pan and not re.match(r'^[A-Z]{5}\d{4}[A-Z]$', pan.upper()):
            errors.append('PAN must be in format ABCDE1234F.')
        if ifsc_code and not re.match(r'^[A-Z]{4}0[A-Z0-9]{6}$', ifsc_code):
            errors.append('IFSC Code must be in format ABCD0123456 (4 letters, 0, 6 alphanumeric).')
        try:
            salary_val = float(salary)
            if salary_val < 0:
                errors.append('Salary must be a positive number.')
        except ValueError:
            errors.append('Salary must be a valid number.')
            salary_val = 0
        try:
            basic_val = float(basic_salary) if basic_salary else 0
        except ValueError:
            basic_val = 0
        try:
            epf_val = float(epf_amount) if epf_amount and not epf_exempt else 0
        except ValueError:
            epf_val = 0
        if epf_val > _EPF_MAX:
            epf_val = _EPF_MAX
        try:
            esic_val = float(esic_amount) if esic_amount and not esic_exempt else 0
        except ValueError:
            esic_val = 0
        try:
            cr_val = float(credit_repayment) if credit_repayment else 0
        except ValueError:
            cr_val = 0
        try:
            total_off_val = int(total_off_raw) if total_off_raw else 0
            if total_off_val < 0 or total_off_val > 31:
                errors.append('Total Off must be between 0 and 31.')
        except ValueError:
            errors.append('Total Off must be a whole number.')
            total_off_val = 0
        if status not in ('active', 'inactive'):
            status = 'active'
        if sex not in ('Male', 'Female', ''):
            sex = ''
        if weekday_shift not in ('shift1', 'shift2', 'shift3', 'shift4', ''):
            weekday_shift = ''
        if sunday_shift not in ('shift1', 'shift2', 'shift3', ''):
            sunday_shift = ''

        payroll_fields_locked = _employee_has_locked_month_data(conn, emp_id)
        if payroll_fields_locked and _wage_fields_changed(existing, {
            'gross_salary': salary_val,
            'basic_salary': basic_val,
            'epf_amount': epf_val,
            'esic_amount': esic_val,
            'epf_exempt': epf_exempt,
            'esic_exempt': esic_exempt,
            'total_off': total_off_val,
        }):
            errors.append(
                'This employee has data in a locked payroll month. '
                'Salary, statutory, and Total Off fields cannot be changed.'
            )

        if errors:
            form_data = dict(request.form)
            form_data['id'] = emp_id
            form_data['emp_code'] = emp_code
            conn.close()
            return _emp_render('employees.html', errors=errors, form=form_data,
                                   mode='edit', employees=[], search='',
                                   payroll_fields_locked=payroll_fields_locked)

        conn.execute(
            f"UPDATE employees SET emp_code=?, name=?, company=?, location=?, mobile=?, guardian_mobile=?, sex=?, address=?, aadhar=?, pan=?, epf_number=?, esic_number=?, gross_salary=?, basic_salary=?, epf_amount=?, esic_amount=?, credit_repayment=?, epf_exempt=?, esic_exempt=?, weekday_shift=?, sunday_shift=?, bank_name=?, account_holder_name=?, account_number=?, ifsc_code=?, total_off=?, status=?, updated_at={SQL_NOW} WHERE id=?",
            (emp_code, name, company, location, mobile, guardian_mobile, sex, address, aadhar, pan.upper(), epf_number, esic_number, salary_val, basic_val, epf_val, esic_val, cr_val, epf_exempt, esic_exempt, weekday_shift, sunday_shift, bank_name, account_holder_name, account_number, ifsc_code, total_off_val, status, emp_id)
        )
        conn.commit()
        conn.close()
        return redirect(url_for('employees'))

    payroll_fields_locked = _employee_has_locked_month_data(conn, emp_id)
    conn.close()
    return _emp_render('employees.html', mode='edit', form=dict(existing), employees=[], search='',
                       payroll_fields_locked=payroll_fields_locked)


@payroll_bp.route('/delete_employee/<int:emp_id>')
def delete_employee(emp_id):
    conn = get_db()
    if _employee_has_locked_month_data(conn, emp_id):
        conn.close()
        return _permission_denied_response(
            'This employee has data in a locked payroll month and cannot be deleted — including by administrators.'
        )
    conn.execute("DELETE FROM credits WHERE employee_id=?", (emp_id,))
    conn.execute("DELETE FROM attendance WHERE employee_id=?", (emp_id,))
    conn.execute("DELETE FROM employees WHERE id=?", (emp_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('employees'))


@payroll_bp.route('/attendance_overview')
def attendance_overview():
    """Attendance overview — all active employees for a given month."""
    year, month = _period_from_source(request.args)
    q        = request.args.get('q', '').strip()
    location = request.args.get('location', '').strip()
    conn  = get_db()
    user = get_current_user()

    base_conditions = ["status='active'"]
    base_params = []
    _append_attendance_scope_conditions(base_conditions, base_params, user)
    if location:
        base_conditions.append("location = ?")
        base_params.append(location)

    autocomplete_rows = conn.execute(
        f"SELECT name, emp_code, location FROM employees WHERE {' AND '.join(base_conditions)} ORDER BY {_EMPLOYEE_DISPLAY_ORDER}",
        tuple(base_params)
    ).fetchall()
    attendance_autocomplete_emps = [
        {
            'name': r['name'],
            'emp_code': r['emp_code'] or '',
            'location': r['location'] or '',
        }
        for r in autocomplete_rows
    ]

    conditions = list(base_conditions)
    params = list(base_params)
    if q:
        conditions.append("(name LIKE ? OR emp_code LIKE ?)")
        params.extend([f'%{q}%', f'%{q}%'])
    where = " WHERE " + " AND ".join(conditions)
    rows = conn.execute(
        f"SELECT * FROM employees{where} ORDER BY {_EMPLOYEE_DISPLAY_ORDER}",
        tuple(params)
    ).fetchall()

    emps = []
    tracked = total_present = total_absent = total_half = 0
    payroll_state = _get_payroll_month_state(conn, year, month)
    for r in rows:
        e = _attach_employee_month_context(conn, r, year, month, payroll_state=payroll_state)
        att = e['att']
        if att['tracked']:
            tracked += 1
            total_present += float(att.get('present_effective', 0) or 0)
            total_absent  += float(att.get('absent', 0) or 0)
            total_half    += att.get('half_day_marked', att.get('half_day', 0))
        emps.append(e)

    if float(total_present).is_integer():
        total_present = int(total_present)
    else:
        total_present = round(total_present, 1)
    if float(total_absent).is_integer():
        total_absent = int(total_absent)
    else:
        total_absent = round(total_absent, 1)

    if month == 1:   prev_y, prev_m = year - 1, 12
    else:            prev_y, prev_m = year, month - 1
    if month == 12:  next_y, next_m = year + 1, 1
    else:            next_y, next_m = year, month + 1
    date_view_date = _period_anchor_date(year, month).isoformat()
    allowed_locations = _attendance_filter_options(conn, user)

    conn.close()
    return _emp_render('employees.html', mode='attendance_list',
                       employees=emps, search=q,
                       sel_location=location,
                       locations=allowed_locations,
                       attendance_autocomplete_emps=attendance_autocomplete_emps,
                       sel_year=year, sel_month=month,
                       month_name=calendar.month_name[month],
                       tracked_count=tracked,
                       total_present=total_present,
                       total_absent=total_absent,
                       total_half=total_half,
                       date_view_date=date_view_date,
                       prev_y=prev_y, prev_m=prev_m,
                       next_y=next_y, next_m=next_m)


@payroll_bp.route('/attendance_date')
def attendance_date_view():
    """Date-wise attendance — mark all active employees for a single date."""
    from datetime import timedelta
    today_dt = date.today()
    date_str = request.args.get('date', _default_reporting_date(today_dt).isoformat()).strip()
    try:
        sel_dt = date.fromisoformat(date_str)
    except ValueError:
        sel_dt = today_dt
    date_str = sel_dt.isoformat()

    prev_date = (sel_dt - timedelta(days=1)).isoformat()
    next_date = (sel_dt + timedelta(days=1)).isoformat()

    is_sunday = sel_dt.weekday() == 6
    is_future = sel_dt > today_dt
    user = get_current_user()
    day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    day_name  = day_names[sel_dt.weekday()]

    location = request.args.get('location', '').strip()

    conn = get_db()
    _conds, _params = ["status='active'"], []
    _append_attendance_scope_conditions(_conds, _params, user)
    if location:
        _conds.append("location=?"); _params.append(location)
    rows = conn.execute(
        f"SELECT * FROM employees WHERE {' AND '.join(_conds)} ORDER BY {_EMPLOYEE_DISPLAY_ORDER}",
        _params
    ).fetchall()

    payroll_state = _get_payroll_month_state(conn, sel_dt.year, sel_dt.month)
    payroll_locked = bool(payroll_state['locked'])
    emps = []
    present_count = absent_count = half_count = unmarked_count = 0
    can_modify_attendance_date = False
    now_dt = datetime.now()
    for r in rows:
        e = dict(r)
        rec = None
        if is_future:
            e['date_status'] = ''
        else:
            rec = conn.execute(
                "SELECT status, updated_at FROM attendance WHERE employee_id=? AND date=?",
                (e['id'], date_str)
            ).fetchone()
            e['date_status'] = rec['status'] if rec else ''
        e['can_modify_date_status'] = (not is_future) and _can_modify_attendance_record(
            user, sel_dt, rec, today=today_dt, now=now_dt, payroll_locked=payroll_locked
        )
        if e['can_modify_date_status']:
            can_modify_attendance_date = True
        if e['date_status'] == 'present':    present_count  += 1
        elif e['date_status'] == 'absent':   absent_count   += 1
        elif e['date_status'] == 'half_day': half_count     += 1
        else:                                unmarked_count += 1
        emps.append(e)

    allowed_locations = _attendance_filter_options(conn, user)
    conn.close()
    return _emp_render('employees.html', mode='attendance_date',
                       employees=emps,
                       sel_date=date_str,
                       sel_year=sel_dt.year, sel_month=sel_dt.month,
                       month_name=calendar.month_name[sel_dt.month],
                       day_name=day_name,
                       prev_date=prev_date,
                       next_date=next_date,
                       is_sunday=is_sunday,
                       is_future=is_future,
                       can_modify_attendance_date=can_modify_attendance_date,
                       attendance_date_lock_message=_attendance_date_lock_message(
                           today_dt,
                           payroll_locked=payroll_locked,
                           year=sel_dt.year,
                           month=sel_dt.month,
                       ),
                       payroll_state=payroll_state,
                       today=today_dt.isoformat(),
                       present_count=present_count,
                       absent_count=absent_count,
                       half_count=half_count,
                       unmarked_count=unmarked_count,
                       sel_location=location,
                       locations=allowed_locations)


@payroll_bp.route('/attendance/<int:emp_id>')
def attendance(emp_id):
    """Show monthly attendance calendar for an employee."""
    year, month = _period_from_source(request.args)
    user = get_current_user()
    conn = get_db()
    emp = conn.execute("SELECT * FROM employees WHERE id=?", (emp_id,)).fetchone()
    if not emp:
        conn.close()
        return redirect(url_for('employees'))
    if not _user_can_access_attendance_employee(conn, emp_id):
        conn.close()
        return _permission_denied_response('You do not have access to this employee attendance.')
    payroll_state = _get_payroll_month_state(conn, year, month)
    payroll_locked = bool(payroll_state['locked'])
    emp = _attach_employee_month_context(conn, emp, year, month, payroll_state=payroll_state)
    att = emp['att']

    # Build calendar weeks
    _, num_days = calendar.monthrange(year, month)
    first_weekday = calendar.weekday(year, month, 1)  # 0=Mon
    cal_weeks = []
    week = [None] * first_weekday
    today_dt = date.today()
    now_dt = datetime.now()
    for day in range(1, num_days + 1):
        d_str = f'{year}-{month:02d}-{day:02d}'
        status = att['records'].get(d_str, '')
        att_dt = date(year, month, day)
        record = att.get('record_meta', {}).get(d_str)
        can_edit = _can_modify_attendance_record(
            user, att_dt, record, today=today_dt, now=now_dt, payroll_locked=payroll_locked
        )
        week.append({'day': day, 'date': d_str, 'status': status, 'can_edit': can_edit})
        if len(week) == 7:
            cal_weeks.append(week)
            week = []
    if week:
        week += [None] * (7 - len(week))
        cal_weeks.append(week)

    # Previous/next month
    if month == 1:
        prev_y, prev_m = year - 1, 12
    else:
        prev_y, prev_m = year, month - 1
    if month == 12:
        next_y, next_m = year + 1, 1
    else:
        next_y, next_m = year, month + 1

    conn.close()
    return _emp_render('employees.html', mode='attendance', emp=emp,
                       cal_weeks=cal_weeks, sel_year=year, sel_month=month,
                       month_name=calendar.month_name[month],
                       payroll_state=payroll_state,
                       prev_y=prev_y, prev_m=prev_m,
                       next_y=next_y, next_m=next_m,
                       today=date.today().isoformat(),
                       can_modify_past_attendance=(
                           bool(user and user.get('is_admin')) and not payroll_locked
                       ),
                       attendance_date_lock_message=_attendance_date_lock_message(
                           payroll_locked=payroll_locked, year=year, month=month
                       ),
                       employees=[], search='')


@payroll_bp.route('/mark_attendance', methods=['POST'])
def mark_attendance():
    """Mark attendance for an employee on a specific date."""
    emp_id   = int(request.form.get('employee_id', 0))
    att_date = request.form.get('date', '').strip()
    status   = request.form.get('status', '').strip()
    year, month = _period_from_source(request.form)

    if not emp_id or not att_date or status not in ('present', 'absent', 'half_day', ''):
        return redirect(url_for('attendance', emp_id=emp_id, year=year, month=month))

    att_dt = _parse_attendance_date(att_date)
    if not att_dt:
        return redirect(url_for('attendance', emp_id=emp_id, year=year, month=month))
    today_dt = date.today()
    if att_dt > today_dt:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': 'Future attendance cannot be marked.'}), 400
        return redirect(url_for('attendance', emp_id=emp_id, year=year, month=month))
    user = get_current_user()

    conn = get_db()
    if not _user_can_access_attendance_employee(conn, emp_id):
        conn.close()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': 'You do not have access to this employee attendance.'}), 403
        return _permission_denied_response('You do not have access to this employee attendance.')
    payroll_locked = _is_payroll_month_locked(conn, att_dt.year, att_dt.month)
    if payroll_locked:
        conn.close()
        message = _payroll_month_frozen_message(att_dt.year, att_dt.month)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': message, 'locked': True}), 403
        return _permission_denied_response(message)
    rec = conn.execute(
        "SELECT status, updated_at FROM attendance WHERE employee_id=? AND date=?",
        (emp_id, att_date)
    ).fetchone()
    if not _can_modify_attendance_record(
        user, att_dt, rec, today=today_dt, now=datetime.now(), payroll_locked=payroll_locked
    ):
        conn.close()
        message = _attendance_date_lock_message(today_dt)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': message}), 403
        return _permission_denied_response(message)
    # Sundays are regular working days — mark present / half_day / absent the same as other days.
    if status == '':
        conn.execute("DELETE FROM attendance WHERE employee_id=? AND date=?", (emp_id, att_date))
    else:
        conn.execute(
            f"INSERT INTO attendance (employee_id, date, status, created_at, updated_at) VALUES (?,?,?,{SQL_NOW},{SQL_NOW}) "
            f"ON CONFLICT(employee_id, date) DO UPDATE SET status=excluded.status, updated_at={SQL_NOW}",
            (emp_id, att_date, status)
        )
    conn.commit()
    conn.close()

    # Return JSON for AJAX calls
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'ok': True})
    return redirect(url_for('attendance', emp_id=emp_id, year=year, month=month))


@payroll_bp.route('/bulk_attendance/<int:emp_id>', methods=['POST'])
def bulk_attendance(emp_id):
    """Mark all unmarked days in a month as present."""
    year, month = _period_from_source(request.form)
    action = request.form.get('action', 'mark_all_present')

    conn = get_db()
    user = get_current_user()
    if not _user_can_access_attendance_employee(conn, emp_id):
        conn.close()
        return _permission_denied_response('You do not have access to this employee attendance.')
    if _is_payroll_month_locked(conn, year, month):
        conn.close()
        return _permission_denied_response(_payroll_month_frozen_message(year, month))
    if not (user and user.get('is_admin')):
        conn.close()
        return _permission_denied_response(_attendance_date_lock_message())
    _, num_days = calendar.monthrange(year, month)
    today = date.today()
    if (year, month) > (today.year, today.month):
        last_markable_day = 0
    elif (year, month) == (today.year, today.month):
        last_markable_day = today.day
    else:
        last_markable_day = num_days

    if action == 'mark_all_present':
        for day in range(1, last_markable_day + 1):
            d_str = f'{year}-{month:02d}-{day:02d}'
            conn.execute(
                f"INSERT INTO attendance (employee_id, date, status, created_at, updated_at) VALUES (?,?,?,{SQL_NOW},{SQL_NOW}) "
                "ON CONFLICT(employee_id, date) DO NOTHING",
                (emp_id, d_str, 'present')
            )
    elif action == 'clear_all':
        conn.execute(
            "DELETE FROM attendance WHERE employee_id=? AND date LIKE ?",
            (emp_id, f'{year}-{month:02d}-%')
        )
    conn.commit()
    conn.close()
    return redirect(url_for('attendance', emp_id=emp_id, year=year, month=month))


def _get_employee_credits(conn, emp_id, year=None, month=None, date_from=None, date_to=None):
    """Return credits for an employee, optionally filtered by month or date range."""
    if date_from and date_to:
        rows = conn.execute(
            "SELECT * FROM credits WHERE employee_id=? AND date>=? AND date<=? ORDER BY date DESC",
            (emp_id, date_from, date_to)
        ).fetchall()
    elif year and month:
        rows = conn.execute(
            "SELECT * FROM credits WHERE employee_id=? AND date LIKE ? ORDER BY date DESC",
            (emp_id, f'{year}-{month:02d}-%')
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM credits WHERE employee_id=? ORDER BY date DESC",
            (emp_id,)
        ).fetchall()
    credits_list = [dict(r) for r in rows]
    total = sum(c['amount'] for c in credits_list)
    return {'items': credits_list, 'total': round(total, 2)}


def _get_total_credits(conn, emp_id):
    """Return overall total credit amount for an employee."""
    row = conn.execute(
        "SELECT COALESCE(SUM(amount), 0) as total FROM credits WHERE employee_id=?",
        (emp_id,)
    ).fetchone()
    return round(row['total'], 2)


@payroll_bp.route('/credits')
def credits_dashboard():
    """Standalone Credit Dashboard — all employees with credit history."""
    year, month = _period_from_source(request.args)
    conn = get_db()
    payroll_state = _get_payroll_month_state(conn, year, month)

    credit_emps = conn.execute("""
        SELECT e.*,
               COALESCE(SUM(c.amount), 0)        AS credit_balance,
               COUNT(c.id)                       AS credit_entries
        FROM employees e
        LEFT JOIN credits c ON e.id = c.employee_id
        GROUP BY e.id
        HAVING COUNT(c.id) > 0
        ORDER BY
            CASE WHEN COALESCE(SUM(c.amount), 0) > 0 THEN 0 ELSE 1 END,
            LOWER(e.company), LOWER(e.location), LOWER(e.name), e.id DESC
    """).fetchall()

    recent = conn.execute("""
        SELECT c.id, c.date, c.description, c.amount,
               e.id AS emp_id, e.name AS emp_name, e.company
        FROM credits c
        JOIN employees e ON c.employee_id = e.id
        ORDER BY c.date DESC, c.id DESC
        LIMIT 15
    """).fetchall()
    credit_emps = [
        {
            **_attach_employee_month_context(conn, row, year, month, payroll_state=payroll_state),
            'credit_balance': _round_half_up(row['credit_balance'] or 0, 2),
            'credit_entries': int(row['credit_entries'] or 0),
        }
        for row in credit_emps
    ]
    total_credit_amount = _round_half_up(sum(e['credit_balance'] for e in credit_emps), 2)
    total_credit_entries = int(conn.execute(
        "SELECT COUNT(*) FROM credits"
    ).fetchone()[0])
    employees_with_credit = len(credit_emps)
    all_employees = conn.execute(
        f"SELECT id, name, emp_code, company FROM employees WHERE status='active' ORDER BY {_EMPLOYEE_DISPLAY_ORDER}"
    ).fetchall()
    recent_credits = _annotate_credit_editability(conn, recent)
    conn.close()

    return _emp_render('employees.html', mode='credits_dashboard',
                       credit_emps=credit_emps,
                       recent_credits=recent_credits,
                       total_credit_amount=total_credit_amount,
                       total_credit_entries=total_credit_entries,
                       employees_with_credit=employees_with_credit,
                       all_employees=[dict(r) for r in all_employees],
                       sel_year=year, sel_month=month,
                       month_name=calendar.month_name[month],
                       payroll_state=payroll_state)


@payroll_bp.route('/add_credit_global', methods=['POST'])
def add_credit_global():
    """Add a credit or repayment entry from the Credit Dashboard."""
    emp_id      = request.form.get('employee_id', '0').strip()
    cr_date     = request.form.get('date', '').strip()
    description = request.form.get('description', '').strip()
    amount      = request.form.get('amount', '0').strip()
    txn_type    = request.form.get('transaction_type', 'credit').strip().lower()

    year, month = _period_from_source(request.form)

    try:
        emp_id_val = int(emp_id)
        raw_amount = abs(float(amount))
    except (ValueError, TypeError):
        return redirect(url_for('credits_dashboard', year=year, month=month))

    if not cr_date or raw_amount <= 0 or emp_id_val <= 0:
        return redirect(url_for('credits_dashboard', year=year, month=month))

    is_repayment = txn_type == 'repayment'
    amount_val = -raw_amount if is_repayment else raw_amount
    entry_type = 'manual_repayment' if is_repayment else 'manual'
    if is_repayment and not description:
        description = 'Repayment'

    conn = get_db()
    if _is_credit_date_locked(conn, cr_date):
        conn.close()
        return redirect(url_for('credits_dashboard', year=year, month=month))
    emp = conn.execute("SELECT id FROM employees WHERE id=?", (emp_id_val,)).fetchone()
    if emp:
        conn.execute(
            "INSERT INTO credits (employee_id, date, description, amount, entry_type) VALUES (?,?,?,?,?)",
            (emp_id_val, cr_date, description, amount_val, entry_type)
        )
        conn.commit()
    conn.close()
    return redirect(url_for('credits_dashboard', year=year, month=month))


@payroll_bp.route('/credits/<int:emp_id>')
def employee_credits(emp_id):
    """Show credits/advances for an employee."""
    date_from = (request.args.get('date_from') or '').strip()
    date_to = (request.args.get('date_to') or '').strip()
    active_date_filter = bool(date_from and date_to)
    year, month = _period_from_source(request.args)
    if active_date_filter:
        try:
            period_end = datetime.strptime(date_to, '%Y-%m-%d').date()
            year, month = period_end.year, period_end.month
        except ValueError:
            active_date_filter = False
            date_from = date_to = ''
    conn = get_db()
    emp = conn.execute("SELECT * FROM employees WHERE id=?", (emp_id,)).fetchone()
    if not emp:
        conn.close()
        return redirect(url_for('employees'))
    payroll_state = _get_payroll_month_state(conn, year, month)
    emp = _attach_employee_month_context(conn, emp, year, month, payroll_state=payroll_state)
    cr = _get_employee_credits(
        conn, emp_id,
        date_from=date_from if active_date_filter else None,
        date_to=date_to if active_date_filter else None,
    )
    credit_items = _annotate_credit_editability(conn, cr['items'])
    overall_total = _get_total_credits(conn, emp_id)
    today_iso = date.today().isoformat()
    conn.close()
    return _emp_render('employees.html', mode='credits', emp=emp,
                       credit_items=credit_items, credit_total=cr['total'],
                       overall_credit=overall_total,
                       sel_year=year, sel_month=month,
                       month_name=calendar.month_name[month],
                       payroll_state=payroll_state,
                       date_from=date_from if active_date_filter else '',
                       date_to=date_to if active_date_filter else '',
                       active_date_filter=active_date_filter,
                       today_iso=today_iso,
                       employees=[], search='')


@payroll_bp.route('/add_credit/<int:emp_id>', methods=['POST'])
def add_credit(emp_id):
    """Add a credit/advance or repayment entry for an employee."""
    year, month = _period_from_source(request.form)
    cr_date     = request.form.get('date', '').strip()
    description = request.form.get('description', '').strip()
    amount      = request.form.get('amount', '0').strip()
    txn_type    = request.form.get('transaction_type', 'credit').strip().lower()

    try:
        raw_amount = abs(float(amount))
    except ValueError:
        raw_amount = 0

    if not cr_date or raw_amount <= 0:
        return redirect(url_for('employee_credits', emp_id=emp_id, year=year, month=month))

    is_repayment = txn_type == 'repayment'
    amount_val = -raw_amount if is_repayment else raw_amount
    entry_type = 'manual_repayment' if is_repayment else 'manual'
    if is_repayment and not description:
        description = 'Repayment'

    conn = get_db()
    if _is_credit_date_locked(conn, cr_date):
        conn.close()
        return redirect(url_for('employee_credits', emp_id=emp_id, year=year, month=month))
    conn.execute(
        "INSERT INTO credits (employee_id, date, description, amount, entry_type) VALUES (?,?,?,?,?)",
        (emp_id, cr_date, description, amount_val, entry_type)
    )
    conn.commit()
    conn.close()
    return redirect(url_for('employee_credits', emp_id=emp_id, year=year, month=month))


@payroll_bp.route('/edit_credit/<int:credit_id>', methods=['POST'])
def edit_credit(credit_id):
    """Edit an existing credit entry."""
    year, month = _period_from_source(request.form)
    cr_date     = request.form.get('date', '').strip()
    description = request.form.get('description', '').strip()
    amount      = request.form.get('amount', '0').strip()

    try:
        amount_val = float(amount)
    except ValueError:
        amount_val = 0

    conn = get_db()
    row = conn.execute(
        "SELECT employee_id, date FROM credits WHERE id=?", (credit_id,)
    ).fetchone()
    if row and cr_date and amount_val != 0:
        emp_id = row['employee_id']
        if _is_credit_date_locked(conn, row['date']) or _is_credit_date_locked(conn, cr_date):
            conn.close()
            return redirect(url_for('employee_credits', emp_id=emp_id, year=year, month=month))
        conn.execute(
            "UPDATE credits SET date=?, description=?, amount=? WHERE id=?",
            (cr_date, description, amount_val, credit_id)
        )
        conn.commit()
        conn.close()
        return redirect(url_for('employee_credits', emp_id=emp_id, year=year, month=month))
    conn.close()
    return redirect(url_for('employees'))


@payroll_bp.route('/delete_credit/<int:credit_id>')
def delete_credit(credit_id):
    """Delete a credit entry."""
    year, month = _period_from_source(request.args)
    conn = get_db()
    row = conn.execute(
        "SELECT employee_id, date FROM credits WHERE id=?", (credit_id,)
    ).fetchone()
    if row:
        emp_id = row['employee_id']
        if _is_credit_date_locked(conn, row['date']):
            conn.close()
            return redirect(url_for('employee_credits', emp_id=emp_id, year=year, month=month))
        conn.execute("DELETE FROM credits WHERE id=?", (credit_id,))
        conn.commit()
        conn.close()
        return redirect(url_for('employee_credits', emp_id=emp_id, year=year, month=month))
    conn.close()
    return redirect(url_for('employees'))


@payroll_bp.route('/download_employee_template')
def download_employee_template():
    """Download a pre-built Excel template for employee import."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = 'Employee Import Template'

    # ── Column definitions ──────────────────────────────────────────────────
    # (header, description, sample_value, is_required, has_dropdown)
    columns = [
        ('EMP ID',           'Employee Code / ID\n(optional, must be unique)',           'NL-001',                          False, False),
        ('NAME',             'Full Employee Name\n(REQUIRED)',                            'Ravi Kumar',                      True,  False),
        ('COMPANY',          'Company / Brand Name\nSelect from dropdown',               'NL',                              False, True),
        ('LOCATION',         'Work Location / Branch\nSelect from dropdown',             'Gandhi Market',                   False, True),
        ('MOBILE',           '10-digit Mobile Number\n(digits only, no spaces)',         '9876543210',                      False, False),
        ('GUARDIAN MOBILE',  'Guardian / Emergency Contact\n10-digit number (optional)', '9123456789',                      False, False),
        ('SEX',              'Employee Gender\nSelect from dropdown',                    'Male',                            False, True),
        ('ADDRESS',          'Home / Residential Address\n(optional)',                   '12 Main St, Chennai',             False, False),
        ('AADHAR',           '12-digit Aadhar Number\n(digits only)',                    '123456789012',                    False, False),
        ('PAN',              'PAN Card Number\n(format: ABCDE1234F)',                    'ABCDE1234F',                      False, False),
        ('EPF NUMBER',       'EPF Registration Number\n(optional)',                      'TN/CHN/12345',                    False, False),
        ('ESIC NUMBER',      'ESIC Registration Number\n(optional)',                     'ESI/123456',                      False, False),
        ('SALARY',           'Gross Monthly Salary\n(numbers only, no Rs symbol)',       '30000',                           False, False),
        ('BASIC PAY',        'Basic Pay Component\n(numbers only)',                      '15000',                           False, False),
        ('EPF',              'EPF Deduction Amount\n(numbers only)',                     '1800',                            False, False),
        ('ESIC',             'ESIC Deduction Amount\n(numbers only)',                    '525',                             False, False),
        ('CREDIT REPAYMENT', 'Monthly Credit Repayment\n(numbers only, 0 if none)',      '0',                               False, False),
        ('STATUS',           'Employment Status\nSelect from dropdown (default: active)','active',                          False, True),
        ('WEEKDAY SHIFT',    'Weekday working shift\nSelect from dropdown (optional)',   'Shift 1 - 8:30 AM to 7:30 PM',   False, True),
        ('SUNDAY SHIFT',     'Sunday shift with fixed pay\nSelect from dropdown (opt.)','Shift 1 - 9:00 AM to 7:00 PM',   False, True),
        ('BANK NAME',        'Name of the employee\'s bank\n(optional)',                 'State Bank of India',             False, False),
        ('ACCOUNT HOLDER NAME', 'Account holder name as per bank\n(optional)',          'Ravi Kumar',                      False, False),
        ('ACCOUNT NUMBER',   'Bank account number\n(digits only, optional)',             '1234567890',                      False, False),
        ('IFSC CODE',        'Branch IFSC Code\n(format: ABCD0123456)',                 'SBIN0001234',                     False, False),
    ]

    # ── Hidden "Lists" sheet for dropdown values ─────────────────────────────
    ls = wb.create_sheet('Lists')
    ls.sheet_state = 'hidden'

    company_list   = ['NL', 'TLNT']
    location_list  = list(_PAYROLL_DEPARTMENTS)
    sex_list       = ['Male', 'Female']
    status_list    = ['active', 'inactive']
    wd_shift_list  = [
        'Shift 1 - 8:30 AM to 7:30 PM',
        'Shift 2 - 8:30 AM to 7:50 PM',
        'Shift 3 - 8:45 AM to 7:50 PM',
        'Shift 4 - 9:00 AM to 8:30 PM',
    ]
    sun_shift_list = [
        'Shift 1 - 9:00 AM to 7:00 PM (Rs.550)',
        'Shift 2 - 9:00 AM to 8:00 PM (Rs.600)',
        'Shift 3 - 9:00 AM to 8:30 PM (Rs.650)',
    ]
    list_cols = [company_list, location_list, sex_list, status_list, wd_shift_list, sun_shift_list]
    list_names = ['Company', 'Location', 'Sex', 'Status', 'WeekdayShift', 'SundayShift']
    for c_idx, (name, vals) in enumerate(zip(list_names, list_cols), 1):
        ls.cell(row=1, column=c_idx, value=name).font = Font(bold=True)
        for r_idx, v in enumerate(vals, 2):
            ls.cell(row=r_idx, column=c_idx, value=v)

    # Column letters in Lists sheet
    def _lists_range(col_idx, count):
        col = get_column_letter(col_idx)
        return f"Lists!${col}$2:${col}${count + 1}"

    company_range  = _lists_range(1, len(company_list))
    location_range = _lists_range(2, len(location_list))
    sex_range      = _lists_range(3, len(sex_list))
    status_range   = _lists_range(4, len(status_list))
    wd_range       = _lists_range(5, len(wd_shift_list))
    sun_range      = _lists_range(6, len(sun_shift_list))

    # Styles
    hdr_fill    = PatternFill('solid', fgColor='1F4E79')
    hdr_font    = Font(name='Segoe UI', bold=True, color='FFFFFF', size=10)
    req_fill    = PatternFill('solid', fgColor='DC2626')
    req_font    = Font(name='Segoe UI', bold=True, color='FFFFFF', size=10)
    drop_fill   = PatternFill('solid', fgColor='1F4E79')  # uniform brand header colour
    drop_font   = Font(name='Segoe UI', bold=True, color='FFFFFF', size=10)
    desc_fill   = PatternFill('solid', fgColor='EEF5FF')
    desc_font   = Font(name='Segoe UI', color='1560D4', size=9, italic=True)
    ddesc_fill  = PatternFill('solid', fgColor='E8F0FE')
    ddesc_font  = Font(name='Segoe UI', color='0D47A1', size=9, italic=True)
    data_font   = Font(name='Segoe UI', size=10)
    data2_font  = Font(name='Segoe UI', size=10, color='94A3B8', italic=True)
    thin        = Side(style='thin', color='CBD5E0')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # ── Row 1: Title banner ──────────────────────────────────────────────────
    ws.merge_cells(f'A1:{get_column_letter(len(columns))}1')
    title_cell = ws['A1']
    title_cell.value = '  NEERAJ TEXTILE — EMPLOYEE IMPORT TEMPLATE    |    Fill data from Row 7 onwards  ·  Use dropdowns where shown  ·  Required: NAME column'
    title_cell.font  = Font(name='Segoe UI', bold=True, color='FFFFFF', size=11)
    title_cell.fill  = PatternFill('solid', fgColor='0F172A')
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24

    # ── Row 2: Column headers ────────────────────────────────────────────────
    for col_idx, (header, desc, sample, required, has_dd) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        if required:
            cell.font, cell.fill = req_font, req_fill
        elif has_dd:
            cell.font, cell.fill = drop_font, drop_fill
        else:
            cell.font, cell.fill = hdr_font, hdr_fill
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[2].height = 26

    # ── Row 3: Descriptions ──────────────────────────────────────────────────
    for col_idx, (header, desc, sample, required, has_dd) in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=desc)
        if has_dd:
            cell.font, cell.fill = ddesc_font, ddesc_fill
        else:
            cell.font, cell.fill = desc_font, desc_fill
        cell.alignment = left_wrap
        cell.border    = border
    ws.row_dimensions[3].height = 40

    # ── Row 4: Dropdown legend ───────────────────────────────────────────────
    ws.merge_cells(f'A4:{get_column_letter(len(columns))}4')
    leg = ws['A4']
    leg.value = (
        'DROPDOWN COLUMNS (dark blue headers) — Click a cell and use the dropdown arrow to pick a value:  '
        'COMPANY: NL / TLNT  |  LOCATION: Gandhi Market / Bathubasthi / New Main  |  STATUS: active / inactive  |  '
        'WEEKDAY SHIFT: 4 shifts  |  SUNDAY SHIFT: 3 shifts (fixed Sunday pay)'
    )
    leg.font      = Font(name='Segoe UI', size=9, color='0D47A1', italic=True)
    leg.fill      = PatternFill('solid', fgColor='E8F0FE')
    leg.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[4].height = 30

    # ── Row 5-6: Sample data rows ────────────────────────────────────────────
    sample_rows = [
        ['NL-001', 'Ravi Kumar',     'NL',   'Gandhi Market', '9876543210', '9123456789', 'Male',
         '12 Main St, Chennai', '123456789012', 'ABCDE1234F', 'TN/CHN/12345', 'ESI/123456',
         30000, 15000, 1800, 525, 0, 'active',
         'Shift 1 - 8:30 AM to 7:30 PM', 'Shift 1 - 9:00 AM to 7:00 PM (Rs.550)',
         'State Bank of India', 'Ravi Kumar', '1234567890', 'SBIN0001234'],
        ['NL-002', 'Priya Sundaram', 'TLNT', 'Bathubasthi',   '8765432109', '', 'Female',
         '',                    '',              '',          '',            '',
         25000, 12000, 1440, 438, 500, 'active',
         'Shift 2 - 8:30 AM to 7:50 PM', '',
         '', '', '', ''],
    ]
    note_fill = PatternFill('solid', fgColor='F8FAFC')
    for r_idx, row_data in enumerate(sample_rows, 5):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.font      = data2_font if r_idx == 6 else data_font
            cell.fill      = note_fill
            cell.alignment = left_wrap
            cell.border    = border
        ws.row_dimensions[r_idx].height = 18

    # ── Row 7: "Add your data below" marker ─────────────────────────────────
    ws.merge_cells(f'A7:{get_column_letter(len(columns))}7')
    marker = ws['A7']
    marker.value     = '▼  Add your employee data from this row onwards  ▼'
    marker.font      = Font(name='Segoe UI', bold=True, color='16A34A', size=10)
    marker.fill      = PatternFill('solid', fgColor='F0FDF4')
    marker.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[7].height = 20

    # ── Data validation dropdowns (rows 8 to 1007) ──────────────────────────
    DATA_START = 8
    DATA_END   = 1007

    # Map header → Lists range
    col_map = {}
    for col_idx, (header, _, _, _, _) in enumerate(columns, 1):
        col_map[header] = get_column_letter(col_idx)

    def _dv(formula, col_letter, field_name):
        dv = DataValidation(
            type='list',
            formula1=formula,
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorStyle='stop',
            errorTitle='Invalid Value',
            error=f'"{field_name}" only accepts values from the dropdown list. '
                  f'Please click the cell and select an option from the dropdown arrow.',
            showInputMessage=True,
            promptTitle=field_name,
            prompt=f'Click the dropdown arrow to select a valid {field_name} option.',
        )
        dv.sqref = f'{col_letter}{DATA_START}:{col_letter}{DATA_END}'
        ws.add_data_validation(dv)

    _dv(f'={company_range}',  col_map['COMPANY'],       'Company')
    _dv(f'={location_range}', col_map['LOCATION'],      'Work Location')
    _dv(f'={sex_range}',      col_map['SEX'],           'Sex')
    _dv(f'={status_range}',   col_map['STATUS'],        'Status')
    _dv(f'={wd_range}',       col_map['WEEKDAY SHIFT'], 'Weekday Shift')
    _dv(f'={sun_range}',      col_map['SUNDAY SHIFT'],  'Sunday Shift')

    # ── Column widths ────────────────────────────────────────────────────────
    col_widths = [12, 24, 12, 18, 14, 18, 10, 28, 16, 14, 18, 16, 12, 12, 10, 10, 18, 10, 32, 36]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes so headers stay visible while scrolling
    ws.freeze_panes = 'A8'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name='Employee_Import_Template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@payroll_bp.route('/upload_employees', methods=['POST'])
def upload_employees():
    """Import employees from an uploaded Excel file.

    Expected columns (case-insensitive, spaces trimmed):
        EMP ID, NAME, COMPANY, LOCATION, MOBILE, ADDRESS, AADHAR, PAN,
        EPF NUMBER, ESIC NUMBER, SALARY, BASIC PAY, EPF, ESIC,
        CREDIT REPAYMENT, STATUS

    Rows with a blank NAME are skipped. All other validation errors
    are counted but do not abort the import.
    """
    f = request.files.get('file')
    if not f or not f.filename:
        return redirect(url_for('employees'))

    import pandas as pd

    def _norm_header(value):
        return str(value or '').strip().upper()

    def _load_employee_import_df(fileobj):
        # First pass: inspect raw rows so we can detect the actual header row in
        # the styled template workbook instead of assuming row 1 contains headers.
        raw = pd.read_excel(fileobj, dtype=str, header=None, sheet_name=0)
        header_idx = None
        marker_idx = None
        for idx, row in raw.iterrows():
            cells = [_norm_header(v) for v in row.tolist()]
            cell_set = {c for c in cells if c and c != 'NAN'}
            if header_idx is None and 'NAME' in cell_set:
                if len(cell_set & {'NAME', 'COMPANY', 'LOCATION', 'SALARY'}) >= 2:
                    header_idx = idx
            if marker_idx is None:
                joined = ' '.join(cells)
                if 'ADD YOUR EMPLOYEE DATA FROM THIS ROW ONWARDS' in joined:
                    marker_idx = idx
        if header_idx is None:
            header_idx = 0

        fileobj.seek(0)
        df = pd.read_excel(fileobj, dtype=str, header=header_idx, sheet_name=0)
        df.columns = [_norm_header(c) for c in df.columns]

        if marker_idx is not None and marker_idx >= header_idx:
            first_data_pos = marker_idx - header_idx
            df = df.iloc[first_data_pos:].copy()

        # Drop template helper rows and fully empty rows.
        if 'NAME' in df.columns:
            name_series = df['NAME'].astype(str).str.strip()
            helper_mask = (
                name_series.str.upper().isin({
                    '',
                    'NAN',
                    'FULL EMPLOYEE NAME\n(REQUIRED)'.upper(),
                })
                | name_series.str.contains('FULL EMPLOYEE NAME', case=False, na=False)
                | name_series.str.contains('ADD YOUR EMPLOYEE DATA FROM THIS ROW ONWARDS', case=False, na=False)
            )
            df = df.loc[~helper_mask].copy()

        return df

    try:
        df = _load_employee_import_df(f)
    except Exception:
        return redirect(url_for('employees', import_ok=0, import_err=1))

    def _flt(row, key, default=0.0):
        try:
            v = str(row.get(key, '') or '').strip()
            return float(v) if v and v.lower() != 'nan' else default
        except (ValueError, TypeError):
            return default

    def _str(row, key):
        v = str(row.get(key, '') or '').strip()
        return '' if v.lower() == 'nan' else v

    def _parse_shift(val, valid_keys):
        """Map a dropdown label like 'Shift 2 - ...' to 'shift2', or '' if blank."""
        v = str(val or '').strip().lower()
        if not v or v == 'nan':
            return ''
        for i, _ in enumerate(valid_keys, 1):
            if f'shift {i}' in v or v == f'shift{i}':
                return f'shift{i}'
        return ''

    # Validate required columns are present
    required = {'NAME'}
    missing = required - set(df.columns)
    if missing:
        return redirect(url_for('employees', import_ok=0, import_err=len(df)))

    conn = get_db()
    inserted = 0
    skipped = 0
    for _, row in df.iterrows():
        name = _str(row, 'NAME')
        if not name:
            skipped += 1
            continue
        try:
            emp_code         = _str(row, 'EMP ID')
            company          = _str(row, 'COMPANY')
            location         = _str(row, 'LOCATION')
            mobile           = _str(row, 'MOBILE')
            guardian_mobile  = _str(row, 'GUARDIAN MOBILE')
            raw_sex          = _str(row, 'SEX').strip().capitalize()
            sex              = raw_sex if raw_sex in ('Male', 'Female') else ''
            address          = _str(row, 'ADDRESS')
            aadhar           = _str(row, 'AADHAR')
            pan              = _str(row, 'PAN').upper()
            epf_number       = _str(row, 'EPF NUMBER')
            esic_number      = _str(row, 'ESIC NUMBER')
            gross_salary     = _flt(row, 'SALARY')
            basic_salary     = _flt(row, 'BASIC PAY')
            epf_amount       = _flt(row, 'EPF')
            if epf_amount > _EPF_MAX:
                epf_amount = _EPF_MAX
            esic_amount      = _flt(row, 'ESIC')
            credit_repayment = _flt(row, 'CREDIT REPAYMENT')
            raw_status       = _str(row, 'STATUS').lower()
            status           = raw_status if raw_status in ('active', 'inactive') else 'active'
            weekday_shift    = _parse_shift(_str(row, 'WEEKDAY SHIFT'), ['shift1','shift2','shift3','shift4'])
            sunday_shift     = _parse_shift(_str(row, 'SUNDAY SHIFT'),  ['shift1','shift2','shift3'])
            bank_name            = _str(row, 'BANK NAME')
            account_holder_name  = _str(row, 'ACCOUNT HOLDER NAME')
            account_number       = _str(row, 'ACCOUNT NUMBER')
            ifsc_code            = _str(row, 'IFSC CODE').upper()

            # Skip if a duplicate employee code already exists (non-empty codes only)
            if emp_code:
                dup = conn.execute(
                    "SELECT id FROM employees WHERE emp_code=?", (emp_code,)
                ).fetchone()
                if dup:
                    skipped += 1
                    continue
            else:
                emp_code = _next_emp_code(conn)

            conn.execute(
                """INSERT INTO employees
                   (emp_code, name, company, location, mobile, guardian_mobile, sex, address,
                    aadhar, pan, epf_number, esic_number,
                    gross_salary, basic_salary, epf_amount, esic_amount,
                    credit_repayment, weekday_shift, sunday_shift,
                    bank_name, account_holder_name, account_number, ifsc_code, status)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (emp_code, name, company, location, mobile, guardian_mobile, sex, address,
                 aadhar, pan, epf_number, esic_number,
                 gross_salary, basic_salary, epf_amount, esic_amount,
                 credit_repayment, weekday_shift, sunday_shift,
                 bank_name, account_holder_name, account_number, ifsc_code, status)
            )
            inserted += 1
        except Exception:
            skipped += 1

    conn.commit()
    conn.close()
    return redirect(url_for('employees',
                            import_ok=inserted, import_err=skipped))


@payroll_bp.route('/export_employees')
def export_employees():
    """Export monthly payroll for Hotel Bell Elite (single company)."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

    year, month = _period_from_source(request.args)

    conn = get_db()
    rows = conn.execute(
        f"SELECT * FROM employees ORDER BY {_EMPLOYEE_DISPLAY_ORDER}"
    ).fetchall()

    wb = Workbook()
    headers = ['Emp ID', 'Name', 'Department',
               'Gross Salary', 'Total Days', 'Holiday', 'Present', 'Weekday Leave',
               'Sunday Worked',
               'Total Off', 'LOP Days', 'Paid Days', 'Actual Gross', 'Basic', 'EPF', 'ESIC',
               'Overtime', 'Credit Repayment', 'Total Credit', 'Net Salary']
    non_epf_headers = ['Emp ID', 'Name', 'Department',
                       'Gross Salary', 'Total Days', 'Holiday', 'Present', 'Weekday Leave',
                       'Sunday Worked',
                       'Total Off', 'LOP Days', 'Paid Days', 'Actual Gross',
                       'Overtime', 'Credit Repayment', 'Total Credit', 'Net Salary']

    hf = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    df = Font(name='Segoe UI', size=10)
    tb = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
    la = Alignment(horizontal='left', vertical='center', wrap_text=True)
    hfill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')

    def _write_payroll_sheet(ws, sheet_title, company_rows, include_epf_esic=True):
        sheet_headers = headers if include_epf_esic else non_epf_headers
        ws.title = sheet_title
        ws.append(sheet_headers)

        for col in range(1, len(sheet_headers) + 1):
            c = ws.cell(row=1, column=col)
            c.font = hf
            c.border = tb
            c.alignment = ca
            c.fill = hfill

        for e in company_rows:
            view = _attach_employee_month_context(conn, e, year, month)
            cr = view['credit_repayment']
            att = view['att']
            credit_total = view['credit_total']
            total_absent = float(att['weekday_leave_days'])
            if total_absent.is_integer():
                total_absent = int(total_absent)
            paid_days = view.get('paid_calendar_days', 0)
            if isinstance(paid_days, float) and paid_days.is_integer():
                paid_days = int(paid_days)
            row_data = [
                view.get('emp_code', ''), view['name'], view['location'],
                _round_rupee(view['gross_salary']), att.get('num_days', 0), att.get('holiday', 0),
                att['present'], total_absent, att.get('sunday_effective', 0),
                view.get('total_off', 0), view.get('lop_days', 0), paid_days, view['gross_actual'],
            ]
            if include_epf_esic:
                row_data.extend([view['basic'], view['epf'], view['esic']])
            row_data.extend([
                view.get('sunday_incentive', 0), _round_rupee(cr), credit_total, view['net']
            ])
            ws.append(row_data)

        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, len(sheet_headers) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.font = df
                c.border = tb
                c.alignment = ca if col_idx >= 4 else la

        col_widths = [12, 22, 18, 14, 10, 10, 10, 12, 12, 10, 10, 12, 14, 14, 14, 14, 14, 14, 14, 14]
        if not include_epf_esic:
            col_widths = [12, 22, 18, 14, 10, 10, 10, 12, 12, 10, 10, 12, 14, 14, 14, 14, 14]
        for i, w in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

    regular_rows = []
    non_epf_rows = []
    for row in rows:
        item = dict(row) if not isinstance(row, dict) else dict(row)
        if bool(item.get('epf_exempt', 0)):
            non_epf_rows.append(item)
        else:
            regular_rows.append(item)

    _write_payroll_sheet(wb.active, _DEFAULT_COMPANY, regular_rows)
    if non_epf_rows:
        non_epf_ws = wb.create_sheet()
        _write_payroll_sheet(non_epf_ws, 'Non EPF Employees', non_epf_rows, include_epf_esic=False)

    conn.close()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'payroll_{calendar.month_abbr[month].lower()}_{year}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Additional Export Routes ──────────────────────────────────────────────

def _xl_style_header(ws, headers, fill_hex='1F4E79'):
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    hf = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
    hfill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = hf; c.border = tb; c.alignment = ca; c.fill = hfill

def _xl_style_data(ws, headers, center_from=6):
    from openpyxl.styles import Font, Border, Side, Alignment
    df = Font(name='Segoe UI', size=10)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    ca = Alignment(horizontal='center', vertical='center')
    la = Alignment(horizontal='left', vertical='center')
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = df; c.border = tb
            c.alignment = ca if col_idx >= center_from else la

def _xl_col_widths(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

def _xl_send(wb, filename):
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


_REPORT_COMPANY_ORDER = ('NL', 'TLNT')


def _normalize_company_code(value):
    return (value or '').strip().upper()


def _group_report_rows_by_company(rows, company_filter=''):
    """Split rows into ordered company buckets for report exports."""
    buckets = {}
    for row in rows:
        item = dict(row) if not isinstance(row, dict) else dict(row)
        company_code = _normalize_company_code(item.get('company', ''))
        buckets.setdefault(company_code, []).append(item)

    if company_filter:
        ordered_codes = [_normalize_company_code(company_filter)]
    else:
        extras = sorted(code for code in buckets if code and code not in _REPORT_COMPANY_ORDER)
        ordered_codes = list(_REPORT_COMPANY_ORDER) + extras

    pages = []
    for code in ordered_codes:
        label = code or 'OTHERS'
        pages.append((label, buckets.get(code, [])))
    return pages


_WD_SHIFT_LABELS = {
    'shift1': 'Shift 1 - 8:30 AM to 7:30 PM',
    'shift2': 'Shift 2 - 8:30 AM to 7:50 PM',
    'shift3': 'Shift 3 - 8:45 AM to 7:50 PM',
    'shift4': 'Shift 4 - 9:00 AM to 8:30 PM',
}
_SUN_SHIFT_LABELS = {
    'shift1': 'Shift 1 - 9:00 AM to 7:00 PM (Rs.550)',
    'shift2': 'Shift 2 - 9:00 AM to 8:00 PM (Rs.600)',
    'shift3': 'Shift 3 - 9:00 AM to 8:30 PM (Rs.650)',
}

@payroll_bp.route('/export/employee_master')
def export_employee_master():
    """Employee Master Report — Hotel Bell Elite only."""
    from openpyxl import Workbook
    conn = get_db()
    rows = conn.execute(
        f"SELECT * FROM employees ORDER BY {_EMPLOYEE_DISPLAY_ORDER}"
    ).fetchall()
    conn.close()

    COL_WIDTHS = [
        12, 26, 16,                # Emp ID, Name, Department
        14, 16, 10, 10,            # Mobile, Guardian Mobile, Sex, Status
        16, 14, 18, 16,            # Aadhar, PAN, EPF No, ESIC No
        14, 14,                    # Gross, Basic
        14, 10,                    # EPF Amount, EPF Exempt
        14, 10,                    # ESIC Amount, ESIC Exempt
        20, 24, 20, 14,            # Bank Name, Account Holder, Account Number, IFSC
        34,                        # Address
    ]
    headers = [
        'Emp ID', 'Name', 'Department',
        'Mobile', 'Guardian Mobile', 'Sex', 'Status',
        'Aadhar', 'PAN', 'EPF No', 'ESIC No',
        'Gross Salary', 'Basic Salary',
        'EPF Amount', 'EPF Exempt',
        'ESIC Amount', 'ESIC Exempt',
        'Bank Name', 'Account Holder Name', 'Account Number', 'IFSC Code',
        'Address',
    ]

    def _fill_sheet(ws, sheet_rows):
        ws.append(headers)
        _xl_style_header(ws, headers, '1F4E79')
        for r in sheet_rows:
            e = dict(r)
            ws.append([
                e.get('emp_code', ''),
                e['name'],
                e['location'],
                e.get('mobile', ''),
                e.get('guardian_mobile', ''),
                e.get('sex', ''),
                e.get('status', 'active'),
                e.get('aadhar', ''),
                e.get('pan', ''),
                e.get('epf_number', ''),
                e.get('esic_number', ''),
                _round_rupee(e['gross_salary']),
                _round_rupee(e.get('basic_salary', 0) or 0),
                _round_rupee(e.get('epf_amount', 0) or 0),
                'Yes' if e.get('epf_exempt') else 'No',
                _round_rupee(e.get('esic_amount', 0) or 0),
                'Yes' if e.get('esic_exempt') else 'No',
                e.get('bank_name', '') or '',
                e.get('account_holder_name', '') or '',
                e.get('account_number', '') or '',
                e.get('ifsc_code', '') or '',
                e.get('address', ''),
            ])
        _xl_style_data(ws, headers, center_from=5)
        _xl_col_widths(ws, COL_WIDTHS)

    wb = Workbook()
    ws = wb.active
    ws.title = _DEFAULT_COMPANY[:31]
    _fill_sheet(ws, rows)

    return _xl_send(wb, 'employee_master.xlsx')


@payroll_bp.route('/export/attendance_report')
def export_attendance_report():
    """Redirects to the combined Attendance Register (backward compatibility)."""
    year, month = _period_from_source(request.args)
    location = request.args.get('location', '').strip()
    return redirect(url_for('export_attendance_register',
                            year=year,
                            month=month,
                            location=location))


@payroll_bp.route('/export/attendance_register')
def export_attendance_register():
    """Attendance Register — day-wise grid + summary for Hotel Bell Elite."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    year, month = _period_from_source(request.args)
    location = request.args.get('location', '').strip()

    conn = get_db()
    user = get_current_user()

    # ── Fetch employees ──────────────────────────────────────────────────────
    conds, params = ["status='active'"], []
    _append_attendance_scope_conditions(conds, params, user)
    if location:
        conds.append("location=?"); params.append(location)
    where = ' AND '.join(conds)
    emps = [dict(r) for r in conn.execute(
        f"SELECT id, emp_code, name, sex, location, total_off, gross_salary "
        f"FROM employees WHERE {where} ORDER BY {_EMPLOYEE_DISPLAY_ORDER}",
        params
    ).fetchall()]

    # ── Calendar info ────────────────────────────────────────────────────────
    _, num_days = calendar.monthrange(year, month)
    month_name_str = calendar.month_name[month]

    # ── Status codes & colours ───────────────────────────────────────────────
    FILL_P   = PatternFill('solid', fgColor='C6EFCE')   # green
    FILL_A   = PatternFill('solid', fgColor='FFCCCC')   # red
    FILL_H   = PatternFill('solid', fgColor='FFEB9C')   # orange/yellow
    FILL_OUT = PatternFill('solid', fgColor='F1F5F9')   # out-of-month
    FILL_HDR = PatternFill('solid', fgColor='1F4E79')   # header
    FILL_TTL = PatternFill('solid', fgColor='D6E4F0')   # total column

    thin   = Side(style='thin', color='B0BEC5')
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    hdr_font  = Font(name='Segoe UI', bold=True, color='FFFFFF', size=9)
    data_font = Font(name='Segoe UI', size=9)
    bold9     = Font(name='Segoe UI', bold=True, size=9)

    # Emp ID, Name, Sex, Department + days 1–31 + Total Present
    fixed_cols = 4
    total_present_ci = fixed_cols + 32
    total_cols = total_present_ci

    sum_hdrs = ['Emp ID', 'Name', 'Department',
                'Working Days', 'Present', 'Half Day',
                'Absent', 'Effective Days', 'Attendance %']

    wb = Workbook()

    def _write_register_sheet(ws, company_emps):
        ws.title = 'Attendance Register'[:31]

        title_fill = PatternFill('solid', fgColor='0F172A')
        loc_label = f'  |  {location}' if location else ''
        title_text = (
            f'  {_DEFAULT_COMPANY}{loc_label}'
            f'  —  ATTENDANCE REGISTER  |  {month_name_str} {year}'
        )
        ws.cell(row=1, column=1, value=title_text).font = Font(
            name='Segoe UI', bold=True, color='FFFFFF', size=11)
        ws.cell(row=1, column=1).fill = title_fill
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws.row_dimensions[1].height = 22

        fixed_hdrs = ['Emp ID', 'Name', 'Sex', 'Department']
        for ci, h in enumerate(fixed_hdrs, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font = hdr_font
            c.fill = FILL_HDR
            c.alignment = center
            c.border = bdr

        for d in range(1, num_days + 1):
            ci = fixed_cols + d
            c = ws.cell(row=2, column=ci, value=d)
            c.fill = FILL_HDR
            c.font = hdr_font
            c.alignment = center
            c.border = bdr

        for d in range(num_days + 1, 32):
            ci = fixed_cols + d
            c = ws.cell(row=2, column=ci, value='')
            c.fill = PatternFill('solid', fgColor='475569')
            c.font = hdr_font
            c.alignment = center
            c.border = bdr

        c = ws.cell(row=2, column=total_present_ci, value='Total Present')
        c.font = hdr_font
        c.fill = FILL_HDR
        c.alignment = center
        c.border = bdr
        ws.row_dimensions[2].height = 20

        for row_idx, e in enumerate(company_emps, 3):
            att = _get_month_attendance(conn, e['id'], year, month)
            records = att['records']

            for col_idx, value, align in [
                (1, e.get('emp_code', '') or '', center),
                (2, e.get('name', ''), left),
                (3, e.get('sex', '') or '', center),
                (4, e.get('location', '') or '', center),
            ]:
                c = ws.cell(row=row_idx, column=col_idx, value=value)
                c.font = data_font
                c.alignment = align
                c.border = bdr

            for d in range(1, num_days + 1):
                ci = fixed_cols + d
                date_str = f'{year}-{month:02d}-{d:02d}'
                status = records.get(date_str)

                if status == 'present':
                    code = 'P'; fill = FILL_P
                elif status == 'half_day':
                    code = 'H'; fill = FILL_H
                elif status == 'absent':
                    code = 'A'; fill = FILL_A
                else:
                    code = ''; fill = None

                c = ws.cell(row=row_idx, column=ci, value=code)
                c.font = Font(
                    name='Segoe UI', size=8,
                    bold=(code in ('P', 'A', 'H')),
                    color=('C00000' if code == 'A' else
                           '375623' if code == 'P' else
                           '7D4800' if code == 'H' else '000000'))
                if fill:
                    c.fill = fill
                c.alignment = center
                c.border = bdr

            for d in range(num_days + 1, 32):
                ci = fixed_cols + d
                c = ws.cell(row=row_idx, column=ci, value='—')
                c.fill = FILL_OUT
                c.font = Font(name='Segoe UI', size=8, color='CBD5E0')
                c.alignment = center
                c.border = bdr

            total_present = _calc_total_off_lop(
                att.get('weekday_leave_days', 0),
                e.get('total_off', 0),
                e.get('gross_salary', 0),
                att.get('num_days', 0),
            )['paid_calendar_days']
            total_val = int(total_present) if total_present == int(total_present) else total_present
            c = ws.cell(row=row_idx, column=total_present_ci, value=total_val)
            c.font = bold9
            c.fill = FILL_TTL
            c.alignment = center
            c.border = bdr
            ws.row_dimensions[row_idx].height = 16

        ttl_row = 3 + len(company_emps)
        ws.merge_cells(start_row=ttl_row, start_column=1, end_row=ttl_row, end_column=fixed_cols)
        c = ws.cell(row=ttl_row, column=1, value='TOTAL')
        c.font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=9)
        c.fill = FILL_HDR
        c.alignment = center
        c.border = bdr
        for col in range(fixed_cols + 1, total_present_ci + 1):
            c = ws.cell(row=ttl_row, column=col)
            c.fill = FILL_TTL
            c.border = bdr
            c.font = bold9
            c.alignment = center
        if company_emps:
            c = ws.cell(
                row=ttl_row, column=total_present_ci,
                value=f'=SUM({get_column_letter(total_present_ci)}3:{get_column_letter(total_present_ci)}{ttl_row-1})',
            )
            c.fill = FILL_TTL
            c.border = bdr
            c.font = bold9
            c.alignment = center
        ws.row_dimensions[ttl_row].height = 18

        LEG_COL = 40
        legends = [
            ('P  =  Present',  FILL_P, '375623'),
            ('A  =  Absent',   FILL_A, 'C00000'),
            ('H  =  Half Day', FILL_H, '7D4800'),
        ]
        lhdr = ws.cell(row=2, column=LEG_COL, value='Legend:')
        lhdr.font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=9)
        lhdr.fill = FILL_HDR
        lhdr.alignment = center
        lhdr.border = bdr
        for li, (txt, lfill, lcol) in enumerate(legends, 3):
            c = ws.cell(row=li, column=LEG_COL, value=txt)
            c.fill = lfill
            c.font = Font(name='Segoe UI', size=8, bold=True, color=lcol)
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            c.border = bdr

        ws.column_dimensions[get_column_letter(1)].width = 10
        ws.column_dimensions[get_column_letter(2)].width = 22
        ws.column_dimensions[get_column_letter(3)].width = 5
        ws.column_dimensions[get_column_letter(4)].width = 14
        for d in range(1, 32):
            ws.column_dimensions[get_column_letter(fixed_cols + d)].width = 3.5
        ws.column_dimensions[get_column_letter(total_present_ci)].width = 14
        ws.column_dimensions[get_column_letter(LEG_COL)].width = 22
        ws.freeze_panes = 'E3'

    def _write_summary_sheet(ws, company_emps):
        ws.title = 'Attendance Summary'[:31]
        ws.append(sum_hdrs)
        _xl_style_header(ws, sum_hdrs)

        for e in company_emps:
            att = _get_month_attendance(conn, e['id'], year, month)
            pct = (round(att['effective'] / att['total_working'] * 100, 1)
                   if att['total_working'] > 0 else 0.0)
            ws.append([
                e.get('emp_code', ''), e.get('name', ''),
                e.get('location', ''),
                att['total_working'],
                att['present'], att['half_day'], att['absent'],
                att['effective'], pct,
            ])

        _xl_style_data(ws, sum_hdrs, center_from=4)
        _xl_col_widths(ws, [12, 24, 16, 12, 10, 10, 10, 14, 12])
        ws.freeze_panes = 'A2'

    _write_register_sheet(wb.active, emps)
    _write_summary_sheet(wb.create_sheet(), emps)

    conn.close()

    fname = f'attendance_register_{calendar.month_abbr[month].lower()}_{year}.xlsx'
    return _xl_send(wb, fname)


@payroll_bp.route('/export/wage_register')
def export_wage_register():
    """Form XII [Rule 28(1)] — Register of Wages for EPF employees only."""
    from openpyxl import Workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    year, month = _period_from_source(request.args)
    location = request.args.get('location', '').strip()

    conn = get_db()
    conds, params = ["status='active'", "COALESCE(epf_exempt, 0)=0"], []
    if location:
        conds.append("location=?"); params.append(location)
    rows = conn.execute(
        f"SELECT * FROM employees WHERE {' AND '.join(conds)} ORDER BY {_EMPLOYEE_DISPLAY_ORDER}",
        params
    ).fetchall()

    _, num_days    = calendar.monthrange(year, month)
    month_name_str = calendar.month_name[month]
    period_from    = f'01 {month_name_str} {year}'
    period_to      = f'{num_days:02d} {month_name_str} {year}'
    loc_label      = location if location else 'All Locations'

    # ── Styles ────────────────────────────────────────────────────────────────
    HDR_FILL   = PatternFill('solid', fgColor='1F4E79')
    SUB_FILL   = PatternFill('solid', fgColor='2E6DA4')
    TITLE_FILL = PatternFill('solid', fgColor='0F172A')
    INFO_FILL  = PatternFill('solid', fgColor='EFF6FF')
    TTL_FILL   = PatternFill('solid', fgColor='D6E4F0')
    J_WARN_FILL = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
    thin    = Side(style='thin',   color='B0BEC5')
    bdr     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_al = Alignment(horizontal='left',   vertical='center', wrap_text=False, indent=1)
    hdr_f   = Font(name='Segoe UI', bold=True, color='FFFFFF', size=9)
    sub_f   = Font(name='Segoe UI', bold=True, color='FFFFFF', size=8)
    data_f  = Font(name='Segoe UI', size=9)
    bold_f  = Font(name='Segoe UI', bold=True, size=9)
    info_f  = Font(name='Segoe UI', bold=True, size=9, color='1F4E79')

    TOTAL_COLS = 18

    wb = Workbook()
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    def _calc_effective_days(view):
        paid_days = view.get('paid_calendar_days', 0)
        return int(paid_days) if isinstance(paid_days, float) and paid_days.is_integer() else paid_days

    def _write_wage_sheet(ws, sheet_title, sheet_company, company_rows):
        ws.title = sheet_title[:31]
        est_label = f'NEERAJ TEXTILE — {sheet_company}' if sheet_company else 'NEERAJ TEXTILE'

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
        c = ws.cell(row=1, column=1,
                    value='  FORM XII  [Rule 28(1)]  —  REGISTER OF WAGES')
        c.font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=11)
        c.fill = TITLE_FILL
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 22

        ws.merge_cells(start_row=2, start_column=1,  end_row=2, end_column=7)
        ws.merge_cells(start_row=2, start_column=8,  end_row=2, end_column=13)
        ws.merge_cells(start_row=2, start_column=14, end_row=2, end_column=TOTAL_COLS)
        for col, val in [(1,  f'Name of Establishment:   {est_label}'),
                         (8,  f'Wages Period:   {period_from}  to  {period_to}'),
                         (14, f'Place:   {loc_label}')]:
            c = ws.cell(row=2, column=col, value=val)
            c.font = info_f
            c.fill = INFO_FILL
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            c.border = bdr
        ws.row_dimensions[2].height = 20

        def hc(row, col, val, fill=HDR_FILL, fnt=None, span_rows=1, span_cols=1):
            if span_rows > 1 or span_cols > 1:
                ws.merge_cells(start_row=row, start_column=col,
                               end_row=row+span_rows-1, end_column=col+span_cols-1)
            c = ws.cell(row=row, column=col, value=val)
            c.font = fnt or hdr_f
            c.fill = fill
            c.alignment = center
            c.border = bdr

        for col, lbl in [
            (1,  'Name of\nEmployee'),
            (2,  'Company'),
            (3,  'Location'),
            (4,  "Father's /\nHusband's Name"),
            (5,  'Designation'),
            (10, 'Total\nAttendance'),
            (11, 'Over-time /\nSun. Worked'),
            (12, 'Gross Wages\nPayable (₹)'),
            (13, 'EPF\n(₹)'),
            (14, 'ESIC\n(₹)'),
            (15, 'Total\nDeductions (₹)'),
            (16, 'Wages\nPaid (₹)'),
            (17, 'Dis. of\nPayment'),
            (18, 'Signature /\nThumb Impression'),
        ]:
            hc(3, col, lbl, span_rows=2)

        hc(3, 6, 'Minimum Rates of\nWages Payable', span_cols=2)
        hc(3, 8, 'Rates of Wages\nPayable', span_cols=2)
        for col, lbl in [(6, 'Basic (₹)'), (7, 'DA'), (8, 'Basic (₹)'), (9, 'DA')]:
            hc(4, col, lbl, fill=SUB_FILL, fnt=sub_f)

        ws.row_dimensions[3].height = 32
        ws.row_dimensions[4].height = 18

        DATA_START = 5
        for i, e in enumerate(company_rows, 1):
            view = _attach_employee_month_context(conn, e, year, month)
            epf_val = _round_rupee(view.get('epf', 0) or 0)
            esic_val = _round_rupee(view.get('esic', 0) or 0)
            tot_ded = _round_rupee(epf_val + esic_val)
            net_pay = _round_rupee(view.get('net', 0) or 0)
            row_idx = DATA_START + i - 1

            row_data = [
                view.get('name', ''),
                view.get('company', ''),
                view.get('location', ''),
                '',
                '',
                '',
                '',
                '',
                '',
                f'=INT(L{row_idx}/H{row_idx})',
                f'=L{row_idx}-(H{row_idx}*J{row_idx})',
                f'=SUM(M{row_idx},N{row_idx},P{row_idx})',
                epf_val,
                esic_val,
                tot_ded,
                net_pay,
                '',
                '',
            ]

            for col_idx, val in enumerate(row_data, 1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.font = data_f
                c.border = bdr
                c.alignment = left_al if col_idx == 1 else center
            ws.row_dimensions[row_idx].height = 16

        ttl_row = DATA_START + len(company_rows)
        data_end = ttl_row - 1
        if data_end >= DATA_START:
            ws.conditional_formatting.add(
                f'J{DATA_START}:J{data_end}',
                FormulaRule(formula=[f'J{DATA_START}>26'], fill=J_WARN_FILL),
            )
        ws.merge_cells(start_row=ttl_row, start_column=1, end_row=ttl_row, end_column=5)
        c = ws.cell(row=ttl_row, column=1, value='TOTAL')
        c.font = bold_f
        c.fill = TTL_FILL
        c.alignment = center
        c.border = bdr

        sum_cols = {10, 11, 12, 13, 14, 15, 16}
        for col in range(6, TOTAL_COLS + 1):
            c = ws.cell(row=ttl_row, column=col)
            c.border = bdr
            c.fill = TTL_FILL
            c.font = bold_f
            c.alignment = center
            if col in sum_cols and data_end >= DATA_START:
                letter = get_column_letter(col)
                c.value = f'=SUM({letter}{DATA_START}:{letter}{data_end})'
        ws.row_dimensions[ttl_row].height = 18

        col_widths = [22, 12, 12, 20, 14, 13, 8, 13, 8, 11, 11, 14, 10, 10, 14, 12, 12, 18]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = 'A5'

    company_pages = _group_report_rows_by_company(rows)
    for idx, (company_name, company_rows) in enumerate(company_pages):
        ws = wb.active if idx == 0 else wb.create_sheet()
        _write_wage_sheet(ws, company_name, company_name, company_rows)

    conn.close()

    fname = f'wage_register_{calendar.month_abbr[month].lower()}_{year}.xlsx'
    return _xl_send(wb, fname)


@payroll_bp.route('/export/credits_report')
def export_credits_report():
    """Credit/Advance Report — Credit Summary (Sheet 1) + All Entries (Sheet 2)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    year, month = _period_from_source(request.args)
    month_label = _period_label(year, month)
    conn = get_db()
    emp_totals = conn.execute("""
        SELECT e.id, e.name, e.emp_code, e.company, e.location,
               COALESCE(SUM(c.amount), 0)       AS total_credit,
               COUNT(c.id)                      AS entry_count
        FROM employees e
        LEFT JOIN credits c ON e.id = c.employee_id
        GROUP BY e.id
        HAVING COUNT(c.id) > 0
        ORDER BY total_credit DESC
    """).fetchall()
    entries = conn.execute("""
        SELECT c.id, c.date, c.description, c.amount,
               e.name AS emp_name, e.emp_code, e.company, e.location
        FROM credits c
        JOIN employees e ON c.employee_id = e.id
        ORDER BY c.date DESC, e.name
    """).fetchall()
    repayment_map = _get_month_credit_repayment_map(conn, year, month)
    conn.close()

    emp_totals = [
        {
            **dict(r),
            'credit_repayment': repayment_map.get(r['id'], 0),
        }
        for r in emp_totals
    ]
    entries    = [dict(r) for r in entries]

    # ── KPI figures ───────────────────────────────────────────────────────────
    total_outstanding = sum(r['total_credit']     for r in emp_totals)
    total_employees   = len(emp_totals)
    # ── Shared styles ─────────────────────────────────────────────────────────
    HDR_FILL   = PatternFill('solid', fgColor='1F4E79')
    TITLE_FILL = PatternFill('solid', fgColor='0F172A')
    TTL_FILL   = PatternFill('solid', fgColor='D6E4F0')
    ADV_FILL   = PatternFill('solid', fgColor='FEE2E2')   # red  – advance
    REP_FILL   = PatternFill('solid', fgColor='DCFCE7')   # green – repayment
    thin  = Side(style='thin', color='B0BEC5')
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_a = Alignment(horizontal='left',   vertical='center', indent=1)

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — CREDIT SUMMARY  (first tab)
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Credit Summary'
    NCOLS = 7   # Emp Code, Name, Company, Location, Total Credit, Monthly Repay, Entries

    # Row 1 – Title bar
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    c = ws1.cell(row=1, column=1,
                 value='  NEERAJ TEXTILE  —  CREDIT / ADVANCE REGISTER')
    c.font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=11)
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws1.row_dimensions[1].height = 24

    # Row 2 – sub-caption
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
    c = ws1.cell(row=2, column=1,
                 value=f'  Generated: {date.today().strftime("%d %B %Y")}   |   Selected repayment month: {month_label}   |   {total_employees} employees with credit history')
    c.font = Font(name='Segoe UI', italic=True, size=9, color='64748B')
    c.fill = PatternFill('solid', fgColor='F8FAFC')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ws1.row_dimensions[2].height = 16

    # Rows 3-4 – KPI pair
    # Left half = cols 1-4, Right half = cols 5-7
    # NO borders on KPI cells — fill colour defines each block; no grey lines
    KPI = [
        ('Total Outstanding',       _inr_format(total_outstanding, 2),  'EFF6FF', '1F4E79'),
        ('Employees with History',  str(total_employees),            'FFF7ED', 'C2410C'),
    ]
    KPI_POS = [
        (3, 4, 1, 4),   # top-left
        (3, 4, 5, 7),   # top-right
    ]
    for (lbl_row, val_row, sc, ec), (lbl, val, bg, fc) in zip(KPI_POS, KPI):
        kpi_fill = PatternFill('solid', fgColor=bg)
        # Label row – no border
        ws1.merge_cells(start_row=lbl_row, start_column=sc,
                        end_row=lbl_row, end_column=ec)
        c = ws1.cell(row=lbl_row, column=sc, value=lbl)
        c.font = Font(name='Segoe UI', bold=True, size=8, color=fc)
        c.fill = kpi_fill
        c.alignment = Alignment(horizontal='center', vertical='bottom')
        # Value row – no border
        ws1.merge_cells(start_row=val_row, start_column=sc,
                        end_row=val_row, end_column=ec)
        c = ws1.cell(row=val_row, column=sc, value=val)
        c.font = Font(name='Segoe UI', bold=True, size=14, color=fc)
        c.fill = kpi_fill
        c.alignment = Alignment(horizontal='center', vertical='top')
    ws1.row_dimensions[3].height = 16
    ws1.row_dimensions[4].height = 26

    # Row 5 – Column headers (immediately after KPIs, no spacer)
    HDR_ROW = 5
    s1_hdrs = ['Emp Code', 'Name', 'Company', 'Location',
               'Total Credit (₹)', 'Monthly Repay (₹)', 'Entries']
    for col, h in enumerate(s1_hdrs, 1):
        c = ws1.cell(row=HDR_ROW, column=col, value=h)
        c.font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=9)
        c.fill = HDR_FILL; c.alignment = center; c.border = bdr
    ws1.row_dimensions[HDR_ROW].height = 22

    # AutoFilter on header row
    DATA_S1 = HDR_ROW + 1
    ws1.auto_filter.ref = (f'A{HDR_ROW}:'
                           f'{get_column_letter(NCOLS)}{HDR_ROW + len(emp_totals)}')

    for i, r in enumerate(emp_totals, 1):
        row_idx = DATA_S1 + i - 1
        tc = r['total_credit'];  cr = r['credit_repayment']
        row_data = [r['emp_code'], r['name'], r['company'], r['location'],
                    tc, cr, r['entry_count']]
        for col, val in enumerate(row_data, 1):
            c = ws1.cell(row=row_idx, column=col, value=val)
            c.font = Font(name='Segoe UI', size=9,
                          bold=(col == 5),
                          color=('BE123C' if col == 5 else '000000'))
            c.border = bdr
            c.alignment = left_a if col == 2 else center
        ws1.row_dimensions[row_idx].height = 16

    # Total row
    ttl_s1 = DATA_S1 + len(emp_totals)
    ws1.merge_cells(start_row=ttl_s1, start_column=1, end_row=ttl_s1, end_column=4)
    c = ws1.cell(row=ttl_s1, column=1, value='TOTAL')
    c.font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=9)
    c.fill = HDR_FILL; c.alignment = center; c.border = bdr
    for col in range(5, NCOLS + 1):
        c = ws1.cell(row=ttl_s1, column=col)
        c.fill = TTL_FILL; c.border = bdr
        c.font = Font(name='Segoe UI', bold=True, size=9); c.alignment = center
        if col in (5, 6, 7) and ttl_s1 > DATA_S1:
            letter = get_column_letter(col)
            c.value = f'=SUM({letter}{DATA_S1}:{letter}{ttl_s1-1})'
    ws1.row_dimensions[ttl_s1].height = 18

    _xl_col_widths(ws1, [14, 20, 14, 14, 18, 18, 10])
    ws1.freeze_panes = f'A{DATA_S1}'

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — ALL ENTRIES  (second tab)
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('All Entries')
    NCOLS2 = 8

    # Title bar
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS2)
    c = ws2.cell(row=1, column=1,
                 value='  NEERAJ TEXTILE  —  ALL CREDIT / ADVANCE ENTRIES')
    c.font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=11)
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws2.row_dimensions[1].height = 22

    # Column headers
    s2_hdrs = ['Date', 'Employee Name', 'Emp Code',
               'Company', 'Location', 'Description', 'Amount (₹)', 'Type']
    for col, h in enumerate(s2_hdrs, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=9)
        c.fill = HDR_FILL; c.alignment = center; c.border = bdr
    ws2.row_dimensions[2].height = 20

    # AutoFilter
    ws2.auto_filter.ref = f'A2:{get_column_letter(NCOLS2)}{2 + len(entries)}'

    for i, r in enumerate(entries, 1):
        amt  = r['amount']
        etype = 'Advance' if amt > 0 else ('Repayment' if amt < 0 else '—')
        afill = ADV_FILL if amt > 0 else (REP_FILL if amt < 0 else None)
        acol  = 'BE123C' if amt > 0 else ('166534' if amt < 0 else '000000')
        row_data = [r['date'], r['emp_name'], r['emp_code'],
                    r['company'], r.get('location', ''),
                    r.get('description', '') or '—', amt, etype]
        row_idx = i + 2
        for col, val in enumerate(row_data, 1):
            c = ws2.cell(row=row_idx, column=col, value=val)
            c.border = bdr
            c.alignment = left_a if col in (2, 6) else center
            if col in (7, 8):
                c.font = Font(name='Segoe UI', size=9, bold=(col == 7), color=acol)
                if afill: c.fill = afill
            else:
                c.font = Font(name='Segoe UI', size=9)
        ws2.row_dimensions[row_idx].height = 15

    # Total row for amounts
    ttl_s2 = 3 + len(entries)
    ws2.merge_cells(start_row=ttl_s2, start_column=1, end_row=ttl_s2, end_column=6)
    c = ws2.cell(row=ttl_s2, column=1, value='TOTAL')
    c.font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=9)
    c.fill = HDR_FILL; c.alignment = center; c.border = bdr
    if ttl_s2 > 3:
        c = ws2.cell(row=ttl_s2, column=7,
                     value=f'=SUM(G3:G{ttl_s2-1})')
        c.font = Font(name='Segoe UI', bold=True, size=9)
        c.fill = TTL_FILL; c.alignment = center; c.border = bdr
    ws2.cell(row=ttl_s2, column=8).border = bdr
    ws2.row_dimensions[ttl_s2].height = 18

    _xl_col_widths(ws2, [14, 26, 12, 14, 16, 34, 14, 12])
    ws2.freeze_panes = 'A3'

    fname = f'credit_advance_{date.today().strftime("%Y%m%d")}.xlsx'
    return _xl_send(wb, fname)


_BANK_REPORT_TEMPLATE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    'templates', 'payroll', 'bank_report_template.xlsx',
)
_BANK_DEBIT_ACC_NO = '387905000829'
_BANK_EMAIL_ID = 'mithra.varma@gmail.com'


@payroll_bp.route('/export/bank_report')
def export_bank_report():
    """ICICI fund-transfer Excel for active EPF employees (selected payroll month)."""
    from openpyxl import load_workbook

    year, month = _period_from_source(request.args)
    if not os.path.isfile(_BANK_REPORT_TEMPLATE):
        return ('Bank report template is missing.', 500)

    conn = get_db()
    rows = conn.execute(
        f"""SELECT * FROM employees
            WHERE status='active' AND COALESCE(epf_exempt, 0)=0
            ORDER BY {_EMPLOYEE_DISPLAY_ORDER}"""
    ).fetchall()

    payment_date = date.today()
    bank_rows = []
    for row in rows:
        view = _attach_employee_month_context(conn, row, year, month)
        amount = _round_rupee(view.get('net', 0) or 0)
        if amount <= 0:
            continue
        ifsc = (view.get('ifsc_code') or '').strip().upper()
        holder = (view.get('account_holder_name') or '').strip() or (view.get('name') or '')
        bank_rows.append({
            'name': holder,
            'account': (view.get('account_number') or '').strip(),
            'ifsc': ifsc,
            'amount': amount,
            'mobile': (view.get('mobile') or '').strip(),
            'mode': 'FT' if ifsc.startswith('ICIC') else 'NEFT',
        })
    conn.close()

    wb = load_workbook(_BANK_REPORT_TEMPLATE)
    ws = wb.active

    # Keep header row; clear sample data rows.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # Clone formatting from the original first sample row if available via template styles;
    # after delete, append values with shared constants.
    for item in bank_rows:
        ws.append([
            'PAB_VENDOR',                 # A
            item['mode'],                 # B PYMT_MODE
            _BANK_DEBIT_ACC_NO,           # C
            item['name'],                 # D BNF_NAME
            item['account'],              # E BENE_ACC_NO
            item['ifsc'],                 # F BENE_IFSC
            item['amount'],               # G AMOUNT
            'SALARRY',                    # H
            'SALARRY',                    # I
            item['mobile'],               # J MOBILE_NUM
            _BANK_EMAIL_ID,               # K
            'NIL',                        # L
            payment_date,                 # M PYMT_DATE
            'NIL',                        # N
            'NIL',                        # O
            'NIL',                        # P
            'NIL',                        # Q
            'NIL',                        # R
            'NIL',                        # S
        ])
        # Ensure payment date is a real date cell (not text)
        ws.cell(row=ws.max_row, column=13).value = payment_date

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'bank_report_{calendar.month_abbr[month].lower()}_{year}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ── Quick inline update for salary fields ────────────────────────────────

@payroll_bp.route('/lock_payroll_month', methods=['POST'])
def lock_payroll_month():
    year, month = _period_from_source(request.form)
    if request.form.get('lock_action') != 'manual_lock':
        return redirect(request.referrer or url_for('employees', year=year, month=month))
    conn = get_db()
    try:
        state = _get_payroll_month_state(conn, year, month)
        if state['can_lock'] and not state['locked']:
            existing = conn.execute(
                "SELECT 1 FROM payroll_month_locks WHERE year=? AND month=?",
                (year, month)
            ).fetchone()
            if not existing:
                conn.execute(
                    "INSERT INTO payroll_month_locks (year, month) VALUES (?, ?)",
                    (year, month)
                )
                conn.commit()
    finally:
        conn.close()

    return redirect(request.referrer or url_for('employees', year=year, month=month))


@payroll_bp.route('/update_salary/<int:emp_id>', methods=['POST'])
def update_salary(emp_id):
    """AJAX endpoint: update month-scoped credit repayment for an employee."""
    data = request.get_json(silent=True) or {}
    cr = data.get('credit_repayment')
    year, month = _period_from_source(data)

    try:
        cr_val = float(cr) if cr not in (None, '') else 0
    except (TypeError, ValueError):
        return jsonify({'error': 'Invalid number'}), 400
    if cr_val < 0:
        return jsonify({'error': 'Credit repayment cannot be negative.'}), 400

    conn = get_db()
    emp = conn.execute("SELECT * FROM employees WHERE id=?", (emp_id,)).fetchone()
    if not emp:
        conn.close()
        return jsonify({'error': 'Employee not found'}), 404

    payroll_state = _get_payroll_month_state(conn, year, month)
    if payroll_state['locked'] or not payroll_state['can_edit']:
        conn.close()
        return jsonify({
            'error': payroll_state['message'],
            'locked': bool(payroll_state['locked']),
        }), 403

    limits = _get_month_repayment_limits(conn, emp, year, month)
    max_allowed = limits['repayment_max']
    if cr_val > max_allowed + 1e-9:
        conn.close()
        return jsonify({
            'error': f"Maximum allowed for {payroll_state['label']} is ₹{_round_rupee(max_allowed)}.",
            'max_repayment': max_allowed,
            'salary_cap': limits['salary_cap'],
            'credit_cap': limits['credit_total_before_month'],
        }), 400

    _upsert_month_credit_repayment(conn, emp_id, year, month, cr_val)
    conn.commit()

    e = _attach_employee_month_context(conn, emp, year, month, payroll_state=payroll_state)
    conn.close()
    return jsonify({
        'ok': True,
        'basic': e['basic'],
        'epf': e['epf'],
        'esic': e['esic'],
        'net': e['net'],
        'gross_actual': e['gross_actual'],
        'sunday_incentive': e.get('sunday_incentive', 0),
        'lop_deduction': e.get('lop_deduction', 0),
        'lop_days': e.get('lop_days', 0),
        'total_off': e.get('total_off', 0),
        'weekday_leave_days': e.get('weekday_leave_days', 0),
        'paid_calendar_days': e.get('paid_calendar_days', 0),
        'credit_repayment': e['credit_repayment'],
        'credit_balance': e['credit_total'],
        'max_repayment': e['repayment_max'],
        'salary_cap': e['salary_cap'],
        'credit_cap': e['credit_total_before_month'],
        'payroll_month': payroll_state['label'],
        'locked': payroll_state['locked'],
    })

def register_employee_payroll(app, *, pop_auth_notice, permission_denied_response, get_user):
    _bind_app_helpers(pop_auth_notice, permission_denied_response, get_user)
    app.register_blueprint(payroll_bp)
    for rule in list(app.url_map.iter_rules()):
        if not rule.endpoint.startswith("payroll."):
            continue
        bare_endpoint = rule.endpoint.split(".", 1)[1]
        if bare_endpoint in app.view_functions:
            continue
        app.add_url_rule(
            rule.rule,
            endpoint=bare_endpoint,
            view_func=app.view_functions[rule.endpoint],
            methods=rule.methods,
        )
