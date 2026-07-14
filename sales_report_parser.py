"""Parse order-invoice Excel reports into Bar / Restaurant sales aggregates."""

from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from typing import BinaryIO

import openpyxl

OUTLET_BAR = "Bar"
OUTLET_RESTAURANT = "Restaurant"

PAYMENT_FIELDS = ("cash", "card", "upi", "room_credit")


def parse_amount(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    return round(float(text), 2)


def parse_row_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in (
        "%d-%b-%Y %I:%M %p",
        "%d-%b-%Y %H:%M",
        "%d-%b-%Y",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.strptime(text[:11], "%d-%b-%Y").date()
    except ValueError:
        return None


def classify_outlet(invoice_number: str) -> str | None:
    inv = (invoice_number or "").strip().upper()
    if inv.startswith("INV"):
        return OUTLET_BAR
    if inv.startswith("SPC"):
        return OUTLET_RESTAURANT
    return None


def map_payment_mode(pay_mode: str) -> str | None:
    mode = (pay_mode or "").strip()
    if not mode or mode.lower() == "grand total":
        return None
    if mode == "Cash":
        return "cash"
    if mode in ("Credit Card", "Debit Card"):
        return "card"
    if mode == "UPI":
        return "upi"
    if mode == "Room Credit":
        return "room_credit"
    return None


def map_ledger(ledger: str) -> str | None:
    """Map Report - Collections ledger text (column G) to sales entry field."""
    text = (ledger or "").strip()
    if not text:
        return None
    if text == "Cash":
        return "cash"
    if text.startswith("Room Credit"):
        return "room_credit"
    if "Credit" in text or "Debit" in text:
        return "card"
    if text.startswith("UPI"):
        return "upi"
    return None


SECTION_LABELS = frozenset({
    "cash",
    "debit card",
    "credit card",
    "room credit",
    "upi",
    "total",
})


def _is_data_row(row) -> bool:
    if not row or len(row) < 8:
        return False
    invoice = row[1]
    if invoice is None or str(invoice).strip() == "":
        return False
    if str(invoice).strip().lower() == "total":
        return False
    date_cell = row[0]
    if date_cell is None:
        return False
    if isinstance(date_cell, str) and date_cell.strip().lower() in SECTION_LABELS:
        return False
    return True


def _collect_collections_dates(ws) -> list[str]:
    found = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not _is_data_row(row):
            continue
        row_date = parse_row_date(row[0])
        if row_date:
            found.add(row_date.isoformat())
    return sorted(found)


def _collect_order_invoice_dates(ws) -> list[str]:
    found = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 15:
            continue
        row_date = parse_row_date(row[0])
        invoice = row[1]
        if row_date and invoice and classify_outlet(str(invoice)):
            found.add(row_date.isoformat())
    return sorted(found)


def empty_outlet_totals() -> dict[str, float]:
    return {
        "total_sales": 0.0,
        "cash": 0.0,
        "card": 0.0,
        "upi": 0.0,
        "room_credit": 0.0,
    }


def parse_order_invoice_report(file_stream: BinaryIO, sales_date: date) -> dict:
    """
    Parse an order-invoice Excel export.

    Column B = Invoice #, N (14) = Pay Mode, O (15) = Pay Amount.
    """
    wb = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
    ws = wb.active

    totals = {
        OUTLET_BAR: empty_outlet_totals(),
        OUTLET_RESTAURANT: empty_outlet_totals(),
    }
    meta = {
        "rows_bar": 0,
        "rows_restaurant": 0,
        "skipped": 0,
        "sales_date": sales_date.isoformat(),
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 15:
            meta["skipped"] += 1
            continue

        row_date = parse_row_date(row[0])
        if row_date != sales_date:
            meta["skipped"] += 1
            continue

        invoice = row[1]
        pay_mode = row[13]
        pay_field = map_payment_mode(str(pay_mode or ""))
        if pay_field is None:
            meta["skipped"] += 1
            continue

        outlet = classify_outlet(str(invoice or ""))
        if outlet is None:
            meta["skipped"] += 1
            continue

        amount = parse_amount(row[14])
        bucket = totals[outlet]
        bucket["total_sales"] = round(bucket["total_sales"] + amount, 2)
        bucket[pay_field] = round(bucket[pay_field] + amount, 2)

        if outlet == OUTLET_BAR:
            meta["rows_bar"] += 1
        else:
            meta["rows_restaurant"] += 1

    meta["available_dates"] = _collect_order_invoice_dates(ws)
    wb.close()
    return {
        OUTLET_BAR: totals[OUTLET_BAR],
        OUTLET_RESTAURANT: totals[OUTLET_RESTAURANT],
        "meta": meta,
    }


def parse_collections_report(file_stream: BinaryIO, sales_date: date) -> dict:
    """
    Parse Report - Collections Excel export.

    Column B = Invoice # (INV = Bar, SPC = Restaurant)
    Column G = Ledger (payment mode)
    Column H = Amount
    """
    wb = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
    try:
        ws = wb["Report - Collections"] if "Report - Collections" in wb.sheetnames else wb.active
        totals = {
            OUTLET_BAR: empty_outlet_totals(),
            OUTLET_RESTAURANT: empty_outlet_totals(),
        }
        meta = {
            "rows_bar": 0,
            "rows_restaurant": 0,
            "rows_room_transfer": 0,
            "skipped": 0,
            "sales_date": sales_date.isoformat(),
            "sheet": ws.title,
            "format": "collections",
        }
        room_transfer_lines = []
        room_sort_order = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not _is_data_row(row):
                meta["skipped"] += 1
                continue

            row_date = parse_row_date(row[0])
            if row_date != sales_date:
                meta["skipped"] += 1
                continue

            outlet = classify_outlet(str(row[1] or ""))
            if outlet is None:
                meta["skipped"] += 1
                continue

            pay_field = map_ledger(str(row[6] or ""))
            if pay_field is None:
                meta["skipped"] += 1
                continue

            amount = parse_amount(row[7])
            bucket = totals[outlet]
            bucket["total_sales"] = round(bucket["total_sales"] + amount, 2)
            bucket[pay_field] = round(bucket[pay_field] + amount, 2)

            if outlet == OUTLET_BAR:
                meta["rows_bar"] += 1
            else:
                meta["rows_restaurant"] += 1

            if pay_field == "room_credit":
                room_sort_order += 1
                meta["rows_room_transfer"] += 1
                room_transfer_lines.append({
                    "location": outlet,
                    "invoice_number": str(row[1] or "").strip(),
                    "outlet_name": str(row[2] or "").strip() if len(row) > 2 else "",
                    "table_room": str(row[4] or "").strip() if len(row) > 4 else "",
                    "guest_name": str(row[5] or "").strip() if len(row) > 5 else "",
                    "ledger_detail": str(row[6] or "").strip(),
                    "amount": amount,
                    "payment_status": "unpaid",
                    "sort_order": room_sort_order,
                    "source_row": row_idx,
                })

        meta["available_dates"] = _collect_collections_dates(ws)
        return {
            OUTLET_BAR: totals[OUTLET_BAR],
            OUTLET_RESTAURANT: totals[OUTLET_RESTAURANT],
            "room_transfer_lines": room_transfer_lines,
            "meta": meta,
        }
    finally:
        wb.close()


def parse_sales_report(file_stream: BinaryIO, sales_date: date) -> dict:
    """Parse Bar/Restaurant sales upload (collections format, with order-invoice fallback)."""
    file_stream.seek(0)
    wb = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames
    finally:
        wb.close()

    file_stream.seek(0)
    if "Report - Collections" in sheet_names:
        return parse_collections_report(file_stream, sales_date)

    file_stream.seek(0)
    result = parse_order_invoice_report(file_stream, sales_date)
    result["room_transfer_lines"] = []
    return result
