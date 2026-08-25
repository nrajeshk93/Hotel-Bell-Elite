"""KOT report — kitchen-sent Restaurant and Bar orders with lifecycle status."""

from __future__ import annotations

from datetime import date
from io import BytesIO

from db import (
    POS_OUTLET_BAR,
    POS_OUTLET_RESTAURANT,
    ensure_pos_schema,
    normalize_pos_outlet,
    pos_kot_display_no,
)
from reports import format_report_datetime

STATUS_OPEN = "open"
STATUS_INVOICE_GENERATED = "invoice_generated"
STATUS_CANCELLED = "cancelled"

STATUS_LABELS = {
    STATUS_OPEN: "Open",
    STATUS_INVOICE_GENERATED: "Invoice generated",
    STATUS_CANCELLED: "Cancelled",
}

OUTLET_LABELS = {
    POS_OUTLET_RESTAURANT: "Restaurant",
    POS_OUTLET_BAR: "Bar",
}


def kot_display_no(order_no, kot_no=""):
    """Match live Tables KOT display: KOT/SPC|{INV}/{yy-yy}/{n} series."""
    return pos_kot_display_no(order_no, kot_no)

def _parse_iso_date(value):
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _kot_activity_date(row):
    """Prefer first kitchen send, then order_date, then saved_at."""
    for key in ("first_kot_at", "order_date", "saved_at"):
        parsed = _parse_iso_date(row.get(key))
        if parsed is not None:
            return parsed
    return None


def classify_kot_status(row):
    """Return (status_key, status_label) for a POS invoice that had a KOT."""
    is_active = int(row.get("is_active") or 0)
    status = str(row.get("status") or "open").strip().lower() or "open"
    bill_sent = bool(int(row.get("customer_bill_sent") or 0))
    if not is_active or status == "cancelled":
        return STATUS_CANCELLED, STATUS_LABELS[STATUS_CANCELLED]
    if bill_sent:
        return STATUS_INVOICE_GENERATED, STATUS_LABELS[STATUS_INVOICE_GENERATED]
    return STATUS_OPEN, STATUS_LABELS[STATUS_OPEN]


def kot_invoice_no(status_key, order_no, *, bill_sent):
    """Invoice column value — only for invoice-generated KOTs; never cancelled/open."""
    text = " ".join(str(order_no or "").split()).strip()
    if status_key != STATUS_INVOICE_GENERATED or not bill_sent or not text:
        return ""
    return text


def _normalize_outlet_filter(outlet):
    raw = str(outlet or "all").strip().lower()
    if raw in ("all", ""):
        return "all"
    return normalize_pos_outlet(raw)


def load_kot_rows(conn, *, date_from=None, date_to=None, outlet="all"):
    """Load kitchen-sent orders (active + soft-deleted) for the report."""
    ensure_pos_schema(conn)
    outlet_key = _normalize_outlet_filter(outlet)
    params = []
    outlet_sql = ""
    if outlet_key == POS_OUTLET_BAR:
        outlet_sql = " AND LOWER(COALESCE(i.outlet, 'restaurant')) = 'bar' "
    elif outlet_key == POS_OUTLET_RESTAURANT:
        outlet_sql = " AND LOWER(COALESCE(i.outlet, 'restaurant')) = 'restaurant' "
    else:
        outlet_sql = (
            " AND LOWER(COALESCE(i.outlet, 'restaurant')) IN ('restaurant', 'bar') "
        )

    rows = conn.execute(
        f"""
        SELECT
            i.id,
            i.order_no,
            i.outlet,
            i.table_label,
            i.captain,
            i.order_type,
            i.order_date,
            i.saved_at,
            i.first_kot_at,
            i.kot_sent,
            i.kot_no,
            i.customer_bill_sent,
            i.customer_bill_at,
            i.status,
            i.is_active,
            i.cancel_reason,
            i.cancelled_at,
            COALESCE(SUM(COALESCE(l.sent_qty, 0)), 0) AS sent_qty,
            COUNT(CASE WHEN COALESCE(l.sent_qty, 0) > 0 THEN 1 END) AS sent_item_count
        FROM pos_invoices i
        LEFT JOIN pos_invoice_lines l ON l.invoice_id = i.id
        WHERE (
            COALESCE(i.kot_sent, 0) = 1
            OR EXISTS (
                SELECT 1 FROM pos_invoice_lines lx
                WHERE lx.invoice_id = i.id
                  AND COALESCE(lx.sent_qty, 0) > 0
            )
        )
        {outlet_sql}
        GROUP BY i.id
        HAVING COALESCE(SUM(COALESCE(l.sent_qty, 0)), 0) > 0
            OR COALESCE(i.kot_sent, 0) = 1
        ORDER BY
            COALESCE(NULLIF(TRIM(i.first_kot_at), ''), i.order_date, i.saved_at) DESC,
            i.id DESC
        """
    ).fetchall()

    result = []
    for row in rows:
        raw = dict(row)
        activity = _kot_activity_date(raw)
        if date_from and activity and activity < date_from:
            continue
        if date_to and activity and activity > date_to:
            continue
        if (date_from or date_to) and activity is None:
            continue

        outlet_norm = normalize_pos_outlet(raw.get("outlet"))
        status_key, status_label = classify_kot_status(raw)
        order_no = str(raw.get("order_no") or "").strip()
        bill_sent = bool(int(raw.get("customer_bill_sent") or 0))
        invoice_no = kot_invoice_no(status_key, order_no, bill_sent=bill_sent)

        sent_qty = float(raw.get("sent_qty") or 0)
        sent_items = int(raw.get("sent_item_count") or 0)
        cancel_reason = " ".join(str(raw.get("cancel_reason") or "").split()).strip()
        result.append(
            {
                "id": int(raw["id"]),
                "order_no": order_no,
                "kot_no": kot_display_no(order_no, raw.get("kot_no") or ""),
                "outlet": outlet_norm,
                "outlet_label": OUTLET_LABELS.get(outlet_norm, "Restaurant"),
                "table_label": str(raw.get("table_label") or "").strip(),
                "captain": str(raw.get("captain") or "").strip(),
                "order_type": str(raw.get("order_type") or "").strip(),
                "first_kot_at": str(raw.get("first_kot_at") or "").strip(),
                "first_kot_at_display": format_report_datetime(
                    raw.get("first_kot_at") or raw.get("saved_at") or raw.get("order_date")
                ),
                "activity_date": activity.isoformat() if activity else "",
                "sent_qty": sent_qty,
                "sent_qty_display": int(sent_qty) if abs(sent_qty - int(sent_qty)) < 1e-9 else round(sent_qty, 2),
                "sent_item_count": sent_items,
                "status": status_key,
                "status_label": status_label,
                "invoice_no": invoice_no,
                "cancel_reason": cancel_reason,
                "customer_bill_sent": bill_sent,
                "is_active": int(raw.get("is_active") or 0),
            }
        )
    return result


def summarize_kot_rows(rows):
    kpis = {
        "total": 0,
        "open": 0,
        "invoice_generated": 0,
        "cancelled": 0,
        "restaurant": 0,
        "bar": 0,
        "sent_qty": 0.0,
    }
    for row in rows or []:
        kpis["total"] += 1
        status = row.get("status")
        if status == STATUS_OPEN:
            kpis["open"] += 1
        elif status == STATUS_INVOICE_GENERATED:
            kpis["invoice_generated"] += 1
        elif status == STATUS_CANCELLED:
            kpis["cancelled"] += 1
        if row.get("outlet") == POS_OUTLET_BAR:
            kpis["bar"] += 1
        else:
            kpis["restaurant"] += 1
        kpis["sent_qty"] += float(row.get("sent_qty") or 0)
    kpis["sent_qty"] = int(kpis["sent_qty"]) if abs(kpis["sent_qty"] - int(kpis["sent_qty"])) < 1e-9 else round(kpis["sent_qty"], 2)
    return kpis


def build_kot_report(conn, *, date_from=None, date_to=None, outlet="all"):
    rows = load_kot_rows(
        conn,
        date_from=date_from,
        date_to=date_to,
        outlet=outlet,
    )
    return {
        "rows": rows,
        "kpis": summarize_kot_rows(rows),
        "outlet": _normalize_outlet_filter(outlet),
        "date_from": date_from,
        "date_to": date_to,
    }


def build_kot_workbook(payload, *, title_date=""):
    """Return an openpyxl Workbook for the KOT report export."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    rows = payload.get("rows") or []
    kpis = payload.get("kpis") or {}

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    details = wb.create_sheet("KOTs")

    # Match Purchase / Expense ledger Excel chrome (Summary + column headers).
    header_fill = PatternFill(
        fill_type="solid",
        start_color="FF315A78",
        end_color="FF315A78",
    )
    summary_title_font = Font(name="Calibri", bold=True, size=16, color="FFFFFFFF")
    summary_font = Font(name="Calibri", size=12, color="FF000000")
    summary_bold_font = Font(name="Calibri", bold=True, size=12, color="FF000000")
    chrome_header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFFFF")
    body_font = Font(name="Calibri", size=11, color="FF000000")
    thin = Side(style="thin", color="FF000000")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    summary["A1"] = f"Hotel Bell Elite — KOT{title_date}"
    summary.merge_cells("A1:B1")

    summary_rows = (
        ("Total KOTs", int(kpis.get("total") or 0)),
        ("Open", int(kpis.get("open") or 0)),
        ("Invoice generated", int(kpis.get("invoice_generated") or 0)),
        ("Cancelled", int(kpis.get("cancelled") or 0)),
        ("Restaurant", int(kpis.get("restaurant") or 0)),
        ("Bar", int(kpis.get("bar") or 0)),
        ("Items sent", kpis.get("sent_qty") or 0),
    )
    for idx, (label, value) in enumerate(summary_rows, start=2):
        summary.cell(row=idx, column=1, value=label)
        summary.cell(row=idx, column=2, value=value)

    last_summary_row = 1 + len(summary_rows)
    for r in range(1, last_summary_row + 1):
        for col in range(1, 3):
            cell = summary.cell(row=r, column=col)
            cell.border = grid
            cell.alignment = center
            if r == 1:
                cell.fill = header_fill
                cell.font = summary_title_font
            elif r == 2:
                cell.font = summary_bold_font
            else:
                cell.font = summary_font

    summary.row_dimensions[1].height = 22
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 14

    headers = (
        "KOT time",
        "Outlet",
        "KOT no",
        "Table",
        "Captain",
        "Items sent",
        "Qty sent",
        "Status",
        "Invoice no",
        "Cancel reason",
    )
    for col, title in enumerate(headers, start=1):
        cell = details.cell(row=1, column=col, value=title)
        cell.font = chrome_header_font
        cell.fill = header_fill
        cell.border = grid
        cell.alignment = center

    for row_idx, entry in enumerate(rows, start=2):
        values = (
            entry.get("first_kot_at_display") or "",
            entry.get("outlet_label") or "",
            entry.get("kot_no") or "",
            entry.get("table_label") or "",
            entry.get("captain") or "",
            int(entry.get("sent_item_count") or 0),
            entry.get("sent_qty_display") if entry.get("sent_qty_display") is not None else entry.get("sent_qty") or 0,
            entry.get("status_label") or "",
            entry.get("invoice_no") or "",
            entry.get("cancel_reason") or "",
        )
        for col, value in enumerate(values, start=1):
            cell = details.cell(row=row_idx, column=col, value=value)
            cell.font = body_font
            cell.border = grid
            cell.alignment = right if col in (6, 7) else left

    details.row_dimensions[1].height = 20
    widths = (18, 12, 18, 12, 16, 12, 12, 18, 18, 28)
    for col, width in enumerate(widths, start=1):
        details.column_dimensions[get_column_letter(col)].width = width

    return wb


def kot_report_excel_bytes(payload, *, title_date=""):
    wb = build_kot_workbook(payload, title_date=title_date)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
